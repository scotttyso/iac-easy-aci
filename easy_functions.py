#!/usr/bin/env python3

#========================================================
# Source Modules
#========================================================
from copy import deepcopy
from dotmap import DotMap
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side 
from ordered_set import OrderedSet
from textwrap import fill
import git
import jinja2
import json
import itertools
import openpyxl
import os
import pkg_resources
import platform
import re
import requests
import shutil
import subprocess
import sys
import stdiomask
import time
import validating
import yaml

# Global options for debugging
print_payload = False
print_response_always = False
print_response_on_fail = True

#========================================================
# Log Level - 0 = None, 1 = Class only, 2 = Line
#========================================================
log_level = 2

#========================================================
# Exception Classes
#========================================================
class InsufficientArgs(Exception):
    pass

#========================================================
# Function to Connect to the APIC API
#========================================================
def apic_api(apic, method, payload, cookies, uri, section=''):
    s = requests.Session()
    r = ''
    while r == '':
        try:
            if method == 'conf':
                r = s.post('https://{}/{}.json'.format(apic, uri),
                data=payload, cookies=cookies, verify=False)
            elif 'get':
                r = s.get('https://{}/{}.json'.format(apic, uri), cookies=cookies, verify=False)
            status = r.status_code
        except requests.exceptions.ConnectionError as e:
            print("Connection error, pausing before retrying. Error: {}"
                .format(e))
            time.sleep(5)
        except Exception as e:
            print("Method {} failed. Exception: {}".format(section[:-5], e))
            status = 666
            return(status)
    if print_response_always: print(r.text)
    if status != 200 and print_response_on_fail: print(r.text)
    return r

#========================================================
# Function to Connect to the APIC API
#========================================================
def apic_api_with_filter(apic, cookies, uri, uriFilter, section=''):
    s = requests.Session()
    r = ''
    while r == '':
        try:
            print('https://{}/{}.json?{}'.format(apic, uri, uriFilter))
            r = s.get('https://{}/{}.json?{}'.format(apic, uri, uriFilter), cookies=cookies, verify=False)
            status = r.status_code
            print(status)
            print(r.json())
        except requests.exceptions.ConnectionError as e:
            print("Connection error, pausing before retrying. Error: {}"
                .format(e))
            time.sleep(5)
        except Exception as e:
            print("Method {} failed. Exception: {}".format(section[:-5], e))
            status = 666
            return(status)
    if print_response_always: print(r.text)
    if status != 200 and print_response_on_fail: print(r.text)
    return r

#========================================================
# Function to Split Annotations to key and Value
#========================================================
def annotations_split(annotations):
    ann_list = []
    ann_split = annotations.split('')
    for i in ann_split:
        x = i.split('_')
        if len(x) == 2:
            ann_list.append({'key':x[0],'value':x[1]})
    annotations = ann_list
    return annotations

#========================================================
# Function to run 'terraform plan' and 'terraform apply'
# in the each folder of the Destination Directory.
#========================================================
def apply_terraform(args, path_sep, **easyDict):
    folders           = easyDict['changed_folders']
    base_dir          = args.dir
    running_directory = os.getcwd()
    sites             = easyDict['site_names']
    print(f'\n-----------------------------------------------------------------------------\n')
    print(f'  Found the Followng Folders with uncommitted changes:\n')
    for folder in folders: print(f'  - {folder}')
    print(f'\n  Beginning Terraform Proceedures.')
    print(f'\n-----------------------------------------------------------------------------\n')
    tfe_dir = f'.terraform{path_sep}providers{path_sep}'
    tfe_cmd = subprocess.Popen(['terraform', 'init', '-upgrade'],
        cwd=running_directory, stdout=subprocess.PIPE, stderr=subprocess.STDOUT
    )
    output, err = tfe_cmd.communicate()
    tfe_cmd.wait()

    print(output.decode('utf-8'))
    response_p = ''
    response_a = ''
    sites = easyDict['sites'].keys()
    for site in sites:
        run_loc = easyDict['sites'][site]['site_settings']['run_location']
        site_name = easyDict['sites'][site]['site_settings']['site_name']
        path = f'{base_dir}{path_sep}{site_name}'
        run_count = 0
        for folder in folders:
            if site_name in folder:
                run_count +=1
        if run_loc == 'local' and run_count > 0:
            if os.path.isfile(os.path.join(path, '.terraform.lock.hcl')):
                os.remove(os.path.join(path, '.terraform.lock.hcl'))
            if os.path.isdir(os.path.join(path, '.terraform')):
                shutil.rmtree(os.path.join(path, '.terraform'))
            lock_count = 0
            tfe_cmd = subprocess.Popen(
                ['terraform', 'init', f'-plugin-dir={running_directory}/{tfe_dir}'],
                cwd=path, stdout=subprocess.PIPE, stderr=subprocess.STDOUT
            )
            output, err = tfe_cmd.communicate()
            tfe_cmd.wait()
            print(output.decode('utf-8'))
            if 'does not match configured version' in output.decode('utf-8'): lock_count =+ 1
            if lock_count > 0:
                tfe_cmd = subprocess.Popen(
                    ['terraform', 'init', '-upgrade', f'-plugin-dir={running_directory}/{tfe_dir}'], cwd=path
                )
                tfe_cmd.wait()
                print(output.decode('utf-8'))
            tfe_cmd = subprocess.Popen(['terraform', 'plan', '-out=main.plan'], cwd=path)
            output, err = tfe_cmd.communicate()
            tfe_cmd.wait()
            if not output == None: print(output.decode('utf-8'))
            while True:
                print(f'\n-----------------------------------------------------------------------------\n')
                print(f'  Terraform Plan Complete.  Please Review the Plan and confirm if you want')
                print(f'  to move forward.  "A" to Apply the Plan. "S" to Skip.  "Q" to Quit.')
                print(f'  Current Working Directory: {path}')
                print(f'\n-----------------------------------------------------------------------------\n')
                response_p = input('  Please Enter ["A", "S" or "Q"]: ')
                if re.search('^(A|S)$', response_p): break
                elif response_p == 'Q': exit()
                else:
                    print(f'\n-----------------------------------------------------------------------------\n')
                    print(f'  A Valid Response is either "A", "S" or "Q"...')
                    print(f'\n-----------------------------------------------------------------------------\n')

            if response_p == 'A':
                tfe_cmd = subprocess.Popen(['terraform', 'apply', '-parallelism=1', 'main.plan'], cwd=path)
                tfe_cmd.wait()
                output, err = tfe_cmd.communicate()
                tfe_cmd.wait()
                if not output == None: print(output.decode('utf-8'))
                print(f'\n--------------------------------------------------------------------------------\n')
                print(f'  Terraform Apply Complete.  Please Review for any errors and confirm next steps')
                print(f'\n--------------------------------------------------------------------------------\n')

        if args.git_check == True:
            while True:
                print(f'\n-----------------------------------------------------------------------------\n')
                print(f'  Folder: {path}.')
                print(f'  Please confirm if you want to commit {site_name} or just move forward.')
                print(f'  "C" to Commit {site_name} and move forward.')
                print(f'  "S" to Skip the Commit for {site_name}.')
                print(f'  "Q" to Quit..')
                print(f'\n-----------------------------------------------------------------------------\n')
                response_a = input('  Please Enter ["C", "S" or "Q"]: ')
                if response_a == 'C': break
                elif response_a == 'S': break
                elif response_a == 'Q': exit()
                else:
                    print(f'\n-----------------------------------------------------------------------------\n')
                    print(f'  A Valid Response is either "C", "S" or "Q"...')
                    print(f'\n-----------------------------------------------------------------------------\n')

            while True:
                if response_a == 'C':
                    print(f'\n-----------------------------------------------------------------------------\n')
                    commit_message = input(f'  Please Enter your Commit Message for the folder {folder}: ')
                    baseRepo = git.Repo(args.dir)
                    baseRepo.git.add(all=True)
                    baseRepo.git.commit('-m', f'{commit_message}', '--', folder)
                    baseRepo.git.push()
                    break
                else: break

#========================================================
# Function to Add Required Arguments
#========================================================
def args_add(args_list, jsonData):
    for i in args_list:
        jsonData['required_args'].update({f'{i}': ''})
        jsonData['optional_args'].pop(i)
    return jsonData

#========================================================
# Function to Remove Required Arguments
#========================================================
def args_remove(args_list, jsonData):
    for i in args_list:
        jsonData['optional_args'].update({f'{i}': ''})
        jsonData['required_args'].pop(i)
    return jsonData

#========================================================
# Function to Check the Existance of a Template
#========================================================
def confirm_templates_exist(template_type, template_name, **kwargs):
    def template_check(site, template_type, template_name, **kwargs):
        template_count = 0
        if kwargs['easyDict']['sites'][site]['templates'].get(template_type):
            for i in kwargs['easyDict']['sites'][site]['templates'][template_type]:
                if i['template_name'] == template_name:
                    template_count += 1
                    if template_type == 'bridge_domains':
                        if len(kwargs['l3outs']) > 0:
                            if not i.get('l3_configurations'):
                                i['l3_configurations'] = deepcopy({})
                                if not i['l3_configurations'].get('l3outs'):
                                    i['l3_configurations']['l3outs'] = deepcopy([])
                                i['l3_configurations']['l3outs'].append(deepcopy(kwargs['l3outs']))
        if template_count == 0:
            if template_type == 'application_epgs':
                validating.error_template_not_found('epg_template', **kwargs)
            elif template_type == 'bridge_domains':
                validating.error_template_not_found('bd_template', **kwargs)
            else: validating.error_template_not_found('subnet_templates', **kwargs)
    
    if 'Grp_' in kwargs['site_group']:
        sites = kwargs['easyDict']['site_groups'][kwargs['site_group']]['sites']
        for site in sites:
            template_check(site, template_type, template_name, **kwargs)
    else: template_check(kwargs['site_group'], template_type, template_name, **kwargs)

#========================================================
# Function to Count the Number of Keys/Columns
#========================================================
def countKeys(ws, func):
    count = 0
    for i in ws.rows:
        if any(i):
            if str(i[0].value) == func:
                count += 1
    return count

#========================================================
# Function to Create Interface Selectors
#========================================================
def create_selector(ws_sw, ws_sw_row_count, **polVars):
    port_selector = ''
    for port in range(1, int(polVars['port_count']) + 1):
        if port < 10:
            port_selector = 'Eth%s-0%s' % (polVars['module'], port)
        elif port < 100:
            port_selector = 'Eth%s-%s' % (polVars['module'], port)
        elif port > 99:
            port_selector = 'Eth%s_%s' % (polVars['module'], port)
        modport = '%s/%s' % (polVars['module'],port)
        # Copy the Port Selector to the Worksheet
        if polVars['node_type'] == 'spine':
            data = [
                'intf_selector',
                polVars['site_group'],
                polVars['pod_id'],
                polVars['node_id'],
                polVars['switch_name'],
                port_selector,modport,
                'spine_pg','','','','',''
            ]
        else:
            data = [
                'intf_selector',
                polVars['site_group'],
                polVars['pod_id'],
                polVars['node_id'],
                polVars['switch_name'],
                port_selector,modport,
                '','','','','',''
            ]
        ws_sw.append(data)
        rc = f'{ws_sw_row_count}:{ws_sw_row_count}'
        for cell in ws_sw[rc]:
            if ws_sw_row_count % 2 == 0:
                cell.style = 'ws_odd'
            else:
                cell.style = 'ws_even'
        if polVars['node_type'] == 'spine':
            polVars['dv3'] = DataValidation(type="list", formula1='spine_pg', allow_blank=True)
        else:
            polVars['dv3'] = DataValidation(type="list", formula1=f'INDIRECT(H{ws_sw_row_count})', allow_blank=True)
        ws_sw.add_data_validation(polVars['dv3'])
        dv1_cell = f'A{ws_sw_row_count}'
        dv2_cell = f'H{ws_sw_row_count}'
        dv3_cell = f'I{ws_sw_row_count}'
        dv4_cell = f'K{ws_sw_row_count}'
        polVars['dv1'].add(dv1_cell)
        polVars['dv2'].add(dv2_cell)
        polVars['dv3'].add(dv3_cell)
        polVars['dv4'].add(dv4_cell)
        ws_sw_row_count += 1
    return ws_sw_row_count

#========================================================
# Function to Create Static Paths within EPGs
#========================================================
def create_static_paths(**kwargs):
    dest_dir  = kwargs['dest_dir']
    dest_file = kwargs['dest_file']
    polVars = {}
    wb    = kwargs['wb']
    wb_sw = kwargs['wb_sw']
    wsheets = wb_sw.get_sheet_names()
    tf_file = ''
    for wsheet in wsheets:
        ws = wb_sw[wsheet]
        for row in ws.rows:
            if not (row[12].value == None or row[13].value == None):
                vlan_test = ''
                if re.search('^(individual|port-channel|vpc)$', row[7].value) and (re.search(r'\d+', str(row[12].value)
                ) or re.search(r'\d+', str(row[13].value))):
                    if not row[12].value == None:
                        vlan = row[12].value
                        vlan_test = vlan_range(vlan, **kwargs)
                        if 'true' in vlan_test:
                            polVars['mode'] = 'native'
                    if not 'true' in vlan_test:
                        polVars['mode'] = 'regular'
                        if not row[13].value == None:
                            vlans = row[13].value
                            vlan_test = vlan_range(vlans, **kwargs)
                if vlan_test == 'true':
                    polVars['Pod_ID'] = row[1].value
                    polVars['Node_ID'] = row[2].value
                    polVars['Interface_Profile'] = row[3].value
                    polVars['Interface_Selector'] = row[4].value
                    polVars['Port'] = row[5].value
                    polVars['Policy_Group'] = row[6].value
                    polVars['Port_Type'] = row[7].value
                    polVars['Bundle_ID'] = row[9].value
                    Site_Group = polVars['Site_Group']
                    pod = polVars['Pod_ID']
                    node_id =  polVars['Node_ID']
                    if polVars['Port_Type'] == 'vpc':
                        ws_vpc = wb['Inventory']
                        for rx in ws_vpc.rows:
                            if rx[0].value == 'vpc_pair' and int(rx[1].value) == int(Site_Group) and str(rx[4].value) == str(node_id):
                                node1 = polVars['Node_ID']
                                node2 = rx[5].value
                                polVars['Policy_Group'] = '%s_vpc%s' % (row[3].value, polVars['Bundle_ID'])
                                polVars['tDn'] = 'topology/pod-%s/protpaths-%s-%s/pathep-[%s]' % (
                                    pod, node1, node2, polVars['Policy_Group']
                                )
                                polVars['Static_Path'] = 'rspathAtt-[topology/pod-%s/protpaths-%s-%s/pathep-[%s]' % (
                                    pod, node1, node2, polVars['Policy_Group']
                                )
                                polVars['GUI_Static'] = 'Pod-%s/Node-%s-%s/%s' % (pod, node1, node2, polVars['Policy_Group'])
                                polVars['Static_descr'] = 'Pod-%s_Nodes-%s-%s_%s' % (pod, node1, node2, polVars['Policy_Group'])
                                tf_file = './ACI/%s/%s/%s' % (polVars['Site_Name'], dest_dir, dest_file)
                                read_file = open(tf_file, 'r')
                                read_file.seek(0)
                                static_path_descr = 'resource "aci_epg_to_static_path" "%s_%s_%s"' % (
                                    polVars['App_Profile'], polVars['EPG'], polVars['Static_descr']
                                )
                                # if not static_path_descr in read_file.read():
                                #     create_tf_file(wr_method, dest_dir, dest_file, template, **polVars)

                    elif polVars['Port_Type'] == 'port-channel':
                        polVars['Policy_Group'] = '%s_pc%s' % (row[3].value, polVars['Bundle_ID'])
                        polVars['tDn'] = 'topology/pod-%s/paths-%s/pathep-[%s]' % (pod, polVars['Node_ID'], polVars['Policy_Group'])
                        polVars['Static_Path'] = 'rspathAtt-[topology/pod-%s/paths-%s/pathep-[%s]' % (
                            pod, polVars['Node_ID'], polVars['Policy_Group']
                        )
                        polVars['GUI_Static'] = 'Pod-%s/Node-%s/%s' % (pod, polVars['Node_ID'], polVars['Policy_Group'])
                        polVars['Static_descr'] = 'Pod-%s_Node-%s_%s' % (pod, polVars['Node_ID'], polVars['Policy_Group'])
                        read_file = open(tf_file, 'r')
                        read_file.seek(0)
                        static_path_descr = 'resource "aci_epg_to_static_path" "%s_%s_%s"' % (
                            polVars['App_Profile'], polVars['EPG'], polVars['Static_descr']
                        )
                        # if not static_path_descr in read_file.read():
                        #     create_tf_file(wr_method, dest_dir, dest_file, template, **polVars)

                    elif polVars['Port_Type'] == 'individual':
                        port = 'eth%s' % (polVars['Port'])
                        polVars['tDn'] = 'topology/pod-%s/paths-%s/pathep-[%s]' % (pod, polVars['Node_ID'], port)
                        polVars['Static_Path'] = 'rspathAtt-[topology/pod-%s/paths-%s/pathep-[%s]' % (pod, polVars['Node_ID'], port)
                        polVars['GUI_Static'] = 'Pod-%s/Node-%s/%s' % (pod, polVars['Node_ID'], port)
                        polVars['Static_descr'] = 'Pod-%s_Node-%s_%s' % (pod, polVars['Node_ID'], polVars['Interface_Selector'])
                        read_file = open(tf_file, 'r')
                        read_file.seek(0)
                        static_path_descr = 'resource "aci_epg_to_static_path" "%s_%s_%s"' % (
                            polVars['App_Profile'], polVars['EPG'], polVars['Static_descr']
                        )
                        # if not static_path_descr in read_file.read():
                        #     create_tf_file(wr_method, dest_dir, dest_file, template, **polVars)
                        print('hello')

#==========================================================
# Function for Processing easyDict and Creating YAML Files
#==========================================================
def create_yaml(args, easy_jsonData, **easyDict):
    jsonData = easy_jsonData['easy_aci']['allOf'][1]['properties']
    classes = jsonData['classes']['enum']
    opSystem = platform.system()
    if opSystem == 'Windows': path_sep = '\\'
    else: path_sep = '/'

    def write_file(dest_dir, dest_file, dict, title1):
        class MyDumper(yaml.Dumper):
            def increase_indent(self, flow=False, indentless=False):
                return super(MyDumper, self).increase_indent(flow, False)
        
        if not os.path.exists(os.path.join(dest_dir, dest_file)):
            create_file = f'type nul >> {os.path.join(dest_dir, dest_file)}'
            os.system(create_file)
        wr_file = open(os.path.join(dest_dir, dest_file), 'w')
        wr_file.write('---\n')
        wr_file = open(os.path.join(dest_dir, dest_file), 'a')
        dash_length = '='*(len(title1) + 20)
        wr_file.write(f'#{dash_length}\n')
        wr_file.write(f'#   {title1} - Variables\n')
        wr_file.write(f'#{dash_length}\n')
        stream = yaml.dump(dict, default_flow_style=False)
        wr_file.write(stream.replace('\n- ', '\n\n- '))
        # wr_file.write(yaml.dump(dict, Dumper=MyDumper, default_flow_style=False))
        wr_file.close()

    for k,v in easyDict['sites'].items():
        baseRepo = args.dir
        site_name = easyDict['sites'][k]['site_settings']['site_name']
        
        # Remove Random Lines from Dictionary.  Need to figure out why this happens
        pop_list = ['snmp_client_groups', 'snmp_communities', 'snmp_destinations', 'users']
        if easyDict['sites'][k].get('fabric'):
            if easyDict['sites'][k]['fabric'].get('policies'):
                if easyDict['sites'][k]['fabric']['policies'].get('pod'):
                    if easyDict['sites'][k]['fabric']['policies'].get('pod'):
                        for item in easyDict['sites'][k]['fabric']['policies']['pod']['date_and_time']:
                            for i in pop_list:
                                if item.get(i): item.pop(i)
        if easyDict['sites'][k].get('switch'):
            if easyDict['sites'][k]['switch'].get('vpc_domains'):
                for item in easyDict['sites'][k]['switch']['vpc_domains']:
                    item.pop('interfaces')
        for item in classes:
            if easyDict['sites'][k].get(item):
                if item == 'tenants':
                    tcount = 0
                    for i in easyDict['sites'][k][item]:
                        dest_dir = f"tenants{path_sep}{i['name']}"
                        if not os.path.isdir(os.path.join(baseRepo, site_name, dest_dir)):
                            dest_path = f'{os.path.join(baseRepo, site_name)}{path_sep}{dest_dir}'
                            os.makedirs(dest_path)
                        dest_dir = os.path.join(baseRepo, site_name, dest_dir)
                        dict = {item:{i['name']:easyDict['sites'][k][item][tcount]}}
                        dest_file = f"{i['name']}.yaml"
                        title1 = f"{str.title(item)} -> {i['name']}"
                        write_file(dest_dir, dest_file, dict, title1)
                        tcount += 1
                else:
                    dest_dir = jsonData[f'class.{item}']['directory']
                    if not os.path.isdir(os.path.join(baseRepo, site_name, dest_dir)):
                        dest_path = f'{os.path.join(baseRepo, site_name)}{path_sep}{dest_dir}'
                        os.makedirs(dest_path)
                    dest_dir = os.path.join(baseRepo, site_name, dest_dir)
                    for i in jsonData[f'class.{item}']['enum']:
                        if item == i:
                            dict = {item:easyDict['sites'][k][item]}
                        else:
                            dict = {item:{i:easyDict['sites'][k][item][i]}}
                        if item == 'switch' and i == 'switch_profiles':
                            icount = 0
                            for items in dict['switch']['switch_profiles']:
                                dest_file = f"{items['name']}.yaml"
                                title1 = items['name']
                                dict2 = {item:{i:{items['name']:easyDict['sites'][k][item][i][icount]}}}
                                write_file(dest_dir, dest_file, dict2, title1)
                                icount += 1
                        else:
                            dest_file = f'{i}.yaml'
                            if item == i:
                                title1 = str.title(item.replace('_', ' '))
                            else:
                                title1 = f"{str.title(item.replace('_', ' '))} -> {str.title(i.replace('_', ' '))}"
                            write_file(dest_dir, dest_file, dict, title1)
                        
#========================================================
# Function for Processing Loops to auto.tfvars files
#========================================================
def env_sensitive(sensitive_list, jsonData, polVars, **kwargs):
    pop_list = ['easyDict', 'jsonData', 'Variable']

    # Process Sensitive Variables
    for item in sensitive_list:
        polVars['easyDict'] = kwargs['easyDict']
        polVars['jsonData'] = jsonData
        polVars["Variable"] = item
        kwargs['easyDict'] = sensitive_var_site_group(**polVars)
        for i in pop_list:
            polVars.pop(i)

        return polVars, kwargs

#========================================================
# Function to Append the easyDict Dictionary
#========================================================
def ez_append(polVars, **kwargs):
    class_path = kwargs['class_path']
    cS = class_path.split(',')
    polVars.pop('site_group')
    polVars = ez_remove_empty(polVars)

    def site_append(cS, site, polVars):
        if not kwargs['easyDict']['sites'].get(site):
            validating.error_site_group('site_group', **kwargs)

        # Confirm the Key Exists
        if len(cS) == 1:
            if not kwargs['easyDict']['sites'][site].get(cS[0]):
                kwargs['easyDict']['sites'][site].update(deepcopy({cS[0]:[]}))
        if len(cS) >= 2:
            if not kwargs['easyDict']['sites'][site].get(cS[0]):
                kwargs['easyDict']['sites'][site].update(deepcopy({cS[0]:{}}))
        if len(cS) >= 3:
            if not kwargs['easyDict']['sites'][site][cS[0]].get(cS[1]):
                kwargs['easyDict']['sites'][site][cS[0]].update(deepcopy({cS[1]:{}}))
        if len(cS) >= 4:
            if not kwargs['easyDict']['sites'][site][cS[0]][cS[1]].get(cS[2]):
                kwargs['easyDict']['sites'][site][cS[0]][cS[1]].update(deepcopy({cS[2]:{}}))
        if len(cS) == 1: cs_count = 0
        elif len(cS) == 2:
            if not kwargs['easyDict']['sites'][site][cS[0]].get(cS[1]):
                kwargs['easyDict']['sites'][site][cS[0]].update(deepcopy({cS[1]:[]}))
        elif len(cS) == 3:
            if not kwargs['easyDict']['sites'][site][cS[0]][cS[1]].get(cS[2]):
                kwargs['easyDict']['sites'][site][cS[0]][cS[1]].update(deepcopy({cS[2]:[]}))
        elif len(cS) == 4:
            if not kwargs['easyDict']['sites'][site][cS[0]][cS[1]][cS[2]].get(cS[3]):
                kwargs['easyDict']['sites'][site][cS[0]][cS[1]][cS[2]].update(deepcopy({cS[3]:[]}))
        elif len(cS) == 5:
            if not kwargs['easyDict']['sites'][site][cS[0]][cS[1]][cS[2]].get(cS[3]):
                kwargs['easyDict']['sites'][site][cS[0]][cS[1]][cS[2]].update(deepcopy({cS[3]:{}}))
            if not kwargs['easyDict']['sites'][site][cS[0]][cS[1]][cS[2]][cS[3]].get(cS[4]):
                kwargs['easyDict']['sites'][site][cS[0]][cS[1]][cS[2]][cS[3]].update(deepcopy({cS[4]:[]}))
        
        # append the Dictionary
        if len(cS) == 1:   kwargs['easyDict']['sites'][site][cS[0]].append(deepcopy(polVars))
        elif len(cS) == 2: kwargs['easyDict']['sites'][site][cS[0]][cS[1]].append(deepcopy(polVars))
        elif len(cS) == 3: kwargs['easyDict']['sites'][site][cS[0]][cS[1]][cS[2]].append(deepcopy(polVars))
        elif len(cS) == 4: kwargs['easyDict']['sites'][site][cS[0]][cS[1]][cS[2]][cS[3]].append(deepcopy(polVars))
        elif len(cS) == 5: kwargs['easyDict']['sites'][site][cS[0]][cS[1]][cS[2]][cS[3]][cS[4]].append(deepcopy(polVars))

    if 'Grp_' in kwargs['site_group']:
        if kwargs['easyDict']['site_groups'].get(kwargs['site_group']):
            sites = kwargs['easyDict']['site_groups'][kwargs['site_group']]['sites']
            for site in sites:
                site_append(cS, site, polVars)
        else: validating.error_site_group('site_group', **kwargs)
    else: site_append(cS, kwargs['site_group'], polVars)
        
    return kwargs['easyDict']

#========================================================
# Function to Append Subtype easyDict Dictionary
#========================================================
def ez_append_subtype(polVars, **kwargs):
    class_path   = kwargs['class_path']
    cS = class_path.split(',')
    policy  = kwargs['policy']
    policy_name  = kwargs['policy_name']
    polVars.pop('site_group')
    polVars = ez_remove_empty(polVars)

    def site_append(cS, site, polVars):
        # Assign the Dictionary
        if len(cS) == 3:   dict1 = kwargs['easyDict']['sites'][site][cS[0]]
        elif len(cS) == 4: dict1 = kwargs['easyDict']['sites'][site][cS[0]][cS[1]]
        elif len(cS) == 5: dict1 = kwargs['easyDict']['sites'][site][cS[0]][cS[1]][cS[2]]
        elif len(cS) == 6: dict1 = kwargs['easyDict']['sites'][site][cS[0]][cS[1]][cS[2]][cS[3]]

        for k, v in dict1.items():
            for i in v:
                if not i.get(cS[-1]):
                    i[cS[-1]] = []
                if i[policy] == policy_name:
                    i[cS[-1]].append(deepcopy(polVars))
                    break
        
    if 'Grp_' in kwargs['site_group']:
        if kwargs['easyDict']['site_groups'].get(kwargs['site_group']):
            sites = kwargs['easyDict']['site_groups'][kwargs['site_group']]['sites']
            for site in sites:
                site_append(cS, site, polVars)
        else: validating.error_site_group('site_group', **kwargs)
    else: site_append(cS, kwargs['site_group'], polVars)

    # Return Dictionary
    return kwargs['easyDict']

#========================================================
# Function to Append Subtype easyDict Dictionary
#========================================================
def ez_append_arg(polVars, **kwargs):
    class_path   = kwargs['class_path']
    cS = class_path.split(',')
    policy  = kwargs['policy']
    policy_name  = kwargs['policy_name']
    polVars.pop('site_group')
    polVars = ez_remove_empty(polVars)

    def site_append(cS, site, polVars):
        # Assign the Dictionary
        if len(cS) == 3:   dict1 = kwargs['easyDict']['sites'][site][cS[0]][cS[1]]
        elif len(cS) == 4: dict1 = kwargs['easyDict']['sites'][site][cS[0]][cS[1]][cS[2]]
        elif len(cS) == 5: dict1 = kwargs['easyDict']['sites'][site][cS[0]][cS[1]][cS[2]][cS[3]]
        elif len(cS) == 6: dict1 = kwargs['easyDict']['sites'][site][cS[0]][cS[1]][cS[2]][cS[3]][cS[4]]

        for i in dict1:
            if not i.get(cS[-1]):
                i[cS[-1]] = []
            if i[policy] == policy_name:
                i[cS[-1]].extend(polVars[cS[-1]])
                break
        
    if 'Grp_' in kwargs['site_group']:
        if kwargs['easyDict']['site_groups'].get(kwargs['site_group']):
            sites = kwargs['easyDict']['site_groups'][kwargs['site_group']]['sites']
            for site in sites:
                site_append(cS, site, polVars)
        else: validating.error_site_group('site_group', **kwargs)
    else: site_append(cS, kwargs['site_group'], polVars)
    # Return Dictionary
    return kwargs['easyDict']

#========================================================
# Function to Append the easyDict Dictionary
#========================================================
def ez_merge(polVars, **kwargs):
    class_path = kwargs['class_path']
    cS = class_path.split(',')
    policy  = kwargs['policy']
    policy_name  = kwargs['policy_name']
    polVars.pop('site_group')
    polVars = ez_remove_empty(polVars)

    def site_merge(cS, site, polVars):
        if not kwargs['easyDict']['sites'].get(site):
            validating.error_site_group('site_group', **kwargs)

        # Assign the Dictionary
        if len(cS) == 3:   dict1 = kwargs['easyDict']['sites'][site][cS[0]][cS[1]][cS[2]]
        elif len(cS) == 4: dict1 = kwargs['easyDict']['sites'][site][cS[0]][cS[1]][cS[2]][cS[3]]
        elif len(cS) == 5: dict1 = kwargs['easyDict']['sites'][site][cS[0]][cS[1]][cS[2]][cS[4]][cS[5]]

        for i in dict1:
            if i[policy] == policy_name:
                i.update(deepcopy(polVars))
                break

    if 'Grp_' in kwargs['site_group']:
        if kwargs['easyDict']['site_groups'].get(kwargs['site_group']):
            sites = kwargs['easyDict']['site_groups'][kwargs['site_group']]['sites']
            for site in sites:
                site_merge(cS, site, polVars)
        else: validating.error_site_group('site_group', **kwargs)
    else: site_merge(cS, kwargs['site_group'], polVars)
    return kwargs['easyDict']

#========================================================
# Function to Remove Empty Arguments
#========================================================
def ez_remove_empty(polVars):
    pop_list = []
    for k,v in polVars.items():
        if v == None:
            pop_list.append(k)
    for i in pop_list:
        polVars.pop(i)
    return polVars

#========================================================
# Function to Append the Tenant easyDict Dictionary
#========================================================
def ez_tenants_append(polVars, **kwargs):
    class_path = kwargs['class_path']
    cS = class_path.split(',')
    polVars.pop('site_group')
    polVars = ez_remove_empty(polVars)

    def site_append(cS, site, polVars):
        if not kwargs['easyDict']['sites'].get(site):
            validating.error_site_group('site_group', **kwargs)
        tenant_match = False
        tkey = 0
        for i in kwargs['easyDict']['sites'][site]['tenants']:
            if i['name'] == kwargs['tenant']:
                tenant_match = True
                break
            tkey += 1
        if tenant_match == False:
            validating.error_tenant('tenant', **kwargs)

        # Confirm the Key Exists
        if len(cS) == 1:
            if not kwargs['easyDict']['sites'][site]['tenants'][tkey].get(cS[0]):
                kwargs['easyDict']['sites'][site]['tenants'][tkey].update(deepcopy({cS[0]:[]}))
        if len(cS) >= 2:
            if not kwargs['easyDict']['sites'][site]['tenants'][tkey].get(cS[0]):
                kwargs['easyDict']['sites'][site]['tenants'][tkey].update(deepcopy({cS[0]:{}}))
        if len(cS) >= 3:
            if not kwargs['easyDict']['sites'][site]['tenants'][tkey][cS[0]].get(cS[1]):
                kwargs['easyDict']['sites'][site]['tenants'][tkey][cS[0]].update(deepcopy({cS[1]:{}}))
        if len(cS) >= 4:
            if not kwargs['easyDict']['sites'][site]['tenants'][tkey][cS[0]][cS[1]].get(cS[2]):
                kwargs['easyDict']['sites'][site]['tenants'][tkey][cS[0]][cS[1]].update(deepcopy({cS[2]:{}}))
        if len(cS) == 1: cs_count = 0
        elif len(cS) == 2:
            if not kwargs['easyDict']['sites'][site]['tenants'][tkey][cS[0]].get(cS[1]):
                kwargs['easyDict']['sites'][site]['tenants'][tkey][cS[0]].update(deepcopy({cS[1]:[]}))
        elif len(cS) == 3:
            if not kwargs['easyDict']['sites'][site]['tenants'][tkey][cS[0]][cS[1]].get(cS[2]):
                kwargs['easyDict']['sites'][site]['tenants'][tkey][cS[0]][cS[1]].update(deepcopy({cS[2]:[]}))
        elif len(cS) == 4:
            if not kwargs['easyDict']['sites'][site]['tenants'][tkey][cS[0]][cS[1]][cS[2]].get(cS[3]):
                kwargs['easyDict']['sites'][site]['tenants'][tkey][cS[0]][cS[1]][cS[2]].update(deepcopy({cS[3]:[]}))
        
        # append the Dictionary
        if len(cS) ==   1: kwargs['easyDict']['sites'][site]['tenants'][tkey][cS[0]].append(deepcopy(polVars))
        elif len(cS) == 2: kwargs['easyDict']['sites'][site]['tenants'][tkey][cS[0]][cS[1]].append(deepcopy(polVars))
        elif len(cS) == 3: kwargs['easyDict']['sites'][site]['tenants'][tkey][cS[0]][cS[1]][cS[2]].append(deepcopy(polVars))
        elif len(cS) == 4: kwargs['easyDict']['sites'][site]['tenants'][tkey][cS[0]][cS[1]][cS[2]][cS[3]].append(deepcopy(polVars))

    if 'Grp_' in kwargs['site_group']:
        if kwargs['easyDict']['site_groups'].get(kwargs['site_group']):
            sites = kwargs['easyDict']['site_groups'][kwargs['site_group']]['sites']
            for site in sites:
                site_append(cS, site, polVars)
        else: validating.error_site_group('site_group', **kwargs)
    else: site_append(cS, kwargs['site_group'], polVars)
    return kwargs['easyDict']

#========================================================
# Function to Append L3Out easyDict Dictionary
#========================================================
def ez_append_l3out(polVars, **kwargs):
    class_path  = kwargs['class_path']
    cS          = class_path.split(',')
    policy      = kwargs['policy']
    policy_name = kwargs['policy_name']
    polVars = ez_remove_empty(polVars)

    def site_append(cS, site, polVars):
        # Assign the Dictionary
        dict1 = kwargs['easyDict']['sites'][site][cS[0]][cS[1]]

        for i in dict1:
            if i[policy] == policy_name:
                i[cS[-2]][cS[-1]][0].update(deepcopy(polVars))
                break
        
    if 'Grp_' in kwargs['site_group']:
        if kwargs['easyDict']['site_groups'].get(kwargs['site_group']):
            sites = kwargs['easyDict']['site_groups'][kwargs['site_group']]['sites']
            for site in sites:
                site_append(cS, site, polVars)
        else: validating.error_site_group('site_group', **kwargs)
    else: site_append(cS, kwargs['site_group'], polVars)

    # Return Dictionary
    return kwargs['easyDict']

#========================================================
# Function to Append Subtype easyDict Dictionary
#========================================================
def ez_tenants_append_subtype(polVars, **kwargs):
    class_path  = kwargs['class_path']
    cS          = class_path.split(',')
    policy      = kwargs['policy']
    policy_name = kwargs['policy_name']
    polVars.pop('site_group')
    polVars = ez_remove_empty(polVars)

    def site_append(cS, site, polVars):
        tenant_match = False
        tkey = 0
        for i in kwargs['easyDict']['sites'][site]['tenants']:
            if i['name'] == kwargs['tenant']:
                tenant_match = True
                break
            tkey += 1
        if tenant_match == False:
            validating.error_tenant('tenant', **kwargs)

        # Assign the Dictionary
        if len(cS) == 2:   dict1 = kwargs['easyDict']['sites'][site]['tenants'][tkey][cS[0]]
        elif len(cS) == 3: dict1 = kwargs['easyDict']['sites'][site]['tenants'][tkey][cS[0]][cS[1]]
        elif len(cS) == 4: dict1 = kwargs['easyDict']['sites'][site]['tenants'][tkey][cS[0]][cS[1]][cS[2]]
        elif len(cS) == 5: dict1 = kwargs['easyDict']['sites'][site]['tenants'][tkey][cS[0]][cS[1]][cS[2]][cS[3]]

        for i in dict1:
            if not i.get(cS[-1]):
                i[cS[-1]] = []
            if i[policy] == policy_name:
                i[cS[-1]].append(deepcopy(polVars))
                break
        
    if 'Grp_' in kwargs['site_group']:
        if kwargs['easyDict']['site_groups'].get(kwargs['site_group']):
            sites = kwargs['easyDict']['site_groups'][kwargs['site_group']]['sites']
            for site in sites:
                site_append(cS, site, polVars)
        else: validating.error_site_group('site_group', **kwargs)
    else: site_append(cS, kwargs['site_group'], polVars)

    # Return Dictionary
    return kwargs['easyDict']

#========================================================
# Function to Append Subtype easyDict Dictionary
#========================================================
def ez_tenants_append_sub_subtype(polVars, **kwargs):
    class_path  = kwargs['class_path']
    cS          = class_path.split(',')
    policy1      = kwargs['policy1']
    policy_name1 = kwargs['policy_name1']
    policy2      = kwargs['policy2']
    policy_name2 = kwargs['policy_name2']
    polVars.pop('site_group')
    polVars = ez_remove_empty(polVars)

    def site_append(cS, site, polVars):
        tenant_match = False
        tkey = 0
        for i in kwargs['easyDict']['sites'][site]['tenants']:
            if i['name'] == kwargs['tenant']:
                tenant_match = True
                break
            tkey += 1
        if tenant_match == False:
            validating.error_tenant('tenant', **kwargs)

        # Assign the Dictionary
        if policy2 == 'epg_esg_collection_for_vrfs':
            dict1 = kwargs['easyDict']['sites'][site]['tenants'][tkey][cS[0]][cS[1]]
        else:
            if len(cS) == 3:   dict1 = kwargs['easyDict']['sites'][site]['tenants'][tkey][cS[0]]
            elif len(cS) == 4: dict1 = kwargs['easyDict']['sites'][site]['tenants'][tkey][cS[0]][cS[1]]
            elif len(cS) == 5: dict1 = kwargs['easyDict']['sites'][site]['tenants'][tkey][cS[0]][cS[1]][cS[2]]

        itcount = 0
        for item in dict1:
            if item[policy1] == policy_name1:
                if policy2 == 'epg_esg_collection_for_vrfs':
                    item[policy2][cS[-1]].append(deepcopy(polVars))
                    break
                else:
                    icount = 0
                    #print(dict1[itcount][cS[-2]])
                    #exit()
                    for i in dict1[itcount][cS[-2]]:
                        if not i.get(cS[-1]):
                            i[cS[-1]] = []
                        if i[policy2] == policy_name2:
                            dict1[itcount][cS[-2]][icount][cS[-1]].append(deepcopy(polVars))
                            break
                        icount +=1
            itcount += 1
        return kwargs['easyDict'] 
        
    if 'Grp_' in kwargs['site_group']:
        if kwargs['easyDict']['site_groups'].get(kwargs['site_group']):
            sites = kwargs['easyDict']['site_groups'][kwargs['site_group']]['sites']
            for site in sites:
                kwargs['easyDict']  = site_append(cS, site, polVars)
        else: validating.error_site_group('site_group', **kwargs)
    else: kwargs['easyDict']  = site_append(cS, kwargs['site_group'], polVars)

    # Return Dictionary
    return kwargs['easyDict']

#========================================================
# Function to Append the easyDict Dictionary
#========================================================
def ez_update(polVars, **kwargs):
    class_path = kwargs['class_path']
    cS = class_path.split(',')
    polVars.pop('site_group')
    polVars = ez_remove_empty(polVars)

    def site_update(cS, site, polVars):
        if not kwargs['easyDict']['sites'].get(site):
            validating.error_site_group('site_group', **kwargs)

        # Confirm the Key(s) Exists
        if not kwargs['easyDict']['sites'][site].get(cS[0]):
            kwargs['easyDict']['sites'][site].update(deepcopy({cS[0]:{}}))
        if len(cS) >= 2:
            if not kwargs['easyDict']['sites'][site][cS[0]].get(cS[1]):
                kwargs['easyDict']['sites'][site][cS[0]].update(deepcopy({cS[1]:{}}))
        if len(cS) >= 3:
            if not kwargs['easyDict']['sites'][site][cS[0]][cS[1]].get(cS[2]):
                kwargs['easyDict']['sites'][site][cS[0]][cS[1]].update(deepcopy({cS[2]:{}}))
        if len(cS) >= 4:
            if not kwargs['easyDict']['sites'][site][cS[0]][cS[1]][cS[2]].get(cS[3]):
                kwargs['easyDict']['sites'][site][cS[0]][cS[1]][cS[2]].update(deepcopy({cS[3]:{}}))
        if len(cS) == 5:
            if not kwargs['easyDict']['sites'][site][cS[0]][cS[1]][cS[2]][cS[3]].get(cS[4]):
                kwargs['easyDict']['sites'][site][cS[0]][cS[1]][cS[2]][cS[3]].update(deepcopy({cS[4]:{}}))

        # Update the Dictionary
        if len(cS) == 1:   kwargs['easyDict']['sites'][site][cS[0]].update(deepcopy(polVars))
        elif len(cS) == 2:   kwargs['easyDict']['sites'][site][cS[0]][cS[1]].update(deepcopy(polVars))
        elif len(cS) == 3: kwargs['easyDict']['sites'][site][cS[0]][cS[1]][cS[2]].update(deepcopy(polVars))
        elif len(cS) == 4: kwargs['easyDict']['sites'][site][cS[0]][cS[1]][cS[2]][cS[3]].update(deepcopy(polVars))
        elif len(cS) == 5: kwargs['easyDict']['sites'][site][cS[0]][cS[1]][cS[2]][cS[3]][cS[4]].update(deepcopy(polVars))

    if 'Grp_' in kwargs['site_group']:
        if kwargs['easyDict']['site_groups'].get(kwargs['site_group']):
            sites = kwargs['easyDict']['site_groups'][kwargs['site_group']]['sites']
            for site in sites:
                site_update(cS, site, polVars)
        else: validating.error_site_group('site_group', **kwargs)
    else: site_update(cS, kwargs['site_group'], polVars)
    return kwargs['easyDict']

#========================================================
# Function to Append Subtype easyDict Dictionary
#========================================================
def ez_update_subtype(polVars, **kwargs):
    class_path   = kwargs['class_path']
    cS = class_path.split(',')
    policy  = kwargs['policy']
    policy_name  = kwargs['policy_name']
    polVars.pop('site_group')
    polVars = ez_remove_empty(polVars)

    def site_append(cS, site, polVars):
        # Assign the Dictionary
        if len(cS) == 3: dict1 = kwargs['easyDict']['sites'][site][cS[0]]
        elif len(cS) == 4: dict1 = kwargs['easyDict']['sites'][site][cS[0]][cS[1]]
        elif len(cS) == 5: dict1 = kwargs['easyDict']['sites'][site][cS[0]][cS[1]][cS[2]]
        elif len(cS) == 6: dict1 = kwargs['easyDict']['sites'][site][cS[0]][cS[1]][cS[2]][cS[3]]

        for k, v in dict1.items():
            for i in v:
                if not i.get(cS[-1]):
                    i[cS[-1]] = {}
                if i[policy] == policy_name:
                    i[cS[-1]].update(deepcopy(polVars))
                    break
        
    if 'Grp_' in kwargs['site_group']:
        if kwargs['easyDict']['site_groups'].get(kwargs['site_group']):
            sites = kwargs['easyDict']['site_groups'][kwargs['site_group']]['sites']
            for site in sites:
                site_append(cS, site, polVars)
        else: validating.error_site_group('site_group', **kwargs)
    else: site_append(cS, kwargs['site_group'], polVars)

    # Return Dictionary
    return kwargs['easyDict']

#========================================================
# Function to find the Keys for each Worksheet
#========================================================
def findKeys(ws, func_regex):
    func_list = OrderedSet()
    for i in ws.rows:
        if any(i):
            if re.search(func_regex, str(i[0].value)):
                func_list.add(str(i[0].value))
    return func_list

#========================================================
# Function to Create Terraform auto.tfvars files
#========================================================
def findVars(ws, func, rows, count):
    var_list = []
    var_dict = {}
    for i in range(1, rows + 1):
        if (ws.cell(row=i, column=1)).value == func:
            try:
                for x in range(2, 36):
                    if (ws.cell(row=i - 1, column=x)).value:
                        var_list.append(str(ws.cell(row=i - 1, column=x).value))
                    else: x += 1
            except Exception as e:
                e = e
                pass
            break
    vcount = 1
    while vcount <= count:
        var_dict[vcount] = {}
        var_count = 0
        for z in var_list:
            var_dict[vcount][z] = ws.cell(row=i + vcount - 1, column=2 + var_count).value
            var_count += 1
        var_dict[vcount]['row'] = i + vcount - 1
        vcount += 1
    return var_dict

#========================================================
# Function to Merge Easy ACI Repository to Dest Folder
#========================================================
def get_latest_versions(args, easyDict):

    url_list = [
        'https://github.com/CiscoDevNet/terraform-provider-aci/tags/',
        'https://github.com/CiscoDevNet/terraform-provider-mso/tags/',
        'https://github.com/hashicorp/terraform/tags',
        'https://github.com/netascode/terraform-provider-utils/tags/'
    ]
    if args.skip_version_check == True:
        easyDict['latest_versions']['aci_provider_version'] = '2.6.1'
        easyDict['latest_versions']['ndo_provider_version'] = '0.8.1'
        easyDict['latest_versions']['terraform_version'] = '1.3.6'
        easyDict['latest_versions']['utils_provider_version'] = '0.2.4'
    else:
        for url in url_list:
            # Get the Latest Release Tag for the Provider
            r = requests.get(url, stream=True)
            repoVer = 'BLANK'
            stringMatch = False
            while stringMatch == False:
                for line in r.iter_lines():
                    toString = line.decode("utf-8")
                    if re.search(r'/releases/tag/v(\d+\.\d+\.\d+)\"', toString):
                        repoVer = re.search('/releases/tag/v(\d+\.\d+\.\d+)', toString).group(1)
                        break
                stringMatch = True
            
            # Make sure the latest_versions Key exists
            if easyDict.get('latest_versions') == None:
                easyDict['latest_versions'] = {}
            
            # Set Provider Version
            if   'terraform-provider-aci' in url:
                easyDict['latest_versions']['aci_provider_version'] = repoVer
            elif   'terraform-provider-mso' in url:
                easyDict['latest_versions']['ndo_provider_version'] = repoVer
            elif 'netascode' in url:
                easyDict['latest_versions']['utils_provider_version'] = repoVer
            else: easyDict['latest_versions']['terraform_version'] = repoVer
    
    # Return kwargs
    return easyDict

#========================================================
# Function to Check the Git Status of the Folders
#========================================================
def git_base_repo(args, wb):
    repoName = args.dir
    if not os.path.isdir(repoName):
        baseRepo = git.Repo.init(repoName, bare=True, mkdir=True)
        assert baseRepo.bare
    else:
        try: 
            baseRepo = git.Repo.init(repoName)
        except:
            baseRepo = git.Repo.init(repoName, bare=True, mkdir=True)
    base_dir = args.dir
    with baseRepo.config_reader() as git_config:
        try:
            git_config.get_value('user', 'email')
            git_config.get_value('user', 'name')
        except:
            valid = False
            while valid == False:
                polVars = {}
                polVars["Description"] = f'Git Email Configuration. i.e. user@example.com'
                polVars["varInput"] = f'What is your Git email?'
                polVars["minimum"] = 5
                polVars["maximum"] = 128
                polVars["pattern"] = '[\\S]+'
                polVars["varName"] = 'Git Email'
                repoName = varStringLoop(**polVars)
                valid = True

            valid = False
            while valid == False:
                polVars = {}
                polVars["Description"] = f'Git Username Configuration. i.e. user'
                polVars["varInput"] = f'What is your Git Username?'
                polVars["minimum"] = 5
                polVars["maximum"] = 64
                polVars["pattern"] = '[\\S]+'
                polVars["varName"] = 'Git Email'
                repoName = varStringLoop(**polVars)
                valid = True
    result = subprocess.Popen(['python3', '-m', 'git_status_checker', base_dir], stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
    while(True):
        # returns None while subprocess is running
        retcode = result.poll()
        line = result.stdout.readline()
        line = line.decode('utf-8')
        if 'this operation must be run in a work tree' in line:
            arg_folder = os.path.basename(os.path.normpath(args.dir))
            git_user = git_config.get_value('user', 'name')
            defaultUrl = f"github.com/{git_user}/{arg_folder}.git"
            valid = False
            while valid == False:
                polVars = {}
                polVars["varDefault"] = defaultUrl
                polVars["Description"] = f'The Destination Directory is not currently a Git Repository.'
                polVars["varInput"] = f'What is the Git URL (without https://) for "{args.dir}"? [{defaultUrl}]'
                polVars["minimum"] = 5
                polVars["maximum"] = 64
                polVars["pattern"] = '[\\S]+'
                polVars["varName"] = 'Git URL'
                gitUrl = varStringLoop(**polVars)
                valid = True
            ws = wb['Sites']
            kwargs = {'ws':ws, 'row_num':0, 'url':gitUrl}
            validating.url('url', **kwargs)
            gitUrl = f'https://{gitUrl}'
            baseRepo.create_remote('origin', gitUrl)
            baseRepo.remotes.origin.push('master:master')
            break
        elif 'has outstanding commits' in line:
            break
        elif 'has outstanding pushes' in line:
            break
        
        print(baseRepo.is_dirty(untracked_files=True))
        try:
            baseRepo.remotes.origin.push('master:master')
        except Exception as e:
            print(f'Script Errored {e}')
            exit()
        break

#========================================================
# Function to Check the Git Status of the Folders
#========================================================
def git_check_status(args, site_names, site_directories):
    gitRepo = True
    args.git_check = True
    try:
        baseRepo = git.Repo(args.dir)
    except git.exc.GitError as e:
        print(f'\nError: {args.dir} is not a Git Repository\n Error {e}')
        args.git_check = False
        gitRepo = False
    if gitRepo == True:
        untrackedFiles = baseRepo.untracked_files
        random_folders = []
        modified = baseRepo.git.status()
        modifiedList = [y for y in (x.strip() for x in modified.splitlines()) if y]
        for site in site_names:
            for x in site_directories:
                if site in x:
                    for line in modifiedList:
                        if not x in random_folders and site in line: random_folders.append(x)  
                    for line in untrackedFiles:
                        if not x in random_folders and site in line: random_folders.append(x)
        strict_folders = list(set(random_folders))
        strict_folders.sort()
        if not len(strict_folders) > 0:
            print(f'\n-----------------------------------------------------------------------------\n')
            print(f'   There were no uncommitted changes in the environment.')
            print(f'   Proceedures Complete!!! Closing Environment and Exiting Script.')
            print(f'\n-----------------------------------------------------------------------------\n')
            exit()
        return args, strict_folders, True
    else: return args, [], False


#========================================================
# Function to Create Interface Selector Workbooks
#========================================================
def interface_selector_workbook(polVars, **kwargs):
    # Set the Workbook var
    wb_sw = kwargs['wb_sw']
    site_group = kwargs['site_group']

    # Use Switch_Type to Determine the Number of ports on the switch
    modules,port_count = switch_model_ports(kwargs['row_num'], polVars['switch_model'])

    # Get the Interface Policy Groups from EasyDict
    if polVars['node_type'] == 'spine':
        pg_list = ['spine_pg']
    else:
        pg_list = ['access', 'breakout', 'bundle']
    switch_pgs = {}
    for pgroup in pg_list:
        switch_pgs[pgroup] = []
        if 'spine_pg' in pgroup:
            pgroups = kwargs['easyDict']['sites'][site_group]['access']['interfaces']['spine']['policy_groups']
        else:
            pgroups = kwargs['easyDict']['sites'][site_group]['access']['interfaces']['leaf']['policy_groups'][pgroup]
        for i in pgroups:
            if pgroup == 'bundle':
                if i.get('names'):
                    switch_pgs[pgroup].extend(i['names'])
                else:
                    kwargs.update({'template_name':i['template_name']})
                    validating.error_bundle_names('template_name', **kwargs)
            else:
                switch_pgs[pgroup].append(i['name'])

    # Sort the Policy Group List and Convert to a string
    for pgroup in pg_list:
        switch_pgs[pgroup].sort()

    if not 'formulas' in wb_sw.sheetnames:
        ws_sw = wb_sw.create_sheet(title = 'formulas')
        ws_sw.column_dimensions['A'].width = 30
        ws_sw.column_dimensions['B'].width = 30
        ws_sw.column_dimensions['C'].width = 30
        ws_sw.column_dimensions['D'].width = 30
        data = ['access', 'breakout', 'bundle', 'spine_pg']
        ws_sw.append(data)
        for cell in ws_sw['1:1']:
            cell.style = 'Heading 1'
        wb_sw.save(kwargs['excel_workbook'])

    ws_sw = wb_sw['formulas']
    for pgroup in pg_list:
        if pgroup == 'access': x = 1
        elif pgroup == 'breakout': x = 2
        elif pgroup == 'bundle': x = 3
        elif polVars['node_type'] == 'spine': x = 4
        if len(switch_pgs[pgroup]) > 0:
            row_start = 2
            row_end = len(switch_pgs[pgroup]) + row_start - 1
            sw_pgs_count = 0
            for row_num in range(row_start, row_end + 1):
                rc = f'{row_num}:{row_num}'
                for cell in ws_sw[rc]:
                    if row_num % 2 == 0:
                        cell.style = 'ws_odd'
                    else:
                        cell.style = 'ws_even'
                ws_sw.cell(row=row_num, column=x, value=switch_pgs[pgroup][sw_pgs_count])
                sw_pgs_count += 1

        wb_sw.save(kwargs['excel_workbook'])
        last_row = len(switch_pgs[pgroup]) + 1
        dnames = ['DSCP', 'leaf', 'spine', 'spine_modules', 'spine_type', 'switch_role', 'tag', 'Time_Zone']
        for dname in dnames:
            if dname in wb_sw.defined_names:
                if type(wb_sw.defined_names) is dict:
                    wb_sw.defined_names.pop(dname)
                elif type(wb_sw.defined_names) is list:
                    wb_sw.defined_names.delete(dname)
        if polVars['node_type'] == 'spine':
            new_range = openpyxl.workbook.defined_name.DefinedName('spine_pg',attr_text=f"formulas!$D$2:$D{last_row}")
            if not 'spine_pg' in wb_sw.defined_names:
                # wb_sw.defined_names.delete('spine_pg')
                if type(wb_sw.defined_names) is dict:
                    wb_sw.defined_names.update(new_range)
                elif type(wb_sw.defined_names) is list:
                    wb_sw.defined_names.append(new_range)
        elif pgroup == 'access':
            new_range = openpyxl.workbook.defined_name.DefinedName('access',attr_text=f"formulas!$A$2:$A{last_row}")
            if not 'access' in wb_sw.defined_names:
                # wb_sw.defined_names.delete('access')
                if type(wb_sw.defined_names) is dict:
                    wb_sw.defined_names.update(new_range)
                elif type(wb_sw.defined_names) is list:
                    wb_sw.defined_names.append(new_range)
        elif pgroup == 'breakout':
            new_range = openpyxl.workbook.defined_name.DefinedName('breakout',attr_text=f"formulas!$B$2:$B{last_row}")
            if not 'breakout' in wb_sw.defined_names:
                # wb_sw.defined_names.delete('breakout')
                if type(wb_sw.defined_names) is dict:
                    wb_sw.defined_names.update(new_range)
                elif type(wb_sw.defined_names) is list:
                    wb_sw.defined_names.append(new_range)
        elif pgroup == 'bundle':
            new_range = openpyxl.workbook.defined_name.DefinedName('bundle',attr_text=f"formulas!$C$2:$C{last_row}")
            if not 'bundle' in wb_sw.defined_names:
                # wb_sw.defined_names.delete('bundle')
                if type(wb_sw.defined_names) is dict:
                    wb_sw.defined_names.update(new_range)
                elif type(wb_sw.defined_names) is list:
                    wb_sw.defined_names.append(new_range)
        wb_sw.save(kwargs['excel_workbook'])

    # Check if there is a Worksheet for the Switch Already
    if not polVars['switch_name'] in wb_sw.sheetnames:
        ws_sw = wb_sw.create_sheet(title = polVars['switch_name'])
        ws_sw = wb_sw[polVars['switch_name']]
        ws_sw.column_dimensions['A'].width = 15
        ws_sw.column_dimensions['B'].width = 15
        ws_sw.column_dimensions['C'].width = 10
        ws_sw.column_dimensions['D'].width = 10
        ws_sw.column_dimensions['E'].width = 20
        ws_sw.column_dimensions['F'].width = 20
        ws_sw.column_dimensions['G'].width = 20
        ws_sw.column_dimensions['H'].width = 20
        ws_sw.column_dimensions['I'].width = 20
        ws_sw.column_dimensions['J'].width = 20
        ws_sw.column_dimensions['K'].width = 20
        ws_sw.column_dimensions['L'].width = 25
        ws_sw.column_dimensions['M'].width = 30
        dv1 = DataValidation(type="list", formula1='"intf_selector"', allow_blank=True)
        if polVars['node_type'] == 'spine':
            dv2 = DataValidation(type="list", formula1='"spine_pg"', allow_blank=True)
        else:
            dv2 = DataValidation(type="list", formula1='"access,breakout,bundle"', allow_blank=True)
        dv4 = DataValidation(type="list", formula1='"access,aaep_encap,trunk"', allow_blank=True)
        ws_sw.add_data_validation(dv1)
        ws_sw.add_data_validation(dv2)
        ws_sw.add_data_validation(dv4)
        ws_header = '%s Interface Selectors' % (kwargs['switch_name'])
        data = [ws_header]
        ws_sw.append(data)
        ws_sw.merge_cells('A1:M1')
        for cell in ws_sw['1:1']:
            cell.style = 'Heading 1'
        data = ['','Notes:']
        ws_sw.append(data)
        ws_sw.merge_cells('B2:M2')
        for cell in ws_sw['2:2']:
            cell.style = 'Heading 2'
        data = [
            'Type','site_group','pod_id','node_id','interface_profile','interface_selector','interface','policy_group_type',
            'policy_group','description','switchport_mode','access_or_native_vlan','trunk_port_allowed_vlans'
        ]
        ws_sw.append(data)
        for cell in ws_sw['3:3']:
            cell.style = 'Heading 3'

        ws_sw_row_count = 4
        polVars['dv1'] = dv1
        polVars['dv2'] = dv2
        polVars['dv4'] = dv4
        polVars['port_count'] = port_count
        sw_type = str(polVars['switch_model'])
        if re.search('^(95[0-1][4-8])', sw_type):
            if kwargs['easyDict']['sites'][site_group]['switch'].get('spine_modules'):
                modCount = 0
                for i in kwargs['easyDict']['sites'][site_group]['switch']['spine_modules']:
                    modDict = {}
                    if str(polVars['node_id']) in i['node_list']:
                        modDict = i
                        modCount += 1
                if modCount == 0:
                    print(f"Error, Could not find the Module list for spine {polVars['node_id']}")
                    exit()
                
                for k, v in modDict.items():
                    if 'module' in k:
                        polVars['module'] = k.split('_')[1]
                        if re.search('^X97', v):
                            polVars['port_count'] = spine_module_port_count(v)
                            ws_sw_row_count = create_selector(ws_sw, ws_sw_row_count, **polVars)
        else:
            polVars['module'] = 1
            ws_sw_row_count = create_selector(ws_sw, ws_sw_row_count, **polVars)

        # Save the Workbook
        wb_sw.save(kwargs['excel_workbook'])
        wb_sw.close()

#========================================================
# Function to Merge Easy ACI Repository to Dest Folder
#========================================================
def merge_easy_aci_repository(args, easy_jsonData, **easyDict):
    # Setup Operating Environment
    baseRepo = args.dir
    opSystem = platform.system()
    tfe_dir = 'tfe_modules'
    if opSystem == 'Windows': path_sep = '\\'
    else: path_sep = '/'
    tfe_modules = f'{tfe_dir}{path_sep}modules'
    git_url = "https://github.com/terraform-cisco-modules/easy-aci-complete"
    if not os.path.isdir(tfe_dir):
        print('creating directory')
        os.mkdir(tfe_dir)
        git.Repo.clone_from(git_url, tfe_dir)
    if not os.path.isfile(os.path.join(tfe_dir, 'README.md')):
        print('cloning directory')
        git.Repo.clone_from(git_url, tfe_dir)
    else:
        print('pulling directory')
        g = git.cmd.Git(tfe_dir)
        g.pull()

    # Get All sub-folders from tfDir
    site_list = list(easyDict['sites'].keys())
    site_directories = []
    site_names = []
    for item in site_list:
        site_name = easyDict['sites'][item]['site_settings']['site_name']
        site_names.append(site_name)
        site_dir = os.path.join(baseRepo, site_name)
        site_directories.append(site_dir)
        default_dir = os.path.join(baseRepo, site_name, 'defaults')
        if not os.path.isdir(site_dir):
            os.mkdir(site_dir)
        if not os.path.isdir(default_dir):
            os.mkdir(default_dir)
        # Now Loop over the folders and merge the module files
        for folder in [site_name, 'defaults']:
            if folder == 'defaults':
                dest_dir = os.path.join(baseRepo, site_name, folder)
                src_dir = os.path.join(tfe_dir, 'defaults')
            else:
                dest_dir = os.path.join(baseRepo, site_name)
                src_dir = os.path.join(tfe_dir)
            copy_files = os.listdir(src_dir)
            for fname in copy_files:
                if not os.path.isdir(os.path.join(src_dir, fname)):
                    shutil.copy2(os.path.join(src_dir, fname), dest_dir)
        terraform_fmt(site_dir)
    # Return Site Names and Site Directories
    return site_names, site_directories

#========================================================
# Function to GET to the NDO API
#========================================================
def ndo_api(ndo, method, cookies, uri, section=''):
    s = requests.Session()
    r = ''
    while r == '':
        try:
            if method == 'get':
                r = s.get('https://{}/{}'.format(ndo, uri), cookies=cookies, verify=False)
            status = r.status_code
        except requests.exceptions.ConnectionError as e:
            print("Connection error, pausing before retrying. Error: {}"
                .format(e))
            time.sleep(5)
        except Exception as e:
            print("Method {} failed. Exception: {}".format(section[:-5], e))
            status = 666
            return(status)
    if print_response_always: print(r.text)
    if status != 200 and print_response_on_fail: print(r.text)
    return r

#========================================================
# Function to validate input for each method
#========================================================
def process_kwargs(jsonData, **kwargs):
    # Validate User Input
    validate_args(jsonData, **kwargs)
    
    error_count = 0
    error_list = []
    optional_args = jsonData['optional_args']
    required_args = jsonData['required_args']
    row_num = kwargs["row_num"]
    ws = kwargs["ws"]
    for item in required_args:
        if item not in kwargs.keys():
            error_count =+ 1
            error_list += [item]
    if error_count > 0:
        print(f'\n\n***Begin ERROR ***\n\nError on Worksheet {ws.title} row {row_num}\n - The Following'\
            f' REQUIRED Key(s) Were Not Found in kwargs: "{error_list}"\n\n****End ERROR****\n')
        exit()
        #error_ = f'\n\n***Begin ERROR ***\n\nError on Worksheet {ws.title} row {row_num}\n - The Following REQUIRED Key(s) Were Not Found in kwargs: "{error_list}"\n\n****End ERROR****\n'
        #raise InsufficientArgs(error_)

    error_count = 0
    error_list = []
    for item in optional_args:
        if item not in kwargs.keys():
            error_count =+ 1
            error_list += [item]
    if error_count > 0:
        print(f'\n\n***Begin ERROR***\n\nError on Worksheet {ws.title} row {row_num}\n - The Following'\
            f' Optional Key(s) Were Not Found in kwargs: "{error_list}"\n\n****End ERROR****\n')
        exit()
        #error_ = f'\n\n***Begin ERROR***\n\nError on Worksheet {ws.title} row {row_num}\n - The Following Optional Key(s) Were Not Found in kwargs: "{error_list}"\n\n****End ERROR****\n'
        #raise InsufficientArgs(error_)

    # Load all required args values from kwargs
    error_count = 0
    error_list = []
    for item in kwargs:
        if item in required_args.keys():
            required_args[item] = kwargs[item]
            if required_args[item] == None:
                error_count =+ 1
                error_list += [item]
            elif required_args[item] == 'false':
                required_args[item] = False
            elif required_args[item] == 'true':
                required_args[item] = True

    if error_count > 0:
        print(f'\n\n***Begin ERROR***\n\nError on Worksheet {ws.title} row {row_num}\n - The Following'\
            f' REQUIRED Key(s) Argument(s) are Blank:\nPlease Validate "{error_list}"\n\n****End ERROR****\n')
        exit()
        #error_ = 
        #raise InsufficientArgs(error_)

    for item in kwargs:
        if item in optional_args.keys():
            optional_args[item] = kwargs[item]
            if optional_args[item] == 'false':
                optional_args[item] = False
            elif optional_args[item] == 'true':
                optional_args[item] = True

    # Combine option and required dicts for Jinja template render
    polVars = {**required_args, **optional_args}

    if kwargs['easyDict']['remove_default_args'] == True:
        Dicts = deepcopy(polVars)
        for k,v in Dicts.items():
            if jsonData.get(k):
                if not jsonData[k].get('default') == None:
                    if v == jsonData[k]['default']:
                        polVars.pop(k)
    return(polVars)

#========================================================
# Add Static Port Bindings to Bridge Domains
#========================================================
def process_workbook(polVars, **kwargs):
    row_num = kwargs['row_num']
    ws = kwargs['ws']
    def process_site(siteDict, polVars, **kwargs):
        # Create polVars for Site_Name and APIC_URL
        polVars['site_name'] =  siteDict['Site_Name']
        polVars['site_group'] = siteDict['site_group']
        polVars['controller'] =   siteDict['controller']

        # Pull in the Site Workbook
        excel_workbook = '%s_intf_selectors.xlsx' % (polVars['site_name'])
        try:
            kwargs['wb_sw'] = load_workbook(excel_workbook)
        except Exception as e:
            print(f"Something went wrong while opening the workbook - {excel_workbook}... ABORT!")
            sys.exit(e)

        # Process the Interface Selectors for Static Port Paths
        create_static_paths(polVars, **kwargs)

    if re.search('Grp_[A-F]', polVars['site_group']):
        site_group = kwargs['easyDict']['sites']['site_groups'][kwargs['site_group']][0]
        for site in site_group['sites']:
            siteDict = kwargs['easyDict']['sites']['site_settings'][site][0]
            process_site(siteDict, polVars, **kwargs)
    elif re.search(r'\d+', polVars['Site_Group']):
        siteDict = kwargs['easyDict']['sites']['site_settings'][kwargs['site_group']][0]
        process_site(siteDict, polVars, **kwargs)
    else:
        print(f"\n-----------------------------------------------------------------------------\n")
        print(f"   Error on Worksheet {ws.title}, Row {row_num} Site_Group, value {polVars['Site_Group']}.")
        print(f"   Unable to Determine if this is a Single or Group of Site(s).  Exiting....")
        print(f"\n-----------------------------------------------------------------------------\n")
        exit()

#========================================================
# Function to Read Excel Workbook Data
#========================================================
def read_in(excel_workbook):
    try:
        wb = load_workbook(excel_workbook)
    except Exception as e:
        print(f"Something went wrong while opening the workbook - {excel_workbook}... ABORT!")
        sys.exit(e)
    return wb

#========================================================
# Function to Read the Worksheet and Create Templates
#========================================================
def read_worksheet(class_init, class_folder, easyDict, easy_jsonData, func_regex, wb, ws):
    rows = ws.max_row
    func_list = findKeys(ws, func_regex)
    stdout_log(ws, None, 'begin')
    for func in func_list:
        count = countKeys(ws, func)
        var_dict = findVars(ws, func, rows, count)
        for pos in var_dict:
            row_num = var_dict[pos]['row']
            del var_dict[pos]['row']
            for x in list(var_dict[pos].keys()):
                if var_dict[pos][x] == '':
                    del var_dict[pos][x]
            stdout_log(ws, row_num, 'begin')
            var_dict[pos].update(
                {
                    'class_folder':class_folder,
                    'easyDict':easyDict,
                    'easy_jsonData':easy_jsonData,
                    'row_num':row_num,
                    'wb':wb,
                    'ws':ws
                }
            )
            easyDict = eval(f"{class_init}(class_folder).{func}(**var_dict[pos])")
    
    stdout_log(ws, None, 'end')
    # Return the easyDict
    return easyDict

#========================================================
# Function to loop through site_groups for sensitve vars
#========================================================
def sensitive_var_site_group(**kwargs):
    site_group = kwargs['site_group']
    sensitive_var = kwargs['Variable']

    # Loop Through Site Groups to confirm Sensitive Variable in the Environment
    if re.search('Grp_[A-F]', site_group):
       siteGroup = kwargs['easyDict']['site_groups'][site_group]
       for site in siteGroup['sites']:
            kwargs['easyDict']['sites'][site]['sensitive_vars'].append(sensitive_var)
            siteDict = kwargs['easyDict']['sites'][site]['site_settings']
            if siteDict['run_location'] == 'local' or siteDict['configure_terraform_cloud'] == 'true':
                sensitive_var_value(**kwargs)
    else:
        kwargs['easyDict']['sites'][site_group]['sensitive_vars'].append(sensitive_var)
        siteDict = kwargs['easyDict']['sites'][site_group]['site_settings']
        if siteDict['run_location'] == 'local' or siteDict['configure_terraform_cloud'] == 'true':
            sensitive_var_value(**kwargs)
    return kwargs['easyDict']

#========================================================
# Function to add sensitive_var to Environment
#========================================================
def sensitive_var_value(**kwargs):
    sensitive_var = 'TF_VAR_%s' % (kwargs['Variable'])
    # -------------------------------------------------------------------------------------------------------------------------
    # Check to see if the Variable is already set in the Environment, and if not prompt the user for Input.
    #--------------------------------------------------------------------------------------------------------------------------
    if os.environ.get(sensitive_var) is None:
        print(f"\n----------------------------------------------------------------------------------\n")
        print(f"  The Script did not find {sensitive_var} as an 'environment' variable.")
        print(f"  To not be prompted for the value of {kwargs['Variable']} each time")
        print(f"  add the following to your local environemnt:\n")
        print(f"   - export {sensitive_var}='{kwargs['Variable']}_value'")
        print(f"\n----------------------------------------------------------------------------------\n")

    if os.environ.get(sensitive_var) is None:
        valid = False
        while valid == False:
            varValue = input('press enter to continue: ')
            if varValue == '':
                valid = True

        valid = False
        while valid == False:
            if kwargs.get('Multi_Line_Input'):
                print(f'Enter the value for {kwargs["Variable"]}:')
                lines = []
                while True:
                    # line = input('')
                    line = stdiomask.getpass(prompt='')
                    if line:
                        lines.append(line)
                    else:
                        break
                if not re.search('(certificate|private_key)', sensitive_var):
                    secure_value = '\\n'.join(lines)
                else:
                    secure_value = '\n'.join(lines)
            else:
                valid_pass = False
                while valid_pass == False:
                    password1 = stdiomask.getpass(prompt=f'Enter the value for {kwargs["Variable"]}: ')
                    password2 = stdiomask.getpass(prompt=f'Re-Enter the value for {kwargs["Variable"]}: ')
                    if password1 == password2:
                        secure_value = password1
                        valid_pass = True
                    else:
                        print('!!!Error!!! Sensitive Values did not match.  Please re-enter...')

            # Validate Sensitive Passwords
            cert_regex = re.compile(r'^\-{5}BEGIN (CERTIFICATE|PRIVATE KEY)\-{5}.*\-{5}END (CERTIFICATE|PRIVATE KEY)\-{5}$')
            if re.search('(certificate|certName|private_key|privateKey)', sensitive_var):
                if not re.search(cert_regex, secure_value):
                    valid = True
                else:
                    print(f'\n-------------------------------------------------------------------------------------------\n')
                    print(f'    Error!!! Invalid Value for the {sensitive_var}.  Please re-enter the {sensitive_var}.')
                    print(f'\n-------------------------------------------------------------------------------------------\n')
            elif re.search('(apikey|secretkey)', sensitive_var):
                if not sensitive_var == '':
                    valid = True
            else:
                if 'aes_passphrase' in sensitive_var:
                    sKey = 'aes_passphrase'
                    varTitle = 'Global AES Phassphrase'
                elif 'bgp_password' in sensitive_var:
                    sKey = 'password'
                    varTitle = 'BGP Password'
                elif 'apicPass' in sensitive_var:
                    sKey = 'password'
                    varTitle = 'APIC User Password.'
                elif 'eigrp_key' in sensitive_var:
                    sKey = 'eigrp_key'
                    varTitle = 'EIGRP Key.'
                elif 'ndoPass' in sensitive_var:
                    sKey = 'password'
                    varTitle = 'Nexus Dashboard Orchestrator User Password.'
                elif 'ntp_key' in sensitive_var:
                    sKey = 'key'
                    varTitle = 'NTP Key'
                elif 'ospf_key' in sensitive_var:
                    sKey = 'ospf_key'
                    varTitle = 'OSPF Authentication Password.'
                elif 'radius_key' in sensitive_var:
                    sKey = 'radius_key'
                    varTitle = 'The RADIUS shared secret cannot contain backslashes, space, or hashtag "#".'
                elif 'radius_monitoring_password' in sensitive_var:
                    sKey = 'radius_monitoring_password'
                    varTitle = 'RADIUS Monitoring Password.'
                elif 'remote_password' in sensitive_var:
                    sKey = 'remote_password'
                    varTitle = 'The Remote Host password.'
                elif re.search('snmp_(authorization|privacy)_key', sensitive_var):
                    sKey = 'snmp_key'
                    x = sensitive_var.split('_')
                    varType = '%s %s' % (x[0].capitalize(), x[1].capitalize())
                    varTitle = f'{varType}'
                elif 'snmp_community' in sensitive_var:
                    sKey = 'snmp_community'
                    varTitle = 'The Community may only contain letters, numbers and the special characters of \
                    underscore (_), hyphen (-), or period (.). The Community may not contain the @ symbol.'
                elif 'smtp_password' in sensitive_var:
                    sKey = 'smtp_password'
                    varTitle = 'Smart CallHome SMTP Server Password'
                elif 'tacacs_key' in sensitive_var:
                    sKey = 'tacacs_key'
                    varTitle = 'The TACACS+ shared secret cannot contain backslashes, space, or hashtag "#".'
                elif 'tacacs_monitoring_password' in sensitive_var:
                    sKey = 'tacacs_monitoring_password'
                    varTitle = 'TACACS+ Monitoring Password.'
                elif 'vmm_password' in sensitive_var:
                    sKey = 'vmm_password'
                    varTitle = 'Virtual Networking Controller Password.'
                else:
                    print(sensitive_var)
                    print('Could not Match Sensitive Value Type')
                    exit()
                minimum = kwargs['jsonData'][sKey]['minimum']
                maximum = kwargs['jsonData'][sKey]['maximum']
                pattern = kwargs['jsonData'][sKey]['pattern']
                valid = validating.length_and_regex_sensitive(pattern, varTitle, secure_value, minimum, maximum)

        # Add the Variable to the Environment
        os.environ[sensitive_var] = '%s' % (secure_value)
        var_value = secure_value

    else:
        # Add the Variable to the Environment
        if kwargs.get('Multi_Line_Input'):
            var_value = os.environ.get(sensitive_var)
            var_value = var_value.replace('\n', '\\n')
        else:
            var_value = os.environ.get(sensitive_var)

    return var_value

#========================================================
# Function to Define stdout_log output
#========================================================
def stdout_log(ws, row_num, spot):
    if log_level == 0:
        return
    elif ((log_level == (1) or log_level == (2)) and
            (ws) and (row_num is None)) and spot == 'begin':
        print(f'-----------------------------------------------------------------------------\n')
        print(f'   Begin Worksheet "{ws.title}" evaluation...')
        print(f'\n-----------------------------------------------------------------------------\n')
    elif (log_level == (1) or log_level == (2)) and spot == 'end':
        print(f'\n-----------------------------------------------------------------------------\n')
        print(f'   Completed Worksheet "{ws.title}" evaluation...')
        print(f'\n-----------------------------------------------------------------------------')
    elif log_level == (2) and (ws) and (row_num is not None):
        if re.fullmatch('[0-9]', str(row_num)):
            print(f'    - Evaluating Row   {row_num}...')
        elif re.fullmatch('[0-9][0-9]',  str(row_num)):
            print(f'    - Evaluating Row  {row_num}...')
        elif re.fullmatch('[0-9][0-9][0-9]',  str(row_num)):
            print(f'    - Evaluating Row {row_num}...')
    else:
        return

#========================================================
# Function to Determine Port count from Switch Model
#========================================================
def switch_model_ports(row_num, switch_type):
    modules = ''
    switch_type = str(switch_type)
    if re.search('^9396', switch_type):
        modules = '2'
        port_count = '48'
    elif re.search('^93', switch_type):
        modules = '1'

    if re.search('^9316', switch_type):
        port_count = '16'
    elif re.search('^(93120)', switch_type):
        port_count = '102'
    elif re.search('^(93108|93120|93216|93360)', switch_type):
        port_count = '108'
    elif re.search('^(93180|93240|9348|9396)', switch_type):
        port_count = '54'
    elif re.search('^(93240)', switch_type):
        port_count = '60'
    elif re.search('^9332', switch_type):
        port_count = '34'
    elif re.search('^(9336|93600)', switch_type):
        port_count = '36'
    elif re.search('^9364C-GX', switch_type):
        port_count = '64'
    elif re.search('^9364', switch_type):
        port_count = '66'
    elif re.search('^95', switch_type):
        port_count = '36'
        if switch_type == '9504':
            modules = '4'
        elif switch_type == '9508':
            modules = '8'
        elif switch_type == '9516':
            modules = '16'
        else:
            print(f'\n-----------------------------------------------------------------------------\n')
            print(f'   Error on Row {row_num}.  Unknown Switch Model {switch_type}')
            print(f'   Please verify Input Information.  Exiting....')
            print(f'\n-----------------------------------------------------------------------------\n')
            exit()
    else:
        print(f'\n-----------------------------------------------------------------------------\n')
        print(f'   Error on Row {row_num}.  Unknown Switch Model {switch_type}')
        print(f'   Please verify Input Information.  Exiting....')
        print(f'\n-----------------------------------------------------------------------------\n')
        exit()
    return modules,port_count

#========================================================
# Function to Determine Port Count on Modules
#========================================================
def spine_module_port_count(module_type):
    if re.search('X9716D-GX', module_type):
        port_count = '16'
    elif re.search('X9732C-EX', module_type):
        port_count = '32'
    elif re.search('X9736', module_type):
        port_count = '36'
    return port_count

#========================================================
# Function for Terraform Cloud API
#========================================================
def tfc_api(url, method, payload, site_header, section=''):
    r = ''
    while r == '':
        try:
            if method == 'get':
                r = requests.get(url, headers=site_header)
            elif method == 'patch':
                r = requests.patch(url, data=payload, headers=site_header)
            elif method == 'post':
                r = requests.post(url, data=payload, headers=site_header)

            # Use this for Troubleshooting
            if print_response_always:
                print(r.status_code)
                print(r.text)

            # Check Status and Return or Show Error
            if method == 'get':
                if r.status_code == 200 or r.status_code == 404:
                    json_data = r.json()
                    return r.status_code,json_data
                else: validating.error_request(r.status_code, r.json())
            elif method == 'patch':
                if r.status_code == 201 or r.status_code == 200:
                    json_data = r.json()
                    return json_data
                else:
                    validating.error_request(r.status_code, r.json())
            elif method == 'post':
                if r.status_code != 200:
                    json_data = r.json()
                    return json_data
                else:
                    validating.error_request(r.status_code, r.json())
            else: validating.error_request(r.status_code, r.json())

        except requests.exceptions.ConnectionError as e:
            print("Connection error, pausing before retrying. Error: %s" % (e))
            time.sleep(5)
        except Exception as e:
            print("Method %s Failed. Exception: %s" % (section[:-5], e))
            exit()

#========================================================
# Function to Format Terraform Files
#========================================================
def terraform_fmt(folder):
    # Run terraform fmt to cleanup the formating for all of the auto.tfvar files and tf files if needed
    print(f'\n-----------------------------------------------------------------------------\n')
    print(f'  Running "terraform fmt" in folder "{folder}",')
    print(f'  to correct variable formatting!')
    print(f'\n-----------------------------------------------------------------------------\n')
    p = subprocess.Popen(
        ['terraform', 'fmt', folder],
        stdout = subprocess.PIPE,
        stderr = subprocess.PIPE
    )
    print('Format updated for the following Files:')
    for line in iter(p.stdout.readline, b''):
        line = line.decode("utf-8")
        line = line.strip()
        print(f'- {line}')

#========================================================
# Function to Validate Worksheet User Input
#========================================================
def validate_args(jsonData, **kwargs):
    globalData = kwargs['easy_jsonData']['globalData']['allOf'][1]['properties']
    global_args = [
        'admin_state',
        'application_epg',
        'application_profile',
        'annotation',
        'annotations',
        'audit_logs',
        'bridge_domain',
        'cdp_interface_policy',
        'description',
        'events',
        'faults',
        'global_alias',
        'l3out',
        'lldp_interface_policy',
        'login_domain',
        'management_epg',
        'management_epg_type',
        'monitoring_policy',
        'name',
        'name_alias',
        'node_id',
        'pod_id',
        'policies_tenant',
        'policy_name',
        'profile_name',
        'port_channel_policy',
        'qos_class',
        'schema',
        'session_logs',
        'sites',
        'target_dscp',
        'template',
        'tenant',
        'username',
        'vrf'
    ]
    for i in jsonData['required_args']:
        if i in global_args:
            if globalData[i]['type'] == 'boolean':
                if not (kwargs[i] == None or kwargs[i] == ''):
                    validating.boolean(i, **kwargs)
            elif globalData[i]['type'] == 'integer':
                if kwargs[i] == None:
                    kwargs[i] = globalData[i]['default']
                else:
                    validating.number_check(i, globalData, **kwargs)
            elif globalData[i]['type'] == 'key_value':
                if not (kwargs[i] == None or kwargs[i] == ''):
                    validating.key_value(i, globalData, **kwargs)
            elif globalData[i]['type'] == 'list_of_string':
                if not (kwargs[i] == None or kwargs[i] == ''):
                    validating.string_list(i, globalData, **kwargs)
            elif globalData[i]['type'] == 'list_of_values':
                if kwargs[i] == None:
                    kwargs[i] = globalData[i]['default']
                else:
                    validating.list_values(i, globalData, **kwargs)
            elif globalData[i]['type'] == 'string':
                if not (kwargs[i] == None or kwargs[i] == ''):
                    validating.string_pattern(i, globalData, **kwargs)
            else:
                print(f'error validating.  Type not found {i}. 1')
                exit()
        elif i == 'site_group':
            validating.site_group('site_group', **kwargs)
        elif jsonData[i]['type'] == 'boolean':
            if not (kwargs[i] == None or kwargs[i] == ''):
                validating.boolean(i, **kwargs)
        elif jsonData[i]['type'] == 'hostname':
            if not (kwargs[i] == None or kwargs[i] == ''):
                if ':' in kwargs[i]:
                    validating.ip_address(i, **kwargs)
                elif re.search('[a-z]', kwargs[i], re.IGNORECASE):
                    validating.dns_name(i, **kwargs)
                else:
                    validating.ip_address(i, **kwargs)
        elif jsonData[i]['type'] == 'email':
            if not (kwargs[i] == None or kwargs[i] == ''):
                validating.email(i, **kwargs)
        elif jsonData[i]['type'] == 'integer':
            if kwargs[i] == None:
                kwargs[i] = jsonData[i]['default']
            else:
                validating.number_check(i, jsonData, **kwargs)
        elif jsonData[i]['type'] == 'list_of_domains':
            if not (kwargs[i] == None or kwargs[i] == ''):
                count = 1
                for domain in kwargs[i]:
                    kwargs[f'domain_{count}'] = domain
                    validating.domain(f'domain_{count}', **kwargs)
                    kwargs.pop(f'domain_{count}')
                    count += 1
        elif jsonData[i]['type'] == 'list_of_hosts':
            if not (kwargs[i] == None or kwargs[i] == ''):
                count = 1
                for hostname in kwargs[i].split(','):
                    kwargs[f'{i}_{count}'] = hostname
                    if ':' in hostname:
                        validating.ip_address(f'{i}_{count}', **kwargs)
                    elif re.search('[a-z]', hostname, re.IGNORECASE):
                        validating.dns_name(f'{i}_{count}', **kwargs)
                    else:
                        validating.ip_address(f'{i}_{count}', **kwargs)
                    kwargs.pop(f'{i}_{count}')
                    count += 1
        elif jsonData[i]['type'] == 'list_of_integer':
            if kwargs[i] == None:
                kwargs[i] = jsonData[i]['default']
            else:
                validating.number_list(i, jsonData, **kwargs)
        elif jsonData[i]['type'] == 'list_of_string':
            if not (kwargs[i] == None or kwargs[i] == ''):
                validating.string_list(i, jsonData, **kwargs)
        elif jsonData[i]['type'] == 'list_of_values':
            if kwargs[i] == None:
                kwargs[i] = jsonData[i]['default']
            else:
                validating.list_values(i, jsonData, **kwargs)
        elif jsonData[i]['type'] == 'list_of_vlans':
            if not (kwargs[i] == None or kwargs[i] == ''):
                validating.vlans(i, **kwargs)
        elif jsonData[i]['type'] == 'string':
            if not (kwargs[i] == None or kwargs[i] == ''):
                validating.string_pattern(i, jsonData, **kwargs)
        else:
            print(f"error validating.  Type not found {jsonData[i]['type']}. 2.")
            exit()
    for i in jsonData['optional_args']:
        if not (kwargs[i] == None or kwargs[i] == ''):
            if i in global_args:
                if globalData[i]['type'] == 'integer':
                    validating.number_check(i, globalData, **kwargs)
                elif globalData[i]['type'] == 'key_value':
                    validating.key_value(i, globalData, **kwargs)
                elif globalData[i]['type'] == 'list_of_string':
                    validating.string_list(i, globalData, **kwargs)
                elif globalData[i]['type'] == 'list_of_values':
                    validating.list_values(i, globalData, **kwargs)
                elif globalData[i]['type'] == 'string':
                    validating.string_pattern(i, globalData, **kwargs)
                else:
                    validating.validator(i, **kwargs)
            elif re.search(r'^module_[\d]+$', i):
                validating.list_values_key('modules', i, jsonData, **kwargs)
            elif jsonData[i]['type'] == 'boolean':
                validating.boolean(i, **kwargs)
            elif jsonData[i]['type'] == 'domain':
                validating.domain(i, **kwargs)
            elif jsonData[i]['type'] == 'email':
                validating.email(i, **kwargs)
            elif jsonData[i]['type'] == 'hostname':
                if ':' in kwargs[i]:
                    validating.ip_address(i, **kwargs)
                elif re.search('[a-z]', kwargs[i], re.IGNORECASE):
                    validating.dns_name(i, **kwargs)
                else:
                    validating.ip_address(i, **kwargs)
            elif jsonData[i]['type'] == 'integer':
                validating.number_check(i, jsonData, **kwargs)
            elif jsonData[i]['type'] == 'list_of_integer':
                validating.number_list(i, jsonData, **kwargs)
            elif jsonData[i]['type'] == 'list_of_hosts':
                count = 1
                for hostname in kwargs[i].split(','):
                    kwargs[f'{i}_{count}'] = hostname
                    if ':' in hostname:
                        validating.ip_address(f'{i}_{count}', **kwargs)
                    elif re.search('[a-z]', hostname, re.IGNORECASE):
                        validating.dns_name(f'{i}_{count}', **kwargs)
                    else:
                        validating.ip_address(f'{i}_{count}', **kwargs)
                    kwargs.pop(f'{i}_{count}')
                    count += 1
            elif jsonData[i]['type'] == 'list_of_macs':
                count = 1
                for mac in kwargs[i].split(','):
                    kwargs[f'{i}_{count}'] = mac
                    validating.mac_address(f'{i}_{count}', **kwargs)
                    kwargs.pop(f'{i}_{count}')
                    count += 1
            elif jsonData[i]['type'] == 'list_of_string':
                validating.string_list(i, jsonData, **kwargs)
            elif jsonData[i]['type'] == 'list_of_values':
                validating.list_values(i, jsonData, **kwargs)
            elif jsonData[i]['type'] == 'list_of_vlans':
                validating.vlans(i, **kwargs)
            elif jsonData[i]['type'] == 'mac_address':
                validating.mac_address(i, **kwargs)
            elif jsonData[i]['type'] == 'phone_number':
                validating.phone_number(i, **kwargs)
            elif jsonData[i]['type'] == 'string':
                validating.string_pattern(i, jsonData, **kwargs)
            else:
                print(f"error validating.  Type not found {jsonData[i]['type']}. 3.")
                exit()
    return kwargs

#========================================================
# Function to pull variables from easy_jsonData
#========================================================
def variablesFromAPI(**polVars):
    valid = False
    while valid == False:
        json_vars = polVars["jsonVars"]
        if 'popList' in polVars:
            if len(polVars["popList"]) > 0:
                for x in polVars["popList"]:
                    varsCount = len(json_vars)
                    for r in range(0, varsCount):
                        if json_vars[r] == x:
                            json_vars.pop(r)
                            break
        print(f'\n-------------------------------------------------------------------------------------------\n')
        newDescr = polVars["var_description"]
        if '\n' in newDescr:
            newDescr = newDescr.split('\n')
            for line in newDescr:
                if '*' in line:
                    print(fill(f'{line}',width=88, subsequent_indent='    '))
                else:
                    print(fill(f'{line}',88))
        else:
            print(fill(f'{polVars["var_description"]}',88))
        print(f'\n    Select an Option Below:')
        for index, value in enumerate(json_vars):
            index += 1
            if value == polVars["defaultVar"]:
                defaultIndex = index
            if index < 10:
                print(f'     {index}. {value}')
            else:
                print(f'    {index}. {value}')
        print(f'\n-------------------------------------------------------------------------------------------\n')
        if polVars["multi_select"] == True:
            if not polVars["defaultVar"] == '':
                var_selection = input(f'Please Enter the Option Number(s) to Select for {polVars["varType"]}.  [{defaultIndex}]: ')
            else:
                var_selection = input(f'Please Enter the Option Number(s) to Select for {polVars["varType"]}: ')
        else:
            if not polVars["defaultVar"] == '':
                var_selection = input(f'Please Enter the Option Number to Select for {polVars["varType"]}.  [{defaultIndex}]: ')
            else:
                var_selection = input(f'Please Enter the Option Number to Select for {polVars["varType"]}: ')
        if not polVars["defaultVar"] == '' and var_selection == '':
            var_selection = defaultIndex

        if polVars["multi_select"] == False and re.search(r'^[0-9]+$', str(var_selection)):
            for index, value in enumerate(json_vars):
                index += 1
                if int(var_selection) == index:
                    selection = value
                    valid = True
        elif polVars["multi_select"] == True and re.search(r'(^[0-9]+$|^[0-9\-,]+[0-9]$)', str(var_selection)):
            var_list = vlan_list_full(var_selection)
            var_length = int(len(var_list))
            var_count = 0
            selection = []
            for index, value in enumerate(json_vars):
                index += 1
                for vars in var_list:
                    if int(vars) == index:
                        var_count += 1
                        selection.append(value)
            if var_count == var_length:
                valid = True
            else:
                print(f'\n-------------------------------------------------------------------------------------------\n')
                print(f'  The list of Vars {var_list} did not match the available list.')
                print(f'\n-------------------------------------------------------------------------------------------\n')
        else:
            print(f'\n-------------------------------------------------------------------------------------------\n')
            print(f'  Error!! Invalid Selection.  Please Select a valid Option from the List.')
            print(f'\n-------------------------------------------------------------------------------------------\n')
    return selection

#========================================================
# Function to pull variables from easy_jsonData
#========================================================
def varBoolLoop(**polVars):
    print(f'\n-------------------------------------------------------------------------------------------\n')
    newDescr = polVars["Description"]
    if '\n' in newDescr:
        newDescr = newDescr.split('\n')
        for line in newDescr:
            if '*' in line:
                print(fill(f'{line}',width=88, subsequent_indent='    '))
            else:
                print(fill(f'{line}',88))
    else:
        print(fill(f'{polVars["Description"]}',88))
    print(f'\n-------------------------------------------------------------------------------------------\n')
    valid = False
    while valid == False:
        varValue = input(f'{polVars["varInput"]}  [{polVars["varDefault"]}]: ')
        if varValue == '':
            if polVars["varDefault"] == 'Y':
                varValue = True
            elif polVars["varDefault"] == 'N':
                varValue = False
            valid = True
        elif varValue == 'N':
            varValue = False
            valid = True
        elif varValue == 'Y':
            varValue = True
            valid = True
        else:
            print(f'\n-------------------------------------------------------------------------------------------\n')
            print(f'   {polVars["varName"]} value of "{varValue}" is Invalid!!! Please enter "Y" or "N".')
            print(f'\n-------------------------------------------------------------------------------------------\n')
    return varValue

#========================================================
# Function to pull variables from easy_jsonData
#========================================================
def varStringLoop(**polVars):
    maximum = polVars["maximum"]
    minimum = polVars["minimum"]
    varName = polVars["varName"]
    pattern = polVars["pattern"]

    print(f'\n-------------------------------------------------------------------------------------------\n')
    newDescr = polVars["Description"]
    if '\n' in newDescr:
        newDescr = newDescr.split('\n')
        for line in newDescr:
            if '*' in line:
                print(fill(f'{line}',width=88, subsequent_indent='    '))
            else:
                print(fill(f'{line}',88))
    else:
        print(fill(f'{polVars["Description"]}',88))
    print(f'\n-------------------------------------------------------------------------------------------\n')
    valid = False
    while valid == False:
        varValue = input(f'{polVars["varInput"]} ')
        if 'press enter to skip' in polVars["varInput"] and varValue == '':
            valid = True
        elif not polVars["varDefault"] == '' and varValue == '':
            varValue = polVars["varDefault"]
            valid = True
        elif not varValue == '':
            valid = validating.length_and_regex(pattern, varName, varValue, minimum, maximum)
        else:
            print(f'\n-------------------------------------------------------------------------------------------\n')
            print(f'   {varName} value of "{varValue}" is Invalid!!! ')
            print(f'\n-------------------------------------------------------------------------------------------\n')
    return varValue

#======================================================
# Function - Collapse VLAN List
#======================================================
def vlan_list_format(vlan_list_expanded):
    vlanGroups = itertools.groupby(vlan_list_expanded, key=lambda item, c=itertools.count():item-next(c))
    tempvlans = [list(g) for k, g in vlanGroups]
    vlanList = [str(x[0]) if len(x) == 1 else "{}-{}".format(x[0],x[-1]) for x in tempvlans]
    vlan_list = ",".join(vlanList)
    return vlan_list

#========================================================
# Function to Expand the VLAN list
#========================================================
def vlan_list_full(vlan_list):
    full_vlan_list = []
    if re.search(r',', str(vlan_list)):
        vlist = vlan_list.split(',')
        for v in vlist:
            if re.fullmatch('^\\d{1,4}\\-\\d{1,4}$', v):
                a,b = v.split('-')
                a = int(a)
                b = int(b)
                vrange = range(a,b+1)
                for vl in vrange:
                    full_vlan_list.append(int(vl))
            elif re.fullmatch('^\\d{1,4}$', v):
                full_vlan_list.append(int(v))
    elif re.search('\\-', str(vlan_list)):
        a,b = vlan_list.split('-')
        a = int(a)
        b = int(b)
        vrange = range(a,b+1)
        for v in vrange:
            full_vlan_list.append(int(v))
    else:
        full_vlan_list.append(vlan_list)
    return full_vlan_list

#========================================================
# Function to Expand a VLAN Range to a VLAN List
#========================================================
def vlan_range(vlan_list, **polVars):
    results = 'unknown'
    while results == 'unknown':
        if re.search(',', str(vlan_list)):
            vx = vlan_list.split(',')
            for vrange in vx:
                if re.search('-', vrange):
                    vl = vrange.split('-')
                    min_ = int(vl[0])
                    max_ = int(vl[1])
                    if (int(polVars['VLAN']) >= min_ and int(polVars['VLAN']) <= max_):
                        results = 'true'
                        return results
                else:
                    if polVars['VLAN'] == vrange:
                        results = 'true'
                        return results
            results = 'false'
            return results
        elif re.search('-', str(vlan_list)):
            vl = vlan_list.split('-')
            min_ = int(vl[0])
            max_ = int(vl[1])
            if (int(polVars['VLAN']) >= min_ and int(polVars['VLAN']) <= max_):
                results = 'true'
                return results
        else:
            if int(polVars['VLAN']) == int(vlan_list):
                results = 'true'
                return results
        results = 'false'
        return results

#========================================================
# Function to Create Workbook Styles.
#========================================================
def workbook_styles():
    wbstyles = DotMap()
    # Create Workbook Format
    bd1 = Side(style="thick", color="8EA9DB")
    bd2 = Side(style="medium", color="8EA9DB")
    wbstyles.wsh1 = NamedStyle(name="wsh1")
    wbstyles.wsh1.alignment = Alignment(horizontal="center", vertical="center", wrap_text="True")
    wbstyles.wsh1.border = Border(left=bd1, top=bd1, right=bd1, bottom=bd1)
    wbstyles.wsh1.font = Font(bold=True, size=15, color="FFFFFF")
    wbstyles.wsh2 = NamedStyle(name="wsh2")
    wbstyles.wsh2.alignment = Alignment(horizontal="center", vertical="center", wrap_text="True")
    wbstyles.wsh2.border = Border(left=bd2, top=bd2, right=bd2, bottom=bd2)
    wbstyles.wsh2.fill = PatternFill("solid", fgColor="305496")
    wbstyles.wsh2.font = Font(bold=True, size=15, color="FFFFFF")
    wbstyles.ws_odd = NamedStyle(name="ws_odd")
    wbstyles.ws_odd.alignment = Alignment(horizontal="center", vertical="center")
    wbstyles.ws_odd.border = Border(left=bd2, top=bd2, right=bd2, bottom=bd2)
    wbstyles.ws_odd.fill = PatternFill("solid", fgColor="D9E1F2")
    wbstyles.ws_odd.font = Font(bold=False, size=12, color="44546A")
    wbstyles.ws_even = NamedStyle(name="ws_even")
    wbstyles.ws_even.alignment = Alignment(horizontal="center", vertical="center")
    wbstyles.ws_even.border = Border(left=bd2, top=bd2, right=bd2, bottom=bd2)
    wbstyles.ws_even.font = Font(bold=False, size=12, color="44546A")
    return wbstyles


#========================================================
# Function to Determine which sites to write files to.
#========================================================
def write_to_site(polVars, **kwargs):
    args       = kwargs['args']
    baseRepo   = args.dir
    class_type = kwargs['class_type']
    dest_dir   = kwargs["dest_dir"]
    dest_file  = kwargs["tf_file"]
    site_name  = kwargs["site_name"]

    aci_template_path = pkg_resources.resource_filename(f'classes', 'templates/')
    templateLoader = jinja2.FileSystemLoader(
        searchpath=(aci_template_path + '%s/') % (class_type))
    templateEnv = jinja2.Environment(loader=templateLoader)
    
    # Define the Template Source
    template = templateEnv.get_template(kwargs["template_file"])

    # Make sure the Destination Path and Folder Exist
    if not os.path.isdir(os.path.join(baseRepo, site_name, dest_dir)):
        opSystem = platform.system()
        if opSystem == 'Windows': path_sep = '\\'
        else: path_sep = '/'
        dest_path = f'{os.path.join(baseRepo, site_name)}{path_sep}{dest_dir}'
        os.makedirs(dest_path)
    dest_dir = os.path.join(baseRepo, site_name, dest_dir)
    if not os.path.exists(os.path.join(dest_dir, dest_file)):
        create_file = f'type nul >> {os.path.join(dest_dir, dest_file)}'
        os.system(create_file)
    tf_file = os.path.join(dest_dir, dest_file)
    wr_file = open(tf_file, 'w')

    # Render Payload and Write to File
    polVars = json.loads(json.dumps(polVars))
    polVars = {'keys':polVars}
    payload = template.render(polVars)
    wr_file.write(payload)
    wr_file.close()

