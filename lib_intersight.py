#!/usr/bin/env python3

import ast
import ipaddress
import jinja2
import os, re, sys
import pkg_resources
import validating
from openpyxl import load_workbook, workbook, Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, colors, Border, Font, NamedStyle, PatternFill, Protection, Side 
from openpyxl.utils.dataframe import dataframe_to_rows
from ordered_set import OrderedSet
from subprocess import check_output

re_aep = re.compile(r'"uni/infra/attentp-(.*)"\n')
re_cdp = re.compile(r'uni/infra/cdpIfP-(.*)"\n')
re_llp = re.compile(r'uni/infra/hintfpol-(.*)"\n')
re_lldp = re.compile(r'uni/infra/lldpIfP-(.*)"\n')
re_mtu = re.compile(r'uni/infra/attentp-(.*)"\n')
re_stp = re.compile(r'uni/infra/ifPol-(.*)"\n')

# Log levels 0 = None, 1 = Class only, 2 = Line
log_level = 2

# Global path to main Template directory
aci_template_path = pkg_resources.resource_filename('aci_lib', 'ACI/templates/')

# Exception Classes
class InsufficientArgs(Exception):
    pass

class ErrException(Exception):
    pass

class InvalidArg(Exception):
    pass

class LoginFailed(Exception):
    pass

# Terraform ACI Provider - Access Policies
# Class must be instantiated with Variables
class Access_Policies(object):
    def __init__(self, ws):
        self.templateLoader = jinja2.FileSystemLoader(
            searchpath=(aci_template_path + 'Access_Policies/'))
        self.templateEnv = jinja2.Environment(loader=self.templateLoader)
    
    # Method must be called with the following kwargs.
    # Please Refer to the "Excel Spreadsheet Guidance" PDF File  
    # for Detailed information on the Arguments used by this Method.
    def add_apg(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'AAEP': '',
                         'MTU': '',
                         'Speed': '',
                         'CDP': '',
                         'LLDP_Rx': '',
                         'LLDP_Tx': '',
                         'STP': ''}
        optional_args = {'Description': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.mtu(row_num, ws, 'MTU', templateVars['MTU'])
            validating.mtu(row_num, ws, 'MTU', templateVars['MTU'])
            validating.link_level(row_num, ws, 'Speed', templateVars['Speed'])
            validating.stp(row_num, ws, 'STP', templateVars['STP'])
            validating.noyes(row_num, ws, 'CDP', templateVars['CDP'])
            validating.noyes(row_num, ws, 'LLDP_Rx', templateVars['LLDP_Rx'])
            validating.noyes(row_num, ws, 'LLDP_Tx', templateVars['LLDP_Tx'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)
        
        if templateVars['CDP'] == 'no':
            templateVars['CDP'] = 'cdp_Disabled'
        else:
            templateVars['CDP'] = 'cdp_Enabled'
        if templateVars['LLDP_Tx'] == 'no':
            templateVars['LLDP'] = 'lldp_Disabled'
        else:
            templateVars['LLDP'] = 'lldp_Enabled'

        # Define the Template Source
        template_file = "add_apg.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'pg_access_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the "Excel Spreadsheet Guidance" PDF File  
    # for Detailed information on the Arguments used by this Method.
    def add_bundle(self, wb, ws, row_num, wr_file, **kwargs):
        # Dicts for required and optional args
        required_args = {'Port_Type': '',
                         'LACP': '',
                         'Bundle_ID': '',
                         'Site_Group': '',
                         'Site_Name': '',
                         'Name': '',
                         'AAEP': '',
                         'CDP': '',
                         'LLDP': '',
                         'MTU': '',
                         'Speed': '',
                         'STP': '',}
        optional_args = {'Description': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.int_type(row_num, ws, 'Port_Type', templateVars['Port_Type'])
            validating.link_level(row_num, ws, 'Speed', templateVars['Speed'])
            # validating.mtu(row_num, ws, 'MTU', templateVars['MTU'])
            validating.stp(row_num, ws, 'STP', templateVars['STP'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)
        
        
        # Define the Template Source
        template_file = "add_bundle.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        if templateVars['Port_Type'] == 'port-channel':
            dest_file = 'pg_pc%s_%s.tf' % (templateVars['Bundle_ID'], templateVars['Name'])
            templateVars['Name'] = 'pc%s_%s' % (templateVars['Bundle_ID'], templateVars['Name'])
            templateVars['LAG_Type'] = 'link'
        else:
            dest_file = 'pg_vpc%s_%s.tf' % (templateVars['Bundle_ID'], templateVars['Name'])
            templateVars['Name'] = 'vpc%s_%s' % (templateVars['Bundle_ID'], templateVars['Name'])
            templateVars['LAG_Type'] = 'node'
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the "Excel Spreadsheet Guidance" PDF File  
    # for Detailed information on the Arguments used by this Method.
    def apic_inb(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Node_ID': '',
                         'Pod_ID': '',
                         'Inband_IP': '',
                         'Inband_GW': ''}
        optional_args = { }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.hostname(row_num, ws, 'Name', templateVars['Name'])
            validating.node_id_apic(row_num, ws, 'Node_ID', templateVars['Node_ID'])
            validating.pod_id(row_num, ws, 'Pod_ID', templateVars['Pod_ID'])
            validating.mgmt_network(row_num, ws, 'Inband_IP', templateVars['Inband_IP'], 'Inband_GW', templateVars['Inband_GW'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)
        
        templateVars['Inband_GW_'] = templateVars['Inband_GW'].replace('.', '-')
        templateVars['app_Dn'] = 'aci_application_epg.mgmt_inb_ap_default'

        # Define the Template Source
        template_file = "mgmt_inb.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = '%s_inb_%s.tf' % (templateVars['Name'], templateVars['Inband_GW_'])
        dest_dir = 'Tenant_mgmt'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the "Excel Spreadsheet Guidance" PDF File  
    # for Detailed information on the Arguments used by this Method.
    def inb_subnet(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Inband_VLAN': '',
                         'Inband_GW': ''}
        optional_args = { }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.ip_address(row_num, ws, 'Inband_GW', templateVars['Inband_GW'])
            validating.vlans(row_num, ws, 'Inband_VLAN', templateVars['Inband_VLAN'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)
        
        # Define the Template Source
        template_file = "inb_subnet.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'inb_subnet.tf'
        dest_dir = 'Tenant_mgmt'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the "Excel Spreadsheet Guidance" PDF File  
    # for Detailed information on the Arguments used by this Method.
    def intf_selector(self, wb, ws, row_num, wr_file, Site_Group, Site_Name, Switch_Role, **kwargs):
        if not kwargs.get('Policy_Group') == None:
            # Dicts for required and optional args
            required_args = {'Switch_Name': '',
                            'Node_ID': '',
                            'Interface_Selector': '',
                            'Port': '',
                            'Policy_Group': '',
                            'Port_Type': ''}
            optional_args = {'LACP': '',
                            'Bundle_ID': '',
                            'Description': '',
                            'Switchport_Mode': '',
                            'Access_or_Native': '',
                            'Trunk_Allowed_VLANs': ''}

            # Validate inputs, return dict of template vars
            templateVars = process_kwargs(required_args, optional_args, **kwargs)
            # leafx = Name
            templateVars['Site_Group'] = Site_Group
            templateVars['Site_Name'] = Site_Name
            templateVars['Switch_Role'] = Switch_Role
            if not templateVars['Port_Type'] == None:
                if re.search('(port-channel|vpc)', templateVars['Port_Type']):
                    pg_file = './ACI/%s/Access/pg_access_%s.tf' % (templateVars['Site_Name'], templateVars['Policy_Group'])
                    if not os.path.isfile(pg_file):
                        print(f"\n-----------------------------------------------------------------------------\n")
                        print(f"   Error on Worksheet {ws.title}, Row {row_num} Policy Group.")
                        print(f"   {templateVars['Policy_Group']} does not exist in the directory ")
                        print(f"   {pg_file} not found.  Exiting....")
                        print(f"\n-----------------------------------------------------------------------------\n")
                        exit()

                    # Get Policy Group Attributes from the Access Policy Group
                    filename = './ACI/%s/Access/pg_access_%s.tf' % (Site_Name, templateVars['Policy_Group'])
                    child = check_output(['cat', filename])
                    child = child.decode("utf-8")
                    if re.search(re_aep, child):
                        templateVars['AAEP'] = re.search(re_aep, child).group(1)
                    if re.search(re_cdp, child):
                        templateVars['CDP'] = re.search(re_cdp, child).group(1)
                    if re.search(re_lldp, child):
                        templateVars['LLDP'] = re.search(re_lldp, child).group(1)
                    if re.search(re_mtu, child):
                        templateVars['MTU'] = re.search(re_mtu, child).group(1)
                    if re.search(re_llp, child):
                        templateVars['Speed'] = re.search(re_llp, child).group(1)
                    if re.search(re_stp, child):
                        templateVars['STP'] = re.search(re_stp, child).group(1)

                    if templateVars['Port_Type'] == 'vpc':
                        ws_vpc = wb['Inventory']
                        for row in ws_vpc.rows:
                            if row[0].value == 'vpc_pair' and int(row[1].value) == int(Site_Group) and str(row[4].value) == str(templateVars['Node_ID']):
                                templateVars['Name'] = row[3].value
                                templateVars['Policy_Group'] = 'pg_vpc%s_%s' % (templateVars['Bundle_ID'], templateVars['Name'])

                            elif row[0].value == 'vpc_pair' and str(row[1].value) == str(Site_Group) and str(row[5].value) == str(templateVars['Node_ID']):
                                templateVars['Name'] = row[3].value
                                templateVars['Policy_Group'] = 'pg_vpc%s_%s' % (templateVars['Bundle_ID'], templateVars['Name'])
                    elif templateVars['Port_Type'] == 'port-channel':
                        templateVars['Name'] = templateVars['Switch_Name']
                        templateVars['Policy_Group'] = 'pg_pc%s_%s' % (templateVars['Bundle_ID'], templateVars['Name'])
                    
                    # Create the Bundle Policy Group
                    aci_lib_ref = 'Access_Policies'
                    class_init = '%s(ws)' % (aci_lib_ref)
                    func = 'add_bundle'
                    eval("%s.%s(wb, ws, row_num, wr_file, **templateVars)" % (class_init, func))

            xa = templateVars['Port'].split('/')
            xcount = len(xa)
            templateVars['Module'] = xa[0]
            templateVars['Port'] = xa[1]
            if Switch_Role == 'leaf':
                if not templateVars['Port_Type'] == 'breakout':
                    templateVars['Resource_Type'] = 'aci_leaf_access_port_policy_group'
                    if templateVars['Port_Type'] == 'access':
                        templateVars['PG_Type'] = 'accportgrp'
                elif templateVars['Port_Type'] == 'breakout':
                    templateVars['Resource_Type'] = 'aci_rest'
                    templateVars['PG_Type'] = 'brkoutportgrp'

                # If Policy Group Exists then Add Policy Group to templateVars for Port Selector
                if not templateVars['Policy_Group'] == None:
                    if templateVars['Port_Type'] == 'breakout':
                        templateVars['DN_Policy_Group'] = 'uni/infra/funcprof/brkoutportgrp-%s' % (templateVars['Policy_Group'])
                    elif templateVars['Port_Type'] == 'individual':
                        templateVars['DN_Policy_Group'] = 'uni/infra/funcprof/accportgrp-%s' % (templateVars['Policy_Group'])
                    elif templateVars['Port_Type'] == 'port-channel':
                        templateVars['DN_Policy_Group'] = 'uni/infra/funcprof/accbundle-vpc%s_%s' % (templateVars['Bundle_ID'], templateVars['Name'])
                    elif templateVars['Port_Type'] == 'vpc':
                        templateVars['DN_Policy_Group'] = 'uni/infra/funcprof/accbundle-vpc%s_%s' % (templateVars['Bundle_ID'], templateVars['Name'])

                # Define the Template Source
                template_file = "leaf_portselect.template"
                template = self.templateEnv.get_template(template_file)

                # Process the template and write to file
                payload = template.render(templateVars)
                wr_file.write(payload + '\n\n')

                # Define the Template Source
                if xcount == 3:
                    templateVars['Sub_Port'] = xa[2]
                    template_file = "leaf_portblock_sub.template"
                else:
                    template_file = "leaf_portblock.template"
                template = self.templateEnv.get_template(template_file)

                # Process the template and write to file
                payload = template.render(templateVars)
                wr_file.write(payload + '\n\n')

            elif Switch_Role == 'spine':
                # Define the Template Source
                template_file = "spine_portselect.template"
                template = self.templateEnv.get_template(template_file)

                # Process the template and write to file
                payload = template.render(templateVars)
                wr_file.write(payload + '\n\n')

                # Define the Template Source
                if not templateVars['Policy_Group'] == None:
                    if Switch_Role == 'spine':
                        template_file = "spine_pg_to_select.template"
                    template = self.templateEnv.get_template(template_file)

                    # Process the template and write to file
                    payload = template.render(templateVars)
                    wr_file.write(payload + '\n\n')

    # Method must be called with the following kwargs.
    # Please Refer to the "Excel Spreadsheet Guidance" PDF File  
    # for Detailed information on the Arguments used by this Method.
    def port_cnvt(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Node_ID': '',
                         'Port': ''}
        optional_args = { }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.hostname(row_num, ws, 'Name', templateVars['Name'])
            validating.node_id(row_num, ws, 'Node_ID', templateVars['Node_ID'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)
        
        # Create Port Name Var
        zz = templateVars['Port'].split('/')
        templateVars['Port_Name'] = '%s_%s' % (zz[0], zz[1])

        # Define the Template Source
        template_file = "downlink.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'downlink_convert_%s.tf' % (templateVars['Port_Name'])
        dest_dir = 'Access/%s' % (templateVars['Name'])
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the "Excel Spreadsheet Guidance" PDF File  
    # for Detailed information on the Arguments used by this Method.
    def switch(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Serial': '',
                         'Name': '',
                         'Node_ID': '',
                         'Node_Type': '',
                         'Pod_ID': '',
                         'Switch_Role': '',
                         'Switch_Type': '',
                         'Inband_IP': '',
                         'Inband_GW': ''}
        optional_args = {'OOB_IP': '',
                         'OOB_GW': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        # Use Switch_Type to Determine the Number of ports on the switch
        modules,port_count = query_switch_model(row_num, templateVars['Switch_Type'])
        
        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.hostname(row_num, ws, 'Name', templateVars['Name'])
            validating.node_id(row_num, ws, 'Node_ID', templateVars['Node_ID'])
            validating.node_type(row_num, templateVars['Name'], templateVars['Node_Type'])
            validating.pod_id(row_num, ws, 'Pod_ID', templateVars['Pod_ID'])
            validating.role(row_num, templateVars['Name'], templateVars['Switch_Role'])
            validating.modules(row_num, templateVars['Name'], templateVars['Switch_Role'], modules)
            validating.port_count(row_num, templateVars['Name'], templateVars['Switch_Role'], port_count)
            validating.mgmt_network(row_num, ws, 'Inband_IP', templateVars['Inband_IP'], 'Inband_GW', templateVars['Inband_GW'])
            if not templateVars['OOB_IP'] == None:
                validating.mgmt_network(row_num, ws, 'OOB_IP', templateVars['OOB_IP'], 'OOB_GW', templateVars['OOB_GW'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)
        
        if re.search('Grp_[A-F]', templateVars['Site_Group']):
            print(f"\n-----------------------------------------------------------------------------\n")
            print(f"   Error on Worksheet {ws.title}, Row {row_num} Site_Group, value {templateVars['Site_Group']}.")
            print(f"   A Leaf can only be assigned to one Site.  Exiting....")
            print(f"\n-----------------------------------------------------------------------------\n")
            exit()
        elif re.search(r'\d+', templateVars['Site_Group']):
            Site_ID = 'Site_ID_%s' % (templateVars['Site_Group'])
            site_dict = ast.literal_eval(os.environ[Site_ID])

            # Create templateVars for Site_Name and APIC_URL
            templateVars['Site_Name'] = site_dict.get('Site_Name')
            templateVars['APIC_URL'] = site_dict.get('APIC_URL')
        else:
            print(f"\n-----------------------------------------------------------------------------\n")
            print(f"   Error on Worksheet {ws.title}, Row {row_num} Site_Group, value {templateVars['Site_Group']}.")
            print(f"   Unable to Determine if this is a Single or Group of Site(s).  Exiting....")
            print(f"\n-----------------------------------------------------------------------------\n")
            exit()

        # Copy the data file for the Inband EPG Into the Switch Directory
        src_dir = './ACI/templates'
        dest_dir = './ACI/%s/%s' % (templateVars['Site_Name'], templateVars['Name'])

        cp_template = 'cp %s/data_inband_epg.tf %s/' % (src_dir, dest_dir)
        os.system(cp_template)

        dest_dir = templateVars['Name']

        # Copy the Necessary Default terraform files to the switch directory
        dest_dir = templateVars['Name']
        copy_defaults(templateVars['Site_Name'], dest_dir)

        # Write the variables.tf to the Appropriate Directories
        self.templateLoader = jinja2.FileSystemLoader(searchpath=('ACI/templates/'))
        self.templateEnv = jinja2.Environment(loader=self.templateLoader)
        template_file = "variables.tf"
        template = self.templateEnv.get_template(template_file)
        create_tf_file('w', dest_dir, template_file, template, **templateVars)

        self.templateLoader = jinja2.FileSystemLoader(searchpath=(aci_template_path + 'Access_Policies/'))
        self.templateEnv = jinja2.Environment(loader=self.templateLoader)
        excel_wkbook = '%s_intf_selectors.xlsx' % (templateVars['Site_Name'])
        
        wb_sw = load_workbook(excel_wkbook)

        # Check if there is a Worksheet for the Switch Already
        if not templateVars['Name'] in wb_sw.sheetnames:
            ws_sw = wb_sw.create_sheet(title = templateVars['Name'])
            ws_sw = wb_sw[templateVars['Name']]
            ws_sw.column_dimensions['A'].width = 15
            ws_sw.column_dimensions['B'].width = 10
            ws_sw.column_dimensions['c'].width = 10
            ws_sw.column_dimensions['d'].width = 20
            ws_sw.column_dimensions['E'].width = 20
            ws_sw.column_dimensions['F'].width = 10
            ws_sw.column_dimensions['G'].width = 20
            ws_sw.column_dimensions['H'].width = 20
            ws_sw.column_dimensions['I'].width = 20
            ws_sw.column_dimensions['J'].width = 15
            ws_sw.column_dimensions['K'].width = 30
            ws_sw.column_dimensions['L'].width = 20
            ws_sw.column_dimensions['M'].width = 20
            ws_sw.column_dimensions['N'].width = 30
            dv1 = DataValidation(type="list", formula1='"intf_selector"', allow_blank=True)
            dv2 = DataValidation(type="list", formula1='"access,breakout,port-channel,vpc"', allow_blank=True)
            dv3 = DataValidation(type="list", formula1='"lacp_Active,lacp_MacPin,lacp_Passive,lacp_Static"', allow_blank=True)
            ws_sw.add_data_validation(dv1)
            ws_sw.add_data_validation(dv2)
            ws_sw.add_data_validation(dv3)
            ws_header = '%s Interface Selectors' % (templateVars['Name'])
            data = [ws_header]
            ws_sw.append(data)
            ws_sw.merge_cells('A1:N1')
            for cell in ws_sw['1:1']:
                cell.style = 'Heading 1'
            data = ['','Notes: Breakout Policy Group Names are 2x100g_pg, 4x10g_pg, 4x25g_pg, 4x100g_pg, 8x50g_pg.']
            ws_sw.append(data)
            ws_sw.merge_cells('B2:N2')
            for cell in ws_sw['2:2']:
                cell.style = 'Heading 2'
            data = ['Type','Pod_ID','Node_ID','Switch_Name','Interface_Selector','Port','Policy_Group','Port_Type','LACP','Bundle_ID','Description','Switchport_Mode','Access_or_Native','Trunk_Allowed_VLANs']
            ws_sw.append(data)
            for cell in ws_sw['3:3']:
                cell.style = 'Heading 3'

            ws_sw_row_count = 4
            templateVars['dv1'] = dv1
            templateVars['dv2'] = dv2
            templateVars['dv3'] = dv3
            templateVars['port_count'] = port_count
            sw_type = str(templateVars['Switch_Type'])
            sw_name = str(templateVars['Name'])
            if re.search('^(93[0-9][0-9])', sw_type):
                for module in range(1, 2):
                    templateVars['module'] = module
                    ws_sw_row_count = create_selector(ws_sw, ws_sw_row_count, **templateVars)
            if re.search('^(9396|95[0-1][4-8])', sw_type):
                row_count = 1
                for row in ws.rows:
                    if re.search('9396', sw_type):
                        start, end = 2, 2
                    else:
                        start, end = 1, int(modules)
                    if str(row[0].value) == sw_type and str(row[2].value) == sw_name:
                        for module in range(start, end + 2):
                            templateVars['module'] = module
                            module_type = row[module + 2].value
                            if module_type == None:
                                module_type = 'none'
                            elif re.search('(X97|M(4|6|12)P)', module_type):
                                templateVars['port_count'] = query_module_type(row_count, module_type)
                                ws_sw_row_count = create_selector(ws_sw, ws_sw_row_count, **templateVars)
                        row_count += 1
                        break
            wb_sw.save(excel_wkbook)
        else:
            ws_sw = wb_sw[templateVars['Name']]

        # Determine if this is an odd or even switch
        templateVars['Node_ID'] = int(templateVars['Node_ID'])
        if templateVars['Node_ID'] % 2 == 0:
            templateVars['Maint_Grp'] = 'MgB'
        else:
            templateVars['Maint_Grp'] = 'MgA'
        templateVars['Node_ID'] = str(templateVars['Node_ID'])

        templateVars['Inband_GW_'] = templateVars['Inband_GW'].replace('.', '-')
        templateVars['OOB_GW_'] = templateVars['OOB_GW'].replace('.', '-')

        # Define the Template Source
        if templateVars['Switch_Role'] == 'leaf':
            template_file = "leaf.template"
        else:
            template_file = "spine.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = '%s.tf' % (templateVars['Name'])
        dest_dir = '%s' % (templateVars['Name'])
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        templateVars['app_Dn'] = 'data.aci_application_epg.mgmt_inb_ap_default'

        # Define the Template Source
        template_file = "mgmt_inb.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = '%s.tf' % (templateVars['Name'])
        dest_dir = '%s' % (templateVars['Name'])
        process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

        if not templateVars['OOB_IP'] == None:
            # Define the Template Source
            template_file = "mgmt_oob.template"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = '%s.tf' % (templateVars['Name'])
            dest_dir = '%s' % (templateVars['Name'])
            process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

        sw_file = './ACI/%s/%s/%s.tf' % (templateVars['Site_Name'], templateVars['Name'], templateVars['Name'])
        wr_file = open(sw_file, 'a+')
        aci_lib_ref_sw = 'Access_Policies'
        rows_sw = ws_sw.max_row
        func_regex = re.compile('^intf_selector$')
        func_list_sw = findKeys(ws_sw, func_regex)
        class_init_sw = '%s(ws_sw)' % (aci_lib_ref_sw)
        stdout_log(ws_sw, None)
        for func_sw in func_list_sw:
            count_sw = countKeys(ws_sw, func_sw)
            var_dict_sw = findVars(ws_sw, func_sw, rows_sw, count_sw)
            for pos_sw in var_dict_sw:
                row_num_sw = var_dict_sw[pos_sw]['row']
                del var_dict_sw[pos_sw]['row']
                for x_sw in list(var_dict_sw[pos_sw].keys()):
                    if var_dict_sw[pos_sw][x_sw] == '':
                        del var_dict_sw[pos_sw][x_sw]
                stdout_log(ws_sw, row_num_sw)
                site_group = templateVars['Site_Group']
                sw_role = templateVars['Switch_Role']
                site_name = templateVars['Site_Name']
                eval("%s.%s(wb, ws_sw, row_num_sw, wr_file, '%s', '%s', '%s', **var_dict_sw[pos_sw])" % (class_init_sw, func_sw, site_group, site_name, sw_role))
        wr_file.close()
        ws_wr = wb_sw.get_sheet_names()
        for sheetName in ws_wr:
            if sheetName in ['Sites']:
                sheetToDelete = wb_sw.get_sheet_by_name(sheetName)
                wb_sw.remove_sheet(sheetToDelete)
                wb_sw.save(excel_wkbook)
        wb_sw.close()

    # Method must be called with the following kwargs.
    # Please Refer to the "Excel Spreadsheet Guidance" PDF File  
    # for Detailed information on the Arguments used by this Method.
    def vlan_pool(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Allocation_Mode': '',
                         'VLAN_Grp1': '',
                         'VGRP1_Allocation': ''}
        optional_args = {'VLAN_Grp1': '',
                         'VGRP1_Allocation': '',
                         'VLAN_Grp2': '',
                         'VGRP2_Allocation': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.alloc_mode(row_num, ws, 'Allocation_Mode', templateVars['Allocation_Mode'])
            validating.alloc_mode(row_num, ws, 'VGRP1_Allocation', templateVars['VGRP1_Allocation'])
            if not templateVars['VGRP2_Allocation'] == None:
                validating.alloc_mode(row_num, ws, 'VGRP2_Allocation', templateVars['VGRP2_Allocation'])
            validating.vlans(row_num, ws, 'VLAN_Grp1', templateVars['VLAN_Grp1'])
            if not templateVars['VLAN_Grp2'] == None:
                validating.vlans(row_num, ws, 'VLAN_Grp2', templateVars['VLAN_Grp2'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        if templateVars['Name'] == None:
            Error_Return = 'Error on Worksheet %s Row %s.  Could not Determine the Name of the VLAN Pool.' % (ws.title, row_num)
            raise ErrException(Error_Return)
        
        # Define the Template Source
        template_file = "vlan_pool.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'vlp_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "data_vlan_pool.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'data_vlp_%s.tf' % (templateVars['Name'])
        dest_dir = 'VLANs'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # Add VLAN(s) to VLAN Pool FIle
        if re.search('Grp_[A-F]', templateVars['Site_Group']):
            Group_ID = '%s' % (templateVars['Site_Group'])
            site_group = ast.literal_eval(os.environ[Group_ID])
            for x in range(1, 13):
                sitex = 'Site_%s' % (x)
                if not site_group[sitex] == None:
                    Site_ID = 'Site_ID_%s' % (site_group[sitex])
                    site_dict = ast.literal_eval(os.environ[Site_ID])

                    # Create templateVars for Site_Name and APIC_URL
                    templateVars['Site_Name'] = site_dict.get('Site_Name')
                    templateVars['APIC_URL'] = site_dict.get('APIC_URL')

                    # Create Blank VLAN Pool VLAN(s) File
                    dest_file = './ACI/%s/VLANs/vlp_%s.tf' % (templateVars['Site_Name'], templateVars['Name'])
                    wr_file = open(dest_file, 'w')
                    wr_file.close()
                    dest_file = 'vlp_%s.tf' % (templateVars['Name'])
                    dest_dir = 'VLANs'
                    template_file = "add_vlan_to_pool.template"
                    template = self.templateEnv.get_template(template_file)

                    for z in range(1, 3):
                        vgroup = 'VLAN_Grp%s' % (z)
                        vgrp = 'VGRP%s_Allocation' % (z)
                        templateVars['Allocation_Mode'] = templateVars[vgrp]
                        if re.search(r'\d+', str(templateVars[vgroup])):
                            vlan_list = vlan_list_full(templateVars[vgroup])
                            for v in vlan_list:
                                vlan = str(v)
                                if re.fullmatch(r'\d+', vlan):
                                    templateVars['VLAN_ID'] = int(vlan)

                                    # Add VLAN to VLAN Pool File
                                    create_tf_file('a+', dest_dir, dest_file, template, **templateVars)

        elif re.search(r'\d+', templateVars['Site_Group']):
            Site_ID = 'Site_ID_%s' % (templateVars['Site_Group'])
            site_dict = ast.literal_eval(os.environ[Site_ID])

            # Create templateVars for Site_Name and APIC_URL
            templateVars['Site_Name'] = site_dict.get('Site_Name')
            templateVars['APIC_URL'] = site_dict.get('APIC_URL')

            # Create Blank VLAN Pool VLAN(s) File
            dest_file = './ACI/%s/VLANs/vlp_%s.tf' % (templateVars['Site_Name'], templateVars['Name'])
            wr_file = open(dest_file, 'w')
            wr_file.close()
            dest_file = 'vlp_%s.tf' % (templateVars['Name'])
            dest_dir = 'VLANs'
            template_file = "add_vlan_to_pool.template"
            template = self.templateEnv.get_template(template_file)

            for z in range(1, 3):
                vgroup = 'VLAN_Grp%s' % (z)
                vgrp = 'VGRP%s_Allocation' % (z)
                templateVars['Allocation_Mode'] = templateVars[vgrp]
                if re.search(r'\d+', str(templateVars[vgroup])):
                    vlan_list = vlan_list_full(templateVars[vgroup])
                    for v in vlan_list:
                        vlan = str(v)
                        if re.fullmatch(r'\d+', vlan):
                            templateVars['VLAN_ID'] = int(vlan)

                            # Add VLAN to VLAN Pool File
                            create_tf_file('a+', dest_dir, dest_file, template, **templateVars)
        else:
            print(f"\n-----------------------------------------------------------------------------\n")
            print(f"   Error on Worksheet {ws.title}, Row {row_num} Site_Group, value {templateVars['Site_Group']}.")
            print(f"   Unable to Determine if this is a Single or Group of Site(s).  Exiting....")
            print(f"\n-----------------------------------------------------------------------------\n")
            exit()

    # Method must be called with the following kwargs.
    # Please Refer to the "Excel Spreadsheet Guidance" PDF File  
    # for Detailed information on the Arguments used by this Method.
    def vpc_pair(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'VPC_ID': '',
                         'Name': '',
                         'Node1_ID': '',
                         'Node2_ID': ''}
        optional_args = { }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.vpc_id(row_num, ws, 'VPC_ID', templateVars['VPC_ID'])
            validating.node_id(row_num, ws, 'Node1_ID', templateVars['Node1_ID'])
            validating.node_id(row_num, ws, 'Node2_ID', templateVars['Node2_ID'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)
        
        # Define the Template Source
        template_file = "vpc_domain.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'vpc_domain_%s.tf' % (templateVars['VPC_ID'])
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

# Terraform ACI Provider - Admin Policies
# Class must be instantiated with Variables
class Admin_Policies(object):
    def __init__(self, ws):
        self.templateLoader = jinja2.FileSystemLoader(
            searchpath=(aci_template_path + 'Admin_Policies/'))
        self.templateEnv = jinja2.Environment(loader=self.templateLoader)
    
    # Method must be called with the following kwargs.
    # Please Refer to the "Excel Spreadsheet Guidance" PDF File  
    # for Detailed information on the Arguments used by this Method.
    def backup(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Encryption_Key': '',
                         'Backup_Hour': '',
                         'Backup_Minute': '',
                         'Remote_Host': '',
                         'Mgmt_Domain': '',
                         'Protocol': '',
                         'Remote_Path': '',
                         'Port': '',
                         'Auth_Type': '',
                         'Passwd_or_SSH_Pass': ''}
        optional_args = {'Username': '',
                         'SSH_Key': '',
                         'Description': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.auth_type(row_num, ws, 'Auth_Type', templateVars['Auth_Type'])
            validating.encryption_key(row_num, ws, 'Encryption_Key', templateVars['Encryption_Key'])
            validating.hour(row_num, ws, 'Backup_Hour', templateVars['Backup_Hour'])
            validating.minute(row_num, ws, 'Backup_Minute', templateVars['Backup_Minute'])
            validating.port(row_num, ws, 'Port', templateVars['Port'])
            templateVars['Mgmt_Domain'] = validating.mgmt_domain(row_num, ws, 'Mgmt_Domain', templateVars['Mgmt_Domain'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)
        
        templateVars['Remote_Host_'] = templateVars['Remote_Host'].replace('.', '-')
        if templateVars['Auth_Type'] == 'password':
            templateVars['Auth_Type'] = 'usePassword'
        elif templateVars['Auth_Type'] == 'ssh-key':
            templateVars['Auth_Type'] = 'useSshKeyContents'

        # Define the Template Source
        template_file = "global_key.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'backup_remotehost.tf'
        dest_dir = 'Admin'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "backup_host.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        process_method(wb, ws, row_num, 'w+', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "backup_policy.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        process_method(wb, ws, row_num, 'w+', dest_dir, dest_file, template, **templateVars)
    
    # Method must be called with the following kwargs.
    # Please Refer to the "Excel Spreadsheet Guidance" PDF File  
    # for Detailed information on the Arguments used by this Method.
    def realm(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Auth_Realm': '',
                         'Domain_Type': ''}
        optional_args = {'Login_Domain': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.auth_realm(row_num, ws, 'Auth_Realm', templateVars['Auth_Realm'])
            validating.login_type(row_num, ws, 'Auth_Realm', templateVars['Auth_Realm'], 'Domain_Type', templateVars['Domain_Type'])
            if not templateVars['Domain_Type'] == 'local':
                validating.login_domain(row_num, ws, 'Login_Domain', templateVars['Login_Domain'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)
        
        if templateVars['Auth_Realm'] == 'console':
            templateVars['child_class'] = 'aaaConsoleAuth'
        elif templateVars['Auth_Realm'] == 'default':
            templateVars['child_class'] = 'aaaDefaultAuth'

        # Define the Template Source
        template_file = "realm.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        if templateVars['Auth_Realm'] == 'console':
            dest_file = 'realm_console.tf'
        else:
            dest_file = 'realm_default.tf'
        dest_dir = 'Admin'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the "Excel Spreadsheet Guidance" PDF File  
    # for Detailed information on the Arguments used by this Method.
    def radius(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Login_Domain': '',
                         'RADIUS_Server': '',
                         'Port': '',
                         'Shared_Secret': '',
                         'Authz_Proto': '',
                         'Timeout': '',
                         'Retry_Interval': '',
                         'Mgmt_Domain': '',
                         'Domain_Order': ''}
        optional_args = {}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.auth_proto(row_num, ws, 'Authz_Proto', templateVars['Authz_Proto'])  
            validating.ipv4(row_num, ws, 'RADIUS_Server', templateVars['RADIUS_Server'])
            validating.login_domain(row_num, ws, 'Login_Domain', templateVars['Login_Domain'])
            validating.secret(row_num, ws, 'Shared_Secret', templateVars['Shared_Secret'])
            validating.retry(row_num, ws, 'Retry_Interval', templateVars['Retry_Interval'])
            validating.timeout(row_num, ws, 'Timeout', templateVars['Timeout'])
            templateVars['Mgmt_Domain'] = validating.mgmt_domain(row_num, ws, 'Mgmt_Domain', templateVars['Mgmt_Domain'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)
        
        templateVars['RADIUS_Server_'] = templateVars['RADIUS_Server'].replace('.', '-')
        
        # Define the Template Source
        template_file = "radius.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'radius_%s.tf' % (templateVars['RADIUS_Server_'])
        dest_dir = 'Admin'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the "Excel Spreadsheet Guidance" PDF File  
    # for Detailed information on the Arguments used by this Method.
    def tacacs(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Login_Domain': '',
                         'TACACS_Server': '',
                         'Port': '',
                         'Shared_Secret': '',
                         'Auth_Proto': '',
                         'Timeout': '',
                         'Retry_Interval': '',
                         'Mgmt_Domain': '',
                         'Domain_Order': ''}
        optional_args = {}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.auth_proto(row_num, ws, 'Auth_Proto', templateVars['Auth_Proto'])  
            validating.ipv4(row_num, ws, 'TACACS_Server', templateVars['TACACS_Server'])
            validating.login_domain(row_num, ws, 'Login_Domain', templateVars['Login_Domain'])
            validating.secret(row_num, ws, 'Shared_Secret', templateVars['Shared_Secret'])
            validating.retry(row_num, ws, 'Retry_Interval', templateVars['Retry_Interval'])
            validating.timeout(row_num, ws, 'Timeout', templateVars['Timeout'])
            templateVars['Mgmt_Domain'] = validating.mgmt_domain(row_num, ws, 'Mgmt_Domain', templateVars['Mgmt_Domain'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)
        
        templateVars['TACACS_Server_'] = templateVars['TACACS_Server'].replace('.', '-')

        # Define the Template Source
        template_file = "tacacs.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'tacacs_%s.tf' % (templateVars['TACACS_Server_'])
        dest_dir = 'Admin'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "tacacs_src.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'tacacs_src.tf'
        dest_dir = 'Admin'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the "Excel Spreadsheet Guidance" PDF File  
    # for Detailed information on the Arguments used by this Method.
    def web_security(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Passwd_Strength': '',
                         'Enforce_Intv': '',
                         'Passwd_Intv': '',
                         'Number_Allowed': '',
                         'Passwd_Store': '',
                         'Lockout': '',
                         'Failed_Attempts': '',
                         'Time_Period': '',
                         'Dur_Lockout': '',
                         'Token_Timeout': '',
                         'Maximum_Valid': '',
                         'Web_Timeout': ''}
        optional_args = { }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.enable(row_num, ws, 'Enforce_Intv', templateVars['Enforce_Intv'])
            validating.enable(row_num, ws, 'Lockout', templateVars['Lockout'])
            validating.noyes(row_num, ws, 'Passwd_Strength', templateVars['Passwd_Strength'])
            validating.number_check(row_num, ws, 'Passwd_Intv', templateVars['Passwd_Intv'], 0, 745)
            validating.number_check(row_num, ws, 'Number_Allowed', templateVars['Number_Allowed'], 0, 10)
            validating.number_check(row_num, ws, 'Passwd_Store', templateVars['Passwd_Store'], 0, 15)  
            validating.number_check(row_num, ws, 'Failed_Attempts', templateVars['Failed_Attempts'], 1, 15)
            validating.number_check(row_num, ws, 'Time_Period', templateVars['Time_Period'], 1, 720)  
            validating.number_check(row_num, ws, 'Dur_Lockout', templateVars['Dur_Lockout'], 1, 1440)  
            validating.number_check(row_num, ws, 'Token_Timeout', templateVars['Token_Timeout'], 300, 9600)  
            validating.number_check(row_num, ws, 'Maximum_Valid', templateVars['Maximum_Valid'], 0, 24)  
            validating.number_check(row_num, ws, 'Web_Timeout', templateVars['Web_Timeout'], 60, 65525)  
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)
        
        # Define the Template Source
        template_file = "web_security.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'web_security.tf'
        dest_dir = 'Admin'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

# Terraform ACI Provider - Fabric Policies
# Class must be instantiated with Variables
class Fabric_Policies(object):
    def __init__(self, ws):
        self.templateLoader = jinja2.FileSystemLoader(
            searchpath=(aci_template_path + 'Fabric_Policies/'))
        self.templateEnv = jinja2.Environment(loader=self.templateLoader)
    
    # Method must be called with the following kwargs.
    # Please Refer to the "Excel Spreadsheet Guidance" PDF File  
    # for Detailed information on the Arguments used by this Method.
    def bgp_rr(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Node_ID': ''}
        optional_args = {}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.node_id(row_num, ws, 'Node_ID', templateVars['Node_ID'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)
        
        # Define the Template Source
        template_file = "bgp_rr.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'bgp_rr_%s.tf' % (templateVars['Node_ID'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the "Excel Spreadsheet Guidance" PDF File  
    # for Detailed information on the Arguments used by this Method.
    def dns(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'DNS_Server': '',
                         'Preferred': ''}
        optional_args = {}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.ipv4(row_num, ws, 'DNS_Server', templateVars['DNS_Server'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)
        
        templateVars['DNS_Server_'] = templateVars['DNS_Server'].replace('.', '-')

        # Define the Template Source
        template_file = "dns.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'dns_%s.tf' % (templateVars['DNS_Server_'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the "Excel Spreadsheet Guidance" PDF File  
    # for Detailed information on the Arguments used by this Method.
    def dns_mgmt(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Mgmt_Domain': ''}
        optional_args = {}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            templateVars['Mgmt_Domain'] = validating.mgmt_domain(row_num, ws, 'Mgmt_Domain', templateVars['Mgmt_Domain'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)
        
        # Define the Template Source
        template_file = "dns_mgmt.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'dns_mgmt.tf'
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the "Excel Spreadsheet Guidance" PDF File  
    # for Detailed information on the Arguments used by this Method.
    def domain(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Domain': '',
                         'Default_Domain': ''}
        optional_args = {}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.domain(row_num, ws, 'Domain', templateVars['Domain'])
            validating.noyes(row_num, ws, 'Default_Domain', templateVars['Default_Domain'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)
        
        templateVars['Domain_'] = templateVars['Domain'].replace('.', '-')

        # Define the Template Source
        template_file = "domain.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'domain_%s.tf' % (templateVars['Domain_'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the "Excel Spreadsheet Guidance" PDF File  
    # for Detailed information on the Arguments used by this Method.
    def ntp(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'NTP_Server': '',
                         'Preferred': '',
                         'Mgmt_Domain': ''}
        optional_args = {}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.ipv4(row_num, ws, 'NTP_Server', templateVars['NTP_Server'])
            validating.true(row_num, ws, 'Preferred', templateVars['Preferred'])
            templateVars['Mgmt_Domain'] = validating.mgmt_domain(row_num, ws, 'Mgmt_Domain', templateVars['Mgmt_Domain'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)
        
        templateVars['NTP_Server_'] = templateVars['NTP_Server'].replace('.', '-')

        # Define the Template Source
        template_file = "ntp.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'ntp_%s.tf' % (templateVars['NTP_Server_'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the "Excel Spreadsheet Guidance" PDF File  
    # for Detailed information on the Arguments used by this Method.
    def smartcallhome(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'SMTP_Port': '',
                         'SMTP_Relay': '',
                         'Mgmt_Domain': '',
                         'From_Email': '',
                         'Reply_Email': '',
                         'To_Email': ''}
        optional_args = {'Phone_Number': '',
                         'Contact_Info': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.email(row_num, ws, 'From_Email', templateVars['From_Email'])
            validating.email(row_num, ws, 'Reply_Email', templateVars['Reply_Email'])
            validating.email(row_num, ws, 'To_Email', templateVars['To_Email'])
            if not templateVars['Phone_Number'] == None:
                validating.phone(row_num, ws, 'Phone_Number', templateVars['Phone_Number'])
            templateVars['Mgmt_Domain'] = validating.mgmt_domain(row_num, ws, 'Mgmt_Domain', templateVars['Mgmt_Domain'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)
        
        # Define the Template Source
        template_file = "smartcallhome.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'smartcallhome.tf'
        dest_dir = 'Fabric'
        if re.search('Grp_[A-F]', templateVars['Site_Group']):
            Group_ID = '%s' % (templateVars['Site_Group'])
            site_group = ast.literal_eval(os.environ[Group_ID])
            for x in range(1, 13):
                sitex = 'Site_%s' % (x)
                if not site_group[sitex] == None:
                    Site_ID = 'Site_ID_%s' % (site_group[sitex])
                    site_dict = ast.literal_eval(os.environ[Site_ID])

                    # Create templateVars for Site_Name and APIC_URL
                    templateVars['Site_Name'] = site_dict.get('Site_Name')
                    templateVars['APIC_URL'] = site_dict.get('APIC_URL')
                    templateVars['Street_Address'] = site_dict.get('Street_Address')
                    templateVars['Contract_ID'] = site_dict.get('Contract_ID')
                    templateVars['Customer_Identifier'] = site_dict.get('Customer_Identifier')
                    templateVars['Site_Identifier'] = site_dict.get('Site_Identifier')

                    # Create Terraform file from Template
                    create_tf_file('w', dest_dir, dest_file, template, **templateVars)

        elif re.search(r'\d+', templateVars['Site_Group']):
            Site_ID = 'Site_ID_%s' % (templateVars['Site_Group'])
            site_dict = ast.literal_eval(os.environ[Site_ID])

            # Create templateVars for Site_Name and APIC_URL
            templateVars['Site_Name'] = site_dict.get('Site_Name')
            templateVars['APIC_URL'] = site_dict.get('APIC_URL')
            templateVars['Street_Address'] = site_dict.get('Street_Address')
            templateVars['Contract_ID'] = site_dict.get('Contract_ID')
            templateVars['Customer_Identifier'] = site_dict.get('Customer_Identifier')
            templateVars['Site_Identifier'] = site_dict.get('Site_Identifier')

            # Create Terraform file from Template
            create_tf_file('w', dest_dir, dest_file, template, **templateVars)
        else:
            print(f"\n-----------------------------------------------------------------------------\n")
            print(f"   Error on Worksheet {ws.title}, Row {row_num} Site_Group, value {templateVars['Site_Group']}.")
            print(f"   Unable to Determine if this is a Single or Group of Site(s).  Exiting....")
            print(f"\n-----------------------------------------------------------------------------\n")
            exit()

    # Method must be called with the following kwargs.
    # Please Refer to the "Excel Spreadsheet Guidance" PDF File  
    # for Detailed information on the Arguments used by this Method.
    def snmp_client(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'SNMP_Client_Name': '',
                         'SNMP_Client': '',
                         'Mgmt_Domain': ''}
        optional_args = { }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.ipv4(row_num, ws, 'SNMP_Client', templateVars['SNMP_Client'])
            templateVars['Mgmt_Domain'] = validating.snmp_mgmt(row_num, ws, 'Mgmt_Domain', templateVars['Mgmt_Domain'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)
        
        templateVars['SNMP_Client_'] = templateVars['SNMP_Client'].replace('.', '-')

        # Define the Template Source
        template_file = "snmp_client.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'snmp_client_%s.tf' % (templateVars['SNMP_Client_'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the "Excel Spreadsheet Guidance" PDF File  
    # for Detailed information on the Arguments used by this Method.
    def snmp_comm(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'SNMP_Community': ''}
        optional_args = {'Description': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.snmp_string(row_num, ws, 'SNMP_Community', templateVars['SNMP_Community'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)
        
        # Define the Template Source
        template_file = "snmp_comm.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'snmp_comm_%s.tf' % (templateVars['SNMP_Community'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the "Excel Spreadsheet Guidance" PDF File  
    # for Detailed information on the Arguments used by this Method.
    def snmp_info(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'SNMP_Contact': ''}
        optional_args = { }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.snmp_info(row_num, ws, 'SNMP_Contact', templateVars['SNMP_Contact'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)
        
        # Define the Template Source
        template_file = "snmp_info.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'snmp_info.tf'
        dest_dir = 'Fabric'
        if re.search('Grp_[A-F]', templateVars['Site_Group']):
            Group_ID = '%s' % (templateVars['Site_Group'])
            site_group = ast.literal_eval(os.environ[Group_ID])
            for x in range(1, 13):
                sitex = 'Site_%s' % (x)
                if not site_group[sitex] == None:
                    Site_ID = 'Site_ID_%s' % (site_group[sitex])
                    site_dict = ast.literal_eval(os.environ[Site_ID])

                    # Create templateVars for Site_Name and APIC_URL
                    templateVars['Site_Name'] = site_dict.get('Site_Name')
                    templateVars['APIC_URL'] = site_dict.get('APIC_URL')
                    templateVars['SNMP_Location'] = site_dict.get('SNMP_Location')

                    # Create Terraform file from Template
                    create_tf_file('w', dest_dir, dest_file, template, **templateVars)

        elif re.search(r'\d+', templateVars['Site_Group']):
            Site_ID = 'Site_ID_%s' % (templateVars['Site_Group'])
            site_dict = ast.literal_eval(os.environ[Site_ID])

            # Create templateVars for Site_Name and APIC_URL
            templateVars['Site_Name'] = site_dict.get('Site_Name')
            templateVars['APIC_URL'] = site_dict.get('APIC_URL')
            templateVars['SNMP_Location'] = site_dict.get('SNMP_Location')

            # Create Terraform file from Template
            create_tf_file('w', dest_dir, dest_file, template, **templateVars)
        else:
            print(f"\n-----------------------------------------------------------------------------\n")
            print(f"   Error on Worksheet {ws.title}, Row {row_num} Site_Group, value {templateVars['Site_Group']}.")
            print(f"   Unable to Determine if this is a Single or Group of Site(s).  Exiting....")
            print(f"\n-----------------------------------------------------------------------------\n")
            exit()
    
    # Method must be called with the following kwargs.
    # Please Refer to the "Excel Spreadsheet Guidance" PDF File  
    # for Detailed information on the Arguments used by this Method.
    def snmp_trap(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Trap_Server': '',
                         'Destination_Port': '',
                         'Version': '',
                         'Community_or_Username': '',
                         'Security_Level': '',
                         'Mgmt_Domain': ''}
        optional_args = { }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        # Set noauth if v1 or v2c
        if re.search('(v1|v2c)', templateVars['Version']):
            templateVars['Security_Level'] = 'noauth'
    
        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.ipv4(row_num, ws, 'Trap_Server', templateVars['Trap_Server'])
            validating.port(row_num, ws, 'Destination_Port', templateVars['Destination_Port'])
            validating.snmp_ver(row_num, ws, 'Version', templateVars['Version'])
            validating.snmp_sec(row_num, ws, 'Security_Level', templateVars['Security_Level'])
            validating.snmp_string(row_num, ws, 'Community_or_Username', templateVars['Community_or_Username'])
            templateVars['Mgmt_Domain'] = validating.mgmt_domain(row_num, ws, 'Mgmt_Domain', templateVars['Mgmt_Domain'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)
        
        templateVars['Trap_Server_'] = templateVars['Trap_Server'].replace('.', '-')

        # Define the Template Source
        template_file = "snmp_trap.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'snmp_trap_%s.tf' % (templateVars['Trap_Server_'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the "Excel Spreadsheet Guidance" PDF File  
    # for Detailed information on the Arguments used by this Method.
    def snmp_user(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'SNMP_User': '',
                         'Authorization_Type': '',
                         'Authorization_Key': ''}
        optional_args = {'Privacy_Type': '',
                         'Privacy_Key': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            auth_type = templateVars['Authorization_Type']
            auth_key = templateVars['Authorization_Key']
            validating.snmp_auth(row_num, ws, templateVars['Privacy_Type'], templateVars['Privacy_Key'], auth_type, auth_key)
            validating.snmp_string(row_num, ws, 'SNMP_User', templateVars['SNMP_User'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)
        
        # Modify User Input of templateVars['Privacy_Type'] or templateVars['Authorization_Type'] to send to APIC
        if templateVars['Privacy_Type'] == 'none':
            templateVars['Privacy_Type'] = None
        if templateVars['Authorization_Type'] == 'md5':
            templateVars['Authorization_Type'] = None
        if templateVars['Authorization_Type'] == 'sha1':
            templateVars['Authorization_Type'] = 'hmac-sha1-96'
    
        # Define the Template Source
        template_file = "snmp_user.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'snmp_user_%s.tf' % (templateVars['SNMP_User'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the "Excel Spreadsheet Guidance" PDF File  
    # for Detailed information on the Arguments used by this Method.
    def syslog_dg(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Dest_Grp_Name': '',
                         'Minimum_Level': '',
                         'Log_Format': '',
                         'Console': '',
                         'Console_Level': '',
                         'Local': '',
                         'Local_Level': '',
                         'Include_msec': '',
                         'Include_timezone': '',
                         'Audit': '',
                         'Events': '',
                         'Faults': '',
                         'Session': ''}
        optional_args = { }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.enabled(row_num, ws, 'Console', templateVars['Console'])
            validating.enabled(row_num, ws, 'Local', templateVars['Local'])
            validating.log_level(row_num, ws, 'Minimum_Level', templateVars['Minimum_Level'])
            validating.log_level(row_num, ws, 'Local_Level', templateVars['Local_Level'])
            validating.log_level(row_num, ws, 'Console_Level', templateVars['Console_Level'])
            validating.true(row_num, ws, 'Include_msec', templateVars['Include_msec'])
            validating.true(row_num, ws, 'Include_timezone', templateVars['Include_timezone'])
            validating.true(row_num, ws, 'Audit', templateVars['Audit'])
            validating.true(row_num, ws, 'Events', templateVars['Events'])
            validating.true(row_num, ws, 'Faults', templateVars['Faults'])
            validating.true(row_num, ws, 'Session', templateVars['Session'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)
        incl_list = ''
        if not templateVars['Audit'] == 'false':
            incl_list = 'audit'
        if not templateVars['Events'] == 'false':
            if incl_list == '':
                incl_list = 'events'
            else:
                incl_list = incl_list + ',events'
        if not templateVars['Faults'] == 'false':
            if incl_list == '':
                incl_list = 'faults'
            else:
                incl_list = incl_list + ',faults'
        if not templateVars['Session'] == 'false':
            if incl_list == '':
                incl_list = 'session'
            else:
                incl_list = incl_list + ',session'
        templateVars['Included_Types'] = incl_list

        # Define the Template Source
        template_file = "syslog_dg.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'syslog_dg.tf'
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the "Excel Spreadsheet Guidance" PDF File  
    # for Detailed information on the Arguments used by this Method.
    def syslog_rmt(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Dest_Grp_Name': '',
                         'Syslog_Server': '',
                         'Port': '',
                         'Mgmt_Domain': '',
                         'Severity': '',
                         'Facility': ''}
        optional_args = { }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.ipv4(row_num, ws, 'Syslog_Server', templateVars['Syslog_Server'])
            validating.log_level(row_num, ws, 'Severity', templateVars['Severity'])
            validating.port(row_num, ws, 'Port', templateVars['Port'])
            validating.syslog_fac(row_num, ws, 'Facility', templateVars['Facility'])
            templateVars['Mgmt_Domain'] = validating.mgmt_domain(row_num, ws, 'Mgmt_Domain', templateVars['Mgmt_Domain'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)
        
        templateVars['Syslog_Server_'] = templateVars['Syslog_Server'].replace('.', '-')

        # Define the Template Source
        template_file = "syslog_rmt.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'syslog_rmt_%s.tf' % (templateVars['Syslog_Server_'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

# Terraform ACI Provider - Tenants Policies
# Class must be instantiated with Variables
class L3Out_Policies(object):
    def __init__(self, ws):
        self.templateLoader = jinja2.FileSystemLoader(
            searchpath=(aci_template_path + 'Tenant_Policies/'))
        self.templateEnv = jinja2.Environment(loader=self.templateLoader)

    # Method must be called with the following kwargs.
    # Please Refer to the "Excel Spreadsheet Guidance" PDF File  
    # for Detailed information on the Arguments used by this Method.
    def add_l3out(self, wb, ws, row_num, **kwargs):
        # Open the Network Policies Worksheet
        ws_net = wb['Network Policies']
        rowcount = ws_net.max_row

        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Tenant': '',
                         'VRF': '',
                         'Name': '',
                         'L3Out_Policy': '',
                         'L3_Domain': '',
                         'Ext_EPG': '',
                         'Ext_EPG_Policy': '',
                         'Subnet': '',
                         'Ext_Subnet_Policy': '',
                         'target_dscp': '',
                         'enforce_rtctrl': '',
                         'prio': '',
                         'epg_target_dscp': '',
                         'pref_gr_memb': '',
                         'match_t': '',
                         'flood': '',
                         'export-rtctrl': '',
                         'import-rtctrl': '',
                         'import-security': '',
                         'shared-security': '',
                         'shared-rtctrl': '',
                         'agg-export': '',
                         'agg-import': '',
                         'agg-shared': ''}
        optional_args = {'Description': '',
                         'EPG_Description': '',	
                         'annotation': '',
                         'name_alias': '',
                         'leak_rtctrlProfile': '',
                         'damp_rtctrlProfile': '',
                         'fvBDPublicSubnetHolder': '',	
                         'epg_annotation': '',
                         'epg_name_alias': '',
                         'cons_vzBrCP': '',
                         'vzCPIf': '',
                         'Master_fvEPg': '',
                         'prov_vzBrCP': '',
                         'vzTaboo': '',
                         'exception_tag': '',
                         'rtctrlProfile': '',
                         'sub_annotation': '',
                         'sub_name_alias': '',
                         'sub_rtctrlProfile': '',
                         'rtsumARtSummPol': ''}


        # Get the L3Out Policies from the Network Policies Tab
        func = 'L3Out_Policy'
        count = countKeys(ws_net, func)
        l3_count = ''
        var_dict = findVars(ws_net, func, rowcount, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('L3Out_Policy'):
                l3_count = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}

        func = 'Ext_EPG_Policy'
        count = countKeys(ws_net, func)
        epg_count = ''
        var_dict = findVars(ws_net, func, rowcount, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('Ext_EPG_Policy'):
                epg_count = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}

        func = 'Ext_Subnet_Policy'
        count = countKeys(ws_net, func)
        sub_count = ''
        var_dict = findVars(ws_net, func, rowcount, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('Ext_Subnet_Policy'):
                sub_count = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
            validating.name_rule(row_num, ws, 'VRF', templateVars['VRF'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            if not templateVars['Subnet'] == None:
                if re.search(',', templateVars['Subnet']):
                    sx = templateVars['Subnet'].split(',')
                    for x in sx:
                        validating.ip_address(row_num, ws, 'Subnet', x)
            validating.dscp(l3_count, ws_net, 'target_dscp', templateVars['target_dscp'])
            validating.export(l3_count, ws_net, 'enforce_rtctrl', templateVars['enforce_rtctrl'])
            validating.dscp(epg_count, ws_net, 'epg_target_dscp', templateVars['epg_target_dscp'])
            validating.enabled(epg_count, ws_net, 'flood', templateVars['flood'])
            validating.include(epg_count, ws_net, 'pref_gr_memb', templateVars['pref_gr_memb'])
            validating.match_t(epg_count, ws_net, 'match_t', templateVars['match_t'])
            validating.qos_priority(epg_count, ws_net, 'prio', templateVars['prio'])
            validating.noyes(sub_count, ws_net, 'agg-export', templateVars['agg-export'])
            validating.noyes(sub_count, ws_net, 'agg-import', templateVars['agg-import'])
            validating.noyes(sub_count, ws_net, 'agg-shared', templateVars['agg-shared'])
            validating.noyes(sub_count, ws_net, 'export-rtctrl', templateVars['export-rtctrl'])
            validating.noyes(sub_count, ws_net, 'import-rtctrl', templateVars['import-rtctrl'])
            validating.noyes(sub_count, ws_net, 'import-security', templateVars['import-security'])
            validating.noyes(sub_count, ws_net, 'shared-security', templateVars['shared-security'])
            validating.noyes(sub_count, ws_net, 'shared-rtctrl', templateVars['shared-rtctrl'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Create aggregate templateVars
        aggregate = ''
        if templateVars['agg-export'] == 'yes':
            aggregate = aggregate + '"export-rtctrl"'
        if templateVars['agg-import'] == 'yes':
            aggregate = aggregate + ', ' + '"import-rtctrl"'
        if templateVars['agg-shared'] == 'yes':
            aggregate = aggregate + ', ' + '"shared-rtctrl"'

        else:
            templateVars['aggregate'] = '[%s]' % (aggregate)
        
        # Create scope templateVars
        scope = ''
        if templateVars['export-rtctrl'] == 'yes':
            scope = scope + '"export-rtctrl"'
        if templateVars['import-rtctrl'] == 'yes':
            scope = scope + ', ' + '"import-rtctrl"'
        if templateVars['import-security'] == 'yes':
            scope = scope + ', ' + '"import-security"'
        if templateVars['shared-security'] == 'yes':
            scope = scope + ', ' + '"shared-security"'
        if templateVars['shared-rtctrl'] == 'yes':
            scope = scope + ', ' + '"shared-rtctrl"'

        else:
            templateVars['scope'] = '[%s]' % (scope)
        
        # Define the Template Source
        template_file = "l3out.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'l3out_%s.tf' % (templateVars['Name'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "ext_epg.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'l3out_%s_epg_%s.tf' % (templateVars['Name'], templateVars['Ext_EPG'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        if re.search(',', templateVars['Subnet']):
            sx = templateVars['Subnet'].split(',')
            for x in sx:
                templateVars['Subnet'] = x
                templateVars['Subnet_'] = x.replace('.', '-')
                templateVars['Subnet_'] = x.replace('/', '_')
                
                # Define the Template Source
                template_file = "ext_subnet.template"
                template = self.templateEnv.get_template(template_file)

                # Process the template through the Sites
                dest_file = 'l3out_%s_epg_%s_subnet_%s.tf' % (templateVars['Name'], templateVars['Ext_EPG'], templateVars['Subnet'])
                dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
                process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)
        else:
            templateVars['Subnet_'] = templateVars['Subnet'].replace('.', '-')
            templateVars['Subnet_'] = templateVars['Subnet'].replace('/', '_')

            # Define the Template Source
            template_file = "ext_subnet.template"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = 'l3out_%s_epg_%s_subnet_%s.tf' % (templateVars['Name'], templateVars['Ext_EPG'], templateVars['Subnet'])
            dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
            process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the "Excel Spreadsheet Guidance" PDF File  
    # for Detailed information on the Arguments used by this Method.
    def intf_prof(self, wb, ws, row_num, **kwargs):
        # Open the Network Policies Worksheet
        ws_net = wb['Network Policies']
        rowcount = ws_net.max_row

        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'L3_Out': '',
                         'NodeP_Name': '',
                         'Interface_Type': '',
                         'Interface_Policies': '',
                         'Node1_ID': '',
                         'Node1_Intf': '',
                         'Node1_IP': '',
                         'prio': ''}
        optional_args = {'Description': '',
                         'Node2_ID': '',
                         'Node2_Intf': '',
                         'VLAN': '',
                         'Node2_IP': '',
                         'BGP_Profile': '',
                         'EIGRP_Profile': '',
                         'OSPF_Profile': '',
                         'annotation': '',	
                         'name_alias': '',	
                         'mtu': '',	
                         'tag': '',	
                         'arpIfPol': '',	
                         'egress_qosDppPol': '',	
                         'ingress_qosDppPol': '',	
                         'qosCustomPol': '',	
                         'ndIfPol': '',	
                         'netflowMonitorPol': ''}	

        # Get the Node Policies from the Network Policies Tab
        func = 'Interface_Policies'
        count = countKeys(ws_net, func)
        row_count = ''
        var_dict = findVars(ws_net, func, rowcount, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('Interface_Policies'):
                row_count = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.node_id(row_num, ws, 'Node1_ID', templateVars['Node1_ID'])
            if not templateVars['Node2_ID'] == None:
                validating.node_id(row_num, ws, 'Node2_ID', templateVars['Node2_ID'])
                validating.ip_address(row_num, ws, 'Node2_IP', templateVars['Node2_IP'])
            if not templateVars['VLAN'] == None:
                validating.vlans(row_num, ws, 'VLAN', templateVars['VLAN'])
            validating.qos_priority(row_count, ws_net, 'prio', templateVars['prio'])
            if not templateVars['tag'] == None:
                validating.tag_check(row_count, ws_net, 'tag', templateVars['tag'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "nodep.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'l3out_%s_nodep_%s.tf' % (templateVars['L3_Out'], templateVars['NodeP_Name'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "nodep_to_fabric_node.template"
        template = self.templateEnv.get_template(template_file)

        # Modify Variables for Template
        templateVars['Node_ID'] = templateVars['Node1_ID']
        templateVars['rtr_id'] = templateVars['Node1_Rotuer_ID']
        templateVars['rtr_id_loop_back'] = templateVars['Node1_Loopback']
        templateVars['tDn'] = 'topology/pod-%s/node-%s' % (templateVars['Node_ID'], templateVars['Node_ID'])

        # Process the template through the Sites
        dest_file = 'l3out_%s_nodep_%s.tf' % (templateVars['L3_Out'], templateVars['NodeP_Name'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

        if not templateVars['Node2_ID'] == None:
            # Modify Variables for Template
            templateVars['Node_ID'] = templateVars['Node2_ID']
            templateVars['rtr_id'] = templateVars['Node2_Rotuer_ID']
            templateVars['rtr_id_loop_back'] = templateVars['Node2_Loopback']
            templateVars['tDn'] = 'topology/pod-%s/node-%s' % (templateVars['Node_ID'], templateVars['Node_ID'])

            # Process the template through the Sites
            dest_file = 'l3out_%s_nodep_%s.tf' % (templateVars['L3_Out'], templateVars['NodeP_Name'])
            dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
            process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)


    # Method must be called with the following kwargs.
    # Please Refer to the "Excel Spreadsheet Guidance" PDF File  
    # for Detailed information on the Arguments used by this Method.
    def node_prof(self, wb, ws, row_num, **kwargs):
        # Open the Network Policies Worksheet
        ws_net = wb['Network Policies']
        rowcount = ws_net.max_row

        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Pod_ID': '',
                         'NodeP_Name': '',
                         'L3_Out': '',
                         'Node_Policies': '',
                         'Node1_ID': '',
                         'Node1_Rotuer_ID': '',
                         'Node1_Loopback': ''}
        optional_args = {'Description': '',
                         'Node2_ID': '',
                         'Node2_Rotuer_ID': '',
                         'Node2_Loopback': '',
                         'annotation': '',
                         'name_alias': '',
                         'config_issues': '',
                         'dscp': '',
                         'tag': ''}	

        # Get the Node Policies from the Network Policies Tab
        func = 'Node_Policies'
        count = countKeys(ws_net, func)
        row_count = ''
        var_dict = findVars(ws_net, func, rowcount, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('Node_Policies'):
                row_count = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'NodeP_Name', templateVars['NodeP_Name'])
            validating.node_id(row_num, ws, 'Node1_ID', templateVars['Node1_ID'])
            validating.ip_address(row_num, ws, 'Node1_Rotuer_ID', templateVars['Node1_Rotuer_ID'])
            validating.noyes(row_num, ws, 'Node1_Loopback', templateVars['Node1_Loopback'])
            if not templateVars['Node2_ID'] == None:
                validating.node_id(row_num, ws, 'Node2_ID', templateVars['Node2_ID'])
                validating.ip_address(row_num, ws, 'Node2_Rotuer_ID', templateVars['Node2_Rotuer_ID'])
                validating.noyes(row_num, ws, 'Node2_Loopback', templateVars['Node2_Loopback'])
            validating.dscp(row_count, ws_net, 'dscp', templateVars['dscp'])
            validating.tag_check(row_count, ws_net, 'tag', templateVars['tag'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "nodep.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'l3out_%s_nodep_%s.tf' % (templateVars['L3_Out'], templateVars['NodeP_Name'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "nodep_to_fabric_node.template"
        template = self.templateEnv.get_template(template_file)

        # Modify Variables for Template
        templateVars['Node_ID'] = templateVars['Node1_ID']
        templateVars['rtr_id'] = templateVars['Node1_Rotuer_ID']
        templateVars['rtr_id_loop_back'] = templateVars['Node1_Loopback']
        templateVars['tDn'] = 'topology/pod-%s/node-%s' % (templateVars['Node_ID'], templateVars['Node_ID'])

        # Process the template through the Sites
        dest_file = 'l3out_%s_nodep_%s.tf' % (templateVars['L3_Out'], templateVars['NodeP_Name'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

        if not templateVars['Node2_ID'] == None:
            # Modify Variables for Template
            templateVars['Node_ID'] = templateVars['Node2_ID']
            templateVars['rtr_id'] = templateVars['Node2_Rotuer_ID']
            templateVars['rtr_id_loop_back'] = templateVars['Node2_Loopback']
            templateVars['tDn'] = 'topology/pod-%s/node-%s' % (templateVars['Node_ID'], templateVars['Node_ID'])

            # Process the template through the Sites
            dest_file = 'l3out_%s_nodep_%s.tf' % (templateVars['L3_Out'], templateVars['NodeP_Name'])
            dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
            process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

# Terraform ACI Provider - Site Policies
# Class must be instantiated with Variables
class Site_Policies(object):
    def __init__(self, ws):
        self.templateLoader = jinja2.FileSystemLoader(
            searchpath=(aci_template_path))
        self.templateEnv = jinja2.Environment(loader=self.templateLoader)
    
    # Method must be called with the following kwargs.
    # Site_ID: Required.  Number to Represeent the Site
    # Site_Name: Required.  A Name for the Site.  Must only contain alphanumeric and underscore
    # APIC_URL: Required.  URL for the APIC for the Site
    # BGP_AS: Required.  Autonomous System for BGP Process
    # SNMP_Location: Required.  SNMP Location for the APIC Cluster
    # Contract_ID: Required.  Contract for Equipment to be used with Smart CallHome Function
    # Customer_Identifier: Required.  Customer Identifier to be used with Smart CallHome Function
    # Site_Identifier: Required.  Site Identifier to be used with Smart CallHome Function
    # Street_Address: Optional.  Street Address for the Site to be used with Smart CallHome Function
    def site_id(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_ID': '',
                         'Site_Name': '',
                         'APIC_URL': '',
                         'BGP_AS': '',
                         'SNMP_Location': '',
                         'Contract_ID': '',
                         'Customer_Identifier': '',
                         'Site_Identifier': ''}
        optional_args = {'Street_Address': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate BGP AS Number
            validating.bgp_as(row_num, ws, 'BGP_AS', templateVars['BGP_AS'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Save the Site Information into Environment Variables
        Site_ID = 'Site_ID_%s' % (templateVars['Site_ID'])
        os.environ[Site_ID] = '%s' % (templateVars)

        # Copy the Default Templates to the Appropriate Folders
        copy_defaults(templateVars['Site_Name'], 'Access')
        copy_defaults(templateVars['Site_Name'], 'VLANs')
        copy_defaults(templateVars['Site_Name'], 'Admin')
        copy_defaults(templateVars['Site_Name'], 'Fabric')
        copy_defaults(templateVars['Site_Name'], 'Tenant_common')
        copy_defaults(templateVars['Site_Name'], 'Tenant_infra')
        copy_defaults(templateVars['Site_Name'], 'Tenant_mgmt')

        # Write the variables.tf to the Appropriate Directories
        template_file = "variables.tf"
        template = self.templateEnv.get_template(template_file)
        create_tf_file('w', 'Access', template_file, template, **templateVars)
        create_tf_file('w', 'VLANs', template_file, template, **templateVars)
        create_tf_file('w', 'Admin', template_file, template, **templateVars)
        create_tf_file('w', 'Fabric', template_file, template, **templateVars)
        create_tf_file('w', 'Tenant_common', template_file, template, **templateVars)
        create_tf_file('w', 'Tenant_infra', template_file, template, **templateVars)
        create_tf_file('w', 'Tenant_mgmt', template_file, template, **templateVars)

        # Create Directories and default Terraform Files for Tenants in the Tenants and Networks Tab if Needed
        ws_names = ['Tenants', 'Networks']
        for ws_name in ws_names:
            ws_tenants = wb[ws_name]
            rows = ws_tenants.max_row
            func_regex = re.compile('^add_')
            func_list = findKeys(ws_tenants, func_regex)
            for func in func_list:
                count = countKeys(ws_tenants, func)
                var_dict = findVars(ws_tenants, func, rows, count)
                for pos in var_dict:
                    row_num = var_dict[pos]['row']
                    del var_dict[pos]['row']
                    for x in list(var_dict[pos].keys()):
                        if var_dict[pos][x] == '':
                            del var_dict[pos][x]
                    if not var_dict[pos].get('Tenant') == None:
                        tenant_dir = 'Tenant_%s' % var_dict[pos].get('Tenant')
                        copy_defaults(templateVars['Site_Name'], tenant_dir)
                        create_tf_file('w', tenant_dir, template_file, template, **templateVars)
        
        site_wb = '%s_intf_selectors.xlsx' % (templateVars['Site_Name'])
        if not os.path.isfile(site_wb):
            wb.save(filename=site_wb)
            wb_wr = load_workbook(site_wb)
            ws_wr = wb_wr.get_sheet_names()
            for sheetName in ws_wr:
                if sheetName not in ['Sites']:
                    sheetToDelete = wb_wr.get_sheet_by_name(sheetName)
                    wb_wr.remove_sheet(sheetToDelete)
            wb_wr.save(filename=site_wb)

        # Create TF File for the bgp_as in the Fabric Folder
        template_file = "bgp_as.template"
        template = self.templateEnv.get_template(template_file)
        tf_file = 'bgp_as.tf'
        create_tf_file('w', 'Fabric', tf_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Group: Required.  A Group Name to represent a list of Site_ID's
    # Site_1: Required.  The Site_ID for the First Site
    # Site_2: Required.  The Site_ID for the Second Site
    # Site_[3-12]: Optional.  The Site_ID for the 3rd thru the 12th Site(s)
    def grp_id(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Group': '',
                         'Site_1': '',
                         'Site_2': ''}
        optional_args = {'Site_3': '',
                         'Site_4': '',
                         'Site_5': '',
                         'Site_6': '',
                         'Site_7': '',
                         'Site_8': '',
                         'Site_9': '',
                         'Site_10': '',
                         'Site_11': '',
                         'Site_12': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        for x in range(1, 13):
            site = 'Site_%s' % (x)
            if not templateVars[site] == None:
                validating.site_group(wb, ws, site, templateVars[site])

        grp_count = 0
        for x in list(map(chr, range(ord('A'), ord('F')+1))):
            grp = 'Grp_%s' % (x)
            if templateVars['Group'] == grp:
                grp_count += 1
        if grp_count == 0:
            print(f'\n-----------------------------------------------------------------------------\n')
            print(f'   Error on Worksheet {ws.title}, Row {row_num} Group, Group_Name "{templateVars["Group"]}" is invalid.')
            print(f'   A valid Group Name is Grp_A thru Grp_F.  Exiting....')
            print(f'\n-----------------------------------------------------------------------------\n')
            exit()

        # Save the Site Information into Environment Variables
        Group_ID = '%s' % (templateVars['Group'])
        os.environ[Group_ID] = '%s' % (templateVars)

# Terraform ACI Provider - Tenants Policies
# Class must be instantiated with Variables
class Tenant_Policies(object):
    def __init__(self, ws):
        self.templateLoader = jinja2.FileSystemLoader(
            searchpath=(aci_template_path + 'Tenant_Policies/'))
        self.templateEnv = jinja2.Environment(loader=self.templateLoader)

    # Method must be called with the following kwargs.
    # Please Refer to the "Excel Spreadsheet Guidance" PDF File  
    # for Detailed information on the Arguments used by this Method.
    def add_app(self, wb, ws, row_num, **kwargs):
        # Open the Network Policies Worksheet
        ws_net = wb['Network Policies']
        rowcount = ws_net.max_row

        # Dicts for Application Profile; required and optional args
        required_args = {'Site_Group': '',
                        'Tenant': '',
                        'App_Profile': '',
                        'App_Policy': '',
                        'Policy_Name': '',
                        'prio': '',
                        'monEPGPol': ''}
        optional_args = {'annotation': '',
                        'name_alias': ''}

        # Get the Application Profile Policies from the Network Policies Tab
        func = 'app'
        count = countKeys(ws_net, func)
        row_count = ''
        var_dict = findVars(ws_net, func, rowcount, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('App_Policy'):
                row_count = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}
                break

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
            validating.name_rule(row_num, ws, 'App_Profile', templateVars['App_Profile'])
            validating.qos_priority(row_count, ws_net, 'prio', templateVars['prio'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)
        
        if templateVars['monEPGPol'] == 'default':
            templateVars['monEPGPol'] = 'uni/tn-common/monepg-default'

        if re.search('^(common|mgmt|infra)$', templateVars['Tenant']):
            templateVars['Tenant_Dn'] = 'data.aci_tenant.%s' % (templateVars['Tenant'])
        else:
            templateVars['Tenant_Dn'] = 'aci_tenant.%s' % (templateVars['Tenant'])

        # Define the Template Source
        template_file = "app.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'app_%s.tf' % (templateVars['App_Profile'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the "Excel Spreadsheet Guidance" PDF File  
    # for Detailed information on the Arguments used by this Method.
    def add_bd(self, wb, ws, row_num, **kwargs):
        # Open the Network Policies Worksheet
        ws_net = wb['Network Policies']
        rowcount = ws_net.max_row

        # Dicts for Bridge Domain required and optional args
        required_args = {'Site_Group': '',
                        'Tenant': '',
                        'Bridge_Domain': '',
                        'BD_Policy': '',
                        'VRF': '',
                        'VRF_Tenant': '',
                        'Policy_Name': '',
                        'bd_type': '',
                        'host_routing': '',
                        'ep_clear': '',
                        'ep_move': '',
                        'unk_mac': '',
                        'unk_mcast': '',
                        'v6unk_mcast': '',
                        'multi_dst': '',
                        'mcast_allow': '',
                        'ipv6_mcast': '',
                        'arp_flood': '',
                        'limit_learn': '',
                        'fvEpRetPol': '',
                        'unicast_route': '',
                        'intersight_l2': '',
                        'intersight_bum': '',
                        'optimize_wan': '',
                        'monEPGPol': '',
                        'ip_learning': ''}
        optional_args = {'BD_Description': '',
                        'annotation': '',
                        'name_alias': '',
                        'dhcpRelayP': '',
                        'igmpIfPol': '',
                        'igmpSnoopPol': '',
                        'mldSnoopPol': '',
                        'mac': '',
                        'l3extOut': '',
                        'rtctrlProfile': '',
                        'ndIfPol': '',
                        'll_addr': '',
                        'fhsBDPol': '',
                        'netflowMonitorPol': ''}

        # Get the BD Policies from the Network Policies Tab
        func = 'bd'
        row_count = ''
        count = countKeys(ws_net, func)
        var_dict = findVars(ws_net, func, rowcount, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('BD_Policy'):
                row_count = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}
                break

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
            validating.name_rule(row_num, ws, 'VRF', templateVars['VRF'])
            validating.name_rule(row_num, ws, 'VRF_Tenant', templateVars['VRF_Tenant'])
            validating.name_rule(row_num, ws, 'Bridge_Domain', templateVars['Bridge_Domain'])
            validating.bd_type(row_count, ws_net, 'bd_type', templateVars['bd_type'])
            validating.flood(row_count, ws_net, 'unk_mcast', templateVars['unk_mcast'])
            validating.flood(row_count, ws_net, 'v6unk_mcast', templateVars['v6unk_mcast'])
            validating.flood_bd(row_count, ws_net, 'multi_dst', templateVars['multi_dst'])
            validating.garp(row_count, ws_net, 'ep_move', templateVars['ep_move'])
            validating.noyes(row_count, ws_net, 'ep_clear', templateVars['ep_clear'])
            validating.noyes(row_count, ws_net, 'host_routing', templateVars['host_routing'])
            validating.noyes(row_count, ws_net, 'mcast_allow', templateVars['mcast_allow'])
            validating.noyes(row_count, ws_net, 'ipv6_mcast', templateVars['ipv6_mcast'])
            validating.noyes(row_count, ws_net, 'arp_flood', templateVars['arp_flood'])
            validating.noyes(row_count, ws_net, 'limit_learn', templateVars['limit_learn'])
            validating.noyes(row_count, ws_net, 'unicast_route', templateVars['unicast_route'])
            validating.noyes(row_count, ws_net, 'limit_learn', templateVars['limit_learn'])
            validating.noyes(row_count, ws_net, 'intersight_l2', templateVars['intersight_l2'])
            validating.noyes(row_count, ws_net, 'intersight_bum', templateVars['intersight_bum'])
            validating.noyes(row_count, ws_net, 'optimize_wan', templateVars['optimize_wan'])
            validating.noyes(row_count, ws_net, 'ip_learning', templateVars['ip_learning'])
            validating.proxy(row_count, ws_net, 'unk_mac', templateVars['unk_mac'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)
        
        if templateVars['dhcpRelayP'] == 'default':
            templateVars['dhcpRelayP'] = 'uni/tn-common/relayp-default'
        if templateVars['fhsBDPol'] == 'default':
            templateVars['fhsBDPol'] = 'uni/tn-common/bdpol-default'
        if templateVars['fvEpRetPol'] == 'default':
            templateVars['fvEpRetPol'] = 'uni/tn-common/epRPol-default'
        if templateVars['igmpIfPol'] == 'default':
            templateVars['igmpIfPol'] = 'uni/tn-common/igmpIfPol-default'
        if templateVars['igmpSnoopPol'] == 'default':
            templateVars['igmpSnoopPol'] = 'uni/tn-common/snPol-default'
        if templateVars['mldSnoopPol'] == 'default':
            templateVars['mldSnoopPol'] = 'uni/tn-common/mldsnoopPol-default'
        if templateVars['monEPGPol'] == 'default':
            templateVars['monEPGPol'] = 'uni/tn-common/monepg-default'
        if templateVars['ndIfPol'] == 'default':
            templateVars['ndIfPol'] = 'uni/tn-common/ndifpol-default'
        if templateVars['netflowMonitorPol'] == 'default':
            templateVars['netflowMonitorPol'] = 'uni/tn-common/monitorpol-default'

        if re.search('^(common|mgmt|infra)$', templateVars['Tenant']):
            templateVars['Tenant_Dn'] = 'data.aci_tenant.%s' % (templateVars['Tenant'])
        else:
            templateVars['Tenant_Dn'] = 'aci_tenant.%s' % (templateVars['Tenant'])

        if not templateVars['Tenant'] == templateVars['VRF_Tenant']:
            templateVars['vrfDn'] = 'data.aci_tenant.%s,data.aci_vrf.%s' % (templateVars['VRF_Tenant'], templateVars['VRF'])
            templateVars['rel_VRF'] = 'data.aci_vrf.%s' % (templateVars['VRF'])
            # Define the Template Source
            template_file = "data_vrf.template"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = 'data_tenant_%s_vrf_%s.tf' % (templateVars['VRF_Tenant'], templateVars['VRF'])
            dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
            process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

            # Process the template through the Sites
            templateVars['data_Tenant'] = templateVars['VRF_Tenant']
            template_file = "data_tenant.template"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = 'data_tenant_%s.tf' % (templateVars['VRF_Tenant'])
            dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
            process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        else:
            templateVars['vrfDn'] = 'aci_vrf.%s' % (templateVars['VRF'])
            templateVars['rel_VRF'] = 'aci_vrf.%s' % (templateVars['VRF'])


        # Define the Template Source
        template_file = "bd.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'bd_%s.tf' % (templateVars['Bridge_Domain'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the "Excel Spreadsheet Guidance" PDF File  
    # for Detailed information on the Arguments used by this Method.
    def add_epg(self, wb, ws, row_num, **kwargs):
        # Open the Network Policies Worksheet
        ws_net = wb['Network Policies']
        rowcount = ws_net.max_row

        # Dicts for Bridge Domain required and optional args
        required_args = {'Site_Group': '',
                        'Tenant': '',
                        'Bridge_Domain': '',
                        'App_Profile': '',
                        'EPG': '',
                        'EPG_Policy': '',
                        'Policy_Name': '',
                        'is_attr_based': '',
                        'prio': '',
                        'pc_enf_pref': '',
                        'fwd_ctrl': '',
                        'pref_gr_memb': '',
                        'flood': '',
                        'match_t': '',
                        'monEPGPol': '',
                        'shutdown': '',
                        'has_mcast': ''}
        optional_args = {'EPG_Description': '',
                        'VLAN': '',
                        'PVLAN': '',
                        'annotation': '',
                        'name_alias': '',
                        'Physical_Domains': '',
                        'VMM_Domains': '',
                        'cons_vzBrCP': '',
                        'prov_vzBrCP': '',
                        'Master_fvEPg': '',
                        'vzCPIf': '',
                        'vzCtrctEPgCont': '',
                        'vzTaboo': '',
                        'exception_tag': '',
                        'qosCustomPol': '',
                        'qosDppPol': '',
                        'intra_vzBrCP': '',
                        'fhsTrustCtrlPol': '',
                        'fabricNode': '',
                        'fabricPathEp': '',
                        'vzGraphCont': ''}

        # Get the EPG Policies from the Network Policies Tab
        func = 'epg'
        count = countKeys(ws_net, func)
        row_count = ''
        var_dict = findVars(ws_net, func, rowcount, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('EPG_Policy'):
                row_count = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
            validating.name_rule(row_num, ws, 'Bridge_Domain', templateVars['Bridge_Domain'])
            if not templateVars['VLAN'] == None:
                validating.vlans(row_num, ws, 'VLAN', templateVars['VLAN'])
            if not templateVars['PVLAN'] == None:
                validating.vlans(row_num, ws, 'PVLAN', templateVars['PVLAN'])
            validating.enforcement(row_count, ws_net, 'pc_enf_pref', templateVars['pc_enf_pref'])
            validating.enabled(row_count, ws_net, 'flood', templateVars['flood'])
            validating.include(row_count, ws_net, 'pref_gr_memb', templateVars['pref_gr_memb'])
            validating.match_t(row_count, ws_net, 'match_t', templateVars['match_t'])
            validating.noyes(row_count, ws_net, 'is_attr_based', templateVars['is_attr_based'])
            validating.noyes(row_count, ws_net, 'shutdown', templateVars['shutdown'])
            validating.proxy_arp(row_count, ws_net, 'fwd_ctrl', templateVars['fwd_ctrl'])
            validating.qos_priority(row_count, ws_net, 'prio', templateVars['prio'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)
        
        if templateVars['cons_vzBrCP'] == 'default':
            templateVars['cons_vzBrCP'] = 'uni/tn-common/brc-default'
        if templateVars['prov_vzBrCP'] == 'default':
            templateVars['prov_vzBrCP'] = 'uni/tn-common/brc-default'
        if templateVars['vzCPIf'] == 'default':
            templateVars['vzCPIf'] = 'uni/tn-common/cif-default'
        # if templateVars['vzCtrctEPgCont'] == 'default':
        #     templateVars['vzCtrctEPgCont'] = 'uni/tn-common/mldsnoopPol-default'
        if templateVars['vzTaboo'] == 'default':
            templateVars['vzTaboo'] = 'uni/tn-common/taboo-default'
        if templateVars['qosCustomPol'] == 'default':
            templateVars['qosCustomPol'] = 'uni/tn-common/qoscustom-default'
        if templateVars['qosDppPol'] == 'default':
            templateVars['qosDppPol'] = 'uni/tn-common/qosdpppol-default'
        if templateVars['intra_vzBrCP'] == 'default':
            templateVars['intra_vzBrCP'] = 'uni/tn-common/brc-default'
        if templateVars['monEPGPol'] == 'default':
            templateVars['monEPGPol'] = 'uni/tn-common/monepg-default'
        if templateVars['fhsTrustCtrlPol'] == 'default':
            templateVars['fhsTrustCtrlPol'] = 'uni/tn-common/trustctrlpol-default'
        if templateVars['fwd_ctrl'] == 'none':
            templateVars['fwd_ctrl'] = None
        # if templateVars['vzGraphCont'] == 'default':
        #     templateVars['vzGraphCont'] = 'uni/tn-common/monitorpol-default'

        if re.search('^(common|mgmt|infra)$', templateVars['Tenant']):
            templateVars['Tenant_Dn'] = 'data.aci_tenant.%s' % (templateVars['Tenant'])
        else:
            templateVars['Tenant_Dn'] = 'aci_tenant.%s' % (templateVars['Tenant'])

        # Define the Template Source
        template_file = "epg.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'epg_%s_%s.tf' % (templateVars['App_Profile'], templateVars['EPG'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        if not templateVars['Physical_Domains'] == None:
            if ',' in templateVars['Physical_Domains']:
                splitx = templateVars['Physical_Domains'].split(',')
                for x in splitx:
                    templateVars['Domain'] = 'phys-%s' % (x)
                    # Define the Template Source
                    template_file = "domain_to_epg.template"
                    template = self.templateEnv.get_template(template_file)

                    # Process the template through the Sites
                    dest_file = 'epg_%s_%s.tf' % (templateVars['App_Profile'], templateVars['EPG'])
                    dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
                    process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)
            else:
                templateVars['Domain'] = 'phys-%s' % (templateVars['Physical_Domains'])
                # Define the Template Source
                template_file = "domain_to_epg.template"
                template = self.templateEnv.get_template(template_file)

                # Process the template through the Sites
                dest_file = 'epg_%s_%s.tf' % (templateVars['App_Profile'], templateVars['EPG'])
                dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
                process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

        if not templateVars['VMM_Domains'] == None:
            if ',' in templateVars['VMM_Domains']:
                splitx = templateVars['VMM_Domains'].split(',')
                for x in splitx:
                    templateVars['Domain'] = 'vmm-%s' % (x)
                    # Define the Template Source
                    template_file = "domain_to_epg.template"
                    template = self.templateEnv.get_template(template_file)

                    # Process the template through the Sites
                    dest_file = 'epg_%s_%s.tf' % (templateVars['App_Profile'], templateVars['EPG'])
                    dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
                    process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)
            else:
                templateVars['Domain'] = 'vmm-%s' % (templateVars['VMM_Domains'])
                # Define the Template Source
                template_file = "domain_to_epg.template"
                template = self.templateEnv.get_template(template_file)

                # Process the template through the Sites
                dest_file = 'epg_%s_%s.tf' % (templateVars['App_Profile'], templateVars['EPG'])
                dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
                process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)
        
        if not templateVars['VLAN'] == None:
            # Define the Template Source
            template_file = "static_path.template"
            template = self.templateEnv.get_template(template_file)

            dest_file = 'epg_%s_%s.tf' % (templateVars['App_Profile'], templateVars['EPG'])
            dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
            process_workbook(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

        # dest_file = 'epg_%s_%s_static_bindings.tf' % (templateVars['App_Profile'], templateVars['EPG'])
        # dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        # create_file(wb, ws, row_num, 'w', dest_dir, dest_file, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the "Excel Spreadsheet Guidance" PDF File  
    # for Detailed information on the Arguments used by this Method.
    def add_net(self, wb, ws, row_num, **kwargs):
        # Assignt he kwargs to a initial var for each process
        initial_kwargs = kwargs

        # Initialize the Class
        aci_lib_ref = 'Tenant_Policies'
        class_init = '%s(ws)' % (aci_lib_ref)

        # Create Bridge Domain
        eval("%s.%s(wb, ws, row_num, **kwargs)" % (class_init, 'add_bd'))
            
        # Create the Subnet if it Exists
        if not kwargs.get('Subnet') == None:
            eval("%s.%s(wb, ws, row_num, **kwargs)" % (class_init, 'add_subnet'))

        # Reset kwargs back to initial kwargs
        kwargs = initial_kwargs

        # Create the Application Profile if it Exists
        if not kwargs.get('App_Profile') == None:
            eval("%s.%s(wb, ws, row_num, **kwargs)" % (class_init, 'add_app'))

        # Reset kwargs back to initial kwargs
        kwargs = initial_kwargs

        # Create the EPG if it Exists
        if not kwargs.get('EPG') == None:
            eval("%s.%s(wb, ws, row_num, **kwargs)" % (class_init, 'add_epg'))

        # Reset kwargs back to initial kwargs
        kwargs = initial_kwargs

    # Method must be called with the following kwargs.
    # Please Refer to the "Excel Spreadsheet Guidance" PDF File  
    # for Detailed information on the Arguments used by this Method.
    def add_subnet(self, wb, ws, row_num, **kwargs):
        # Open the Network Policies Worksheet
        ws_net = wb['Network Policies']
        rowcount = ws_net.max_row

        # Dicts for Subnet required and optional args
        required_args = {'Site_Group': '',
                        'Tenant': '',
                        'Bridge_Domain': '',
                        'Subnet': '',
                        'Subnet_Policy': '',
                        'Policy_Name': '',
                        'nd': '',
                        'no-default-gateway': '',
                        'querier': '',
                        'preferred': '',
                        'scope': '',
                        'virtual': ''}
        optional_args = {'Subnet_Description': '',
                        'L3Out_Tenant': '',
                        'L3Out': '',
                        'annotation': '',
                        'name_alias': '',
                        'rtctrlProfile': '',
                        'ndPfxPol': ''}

        # Get the Subnet Policies from the Network Policies Tab
        func = 'subnet'
        count = countKeys(ws_net, func)
        row_count = ''
        var_dict = findVars(ws_net, func, rowcount, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('Subnet_Policy'):
                row_count = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}
                break

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.ip_address(row_num, ws, 'Subnet', templateVars['Subnet'])
            validating.name_rule(row_num, ws, 'L3Out', templateVars['L3Out'])
            validating.noyes(row_count, ws_net, 'nd', templateVars['nd'])
            validating.noyes(row_count, ws_net, 'no-default-gateway', templateVars['no-default-gateway'])
            validating.noyes(row_count, ws_net, 'querier', templateVars['querier'])
            validating.noyes(row_count, ws_net, 'preferred', templateVars['preferred'])
            validating.noyes(row_count, ws_net, 'virtual', templateVars['virtual'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)
        
        if templateVars['ndPfxPol'] == 'default':
            templateVars['ndPfxPol'] = 'uni/tn-common/ndpfxpol-default'

        # Create ctrl templateVars
        ctrl = ''
        if templateVars['nd'] == 'yes':
            ctrl = ctrl + '"nd"'
        if templateVars['no-default-gateway'] == 'yes':
            ctrl = ctrl + ', ' + '"no-default-gateway"'
        if templateVars['querier'] == 'yes':
            ctrl = ctrl + ', ' + '"querier"'

        if ctrl == '':
            templateVars['ctrl'] = '["unspecified"]'
        else:
            templateVars['ctrl'] = '[%s]' % (ctrl)
        
        # Modify scope templateVars
        if re.search('^(private|public|shared)$', templateVars['scope']):
            templateVars['scope'] = '"%s"' % (templateVars['scope'])
        elif re.search('^(private|public)\\-shared$', templateVars['scope']):
            x = templateVars['scope'].split('-')
            templateVars['scope'] = '"%s", "%s"' & (x[0], x[1])

        # As period and colon are not allowed in description need to modify Subnet to work for description and filename
        if ':' in templateVars['Subnet']:
            network = "%s" % (ipaddress.IPv6Network(templateVars['Subnet'], strict=False))
            templateVars['Subnet_'] = network
            templateVars['Subnet_'] = templateVars['Subnet_'].replace(':', '-')
            templateVars['Subnet_'] = templateVars['Subnet_'].replace('/', '_')
        else:
            network = "%s" % (ipaddress.IPv4Network(templateVars['Subnet'], strict=False))
            templateVars['Subnet_'] = network
            templateVars['Subnet_'] = templateVars['Subnet_'].replace('.', '-')
            templateVars['Subnet_'] = templateVars['Subnet_'].replace('/', '_')
        
        if not (templateVars['L3Out_Tenant'] == None and templateVars['L3Out'] == None):
            if not templateVars['Tenant'] == templateVars['L3Out_Tenant']:

                # Process the template through the Sites
                templateVars['data_Tenant'] = templateVars['L3Out_Tenant']
                template_file = "data_tenant.template"
                template = self.templateEnv.get_template(template_file)

                # Process the template through the Sites
                dest_file = 'data_tenant_%s.tf' % (templateVars['L3Out_Tenant'])
                dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
                process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

                template_file = "data_l3out.template"
                template = self.templateEnv.get_template(template_file)

                # Process the template through the Sites
                dest_file = 'data_tenant_%s_l3out_%s.tf' % (templateVars['L3Out_Tenant'], templateVars['L3Out'])
                dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
                process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

                # Create the Distinguished Name for the L3Out
                templateVars['L3_Dn'] = 'data.aci_l3_outside.%s_%s' % ((templateVars['L3Out_Tenant'], templateVars['L3Out']))

            else:
                # Create the Distinguished Name for the L3Out
               templateVars['L3_Dn'] = 'aci_l3_outside.%s_%s' % ((templateVars['L3Out_Tenant'], templateVars['L3Out']))

        # Define the Template Source
        template_file = "subnet.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'bd_%s_subnet_%s.tf' % (templateVars['Bridge_Domain'], templateVars['Subnet_'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the "Excel Spreadsheet Guidance" PDF File  
    # for Detailed information on the Arguments used by this Method.
    def add_tenant(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Tenant': ''}
        optional_args = {'Description': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "tenant.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'tenant_%s.tf' % (templateVars['Tenant'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the "Excel Spreadsheet Guidance" PDF File  
    # for Detailed information on the Arguments used by this Method.
    def add_vrf(self, wb, ws, row_num, **kwargs):
        # Open the Network Policies Worksheet
        ws_net = wb['Network Policies']
        rowcount = ws_net.max_row

        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                        'Tenant': '',
                        'VRF': '',
                        'VRF_Policy': '',
                        'Policy_Name': '',
                        'pc_enf_pref': '',
                        'pc_enf_dir': '',
                        'bd_enforce': '',
                        'enf_type': '',
                        'fvEpRetPol': '',
                        'monEPGPol': '',
                        'dp_learning': '',
                        'knw_mcast_act': ''}
        optional_args = {'Description': '',
                        'annotation': '',
                        'name_alias': '',
                        'cons_vzBrCP': '',
                        'vzCPIf': '',
                        'prov_vzBrCP': '',
                        'bgpCtxPol': '',
                        'bgpCtxAfPol': '',
                        'ospfCtxPol': '',
                        'ospfCtxAfPol': '',
                        'eigrpCtxAfPol': '',
                        'l3extRouteTagPol': '',
                        'l3extVrfValidationPol': ''}	


        # Get the VRF Policies from the Network Policies Tab
        func = 'VRF'
        count = countKeys(ws_net, func)
        row_count = ''
        var_dict = findVars(ws_net, func, rowcount, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('VRF_Policy'):
                row_count = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
            validating.name_rule(row_num, ws, 'VRF', templateVars['VRF'])
            validating.deny(row_count, ws_net, 'knw_mcast_act', templateVars['knw_mcast_act'])
            validating.direction(row_count, ws_net, 'pc_enf_dir', templateVars['pc_enf_dir'])
            validating.enabled(row_count, ws_net, 'dp_learning', templateVars['dp_learning'])
            validating.enforcement(row_count, ws_net, 'pc_enf_pref', templateVars['pc_enf_pref'])
            validating.enforce_type(row_count, ws_net, 'enf_type', templateVars['enf_type'])
            validating.noyes(row_count, ws_net, 'bd_enforce', templateVars['bd_enforce'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        if templateVars['cons_vzBrCP'] == 'default':
            templateVars['cons_vzBrCP'] = 'uni/tn-common/brc-default'
        if templateVars['prov_vzBrCP'] == 'default':
            templateVars['prov_vzBrCP'] = 'uni/tn-common/brc-default'
        if templateVars['vzCPIf'] == 'default':
            templateVars['vzCPIf'] = 'uni/tn-common/cif-default'
        if templateVars['bgpCtxPol'] == 'default':
            templateVars['bgpCtxPol'] = 'uni/tn-common/bgpCtxP-default'
        if templateVars['bgpCtxAfPol'] == 'default':
            templateVars['bgpCtxAfPol'] = 'uni/tn-common/bgpCtxAfP-default'
        if templateVars['eigrpCtxAfPol'] == 'default':
            templateVars['eigrpCtxAfPol'] = 'uni/tn-common/eigrpCtxAfP-default'
        if templateVars['ospfCtxPol'] == 'default':
            templateVars['ospfCtxPol'] = 'uni/tn-common/ospfCtxP-default'
        if templateVars['ospfCtxAfPol'] == 'default':
            templateVars['ospfCtxAfPol'] = 'uni/tn-common/ospfCtxP-default'
        if templateVars['fvEpRetPol'] == 'default':
            templateVars['fvEpRetPol'] = 'uni/tn-common/epRPol-default'
        if templateVars['monEPGPol'] == 'default':
            templateVars['monEPGPol'] = 'uni/tn-common/monepg-default'
        if templateVars['l3extVrfValidationPol'] == 'default':
            templateVars['l3extVrfValidationPol'] = 'uni/tn-common/vrfvalidationpol-default'

        if re.search('^(common|mgmt|infra)$', templateVars['Tenant']):
            templateVars['Tenant_Dn'] = 'data.aci_tenant.%s' % (templateVars['Tenant'])
        else:
            templateVars['Tenant_Dn'] = 'aci_tenant.%s' % (templateVars['Tenant'])

        # Define the Template Source
        template_file = "vrf.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'vrf_%s.tf' % (templateVars['VRF'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        if templateVars['enf_type'] == 'pref_grp':
            # Define the Template Source
            template_file = "pref_grp.template"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = 'vrf_%s.tf' % (templateVars['VRF'])
            dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
            process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

        elif templateVars['enf_type'] == 'vzAny':
            # Define the Template Source
            template_file = "vzAny.template"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = 'vrf_%s.tf' % (templateVars['VRF'])
            dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
            process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "snmp_ctx.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'vrf_%s.tf' % (templateVars['VRF'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the "Excel Spreadsheet Guidance" PDF File  
    # for Detailed information on the Arguments used by this Method.
    def ctx_comm(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Tenant': '',
                         'VRF': '',
                         'Ctx_Community': ''}
        optional_args = {'Description': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
            validating.name_rule(row_num, ws, 'VRF', templateVars['VRF'])
            validating.snmp_string(row_num, ws, 'Ctx_Community', templateVars['Ctx_Community'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "snmp_ctx_community.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'vrf_%s.tf' % (templateVars['VRF'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

# Terraform ACI Provider - Tenants Policies
# Class must be instantiated with Variables
class VMM_Policies(object):
    def __init__(self, ws):
        self.templateLoader = jinja2.FileSystemLoader(
            searchpath=(aci_template_path + 'Tenant_Policies/'))
        self.templateEnv = jinja2.Environment(loader=self.templateLoader)

    # Method must be called with the following kwargs.
    # Please Refer to the "Excel Spreadsheet Guidance" PDF File  
    # for Detailed information on the Arguments used by this Method.
    def add_vmm(self, wb, ws, row_num, **kwargs):
        # Open the Network Policies Worksheet
        ws_net = wb['Network Policies']
        rowcount = ws_net.max_row

        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Tenant': '',
                         'VRF': '',
                         'Name': '',
                         'L3Out_Policy': '',
                         'L3_Domain': '',
                         'Ext_EPG': '',
                         'Ext_EPG_Policy': '',
                         'Subnet': '',
                         'Ext_Subnet_Policy': '',
                         'target_dscp': '',
                         'enforce_rtctrl': '',
                         'prio': '',
                         'epg_target_dscp': '',
                         'pref_gr_memb': '',
                         'match_t': '',
                         'flood': '',
                         'export-rtctrl': '',
                         'import-rtctrl': '',
                         'import-security': '',
                         'shared-security': '',
                         'shared-rtctrl': '',
                         'agg-export': '',
                         'agg-import': '',
                         'agg-shared': ''}
        optional_args = {'Description': '',
                         'EPG_Description': '',	
                         'annotation': '',
                         'name_alias': '',
                         'leak_rtctrlProfile': '',
                         'damp_rtctrlProfile': '',
                         'fvBDPublicSubnetHolder': '',	
                         'epg_annotation': '',
                         'epg_name_alias': '',
                         'cons_vzBrCP': '',
                         'vzCPIf': '',
                         'Master_fvEPg': '',
                         'prov_vzBrCP': '',
                         'vzTaboo': '',
                         'exception_tag': '',
                         'rtctrlProfile': '',
                         'sub_annotation': '',
                         'sub_name_alias': '',
                         'sub_rtctrlProfile': '',
                         'rtsumARtSummPol': ''}


        # Get the L3Out Policies from the Network Policies Tab
        func = 'L3Out_Policy'
        count = countKeys(ws_net, func)
        l3_count = ''
        var_dict = findVars(ws_net, func, rowcount, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('L3Out_Policy'):
                l3_count = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}

        func = 'Ext_EPG_Policy'
        count = countKeys(ws_net, func)
        epg_count = ''
        var_dict = findVars(ws_net, func, rowcount, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('Ext_EPG_Policy'):
                epg_count = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}

        func = 'Ext_Subnet_Policy'
        count = countKeys(ws_net, func)
        sub_count = ''
        var_dict = findVars(ws_net, func, rowcount, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('Ext_Subnet_Policy'):
                sub_count = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
            validating.name_rule(row_num, ws, 'VRF', templateVars['VRF'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            if not templateVars['Subnet'] == None:
                if re.search(',', templateVars['Subnet']):
                    sx = templateVars['Subnet'].split(',')
                    for x in sx:
                        validating.ip_address(row_num, ws, 'Subnet', x)
            validating.dscp(l3_count, ws_net, 'target_dscp', templateVars['target_dscp'])
            validating.export(l3_count, ws_net, 'enforce_rtctrl', templateVars['enforce_rtctrl'])
            validating.dscp(epg_count, ws_net, 'epg_target_dscp', templateVars['epg_target_dscp'])
            validating.enabled(epg_count, ws_net, 'flood', templateVars['flood'])
            validating.include(epg_count, ws_net, 'pref_gr_memb', templateVars['pref_gr_memb'])
            validating.match_t(epg_count, ws_net, 'match_t', templateVars['match_t'])
            validating.qos_priority(epg_count, ws_net, 'prio', templateVars['prio'])
            validating.noyes(sub_count, ws_net, 'agg-export', templateVars['agg-export'])
            validating.noyes(sub_count, ws_net, 'agg-import', templateVars['agg-import'])
            validating.noyes(sub_count, ws_net, 'agg-shared', templateVars['agg-shared'])
            validating.noyes(sub_count, ws_net, 'export-rtctrl', templateVars['export-rtctrl'])
            validating.noyes(sub_count, ws_net, 'import-rtctrl', templateVars['import-rtctrl'])
            validating.noyes(sub_count, ws_net, 'import-security', templateVars['import-security'])
            validating.noyes(sub_count, ws_net, 'shared-security', templateVars['shared-security'])
            validating.noyes(sub_count, ws_net, 'shared-rtctrl', templateVars['shared-rtctrl'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Create aggregate templateVars
        aggregate = ''
        if templateVars['agg-export'] == 'yes':
            aggregate = aggregate + '"export-rtctrl"'
        if templateVars['agg-import'] == 'yes':
            aggregate = aggregate + ', ' + '"import-rtctrl"'
        if templateVars['agg-shared'] == 'yes':
            aggregate = aggregate + ', ' + '"shared-rtctrl"'

        else:
            templateVars['aggregate'] = '[%s]' % (aggregate)
        
        # Create scope templateVars
        scope = ''
        if templateVars['export-rtctrl'] == 'yes':
            scope = scope + '"export-rtctrl"'
        if templateVars['import-rtctrl'] == 'yes':
            scope = scope + ', ' + '"import-rtctrl"'
        if templateVars['import-security'] == 'yes':
            scope = scope + ', ' + '"import-security"'
        if templateVars['shared-security'] == 'yes':
            scope = scope + ', ' + '"shared-security"'
        if templateVars['shared-rtctrl'] == 'yes':
            scope = scope + ', ' + '"shared-rtctrl"'

        else:
            templateVars['scope'] = '[%s]' % (scope)
        
        # Define the Template Source
        template_file = "l3out.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'l3out_%s.tf' % (templateVars['Name'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "ext_epg.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'l3out_%s_epg_%s.tf' % (templateVars['Name'], templateVars['Ext_EPG'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        if re.search(',', templateVars['Subnet']):
            sx = templateVars['Subnet'].split(',')
            for x in sx:
                templateVars['Subnet'] = x
                templateVars['Subnet_'] = x.replace('.', '-')
                templateVars['Subnet_'] = x.replace('/', '_')
                
                # Define the Template Source
                template_file = "ext_subnet.template"
                template = self.templateEnv.get_template(template_file)

                # Process the template through the Sites
                dest_file = 'l3out_%s_epg_%s_subnet_%s.tf' % (templateVars['Name'], templateVars['Ext_EPG'], templateVars['Subnet'])
                dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
                process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)
        else:
            templateVars['Subnet_'] = templateVars['Subnet'].replace('.', '-')
            templateVars['Subnet_'] = templateVars['Subnet'].replace('/', '_')

            # Define the Template Source
            template_file = "ext_subnet.template"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = 'l3out_%s_epg_%s_subnet_%s.tf' % (templateVars['Name'], templateVars['Ext_EPG'], templateVars['Subnet'])
            dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
            process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

def copy_defaults(Site_Name, dest_dir):
    src_dir = './ACI/templates'
    dest_dir = './ACI/%s/%s' % (Site_Name, dest_dir)
    if not os.path.isdir(dest_dir):
        mk_dir = 'mkdir -p %s' % (dest_dir)
        os.system(mk_dir)
    cp_main = 'cp %s/main.tf %s/.gitignore %s/variables.tf %s/' % (src_dir, src_dir, src_dir, dest_dir)
    os.system(cp_main)

    if dest_dir.endswith('/Access'):
        cp_template = 'cp %s/defaults_Fabric_Access_Policies.tf %s/vars_Fabric_Access_Policies.tf %s/' % (src_dir, src_dir, dest_dir)
        os.system(cp_template)
    elif dest_dir.endswith('/Admin'):
        cp_template = 'cp %s/defaults_Admin.tf %s/vars_Admin.tf %s/' % (src_dir, src_dir, dest_dir)
        os.system(cp_template)
    elif dest_dir.endswith('/Fabric'):
        cp_template = 'cp %s/defaults_Fabric_Fabric_Policies.tf %s/vars_Fabric_Fabric_Policies.tf %s/' % (src_dir, src_dir, dest_dir)
        os.system(cp_template)
    elif dest_dir.endswith('/Tenant_common'):
        cp_template = 'cp %s/defaults_Tenant_common.tf %s/' % (src_dir, dest_dir)
        os.system(cp_template)
    elif dest_dir.endswith('/Tenant_infra'):
        cp_template = 'cp %s/defaults_Tenant_infra.tf %s/' % (src_dir, dest_dir)
        os.system(cp_template)
    elif dest_dir.endswith('/Tenant_mgmt'):
        cp_template = 'cp %s/defaults_Tenant_mgmt.tf %s/' % (src_dir, dest_dir)
        os.system(cp_template)

def create_file(wb, ws, row_num, wr_method, dest_dir, dest_file, **templateVars):
    if re.search('Grp_[A-F]', templateVars['Site_Group']):
        Group_ID = '%s' % (templateVars['Site_Group'])
        site_group = ast.literal_eval(os.environ[Group_ID])
        for x in range(1, 13):
            sitex = 'Site_%s' % (x)
            if not site_group[sitex] == None:
                Site_ID = 'Site_ID_%s' % (site_group[sitex])
                site_dict = ast.literal_eval(os.environ[Site_ID])

                # Create templateVars for Site_Name and APIC_URL
                templateVars['Site_Name'] = site_dict.get('Site_Name')
                templateVars['APIC_URL'] = site_dict.get('APIC_URL')

                # Create Terraform file from Template
                tf_file = './ACI/%s/%s/%s' % (templateVars['Site_Name'], dest_dir, dest_file)
                wr_file = open(tf_file, wr_method)
                wr_file.close()

    elif re.search(r'\d+', templateVars['Site_Group']):
        Site_ID = 'Site_ID_%s' % (templateVars['Site_Group'])
        site_dict = ast.literal_eval(os.environ[Site_ID])

        # Create templateVars for Site_Name and APIC_URL
        templateVars['Site_Name'] = site_dict.get('Site_Name')
        templateVars['APIC_URL'] = site_dict.get('APIC_URL')

        # Create Terraform file from Template
        tf_file = './ACI/%s/%s/%s' % (templateVars['Site_Name'], dest_dir, dest_file)
        wr_file = open(tf_file, wr_method)
        wr_file.close()
    else:
        print(f"\n-----------------------------------------------------------------------------\n")
        print(f"   Error on Worksheet {ws.title}, Row {row_num} Site_Group, value {templateVars['Site_Group']}.")
        print(f"   Unable to Determine if this is a Single or Group of Site(s).  Exiting....")
        print(f"\n-----------------------------------------------------------------------------\n")
        exit()

def create_tf_file(wr_method, dest_dir, dest_file, template, **templateVars):
    # Create File for the Template in the Destination Folder
    tf_file = './ACI/%s/%s/%s' % (templateVars['Site_Name'], dest_dir, dest_file)
    wr_file = open(tf_file, wr_method)
    
    # Render Payload and Write to File
    payload = template.render(templateVars)
    wr_file.write(payload + '\n\n')
    wr_file.close()

# Function to Create Interface Selectors
def create_selector(ws_sw, ws_sw_row_count, **templateVars):
    print(templateVars['port_count'])
    Port_Selector = ''
    for port in range(1, int(templateVars['port_count']) + 1):
        if port < 10:
            Port_Selector = 'Eth%s-0%s' % (templateVars['module'], port)
        elif port < 100:
            Port_Selector = 'Eth%s-%s' % (templateVars['module'], port)
        elif port > 99:
            Port_Selector = 'Eth%s_%s' % (templateVars['module'], port)
        modp = '%s/%s' % (templateVars['module'],port)
        # Copy the Port Selector to the Worksheet
        data = ['intf_selector',templateVars['Pod_ID'],templateVars['Node_ID'],templateVars['Name'],Port_Selector,modp,'','','','','','']
        ws_sw.append(data)
        rc = '%s:%s' % (ws_sw_row_count, ws_sw_row_count)
        for cell in ws_sw[rc]:
            if ws_sw_row_count % 2 == 0:
                cell.style = 'ws_odd'
            else:
                cell.style = 'ws_even'
        dv1_cell = 'A%s' % (ws_sw_row_count)
        dv2_cell = 'H%s' % (ws_sw_row_count)
        dv3_cell = 'I%s' % (ws_sw_row_count)
        templateVars['dv1'].add(dv1_cell)
        templateVars['dv2'].add(dv2_cell)
        templateVars['dv3'].add(dv3_cell)
        ws_sw_row_count += 1
    return ws_sw_row_count

def create_static_paths(wb, wb_sw, row_num, wr_method, dest_dir, dest_file, template, **templateVars):
    wsheets = wb_sw.get_sheet_names()
    tf_file = ''
    for wsheet in wsheets:
        ws = wb_sw[wsheet]
        for row in ws.rows:
            if not (row[12].value == None or row[13].value == None):
                vlan_test = ''
                if re.search('^(individual|port-channel|vpc)$', row[7].value) and (re.search(r'\d+', str(row[12].value)) or re.search(r'\d+', str(row[13].value))):
                    if not row[12].value == None:
                        vlan = row[12].value
                        vlan_test = vlan_range(vlan, **templateVars)
                        if 'true' in vlan_test:
                            templateVars['mode'] = 'native'
                    if not 'true' in vlan_test:
                        templateVars['mode'] = 'regular'
                        if not row[13].value == None:
                            vlans = row[13].value
                            vlan_test = vlan_range(vlans, **templateVars)
                if vlan_test == 'true':
                    templateVars['Pod_ID'] = row[1].value
                    templateVars['Node_ID'] = row[2].value
                    templateVars['Switch_Name'] = row[3].value
                    templateVars['Interface_Selector'] = row[4].value
                    templateVars['Port'] = row[5].value
                    templateVars['Policy_Group'] = row[6].value
                    templateVars['Port_Type'] = row[7].value
                    templateVars['Bundle_ID'] = row[9].value
                    Site_Group = templateVars['Site_Group']
                    pod = templateVars['Pod_ID']
                    node_id =  templateVars['Node_ID']
                    if templateVars['Port_Type'] == 'vpc':
                        ws_vpc = wb['Inventory']
                        for rx in ws_vpc.rows:
                            if rx[0].value == 'vpc_pair' and int(rx[1].value) == int(Site_Group) and str(rx[4].value) == str(node_id):
                                node1 = templateVars['Node_ID']
                                node2 = rx[5].value
                                templateVars['Policy_Group'] = 'pg_vpc%s_%s' % (templateVars['Bundle_ID'], row[3].value)
                                templateVars['tDn'] = 'topology/pod-%s/protpaths-%s-%s/pathep-[%s]' % (pod, node1, node2, templateVars['Policy_Group'])
                                templateVars['Static_Path'] = 'rspathAtt-[topology/pod-%s/protpaths-%s-%s/pathep-[%s]' % (pod, node1, node2, templateVars['Policy_Group'])
                                templateVars['GUI_Static'] = 'Pod-%s/Node-%s-%s/%s' % (pod, node1, node2, templateVars['Policy_Group'])
                                templateVars['Static_descr'] = 'Pod-%s_Nodes-%s-%s_%s' % (pod, node1, node2, templateVars['Policy_Group'])
                                tf_file = './ACI/%s/%s/%s' % (templateVars['Site_Name'], dest_dir, dest_file)
                                read_file = open(tf_file, 'r')
                                read_file.seek(0)
                                static_path_descr = 'resource "aci_epg_to_static_path" "%s_%s_%s"' % (templateVars['App_Profile'], templateVars['EPG'], templateVars['Static_descr'])
                                if not static_path_descr in read_file.read():
                                    create_tf_file(wr_method, dest_dir, dest_file, template, **templateVars)

                    elif templateVars['Port_Type'] == 'port-channel':
                        templateVars['Policy_Group'] = 'pg_pc%s_%s' % (templateVars['Bundle_ID'], row[3].value)
                        templateVars['tDn'] = 'topology/pod-%s/paths-%s/pathep-[%s]' % (pod, templateVars['Node_ID'], templateVars['Policy_Group'])
                        templateVars['Static_Path'] = 'rspathAtt-[topology/pod-%s/paths-%s/pathep-[%s]' % (pod, templateVars['Node_ID'], templateVars['Policy_Group'])
                        templateVars['GUI_Static'] = 'Pod-%s/Node-%s/%s' % (pod, templateVars['Node_ID'], templateVars['Policy_Group'])
                        templateVars['Static_descr'] = 'Pod-%s_Node-%s_%s' % (pod, templateVars['Node_ID'], templateVars['Policy_Group'])
                        read_file = open(tf_file, 'r')
                        read_file.seek(0)
                        static_path_descr = 'resource "aci_epg_to_static_path" "%s_%s_%s"' % (templateVars['App_Profile'], templateVars['EPG'], templateVars['Static_descr'])
                        if not static_path_descr in read_file.read():
                            create_tf_file(wr_method, dest_dir, dest_file, template, **templateVars)

                    elif templateVars['Port_Type'] == 'individual':
                        port = 'eth%s' % (templateVars['Port'])
                        templateVars['tDn'] = 'topology/pod-%s/paths-%s/pathep-[%s]' % (pod, templateVars['Node_ID'], port)
                        templateVars['Static_Path'] = 'rspathAtt-[topology/pod-%s/paths-%s/pathep-[%s]' % (pod, templateVars['Node_ID'], port)
                        templateVars['GUI_Static'] = 'Pod-%s/Node-%s/%s' % (pod, templateVars['Node_ID'], port)
                        templateVars['Static_descr'] = 'Pod-%s_Node-%s_%s' % (pod, templateVars['Node_ID'], templateVars['Interface_Selector'])
                        read_file = open(tf_file, 'r')
                        read_file.seek(0)
                        static_path_descr = 'resource "aci_epg_to_static_path" "%s_%s_%s"' % (templateVars['App_Profile'], templateVars['EPG'], templateVars['Static_descr'])
                        if not static_path_descr in read_file.read():
                            create_tf_file(wr_method, dest_dir, dest_file, template, **templateVars)

def countKeys(ws, func):
    count = 0
    for i in ws.rows:
        if any(i):
            if str(i[0].value) == func:
                count += 1
    return count

def findKeys(ws, func_regex):
    func_list = OrderedSet()
    for i in ws.rows:
        if any(i):
            if re.search(func_regex, str(i[0].value)):
                func_list.add(str(i[0].value))
    return func_list

def findVars(ws, func, rows, count):
    var_list = []
    var_dict = {}
    for i in range(1, rows + 1):
        if (ws.cell(row=i, column=1)).value == func:
            print()
            try:
                for x in range(2, 34):
                    if (ws.cell(row=i - 1, column=x)).value:
                        var_list.append(str(ws.cell(row=i - 1, column=x).value))
                    else:
                        x += 1
            except Exception as e:
                e = e
                pass
            break
    vcount = 1
    while vcount <= count:
        var_dict[vcount] = {}
        var_count = 0
        for z in var_list:
            var_dict[vcount][z] = ws.cell(row=i + vcount - 1, column=2 + var_count).value
            var_count += 1
        var_dict[vcount]['row'] = i + vcount - 1
        vcount += 1
    return var_dict

# Function to validate input for each method
def process_kwargs(required_args, optional_args, **kwargs):
    # Validate all required kwargs passed
    if all(item in kwargs for item in required_args.keys()) is not True:
        error_ = '\n***ERROR***\nREQUIRED ARGS ARE:\n "%s"\nOPTIONAL ARGS ARE:\n "%s"\nPROVIDED ARGS ARE:\n"%s"\nInsufficient required arguments.' % (required_args, optional_args, kwargs)
        raise InsufficientArgs(error_)

    # Load all required args values from kwargs
    for item in kwargs:
        if item in required_args.keys():
            required_args[item] = kwargs[item]
    for item in kwargs:
        if item in optional_args.keys():
            optional_args[item] = kwargs[item]

    # Combine option and required dicts for Jinja template render
    templateVars = {**required_args, **optional_args}
    return(templateVars)

# Function to Apply Method to Site(s)
def process_method(wb, ws, row_num, wr_method, dest_dir, dest_file, template, **templateVars):
    if re.search('Grp_[A-F]', templateVars['Site_Group']):
        Group_ID = '%s' % (templateVars['Site_Group'])
        site_group = ast.literal_eval(os.environ[Group_ID])
        for x in range(1, 13):
            sitex = 'Site_%s' % (x)
            if not site_group[sitex] == None:
                Site_ID = 'Site_ID_%s' % (site_group[sitex])
                site_dict = ast.literal_eval(os.environ[Site_ID])

                # Create templateVars for Site_Name and APIC_URL
                templateVars['Site_Name'] = site_dict.get('Site_Name')
                templateVars['APIC_URL'] = site_dict.get('APIC_URL')

                # Create Terraform file from Template
                create_tf_file(wr_method, dest_dir, dest_file, template, **templateVars)

    elif re.search(r'\d+', templateVars['Site_Group']):
        Site_ID = 'Site_ID_%s' % (templateVars['Site_Group'])
        site_dict = ast.literal_eval(os.environ[Site_ID])

        # Create templateVars for Site_Name and APIC_URL
        templateVars['Site_Name'] = site_dict.get('Site_Name')
        templateVars['APIC_URL'] = site_dict.get('APIC_URL')

        # Create Terraform file from Template
        create_tf_file(wr_method, dest_dir, dest_file, template, **templateVars)
    else:
        print(f"\n-----------------------------------------------------------------------------\n")
        print(f"   Error on Worksheet {ws.title}, Row {row_num} Site_Group, value {templateVars['Site_Group']}.")
        print(f"   Unable to Determine if this is a Single or Group of Site(s).  Exiting....")
        print(f"\n-----------------------------------------------------------------------------\n")
        exit()

# Function to Add Static Port Bindings to Bridge Domains Terraform Files
def process_workbook(wb, ws, row_num, wr_method, dest_dir, dest_file, template, **templateVars):
    if re.search('Grp_[A-F]', templateVars['Site_Group']):
        Group_ID = '%s' % (templateVars['Site_Group'])
        site_group = ast.literal_eval(os.environ[Group_ID])
        for x in range(1, 13):
            sitex = 'Site_%s' % (x)
            if not site_group[sitex] == None:
                Site_ID = 'Site_ID_%s' % (site_group[sitex])
                site_dict = ast.literal_eval(os.environ[Site_ID])

                # Create templateVars for Site_Name and APIC_URL
                templateVars['Site_Name'] = site_dict.get('Site_Name')
                templateVars['Site_Group'] = site_dict.get('Site_ID')
                templateVars['APIC_URL'] = site_dict.get('APIC_URL')

                # Pull in the Site Workbook
                excel_workbook = '%s_intf_selectors.xlsx' % (templateVars['Site_Name'])
                try:
                    wb_sw = load_workbook(excel_workbook)
                except Exception as e:
                    print(f"Something went wrong while opening the workbook - {excel_workbook}... ABORT!")
                    sys.exit(e)

                # Process the Interface Selectors for Static Port Paths
                create_static_paths(wb, wb_sw, row_num, wr_method, dest_dir, dest_file, template, **templateVars)
                                    
    elif re.search(r'\d+', templateVars['Site_Group']):
        Site_ID = 'Site_ID_%s' % (templateVars['Site_Group'])
        site_dict = ast.literal_eval(os.environ[Site_ID])

        # Pull in the Site Workbook
        excel_workbook = '%s_intf_selectors.xlsx' % (templateVars['Site_Name'])
        try:
            wb_sw = load_workbook(excel_workbook)
        except Exception as e:
            print(f"Something went wrong while opening the workbook - {excel_workbook}... ABORT!")
            sys.exit(e)


        # Process the Interface Selectors for Static Port Paths
        create_static_paths(wb, wb_sw, row_num, wr_method, dest_dir, dest_file, template, **templateVars)
                                    
    else:
        print(f"\n-----------------------------------------------------------------------------\n")
        print(f"   Error on Worksheet {ws.title}, Row {row_num} Site_Group, value {templateVars['Site_Group']}.")
        print(f"   Unable to Determine if this is a Single or Group of Site(s).  Exiting....")
        print(f"\n-----------------------------------------------------------------------------\n")
        exit()

# Function to Determine Port Count on Modules
def query_module_type(row_num, module_type):
    if re.search('^M4', module_type):
        port_count = '4'
    elif re.search('^M6', module_type):
        port_count = '6'
    elif re.search('^M12', module_type):
        port_count = '12'
    elif re.search('X9716D-GX', module_type):
        port_count = '16'
    elif re.search('X9732C-EX', module_type):
        port_count = '32'
    elif re.search('X9736', module_type):
        port_count = '36'
    else:
        print(f'\n-----------------------------------------------------------------------------\n')
        print(f'   Error on Row {row_num}.  Unknown Switch Model {module_type}')
        print(f'   Please verify Input Information.  Exiting....')
        print(f'\n-----------------------------------------------------------------------------\n')
        exit()
    return port_count

def query_switch_model(row_num, switch_type):
    modules = ''
    switch_type = str(switch_type)
    if re.search('^9396', switch_type):
        modules = '2'
        port_count = '48'
    elif re.search('^93', switch_type):
        modules = '1'

    if re.search('^9316', switch_type):
        port_count = '16'
    elif re.search('^(93120)', switch_type):
        port_count = '102'
    elif re.search('^(93108|93120|93216|93360)', switch_type):
        port_count = '108'
    elif re.search('^(93180|93240|9348|9396)', switch_type):
        port_count = '54'
    elif re.search('^(93240)', switch_type):
        port_count = '60'
    elif re.search('^9332', switch_type):
        port_count = '34'
    elif re.search('^(9336|93600)', switch_type):
        port_count = '36'
    elif re.search('^9364C-GX', switch_type):
        port_count = '64'
    elif re.search('^9364', switch_type):
        port_count = '66'
    elif re.search('^95', switch_type):
        port_count = '36'
        if switch_type == '9504':
            modules = '4'
        elif switch_type == '9508':
            modules = '8'
        elif switch_type == '9516':
            modules = '16'
        else:
            print(f'\n-----------------------------------------------------------------------------\n')
            print(f'   Error on Row {row_num}.  Unknown Switch Model {switch_type}')
            print(f'   Please verify Input Information.  Exiting....')
            print(f'\n-----------------------------------------------------------------------------\n')
            exit()
    else:
        print(f'\n-----------------------------------------------------------------------------\n')
        print(f'   Error on Row {row_num}.  Unknown Switch Model {switch_type}')
        print(f'   Please verify Input Information.  Exiting....')
        print(f'\n-----------------------------------------------------------------------------\n')
        exit()
    return modules,port_count

def read_in(excel_workbook):
    try:
        wb = load_workbook(excel_workbook)
        print("Workbook Loaded.")
    except Exception as e:
        print(f"Something went wrong while opening the workbook - {excel_workbook}... ABORT!")
        sys.exit(e)
    return wb

def stdout_log(sheet, line):
    if log_level == 0:
        return
    elif ((log_level == (1) or log_level == (2)) and
            (sheet) and (line is None)):
        #print('*' * 80)
        print(f'\n-----------------------------------------------------------------------------\n')
        print(f'   Starting work on {sheet.title} Worksheet')
        print(f'\n-----------------------------------------------------------------------------\n')
        #print('*' * 80)
    elif log_level == (2) and (sheet) and (line is not None):
        print('Evaluating line %s from %s Worksheet...' % (line, sheet.title))
    else:
        return

def vlan_list_full(vlan_list):
    full_vlan_list = []
    if re.search(r',', str(vlan_list)):
        vlist = vlan_list.split(',')
        for v in vlist:
            if re.fullmatch('^\\d{1,4}\\-\\d{1,4}$', v):
                a,b = v.split('-')
                a = int(a)
                b = int(b)
                vrange = range(a,b+1)
                for vl in vrange:
                    full_vlan_list.append(vl)
            elif re.fullmatch('^\\d{1,4}$', v):
                full_vlan_list.append(v)
    elif re.search('\\-', str(vlan_list)):
        a,b = vlan_list.split('-')
        a = int(a)
        b = int(b)
        vrange = range(a,b+1)
        for v in vrange:
            full_vlan_list.append(v)
    else:
        full_vlan_list.append(vlan_list)
    return full_vlan_list

def vlan_to_netcentric(vlan):
    vlan = int(vlan)
    if vlan < 10:
        vlan = str(vlan)
        netcentric = 'v000' + vlan
        return netcentric
    elif vlan < 100:
        vlan = str(vlan)
        netcentric = 'v00' + vlan
        return netcentric
    elif vlan < 1000:
        vlan = str(vlan)
        netcentric = 'v0' + vlan
        return netcentric
    else:
        vlan = str(vlan)
        netcentric = 'v' + vlan
        return netcentric

def vlan_range(vlan_list, **templateVars):
    results = 'unknown'
    while results == 'unknown':
        if re.search(',', str(vlan_list)):
            vx = vlan_list.split(',')
            for vrange in vx:
                if re.search('-', vrange):
                    vl = vrange.split('-')
                    min_ = int(vl[0])
                    max_ = int(vl[1])
                    if (int(templateVars['VLAN']) >= min_ and int(templateVars['VLAN']) <= max_):
                        results = 'true'
                        return results
                else:
                    if templateVars['VLAN'] == vrange:
                        results = 'true'
                        return results
            results = 'false'
            return results
        elif re.search('-', str(vlan_list)):
            vl = vlan_list.split('-')
            min_ = int(vl[0])
            max_ = int(vl[1])
            if (int(templateVars['VLAN']) >= min_ and int(templateVars['VLAN']) <= max_):
                results = 'true'
                return results
        else:
            if int(templateVars['VLAN']) == int(vlan_list):
                results = 'true'
                return results
        results = 'false'
        return results