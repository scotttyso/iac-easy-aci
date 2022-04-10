#!/usr/bin/env python

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from ordered_set import OrderedSet
from subprocess import check_output
import ast
import ipaddress
import jinja2
import json
import lib_terraform
import os, re, sys
import pkg_resources
import requests
import time
import validating

# Global options for debugging
print_payload = False
print_response_always = True
print_response_on_fail = True

# Log levels 0 = None, 1 = Class only, 2 = Line
log_level = 2

# Global path to main Template directory
aci_template_path = pkg_resources.resource_filename('lib_aci', 'ACI/templates/')

# Exception Classes
class InsufficientArgs(Exception):
    pass

class ErrException(Exception):
    pass

class InvalidArg(Exception):
    pass

class LoginFailed(Exception):
    pass

# Terraform ACI Provider - Access Policies
# Class must be instantiated with Variables
class Access_Policies(object):
    def __init__(self, ws):
        self.templateLoader = jinja2.FileSystemLoader(
            searchpath=(aci_template_path + 'Access_Policies/'))
        self.templateEnv = jinja2.Environment(loader=self.templateLoader)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def aep_profile(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Infra_VLAN': ''}
        optional_args = {'Description': '',
                         'Physical_Domains': '',
                         'L3_Domains': '',
                         'VMM_Domains': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.values(row_num, ws, 'Infra_VLAN', templateVars['Infra_VLAN'], ['no', 'yes'])
            if not templateVars['Physical_Domains'] == None:
                templateVars['domains'] = 'yes'
                if re.search(r',', templateVars['Physical_Domains']):
                    x = templateVars['Physical_Domains'].split(',')
                    for domain in x:
                        validating.name_rule(row_num, ws, 'Physical_Domains', domain)
                else:
                    validating.name_rule(row_num, ws, 'Physical_Domains', templateVars['Physical_Domains'])
            if not templateVars['L3_Domains'] == None:
                templateVars['domains'] = 'yes'
                if re.search(r',', templateVars['L3_Domains']):
                    x = templateVars['L3_Domains'].split(',')
                    for domain in x:
                        validating.name_rule(row_num, ws, 'L3_Domains', domain)
                else:
                    validating.name_rule(row_num, ws, 'L3_Domains', templateVars['L3_Domains'])
            if not templateVars['VMM_Domains'] == None:
                templateVars['domains'] = 'yes'
                if re.search(r',', templateVars['VMM_Domains']):
                    x = templateVars['VMM_Domains'].split(',')
                    for domain in x:
                        validating.name_rule(row_num, ws, 'VMM_Domains', domain)
                else:
                    validating.name_rule(row_num, ws, 'VMM_Domains', templateVars['VMM_Domains'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        templateVars['phys_count'] = 0
        templateVars['l3_count'] = 0
        templateVars['vmm_count'] = 0
        if not templateVars['Physical_Domains'] == None:
            if re.search(r',', templateVars['Physical_Domains']):
                x = templateVars['Physical_Domains'].split(',')
                templateVars['Physical_Domains'] = []
                for domain in x:
                    templateVars['Physical_Domains'].append(domain)
                    templateVars['phys_count'] =+ 1
            else:
                templateVars['Physical_Domains'] = [templateVars['Physical_Domains']]
                templateVars['phys_count'] =+ 1
        if not templateVars['L3_Domains'] == None:
            if re.search(r',', templateVars['L3_Domains']):
                x = templateVars['L3_Domains'].split(',')
                templateVars['L3_Domains'] = []
                for domain in x:
                    templateVars['L3_Domains'].append(domain)
                    templateVars['l3_count'] =+ 1
            else:
                templateVars['L3_Domains'] = [templateVars['L3_Domains']]
                templateVars['l3_count'] =+ 1
        if not templateVars['VMM_Domains'] == None:
            if re.search(r',', templateVars['VMM_Domains']):
                x = templateVars['VMM_Domains'].split(',')
                templateVars['VMM_Domains'] = []
                for domain in x:
                    templateVars['VMM_Domains'].append(domain)
                    templateVars['vmm_count'] =+ 1
            else:
                templateVars['VMM_Domains'] = [templateVars['VMM_Domains']]
                templateVars['vmm_count'] =+ 1

        # Define the Template Source
        template_file = "global_aep.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Policies_Global_AEP_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def apic_inb(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Node_ID': '',
                         'Pod_ID': '',
                         'Inband_EPG': ''}
        optional_args = {'Inband_IPv4': '',
                         'Inband_GWv4': '',
                         'Inband_IPv6': '',
                         'Inband_GWv6': '',}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        # Configure the Generic Template Variables
        templateVars['Device_Type'] = 'apic'
        templateVars['Type'] = 'in_band'
        templateVars['EPG'] = templateVars['Inband_EPG']
        templateVars['IPv4'] = templateVars['Inband_IPv4']
        templateVars['GWv4'] = templateVars['Inband_GWv4']
        templateVars['IPv6'] = templateVars['Inband_IPv6']
        templateVars['GWv6'] = templateVars['Inband_GWv6']

        # Initialize the Class
        lib_aci_ref = 'Access_Policies'
        class_init = '%s(ws)' % (lib_aci_ref)

        # Assign the APIC Inband Management IP's
        eval("%s.%s(wb, ws, row_num, **templateVars)" % (class_init, 'mgmt_static'))

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def cdp(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Admin_State': ''}
        optional_args = {'Description': '',
                         'Alias': ''}
        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.values(row_num, ws, 'Admin_State', templateVars['Admin_State'], ['disabled', 'enabled'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)


        # Define the Template Source
        template_file = "policy_intf_cdp.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Policies_Interface_CDP_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def fibre_channel(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Port_Mode': '',
                         'Trunk_Mode': '',
                         'Speed': '',
                         'Auto_Max_Speed': '',
                         'Fill_Pattern': '',
                         'Buffer_Credit': ''}
        optional_args = {'Description': '',
                         'Alias': ''}
        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.number_check(row_num, ws, 'Buffer_Credit', templateVars['Buffer_Credit'], 16, 64)
            validating.values(row_num, ws, 'Port_Mode', templateVars['Port_Mode'], ['f', 'np'])
            validating.values(row_num, ws, 'Trunk_Mode', templateVars['Trunk_Mode'], ['auto', 'trunk-off', 'trunk-on'])
            validating.values(row_num, ws, 'Speed', templateVars['Speed'], ['auto', '4G', '8G', '16G', '32G'])
            validating.values(row_num, ws, 'Auto_Max_Speed', templateVars['Auto_Max_Speed'], ['4G', '8G', '16G', '32G'])
            validating.values(row_num, ws, 'Fill_Pattern', templateVars['Fill_Pattern'], ['ARBFF', 'IDLE'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)


        # Define the Template Source
        template_file = "policy_intf_fc.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Policies_Interface_FC_Interface_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def intf_profile(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Switch_Role': '',
                         'Name': '',
                         'Dest_Folder': ''}
        optional_args = {'Description': ''}
        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.name_rule(row_num, ws, 'Dest_Folder', templateVars['Dest_Folder'])
            validating.values(row_num, ws, 'Switch_Role', templateVars['Switch_Role'], ['leaf', 'spine'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        if templateVars['Switch_Role'] == 'leaf':
            # Define the Template Source
            template_file = "leaf_interface_profile.jinja2"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = '%s_Interface_Profile.tf' % (templateVars['Name'])
        elif templateVars['Switch_Role'] == 'spine':
            # Define the Template Source
            template_file = "spine_interface_profile.jinja2"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = '%s_Interface_Profile.tf' % (templateVars['Name'])

        if not templateVars['Dest_Folder'] == None:
            dest_dir = '%s' % (templateVars['Dest_Folder'])
        else:
            dest_dir = 'Access'

        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def intf_selector(self, wb, ws, row_num, wr_file, **kwargs):
        if not kwargs.get('Policy_Group') == None:
            # Dicts for required and optional args
            required_args = {'Site_Group': '',
                             'Site_Name': '',
                             'Switch_Role': '',
                             'Pod_ID': '',
                             'Node_ID': '',
                             'Interface_Profile': '',
                             'Interface_Selector': '',
                             'Port': '',
                             'Policy_Group': '',
                             'Port_Type': ''}
            optional_args = {'LACP_Policy': '',
                             'Bundle_ID': '',
                             'Description': '',
                             'Switchport_Mode': '',
                             'Access_or_Native': '',
                             'Trunk_Allowed_VLANs': ''}

            kwargs['Switch_Site'] = kwargs.get('Site_Group')
            if not kwargs.get('Port_Type') == None:
                if re.search('(port-channel|vpc)', kwargs.get('Port_Type')):

                    temp_descr = kwargs['Description']
                    # Open the Access Worksheet and Find the Policy Group
                    ws_pg = wb['Access']
                    rows = ws_pg.max_row
                    row_bundle = ''
                    func = 'pg_access'
                    count = countKeys(ws_pg, func)
                    var_dict = findVars(ws_pg, func, rows, count)
                    for pos in var_dict:
                        if var_dict[pos].get('Name') == kwargs.get('Policy_Group'):
                            row_bundle = var_dict[pos]['row']
                            del var_dict[pos]['row']
                            kwargs = {**kwargs, **var_dict[pos]}
                            break

                    # Open the Network Policies Worksheet to get the Interface_Policy
                    ws_net = wb['Network Policies']
                    rows = ws_net.max_row

                    # Get the Interface Policies from the Network Policies Tab
                    func = 'intf_polgrp'
                    count = countKeys(ws_net, func)
                    row_pg = ''
                    var_dict = findVars(ws_net, func, rows, count)
                    for pos in var_dict:
                        if var_dict[pos].get('Policy_Name') == kwargs.get('Interface_Policy'):
                            row_pg = var_dict[pos]['row']
                            del var_dict[pos]['row']
                            kwargs = {**kwargs, **var_dict[pos]}
                            break

                    # Validate inputs, return dict of template vars

                    if kwargs.get('Port_Type') == 'vpc':
                        kwargs['Lag_Type'] = 'node'
                        ws_vpc = wb['Inventory']
                        for row in ws_vpc.rows:
                            if row[0].value == 'vpc_pair' and int(row[1].value) == int(kwargs.get('Switch_Site')) and str(row[5].value) == str(kwargs.get('Node_ID')):
                                kwargs['VPC_Name'] = row[2].value
                                kwargs['Name'] = '%s_vpc%s' % (row[2].value, kwargs.get('Bundle_ID'))

                            elif row[0].value == 'vpc_pair' and str(row[1].value) == str(kwargs.get('Switch_Site')) and str(row[6].value) == str(kwargs.get('Node_ID')):
                                kwargs['VPC_Name'] = row[2].value
                                kwargs['Name'] = '%s_vpc%s' % (row[2].value, kwargs.get('Bundle_ID'))
                    elif kwargs.get('Port_Type') == 'port-channel':
                        kwargs['Lag_Type'] = 'link'
                        kwargs['Name'] = '%s_pc%s' % (kwargs.get('Interface_Profile'), kwargs.get('Bundle_ID'))

                    kwargs['Description'] = temp_descr
                    # Create the Bundle Policy Group
                    kwargs['Site_Group'] = kwargs.get('Switch_Site')
                    lib_aci_ref = 'Access_Policies'
                    class_init = '%s(ws)' % (lib_aci_ref)
                    func = 'pg_bundle'
                    eval("%s.%s(wb, ws, row_num, **kwargs)" % (class_init, func))

            # Validate inputs, return dict of template vars
            templateVars = process_kwargs(required_args, optional_args, **kwargs)
            # leafx = Name
            xa = templateVars['Port'].split('/')
            xcount = len(xa)
            templateVars['Module_From'] = xa[0]
            templateVars['Module_To'] = xa[0]
            templateVars['Port_From'] = xa[1]
            templateVars['Port_To'] = xa[1]
            templateVars['Selector_Type'] = 'range'

            if templateVars['Switch_Role'] == 'leaf':
                templateVars['Policy_Group'] = kwargs.get('Name')
                # Define the Template Source
                template_file = "leaf_portselect.jinja2"
                template = self.templateEnv.get_template(template_file)

                # Process the template and write to file
                payload = template.render(templateVars)
                wr_file.write(payload + '\n\n')

                # Define the Template Source
                if xcount == 3:
                    templateVars['SubPort_From'] = xa[2]
                    templateVars['SubPort_To'] = xa[2]
                    template_file = "leaf_portblock_sub.jinja2"
                else:
                    template_file = "leaf_portblock.jinja2"
                template = self.templateEnv.get_template(template_file)

                # Process the template and write to file
                payload = template.render(templateVars)
                wr_file.write(payload + '\n\n')

            elif templateVars['Switch_Role'] == 'spine':
                # Define the Template Source
                template_file = "spine_portselect.jinja2"
                template = self.templateEnv.get_template(template_file)

                # Process the template and write to file
                payload = template.render(templateVars)
                wr_file.write(payload + '\n\n')

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def l2_interface(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'QinQ': '',
                         'Reflective_Relay': '',
                         'VLAN_Scope': ''}
        optional_args = {'Description': '',
                         'Alias': ''}
        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.values(row_num, ws, 'QinQ', templateVars['QinQ'], ['disabled', 'enabled'])
            validating.values(row_num, ws, 'Reflective_Relay', templateVars['Reflective_Relay'], ['disabled', 'enabled'])
            validating.values(row_num, ws, 'VLAN_Scope', templateVars['VLAN_Scope'], ['global', 'portlocal'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)


        # Define the Template Source
        template_file = "policy_intf_l2_interface.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Policies_Interface_L2_Interface_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def l3_domain(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'VLAN_Pool': ''}
        optional_args = { }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.name_rule(row_num, ws, 'VLAN_Pool', templateVars['VLAN_Pool'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)


        # Define the Template Source
        template_file = "domain_l3.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Policies_Domain_L3_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def leaf_pg(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Auth_8021X': '',
                         'BFD_IPv4': '',
                         'BFD_IPv6': '',
                         'BFD_MH_IPv4': '',
                         'BFD_MH_IPv6': '',
                         'CDP_Policy': '',
                         'CoPP_Leaf_Policy': '',
                         'CoPP_Pre_Filter': '',
                         'Flash_Config': '',
                         'Fast_Link_Failover': '',
                         'FC_SAN_Policy': '',
                         'FC_Node_Policy': '',
                         'Forward_Scale': '',
                         'LLDP_Policy': '',
                         'Monitoring_Policy': '',
                         'Netflow_Node': '',
                         'PoE_Policy': '',
                         'STP_Policy': ''}
        optional_args = {'Description': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.name_rule(row_num, ws, 'Auth_8021X', templateVars['Auth_8021X'])
            validating.name_rule(row_num, ws, 'BFD_IPv4', templateVars['BFD_IPv4'])
            validating.name_rule(row_num, ws, 'BFD_IPv6', templateVars['BFD_IPv6'])
            validating.name_rule(row_num, ws, 'BFD_MH_IPv4', templateVars['BFD_MH_IPv4'])
            validating.name_rule(row_num, ws, 'BFD_MH_IPv6', templateVars['BFD_MH_IPv6'])
            validating.name_rule(row_num, ws, 'CDP_Policy', templateVars['CDP_Policy'])
            validating.name_rule(row_num, ws, 'CoPP_Leaf_Policy', templateVars['CoPP_Leaf_Policy'])
            validating.name_rule(row_num, ws, 'CoPP_Pre_Filter', templateVars['CoPP_Pre_Filter'])
            validating.name_rule(row_num, ws, 'Flash_Config', templateVars['Flash_Config'])
            validating.name_rule(row_num, ws, 'Fast_Link_Failover', templateVars['Fast_Link_Failover'])
            validating.name_rule(row_num, ws, 'FC_SAN_Policy', templateVars['FC_SAN_Policy'])
            validating.name_rule(row_num, ws, 'FC_Node_Policy', templateVars['FC_Node_Policy'])
            validating.name_rule(row_num, ws, 'Forward_Scale', templateVars['Forward_Scale'])
            validating.name_rule(row_num, ws, 'LLDP_Policy', templateVars['LLDP_Policy'])
            validating.name_rule(row_num, ws, 'Monitoring_Policy', templateVars['Monitoring_Policy'])
            validating.name_rule(row_num, ws, 'Netflow_Node', templateVars['Netflow_Node'])
            validating.name_rule(row_num, ws, 'PoE_Policy', templateVars['PoE_Policy'])
            validating.name_rule(row_num, ws, 'STP_Policy', templateVars['STP_Policy'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "leaf_policy_group.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Leaf_Policy_Group_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def link_level(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Auto_Neg': '',
                         'Speed': '',
                         'Port_Delay': '',
                         'Debounce_Interval': '',
                         'FEC_Mode': ''}
        optional_args = {'Description': '',
                         'Alias': ''}
        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.number_check(row_num, ws, 'Port_Delay', templateVars['Port_Delay'], 0, 10000)
            validating.number_check(row_num, ws, 'Debounce_Interval', templateVars['Debounce_Interval'], 0, 5000)
            validating.values(row_num, ws, 'Auto_Neg', templateVars['Auto_Neg'], ['off', 'on'])
            validating.values(row_num, ws, 'Speed', templateVars['Speed'], ['inherit', '100M', '1G', '10G', '25G', '40G', '50G', '100G', '200G', '400G'])
            validating.values(row_num, ws, 'FEC_Mode', templateVars['FEC_Mode'], ['inherit', 'auto-fec', 'cl74-fc-fec', 'cl91-rs-fec', 'cons16-rs-fec', 'disable-fec', 'ieee-rs-fec', 'kp-fec'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)


        # Define the Template Source
        template_file = "policy_intf_link_level.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Policies_Interface_Link_Level_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def lldp(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Receive_State': '',
                         'Transmit_State': ''}
        optional_args = {'Description': '',
                         'Alias': ''}
        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.values(row_num, ws, 'Receive_State', templateVars['Receive_State'], ['disabled', 'enabled'])
            validating.values(row_num, ws, 'Transmit_State', templateVars['Transmit_State'], ['disabled', 'enabled'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)


        # Define the Template Source
        template_file = "policy_intf_lldp.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Policies_Interface_LLDP_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def mcp(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Admin_State': ''}
        optional_args = {'Description': '',
                         'Alias': ''}
        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.values(row_num, ws, 'Admin_State', templateVars['Admin_State'], ['disabled', 'enabled'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)


        # Define the Template Source
        template_file = "policy_intf_mcp.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Policies_Interface_MCP_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def mgmt_static(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Node_ID': '',
                         'Pod_ID': '',
                         'Device_Type': '',
                         'Type': '',
                         'EPG': ''}
        optional_args = {'IPv4': '',
                         'GWv4': '',
                         'IPv6': '',
                         'GWv6': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.hostname(row_num, ws, 'Name', templateVars['Name'])
            validating.name_rule(row_num, ws, 'EPG', templateVars['EPG'])
            validating.number_check(row_num, ws, 'Pod_ID', templateVars['Pod_ID'], 1, 15)
            if templateVars['Device_Type'] == 'apic':
                validating.number_check(row_num, ws, 'Node_ID', templateVars['Node_ID'], 1, 7)
            else:
                validating.number_check(row_num, ws, 'Node_ID', templateVars['Node_ID'], 101, 4001)
            if not templateVars['IPv4'] == None:
                validating.mgmt_network(row_num, ws, 'IPv4', templateVars['IPv4'], 'GWv4', templateVars['GWv4'])
            if not templateVars['IPv6'] == None:
                validating.mgmt_network(row_num, ws, 'IPv6', templateVars['IPv6'], 'GWv6', templateVars['GWv6'])
            else:
                templateVars['IPv6'] = '::'
                templateVars['GWv6'] = '::'
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "static_node_mgmt_address.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = '%s_%s_EPG_%s_Static_Address.tf' % (templateVars['Name'], templateVars['Type'], templateVars['EPG'])
        dest_dir = 'Tenant_mgmt'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def pg_access(self, wb, ws, row_num, **kwargs):
        # Open the Network Policies Worksheet
        ws_net = wb['Network Policies']
        rows = ws_net.max_row

        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'AEP_Policy': '',
                         'CDP_Policy': '',
                         'Link_Level': '',
                         'LLDP_Policy': '',
                         'MCP_Policy': '',
                         'STP_Policy': '',
                         'Interface_Policy': '',
                         'Policy_Name': ''}
        optional_args = {'Description': '',
                         'Alias': '',
                         'Pol_802_1x': '',
                         'poeIfPol': '',
                         'monFabricPol': '',
                         'dwdmIfPol': '',
                         'coppIfPol': '',
                         'qosDppPol_egress': '',
                         'qosDppPol_ingress': '',
                         'Fibre_Channel': '',
                         'L2_Interface': '',
                         'fabricLinkFlapPol': '',
                         'qosLlfcIfPol': '',
                         'macsecIfPol': '',
                         'netflowMonitorPol': '',
                         'Port_Security': '',
                         'qosPfcIfPol': '',
                         'qosSdIfPol': '',
                         'stormctrlIfPol': ''}

        # Get the Application Profile Policies from the Network Policies Tab
        func = 'intf_polgrp'
        count = countKeys(ws_net, func)
        row_pg = ''
        var_dict = findVars(ws_net, func, rows, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('Interface_Policy'):
                row_pg = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}
                break

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.name_rule(row_num, ws, 'AEP_Policy', templateVars['AEP_Policy'])
            validating.name_rule(row_num, ws, 'CDP_Policy', templateVars['CDP_Policy'])
            validating.name_rule(row_num, ws, 'Link_Level', templateVars['Link_Level'])
            validating.name_rule(row_num, ws, 'LLDP_Policy', templateVars['LLDP_Policy'])
            validating.name_rule(row_num, ws, 'MCP_Policy', templateVars['MCP_Policy'])
            validating.name_rule(row_num, ws, 'STP_Policy', templateVars['STP_Policy'])
            if not templateVars['Fibre_Channel'] == None:
                validating.name_rule(row_pg, ws_net, 'Fibre_Channel', templateVars['Fibre_Channel'])
            validating.name_rule(row_pg, ws_net, 'L2_Interface', templateVars['L2_Interface'])
            validating.name_rule(row_pg, ws_net, 'Port_Security', templateVars['Port_Security'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
            if not templateVars['Pol_802_1x'] == None:
                validating.name_rule(row_num, ws, 'Pol_802_1x', templateVars['Pol_802_1x'])
                templateVars['Pol_802_1x'] = 'uni/infra/portauthpol-%s' % (templateVars['Pol_802_1x'])
            if not templateVars['poeIfPol'] == None:
                validating.name_rule(row_num, ws, 'poeIfPol', templateVars['poeIfPol'])
                templateVars['poeIfPol'] = 'uni/infra/poeIfP-%s' % (templateVars['poeIfPol'])
            if not templateVars['monFabricPol'] == None:
                validating.name_rule(row_num, ws, 'monFabricPol', templateVars['monFabricPol'])
                templateVars['monFabricPol'] = 'uni/fabric/monfab-%s' % (templateVars['monFabricPol'])
            if not templateVars['dwdmIfPol'] == None:
                validating.name_rule(row_num, ws, 'dwdmIfPol', templateVars['dwdmIfPol'])
                templateVars['dwdmIfPol'] = 'uni/infra/dwdmifpol-%s' % (templateVars['dwdmIfPol'])
            if not templateVars['coppIfPol'] == None:
                validating.name_rule(row_pg, ws_net, 'coppIfPol', templateVars['coppIfPol'])
                templateVars['coppIfPol'] = 'uni/infra/coppifpol-%s' % (templateVars['coppIfPol'])
            if not templateVars['qosDppPol_egress'] == None:
                validating.name_rule(row_pg, ws_net, 'qosDppPol_egress', templateVars['qosDppPol_egress'])
                templateVars['qosDppPol_egress'] = 'uni/infra/qosdpppol-%s' % (templateVars['qosDppPol_egress'])
            if not templateVars['qosDppPol_ingress'] == None:
                validating.name_rule(row_pg, ws_net, 'qosDppPol_ingress', templateVars['qosDppPol_ingress'])
                templateVars['qosDppPol_ingress'] = 'uni/infra/qosdpppol-%s' % (templateVars['qosDppPol_ingress'])
            if not templateVars['fabricLinkFlapPol'] == None:
                validating.name_rule(row_pg, ws_net, 'fabricLinkFlapPol', templateVars['fabricLinkFlapPol'])
                templateVars['fabricLinkFlapPol'] = 'uni/infra/linkflappol-%s' % (templateVars['fabricLinkFlapPol'])
            if not templateVars['qosLlfcIfPol'] == None:
                validating.name_rule(row_pg, ws_net, 'qosLlfcIfPol', templateVars['qosLlfcIfPol'])
                templateVars['qosLlfcIfPol'] = 'uni/infra/llfc-%s' % (templateVars['qosLlfcIfPol'])
            if not templateVars['macsecIfPol'] == None:
                validating.name_rule(row_pg, ws_net, 'macsecIfPol', templateVars['macsecIfPol'])
                templateVars['macsecIfPol'] = 'uni/infra/macsecifp-%s' % (templateVars['macsecIfPol'])
            if not templateVars['netflowMonitorPol'] == None:
                validating.name_rule(row_pg, ws_net, 'netflowMonitorPol', templateVars['netflowMonitorPol'])
                templateVars['netflowMonitorPol'] = 'uni/infra/poeIfP-%s' % (templateVars['netflowMonitorPol'])
            if not templateVars['qosPfcIfPol'] == None:
                validating.name_rule(row_pg, ws_net, 'qosPfcIfPol', templateVars['qosPfcIfPol'])
                templateVars['qosPfcIfPol'] = 'uni/infra/pfc-%s' % (templateVars['qosPfcIfPol'])
            if not templateVars['qosSdIfPol'] == None:
                validating.name_rule(row_pg, ws_net, 'qosSdIfPol', templateVars['qosSdIfPol'])
                templateVars['qosSdIfPol'] = 'uni/infra/qossdpol-%s' % (templateVars['qosSdIfPol'])
            if not templateVars['stormctrlIfPol'] == None:
                validating.name_rule(row_pg, ws_net, 'stormctrlIfPol', templateVars['stormctrlIfPol'])
                templateVars['stormctrlIfPol'] = 'uni/infra/stormctrlifp-%s' % (templateVars['stormctrlIfPol'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "leaf_intf_pg_access.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Leaf_Interface_PG_Access_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def pg_bundle(self, wb, ws, row_num, **kwargs):
        # Assign the kwargs to a initial var for each process
        initial_kwargs = kwargs

        # Open the Network Policies Worksheet
        ws_net = wb['Network Policies']
        rows = ws_net.max_row

        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Lag_Type': '',
                         'AEP_Policy': '',
                         'CDP_Policy': '',
                         'Link_Level': '',
                         'LACP_Policy': '',
                         'LLDP_Policy': '',
                         'MCP_Policy': '',
                         'STP_Policy': '',
                         'Interface_Policy': '',
                         'Policy_Name': ''}
        optional_args = {'Description': '',
                         'Alias': '',
                         'monFabricPol': '',
                         'coppIfPol': '',
                         'qosDppPol_egress': '',
                         'qosDppPol_ingress': '',
                         'Fibre_Channel': '',
                         'L2_Interface': '',
                         'fabricLinkFlapPol': '',
                         'qosLlfcIfPol': '',
                         'macsecIfPol': '',
                         'netflowMonitorPol': '',
                         'Port_Security': '',
                         'qosPfcIfPol': '',
                         'qosSdIfPol': '',
                         'stormctrlIfPol': ''}

        # Get the Application Profile Policies from the Network Policies Tab
        func = 'intf_polgrp'
        count = countKeys(ws_net, func)
        row_pg = ''
        var_dict = findVars(ws_net, func, rows, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('Interface_Policy'):
                row_pg = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}
                break

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.name_rule(row_num, ws, 'AEP_Policy', templateVars['AEP_Policy'])
            validating.name_rule(row_num, ws, 'CDP_Policy', templateVars['CDP_Policy'])
            validating.name_rule(row_num, ws, 'Link_Level', templateVars['Link_Level'])
            validating.name_rule(row_num, ws, 'LACP_Policy', templateVars['LACP_Policy'])
            validating.name_rule(row_num, ws, 'LLDP_Policy', templateVars['LLDP_Policy'])
            validating.name_rule(row_num, ws, 'MCP_Policy', templateVars['MCP_Policy'])
            validating.name_rule(row_num, ws, 'STP_Policy', templateVars['STP_Policy'])
            if not templateVars['Fibre_Channel'] == None:
                validating.name_rule(row_pg, ws_net, 'Fibre_Channel', templateVars['Fibre_Channel'])
            validating.name_rule(row_pg, ws_net, 'L2_Interface', templateVars['L2_Interface'])
            validating.name_rule(row_pg, ws_net, 'Port_Security', templateVars['Port_Security'])
            validating.values(row_num, ws, 'Lag_Type', templateVars['Lag_Type'], ['link', 'node'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
            if not templateVars['monFabricPol'] == None:
                validating.name_rule(row_num, ws, 'monFabricPol', templateVars['monFabricPol'])
                templateVars['monFabricPol'] = 'uni/fabric/monfab-%s' % (templateVars['monFabricPol'])
            if not templateVars['coppIfPol'] == None:
                validating.name_rule(row_pg, ws_net, 'coppIfPol', templateVars['coppIfPol'])
                templateVars['coppIfPol'] = 'uni/infra/coppifpol-%s' % (templateVars['coppIfPol'])
            if not templateVars['qosDppPol_egress'] == None:
                validating.name_rule(row_pg, ws_net, 'qosDppPol_egress', templateVars['qosDppPol_egress'])
                templateVars['qosDppPol_egress'] = 'uni/infra/qosdpppol-%s' % (templateVars['qosDppPol_egress'])
            if not templateVars['qosDppPol_ingress'] == None:
                validating.name_rule(row_pg, ws_net, 'qosDppPol_ingress', templateVars['qosDppPol_ingress'])
                templateVars['qosDppPol_ingress'] = 'uni/infra/qosdpppol-%s' % (templateVars['qosDppPol_ingress'])
            if not templateVars['fabricLinkFlapPol'] == None:
                validating.name_rule(row_pg, ws_net, 'fabricLinkFlapPol', templateVars['fabricLinkFlapPol'])
                templateVars['fabricLinkFlapPol'] = 'uni/infra/linkflappol-%s' % (templateVars['fabricLinkFlapPol'])
            if not templateVars['qosLlfcIfPol'] == None:
                validating.name_rule(row_pg, ws_net, 'qosLlfcIfPol', templateVars['qosLlfcIfPol'])
                templateVars['qosLlfcIfPol'] = 'uni/infra/llfc-%s' % (templateVars['qosLlfcIfPol'])
            if not templateVars['macsecIfPol'] == None:
                validating.name_rule(row_pg, ws_net, 'macsecIfPol', templateVars['macsecIfPol'])
                templateVars['macsecIfPol'] = 'uni/infra/macsecifp-%s' % (templateVars['macsecIfPol'])
            if not templateVars['netflowMonitorPol'] == None:
                validating.name_rule(row_pg, ws_net, 'netflowMonitorPol', templateVars['netflowMonitorPol'])
                templateVars['netflowMonitorPol'] = 'uni/infra/poeIfP-%s' % (templateVars['netflowMonitorPol'])
            if not templateVars['qosPfcIfPol'] == None:
                validating.name_rule(row_pg, ws_net, 'qosPfcIfPol', templateVars['qosPfcIfPol'])
                templateVars['qosPfcIfPol'] = 'uni/infra/pfc-%s' % (templateVars['qosPfcIfPol'])
            if not templateVars['qosSdIfPol'] == None:
                validating.name_rule(row_pg, ws_net, 'qosSdIfPol', templateVars['qosSdIfPol'])
                templateVars['qosSdIfPol'] = 'uni/infra/qossdpol-%s' % (templateVars['qosSdIfPol'])
            if not templateVars['stormctrlIfPol'] == None:
                validating.name_rule(row_pg, ws_net, 'stormctrlIfPol', templateVars['stormctrlIfPol'])
                templateVars['stormctrlIfPol'] = 'uni/infra/stormctrlifp-%s' % (templateVars['stormctrlIfPol'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "leaf_intf_pg_bundle.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Leaf_Interface_PG_Bundle_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def pg_breakout(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Breakout_Map': ''}
        optional_args = {'Description': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.values(row_num, ws, 'Breakout_Map', templateVars['Breakout_Map'], ['100g-2x', '100g-4x', '10g-4x', '25g-4x', '50g-8x'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "leaf_intf_pg_breakout.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Leaf_Interface_PG_Breakout_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def pg_spine(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'AEP_Policy': '',
                         'CDP_Policy': '',
                         'Link_Level': ''}
        optional_args = {'Description': '',
                         'Alias': '',
                         'fabricLinkFlapPol': '',
                         'macsecIfPol': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.name_rule(row_num, ws, 'AEP_Policy', templateVars['AEP_Policy'])
            validating.name_rule(row_num, ws, 'CDP_Policy', templateVars['CDP_Policy'])
            validating.name_rule(row_num, ws, 'Link_Level', templateVars['Link_Level'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
            if not templateVars['fabricLinkFlapPol'] == None:
                validating.name_rule(row_num, ws, 'fabricLinkFlapPol', templateVars['fabricLinkFlapPol'])
                templateVars['fabricLinkFlapPol'] = 'uni/infra/linkflappol-%s' % (templateVars['fabricLinkFlapPol'])
            if not templateVars['macsecIfPol'] == None:
                validating.name_rule(row_num, ws, 'macsecIfPol', templateVars['macsecIfPol'])
                templateVars['macsecIfPol'] = 'uni/infra/macsecifpol-%s' % (templateVars['macsecIfPol'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "spine_intf_pg_access.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Spine_Interface_PG_Access_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def phys_dom(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'VLAN_Pool': ''}
        optional_args = { }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.name_rule(row_num, ws, 'VLAN_Pool', templateVars['VLAN_Pool'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)


        # Define the Template Source
        template_file = "domain_phys.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Policies_Domain_Phys_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def port_channel(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Mode': '',
                         'Min_Links': '',
                         'Max_Links': '',
                         'Fast_Select': '',
                         'Graceful': '',
                         'Load_Defer': '',
                         'Suspend_Individual': '',
                         'Symmetric_Hash': ''}
        optional_args = {'Description': '',
                         'Alias': ''}
        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.number_check(row_num, ws, 'Min_Links', templateVars['Min_Links'], 1, 16)
            validating.number_check(row_num, ws, 'Max_Links', templateVars['Max_Links'], 1, 16)
            validating.values(row_num, ws, 'Mode', templateVars['Mode'], ['active', 'explicit-failover', 'mac-pin', 'mac-pin-nicload', 'off', 'passive'])
            validating.values(row_num, ws, 'Fast_Select', templateVars['Fast_Select'], ['no', 'yes'])
            validating.values(row_num, ws, 'Graceful', templateVars['Graceful'], ['no', 'yes'])
            validating.values(row_num, ws, 'Load_Defer', templateVars['Load_Defer'], ['no', 'yes'])
            validating.values(row_num, ws, 'Suspend_Individual', templateVars['Suspend_Individual'], ['no', 'yes'])
            validating.values(row_num, ws, 'Symmetric_Hash', templateVars['Symmetric_Hash'], ['no', 'yes'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Create ctrl templateVars
        ctrl_count = 0
        Ctrl = ''
        if templateVars['Fast_Select'] == 'yes':
            Ctrl = '"fast-sel-hot-stdby"'
            ctrl_count =+ 1
        if templateVars['Graceful'] == 'yes' and ctrl_count > 0:
            Ctrl = Ctrl + ',' + '"graceful-conv"'
            ctrl_count =+ 1
        elif templateVars['Graceful'] == 'yes':
            ctrl = '"graceful-conv"'
            ctrl_count =+ 1
        if templateVars['Load_Defer'] == 'yes' and ctrl_count > 0:
            Ctrl = Ctrl + ',' + '"load-defer"'
            ctrl_count =+ 1
        elif templateVars['Load_Defer'] == 'yes':
            Ctrl = '"load-defer"'
            ctrl_count =+ 1
        if templateVars['Suspend_Individual'] == 'yes' and ctrl_count > 0:
            Ctrl = Ctrl + ',' + '"susp-individual"'
            ctrl_count =+ 1
        elif templateVars['Suspend_Individual'] == 'yes':
            Ctrl = '"susp-individual"'
            ctrl_count =+ 1
        if templateVars['Symmetric_Hash'] == 'yes' and ctrl_count > 0:
            Ctrl = Ctrl + ',' + '"symmetric-hash"'
            ctrl_count =+ 1
        elif templateVars['Symmetric_Hash'] == 'yes':
            Ctrl = '"symmetric-hash"'
            ctrl_count =+ 1
        if ctrl_count > 0:
            templateVars['Ctrl'] = '[%s]' % (Ctrl)
        else:
            templateVars['Ctrl'] = '["unspecified"]'

        # Define the Template Source
        template_file = "policy_intf_lacp.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Policies_Interface_Port_Channel_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def port_cnvt(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Node_ID': '',
                         'Port': ''}
        optional_args = { }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.hostname(row_num, ws, 'Name', templateVars['Name'])
            validating.number_check(row_num, ws, 'Node_ID', templateVars['Node_ID'], 101, 4001)
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Create Port Name Var
        zz = templateVars['Port'].split('/')
        templateVars['Port_Name'] = '%s_%s' % (zz[0], zz[1])

        # Define the Template Source
        template_file = "downlink.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Downlink_Convert_%s.tf' % (templateVars['Port_Name'])
        dest_dir = 'Access/%s' % (templateVars['Name'])
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def port_security(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Timeout': '',
                         'Maximum_Endpoints': ''}
        optional_args = {'Description': '',
                         'Alias': ''}
        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.number_check(row_num, ws, 'Timeout', templateVars['Timeout'], 60, 3600)
            validating.number_check(row_num, ws, 'Maximum_Endpoints', templateVars['Maximum_Endpoints'], 0, 12000)
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)


        # Define the Template Source
        template_file = "policy_intf_port_sec.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Policies_Interface_Port_Security_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def spine_pg(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'BFD_IPv4': '',
                         'BFD_IPv6': '',
                         'CDP_Policy': '',
                         'CoPP_Pre_Filter': '',
                         'CoPP_Spine_Policy': '',
                         'LLDP_Policy': ''}
        optional_args = {'Description': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.name_rule(row_num, ws, 'BFD_IPv4', templateVars['BFD_IPv4'])
            validating.name_rule(row_num, ws, 'BFD_IPv6', templateVars['BFD_IPv6'])
            validating.name_rule(row_num, ws, 'CDP_Policy', templateVars['CDP_Policy'])
            validating.name_rule(row_num, ws, 'CoPP_Pre_Filter', templateVars['CoPP_Pre_Filter'])
            validating.name_rule(row_num, ws, 'CoPP_Spine_Policy', templateVars['CoPP_Spine_Policy'])
            validating.name_rule(row_num, ws, 'LLDP_Policy', templateVars['LLDP_Policy'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "spine_policy_group.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Spine_Policy_Group_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def stp(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Filter': '',
                         'Guard': ''}
        optional_args = {'Description': '',
                         'Alias': ''}
        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.values(row_num, ws, 'Filter', templateVars['Filter'], ['disabled', 'enabled'])
            validating.values(row_num, ws, 'Guard', templateVars['Guard'], ['disabled', 'enabled'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Create ctrl templateVars
        ctrl_count = 0
        Ctrl = ''
        if templateVars['Filter'] == 'enabled':
            Ctrl = '"bpdu-filter"'
            ctrl_count =+ 1
        if templateVars['Guard'] == 'enabled' and ctrl_count > 0:
            Ctrl = Ctrl + ',' + '"bpdu-guard"'
            ctrl_count =+ 1
        elif templateVars['Guard'] == 'enabled':
            Ctrl = '"bpdu-guard"'
            ctrl_count =+ 1
        if ctrl_count > 0:
            templateVars['Ctrl'] = '[%s]' % (Ctrl)
        else:
            templateVars['Ctrl'] = '["unspecified"]'

        # Define the Template Source
        template_file = "policy_intf_stp.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Policies_Interface_STP_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def switch(self, wb, ws, row_num, **kwargs):
        # Initialize the Class
        lib_aci_ref = 'Access_Policies'
        class_init = '%s(ws)' % (lib_aci_ref)

        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Serial': '',
                         'Node_ID': '',
                         'Name': '',
                         'Profiles': '',
                         'Node_Type': '',
                         'Pod_ID': '',
                         'Switch_Role': '',
                         'Switch_Type': '',
                         'Is_Virtual': '',
                         'Tier-2': '',
                         'Inband_EPG': '',
                         'OOB_EPG': ''}
        optional_args = {'Policy_Group': '',
                         'Remote_ID': '',
                         'Fabric_ID': '',
                         'MG_Name': '',
                         'Inband_IPv4': '',
                         'Inband_GWv4': '',
                         'Inband_IPv6': '',
                         'Inband_GWv6': '',
                         'OOB_IPv4': '',
                         'OOB_GWv4': '',
                         'OOB_IPv6': '',
                         'OOB_GWv6': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        # Use Switch_Type to Determine the Number of ports on the switch
        modules,port_count = query_switch_model(row_num, templateVars['Switch_Type'])

        try:
            # Validate Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.hostname(row_num, ws, 'Name', templateVars['Name'])
            validating.modules(row_num, templateVars['Name'], templateVars['Switch_Role'], modules)
            if not templateVars['MG_Name'] == None:
                validating.name_rule(row_num, ws, 'MG_Name', templateVars['MG_Name'])
            validating.name_rule(row_num, ws, 'Inband_EPG', templateVars['Inband_EPG'])
            validating.name_rule(row_num, ws, 'OOB_EPG', templateVars['OOB_EPG'])
            validating.number_check(row_num, ws, 'Node_ID', templateVars['Node_ID'], 101, 4001)
            validating.number_check(row_num, ws, 'Pod_ID', templateVars['Pod_ID'], 1, 12)
            validating.number_check(row_num, ws, 'Fabric_ID', templateVars['Fabric_ID'], 1, 12)
            validating.port_count(row_num, templateVars['Name'], templateVars['Switch_Role'], port_count)
            validating.values(row_num, ws, 'Profiles', templateVars['Profiles'], ['no', 'yes'])
            validating.values(row_num, ws, 'Node_Type', templateVars['Node_Type'], ['remote-leaf-wan', 'unspecified'])
            validating.values(row_num, ws, 'Switch_Role', templateVars['Switch_Role'], ['leaf', 'spine'])
            validating.values(row_num, ws, 'Is_Virtual', templateVars['Is_Virtual'], ['no', 'yes'])
            validating.values(row_num, ws, 'Tier-2', templateVars['Tier-2'], ['no', 'yes'])
            if templateVars['Profiles'] == 'yes':
                validating.name_rule(row_num, ws, 'Policy_Group', templateVars['Policy_Group'])
            if not templateVars['Remote_ID'] == None:
                validating.number_check(row_num, ws, 'Remote_ID', templateVars['Remote_ID'], 1, 255)
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Configure the Generic Template Variables for Node Management Inband IP's
        templateVars['Device_Type'] = 'switch'
        templateVars['Type'] = 'in_band'
        templateVars['EPG'] = templateVars['Inband_EPG']
        templateVars['IPv4'] = templateVars['Inband_IPv4']
        templateVars['GWv4'] = templateVars['Inband_GWv4']
        templateVars['IPv6'] = templateVars['Inband_IPv6']
        templateVars['GWv6'] = templateVars['Inband_GWv6']

        if re.search('\d', templateVars['IPv4']) or re.search('\d', templateVars['IPv6']):
            # Assign the Switch Inband Management IP's
            eval("%s.%s(wb, ws, row_num, **templateVars)" % (class_init, 'mgmt_static'))

        # Configure the Generic Template Variables for Node Management Out-of-Band IP's
        templateVars['Device_Type'] = 'switch'
        templateVars['Type'] = 'out_of_band'
        templateVars['EPG'] = templateVars['OOB_EPG']
        templateVars['IPv4'] = templateVars['OOB_IPv4']
        templateVars['GWv4'] = templateVars['OOB_GWv4']
        templateVars['IPv6'] = templateVars['OOB_IPv6']
        templateVars['GWv6'] = templateVars['OOB_GWv6']

        if re.search('\d', templateVars['IPv4']) or re.search('\d', templateVars['IPv6']):
            # Assign the Switch Out-of-Band Management IP's
            eval("%s.%s(wb, ws, row_num, **templateVars)" % (class_init, 'mgmt_static'))

        if not templateVars['MG_Name'] == None:
            # Define the Template Source
            template_file = "maint_group_nodeblk.jinja2"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = 'Maintenance_Group_%s.tf' % (templateVars['MG_Name'])
            dest_dir = 'Admin'
            process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

        Site_ID = 'Site_ID_%s' % (templateVars['Site_Group'])
        site_dict = ast.literal_eval(os.environ[Site_ID])

        # Create kwargs for Site Variables
        templateVars['Site_ID'] = site_dict.get('Site_ID')
        templateVars['Site_Name'] = site_dict.get('Site_Name')
        templateVars['APIC_URL'] = site_dict.get('APIC_URL')
        templateVars['APIC_Version'] = site_dict.get('APIC_Version')
        templateVars['APIC_Auth_Type'] = site_dict.get('APIC_Auth_Type')
        templateVars['Terraform_EQ'] = site_dict.get('Terraform_EQ')
        templateVars['Terraform_Version'] = site_dict.get('Terraform_Version')
        templateVars['Provider_EQ'] = site_dict.get('Provider_EQ')
        templateVars['Provider_Version'] = site_dict.get('Provider_Version')
        templateVars['Run_Location'] = site_dict.get('Run_Location')
        templateVars['State_Location'] = site_dict.get('State_Location')
        templateVars['Terraform_Cloud_Org'] = site_dict.get('Terraform_Cloud_Org')
        templateVars['Workspace_Prefix'] = site_dict.get('Workspace_Prefix')
        templateVars['VCS_Base_Repo'] = site_dict.get('VCS_Base_Repo')
        templateVars['Terraform_Agent_Pool_ID'] = site_dict.get('Terraform_Agent_Pool_ID')

        self.templateLoader = jinja2.FileSystemLoader(searchpath=(aci_template_path + 'Access_Policies/'))
        self.templateEnv = jinja2.Environment(loader=self.templateLoader)
        excel_wkbook = '%s_intf_selectors.xlsx' % (templateVars['Site_Name'])

        wb_sw = load_workbook(excel_wkbook)

        # Check if there is a Worksheet for the Switch Already
        if not templateVars['Name'] in wb_sw.sheetnames:
            ws_sw = wb_sw.create_sheet(title = templateVars['Name'])
            ws_sw = wb_sw[templateVars['Name']]
            ws_sw.column_dimensions['A'].width = 15
            ws_sw.column_dimensions['B'].width = 10
            ws_sw.column_dimensions['c'].width = 10
            ws_sw.column_dimensions['d'].width = 20
            ws_sw.column_dimensions['E'].width = 20
            ws_sw.column_dimensions['F'].width = 10
            ws_sw.column_dimensions['G'].width = 20
            ws_sw.column_dimensions['H'].width = 20
            ws_sw.column_dimensions['I'].width = 20
            ws_sw.column_dimensions['J'].width = 15
            ws_sw.column_dimensions['K'].width = 30
            ws_sw.column_dimensions['L'].width = 20
            ws_sw.column_dimensions['M'].width = 20
            ws_sw.column_dimensions['N'].width = 30
            dv1 = DataValidation(type="list", formula1='"intf_selector"', allow_blank=True)
            dv2 = DataValidation(type="list", formula1='"access,breakout,port-channel,vpc"', allow_blank=True)
            ws_sw.add_data_validation(dv1)
            ws_sw.add_data_validation(dv2)
            ws_header = '%s Interface Selectors' % (templateVars['Name'])
            data = [ws_header]
            ws_sw.append(data)
            ws_sw.merge_cells('A1:N1')
            for cell in ws_sw['1:1']:
                cell.style = 'Heading 1'
            data = ['','Notes: Breakout Policy Group Names are 2x100g_pg, 4x10g_pg, 4x25g_pg, 4x100g_pg, 8x50g_pg.']
            ws_sw.append(data)
            ws_sw.merge_cells('B2:N2')
            for cell in ws_sw['2:2']:
                cell.style = 'Heading 2'
            data = ['Type','Pod_ID','Node_ID','Interface_Profile','Interface_Selector','Port','Policy_Group','Port_Type','LACP_Policy','Bundle_ID','Description','Switchport_Mode','Access_or_Native','Trunk_Allowed_VLANs']
            ws_sw.append(data)
            for cell in ws_sw['3:3']:
                cell.style = 'Heading 3'

            ws_sw_row_count = 4
            templateVars['dv1'] = dv1
            templateVars['dv2'] = dv2
            templateVars['port_count'] = port_count
            sw_type = str(templateVars['Switch_Type'])
            sw_name = str(templateVars['Name'])
            if re.search('^(93[0-9][0-9])', sw_type):
                for module in range(1, 2):
                    templateVars['module'] = module
                    ws_sw_row_count = create_selector(ws_sw, ws_sw_row_count, **templateVars)
            if re.search('^(9396|95[0-1][4-8])', sw_type):
                row_count = 1
                for row in ws.rows:
                    if re.search('9396', sw_type):
                        start, end = 2, 2
                    else:
                        start, end = 1, int(modules)
                    if str(row[0].value) == sw_type and str(row[2].value) == sw_name:
                        for module in range(start, end + 2):
                            templateVars['module'] = module
                            module_type = row[module + 2].value
                            if module_type == None:
                                module_type = 'none'
                            elif re.search('(X97|M(4|6|12)P)', module_type):
                                templateVars['port_count'] = query_module_type(row_count, module_type)
                                ws_sw_row_count = create_selector(ws_sw, ws_sw_row_count, **templateVars)
                        row_count += 1
                        break
            wb_sw.save(excel_wkbook)
        else:
            ws_sw = wb_sw[templateVars['Name']]

        # Define the Template Source
        template_file = "inventory.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = '%s.tf' % (templateVars['Name'])
        dest_dir = '%s' % (templateVars['Name'])
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        if templateVars['Profiles'] == 'yes':
            templateVars['Description'] = None
            templateVars['Dest_Folder'] = templateVars['Name']
            eval("%s.%s(wb, ws, row_num, **templateVars)" % (class_init, 'intf_profile'))

            templateVars['Selector_Name'] = templateVars['Name']
            templateVars['Association_Type'] = 'range'
            templateVars['Nodeblk_Name'] = 'blk%s-%s' % (templateVars['Node_ID'], templateVars['Node_ID'])
            templateVars['Node_ID_From'] = templateVars['Node_ID']
            templateVars['Node_ID_To'] = templateVars['Node_ID']
            templateVars['Interface_Profile'] = templateVars['Name']
            eval("%s.%s(wb, ws, row_num, **templateVars)" % (class_init, 'sw_profile'))

            sw_intf_profile = './ACI/%s/%s/%s_Interface_Profile.tf' % (templateVars['Site_Name'], templateVars['Name'], templateVars['Name'])
            wr_file = open(sw_intf_profile, 'a+')
            lib_aci_ref = 'Access_Policies'
            rows_sw = ws_sw.max_row
            func_regex = re.compile('^intf_selector$')
            func_list = findKeys(ws_sw, func_regex)
            class_init = '%s(ws_sw)' % (lib_aci_ref)
            stdout_log(ws_sw, None)
            for func in func_list:
                count = countKeys(ws_sw, func)
                var_dict = findVars(ws_sw, func, rows_sw, count)
                for pos in var_dict:
                    row_num = var_dict[pos]['row']
                    del var_dict[pos]['row']
                    for x in list(var_dict[pos].keys()):
                        if var_dict[pos][x] == '':
                            del var_dict[pos][x]
                    stdout_log(ws_sw, row_num)
                    var_dict[pos]['Site_Group'] = templateVars['Site_ID']
                    var_dict[pos]['Switch_Role'] = templateVars['Switch_Role']
                    var_dict[pos]['Site_Name'] = templateVars['Site_Name']
                    eval("%s.%s(wb, ws_sw, row_num, wr_file, **var_dict[pos])" % (class_init, func))
            wr_file.close()
            ws_wr = wb_sw.get_sheet_names()
            for sheetName in ws_wr:
                if sheetName in ['Sites']:
                    sheetToDelete = wb_sw.get_sheet_by_name(sheetName)
                    wb_sw.remove_sheet(sheetToDelete)
                    wb_sw.save(excel_wkbook)
            wb_sw.close()

        if re.search('Grp_[A-F]', templateVars['Site_Group']):
            print(f"\n-----------------------------------------------------------------------------\n")
            print(f"   Error on Worksheet {ws.title}, Row {row_num} Site_Group, value {templateVars['Site_Group']}.")
            print(f"   A Leaf can only be assigned to one Site.  Exiting....")
            print(f"\n-----------------------------------------------------------------------------\n")
            exit()
        elif re.search(r'\d+', templateVars['Site_Group']):
            Site_ID = 'Site_ID_%s' % (templateVars['Site_Group'])
            site_dict = ast.literal_eval(os.environ[Site_ID])

            # Set Destination Directory
            dest_dir = '%s' % (templateVars['Name'])

            # Create kwargs for Site Variables
            kwargs['Site_ID'] = site_dict.get('Site_ID')
            kwargs['Site_Name'] = site_dict.get('Site_Name')
            kwargs['APIC_URL'] = site_dict.get('APIC_URL')
            kwargs['APIC_Version'] = site_dict.get('APIC_Version')
            kwargs['APIC_Auth_Type'] = site_dict.get('APIC_Auth_Type')
            kwargs['Terraform_EQ'] = site_dict.get('Terraform_EQ')
            kwargs['Terraform_Version'] = site_dict.get('Terraform_Version')
            kwargs['Provider_EQ'] = site_dict.get('Provider_EQ')
            kwargs['Provider_Version'] = site_dict.get('Provider_Version')
            kwargs['Run_Location'] = site_dict.get('Run_Location')
            kwargs['State_Location'] = site_dict.get('State_Location')
            kwargs['Terraform_Cloud_Org'] = site_dict.get('Terraform_Cloud_Org')
            kwargs['Workspace_Prefix'] = site_dict.get('Workspace_Prefix')
            kwargs['VCS_Base_Repo'] = site_dict.get('VCS_Base_Repo')
            kwargs['Terraform_Agent_Pool_ID'] = site_dict.get('Terraform_Agent_Pool_ID')

            # Dicts for required and optional args
            required_args = {'Site_ID': '',
                                'Site_Name': '',
                                'APIC_URL': '',
                                'APIC_Version': '',
                                'APIC_Auth_Type': '',
                                'Terraform_EQ': '',
                                'Terraform_Version': '',
                                'Provider_EQ': '',
                                'Provider_Version': '',
                                'Run_Location': '',
                                'State_Location': ''}
            optional_args = {'Terraform_Cloud_Org': '',
                                'Workspace_Prefix': '',
                                'VCS_Base_Repo': '',
                                'Terraform_Agent_Pool_ID': ''}

            # Validate inputs, return dict of template vars
            templateVars = process_kwargs(required_args, optional_args, **kwargs)

            # If the State_Location is Terraform_Cloud Configure Workspaces in the Cloud
            if templateVars['State_Location'] == 'Terraform_Cloud':
                # Initialize the Class
                lib_tf_ref = 'lib_terraform.Terraform_Cloud'
                class_init = '%s()' % (lib_tf_ref)

                # Get terraform_cloud_token
                kwargs['terraform_cloud_token'] = eval("%s.%s()" % (class_init, 'terraform_token'))

                # Get terraform_cloud_token
                kwargs['terraform_oath_token'] = eval("%s.%s(**kwargs)" % (class_init, 'oath_token'))

                # Get workspace_ids
                workspace_dict = {}
                workspace_dict = tf_workspace(class_init, dest_dir, workspace_dict, **kwargs)

            # If the Run_Location is Terraform_Cloud Configure Variables in the Cloud
            if templateVars['Run_Location'] == 'Terraform_Cloud':
                # Set Variable List
                if templateVars['APIC_Auth_Type'] == 'user_pass':
                    var_list = ['aciUrl', 'aciUser', 'aciPass']
                else:
                    var_list = ['aciUrl', 'aciCertName', 'aciPrivateKey']

                # Get var_ids
                tf_var_dict = {}
                folder_id = 'Site_ID_%s_%s' % (templateVars['Site_ID'], dest_dir)
                kwargs['workspace_id'] = workspace_dict[folder_id]
                kwargs['Description'] = ''
                for var in var_list:
                    tf_var_dict = tf_variables(class_init, dest_dir, var, tf_var_dict, **kwargs)

        else:
            print(f"\n-----------------------------------------------------------------------------\n")
            print(f"   Error on Worksheet {ws.title}, Row {row_num} Site_Group, value {templateVars['Site_Group']}.")
            print(f"   Unable to Determine if this is a Single or Group of Site(s).  Exiting....")
            print(f"\n-----------------------------------------------------------------------------\n")
            exit()

        self.templateLoader = jinja2.FileSystemLoader(
            searchpath=(aci_template_path))
        self.templateEnv = jinja2.Environment(loader=self.templateLoader)

        # Add the Default Files to the Tenant Directory
        file_list = ['.gitignore_.gitignore', 'main.jinja2_main.tf', 'variables.jinja2_variables.tf']
        for file in file_list:
            x = file.split('_')
            template_file = x[0]
            dest_file = x[1]
            template = self.templateEnv.get_template(template_file)
            create_tf_file('w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def sw_profile(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Switch_Role': '',
                         'Name': '',
                         'Selector_Name': '',
                         'Association_Type': '',
                         'Nodeblk_Name': '',
                         'Node_ID_From': '',
                         'Node_ID_To': '',
                         'Policy_Group': '',
                         'Interface_Profile': '',
                         'Dest_Folder': ''}
        optional_args = {'Description': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.name_rule(row_num, ws, 'Selector_Name', templateVars['Selector_Name'])
            validating.name_rule(row_num, ws, 'Nodeblk_Name', templateVars['Nodeblk_Name'])
            validating.name_rule(row_num, ws, 'Policy_Group', templateVars['Policy_Group'])
            validating.name_rule(row_num, ws, 'Interface_Profile', templateVars['Interface_Profile'])
            validating.name_rule(row_num, ws, 'Dest_Folder', templateVars['Dest_Folder'])
            validating.number_check(row_num, ws, 'Node_ID_From', templateVars['Node_ID_From'], 101, 4001)
            validating.number_check(row_num, ws, 'Node_ID_To', templateVars['Node_ID_To'], 101, 4001)
            validating.values(row_num, ws, 'Switch_Role', templateVars['Switch_Role'], ['leaf', 'spine'])
            validating.values(row_num, ws, 'Association_Type', templateVars['Association_Type'], ['ALL', 'range', 'ALL_IN_POD'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        if templateVars['Switch_Role'] == 'leaf':
            template_file = "leaf_profile.jinja2"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = '%s_leaf_profile.tf' % (templateVars['Name'])
        else:
            template_file = "spine_profile.jinja2"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = '%s_spine_profile.tf' % (templateVars['Name'])

        if not templateVars['Dest_Folder'] == None:
            dest_dir = '%s' % (templateVars['Dest_Folder'])
        else:
            dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def vlan_pool(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Allocation_Mode': '',
                         'VLAN_Grp1': '',
                         'VGRP1_Allocation': ''}
        optional_args = {'Description': '',
                         'Alias': '',
                         'VLAN_Grp1': '',
                         'VGRP1_Allocation': '',
                         'VLAN_Grp2': '',
                         'VGRP2_Allocation': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.values(row_num, ws, 'Allocation_Mode', templateVars['Allocation_Mode'], ['dynamic', 'static'])
            validating.values(row_num, ws, 'VGRP1_Allocation', templateVars['VGRP1_Allocation'], ['dynamic', 'static'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
            if not templateVars['VGRP2_Allocation'] == None:
                validating.values(row_num, ws, 'VGRP2_Allocation', templateVars['VGRP2_Allocation'], ['dynamic', 'static'])
            validating.vlans(row_num, ws, 'VLAN_Grp1', templateVars['VLAN_Grp1'])
            if not templateVars['VLAN_Grp2'] == None:
                validating.vlans(row_num, ws, 'VLAN_Grp2', templateVars['VLAN_Grp2'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        if templateVars['Name'] == None:
            Error_Return = 'Error on Worksheet %s Row %s.  Could not Determine the Name of the VLAN Pool.' % (ws.title, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "vlan_pool.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'vlan_pool_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "data_vlan_pool.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'data_vlan_pool_%s.tf' % (templateVars['Name'])
        dest_dir = 'VLANs'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # Add VLAN(s) to VLAN Pool FIle
        if re.search('Grp_[A-F]', templateVars['Site_Group']):
            Group_ID = '%s' % (templateVars['Site_Group'])
            site_group = ast.literal_eval(os.environ[Group_ID])
            for x in range(1, 13):
                sitex = 'Site_%s' % (x)
                if not site_group[sitex] == None:
                    Site_ID = 'Site_ID_%s' % (site_group[sitex])
                    site_dict = ast.literal_eval(os.environ[Site_ID])

                    # Create templateVars for Site_Name and APIC_URL
                    templateVars['Site_Name'] = site_dict.get('Site_Name')
                    templateVars['APIC_URL'] = site_dict.get('APIC_URL')

                    # Create Blank VLAN Pool VLAN(s) File
                    dest_file = './ACI/%s/VLANs/vlp_%s.tf' % (templateVars['Site_Name'], templateVars['Name'])
                    wr_file = open(dest_file, 'w')
                    wr_file.close()
                    dest_file = 'vlan_pool_%s.tf' % (templateVars['Name'])
                    dest_dir = 'VLANs'
                    template_file = "add_vlan_to_pool.jinja2"
                    template = self.templateEnv.get_template(template_file)

                    for z in range(1, 3):
                        vgroup = 'VLAN_Grp%s' % (z)
                        vgrp = 'VGRP%s_Allocation' % (z)
                        templateVars['Allocation_Mode'] = templateVars[vgrp]
                        if re.search(r'\d+', str(templateVars[vgroup])):
                            vlan_list = vlan_list_full(templateVars[vgroup])
                            for v in vlan_list:
                                vlan = str(v)
                                if re.fullmatch(r'\d+', vlan):
                                    templateVars['VLAN_ID'] = int(vlan)

                                    # Add VLAN to VLAN Pool File
                                    create_tf_file('a+', dest_dir, dest_file, template, **templateVars)

        elif re.search(r'\d+', templateVars['Site_Group']):
            Site_ID = 'Site_ID_%s' % (templateVars['Site_Group'])
            site_dict = ast.literal_eval(os.environ[Site_ID])

            # Create templateVars for Site_Name and APIC_URL
            templateVars['Site_Name'] = site_dict.get('Site_Name')
            templateVars['APIC_URL'] = site_dict.get('APIC_URL')

            # Create Blank VLAN Pool VLAN(s) File
            dest_file = './ACI/%s/VLANs/vlan_pool_%s.tf' % (templateVars['Site_Name'], templateVars['Name'])
            wr_file = open(dest_file, 'w')
            wr_file.close()
            dest_file = 'vlan_pool_%s.tf' % (templateVars['Name'])
            dest_dir = 'VLANs'
            template_file = "add_vlan_to_pool.jinja2"
            template = self.templateEnv.get_template(template_file)

            for z in range(1, 3):
                vgroup = 'VLAN_Grp%s' % (z)
                vgrp = 'VGRP%s_Allocation' % (z)
                templateVars['Allocation_Mode'] = templateVars[vgrp]
                if re.search(r'\d+', str(templateVars[vgroup])):
                    vlan_list = vlan_list_full(templateVars[vgroup])
                    for v in vlan_list:
                        vlan = str(v)
                        if re.fullmatch(r'\d+', vlan):
                            templateVars['VLAN_ID'] = int(vlan)

                            # Add VLAN to VLAN Pool File
                            create_tf_file('a+', dest_dir, dest_file, template, **templateVars)
        else:
            print(f"\n-----------------------------------------------------------------------------\n")
            print(f"   Error on Worksheet {ws.title}, Row {row_num} Site_Group, value {templateVars['Site_Group']}.")
            print(f"   Unable to Determine if this is a Single or Group of Site(s).  Exiting....")
            print(f"\n-----------------------------------------------------------------------------\n")
            exit()

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def vpc_pair(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'VPC_ID': '',
                         'Name': '',
                         'Node1_ID': '',
                         'Node2_ID': ''}
        optional_args = { }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.number_check(row_num, ws, 'VPC_ID', templateVars['VPC_ID'], 1, 1000)
            validating.number_check(row_num, ws, 'Node1_ID', templateVars['Node1_ID'], 101, 4001)
            validating.number_check(row_num, ws, 'Node2_ID', templateVars['Node2_ID'], 101, 4001)
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "vpc_domain.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'vpc_domain_%s.tf' % (templateVars['VPC_ID'])
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

# Terraform ACI Provider - Admin Policies
# Class must be instantiated with Variables
class Admin_Policies(object):
    def __init__(self, ws):
        self.templateLoader = jinja2.FileSystemLoader(
            searchpath=(aci_template_path + 'Admin_Policies/'))
        self.templateEnv = jinja2.Environment(loader=self.templateLoader)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def export_policy(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Scheduler_Name': '',
                         'Days': '',
                         'Backup_Hour': '',
                         'Backup_Minute': '',
                         'Concurrent_Capacity': '',
                         'Export_Name': '',
                         'Format': '',
                         'Start_Now': '',
                         'Snapshot': '',
                         'Remote_Host': '',
                         'Encryption_Key': ''}
        optional_args = {'Scheduler_Descr': '',
                         'Export_Descr': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.days(row_num, ws, 'Days', templateVars['Days'])
            validating.number_check(row_num, ws, 'Backup_Hour', templateVars['Backup_Hour'], 0, 23)
            validating.number_check(row_num, ws, 'Backup_Minute', templateVars['Backup_Minute'], 0, 59)
            validating.values(row_num, ws, 'Concurrent_Capacity', templateVars['Concurrent_Capacity'], ['unlimited'])
            validating.values(row_num, ws, 'Format', templateVars['Format'], ['json', 'xml'])
            validating.values(row_num, ws, 'Start_Now', templateVars['Start_Now'], ['triggered', 'untriggered'])
            validating.values(row_num, ws, 'Snapshot', templateVars['Snapshot'], ['no', 'yes'])
            validating.sensitive_var(row_num, ws, 'Encryption_Key', templateVars['Encryption_Key'])
            if not templateVars['Scheduler_Descr'] == None:
                validating.description(row_num, ws, 'Scheduler_Descr', templateVars['Scheduler_Descr'])
            if not templateVars['Export_Descr'] == None:
                validating.description(row_num, ws, 'Export_Descr', templateVars['Export_Descr'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        if re.search(r'\.', templateVars['Remote_Host']):
            templateVars['Remote_Host_'] = templateVars['Remote_Host'].replace('.', '-')
        else:
            templateVars['Remote_Host_'] = templateVars['Remote_Host'].replace(':', '-')

        if not templateVars['Encryption_Key'] == None:
            x = templateVars['Encryption_Key'].split('r')
            key_number = x[1]
            templateVars['sensitive_var'] = 'Encryption_Key%s' % (key_number)

        # Define the Template Source
        template_file = "global_key.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Global_Key.tf'
        dest_dir = 'Admin'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "variables.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Create Variables File for the Sensitive Variables
        dest_file = 'variable_%s.tf' % (templateVars['sensitive_var'])
        dest_dir = 'Admin'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        process_sensitive_var(wb, ws, row_num, dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "export_policy.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Configuration_Export_Policy_%s.tf' % (templateVars['Scheduler_Name'])
        dest_dir = 'Admin'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def login_domain(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Login_Domain': '',
                         'Realm_Type': ''}
        optional_args = {'Description': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_complexity(row_num, ws, 'Login_Domain', templateVars['Login_Domain'])
            validating.values(row_num, ws, 'Realm_Type', templateVars['Realm_Type'], ['RADIUS', 'TACACS'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "Login_Domain_%s.jinja2" % (templateVars['Realm_Type'])
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Login_Domain_%s_%s.tf' % (templateVars['Realm_Type'], templateVars['Login_Domain'])
        dest_dir = 'Admin'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def maint_group(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'MG_Name': '',
                         'Admin_State': '',
                         'Admin_Notify': '',
                         'Graceful': '',
                         'Ignore_Compatability': '',
                         'Run_Mode': '',
                         'SW_Version': '',
                         'Ver_Check_Override': '',
                         'FW_Type': '',
                         'MG_Type': ''}
        optional_args = { }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'MG_Name', templateVars['MG_Name'])
            validating.sw_version(row_num, ws, 'SW_Version', templateVars['SW_Version'])
            validating.values(row_num, ws, 'Admin_State', templateVars['Admin_State'], ['triggered', 'untriggered'])
            validating.values(row_num, ws, 'Admin_Notify', templateVars['Admin_Notify'], ['notifyAlwaysBetweenSets', 'notifyNever', 'notifyOnlyOnFailures'])
            validating.values(row_num, ws, 'Graceful', templateVars['Graceful'], ['no', 'yes'])
            validating.values(row_num, ws, 'Ignore_Compatability',templateVars['Ignore_Compatability'], ['no', 'yes'])
            validating.values(row_num, ws, 'Run_Mode', templateVars['Run_Mode'], ['pauseAlwaysBetweenSets', 'pauseNever', 'pauseOnlyOnFailures'])
            validating.values(row_num, ws, 'Ver_Check_Override', templateVars['Ver_Check_Override'], ['trigger', 'trigger-immediate', 'triggered', 'untriggered'])
            validating.values(row_num, ws, 'MG_Type', templateVars['MG_Type'], ['ALL', 'range'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "maintenance_group.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Maintenance_Group_%s.tf' % (templateVars['MG_Name'])
        dest_dir = 'Admin'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def radius(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'RADIUS_Server': '',
                         'Port': '',
                         'RADIUS_Secret': '',
                         'Authz_Proto': '',
                         'Timeout': '',
                         'Retry_Interval': '',
                         'Mgmt_EPG': '',
                         'Login_Domain': '',
                         'Domain_Order': ''}
        optional_args = {'Description': '',
                         'Domain_Descr': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.ip_address(row_num, ws, 'RADIUS_Server', templateVars['RADIUS_Server'])
            validating.name_complexity(row_num, ws, 'Login_Domain', templateVars['Login_Domain'])
            validating.number_check(row_num, ws, 'Domain_Order', templateVars['Domain_Order'], 0, 17)
            validating.number_check(row_num, ws, 'Port', templateVars['Port'], 1, 65535)
            validating.number_check(row_num, ws, 'Retry_Interval', templateVars['Retry_Interval'], 1, 5)
            validating.sensitive_var(row_num, ws, 'RADIUS_Secret', templateVars['RADIUS_Secret'])
            validating.timeout(row_num, ws, 'Timeout', templateVars['Timeout'])
            validating.values(row_num, ws, 'Authz_Proto', templateVars['Authz_Proto'], ['chap', 'mschap', 'pap'])
            templateVars['Mgmt_EPG'] = validating.mgmt_epg(row_num, ws, 'Mgmt_EPG', templateVars['Mgmt_EPG'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
            if not templateVars['Domain_Descr'] == None:
                validating.description(row_num, ws, 'Domain_Descr', templateVars['Domain_Descr'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        if re.search(r'\.', templateVars['RADIUS_Server']):
            templateVars['RADIUS_Server_'] = templateVars['RADIUS_Server'].replace('.', '-')
        else:
            templateVars['RADIUS_Server_'] = templateVars['RADIUS_Server'].replace(':', '-')

        if not templateVars['RADIUS_Secret'] == None:
            x = templateVars['RADIUS_Secret'].split('r')
            key_number = x[1]
            templateVars['sensitive_var'] = 'RADIUS_Secret%s' % (key_number)

        # Define the Template Source
        template_file = "radius.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'RADIUS_Provider_%s.tf' % (templateVars['RADIUS_Server_'])
        dest_dir = 'Admin'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "variables.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Create Variables File for the Sensitive Variables
        dest_file = 'variable_%s.tf' % (templateVars['sensitive_var'])
        dest_dir = 'Admin'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)
        process_sensitive_var(wb, ws, row_num, dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def realm(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Auth_Realm': '',
                         'Domain_Type': ''}
        optional_args = {'Login_Domain': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.login_type(row_num, ws, 'Auth_Realm', templateVars['Auth_Realm'], 'Domain_Type', templateVars['Domain_Type'])
            if not templateVars['Domain_Type'] == 'local':
                validating.name_complexity(row_num, ws, 'Login_Domain', templateVars['Login_Domain'])
            validating.values(row_num, ws, 'Auth_Realm', templateVars['Auth_Realm'], ['console', 'default'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        if templateVars['Auth_Realm'] == 'console':
            templateVars['child_class'] = 'aaaConsoleAuth'
        elif templateVars['Auth_Realm'] == 'default':
            templateVars['child_class'] = 'aaaDefaultAuth'

        # Define the Template Source
        template_file = "realm.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        if templateVars['Auth_Realm'] == 'console':
            dest_file = 'REALM_Console.tf'
        else:
            dest_file = 'REALM_default.tf'
        dest_dir = 'Admin'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def remote_host(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Remote_Host': '',
                         'Mgmt_EPG': '',
                         'Protocol': '',
                         'Remote_Path': '',
                         'Port': '',
                         'Auth_Type': '',
                         'Pwd_or_SSHPhrase': ''}
        optional_args = {'Description': '',
                         'Backup_User': '',
                         'SSH_Key': '',
                         'Description': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])

            if re.match(r'\:', templateVars['Remote_Host']):
                validating.ip_address(row_num, ws, 'Remote_Host', templateVars['Remote_Host'])
            elif re.match('[a-z]', templateVars['Remote_Host'], re.IGNORECASE):
                validating.dns_name(row_num, ws, 'Remote_Host', templateVars['Remote_Host'])
            else:
                validating.ip_address(row_num, ws, 'Remote_Host', templateVars['Remote_Host'])

            validating.sensitive_var(row_num, ws, 'Pwd_or_SSHPhrase', templateVars['Pwd_or_SSHPhrase'])
            if templateVars['Auth_Type'] == 'password':
                validating.sensitive_var(row_num, ws, 'Backup_User', templateVars['Backup_User'])
            else:
                validating.sensitive_var(row_num, ws, 'SSH_Key', templateVars['SSH_Key'])

            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])

            validating.number_check(row_num, ws, 'Port', templateVars['Port'], 1, 65535)
            validating.values(row_num, ws, 'Auth_Type', templateVars['Auth_Type'], ['password', 'ssh-key'])
            validating.values(row_num, ws, 'Protocol', templateVars['Protocol'], ['ftp', 'scp', 'sftp'])
            templateVars['Mgmt_EPG'] = validating.mgmt_epg(row_num, ws, 'Mgmt_EPG', templateVars['Mgmt_EPG'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        if re.search(':', templateVars['Remote_Host']):
            templateVars['Remote_Host_'] = templateVars['Remote_Host'].replace(':', '-')
        else:
            templateVars['Remote_Host_'] = templateVars['Remote_Host'].replace('.', '-')
        if templateVars['Auth_Type'] == 'password':
            templateVars['Auth_Type'] = 'usePassword'
        elif templateVars['Auth_Type'] == 'ssh-key':
            templateVars['Auth_Type'] = 'useSshKeyContents'

        if not templateVars['Pwd_or_SSHPhrase'] == None:
            x = templateVars['Pwd_or_SSHPhrase'].split('r')
            key_number = x[1]
            templateVars['sensitive_var1'] = 'Pwd_or_SSHPhrase%s' % (key_number)

        if templateVars['Auth_Type'] == 'usePassword':
            if not templateVars['Backup_User'] == None:
                x = templateVars['Backup_User'].split('r')
                key_number = x[1]
                templateVars['sensitive_var2'] = 'Backup_User%s' % (key_number)
        else:
            if not templateVars['SSH_Key'] == None:
                x = templateVars['SSH_Key'].split('r')
                key_number = x[1]
                templateVars['sensitive_var2'] = 'SSH_Key%s' % (key_number)

        # Define the Template Source
        template_file = "remote_host.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Remote_Location_%s.tf' % (templateVars['Remote_Host_'])
        dest_dir = 'Admin'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "variables.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Create Variables File for the Sensitive Variables
        templateVars['sensitive_var'] = templateVars['sensitive_var1']
        dest_file = 'variable_%s.tf' % (templateVars['sensitive_var1'])
        dest_dir = 'Admin'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        process_sensitive_var(wb, ws, row_num, dest_dir, dest_file, template, **templateVars)

        templateVars['sensitive_var'] = templateVars['sensitive_var2']
        dest_file = 'variable_%s.tf' % (templateVars['sensitive_var2'])
        dest_dir = 'Admin'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        process_sensitive_var(wb, ws, row_num, dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def security(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Passwd_Strength': '',
                         'Enforce_Intv': '',
                         'Expiration_Warn': '',
                         'Passwd_Intv': '',
                         'Number_Allowed': '',
                         'Passwd_Store': '',
                         'Lockout': '',
                         'Failed_Attempts': '',
                         'Time_Period': '',
                         'Dur_Lockout': '',
                         'Token_Timeout': '',
                         'Maximum_Valid': '',
                         'Web_Timeout': ''}
        optional_args = { }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.number_check(row_num, ws, 'Expiration_Warn', templateVars['Expiration_Warn'], 0, 30)
            validating.number_check(row_num, ws, 'Passwd_Intv', templateVars['Passwd_Intv'], 0, 745)
            validating.number_check(row_num, ws, 'Number_Allowed', templateVars['Number_Allowed'], 0, 10)
            validating.number_check(row_num, ws, 'Passwd_Store', templateVars['Passwd_Store'], 0, 15)
            validating.number_check(row_num, ws, 'Failed_Attempts', templateVars['Failed_Attempts'], 1, 15)
            validating.number_check(row_num, ws, 'Time_Period', templateVars['Time_Period'], 1, 720)
            validating.number_check(row_num, ws, 'Dur_Lockout', templateVars['Dur_Lockout'], 1, 1440)
            validating.number_check(row_num, ws, 'Token_Timeout', templateVars['Token_Timeout'], 300, 9600)
            validating.number_check(row_num, ws, 'Maximum_Valid', templateVars['Maximum_Valid'], 0, 24)
            validating.number_check(row_num, ws, 'Web_Timeout', templateVars['Web_Timeout'], 60, 65525)
            validating.values(row_num, ws, 'Enforce_Intv', templateVars['Enforce_Intv'], ['disable', 'enable'])
            validating.values(row_num, ws, 'Lockout', templateVars['Lockout'], ['disable', 'enable'])
            validating.values(row_num, ws, 'Passwd_Strength', templateVars['Passwd_Strength'], ['no', 'yes'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "security.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Global_Security.tf'
        dest_dir = 'Admin'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def tacacs(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'TACACS_Server': '',
                         'Port': '',
                         'TACACS_Secret': '',
                         'Auth_Proto': '',
                         'Timeout': '',
                         'Retry_Interval': '',
                         'Mgmt_EPG': '',
                         'Login_Domain': '',
                         'Domain_Order': '',
                         'Acct_DestGrp_Name': ''}
        optional_args = {'Description': '',
                         'Domain_Descr': '',
                         'Login_Domain_Descr': ''}

        # Temporarily Move the Provider Description
        tacacs_descr = kwargs['Description']

        ws_admin = wb['Admin']
        rows = ws_admin.max_row
        row_bundle = ''
        func = 'login_domain'
        count = countKeys(ws_admin, func)
        var_dict = findVars(ws_admin, func, rows, count)
        for pos in var_dict:
            if var_dict[pos].get('Name') == kwargs.get('Policy_Group'):
                row_bundle = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}
                break

        kwargs['Login_Domain_Descr'] = kwargs['Description']
        kwargs['Description'] = tacacs_descr

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.ip_address(row_num, ws, 'TACACS_Server', templateVars['TACACS_Server'])
            validating.name_complexity(row_num, ws, 'Login_Domain', templateVars['Login_Domain'])
            validating.number_check(row_num, ws, 'Domain_Order', templateVars['Domain_Order'], 0, 17)
            validating.number_check(row_num, ws, 'Port', templateVars['Port'], 1, 65535)
            validating.number_check(row_num, ws, 'Retry_Interval', templateVars['Retry_Interval'], 1, 5)
            validating.sensitive_var(row_num, ws, 'TACACS_Secret', templateVars['TACACS_Secret'])
            validating.timeout(row_num, ws, 'Timeout', templateVars['Timeout'])
            validating.values(row_num, ws, 'Auth_Proto', templateVars['Auth_Proto'], ['chap', 'mschap', 'pap'])
            templateVars['Mgmt_EPG'] = validating.mgmt_epg(row_num, ws, 'Mgmt_EPG', templateVars['Mgmt_EPG'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
            if not templateVars['Domain_Descr'] == None:
                validating.description(row_num, ws, 'Domain_Descr', templateVars['Domain_Descr'])
            if not templateVars['Login_Domain_Descr'] == None:
                validating.description(row_num, ws, 'Login_Domain_Descr', templateVars['Login_Domain_Descr'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        if re.search(r'\.', templateVars['TACACS_Server']):
            templateVars['TACACS_Server_'] = templateVars['TACACS_Server'].replace('.', '-')
        else:
            templateVars['TACACS_Server_'] = templateVars['TACACS_Server'].replace(':', '-')

        if not templateVars['TACACS_Secret'] == None:
            x = templateVars['TACACS_Secret'].split('r')
            key_number = x[1]
            templateVars['sensitive_var'] = 'TACACS_Secret%s' % (key_number)

        # Define the Template Source
        template_file = "tacacs.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'TACACS_Provider_%s.tf' % (templateVars['TACACS_Server_'])
        dest_dir = 'Admin'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "variables.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Create Variables File for the Sensitive Variables
        dest_file = 'variable_%s.tf' % (templateVars['sensitive_var'])
        dest_dir = 'Admin'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)
        process_sensitive_var(wb, ws, row_num, dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def tacacs_acct(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Acct_DestGrp_Name': '',
                         'Acct_SrcGrp_Name': ''}
        optional_args = {'Description': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Acct_DestGrp_Name', templateVars['Acct_DestGrp_Name'])
            validating.name_rule(row_num, ws, 'Acct_SrcGrp_Name', templateVars['Acct_SrcGrp_Name'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "tacacs_accounting.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'TACACS_Accounting_DestGrp_%s.tf' % (templateVars['Acct_DestGrp_Name'])
        dest_dir = 'Admin'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

# Terraform ACI Provider - Admin Policies
# Class must be instantiated with Variables
class Best_Practices(object):
    def __init__(self, ws):
        self.templateLoader = jinja2.FileSystemLoader(
            searchpath=(aci_template_path + 'Best_Practices/'))
        self.templateEnv = jinja2.Environment(loader=self.templateLoader)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def bgp_asn(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'BGP_ASN': ''}
        optional_args = {}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.number_check(row_num, ws, 'BGP_ASN', templateVars['BGP_ASN'], 1, 4294967295)
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "bgp_asn.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'bgp_asn_%s.tf' % (templateVars['BGP_ASN'])
        dest_dir = 'System'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def bgp_rr(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Node_ID': ''}
        optional_args = {}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.number_check(row_num, ws, 'Node_ID', templateVars['Node_ID'], 101, 4001)
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "bgp_rr.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'bgp_rr_%s.tf' % (templateVars['Node_ID'])
        dest_dir = 'System'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def ep_controls(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'EP_Loop_State': '',
                         'EP_Loop_Interval': '',
                         'EP_Loop_Multiplier': '',
                         'BD_Learn_Disable': '',
                         'Port_Disable': '',
                         'Rogue_State': '',
                         'Rogue_Interval': '',
                         'Rogue_Multiplier': '',
                         'Hold_Interval': '',
                         'IP_Aging_State': ''}
        optional_args = { }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.number_check(row_num, ws, 'EP_Loop_Interval', templateVars['EP_Loop_Interval'], 30, 300)
            validating.number_check(row_num, ws, 'EP_Loop_Multiplier', templateVars['EP_Loop_Multiplier'], 1, 255)
            validating.number_check(row_num, ws, 'Hold_Interval', templateVars['Hold_Interval'], 1800, 3600)
            validating.number_check(row_num, ws, 'Rogue_Interval', templateVars['Rogue_Interval'], 0, 65535)
            validating.number_check(row_num, ws, 'Rogue_Multiplier', templateVars['Rogue_Multiplier'], 2, 10)
            validating.values(row_num, ws, 'BD_Learn_Disable', templateVars['BD_Learn_Disable'], ['no', 'yes'])
            validating.values(row_num, ws, 'EP_Loop_State', templateVars['EP_Loop_State'], ['disabled', 'enabled'])
            validating.values(row_num, ws, 'IP_Aging_State', templateVars['IP_Aging_State'], ['disabled', 'enabled'])
            validating.values(row_num, ws, 'Port_Disable', templateVars['Port_Disable'], ['no', 'yes'])
            validating.values(row_num, ws, 'Rogue_State', templateVars['Rogue_State'], ['disabled', 'enabled'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        #Combine the Control Elements
        templateVars['action'] = ''
        action_count = 0
        if templateVars['BD_Learn_Disable'] == 'yes':
            templateVars['action'] = 'bd-learn-disable'
            action_count =+ 1
        if templateVars['Port_Disable'] == 'yes':
            if action_count == 0:
                templateVars['action'] = 'port-disable'
                scope_count =+ 1
            else:
                templateVars['action'] = 'bd-learn-disable,port-disable'

        # Define the Template Source
        template_file = "ep_controls.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'ep_controls.tf'
        dest_dir = 'System'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def error_recovery(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Recovery_Interval': '',
                         'EP_Move': '',
                         'BPDU_Guard': '',
                         'MCP_Loop': ''}
        optional_args = { }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.number_check(row_num, ws, 'Recovery_Interval', templateVars['Recovery_Interval'], 30, 65535)
            validating.values(row_num, ws, 'EP_Move', templateVars['EP_Move'], ['no', 'yes'])
            validating.values(row_num, ws, 'BPDU_Guard', templateVars['BPDU_Guard'], ['no', 'yes'])
            validating.values(row_num, ws, 'MCP_Loop', templateVars['MCP_Loop'], ['no', 'yes'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "error_recovery.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'error_recovery.tf'
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def fabric_settings(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'MGMT_Pref': '',
                         'Type': '',
                         'Enable_DOM': '',
                         'Feature_Selection': '',
                         'BFD_ISIS_Policy': '',
                         'Preserve_CoS': ''}
        optional_args = {'Description': '',
                         'Alias': '',
                         'L3_Description': '',}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.number_check(row_num, ws, 'Enable_DOM', templateVars['Enable_DOM'], 0, 1)
            validating.values(row_num, ws, 'BFD_ISIS_Policy', templateVars['BFD_ISIS_Policy'], ['disabled', 'enabled'])
            validating.values(row_num, ws, 'MGMT_Pref', templateVars['MGMT_Pref'], ['inband', 'ooband'])
            validating.values(row_num, ws, 'Preserve_CoS', templateVars['Preserve_CoS'], ['no', 'yes'])
            validating.values(row_num, ws, 'Type', templateVars['Type'], ['compatible', 'strict'])
            validating.values(row_num, ws, 'Feature_Selection', templateVars['Feature_Selection'], ['analytics', 'netflow', 'telemetry'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
            if not templateVars['L3_Description'] == None:
                validating.description(row_num, ws, 'L3_Description', templateVars['L3_Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Convert the Preserve_CoS value to APIC Format
        if templateVars['Preserve_CoS'] == 'yes':
            templateVars['Preserve_CoS'] = 'dot1p-preserve'
        else:
            templateVars['Preserve_CoS'] = None

        # Define the Template Source
        template_file = "apic_preference.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'apic_preference.tf'
        dest_dir = 'System'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "coop_policy.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'coop_policy.tf'
        dest_dir = 'System'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "l3_interface.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'l3_interface.tf'
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "node_control.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'node_control.tf'
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "preserve_cos.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'preserve_cos.tf'
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def fabric_wide(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Disable_Remote_EP_Learn': '',
                         'Enforce_Subnet': '',
                         'EPG_VLAN_Validate': '',
                         'Domain_Validation': '',
                         'Opflex_Auth': '',
                         'Reallocate_Gipo': '',
                         'Restrict_Infra_VLAN': '',
                         'Tracking_State': '',
                         'Delay_Timer': '',
                         'Min_Links': '',
                         'APIC_Ports': ''}
        optional_args = { }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.number_check(row_num, ws, 'Delay_Timer', templateVars['Delay_Timer'], 1, 300)
            validating.number_check(row_num, ws, 'Min_Links', templateVars['Min_Links'], 0, 12)
            validating.values(row_num, ws, 'Disable_Remote_EP_Learn', templateVars['Disable_Remote_EP_Learn'], ['no', 'yes'])
            validating.values(row_num, ws, 'Enforce_Subnet', templateVars['Enforce_Subnet'], ['no', 'yes'])
            validating.values(row_num, ws, 'EPG_VLAN_Validate', templateVars['EPG_VLAN_Validate'], ['no', 'yes'])
            validating.values(row_num, ws, 'Domain_Validation', templateVars['Domain_Validation'], ['no', 'yes'])
            validating.values(row_num, ws, 'Opflex_Auth', templateVars['Opflex_Auth'], ['no', 'yes'])
            validating.values(row_num, ws, 'Reallocate_Gipo', templateVars['Reallocate_Gipo'], ['no', 'yes'])
            validating.values(row_num, ws, 'Restrict_Infra_VLAN', templateVars['Restrict_Infra_VLAN'], ['no', 'yes'])
            validating.values(row_num, ws, 'Tracking_State', templateVars['Tracking_State'], ['on', 'off'])
            validating.values(row_num, ws, 'APIC_Ports', templateVars['APIC_Ports'], ['no', 'yes'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "fabric_wide.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'fabric_wide.tf'
        dest_dir = 'System'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "port_tracking.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'port_tracking.tf'
        dest_dir = 'System'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def isis_policy(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'ISIS_MTU': '',
                         'ISIS_Metric': '',
                         'LSP_Flood_Mode': '',
                         'LSP_Initial_Interval': '',
                         'LSP_Max_Interval': '',
                         'LSP_Second_Interval': '',
                         'SPF_Initial_Interval': '',
                         'SPF_Max_Interval': '',
                         'SPF_Second_Interval': ''}
        optional_args = { }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.number_check(row_num, ws, 'ISIS_MTU', templateVars['ISIS_MTU'], 128, 4352)
            validating.number_check(row_num, ws, 'ISIS_Metric', templateVars['ISIS_Metric'], 1, 63)
            validating.number_check(row_num, ws, 'LSP_Initial_Interval', templateVars['LSP_Initial_Interval'], 50, 120000)
            validating.number_check(row_num, ws, 'LSP_Max_Interval', templateVars['LSP_Max_Interval'], 50, 120000)
            validating.number_check(row_num, ws, 'LSP_Second_Interval', templateVars['LSP_Second_Interval'], 50, 120000)
            validating.number_check(row_num, ws, 'SPF_Initial_Interval', templateVars['SPF_Initial_Interval'], 50, 120000)
            validating.number_check(row_num, ws, 'SPF_Max_Interval', templateVars['SPF_Max_Interval'], 50, 120000)
            validating.number_check(row_num, ws, 'SPF_Second_Interval', templateVars['SPF_Second_Interval'], 50, 120000)
            validating.values(row_num, ws, 'LSP_Flood_Mode', templateVars['LSP_Flood_Mode'], ['disabled', 'enabled'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "isis_policy.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'isis_policy.tf'
        dest_dir = 'System'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def mcp_policy(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Admin_State': '',
                         'Control': '',
                         'MCP_Key': '',
                         'Detect_Multiplier': '',
                         'Loop_Action': '',
                         'Initial_Delay': '',
                         'Frequency_Seconds': '',
                         'Frequency_msec': ''}
        optional_args = {'Description': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.number_check(row_num, ws, 'Detect_Multiplier', templateVars['Detect_Multiplier'], 1, 255)
            validating.number_check(row_num, ws, 'Initial_Delay', templateVars['Initial_Delay'], 0, 1800)
            validating.number_check(row_num, ws, 'Frequency_Seconds', templateVars['Frequency_Seconds'], 0, 300)
            validating.number_check(row_num, ws, 'Frequency_msec', templateVars['Frequency_msec'], 0, 999)
            validating.values(row_num, ws, 'Admin_State', templateVars['Admin_State'], ['disabled', 'enabled'])
            validating.values(row_num, ws, 'Control', templateVars['Control'], ['no', 'yes'])
            validating.values(row_num, ws, 'Loop_Action', templateVars['Loop_Action'], ['no', 'yes'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        #Modify the Control Value
        if templateVars['Control'] == 'yes':
            templateVars['Control'] = 'pdu-per-vlan'
        else:
            templateVars['Control'] = None

        # Convert the Loop_Action value to APIC Format
        if templateVars['Loop_Action'] == 'yes':
            templateVars['Loop_Action'] = 'port-disable'
        else:
            templateVars['Loop_Action'] = None

        if not templateVars['MCP_Key'] == None:
            x = templateVars['MCP_Key'].split('r')
            key_number = x[1]
            templateVars['sensitive_var'] = 'MCP_Key%s' % (key_number)

        # Define the Template Source
        template_file = "mcp_policy.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'mcp_policy.tf'
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # site_dict = ast.literal_eval(os.environ[Site_ID])

        # Define the Template Source
        template_file = "variables.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'variable_%s.tf' % (templateVars['sensitive_var'])
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        process_sensitive_var(wb, ws, row_num, dest_dir, dest_file, template, **templateVars)

# Terraform ACI Provider - Fabric Policies
# Class must be instantiated with Variables
class Fabric_Policies(object):
    def __init__(self, ws):
        self.templateLoader = jinja2.FileSystemLoader(
            searchpath=(aci_template_path + 'Fabric_Policies/'))
        self.templateEnv = jinja2.Environment(loader=self.templateLoader)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def date_time(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Admin_State': '',
                         'Server_State': '',
                         'Master_Mode': '',
                         'Auth_State': ''}
        optional_args = {'Description': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.values(row_num, ws, 'Admin_State', templateVars['Admin_State'], ['disabled', 'enabled'])
            validating.values(row_num, ws, 'Server_State', templateVars['Server_State'], ['disabled', 'enabled'])
            validating.values(row_num, ws, 'Master_Mode', templateVars['Master_Mode'], ['disabled', 'enabled'])
            validating.values(row_num, ws, 'Auth_State', templateVars['Auth_State'], ['disabled', 'enabled'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        if templateVars['Server_State'] == 'disabled':
            templateVars['Master_Mode'] = 'disabled'

        # Define the Template Source
        template_file = "date_time_profile.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Date_and_Time_Profile_%s.tf' % (templateVars['Name'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def dns(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'DNS_Profile': '',
                         'DNS_Server': '',
                         'Preferred': ''}
        optional_args = {}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.ip_address(row_num, ws, 'DNS_Server', templateVars['DNS_Server'])
            validating.values(row_num, ws, 'Preferred', templateVars['Preferred'], ['no', 'yes'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        if re.search(r'\.', templateVars['DNS_Server']):
            templateVars['DNS_Server_'] = templateVars['DNS_Server'].replace('.', '-')
        else:
            templateVars['DNS_Server_'] = templateVars['DNS_Server'].replace(':', '-')

        # Define the Template Source
        template_file = "dns.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'DNS_Profile_%s.tf' % (templateVars['DNS_Profile'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def dns_profile(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Mgmt_EPG': ''}
        optional_args = {'Description': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            templateVars['Mgmt_EPG'] = validating.mgmt_epg(row_num, ws, 'Mgmt_EPG', templateVars['Mgmt_EPG'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "dns_profile.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'DNS_Profile_%s.tf' % (templateVars['Name'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def domain(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'DNS_Profile': '',
                         'Default_Domain': '',
                         'Domain': ''}
        optional_args = {}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.domain(row_num, ws, 'Domain', templateVars['Domain'])
            validating.values(row_num, ws, 'Default_Domain', templateVars['Default_Domain'], ['no', 'yes'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        templateVars['Domain_'] = templateVars['Domain'].replace('.', '-')

        # Define the Template Source
        template_file = "domain.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'DNS_Profile_%s.tf' % (templateVars['DNS_Profile'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def pod_policy(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Policy_Group': '',
                         'Pod_Profile': '',
                         'Date_Time_Policy': '',
                         'ISIS_Policy': '',
                         'COOP_Group_Policy': '',
                         'BGP_RR_Policy': '',
                         'Mgmt_Access_Policy': '',
                         'SNMP_Policy': '',
                         'MACsec_Policy': ''}
        optional_args = {'PG_Description': '',
                         'Profile_Descr': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Policy_Group', templateVars['Policy_Group'])
            validating.name_rule(row_num, ws, 'Pod_Profile', templateVars['Pod_Profile'])
            if not templateVars['PG_Description'] == None:
                validating.description(row_num, ws, 'PG_Description', templateVars['PG_Description'])
            if not templateVars['Profile_Descr'] == None:
                validating.description(row_num, ws, 'Profile_Descr', templateVars['Profile_Descr'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        templateVars['ISIS_Policy'] = 'default'
        templateVars['COOP_Group_Policy'] = 'default'
        templateVars['BGP_RR_Policy'] = 'default'

        # Define the Template Source
        template_file = "pod_profile.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Pod_Profile_%s.tf' % (templateVars['Pod_Profile'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def ntp(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Date_Policy': '',
                         'NTP_Server': '',
                         'Preferred': '',
                         'Min_Poll': '',
                         'Max_Poll': '',
                         'Mgmt_EPG': ''}
        optional_args = {'Description': '',
                         'Key_ID': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Date_Policy', templateVars['Date_Policy'])
            if ':' in templateVars['NTP_Server']:
                validating.ip_address(row_num, ws, 'NTP_Server', templateVars['NTP_Server'])
            elif re.search('[a-z]', templateVars['NTP_Server'], re.IGNORECASE):
                validating.dns_name(row_num, ws, 'NTP_Server', templateVars['NTP_Server'])
            else:
                validating.ip_address(row_num, ws, 'NTP_Server', templateVars['NTP_Server'])
            validating.number_check(row_num, ws, 'Min_Poll', templateVars['Min_Poll'], 4, 16)
            validating.number_check(row_num, ws, 'Max_Poll', templateVars['Max_Poll'], 4, 16)
            validating.values(row_num, ws, 'Preferred', templateVars['Preferred'], ['no', 'yes'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
            if not templateVars['Key_ID'] == None:
                validating.number_check(row_num, ws, 'Key_ID', templateVars['Key_ID'], 1, 65535)
            templateVars['Mgmt_EPG'] = validating.mgmt_epg(row_num, ws, 'Mgmt_EPG', templateVars['Mgmt_EPG'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        if re.search(r'\.', templateVars['NTP_Server']):
            templateVars['NTP_Server_'] = templateVars['NTP_Server'].replace('.', '-')
        else:
            templateVars['NTP_Server_'] = templateVars['NTP_Server'].replace(':', '-')

        # Define the Template Source
        template_file = "ntp.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Date_and_Time_Profile_%s.tf' % (templateVars['Date_Policy'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def ntp_key(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Date_Policy': '',
                         'Key_ID': '',
                         'NTP_Key': '',
                         'Key_Type': ''}
        optional_args = {}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Date_Policy', templateVars['Date_Policy'])
            validating.number_check(row_num, ws, 'Key_ID', templateVars['Key_ID'], 1, 65535)
            validating.values(row_num, ws, 'Key_Type', templateVars['Key_Type'], ['md5', 'sha1'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        x = templateVars['NTP_Key'].split('r')
        key_number = x[1]
        templateVars['sensitive_var'] = 'NTP_Key%s' % (key_number)

        # Define the Template Source
        template_file = "variables.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'variable_%s.tf' % (templateVars['sensitive_var'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        process_sensitive_var(wb, ws, row_num, dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "ntp_key.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Date_and_Time_Profile_%s.tf' % (templateVars['Date_Policy'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def sch_dstgrp(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'DestGrp_Name': '',
                         'Admin_State': '',
                         'SMTP_Port': '',
                         'SMTP_Relay': '',
                         'Mgmt_EPG': '',
                         'From_Email': '',
                         'Reply_Email': '',
                         'To_Email': '',
                         'Contract_Id': '',
                         'Customer_Id': '',
                         'Site_Id': ''}
        optional_args = {'Description': '',
                         'Phone_Number': '',
                         'Contact_Info': '',
                         'Street_Address': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            if re.match(r'\:', templateVars['SMTP_Relay']):
                validating.ip_address(row_num, ws, 'SMTP_Relay', templateVars['SMTP_Relay'])
            elif re.match(r'[a-zA-Z]', templateVars['SMTP_Relay']):
                validating.dns_name(row_num, ws, 'SMTP_Relay', templateVars['SMTP_Relay'])
            else:
                validating.ip_address(row_num, ws, 'SMTP_Relay', templateVars['SMTP_Relay'])
            validating.email(row_num, ws, 'From_Email', templateVars['From_Email'])
            validating.email(row_num, ws, 'Reply_Email', templateVars['Reply_Email'])
            validating.email(row_num, ws, 'To_Email', templateVars['To_Email'])
            validating.name_rule(row_num, ws, 'DestGrp_Name', templateVars['DestGrp_Name'])
            validating.values(row_num, ws, 'Admin_State', templateVars['Admin_State'], ['disabled', 'enabled'])
            templateVars['Mgmt_EPG'] = validating.mgmt_epg(row_num, ws, 'Mgmt_EPG', templateVars['Mgmt_EPG'])
            if not templateVars['Contact_Info'] == None:
                validating.description(row_num, ws, 'Contact_Info', templateVars['Contact_Info'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
            if not templateVars['Phone_Number'] == None:
                validating.phone(row_num, ws, 'Phone_Number', templateVars['Phone_Number'])
            if not templateVars['Street_Address'] == None:
                validating.description(row_num, ws, 'Street_Address', templateVars['Street_Address'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "smartcallhome_dg.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Smart_Callhome_%s.tf' % (templateVars['DestGrp_Name'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def sch_receiver(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'DestGrp_Name': '',
                         'Receiver_Name': '',
                         'Admin_State': '',
                         'Email': '',
                         'Format': '',
                         'RFC_Compliant': '',
                         'Audit': '',
                         'Events': '',
                         'Faults': '',
                         'Session': ''}
        optional_args = { }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.email(row_num, ws, 'Email', templateVars['Email'])
            validating.name_rule(row_num, ws, 'DestGrp_Name', templateVars['DestGrp_Name'])
            validating.name_rule(row_num, ws, 'Receiver_Name', templateVars['Receiver_Name'])
            validating.values(row_num, ws, 'Admin_State', templateVars['Admin_State'], ['disabled', 'enabled'])
            validating.values(row_num, ws, 'RFC_Compliant', templateVars['RFC_Compliant'], ['no', 'yes'])
            validating.values(row_num, ws, 'Audit', templateVars['Audit'], ['no', 'yes'])
            validating.values(row_num, ws, 'Events', templateVars['Events'], ['no', 'yes'])
            validating.values(row_num, ws, 'Faults', templateVars['Faults'], ['no', 'yes'])
            validating.values(row_num, ws, 'Session', templateVars['Session'], ['no', 'yes'])
            validating.values(row_num, ws, 'Format', templateVars['Format'], ['aml', 'short-txt', 'xml'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        incl_list = ''
        if not templateVars['Audit'] == 'no':
            incl_list = 'audit'
        if not templateVars['Events'] == 'no':
            if incl_list == '':
                incl_list = 'events'
            else:
                incl_list = incl_list + ',events'
        if not templateVars['Faults'] == 'no':
            if incl_list == '':
                incl_list = 'faults'
            else:
                incl_list = incl_list + ',faults'
        if not templateVars['Session'] == 'no':
            if incl_list == '':
                incl_list = 'session'
            else:
                incl_list = incl_list + ',session'
        templateVars['Included_Types'] = incl_list

        # Define the Template Source
        template_file = "smartcallhome_receiver.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Smart_Callhome_%s.tf' % (templateVars['DestGrp_Name'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "smartcallhome_source.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Smart_Callhome_%s.tf' % (templateVars['DestGrp_Name'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def snmp_client(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'SNMP_Policy': '',
                         'Client_Group': '',
                         'SNMP_Client': '',
                         'SNMP_Client_Name': ''}
        optional_args = { }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.ip_address(row_num, ws, 'SNMP_Client', templateVars['SNMP_Client'])
            validating.name_rule(row_num, ws, 'SNMP_Client_Name', templateVars['SNMP_Client_Name'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        if re.search(r'\.', templateVars['SNMP_Client']):
            templateVars['SNMP_Client_'] = templateVars['SNMP_Client'].replace('.', '-')
        else:
            templateVars['SNMP_Client_'] = templateVars['SNMP_Client'].replace(':', '-')

        # Define the Template Source
        template_file = "snmp_client.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'SNMP_Policy_%s.tf' % (templateVars['SNMP_Policy'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def snmp_comm(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'SNMP_Policy': '',
                         'SNMP_Community': ''}
        optional_args = {'Description': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.snmp_string(row_num, ws, 'SNMP_Community', templateVars['SNMP_Community'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        if not templateVars['SNMP_Community'] == None:
            x = templateVars['SNMP_Community'].split('r')
            key_number = x[1]
            templateVars['sensitive_var'] = 'SNMP_Community%s' % (key_number)

        # Define the Template Source
        template_file = "snmp_comm.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'SNMP_Policy_%s.tf' % (templateVars['SNMP_Policy'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

        # site_dict = ast.literal_eval(os.environ[Site_ID])

        # Define the Template Source
        template_file = "variables.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'variable_%s.tf' % (templateVars['sensitive_var'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        process_sensitive_var(wb, ws, row_num, dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def snmp_clgrp(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'SNMP_Policy': '',
                         'Client_Group': '',
                         'Mgmt_EPG': ''}
        optional_args = {'Description': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Client_Group', templateVars['Client_Group'])
            validating.name_rule(row_num, ws, 'SNMP_Policy', templateVars['SNMP_Policy'])
            templateVars['Mgmt_EPG'] = validating.mgmt_epg(row_num, ws, 'Mgmt_EPG', templateVars['Mgmt_EPG'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "snmp_client_group.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'SNMP_Policy_%s.tf' % (templateVars['SNMP_Policy'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def snmp_policy(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'SNMP_Policy': '',
                         'Admin_State': ''}
        optional_args = {'Description': '',
                         'SNMP_Contact': '',
                         'SNMP_Location': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'SNMP_Policy', templateVars['SNMP_Policy'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
            if not templateVars['SNMP_Contact'] == None:
                validating.description(row_num, ws, 'SNMP_Contact', templateVars['SNMP_Contact'])
            if not templateVars['SNMP_Location'] == None:
                validating.description(row_num, ws, 'SNMP_Location', templateVars['SNMP_Location'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "snmp_policy.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'SNMP_Policy_%s.tf' % (templateVars['SNMP_Policy'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def snmp_trap(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'SNMP_Policy': '',
                         'SNMP_Trap_DG': '',
                         'Trap_Server': '',
                         'Destination_Port': '',
                         'Version': '',
                         'Community_or_Username': '',
                         'Security_Level': '',
                         'Mgmt_EPG': ''}
        optional_args = { }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        # Set noauth if v1 or v2c
        if re.search('(v1|v2c)', templateVars['Version']):
            templateVars['Security_Level'] = 'noauth'

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.ip_address(row_num, ws, 'Trap_Server', templateVars['Trap_Server'])
            validating.number_check(row_num, ws, 'Destination_Port', templateVars['Destination_Port'], 1, 65535)
            validating.values(row_num, ws, 'Version', templateVars['Version'], ['v1', 'v2c', 'v3'])
            validating.values(row_num, ws, 'Security_Level', templateVars['Security_Level'], ['auth', 'noauth', 'priv'])
            validating.snmp_string(row_num, ws, 'Community_or_Username', templateVars['Community_or_Username'])
            templateVars['Mgmt_EPG'] = validating.mgmt_epg(row_num, ws, 'Mgmt_EPG', templateVars['Mgmt_EPG'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        if re.search(r'\.', templateVars['Trap_Server']):
            templateVars['Trap_Server_'] = templateVars['Trap_Server'].replace('.', '-')
        else:
            templateVars['Trap_Server_'] = templateVars['Trap_Server'].replace(':', '-')

        if not templateVars['Community_or_Username'] == None:
            x = templateVars['Community_or_Username'].split('r')
            key_number = x[1]
            templateVars['sensitive_var'] = 'Community_or_Username%s' % (key_number)

        # Define the Template Source
        template_file = "snmp_trap.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'SNMP_Policy_%s.tf' % (templateVars['SNMP_Policy'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "snmp_trap_destgrp_reciever.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'SNMP_Trap_DestGrp_%s.tf' % (templateVars['SNMP_Trap_DG'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

        # site_dict = ast.literal_eval(os.environ[Site_ID])

        # Define the Template Source
        template_file = "variables.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'variable_%s.tf' % (templateVars['sensitive_var'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        process_sensitive_var(wb, ws, row_num, dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def snmp_user(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'SNMP_Policy': '',
                         'SNMP_User': '',
                         'Authorization_Type': '',
                         'Authorization_Key': ''}
        optional_args = {'Privacy_Type': '',
                         'Privacy_Key': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            auth_type = templateVars['Authorization_Type']
            auth_key = templateVars['Authorization_Key']
            validating.snmp_auth(row_num, ws, templateVars['Privacy_Type'], templateVars['Privacy_Key'], auth_type, auth_key)
            validating.snmp_string(row_num, ws, 'SNMP_User', templateVars['SNMP_User'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Modify User Input of templateVars['Privacy_Type'] or templateVars['Authorization_Type'] to send to APIC
        if templateVars['Privacy_Type'] == 'none':
            templateVars['Privacy_Type'] = None
        if templateVars['Authorization_Type'] == 'md5':
            templateVars['Authorization_Type'] = None
        if templateVars['Authorization_Type'] == 'sha1':
            templateVars['Authorization_Type'] = 'hmac-sha1-96'

        if not templateVars['Privacy_Key'] == None:
            x = templateVars['Privacy_Key'].split('r')
            key_number = x[1]
            templateVars['sensitive_var1'] = 'Privacy_Key%s' % (key_number)

        if not templateVars['Authorization_Key'] == None:
            x = templateVars['Authorization_Key'].split('r')
            key_number = x[1]
            templateVars['sensitive_var2'] = 'Authorization_Key%s' % (key_number)

        # Define the Template Source
        template_file = "snmp_user.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'SNMP_Policy_%s.tf' % (templateVars['SNMP_Policy'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

        # site_dict = ast.literal_eval(os.environ[Site_ID])

        # Define the Template Source
        template_file = "variables.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Create Variables File for the Privacy Key & Authorization Key
        if not templateVars['Privacy_Key'] == None:
            dest_file = 'variable_%s.tf' % (templateVars['sensitive_var1'])
            dest_dir = 'Fabric'
            templateVars['sensitive_var'] = templateVars['sensitive_var1']
            process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

            process_sensitive_var(wb, ws, row_num, dest_dir, dest_file, template, **templateVars)

        if not templateVars['Authorization_Key'] == None:
            dest_file = 'variable_%s.tf' % (templateVars['sensitive_var2'])
            dest_dir = 'Fabric'
            templateVars['sensitive_var'] = templateVars['sensitive_var2']
            process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

            process_sensitive_var(wb, ws, row_num, dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def syslog_dg(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Dest_Grp_Name': '',
                         'Minimum_Level': '',
                         'Log_Format': '',
                         'Console': '',
                         'Console_Level': '',
                         'Local': '',
                         'Local_Level': '',
                         'Include_msec': '',
                         'Include_timezone': '',
                         'Audit': '',
                         'Events': '',
                         'Faults': '',
                         'Session': ''}
        optional_args = { }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.log_level(row_num, ws, 'Minimum_Level', templateVars['Minimum_Level'])
            validating.log_level(row_num, ws, 'Local_Level', templateVars['Local_Level'])
            validating.log_level(row_num, ws, 'Console_Level', templateVars['Console_Level'])
            validating.name_rule(row_num, ws, 'Dest_Grp_Name', templateVars['Dest_Grp_Name'])
            validating.values(row_num, ws, 'Console', templateVars['Console'], ['disabled', 'enabled'])
            validating.values(row_num, ws, 'Local', templateVars['Local'], ['disabled', 'enabled'])
            validating.values(row_num, ws, 'Include_msec', templateVars['Include_msec'], ['no', 'yes'])
            validating.values(row_num, ws, 'Include_timezone', templateVars['Include_timezone'], ['no', 'yes'])
            validating.values(row_num, ws, 'Audit', templateVars['Audit'], ['no', 'yes'])
            validating.values(row_num, ws, 'Events', templateVars['Events'], ['no', 'yes'])
            validating.values(row_num, ws, 'Faults', templateVars['Faults'], ['no', 'yes'])
            validating.values(row_num, ws, 'Session', templateVars['Session'], ['no', 'yes'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)
        incl_list = ''
        if not templateVars['Audit'] == 'no':
            incl_list = 'audit'
        if not templateVars['Events'] == 'no':
            if incl_list == '':
                incl_list = 'events'
            else:
                incl_list = incl_list + ',events'
        if not templateVars['Faults'] == 'no':
            if incl_list == '':
                incl_list = 'faults'
            else:
                incl_list = incl_list + ',faults'
        if not templateVars['Session'] == 'no':
            if incl_list == '':
                incl_list = 'session'
            else:
                incl_list = incl_list + ',session'
        templateVars['Included_Types'] = incl_list

        # Define the Template Source
        template_file = "syslog_dg.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Syslog_DestGrp_%s.tf' % (templateVars['Dest_Grp_Name'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def syslog_rmt(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Dest_Grp_Name': '',
                         'Syslog_Server': '',
                         'Port': '',
                         'Mgmt_EPG': '',
                         'Severity': '',
                         'Facility': ''}
        optional_args = { }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.ip_address(row_num, ws, 'Syslog_Server', templateVars['Syslog_Server'])
            validating.log_level(row_num, ws, 'Severity', templateVars['Severity'])
            validating.number_check(row_num, ws, 'Port', templateVars['Port'], 1, 65535)
            validating.syslog_fac(row_num, ws, 'Facility', templateVars['Facility'])
            templateVars['Mgmt_EPG'] = validating.mgmt_epg(row_num, ws, 'Mgmt_EPG', templateVars['Mgmt_EPG'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        if re.search(r'\.', templateVars['Syslog_Server']):
            templateVars['Syslog_Server_'] = templateVars['Syslog_Server'].replace('.', '-')
        else:
            templateVars['Syslog_Server_'] = templateVars['Syslog_Server'].replace(':', '-')

        # Define the Template Source
        template_file = "syslog_rmt.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Syslog_DestGrp_%s.tf' % (templateVars['Dest_Grp_Name'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def trap_groups(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'SNMP_Trap_DG': '',
                         'SNMP_Source': '',
                         'Audit': '',
                         'Events': '',
                         'Faults': '',
                         'Session': ''}
        optional_args = {'Description': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'SNMP_Trap_DG', templateVars['SNMP_Trap_DG'])
            validating.name_rule(row_num, ws, 'SNMP_Source', templateVars['SNMP_Source'])
            validating.values(row_num, ws, 'Audit', templateVars['Audit'], ['no', 'yes'])
            validating.values(row_num, ws, 'Events', templateVars['Events'], ['no', 'yes'])
            validating.values(row_num, ws, 'Faults', templateVars['Faults'], ['no', 'yes'])
            validating.values(row_num, ws, 'Session', templateVars['Session'], ['no', 'yes'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        incl_list = ''
        if not templateVars['Audit'] == 'no':
            incl_list = 'audit'
        if not templateVars['Events'] == 'no':
            if incl_list == '':
                incl_list = 'events'
            else:
                incl_list = incl_list + ',events'
        if not templateVars['Faults'] == 'no':
            if incl_list == '':
                incl_list = 'faults'
            else:
                incl_list = incl_list + ',faults'
        if not templateVars['Session'] == 'no':
            if incl_list == '':
                incl_list = 'session'
            else:
                incl_list = incl_list + ',session'
        templateVars['Included_Types'] = incl_list

        # Define the Template Source
        template_file = "snmp_trap_destgrp.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'SNMP_Trap_DestGrp_%s.tf' % (templateVars['SNMP_Trap_DG'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "snmp_trap_source.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'SNMP_Trap_DestGrp_%s.tf' % (templateVars['SNMP_Trap_DG'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

# Terraform ACI Provider - Site Policies
# Class must be instantiated with Variables
class Site_Policies(object):
    def __init__(self, ws):
        self.templateLoader = jinja2.FileSystemLoader(
            searchpath=(aci_template_path))
        self.templateEnv = jinja2.Environment(loader=self.templateLoader)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def site_id(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_ID': '',
                         'Site_Name': '',
                         'APIC_URL': '',
                         'APIC_Version': '',
                         'APIC_Auth_Type': '',
                         'Terraform_EQ': '',
                         'Terraform_Version': '',
                         'Provider_EQ': '',
                         'Provider_Version': '',
                         'Run_Location': '',
                         'State_Location': ''}
        optional_args = {'Terraform_Cloud_Org': '',
                         'Workspace_Prefix': '',
                         'VCS_Base_Repo': '',
                         'Terraform_Agent_Pool_ID': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Variables
            validating.name_complexity(row_num, ws, 'Site_Name', templateVars['Site_Name'])
            apic_url = 'https://%s' % (templateVars['APIC_URL'])
            validating.url(row_num, ws, 'APIC_URL', apic_url)
            validating.values(row_num, ws, 'APIC_Version', templateVars['APIC_Version'], ['3.X', '4.X', '5.X'])
            validating.values(row_num, ws, 'APIC_Auth_Type', templateVars['APIC_Auth_Type'], ['ssh-key', 'user_pass'])
            validating.values(row_num, ws, 'Provider_EQ', templateVars['Provider_EQ'], ['>=', '=', '<='])
            validating.values(row_num, ws, 'State_Location', templateVars['State_Location'], ['Local', 'Terraform_Cloud'])
            validating.values(row_num, ws, 'Terraform_EQ', templateVars['Terraform_EQ'], ['>=', '=', '<='])
            validating.values(row_num, ws, 'Run_Location', templateVars['Run_Location'], ['Local', 'Terraform_Cloud'])
            validating.not_empty(row_num, ws, 'Provider_Version', templateVars['Provider_Version'])
            validating.not_empty(row_num, ws, 'Terraform_Version', templateVars['Terraform_Version'])
            if templateVars['State_Location'] == 'Terraform_Cloud':
                validating.not_empty(row_num, ws, 'Terraform_Cloud_Org', templateVars['Terraform_Cloud_Org'])
                validating.not_empty(row_num, ws, 'VCS_Base_Repo', templateVars['VCS_Base_Repo'])
                validating.not_empty(row_num, ws, 'Terraform_Agent_Pool_ID', templateVars['Terraform_Agent_Pool_ID'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Save the Site Information into Environment Variables
        Site_ID = 'Site_ID_%s' % (templateVars['Site_ID'])
        os.environ[Site_ID] = '%s' % (templateVars)

        folder_list = ['Access', 'Admin', 'Fabric', 'System', 'VLANs']
        file_list = ['.gitignore_.gitignore', 'main.jinja2_main.tf', 'variables.jinja2_variables.tf']

        # Write the .gitignore to the Appropriate Directories
        for folder in folder_list:
            for file in file_list:
                x = file.split('_')
                template_file = x[0]
                dest_file = x[1]
                template = self.templateEnv.get_template(template_file)
                create_tf_file('w', folder, dest_file, template, **templateVars)

        # If the State_Location is Terraform_Cloud Configure Workspaces in the Cloud
        if templateVars['State_Location'] == 'Terraform_Cloud':
            # Initialize the Class
            lib_tf_ref = 'lib_terraform.Terraform_Cloud'
            class_init = '%s()' % (lib_tf_ref)

            # Get terraform_cloud_token
            kwargs['terraform_cloud_token'] = eval("%s.%s()" % (class_init, 'terraform_token'))

            # Get terraform_cloud_token
            kwargs['terraform_oath_token'] = eval("%s.%s(**kwargs)" % (class_init, 'oath_token'))

            # Get workspace_ids
            workspace_dict = {}
            for folder in folder_list:
                workspace_dict = tf_workspace(class_init, folder, workspace_dict, **kwargs)

        # If the Run_Location is Terraform_Cloud Configure Variables in the Cloud
        if templateVars['Run_Location'] == 'Terraform_Cloud':
            if templateVars['APIC_Auth_Type'] == 'user_pass':
                var_list = ['aciUrl', 'aciUser', 'aciPass']
            else:
                var_list = ['aciUrl', 'aciCertName', 'aciPrivateKey']

            # Get var_ids
            tf_var_dict = {}
            for folder in folder_list:
                folder_id = 'Site_ID_%s_%s' % (templateVars['Site_ID'], folder)
                kwargs['workspace_id'] = workspace_dict[folder_id]
                kwargs['Description'] = ''
                for var in var_list:
                    tf_var_dict = tf_variables(class_init, folder, var, tf_var_dict, **kwargs)

        site_wb = '%s_intf_selectors.xlsx' % (templateVars['Site_Name'])
        if not os.path.isfile(site_wb):
            wb.save(filename=site_wb)
            wb_wr = load_workbook(site_wb)
            ws_wr = wb_wr.get_sheet_names()
            for sheetName in ws_wr:
                if sheetName not in ['Sites']:
                    sheetToDelete = wb_wr.get_sheet_by_name(sheetName)
                    wb_wr.remove_sheet(sheetToDelete)
            wb_wr.save(filename=site_wb)

    # Method must be called with the following kwargs.
    # Group: Required.  A Group Name to represent a list of Site_ID's
    # Site_1: Required.  The Site_ID for the First Site
    # Site_2: Required.  The Site_ID for the Second Site
    # Site_[3-12]: Optional.  The Site_ID for the 3rd thru the 12th Site(s)
    def group_id(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Group': '',
                         'Site_1': ''}
        optional_args = {'Site_2': '',
                         'Site_3': '',
                         'Site_4': '',
                         'Site_5': '',
                         'Site_6': '',
                         'Site_7': '',
                         'Site_8': '',
                         'Site_9': '',
                         'Site_10': '',
                         'Site_11': '',
                         'Site_12': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        for x in range(1, 13):
            site = 'Site_%s' % (x)
            if not templateVars[site] == None:
                validating.site_group(wb, ws, site, templateVars[site])

        grp_count = 0
        for x in list(map(chr, range(ord('A'), ord('F')+1))):
            grp = 'Grp_%s' % (x)
            if templateVars['Group'] == grp:
                grp_count += 1
        if grp_count == 0:
            print(f'\n-----------------------------------------------------------------------------\n')
            print(f'   Error on Worksheet {ws.title}, Row {row_num} Group, Group_Name "{templateVars["Group"]}" is invalid.')
            print(f'   A valid Group Name is Grp_A thru Grp_F.  Exiting....')
            print(f'\n-----------------------------------------------------------------------------\n')
            exit()

        # Save the Site Information into Environment Variables
        Group_ID = '%s' % (templateVars['Group'])
        os.environ[Group_ID] = '%s' % (templateVars)

# Terraform ACI Provider - Tenants Policies
# Class must be instantiated with Variables
class Tenant_Policies(object):
    def __init__(self, ws):
        self.templateLoader = jinja2.FileSystemLoader(
            searchpath=(aci_template_path + 'Tenant_Policies/'))
        self.templateEnv = jinja2.Environment(loader=self.templateLoader)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def add_app(self, wb, ws, row_num, **kwargs):
        # Open the Network Policies Worksheet
        ws_net = wb['Network Policies']
        rows = ws_net.max_row

        # Dicts for Application Profile; required and optional args
        required_args = {'Site_Group': '',
                         'Tenant': '',
                         'App_Profile': '',
                         'App_Policy': '',
                         'Policy_Name': '',
                         'prio': '',
                         'monEPGPol': ''}
        optional_args = {'App_Alias': '',
                         'App_Description': '',
                         'App_Tags': ''}

        # Get the Application Profile Policies from the Network Policies Tab
        func = 'app'
        count = countKeys(ws_net, func)
        row_app = ''
        var_dict = findVars(ws_net, func, rows, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('App_Policy'):
                row_app = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}
                break

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
            validating.name_rule(row_num, ws, 'App_Profile', templateVars['App_Profile'])
            validating.qos_priority(row_app, ws_net, 'prio', templateVars['prio'])
            if not templateVars['App_Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['App_Description'] == None:
                validating.description(row_num, ws, 'App_Description', templateVars['App_Description'])
            if not templateVars['App_Tags'] == None:
                if re.match(',', templateVars['Tags']):
                    for tag in templateVars['Tags'].split(','):
                        validating.name_rule(row_num, ws, 'Tags', tag)
                else:
                    validating.name_rule(row_num, ws, 'Tags', templateVars['Tags'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        if templateVars['monEPGPol'] == 'default':
            templateVars['monEPGPol'] = 'uni/tn-common/monepg-default'

        if re.search('^(common|mgmt|infra)$', templateVars['Tenant']):
            templateVars['Tenant_Dn'] = 'data.aci_tenant.%s' % (templateVars['Tenant'])
        else:
            templateVars['Tenant_Dn'] = 'aci_tenant.%s' % (templateVars['Tenant'])

        # Define the Template Source
        template_file = "app.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'App_Profile_%s.tf' % (templateVars['App_Profile'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def add_bd(self, wb, ws, row_num, **kwargs):
        # Assign the kwargs to a initial var for each process
        initial_kwargs = kwargs

        # Open the Network Policies Worksheet
        ws_net = wb['Network Policies']
        rows = ws_net.max_row

        # Dicts for Bridge Domain required and optional args
        required_args = {'Site_Group': '',
                         'Tenant': '',
                         'Bridge_Domain': '',
                         'BD_Policy': '',
                         'Policy_Name': '',
                         # BD Policy Required Args
                         'bd_type': '',
                         'host_routing': '',
                         'ep_clear': '',
                         'unk_mac': '',
                         'unk_mcast': '',
                         'v6unk_mcast': '',
                         'multi_dst': '',
                         'mcast_allow': '',
                         'ipv6_mcast': '',
                         'arp_flood': '',
                         'limit_learn': '',
                         'fvEpRetPol': '',
                         'unicast_route': '',
                         'intersight_l2': '',
                         'intersight_bum': '',
                         'optimize_wan': '',
                         'monEPGPol': '',
                         'ip_dp_learning': ''}
        optional_args = {'Alias': '',
                         'Description': '',
                         'Tags': '',
                         'Custom_MAC': '',
                         'Link_Local_IPv6': '',
                         'VRF_Tenant': '',
                         'VRF': '',
                         'Subnet': '',
                         'Subnet_Description': '',
                         'Subnet_Policy': '',
                         'L3Out_Tenant': '',
                         'L3Out': '',
                         # BD Policy Optional Args
                         'dhcpRelayP': '',
                         'igmpIfPol': '',
                         'igmpSnoopPol': '',
                         'mldSnoopPol': '',
                         'ep_move': '',
                         'rtctrlProfile': '',
                         'ndIfPol': '',
                         'fhsBDPol': '',
                         'netflowMonitorPol': ''}

        # Get the BD Policies from the Network Policies Tab
        func = 'bd'
        count = countKeys(ws_net, func)
        row_bd = ''
        var_dict = findVars(ws_net, func, rows, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('BD_Policy'):
                row_bd = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}
                break

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
            validating.name_rule(row_num, ws, 'Bridge_Domain', templateVars['Bridge_Domain'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
            if not templateVars['Tags'] == None:
                if re.match(',', templateVars['Tags']):
                    for tag in templateVars['Tags'].split(','):
                        validating.name_rule(row_num, ws, 'Tags', tag)
                else:
                    validating.name_rule(row_num, ws, 'Tags', templateVars['Tags'])
            if not templateVars['Custom_MAC'] == None:
                validating.mac_address(row_num, ws, 'Custom_MAC', templateVars['Custom_MAC'])
            if not templateVars['Link_Local_IPv6'] == None:
                validating.ip_address(row_num, ws, 'Link_Local_IPv6', templateVars['Link_Local_IPv6'])
            if not templateVars['BD_Policy'] == templateVars['Policy_Name']:
                validating.error_policy_names(row_num, ws, templateVars['BD_Policy'], templateVars['Policy_Name'])
            if not templateVars['VRF'] == None:
                validating.name_rule(row_num, ws, 'VRF_Tenant', templateVars['VRF_Tenant'])
                validating.name_rule(row_num, ws, 'VRF', templateVars['VRF'])
            if not templateVars['Subnet'] == None:
                validating.ip_address(row_num, ws, 'Subnet', templateVars['Subnet'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
            if not templateVars['L3Out'] == None:
                validating.name_rule(row_num, ws, 'L3Out_Tenant', templateVars['L3Out_Tenant'])
                validating.name_rule(row_num, ws, 'L3Out', templateVars['L3Out'])
            validating.values(row_bd, ws_net, 'bd_type', templateVars['bd_type'], ['fc', 'regular'])
            validating.values(row_bd, ws_net, 'ep_clear', templateVars['ep_clear'], ['no', 'yes'])
            validating.values(row_bd, ws_net, 'host_routing', templateVars['host_routing'], ['no', 'yes'])
            validating.values(row_bd, ws_net, 'mcast_allow', templateVars['mcast_allow'], ['no', 'yes'])
            validating.values(row_bd, ws_net, 'ipv6_mcast', templateVars['ipv6_mcast'], ['no', 'yes'])
            validating.values(row_bd, ws_net, 'arp_flood', templateVars['arp_flood'], ['no', 'yes'])
            validating.values(row_bd, ws_net, 'limit_learn', templateVars['limit_learn'], ['no', 'yes'])
            validating.values(row_bd, ws_net, 'unicast_route', templateVars['unicast_route'], ['no', 'yes'])
            validating.values(row_bd, ws_net, 'limit_learn', templateVars['limit_learn'], ['no', 'yes'])
            validating.values(row_bd, ws_net, 'intersight_l2', templateVars['intersight_l2'], ['no', 'yes'])
            validating.values(row_bd, ws_net, 'intersight_bum', templateVars['intersight_bum'], ['no', 'yes'])
            validating.values(row_bd, ws_net, 'optimize_wan', templateVars['optimize_wan'], ['no', 'yes'])
            validating.values(row_bd, ws_net, 'ip_dp_learning', templateVars['ip_dp_learning'], ['no', 'yes'])
            validating.values(row_bd, ws_net, 'unk_mac', templateVars['unk_mac'], ['flood', 'proxy'])
            validating.values(row_bd, ws_net, 'unk_mcast', templateVars['unk_mcast'], ['flood', 'opt-flood'])
            validating.values(row_bd, ws_net, 'v6unk_mcast', templateVars['v6unk_mcast'], ['flood', 'opt-flood'])
            validating.values(row_bd, ws_net, 'multi_dst', templateVars['multi_dst'], ['bd-flood', 'drop', 'encap-flood'])
            if not templateVars['ep_move'] == None:
                validating.values(row_bd, ws_net, 'ep_move', templateVars['ep_move'], ['garp'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        if templateVars['dhcpRelayP'] == 'default':
            templateVars['dhcpRelayP'] = 'uni/tn-common/relayp-default'
        if templateVars['fhsBDPol'] == 'default':
            templateVars['fhsBDPol'] = 'uni/tn-common/bdpol-default'
        if templateVars['fvEpRetPol'] == 'default':
            templateVars['fvEpRetPol'] = 'uni/tn-common/epRPol-default'
        if templateVars['igmpIfPol'] == 'default':
            templateVars['igmpIfPol'] = 'uni/tn-common/igmpIfPol-default'
        if templateVars['igmpSnoopPol'] == 'default':
            templateVars['igmpSnoopPol'] = 'uni/tn-common/snPol-default'
        if templateVars['mldSnoopPol'] == 'default':
            templateVars['mldSnoopPol'] = 'uni/tn-common/mldsnoopPol-default'
        if templateVars['monEPGPol'] == 'default':
            templateVars['monEPGPol'] = 'uni/tn-common/monepg-default'
        if templateVars['ndIfPol'] == 'default':
            templateVars['ndIfPol'] = 'uni/tn-common/ndifpol-default'
        if templateVars['netflowMonitorPol'] == 'default':
            templateVars['netflowMonitorPol'] = 'uni/tn-common/monitorpol-default'

        # Define the Template Source
        template_file = "bd.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Bridge_Domain_%s.tf' % (templateVars['Bridge_Domain'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

       # Reset kwargs back to initial kwargs
        kwargs = initial_kwargs

        # Initialize the Class
        lib_aci_ref = 'Tenant_Policies'
        class_init = '%s(ws)' % (lib_aci_ref)

        # Create the Subnet if it Exists
        if not kwargs.get('Subnet') == None:
            eval("%s.%s(wb, ws, row_num, **kwargs)" % (class_init, 'add_subnet'))

        if not templateVars['Tenant'] == templateVars['VRF_Tenant']:
            templateVars['bd_Tenant'] = templateVars['Tenant']

            # Process the template through the Sites
            templateVars['Tenant'] = templateVars['VRF_Tenant']
            template_file = "data_tenant.jinja2"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = 'data_Tenant_%s.tf' % (templateVars['VRF_Tenant'])
            dest_dir = 'Tenant_%s' % (templateVars['bd_Tenant'])
            process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

            # Define the Template Source
            template_file = "data_vrf.jinja2"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = 'data_Tenant_%s_VRF_%s.tf' % (templateVars['VRF_Tenant'], templateVars['VRF'])
            dest_dir = 'Tenant_%s' % (templateVars['bd_Tenant'])
            process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

            templateVars['Tenant'] = templateVars['bd_Tenant']

        if not templateVars['L3Out'] == None:
            if not templateVars['Tenant'] == templateVars['L3Out_Tenant']:
                templateVars['bd_Tenant'] = templateVars['Tenant']

                # Process the template through the Sites
                templateVars['Tenant'] = templateVars['L3Out_Tenant']
                template_file = "data_tenant.jinja2"
                template = self.templateEnv.get_template(template_file)

                # Process the template through the Sites
                dest_file = 'data_Tenant_%s.tf' % (templateVars['L3Out_Tenant'])
                dest_dir = 'Tenant_%s' % (templateVars['bd_Tenant'])
                process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

                # Process the template through the Sites
                template_file = "data_l3out.jinja2"
                template = self.templateEnv.get_template(template_file)

                # Process the template through the Sites
                dest_file = 'data_Tenant_%s_L3Out_%s.tf' % (templateVars['L3Out_Tenant'], templateVars['L3Out'])
                dest_dir = 'Tenant_%s' % (templateVars['bd_Tenant'])
                process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def add_epg(self, wb, ws, row_num, **kwargs):
        # Assign the kwargs to a initial var for each process
        initial_kwargs = kwargs

        # Initialize the Class
        lib_aci_ref = 'Tenant_Policies'
        class_init = '%s(ws)' % (lib_aci_ref)

        # Create the Application Profile if it Exists
        if not kwargs.get('App_Profile') == None:
            eval("%s.%s(wb, ws, row_num, **kwargs)" % (class_init, 'add_app'))

        # Reset kwargs back to initial kwargs
        kwargs = initial_kwargs

        # Open the Network Policies Worksheet
        ws_net = wb['Network Policies']
        rows = ws_net.max_row

        # Dicts for Bridge Domain required and optional args
        required_args = {'Site_Group': '',
                         'Tenant': '',
                         'App_Profile': '',
                         'EPG': '',
                         'BD_Tenant': '',
                         'Bridge_Domain': '',
                         'EPG_Policy': '',
                         'Policy_Name': '',
                         'is_attr_based': '',
                         'prio': '',
                         'pc_enf_pref': '',
                         'fwd_ctrl': '',
                         'pref_gr_memb': '',
                         'flood': '',
                         'match_t': '',
                         'monEPGPol': '',
                         'shutdown': '',
                         'has_mcast': ''}
        optional_args = {'Alias': '',
                         'Description': '',
                         'Tags': '',
                         'Physical_Domains': '',
                         'VMM_Domains': '',
                         'VLAN': '',
                         'PVLAN': '',
                         'EPG_to_AAEP': '',
                         'Master_fvEPg': '',
                         'vzCPIf': '',
                         'vzCtrctEPgCont': '',
                         'vzTaboo': '',
                         'exception_tag': '',
                         'qosCustomPol': '',
                         'qosDppPol': '',
                         'intra_vzBrCP': '',
                         'fhsTrustCtrlPol': '',
                         'vzGraphCont': '',
                         'FC_Domain': '',}

        # Get the EPG Policies from the Network Policies Tab
        func = 'epg'
        count = countKeys(ws_net, func)
        row_epg = ''
        var_dict = findVars(ws_net, func, rows, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('EPG_Policy'):
                row_epg = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
            validating.name_rule(row_num, ws, 'App_Profile', templateVars['App_Profile'])
            validating.name_rule(row_num, ws, 'EPG', templateVars['EPG'])
            validating.name_rule(row_num, ws, 'BD_Tenant', templateVars['BD_Tenant'])
            validating.name_rule(row_num, ws, 'Bridge_Domain', templateVars['Bridge_Domain'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
            if not templateVars['Tags'] == None:
                if re.match(',', templateVars['Tags']):
                    for tag in templateVars['Tags'].split(','):
                        validating.name_rule(row_num, ws, 'Tags', tag)
                else:
                    validating.name_rule(row_num, ws, 'Tags', templateVars['Tags'])
            if not templateVars['Physical_Domains'] == None:
                if re.match(',', templateVars['Physical_Domains']):
                    for phys in templateVars['Physical_Domains'].split(','):
                        validating.name_rule(row_num, ws, 'Physical_Domains', phys)
                else:
                    validating.name_rule(row_num, ws, 'Physical_Domains', templateVars['Physical_Domains'])
            if not templateVars['VMM_Domains'] == None:
                if re.match(',', templateVars['VMM_Domains']):
                    for phys in templateVars['VMM_Domains'].split(','):
                        validating.name_rule(row_num, ws, 'VMM_Domains', phys)
                else:
                    validating.name_rule(row_num, ws, 'VMM_Domains', templateVars['VMM_Domains'])
            if not templateVars['VLAN'] == None:
                validating.vlans(row_num, ws, 'VLAN', templateVars['VLAN'])
            if not templateVars['PVLAN'] == None:
                validating.vlans(row_num, ws, 'PVLAN', templateVars['PVLAN'])
            if not templateVars['EPG_to_AAEP'] == None:
                validating.name_rule(row_num, ws, 'EPG_to_AAEP', templateVars['EPG_to_AAEP'])
            validating.match_t(row_epg, ws_net, 'match_t', templateVars['match_t'])
            validating.values(row_epg, ws_net, 'fwd_ctrl', templateVars['fwd_ctrl'], ['none', 'proxy-arp'])
            validating.qos_priority(row_epg, ws_net, 'prio', templateVars['prio'])
            validating.values(row_epg, ws_net, 'flood', templateVars['flood'], ['disabled', 'enabled'])
            validating.values(row_epg, ws_net, 'is_attr_based', templateVars['is_attr_based'], ['no', 'yes'])
            validating.values(row_epg, ws_net, 'pc_enf_pref', templateVars['pc_enf_pref'], ['enforced', 'unenforced'])
            validating.values(row_epg, ws_net, 'pref_gr_memb', templateVars['pref_gr_memb'], ['exclude', 'include'])
            validating.values(row_epg, ws_net, 'shutdown', templateVars['shutdown'], ['no', 'yes'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        if templateVars['vzCPIf'] == 'default':
            templateVars['vzCPIf'] = 'uni/tn-common/cif-default'
        # if templateVars['vzCtrctEPgCont'] == 'default':
        #     templateVars['vzCtrctEPgCont'] = 'uni/tn-common/mldsnoopPol-default'
        if templateVars['vzTaboo'] == 'default':
            templateVars['vzTaboo'] = 'uni/tn-common/taboo-default'
        if templateVars['qosCustomPol'] == 'default':
            templateVars['qosCustomPol'] = 'uni/tn-common/qoscustom-default'
        if templateVars['qosDppPol'] == 'default':
            templateVars['qosDppPol'] = 'uni/tn-common/qosdpppol-default'
        if templateVars['intra_vzBrCP'] == 'default':
            templateVars['intra_vzBrCP'] = 'uni/tn-common/brc-default'
        if templateVars['monEPGPol'] == 'default':
            templateVars['monEPGPol'] = 'uni/tn-common/monepg-default'
        if templateVars['fhsTrustCtrlPol'] == 'default':
            templateVars['fhsTrustCtrlPol'] = 'uni/tn-common/trustctrlpol-default'
        if templateVars['fwd_ctrl'] == 'none':
            templateVars['fwd_ctrl'] = None
        # if templateVars['vzGraphCont'] == 'default':
        #     templateVars['vzGraphCont'] = 'uni/tn-common/monitorpol-default'

        # Define the Template Source
        template_file = "epg.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'App_Profile_%s_EPG_%s.tf' % (templateVars['App_Profile'], templateVars['EPG'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        if not templateVars['Physical_Domains'] == None:
            if ',' in templateVars['Physical_Domains']:
                splitx = templateVars['Physical_Domains'].split(',')
                for x in splitx:
                    templateVars['Domain'] = 'phys-%s' % (x)
                    # Define the Template Source
                    template_file = "domain_to_epg.jinja2"
                    template = self.templateEnv.get_template(template_file)

                    # Process the template through the Sites
                    dest_file = 'App_Profile_%s_EPG_%s.tf' % (templateVars['App_Profile'], templateVars['EPG'])
                    dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
                    process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)
            else:
                templateVars['Domain'] = 'phys-%s' % (templateVars['Physical_Domains'])
                # Define the Template Source
                template_file = "domain_to_epg.jinja2"
                template = self.templateEnv.get_template(template_file)

                # Process the template through the Sites
                dest_file = 'App_Profile_%s_EPG_%s.tf' % (templateVars['App_Profile'], templateVars['EPG'])
                dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
                process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

        if not templateVars['VMM_Domains'] == None:
            if ',' in templateVars['VMM_Domains']:
                splitx = templateVars['VMM_Domains'].split(',')
                for x in splitx:
                    templateVars['Domain'] = 'vmm-%s' % (x)
                    # Define the Template Source
                    template_file = "domain_to_epg.jinja2"
                    template = self.templateEnv.get_template(template_file)

                    # Process the template through the Sites
                    dest_file = 'App_Profile_%s_EPG_%s.tf' % (templateVars['App_Profile'], templateVars['EPG'])
                    dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
                    process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)
            else:
                templateVars['Domain'] = 'vmm-%s' % (templateVars['VMM_Domains'])
                # Define the Template Source
                template_file = "domain_to_epg.jinja2"
                template = self.templateEnv.get_template(template_file)

                # Process the template through the Sites
                dest_file = 'App_Profile_%s_EPG_%s.tf' % (templateVars['App_Profile'], templateVars['EPG'])
                dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
                process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

        if not templateVars['VLAN'] == None:
            # Define the Template Source
            template_file = "static_path.jinja2"
            template = self.templateEnv.get_template(template_file)

            dest_file = 'App_Profile_%s_EPG_%s.tf' % (templateVars['App_Profile'], templateVars['EPG'])
            dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
            process_workbook(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

        if not templateVars['EPG_to_AAEP'] == None:
            if re.search(',', templateVars['EPG_to_AAEP']):
                # Define the Template Source
                aep_list = templateVars['EPG_to_AAEP'].split(',')
                for aep in aep_list:
                    templateVars['AAEP'] = aep

                    # Define the Template Source
                    template_file = "policies_global_aep_generic.jinja2"
                    template = self.templateEnv.get_template(template_file)

                    # Process the template through the Sites
                    dest_file = 'Policies_Global_AEP_%s_generic.tf' % (templateVars['AAEP'])
                    dest_dir = 'Access'
                    process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

                    # Define the Template Source
                    template_file = "data_access_generic.jinja2"
                    template = self.templateEnv.get_template(template_file)

                    # Process the template through the Sites
                    dest_file = 'data_AEP_%s.tf' % (templateVars['AAEP'])
                    dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
                    process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

                    # Define the Template Source
                    template_file = "epgs_using_function.jinja2"
                    template = self.templateEnv.get_template(template_file)

                    # Process the template through the Sites
                    dest_file = 'App_Profile_%s_EPG_%s.tf' % (templateVars['App_Profile'], templateVars['EPG'])
                    dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
                    process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)
            else:
                templateVars['AAEP'] = templateVars['EPG_to_AAEP']
                # Define the Template Source
                template_file = "policies_global_aep_generic.jinja2"
                template = self.templateEnv.get_template(template_file)

                # Process the template through the Sites
                dest_file = 'Policies_Global_AEP_%s_generic.tf' % (templateVars['AAEP'])
                dest_dir = 'Access'
                process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

                # Define the Template Source
                template_file = "data_access_generic.jinja2"
                template = self.templateEnv.get_template(template_file)

                # Process the template through the Sites
                dest_file = 'data_AEP_%s.tf' % (templateVars['AAEP'])
                dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
                process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

                # Define the Template Source
                template_file = "epgs_using_function.jinja2"
                template = self.templateEnv.get_template(template_file)

                # Process the template through the Sites
                dest_file = 'App_Profile_%s_EPG_%s.tf' % (templateVars['App_Profile'], templateVars['EPG'])
                dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
                process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

        # dest_file = 'epg_%s_%s_static_bindings.tf' % (templateVars['App_Profile'], templateVars['EPG'])
        # dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        # create_file(wb, ws, row_num, 'w', dest_dir, dest_file, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def add_l3out(self, wb, ws, row_num, **kwargs):
        # Open the Network Policies Worksheet
        ws_net = wb['Network Policies']
        rows = ws_net.max_row

        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Tenant': '',
                         'L3Out': '',
                         'VRF_Tenant': '',
                         'VRF': '',
                         'L3_Domain': '',
                         'target_dscp': '',
                         'Run_BGP': '',
                         'export': '',
                         'import': '',
                         'target_dscp': ''}
        optional_args = {'Description': '',
                         'Alias': '',
                         'Tags': '',
                         'EIGRP_Routing_Policy': '',
                         'OSPF_Routing_Policy': '',
                         'leak_rtctrlProfile': '',
                         'damp_rtctrlProfile': '',
                         'fvBDPublicSubnetHolder': ''}


        # Get the L3Out Policies from the Network Policies Tab
        func = 'l3Out'
        count = countKeys(ws_net, func)
        row_l3out = ''
        var_dict = findVars(ws_net, func, rows, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('L3Out_Policy'):
                row_l3out = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.dscp(row_l3out, ws_net, 'target_dscp', templateVars['target_dscp'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
            validating.name_rule(row_num, ws, 'L3Out', templateVars['L3Out'])
            validating.name_rule(row_num, ws, 'VRF', templateVars['VRF'])
            validating.name_rule(row_num, ws, 'VRF_Tenant', templateVars['VRF_Tenant'])
            validating.values(row_num, ws, 'export', templateVars['export'], ['no', 'yes'])
            validating.values(row_num, ws, 'import', templateVars['import'], ['no', 'yes'])
            validating.values(row_num, ws, 'Run_BGP', templateVars['Run_BGP'], ['no', 'yes'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Process the template through the Sites
        template_file = "data_domain_l3_profile.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'data_domain_l3_profile_%s.tf' % (templateVars['L3_Domain'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        ctrl_count = 0
        Ctrl = ''
        if templateVars['export'] == 'yes':
            Ctrl = '"export"'
            ctrl_count =+ 1
        if templateVars['import'] == 'yes' and ctrl_count > 0:
            Ctrl = Ctrl + ',' + '"import"'
            ctrl_count =+ 1
        elif templateVars['import'] == 'yes':
            Ctrl = '"import"'
            ctrl_count =+ 1
        if ctrl_count > 0:
            templateVars['enforce_rtctrl'] = '[%s]' % (Ctrl)
        else:
            templateVars['enforce_rtctrl'] = '["unspecified"]'

        # Define the Template Source
        template_file = "l3out.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'L3Out_%s.tf' % (templateVars['L3Out'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        if templateVars['Run_BGP'] == 'yes':
            # Define the Template Source
            template_file = "bgp_external_policy.jinja2"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = 'L3Out_%s.tf' % (templateVars['L3Out'])
            dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
            process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

        #--------------------------
        # EIGRP Routing Policy
        #--------------------------
        if not templateVars['EIGRP_Routing_Policy'] == None:

            # Dicts for required and optional args
            required_args = {'Site_Group': '',
                            'Tenant': '',
                            'L3Out': '',
                            'Policy_Name': '',
                            'AS_Number': ''}
            optional_args = { }
            # Get the L3Out Policies from the Network Policies Tab
            func = 'eigrp_routing'
            count = countKeys(ws_net, func)
            row_eigrp = ''
            var_dict = findVars(ws_net, func, rows, count)
            for pos in var_dict:
                if var_dict[pos].get('Policy_Name') == kwargs.get('EIGRP_Routing_Policy'):
                    row_eigrp = var_dict[pos]['row']
                    del var_dict[pos]['row']
                    kwargs = {**kwargs, **var_dict[pos]}

            # Validate inputs, return dict of template vars
            templateVars = process_kwargs(required_args, optional_args, **kwargs)

            try:
                # Validate Required Arguments
                validating.number_check(row_eigrp, ws_net, 'AS_Number', templateVars['AS_Number'], 1, 65534)
            except Exception as err:
                Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws_net, row_eigrp)
                raise ErrException(Error_Return)

            # Define the Template Source
            template_file = "l3out_eigrp_external_policy.jinja2"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = 'L3Out_%s.tf' % (templateVars['L3Out'])
            dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
            process_method(wb, ws_net, row_eigrp, 'a+', dest_dir, dest_file, template, **templateVars)

        #--------------------------
        # OSPF Routing Policy
        #--------------------------
        if not templateVars['OSPF_Routing_Policy'] == None:

            # Dicts for required and optional args
            required_args = {'Site_Group': '',
                            'Tenant': '',
                            'L3Out': '',
                            'Policy_Name': '',
                            'Area_ID': '',
                            'Area_Type': '',
                            'Cost': '',
                            'Redistribute_NSSA': '',
                            'Originate_Summary': '',
                            'Suppress_FA': ''}
            optional_args = { }
            # Get the L3Out Policies from the Network Policies Tab
            func = 'ospf_routing'
            count = countKeys(ws_net, func)
            row_ospf = ''
            var_dict = findVars(ws_net, func, rows, count)
            for pos in var_dict:
                if var_dict[pos].get('Policy_Name') == kwargs.get('OSPF_Routing_Policy'):
                    row_ospf = var_dict[pos]['row']
                    del var_dict[pos]['row']
                    kwargs = {**kwargs, **var_dict[pos]}

            # Validate inputs, return dict of template vars
            templateVars = process_kwargs(required_args, optional_args, **kwargs)

            try:
                # Validate Required Arguments
                validating.number_check(row_ospf, ws_net, 'Cost', templateVars['Cost'], 0, 16777215)
                validating.values(row_ospf, ws_net, 'Area_Type', templateVars['Area_Type'], ['nssa', 'regular', 'stub'])
                validating.values(row_ospf, ws_net, 'Redistribute_NSSA', templateVars['Redistribute_NSSA'], ['no', 'yes'])
                validating.values(row_ospf, ws_net, 'Originate_Summary', templateVars['Originate_Summary'], ['no', 'yes'])
                validating.values(row_ospf, ws_net, 'Suppress_FA', templateVars['Suppress_FA'], ['no', 'yes'])
            except Exception as err:
                Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws_net, row_ospf)
                raise ErrException(Error_Return)

            ctrl_count = 0
            Ctrl = ''
            if templateVars['Redistribute_NSSA'] == 'yes':
                Ctrl = 'redistribute'
                ctrl_count =+ 1
            if templateVars['Originate_Summary'] == 'yes' and ctrl_count > 0:
                Ctrl = Ctrl + ',' + 'summary'
                ctrl_count =+ 1
            elif templateVars['Originate_Summary'] == 'yes':
                Ctrl = 'summary'
                ctrl_count =+ 1
            if templateVars['Suppress_FA'] == 'yes' and ctrl_count > 0:
                Ctrl = Ctrl + ',' + 'supress-fa'
                ctrl_count =+ 1
            elif templateVars['Suppress_FA'] == 'yes':
                Ctrl = 'supress-fa'
                ctrl_count =+ 1
            if ctrl_count > 0:
                templateVars['Ctrl'] = '%s' % (Ctrl)
            else:
                templateVars['Ctrl'] = 'unspecified'

            # Define the Template Source
            template_file = "l3out_ospf_external_policy.jinja2"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = 'L3Out_%s.tf' % (templateVars['L3Out'])
            dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
            process_method(wb, ws_net, row_ospf, 'a+', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def add_subnet(self, wb, ws, row_num, **kwargs):
        # Open the Network Policies Worksheet
        ws_net = wb['Network Policies']
        rows = ws_net.max_row

        # Dicts for Subnet required and optional args
        required_args = {'Site_Group': '',
                         'Tenant': '',
                         'Bridge_Domain': '',
                         'Subnet': '',
                         'Subnet_Policy': '',
                         'Policy_Name': '',
                         'virtual': '',
                         'preferred': '',
                         'scope': '',
                         'nd': '',
                         'no-default-gateway': '',
                         'querier': ''}
        optional_args = {'Subnet_Description': '',
                         'l3extOut': '',
                         'rtctrlProfile': '',
                         'ndPfxPol': ''}

        # Get the Subnet Policies from the Network Policies Tab
        func = 'subnet'
        count = countKeys(ws_net, func)
        row_subnet = ''
        var_dict = findVars(ws_net, func, rows, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('Subnet_Policy'):
                row_subnet = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}
                break

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.ip_address(row_num, ws, 'Subnet', templateVars['Subnet'])
            if not templateVars['Subnet_Description'] == None:
                validating.description(row_num, ws, 'Subnet_Description', templateVars['Subnet_Description'])
            validating.values(row_subnet, ws_net, 'virtual', templateVars['virtual'], ['no', 'yes'])
            validating.values(row_subnet, ws_net, 'preferred', templateVars['preferred'], ['no', 'yes'])
            validating.values(row_subnet, ws_net, 'scope', templateVars['scope'], ['private', 'public', 'shared', 'private-shared', 'public-shared'])
            validating.values(row_subnet, ws_net, 'nd', templateVars['nd'], ['no', 'yes'])
            validating.values(row_subnet, ws_net, 'no-default-gateway', templateVars['no-default-gateway'], ['no', 'yes'])
            validating.values(row_subnet, ws_net, 'querier', templateVars['querier'], ['no', 'yes'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        if templateVars['l3extOut'] == 'default':
            templateVars['l3extOut'] = 'uni/tn-common/out-default'
        if templateVars['ndPfxPol'] == 'default':
            templateVars['ndPfxPol'] = 'uni/tn-common/ndpfxpol-default'

        # Create ctrl templateVars
        ctrl_count = 0
        Ctrl = ''
        if templateVars['nd'] == 'yes':
            Ctrl = '"nd"'
            ctrl_count =+ 1
        if templateVars['no-default-gateway'] == 'yes' and ctrl_count > 0:
            Ctrl = Ctrl + ', ' + '"no-default-gateway"'
            ctrl_count =+ 1
        elif templateVars['no-default-gateway'] == 'yes':
            Ctrl = '"no-default-gateway"'
            ctrl_count =+ 1
        if templateVars['querier'] == 'yes' and ctrl_count > 0:
            Ctrl = Ctrl + ', ' + '"querier"'
            ctrl_count =+ 1
        elif templateVars['querier'] == 'yes':
            Ctrl = '"querier"'
            ctrl_count =+ 1

        if ctrl_count > 0:
            templateVars['Ctrl'] = '[%s]' % (Ctrl)
        else:
            templateVars['Ctrl'] = '["unspecified"]'

        # Modify scope templateVars
        if re.search('^(private|public|shared)$', templateVars['scope']):
            templateVars['scope'] = '"%s"' % (templateVars['scope'])
        elif re.search('^(private|public)\\-shared$', templateVars['scope']):
            x = templateVars['scope'].split('-')
            templateVars['scope'] = '"%s", "%s"' % (x[0], x[1])

        # As period and colon are not allowed in description need to modify Subnet to work for description and filename
        if ':' in templateVars['Subnet']:
            network = "%s" % (ipaddress.IPv6Network(templateVars['Subnet'], strict=False))
            templateVars['Subnet_'] = network
            templateVars['Subnet_'] = templateVars['Subnet_'].replace(':', '-')
            templateVars['Subnet_'] = templateVars['Subnet_'].replace('/', '_')
        else:
            network = "%s" % (ipaddress.IPv4Network(templateVars['Subnet'], strict=False))
            templateVars['Subnet_'] = network
            templateVars['Subnet_'] = templateVars['Subnet_'].replace('.', '-')
            templateVars['Subnet_'] = templateVars['Subnet_'].replace('/', '_')

        # Define the Template Source
        template_file = "subnet.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Bridge_Domain_%s.tf' % (templateVars['Bridge_Domain'],)
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def add_tenant(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Tenant': ''}
        optional_args = {'Alias': '',
                         'Description': '',
                         'Tags': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
            if not templateVars['Tags'] == None:
                if re.match(',', templateVars['Tags']):
                    for tag in templateVars['Tags'].split(','):
                        validating.name_rule(row_num, ws, 'Tags', tag)
                else:
                    validating.name_rule(row_num, ws, 'Tags', templateVars['Tags'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "tenant.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Tenant_%s.tf' % (templateVars['Tenant'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        if re.search('Grp_[A-F]', templateVars['Site_Group']):
            Group_ID = '%s' % (templateVars['Site_Group'])
            site_group = ast.literal_eval(os.environ[Group_ID])
            for x in range(1, 13):
                sitex = 'Site_%s' % (x)
                if not site_group[sitex] == None:
                    Site_ID = 'Site_ID_%s' % (site_group[sitex])
                    site_dict = ast.literal_eval(os.environ[Site_ID])

                    # Create kwargs for Site Variables
                    kwargs['Site_ID'] = site_dict.get('Site_ID')
                    kwargs['Site_Name'] = site_dict.get('Site_Name')
                    kwargs['APIC_URL'] = site_dict.get('APIC_URL')
                    kwargs['APIC_Version'] = site_dict.get('APIC_Version')
                    kwargs['APIC_Auth_Type'] = site_dict.get('APIC_Auth_Type')
                    kwargs['Terraform_EQ'] = site_dict.get('Terraform_EQ')
                    kwargs['Terraform_Version'] = site_dict.get('Terraform_Version')
                    kwargs['Provider_EQ'] = site_dict.get('Provider_EQ')
                    kwargs['Provider_Version'] = site_dict.get('Provider_Version')
                    kwargs['Run_Location'] = site_dict.get('Run_Location')
                    kwargs['State_Location'] = site_dict.get('State_Location')
                    kwargs['Terraform_Cloud_Org'] = site_dict.get('Terraform_Cloud_Org')
                    kwargs['Workspace_Prefix'] = site_dict.get('Workspace_Prefix')
                    kwargs['VCS_Base_Repo'] = site_dict.get('VCS_Base_Repo')
                    kwargs['Terraform_Agent_Pool_ID'] = site_dict.get('Terraform_Agent_Pool_ID')

                    # Dicts for required and optional args
                    required_args = {'Site_ID': '',
                                        'Site_Name': '',
                                        'APIC_URL': '',
                                        'APIC_Version': '',
                                        'APIC_Auth_Type': '',
                                        'Terraform_EQ': '',
                                        'Terraform_Version': '',
                                        'Provider_EQ': '',
                                        'Provider_Version': '',
                                        'Run_Location': '',
                                        'State_Location': ''}
                    optional_args = {'Terraform_Cloud_Org': '',
                                        'Workspace_Prefix': '',
                                        'VCS_Base_Repo': '',
                                        'Terraform_Agent_Pool_ID': ''}

                    # Validate inputs, return dict of template vars
                    templateVars = process_kwargs(required_args, optional_args, **kwargs)

                    # If the State_Location is Terraform_Cloud Configure Workspaces in the Cloud
                    if templateVars['State_Location'] == 'Terraform_Cloud':
                        # Initialize the Class
                        lib_tf_ref = 'lib_terraform.Terraform_Cloud'
                        class_init = '%s()' % (lib_tf_ref)

                        # Get terraform_cloud_token
                        kwargs['terraform_cloud_token'] = eval("%s.%s()" % (class_init, 'terraform_token'))

                        # Get terraform_cloud_token
                        kwargs['terraform_oath_token'] = eval("%s.%s(**kwargs)" % (class_init, 'oath_token'))

                        # Get workspace_ids
                        workspace_dict = {}
                        workspace_dict = tf_workspace(class_init, dest_dir, workspace_dict, **kwargs)

                    # If the Run_Location is Terraform_Cloud Configure Variables in the Cloud
                    if templateVars['Run_Location'] == 'Terraform_Cloud':
                        # Set Variable List
                        if templateVars['APIC_Auth_Type'] == 'user_pass':
                            var_list = ['aciUrl', 'aciUser', 'aciPass']
                        else:
                            var_list = ['aciUrl', 'aciCertName', 'aciPrivateKey']

                        # Get var_ids
                        tf_var_dict = {}
                        folder_id = 'Site_ID_%s_%s' % (templateVars['Site_ID'], dest_dir)
                        kwargs['workspace_id'] = workspace_dict[folder_id]
                        kwargs['Description'] = ''
                        for var in var_list:
                            tf_var_dict = tf_variables(class_init, dest_dir, var, tf_var_dict, **kwargs)

                    self.templateLoader = jinja2.FileSystemLoader(
                        searchpath=(aci_template_path))
                    self.templateEnv = jinja2.Environment(loader=self.templateLoader)

                    # Add the Default Files to the Tenant Directory
                    file_list = ['.gitignore_.gitignore', 'main.jinja2_main.tf', 'variables.jinja2_variables.tf']
                    for file in file_list:
                        x = file.split('_')
                        template_file = x[0]
                        dest_file = x[1]
                        template = self.templateEnv.get_template(template_file)
                        create_tf_file('w', dest_dir, dest_file, template, **templateVars)

        elif re.search(r'\d+', templateVars['Site_Group']):
            Site_ID = 'Site_ID_%s' % (templateVars['Site_Group'])
            site_dict = ast.literal_eval(os.environ[Site_ID])

            # Create kwargs for Site Variables
            kwargs['Site_ID'] = site_dict.get('Site_ID')
            kwargs['Site_Name'] = site_dict.get('Site_Name')
            kwargs['APIC_URL'] = site_dict.get('APIC_URL')
            kwargs['APIC_Version'] = site_dict.get('APIC_Version')
            kwargs['APIC_Auth_Type'] = site_dict.get('APIC_Auth_Type')
            kwargs['Terraform_EQ'] = site_dict.get('Terraform_EQ')
            kwargs['Terraform_Version'] = site_dict.get('Terraform_Version')
            kwargs['Provider_EQ'] = site_dict.get('Provider_EQ')
            kwargs['Provider_Version'] = site_dict.get('Provider_Version')
            kwargs['Run_Location'] = site_dict.get('Run_Location')
            kwargs['State_Location'] = site_dict.get('State_Location')
            kwargs['Terraform_Cloud_Org'] = site_dict.get('Terraform_Cloud_Org')
            kwargs['Workspace_Prefix'] = site_dict.get('Workspace_Prefix')
            kwargs['VCS_Base_Repo'] = site_dict.get('VCS_Base_Repo')
            kwargs['Terraform_Agent_Pool_ID'] = site_dict.get('Terraform_Agent_Pool_ID')

            # Dicts for required and optional args
            required_args = {'Site_ID': '',
                                'Site_Name': '',
                                'APIC_URL': '',
                                'APIC_Version': '',
                                'APIC_Auth_Type': '',
                                'Terraform_EQ': '',
                                'Terraform_Version': '',
                                'Provider_EQ': '',
                                'Provider_Version': '',
                                'Run_Location': '',
                                'State_Location': ''}
            optional_args = {'Terraform_Cloud_Org': '',
                                'Workspace_Prefix': '',
                                'VCS_Base_Repo': '',
                                'Terraform_Agent_Pool_ID': ''}

            # Validate inputs, return dict of template vars
            templateVars = process_kwargs(required_args, optional_args, **kwargs)

            # If the State_Location is Terraform_Cloud Configure Workspaces in the Cloud
            if templateVars['State_Location'] == 'Terraform_Cloud':
                # Initialize the Class
                lib_tf_ref = 'lib_terraform.Terraform_Cloud'
                class_init = '%s()' % (lib_tf_ref)

                # Get terraform_cloud_token
                kwargs['terraform_cloud_token'] = eval("%s.%s()" % (class_init, 'terraform_token'))

                # Get terraform_cloud_token
                kwargs['terraform_oath_token'] = eval("%s.%s(**kwargs)" % (class_init, 'oath_token'))

                # Get workspace_ids
                workspace_dict = {}
                workspace_dict = tf_workspace(class_init, dest_dir, workspace_dict, **kwargs)

            # If the Run_Location is Terraform_Cloud Configure Variables in the Cloud
            if templateVars['Run_Location'] == 'Terraform_Cloud':
                # Set Variable List
                if templateVars['APIC_Auth_Type'] == 'user_pass':
                    var_list = ['aciUrl', 'aciUser', 'aciPass']
                else:
                    var_list = ['aciUrl', 'aciCertName', 'aciPrivateKey']

                # Get var_ids
                tf_var_dict = {}
                folder_id = 'Site_ID_%s_%s' % (templateVars['Site_ID'], dest_dir)
                kwargs['workspace_id'] = workspace_dict[folder_id]
                kwargs['Description'] = ''
                for var in var_list:
                    tf_var_dict = tf_variables(class_init, dest_dir, var, tf_var_dict, **kwargs)

            self.templateLoader = jinja2.FileSystemLoader(
                searchpath=(aci_template_path))
            self.templateEnv = jinja2.Environment(loader=self.templateLoader)

            # Add the Default Files to the Tenant Directory
            file_list = ['.gitignore_.gitignore', 'main.jinja2_main.tf', 'variables.jinja2_variables.tf']
            for file in file_list:
                x = file.split('_')
                template_file = x[0]
                dest_file = x[1]
                template = self.templateEnv.get_template(template_file)
                create_tf_file('w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def add_vrf(self, wb, ws, row_num, **kwargs):
        # Open the Network Policies Worksheet
        ws_net = wb['Network Policies']
        rows = ws_net.max_row

        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Tenant': '',
                         'VRF': '',
                         'VRF_Policy': '',
                         'Policy_Name': '',
                         'pc_enf_pref': '',
                         'pc_enf_dir': '',
                         'bd_enforce': '',
                         'enf_type': '',
                         'fvEpRetPol': '',
                         'monEPGPol': '',
                         'ip_dp_learning': '',
                         'knw_mcast_act': ''}
        optional_args = {'Alias': '',
                         'Description': '',
                         'Tags': '',
                         'cons_vzBrCP': '',
                         'vzCPIf': '',
                         'prov_vzBrCP': '',
                         'bgpCtxPol': '',
                         'bgpCtxAfPol': '',
                         'ospfCtxPol': '',
                         'ospfCtxAfPol': '',
                         'eigrpCtxAfPol': '',
                         'l3extRouteTagPol': '',
                         'l3extVrfValidationPol': ''}


        # Get the VRF Policies from the Network Policies Tab
        func = 'VRF'
        count = countKeys(ws_net, func)
        row_vrf = ''
        var_dict = findVars(ws_net, func, rows, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('VRF_Policy'):
                row_vrf = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
            validating.name_rule(row_num, ws, 'VRF', templateVars['VRF'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
            if not templateVars['Tags'] == None:
                if re.match(',', templateVars['Tags']):
                    for tag in templateVars['Tags'].split(','):
                        validating.name_rule(row_num, ws, 'Tags', tag)
                else:
                    validating.name_rule(row_num, ws, 'Tags', templateVars['Tags'])
            validating.values(row_vrf, ws_net, 'bd_enforce', templateVars['bd_enforce'], ['no', 'yes'])
            validating.values(row_vrf, ws_net, 'ip_dp_learning', templateVars['ip_dp_learning'], ['disabled', 'enabled'])
            validating.values(row_vrf, ws_net, 'knw_mcast_act', templateVars['knw_mcast_act'], ['deny', 'permit'])
            validating.values(row_vrf, ws_net, 'pc_enf_dir', templateVars['pc_enf_dir'], ['egress', 'ingress'])
            validating.values(row_vrf, ws_net, 'pc_enf_pref', templateVars['pc_enf_pref'], ['enforced', 'unenforced'])
            validating.values(row_vrf, ws_net, 'enf_type', templateVars['enf_type'], ['contract', 'pref_grp', 'vzAny'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        if templateVars['cons_vzBrCP'] == 'default':
            templateVars['cons_vzBrCP'] = 'uni/tn-common/brc-default'
        if templateVars['prov_vzBrCP'] == 'default':
            templateVars['prov_vzBrCP'] = 'uni/tn-common/brc-default'
        if templateVars['vzCPIf'] == 'default':
            templateVars['vzCPIf'] = 'uni/tn-common/cif-default'
        if templateVars['bgpCtxPol'] == 'default':
            templateVars['bgpCtxPol'] = 'uni/tn-common/bgpCtxP-default'
        if templateVars['bgpCtxAfPol'] == 'default':
            templateVars['bgpCtxAfPol'] = 'uni/tn-common/bgpCtxAfP-default'
        if templateVars['eigrpCtxAfPol'] == 'default':
            templateVars['eigrpCtxAfPol'] = 'uni/tn-common/eigrpCtxAfP-default'
        if templateVars['ospfCtxPol'] == 'default':
            templateVars['ospfCtxPol'] = 'uni/tn-common/ospfCtxP-default'
        if templateVars['ospfCtxAfPol'] == 'default':
            templateVars['ospfCtxAfPol'] = 'uni/tn-common/ospfCtxP-default'
        if templateVars['fvEpRetPol'] == 'default':
            templateVars['fvEpRetPol'] = 'uni/tn-common/epRPol-default'
        if templateVars['monEPGPol'] == 'default':
            templateVars['monEPGPol'] = 'uni/tn-common/monepg-default'
        if templateVars['l3extVrfValidationPol'] == 'default':
            templateVars['l3extVrfValidationPol'] = 'uni/tn-common/vrfvalidationpol-default'

        # Define the Template Source
        template_file = "vrf.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'VRF_%s.tf' % (templateVars['VRF'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        if templateVars['enf_type'] == 'pref_grp':
            # Define the Template Source
            template_file = "pref_grp.jinja2"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = 'VRF_%s.tf' % (templateVars['VRF'])
            dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
            process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

        elif templateVars['enf_type'] == 'vzAny':
            # Define the Template Source
            template_file = "vzAny.jinja2"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = 'VRF_%s.tf' % (templateVars['VRF'])
            dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
            process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "snmp_ctx.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'VRF_%s.tf' % (templateVars['VRF'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def bgp_peer(self, wb, ws, row_num, **kwargs):
        # Open the Network Policies Worksheet
        ws_net = wb['Network Policies']
        rows = ws_net.max_row

        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Path_Policy_Name': '',
                         'Peer_Interface': '',
                         'Peer_Address': '',
                         'Remote_ASN': '',
                         'eBGP_Multihop_TTL': '',
                         'Weight': '',
                         'Local_ASN_Config': '',
                         'Admin_State': '',
                         'BGP_Peer_Policy': '',
                         'Tenant': '',
                         'L3Out': '',
                         'Profile_Name': '',
                         'Node_Profile': '',
                         'Interface_Profile': '',
                         'Interface_Type': '',
                         'Pod_ID': '',
                         'Node1_ID': '',
                         'Interface_or_PG': '',
                         'allow_self_as': '',
                         'as_override': '',
                         'disable_peer_as_check': '',
                         'next_hop_self': '',
                         'send_community': '',
                         'send_ext_community': '',
                         'allowed_self_as_count': '',
                         'bfd': '',
                         'disable_connected_check': '',
                         'AF_Mcast': '',
                         'AF_Ucast': '',
                         'remove_all_private_as': '',
                         'remove_private_as': '',
                         'private_to_local': ''}
        optional_args = {'Description': '',
                         'BGP_Password': '',
                         'BGP_Peer_Prefix_Policy': '',
                         'Local_ASN': '',
                         'Node2_ID': '',
                         'Policy_Name': '',
                         'Prefix_Tenant': '',
                         'PFX_Description': '',
                         'Action': '',
                         'Maximum_Prefixes': '',
                         'Restart_Time': '',
                         'Threshold': ''}

        # Get the Node Policies from the Network Policies Tab
        rows = ws.max_row
        func = 'l3out_path'
        count = countKeys(ws, func)
        row_path = ''
        var_dict = findVars(ws, func, rows, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('Path_Policy_Name'):
                row_path = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}

        # Get the Node Policies from the Network Policies Tab
        func = 'bgp_profile'
        bgp_count = countKeys(ws_net, func)
        row_bgp = ''
        var_dict = findVars(ws_net, func, rows, bgp_count)
        for pos in var_dict:
            if var_dict[pos].get('Profile_Name') == kwargs.get('BGP_Peer_Policy'):
                row_bgp = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}

        if not kwargs.get('BGP_Peer_Prefix_Policy') == None:
            # Get the Node Policies from the Network Policies Tab
            func = 'pfx_policy'
            pfx_count = countKeys(ws_net, func)
            row_pfx = ''
            var_dict = findVars(ws_net, func, rows, pfx_count)
            for pos in var_dict:
                if var_dict[pos].get('Policy_Name') == kwargs.get('BGP_Peer_Prefix_Policy'):
                    row_pfx = var_dict[pos]['row']
                    del var_dict[pos]['row']
                    kwargs = {**kwargs, **var_dict[pos]}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.values(row_num, ws, 'Peer_Interface', templateVars['Peer_Interface'], ['Interface', 'Loopback'])
            validating.ip_address(row_num, ws, 'Peer_Address', templateVars['Peer_Address'])
            validating.number_check(row_num, ws, 'Remote_ASN', templateVars['Remote_ASN'], 1, 4294967295)
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
            validating.number_check(row_num, ws, 'eBGP_Multihop_TTL', templateVars['eBGP_Multihop_TTL'], 1, 255)
            validating.number_check(row_num, ws, 'Weight', templateVars['Weight'], 0, 65535)
            if not templateVars['BGP_Peer_Prefix_Policy'] == None:
                validating.name_rule(row_num, ws, 'BGP_Peer_Prefix_Policy', templateVars['BGP_Peer_Prefix_Policy'])
            validating.values(row_num, ws, 'Local_ASN_Config', templateVars['Local_ASN_Config'], ['dual-as', 'no-prepend', 'none', 'replace-as'])
            if not templateVars['Local_ASN'] == None:
                validating.number_check(row_num, ws, 'Local_ASN', templateVars['Local_ASN'], 1, 4294967295)
            validating.values(row_num, ws, 'Admin_State', templateVars['Admin_State'], ['disabled', 'enabled'])
            validating.name_rule(row_num, ws, 'BGP_Peer_Policy', templateVars['BGP_Peer_Policy'])

            validating.name_rule(row_path, ws, 'Tenant', templateVars['Tenant'])
            validating.name_rule(row_path, ws, 'L3Out', templateVars['L3Out'])
            validating.name_rule(row_path, ws, 'Node_Profile', templateVars['Node_Profile'])
            validating.name_rule(row_path, ws, 'Interface_Profile', templateVars['Interface_Profile'])
            validating.values(row_path, ws, 'Interface_Type', templateVars['Interface_Type'], ['ext-svi', 'l3-port', 'sub-interface'])
            validating.number_check(row_path, ws, 'Pod_ID', templateVars['Pod_ID'], 1, 15)
            validating.number_check(row_path, ws, 'Node1_ID', templateVars['Node1_ID'], 101, 4001)
            if not templateVars['Node2_ID'] == None:
                validating.number_check(row_path, ws, 'Node2_ID', templateVars['Node2_ID'], 101, 4001)

            validating.number_check(row_bgp, ws_net, 'allowed_self_as_count', templateVars['allowed_self_as_count'], 1, 10)
            validating.values(row_bgp, ws_net, 'allow_self_as', templateVars['allow_self_as'], ['no', 'yes'])
            validating.values(row_bgp, ws_net, 'as_override', templateVars['as_override'], ['no', 'yes'])
            validating.values(row_bgp, ws_net, 'disable_peer_as_check', templateVars['disable_peer_as_check'], ['no', 'yes'])
            validating.values(row_bgp, ws_net, 'next_hop_self', templateVars['next_hop_self'], ['no', 'yes'])
            validating.values(row_bgp, ws_net, 'send_community', templateVars['send_community'], ['no', 'yes'])
            validating.values(row_bgp, ws_net, 'send_ext_community', templateVars['send_ext_community'], ['no', 'yes'])
            validating.values(row_bgp, ws_net, 'bfd', templateVars['bfd'], ['no', 'yes'])
            validating.values(row_bgp, ws_net, 'AF_Mcast', templateVars['AF_Mcast'], ['no', 'yes'])
            validating.values(row_bgp, ws_net, 'AF_Ucast', templateVars['AF_Ucast'], ['no', 'yes'])
            validating.values(row_bgp, ws_net, 'remove_all_private_as', templateVars['remove_all_private_as'], ['no', 'yes'])
            validating.values(row_bgp, ws_net, 'remove_private_as', templateVars['remove_private_as'], ['no', 'yes'])
            validating.values(row_bgp, ws_net, 'private_to_local', templateVars['private_to_local'], ['no', 'yes'])

            if not templateVars['BGP_Peer_Prefix_Policy'] == None:
                validating.name_rule(row_pfx, ws_net, 'Prefix_Tenant', templateVars['Prefix_Tenant'])
                validating.number_check(row_pfx, ws_net, 'Maximum_Prefixes', templateVars['Maximum_Prefixes'], 1, 300000)
                validating.number_check(row_pfx, ws_net, 'Restart_Time', templateVars['Restart_Time'], 1, 65535)
                validating.number_check(row_pfx, ws_net, 'Threshold', templateVars['Threshold'], 1, 100)
                validating.values(row_pfx, ws_net, 'Action', templateVars['Action'], ['log', 'reject', 'restart', 'shut'])
                if not templateVars['PFX_Description'] == None:
                    validating.description(row_pfx, ws_net, 'PFX_Description', templateVars['PFX_Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        if re.search(r'\.', templateVars['Peer_Address']):
            templateVars['Peer_Address_'] = templateVars['Peer_Address'].replace('.', '-')
        else:
            templateVars['Peer_Address_'] = templateVars['Peer_Address'].replace(':', '-')

        ctrl_count = 0
        Ctrl = ''
        if templateVars['allow_self_as'] == 'yes':
            Ctrl = 'allow-self-as'
            ctrl_count =+ 1
        if templateVars['as_override'] == 'yes' and ctrl_count > 0:
            Ctrl = Ctrl + ',' + 'as-override'
            ctrl_count =+ 1
        elif templateVars['as_override'] == 'yes':
            Ctrl = 'as-override'
            ctrl_count =+ 1
        if templateVars['disable_peer_as_check'] == 'yes' and ctrl_count > 0:
            Ctrl = Ctrl + ',' + 'dis-peer-as-check'
            ctrl_count =+ 1
        elif templateVars['disable_peer_as_check'] == 'yes':
            Ctrl = 'dis-peer-as-check'
            ctrl_count =+ 1
        if templateVars['next_hop_self'] == 'yes' and ctrl_count > 0:
            Ctrl = Ctrl + ',' + 'nh-self'
            ctrl_count =+ 1
        elif templateVars['next_hop_self'] == 'yes':
            Ctrl = 'nh-self'
            ctrl_count =+ 1
        if templateVars['send_community'] == 'yes' and ctrl_count > 0:
            Ctrl = Ctrl + ',' + 'send-com'
            ctrl_count =+ 1
        elif templateVars['send_community'] == 'yes':
            Ctrl = 'send-com'
            ctrl_count =+ 1
        if templateVars['send_ext_community'] == 'yes' and ctrl_count > 0:
            Ctrl = Ctrl + ',' + 'send-ext-com'
            ctrl_count =+ 1
        elif templateVars['send_ext_community'] == 'yes':
            Ctrl = 'send-ext-com'
            ctrl_count =+ 1
        if ctrl_count > 0:
            templateVars['Ctrl'] = '%s' % (Ctrl)
        else:
            templateVars['Ctrl'] = ''

        ctrl_count = 0
        Ctrl = ''
        if templateVars['bfd'] == 'yes':
            Ctrl = 'bfd'
            ctrl_count =+ 1
        if templateVars['disable_connected_check'] == 'yes' and ctrl_count > 0:
            Ctrl = Ctrl + ',' + 'dis-conn-check'
            ctrl_count =+ 1
        elif templateVars['disable_connected_check'] == 'yes':
            Ctrl = 'dis-conn-check'
            ctrl_count =+ 1
        if ctrl_count > 0:
            templateVars['Peer_Ctrl'] = '%s' % (Ctrl)
        else:
            templateVars['Peer_Ctrl'] = ''

        ctrl_count = 0
        Ctrl = ''
        if templateVars['AF_Mcast'] == 'yes':
            Ctrl = 'af-mcast'
            ctrl_count =+ 1
        if templateVars['AF_Ucast'] == 'yes' and ctrl_count > 0:
            Ctrl = Ctrl + ',' + 'af-ucast'
            ctrl_count =+ 1
        elif templateVars['AF_Ucast'] == 'yes':
            Ctrl = 'af-ucast'
            ctrl_count =+ 1
        if ctrl_count > 0:
            templateVars['Address_Fam_Ctrl'] = '%s' % (Ctrl)
        else:
            templateVars['Address_Fam_Ctrl'] = ''

        ctrl_count = 0
        Ctrl = ''
        if templateVars['remove_all_private_as'] == 'yes':
            Ctrl = 'remove-all'
            ctrl_count =+ 1
        if templateVars['remove_private_as'] == 'yes' and ctrl_count > 0:
            Ctrl = Ctrl + ',' + 'remove-exclusive'
            ctrl_count =+ 1
        elif templateVars['remove_private_as'] == 'yes':
            Ctrl = 'remove-exclusive'
            ctrl_count =+ 1
        if templateVars['private_to_local'] == 'yes' and ctrl_count > 0:
            Ctrl = Ctrl + ',' + 'replace-as'
            ctrl_count =+ 1
        elif templateVars['private_to_local'] == 'yes':
            Ctrl = 'replace-as'
            ctrl_count =+ 1
        if ctrl_count > 0:
            templateVars['Private_AS_Ctrl'] = '%s' % (Ctrl)
        else:
            templateVars['Private_AS_Ctrl'] = ''

        if not templateVars['BGP_Password'] == None:
            x = templateVars['BGP_Password'].split('r')
            key_number = x[1]
            templateVars['sensitive_var'] = 'BGP_Password%s' % (key_number)

            # Define the Template Source
            template_file = "variables.jinja2"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = 'variable_%s.tf' % (templateVars['sensitive_var'])
            dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
            process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

            process_sensitive_var(wb, ws, row_num, dest_dir, dest_file, template, **templateVars)

        # Create Global Variables for First Template
        if not templateVars['Node2_ID'] == None:
            templateVars['PATH'] = 'protpaths-%s-%s' % (templateVars['Node1_ID'], templateVars['Node2_ID'])
        else:
            templateVars['PATH'] = 'paths-%s' % (templateVars['Node1_ID'])

        # Define the Template Source
        if templateVars['PATH'] == 'Loopback':
            template_file = "bgp_peer_connectivity_profile.jinja2"
        else:
            if templateVars['Local_ASN_Config'] == 'none':
                templateVars['Local_ASN_Config'] == None
            template_file = "bgp_peer_interface.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'L3Out_%s_Node_Profile_%s.tf' % (templateVars['L3Out'], templateVars['Node_Profile'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

        if not templateVars['BGP_Peer_Prefix_Policy'] == None:

            # Define the Template Source
            if templateVars['Restart_Time'] == 65535:
                templateVars['Restart_Time'] = 'infinite'
            template_file = "bgp_peer_prefix.jinja2"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = 'BGP_Peer_Prefix_%s.tf' % (templateVars['Policy_Name'])
            dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
            print(f'dest file {dest_file} and dest_dir {dest_dir} and Tenant = {templateVars["Tenant"]}')
            process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

            if not templateVars['Tenant'] == templateVars['Prefix_Tenant']:
                # Define the Template Source
                template_file = "data_tenant.jinja2"
                template = self.templateEnv.get_template(template_file)

                temp_Tenant = templateVars['Tenant']
                templateVars['Tenant'] = templateVars['Prefix_Tenant']
                # Process the template through the Sites
                dest_file = 'data_Tenant_%s.tf' % (templateVars['Tenant'])
                dest_dir = 'Tenant_%s' % (temp_Tenant)
                process_method(wb, ws_net, row_pfx, 'w', dest_dir, dest_file, template, **templateVars)

                templateVars['Tenant'] = temp_Tenant

                # Define the Template Source
                template_file = "data_bgp_peer_prefix.jinja2"
                template = self.templateEnv.get_template(template_file)

                # Process the template through the Sites
                dest_file = 'data_BGP_Peer_Prefix_%s.tf' % (templateVars['Policy_Name'])
                dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
                process_method(wb, ws_net, row_pfx, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def ctx_comm(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Tenant': '',
                         'VRF': '',
                         'Ctx_Community': ''}
        optional_args = {'Description': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
            validating.name_rule(row_num, ws, 'VRF', templateVars['VRF'])
            validating.snmp_string(row_num, ws, 'Ctx_Community', templateVars['Ctx_Community'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "snmp_ctx_community.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'VRF_%s.tf' % (templateVars['VRF'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def contract_add(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Tenant': '',
                         'Contract_Type': '',
                         'Contract': '',
                         'Scope': '',
                         'QoS_Class': '',
                         'Target_DSCP': ''}
        optional_args = {'Description': '',
                         'Alias': '',
                         'Tags': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.dscp(row_num, ws, 'Target_DSCP', templateVars['Target_DSCP'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
            validating.name_rule(row_num, ws, 'Contract', templateVars['Contract'])
            validating.qos_priority(row_num, ws, 'QoS_Class', templateVars['QoS_Class'])
            validating.values(row_num, ws, 'Contract_Type', templateVars['Contract_Type'], ['OOB', 'Standard', 'Taboo'])
            validating.values(row_num, ws, 'Scope', templateVars['Scope'], ['application-profile', 'context', 'global', 'tenant'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        if templateVars['Contract_Type'] == 'OOB':
            template_file = "contract_oob.jinja2"
        elif templateVars['Contract_Type'] == 'Standard':
            template_file = "contract.jinja2"
        elif templateVars['Contract_Type'] == 'Taboo':
            template_file = "contract_taboo.jinja2"
        dest_file = 'Contract_Type_%s_%s.tf' % (templateVars['Contract_Type'], templateVars['Contract'])
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def contract_to_epg(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Contract_Tenant': '',
                         'Contract_Type': '',
                         'Contract': '',
                         'Tenant': '',
                         'App_Profile': '',
                         'EPG': ''}
        optional_args = { }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Contract_Tenant', templateVars['Contract_Tenant'])
            validating.name_rule(row_num, ws, 'Contract', templateVars['Contract'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
            validating.name_rule(row_num, ws, 'App_Profile', templateVars['App_Profile'])
            validating.name_rule(row_num, ws, 'EPG', templateVars['EPG'])
            validating.values(row_num, ws, 'Contract_Type', templateVars['Contract_Type'], ['consumer', 'provider'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "epg_to_contract.jinja2"
        dest_file = 'App_Profile_%s_EPG_%s.tf' % (templateVars['App_Profile'], templateVars['EPG'])
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def ext_epg(self, wb, ws, row_num, **kwargs):
        # Open the Network Policies Worksheet
        ws_net = wb['Network Policies']
        rows = ws_net.max_row

        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Tenant': '',
                         'L3Out': '',
                         'Ext_EPG': '',
                         'Ext_EPG_Policy': '',
                         'Subnets': '',
                         'Ext_Subnet_Policy': '',
                         'prio': '',
                         'target_dscp': '',
                         'pref_gr_memb': '',
                         'match_t': '',
                         'flood': '',
                         'export-rtctrl': '',
                         'import-rtctrl': '',
                         'import-security': '',
                         'shared-security': '',
                         'shared-rtctrl': '',
                         'agg-export': '',
                         'agg-import': '',
                         'agg-shared': ''}
        optional_args = {'Alias': '',
                         'Description': '',
                         'Tags': '',
                         'cons_vzBrCP': '',
                         'vzCPIf': '',
                         'Master_fvEPg': '',
                         'prov_vzBrCP': '',
                         'vzTaboo': '',
                         'exception_tag': '',
                         'rtctrlProfile': '',
                         'sub_rtctrlProfile': '',
                         'rtsumARtSummPol': ''}


        # Get the L3Out Policies from the Network Policies Tab
        func = 'ext_epg'
        count = countKeys(ws_net, func)
        row_epg = ''
        var_dict = findVars(ws_net, func, rows, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('Ext_EPG_Policy'):
                row_epg = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}

        func = 'ext_subnet'
        count = countKeys(ws_net, func)
        row_sub = ''
        var_dict = findVars(ws_net, func, rows, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('Ext_Subnet_Policy'):
                row_sub = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
            if not templateVars['Subnets'] == None:
                if re.search(',', templateVars['Subnets']):
                    sx = templateVars['Subnets'].split(',')
                    for x in sx:
                        validating.ip_address(row_num, ws, 'Subnets', x)
                else:
                    validating.ip_address(row_num, ws, 'Subnets', templateVars['Subnets'])
            validating.dscp(row_epg, ws_net, 'target_dscp', templateVars['target_dscp'])
            validating.match_t(row_epg, ws_net, 'match_t', templateVars['match_t'])
            validating.qos_priority(row_epg, ws_net, 'prio', templateVars['prio'])
            validating.values(row_epg, ws_net, 'flood', templateVars['flood'], ['disabled', 'enabled'])
            validating.values(row_epg, ws_net, 'pref_gr_memb', templateVars['pref_gr_memb'], ['exclude', 'include'])
            validating.values(row_sub, ws_net, 'agg-export', templateVars['agg-export'], ['no', 'yes'])
            validating.values(row_sub, ws_net, 'agg-import', templateVars['agg-import'], ['no', 'yes'])
            validating.values(row_sub, ws_net, 'agg-shared', templateVars['agg-shared'], ['no', 'yes'])
            validating.values(row_sub, ws_net, 'export-rtctrl', templateVars['export-rtctrl'], ['no', 'yes'])
            validating.values(row_sub, ws_net, 'import-rtctrl', templateVars['import-rtctrl'], ['no', 'yes'])
            validating.values(row_sub, ws_net, 'import-security', templateVars['import-security'], ['no', 'yes'])
            validating.values(row_sub, ws_net, 'shared-security', templateVars['shared-security'], ['no', 'yes'])
            validating.values(row_sub, ws_net, 'shared-rtctrl', templateVars['shared-rtctrl'], ['no', 'yes'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Create aggregate templateVars
        aggregate = ''
        aggregate_count = 0
        if templateVars['agg-export'] == 'yes':
            aggregate = '"agg-export"'
            aggregate_count =+ 1
        if templateVars['agg-import'] == 'yes':
            if aggregate_count == 0:
                aggregate = '"agg-import"'
                aggregate_count =+ 1
            else:
                aggregate = aggregate + ', ' + '"agg-import"'
                aggregate_count =+ 1
        if templateVars['agg-shared'] == 'yes':
            if aggregate_count == 0:
                aggregate = '"agg-import"'
                aggregate_count =+ 1
            else:
                aggregate = aggregate + ', ' + '"agg-shared"'
                aggregate_count =+ 1

        if aggregate_count == 0:
            templateVars['aggregate'] = None
        else:
            templateVars['aggregate'] = '[%s]' % (aggregate)

        # Create scope templateVars
        scope = ''
        scope_count = 0
        if templateVars['export-rtctrl'] == 'yes':
            scope = '"export-rtctrl"'
            scope_count =+ 1
        if templateVars['import-rtctrl'] == 'yes':
            if scope_count == 0:
                scope = '"import-rtctrl"'
                scope_count =+ 1
            else:
                scope = scope + ', ' + '"import-rtctrl"'
                scope_count =+ 1
        if templateVars['import-security'] == 'yes':
            if scope_count == 0:
                scope = '"import-security"'
                scope_count =+ 1
            else:
                scope = scope + ', ' + '"import-security"'
                scope_count =+ 1
        if templateVars['shared-security'] == 'yes':
            if scope_count == 0:
                scope = '"shared-security"'
                scope_count =+ 1
            else:
                scope = scope + ', ' + '"shared-security"'
                scope_count =+ 1
        if templateVars['shared-rtctrl'] == 'yes':
            if scope_count == 0:
                scope = '"shared-rtctrl"'
                scope_count =+ 1
            else:
                scope = scope + ', ' + '"shared-rtctrl"'
                scope_count =+ 1

        if scope_count == 0:
            templateVars['scope'] = None
        else:
            templateVars['scope'] = '[%s]' % (scope)

        # Define the Template Source
        template_file = "epg_ext.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'L3Out_%s_EXTERNAL_EPG_%s.tf' % (templateVars['L3Out'], templateVars['Ext_EPG'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        if re.search(',', templateVars['Subnets']):
            sx = templateVars['Subnets'].split(',')
            for x in sx:
                templateVars['Subnet'] = x
                if re.search(':', x):
                    templateVars['Subnet_'] = x.replace(':', '-')
                    templateVars['Subnet_'] = templateVars['Subnet_'].replace('/', '_')
                else:
                    templateVars['Subnet_'] = x.replace('.', '-')
                    templateVars['Subnet_'] = templateVars['Subnet_'].replace('/', '_')

                # Define the Template Source
                template_file = "ext_subnet.jinja2"
                template = self.templateEnv.get_template(template_file)

                # Process the template through the Sites
                dest_file = 'L3Out_%s_EXTERNAL_EPG_%s.tf' % (templateVars['L3Out'], templateVars['Ext_EPG'])
                dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
                process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)
        else:
            templateVars['Subnet'] = templateVars['Subnets']
            if re.search(':', templateVars['Subnet']):
                templateVars['Subnet_'] = templateVars['Subnet'].replace(':', '-')
                templateVars['Subnet_'] = templateVars['Subnet_'].replace('/', '_')
            else:
                templateVars['Subnet_'] = templateVars['Subnet'].replace('.', '-')
                templateVars['Subnet_'] = templateVars['Subnet_'].replace('/', '_')

            # Define the Template Source
            template_file = "ext_subnet.jinja2"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = 'L3Out_%s_EXTERNAL_EPG_%s.tf' % (templateVars['L3Out'], templateVars['Ext_EPG'])
            dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
            process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def extepg_oob(self, wb, ws, row_num, **kwargs):
        # Open the Network Policies Worksheet
        ws_net = wb['Network Policies']
        rows = ws_net.max_row

        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Ext_EPG': '',
                         'QoS_Class': '',
                         'Subnets': ''}
        optional_args = {'Tags': '',
                         'consumed_Contracts': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Ext_EPG', templateVars['Ext_EPG'])
            validating.qos_priority(row_num, ws, 'QoS_Class', templateVars['QoS_Class'])
            if not templateVars['Tags'] == None:
                if re.search(',', templateVars['Tags']):
                    for x in templateVars['Tags'].split(','):
                        validating.description(row_num, ws, 'Tags', x)
                else:
                    validating.description(row_num, ws, 'Tags', templateVars['Tags'])
            if not templateVars['consumed_Contracts'] == None:
                if re.search(',', templateVars['consumed_Contracts']):
                    templateVars['provide_count'] =+ 1
                    for x in templateVars['consumed_Contracts'].split(','):
                        validating.name_rule(row_num, ws, 'consumed_Contracts', x)
                else:
                    validating.name_rule(row_num, ws, 'consumed_Contracts', templateVars['consumed_Contracts'])
            if not templateVars['Subnet'] == None:
                if re.search(',', templateVars['Subnets']):
                    sx = templateVars['Subnets'].split(',')
                    for x in sx:
                        validating.ip_address(row_num, ws, 'Subnets', x)
                else:
                    validating.ip_address(row_num, ws, 'Subnets', templateVars['Subnets'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "epg_ext_oob.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'EPG_Mgmt_OOB_External_EPG_%s.tf' % (templateVars['Ext_EPG'])
        dest_dir = 'Tenant_mgmt'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        if re.search(',', templateVars['Subnets']):
            sx = templateVars['Subnets'].split(',')
            for x in sx:
                templateVars['Subnet'] = x
                if re.search(':', x):
                    templateVars['Subnet_'] = x.replace(':', '-')
                    templateVars['Subnet_'] = templateVars['Subnet_'].replace('/', '_')
                else:
                    templateVars['Subnet_'] = x.replace('.', '-')
                    templateVars['Subnet_'] = templateVars['Subnet_'].replace('/', '_')

                # Define the Template Source
                template_file = "ext_subnet_oob.jinja2"
                template = self.templateEnv.get_template(template_file)

                # Process the template through the Sites
                dest_file = 'EPG_Mgmt_OOB_External_EPG_%s.tf' % (templateVars['Ext_EPG'])
                dest_dir = 'Tenant_mgmt'
                process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)
        else:
            templateVars['Subnet'] = templateVars['Subnets']
            if re.search(':', templateVars['Subnet']):
                templateVars['Subnet_'] = templateVars['Subnet'].replace(':', '-')
                templateVars['Subnet_'] = templateVars['Subnet_'].replace('/', '_')
            else:
                templateVars['Subnet_'] = templateVars['Subnet'].replace('.', '-')
                templateVars['Subnet_'] = templateVars['Subnet_'].replace('/', '_')

            # Define the Template Source
            template_file = "ext_subnet_oob.jinja2"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = 'EPG_Mgmt_OOB_External_EPG_%s.tf' % (templateVars['Ext_EPG'])
            dest_dir = 'Tenant_mgmt'
            process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def filter_add(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Tenant': '',
                         'Filter': ''}
        optional_args = {'Description': '',
                         'Alias': '',
                         'Tags': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Filter', templateVars['Filter'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "contract_filter.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Contract_Filter_%s.tf' % (templateVars['Filter'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def filter_entry(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Tenant': '',
                         'Filter': '',
                         'Filter_Entry': '',
                         'EtherType': '',
                         'IP_Protocol': '',
                         'ARP_Flag': '',
                         'ICMPv4_Type': '',
                         'ICMPv6_Type': '',
                         'Match_DSCP': '',
                         'Match_Only_Frags': '',
                         'Source_From': '',
                         'Source_To': '',
                         'Dest_From': '',
                         'Dest_To': '',
                         'Stateful': '',
                         'TCP_Session_Rules': ''}
        optional_args = {'Description': '',
                         'Alias': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.dscp(row_num, ws, 'Match_DSCP', templateVars['Match_DSCP'])
            validating.filter_ports(row_num, ws, 'Source_From', templateVars['Source_From'])
            validating.filter_ports(row_num, ws, 'Source_To', templateVars['Source_To'])
            validating.filter_ports(row_num, ws, 'Dest_From', templateVars['Dest_From'])
            validating.filter_ports(row_num, ws, 'Dest_To', templateVars['Dest_To'])
            validating.name_rule(row_num, ws, 'Filter', templateVars['Filter'])
            validating.name_rule(row_num, ws, 'Filter_Entry', templateVars['Filter_Entry'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
            validating.values(row_num, ws, 'Match_Only_Frags', templateVars['Match_Only_Frags'], ['no', 'yes'])
            validating.values(row_num, ws, 'Stateful', templateVars['Stateful'], ['no', 'yes'])
            validating.values(row_num, ws, 'EtherType', templateVars['EtherType'], ['arp', 'fcoe', 'ip', 'ipv4', 'ipv6', 'trill', 'mac_security', 'mpls_ucast', 'unspecified'])
            validating.values(row_num, ws, 'IP_Protocol', templateVars['IP_Protocol'], ['egp', 'eigrp', 'igp', 'icmp', 'icmpv6', 'igmp', 'l2tp', 'ospfigp', 'pim', 'tcp', 'udp', 'unspecified'])
            validating.values(row_num, ws, 'ARP_Flag', templateVars['ARP_Flag'], ['req', 'reply', 'unspecified'])
            validating.values(row_num, ws, 'ICMPv4_Type', templateVars['ICMPv4_Type'], ['dst-unreach', 'echo', 'echo-rep', 'src-quench', 'time-exceeded', 'unspecified'])
            validating.values(row_num, ws, 'ICMPv6_Type', templateVars['ICMPv6_Type'], ['dst-unreach', 'echo-req', 'echo-rep', 'nbr-solicit', 'nbr-advert', 'redirect', 'time-exceeded', 'unspecified'])
            validating.values(row_num, ws, 'TCP_Session_Rules', templateVars['TCP_Session_Rules'], ['ack', 'est', 'fin', 'rst', 'syn', 'unspecified'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        if templateVars['TCP_Session_Rules'] == 'unspecified':
            templateVars['TCP_Session_Rules'] = None

        # Define the Template Source
        template_file = "contract_filter_entry.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Contract_Filter_%s.tf' % (templateVars['Filter'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def mgmt_epg(self, wb, ws, row_num, **kwargs):
        # Dicts for Bridge Domain required and optional args
        required_args = {'Site_Group': '',
                         'Type': '',
                         'EPG': '',
                         'QoS_Class': ''}
        optional_args = {'Tags': '',
                         'VLAN': '',
                         'Bridge_Domain': '',
                         'Tenant': '',
                         'consumed_Contracts': '',
                         'provided_Contracts': '',
                         'match_t': '',
                         'Contract_Interfaces': '',
                         'Taboo_Contracts': '',
                         'Subnets': '',
                         'Static_Routes': '',}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        templateVars['consume_count'] = 1
        templateVars['provide_count'] = 1
        templateVars['interface_count'] = 1
        templateVars['taboo_count'] = 1
        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'EPG', templateVars['EPG'])
            validating.qos_priority(row_num, ws, 'QoS_Class', templateVars['QoS_Class'])
            validating.values(row_num, ws, 'Type', templateVars['Type'], ['in_band', 'out_of_band'])
            if templateVars['Type'] == 'in_band':
                validating.vlans(row_num, ws, 'VLAN', templateVars['VLAN'])
                validating.name_rule(row_num, ws, 'Bridge_Domain', templateVars['Bridge_Domain'])
            if not templateVars['Tenant'] == None:
                validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
                if re.search(',', templateVars['provided_Contracts']):
                    templateVars['provide_count'] =+ 1
                    for x in templateVars['provided_Contracts'].split(','):
                        validating.name_rule(row_num, ws, 'provided_Contracts', x)
                else:
                    validating.name_rule(row_num, ws, 'provided_Contracts', templateVars['provided_Contracts'])
                if templateVars['Type'] == 'in_band':
                    if not templateVars['consumed_Contracts'] == None:
                        if re.search(',', templateVars['consumed_Contracts']):
                            templateVars['provide_count'] =+ 1
                            for x in templateVars['consumed_Contracts'].split(','):
                                validating.name_rule(row_num, ws, 'consumed_Contracts', x)
                        else:
                            validating.name_rule(row_num, ws, 'consumed_Contracts', templateVars['consumed_Contracts'])
                    if not templateVars['Contract_Interfaces'] == None:
                        if re.search(',', templateVars['Contract_Interfaces']):
                            for x in templateVars['Contract_Interfaces'].split(','):
                                templateVars['interface_count'] =+ 1
                                validating.not_empty(row_num, ws, 'Contract_Interfaces', x)
                        else:
                            validating.not_empty(row_num, ws, 'Contract_Interfaces', templateVars['Contract_Interfaces'])
                    if not templateVars['Taboo_Contracts'] == None:
                        if re.search(',', templateVars['Taboo_Contracts']):
                            templateVars['taboo_count'] =+ 1
                            for x in templateVars['Taboo_Contracts'].split(','):
                                validating.not_empty(row_num, ws, 'Taboo_Contracts', x)
                        else:
                            validating.not_empty(row_num, ws, 'Taboo_Contracts', templateVars['Taboo_Contracts'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        if not templateVars['Tenant'] == 'mgmt':
            dest_dir = 'Tenant_mgmt'

            template_file = 'data_tenant.jinja2'
            template = self.templateEnv.get_template(template_file)
            dest_file = 'data_Tenant_%s.tf' % (templateVars['Tenant'])
            process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

            templateVars['Contract_Type'] = 'Standard'
            if not templateVars['consumed_Contracts'] == None:
                if re.search(',', templateVars['consumed_Contracts']):
                    for x in templateVars['consumed_Contracts'].split(','):
                        templateVars['Contract'] = x
                        template_file = 'data_contract.jinja2'
                        template = self.templateEnv.get_template(template_file)
                        dest_file = 'data_Contract_Type_%s_%s.tf' % (templateVars['Contract_Type'], x)
                        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)
                else:
                    templateVars['Contract'] = templateVars['consumed_Contracts']
                    template_file = 'data_contract.jinja2'
                    template = self.templateEnv.get_template(template_file)
                    dest_file = 'data_Contract_Type_%s_%s.tf' % (templateVars['Contract_Type'], templateVars['consumed_Contracts'])
                    process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

            if not templateVars['Contract_Interfaces'] == None:
                if re.search(',', templateVars['Contract_Interfaces']):
                    for x in templateVars['Contract_Interfaces'].split(','):
                        templateVars['Contract'] = x
                        template_file = 'data_contract.jinja2'
                        template = self.templateEnv.get_template(template_file)
                        dest_file = 'data_Contract_Type_%s_%s.tf' % (templateVars['Contract_Type'], x)
                        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)
                else:
                    templateVars['Contract'] = templateVars['Contract_Interfaces']
                    template_file = 'data_contract.jinja2'
                    template = self.templateEnv.get_template(template_file)
                    dest_file = 'data_Contract_Type_%s_%s.tf' % (templateVars['Contract_Type'], templateVars['Contract_Interfaces'])
                    process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

            if not templateVars['provided_Contracts'] == None:
                if templateVars['Type'] == 'in_band':
                    if re.search(',', templateVars['provided_Contracts']):
                        for x in templateVars['provided_Contracts'].split(','):
                            templateVars['Contract'] = x
                            template_file = 'data_contract.jinja2'
                            template = self.templateEnv.get_template(template_file)
                            dest_file = 'data_Contract_Type_%s_%s.tf' % (templateVars['Contract_Type'], x)
                            process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)
                    else:
                        templateVars['Contract'] = templateVars['provided_Contracts']
                        template_file = 'data_contract.jinja2'
                        template = self.templateEnv.get_template(template_file)
                        dest_file = 'data_Contract_Type_%s_%s.tf' % (templateVars['Contract_Type'], templateVars['provided_Contracts'])
                        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

            templateVars['Contract_Type'] = 'Taboo'
            if not templateVars['Taboo_Contracts'] == None:
                if re.search(',', templateVars['Taboo_Contracts']):
                    for x in templateVars['Taboo_Contracts'].split(','):
                        templateVars['Contract'] == x
                        template_file = 'data_contract_taboo.jinja2'
                        template = self.templateEnv.get_template(template_file)
                        dest_file = 'data_Contract_Type_%s_%s.tf' % (templateVars['Contract_Type'], x)
                        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)
                else:
                    templateVars['Contract'] = templateVars['Taboo_Contracts']
                    template_file = 'data_contract_taboo.jinja2'
                    template = self.templateEnv.get_template(template_file)
                    dest_file = 'data_Contract_Type_%s_%s.tf' % (templateVars['Contract_Type'], templateVars['Contract_Type'])
                    process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source and Destination File
        template_file = "epg_mgmt.jinja2"
        template = self.templateEnv.get_template(template_file)
        dest_file = 'EPG_Mgmt_Type_%s_EPG_%s.tf' % (templateVars['Type'], templateVars['EPG'])

        # Process the template through the Sites
        dest_dir = 'Tenant_mgmt'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source and Destination File
        template_file = "var_mgmt.jinja2"
        template = self.templateEnv.get_template(template_file)

        if templateVars['Type'] == 'in_band':
            templateVars['var_name'] = 'in_band'
            dest_file = 'var_Mgmt_EPG_%s.tf' % ('inb')
        else:
            templateVars['var_name'] = 'out_of_band'
            dest_file = 'var_Mgmt_EPG_%s.tf' % ('oob')

        # Process the template through the Sites
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)
        dest_dir = 'Admin'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def node_intf(self, wb, ws, row_num, **kwargs):
        # Open the Network Policies Worksheet
        ws_net = wb['Network Policies']
        rows = ws_net.max_row

        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Tenant': '',
                         'L3Out': '',
                         'Node_Profile': '',
                         'Interface_Profile': '',
                         'QoS_Class': '',
                         'Node_Intf_Policy': '',
                         'Policy_Name': '',
                         'tag': ''}
        optional_args = {'Description': '',
                         'Alias': '',
                         'EIGRP_Intf_Profile': '',
                         'OSPF_Intf_Profile': '',
                         'ndIfPol': '',
                         'egress_qosDppPol': '',
                         'ingress_qosDppPol': '',
                         'qosCustomPol': '',
                         'igmp_policy': '',
                         'netflowMonitorPol': ''}

        # Get the Node Policies from the Network Policies Tab
        func = 'node_intf'
        count = countKeys(ws_net, func)
        row_node = ''
        var_dict = findVars(ws_net, func, rows, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('Node_Intf_Policy'):
                row_node = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
            validating.name_rule(row_num, ws, 'L3Out', templateVars['L3Out'])
            validating.name_rule(row_num, ws, 'Node_Profile', templateVars['Node_Profile'])
            validating.name_rule(row_num, ws, 'Interface_Profile', templateVars['Interface_Profile'])
            validating.qos_priority(row_num, ws, 'QoS_Class', templateVars['QoS_Class'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
            if not templateVars['EIGRP_Intf_Profile'] == None:
                validating.name_rule(row_num, ws, 'EIGRP_Intf_Profile', templateVars['EIGRP_Intf_Profile'])
            if not templateVars['OSPF_Intf_Profile'] == None:
                validating.name_rule(row_num, ws, 'OSPF_Intf_Profile', templateVars['OSPF_Intf_Profile'])
            if not templateVars['tag'] == None:
                validating.tag_check(row_node, ws_net, 'tag', templateVars['tag'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "logical_interface_profile.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'L3Out_%s_Node_Profile_%s.tf' % (templateVars['L3Out'], templateVars['Node_Profile'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

        #--------------------------
        # OSPF Interface Profile
        #--------------------------
        if not templateVars['OSPF_Intf_Profile'] == None:

            # Dicts for required and optional args
                             # OSPF Interface Profile
            required_args = {'Site_Group': '',
                             'Tenant': '',
                             'L3Out': '',
                             'Node_Profile': '',
                             'Interface_Profile': '',
                             'Auth_Type': '',
                             'Interface_Policy_Name': '',
                             # OSPF Interface Policy
                             'OSPF_Policy_Name': '',
                             'Policy_Tenant': '',
                             'Network_Type': '',
                             'Priority': '',
                             'Cost': '',
                             'Advertise_Subnet': '',
                             'BFD': '',
                             'MTU_Ignore': '',
                             'Passive_Interface': '',
                             'Hello_Interval': '',
                             'Dead_Interval': '',
                             'Retransmit_Interval': '',
                             'Transmit_Delay': ''}
                             # OSPF Interface Profile
            optional_args = {'Auth_Key_ID': '',
                             'OSPF_Auth_Key': '',
                             'Description': '',
                             # OSPF Interface Policy
                             'OSPF_Description': ''}

            # Get the OSPF Profile Attributes from the Network Policies Tab
            func = 'ospf_profile'
            count = countKeys(ws_net, func)
            row_ospf = ''
            var_dict = findVars(ws_net, func, rows, count)
            for pos in var_dict:
                if var_dict[pos].get('Policy_Name') == kwargs.get('OSPF_Intf_Profile'):
                    row_ospf = var_dict[pos]['row']
                    del var_dict[pos]['row']
                    kwargs = {**kwargs, **var_dict[pos]}

           # Get the OSPF Policy Attributes from the Network Policies Tab
            func = 'ospf_policy'
            count = countKeys(ws_net, func)
            row_intf = ''
            var_dict = findVars(ws_net, func, rows, count)
            for pos in var_dict:
                if var_dict[pos].get('OSPF_Policy_Name') == kwargs.get('Interface_Policy_Name'):
                    row_intf = var_dict[pos]['row']
                    del var_dict[pos]['row']
                    kwargs = {**kwargs, **var_dict[pos]}
            # print(kwargs)
            # Validate inputs, return dict of template vars
            templateVars = process_kwargs(required_args, optional_args, **kwargs)

            try:
                # Validate OSPF Profile Required Arguments
                validating.values(row_ospf, ws_net, 'Auth_Type', templateVars['Auth_Type'], ['md5', 'none', 'simple'])
                if not templateVars['Auth_Key_ID'] == None:
                    validating.number_check(row_ospf, ws_net, 'Auth_Key_ID', templateVars['Auth_Key_ID'], 1, 255)
                if not templateVars['Description'] == None:
                    validating.description(row_ospf, ws_net, 'Description', templateVars['Description'])
                # Validate OSPF Policy Required Arguments
                validating.name_rule(row_intf, ws_net, 'OSPF_Policy_Name', templateVars['OSPF_Policy_Name'])
                validating.number_check(row_intf, ws_net, 'Priority', templateVars['Priority'], 0, 255)
                validating.number_check(row_intf, ws_net, 'Cost', templateVars['Cost'], 0, 65535)
                validating.number_check(row_intf, ws_net, 'Hello_Interval', templateVars['Hello_Interval'], 1, 65535)
                validating.number_check(row_intf, ws_net, 'Dead_Interval', templateVars['Dead_Interval'], 1, 65535)
                validating.number_check(row_intf, ws_net, 'Retransmit_Interval', templateVars['Retransmit_Interval'], 1, 65535)
                validating.number_check(row_intf, ws_net, 'Transmit_Delay', templateVars['Transmit_Delay'], 1, 65535)
                validating.values(row_intf, ws_net, 'Network_Type', templateVars['Network_Type'], ['broadcast', 'p2p', 'unspecified'])
                validating.values(row_intf, ws_net, 'Advertise_Subnet', templateVars['Advertise_Subnet'], ['no', 'yes'])
                validating.values(row_intf, ws_net, 'BFD', templateVars['BFD'], ['no', 'yes'])
                validating.values(row_intf, ws_net, 'MTU_Ignore', templateVars['MTU_Ignore'], ['no', 'yes'])
                validating.values(row_intf, ws_net, 'Passive_Interface', templateVars['Passive_Interface'], ['no', 'yes'])
                if not templateVars['OSPF_Description'] == None:
                    validating.description(row_intf, ws_net, 'OSPF_Description', templateVars['OSPF_Description'])
            except Exception as err:
                Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws_net, row_ospf)
                raise ErrException(Error_Return)

            if templateVars['Auth_Type'] == 'none':
                templateVars['Auth_Key_ID'] = 1

            # Define the Template Source
            template_file = "l3out_ospf_interface_profile.jinja2"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = 'L3Out_%s_Node_Profile_%s.tf' % (templateVars['L3Out'], templateVars['Node_Profile'])
            dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
            process_method(wb, ws_net, row_ospf, 'a+', dest_dir, dest_file, template, **templateVars)

            if not templateVars['OSPF_Auth_Key'] == None:
                x = templateVars['OSPF_Auth_Key'].split('r')
                key_number = x[1]
                templateVars['sensitive_var'] = 'OSPF_Auth_Key%s' % (key_number)

                # Define the Template Source
                template_file = "variables.jinja2"
                template = self.templateEnv.get_template(template_file)

                # Process the template through the Sites
                dest_file = 'variable_%s.tf' % (templateVars['sensitive_var'])
                dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
                process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

                process_sensitive_var(wb, ws, row_num, dest_dir, dest_file, template, **templateVars)

            #--------------------------
            # OSPF Interface Policy
            #--------------------------

            ctrl_count = 0
            Ctrl = ''
            if templateVars['Advertise_Subnet'] == 'yes':
                Ctrl = '"advert-subnet"'
                ctrl_count =+ 1
            if templateVars['BFD'] == 'yes' and ctrl_count > 0:
                Ctrl = Ctrl + ', ' + '"bfd"'
                ctrl_count =+ 1
            elif templateVars['BFD'] == 'yes':
                Ctrl = '"bfd"'
                ctrl_count =+ 1
            if templateVars['MTU_Ignore'] == 'yes' and ctrl_count > 0:
                Ctrl = Ctrl + ', ' + '"mtu-ignore"'
                ctrl_count =+ 1
            elif templateVars['MTU_Ignore'] == 'yes':
                Ctrl = '"mtu-ignore"'
                ctrl_count =+ 1
            if templateVars['Passive_Interface'] == 'yes' and ctrl_count > 0:
                Ctrl = Ctrl + ', ' + '"passive"'
                ctrl_count =+ 1
            elif templateVars['Passive_Interface'] == 'yes':
                Ctrl = '"passive"'
                ctrl_count =+ 1
            if ctrl_count > 0:
                templateVars['Ctrl'] = '[%s]' % (Ctrl)
            else:
                templateVars['Ctrl'] = '[unspecified]'

            if templateVars['Cost'] == 0:
                templateVars['Cost'] = 'unspecified'

            # Define the Template Source
            template_file = "ospf_interface_policy.jinja2"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = 'Policies_OSPF_Interface_%s.tf' % (templateVars['OSPF_Policy_Name'])
            dest_dir = 'Tenant_%s' % (templateVars['Policy_Tenant'])
            process_method(wb, ws_net, row_ospf, 'w', dest_dir, dest_file, template, **templateVars)

            if not templateVars['Tenant'] == templateVars['Policy_Tenant']:
                # Define the Template Source
                template_file = "data_ospf_interface_policy.jinja2"
                template = self.templateEnv.get_template(template_file)

                # Process the template through the Sites
                dest_file = 'data_Policies_OSPF_Interface_%s.tf' % (templateVars['OSPF_Policy_Name'])
                dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
                process_method(wb, ws_net, row_ospf, 'w', dest_dir, dest_file, template, **templateVars)

                templateVars['L3Out_Tenant'] = templateVars['Tenant']
                templateVars['Tenant'] = templateVars['Policy_Tenant']
                # Define the Template Source
                template_file = "data_tenant.jinja2"
                template = self.templateEnv.get_template(template_file)

                # Process the template through the Sites
                dest_file = 'data_Tenant_%s.tf' % (templateVars['Tenant'])
                dest_dir = 'Tenant_%s' % (templateVars['L3Out_Tenant'])
                process_method(wb, ws_net, row_ospf, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def node_path(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
                         # Logical Interface Profile
        required_args = {'Site_Group': '',
                         'Path_Policy_Name': '',
                         'MTU': '',
                         'Target_DSCP': '',
                         'SideA_Address': '',
                         'SideA_IPv6_DAD': '',
                         # L3Out Path Profile
                         'Tenant': '',
                         'L3Out': '',
                         'Policy_Name': '',
                         'Node_Profile': '',
                         'Interface_Profile': '',
                         'Interface_Type': '',
                         'Pod_ID': '',
                         'Node1_ID': '',
                         'Interface_or_PG': ''}
                         # Logical Interface Profile
        optional_args = {'Encap_Scope': '',
                         'Mode': '',
                         'VLAN': '',
                         'Description': '',
                         'Auto_State': '',
                         'MAC_Address': '',
                         'SideA_Secondary': '',
                         'SideA_Link_Local': '',
                         'SideB_Address': '',
                         'SideB_IPv6_DAD': '',
                         'SideB_Secondary': '',
                         'SideB_Link_Local': '',
                         'Node2_ID': ''}

        # Get the Node Policies from the Network Policies Tab
        rows = ws.max_row
        func = 'l3out_path'
        count = countKeys(ws, func)
        row_path = ''
        var_dict = findVars(ws, func, rows, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('Path_Policy_Name'):
                row_path = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments Logincal Interface Profile
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_path, ws, 'Path_Policy_Name', templateVars['Path_Policy_Name'])
            if not templateVars['Encap_Scope'] == None:
                validating.values(row_num, ws, 'Encap_Scope', templateVars['Encap_Scope'], ['ctx', 'local'])
            if not templateVars['Mode'] == None:
                validating.values(row_num, ws, 'Mode', templateVars['Mode'], ['native', 'regular', 'untagged'])
            if not templateVars['VLAN'] == None:
                validating.vlans(row_num, ws, 'VLAN', templateVars['VLAN'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
            if not templateVars['Auto_State'] == None:
                validating.values(row_num, ws, 'Auto_State', templateVars['Auto_State'], ['disabled', 'enabled'])
            if not templateVars['MTU'] == 'inherit':
                validating.number_check(row_path, ws, 'MTU', templateVars['MTU'], 1300, 9216)
            if not templateVars['MAC_Address'] == None:
                validating.mac_address(row_num, ws, 'MAC_Address', templateVars['MAC_Address'])
            validating.dscp(row_num, ws, 'Target_DSCP', templateVars['Target_DSCP'])
            validating.ip_address(row_num, ws, 'SideA_Address', templateVars['SideA_Address'])
            validating.values(row_num, ws, 'SideA_IPv6_DAD', templateVars['SideA_IPv6_DAD'], ['disabled', 'enabled'])
            if not templateVars['SideA_Secondary'] == None:
                validating.ip_address(row_num, ws, 'SideA_Secondary', templateVars['SideA_Secondary'])
            if not templateVars['SideA_Link_Local'] == None:
                validating.ip_address(row_num, ws, 'SideA_Link_Local', templateVars['SideA_Link_Local'])
            if not templateVars['SideB_Address'] == None:
                validating.ip_address(row_num, ws, 'SideB_Address', templateVars['SideB_Address'])
            if not templateVars['SideB_Address'] == None:
                validating.values(row_num, ws, 'SideB_IPv6_DAD', templateVars['SideB_IPv6_DAD'], ['disabled', 'enabled'])
            if not templateVars['SideB_Secondary'] == None:
                validating.ip_address(row_num, ws, 'SideB_Secondary', templateVars['SideB_Secondary'])
            if not templateVars['SideB_Link_Local'] == None:
                validating.ip_address(row_num, ws, 'SideB_Link_Local', templateVars['SideB_Link_Local'])
            validating.name_rule(row_path, ws, 'Tenant', templateVars['Tenant'])
            validating.name_rule(row_path, ws, 'L3Out', templateVars['L3Out'])
            validating.name_rule(row_path, ws, 'Node_Profile', templateVars['Node_Profile'])
            validating.name_rule(row_path, ws, 'Interface_Profile', templateVars['Interface_Profile'])
            validating.values(row_path, ws, 'Interface_Type', templateVars['Interface_Type'], ['ext-svi', 'l3-port', 'sub-interface'])
            validating.number_check(row_path, ws, 'Pod_ID', templateVars['Pod_ID'], 1, 15)
            validating.number_check(row_path, ws, 'Node1_ID', templateVars['Node1_ID'], 101, 4001)
            if not templateVars['Node2_ID'] == None:
                validating.number_check(row_path, ws, 'Node2_ID', templateVars['Node2_ID'], 101, 4001)
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Create Global Variables for First Template
        if not templateVars['Node2_ID'] == None:
            templateVars['Address'] = None
            templateVars['IPv6_DAD'] = templateVars['SideA_IPv6_DAD']
            templateVars['Link_Local'] = None
            templateVars['PATH'] = 'protpaths-%s-%s' % (templateVars['Node1_ID'], templateVars['Node2_ID'])
        else:
            templateVars['Address'] = templateVars['SideA_Address']
            templateVars['IPv6_DAD'] = templateVars['SideA_IPv6_DAD']
            templateVars['Link_Local'] = templateVars['SideA_Link_Local']
            templateVars['PATH'] = 'paths-%s' % (templateVars['Node1_ID'])
        # Define the Template Source
        template_file = "l3out_path_attachment.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'L3Out_%s_Node_Profile_%s.tf' % (templateVars['L3Out'], templateVars['Node_Profile'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

        # Create Global Variables for First Template
        if not templateVars['Node2_ID'] == None:
            templateVars['PATH'] = 'protpaths-%s-%s' % (templateVars['Node1_ID'], templateVars['Node2_ID'])

            templateVars['Address'] = templateVars['SideA_Address']
            templateVars['IPv6_DAD'] = templateVars['SideA_IPv6_DAD']
            templateVars['Link_Local'] = templateVars['SideA_Link_Local']
            templateVars['Side'] = 'A'

            # Define the Template Source
            template_file = "l3out_path_attach_vpc.jinja2"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = 'L3Out_%s_Node_Profile_%s.tf' % (templateVars['L3Out'], templateVars['Node_Profile'])
            dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
            process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

            templateVars['Address'] = templateVars['SideB_Address']
            templateVars['IPv6_DAD'] = templateVars['SideB_IPv6_DAD']
            templateVars['Link_Local'] = templateVars['SideB_Link_Local']
            templateVars['Side'] = 'B'

            # Define the Template Source
            template_file = "l3out_path_attach_vpc.jinja2"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = 'L3Out_%s_Node_Profile_%s.tf' % (templateVars['L3Out'], templateVars['Node_Profile'])
            dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
            process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        if not templateVars['SideA_Secondary'] == None:
            if templateVars['Node2_ID'] == None:
                templateVars['Secondary'] = templateVars['SideA_Secondary']
                templateVars['IPv6_DAD'] = templateVars['SideA_IPv6_DAD']
                templateVars['PATH'] = 'paths-%s' % (templateVars['Node1_ID'])
                template_file = "l3out_path_attach_secondary.jinja2"
                template = self.templateEnv.get_template(template_file)

                # Process the template through the Sites
                dest_file = 'L3Out_%s_Node_Profile_%s.tf' % (templateVars['L3Out'], templateVars['Node_Profile'])
                dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
                process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def node_prof(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Tenant': '',
                         'L3Out': '',
                         'Node_Profile': '',
                         'Target_DSCP': '',
                         'Color_Tag': '',
                         'Pod_ID': '',
                         'Node1_ID': '',
                         'Node1_Router_ID': '',
                         'Node1_Loopback': ''}
        optional_args = {'Alias': '',
                         'Description': '',
                         'Node2_ID': '',
                         'Node2_Router_ID': '',
                         'Node2_Loopback': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
            validating.name_rule(row_num, ws, 'L3Out', templateVars['L3Out'])
            validating.name_rule(row_num, ws, 'Node_Profile', templateVars['Node_Profile'])
            validating.dscp(row_num, ws, 'Target_DSCP', templateVars['Target_DSCP'])
            validating.tag_check(row_num, ws, 'Color_Tag', templateVars['Color_Tag'])
            validating.number_check(row_num, ws, 'Node1_ID', templateVars['Node1_ID'], 101, 4001)
            validating.ip_address(row_num, ws, 'Node1_Router_ID', templateVars['Node1_Router_ID'])
            validating.values(row_num, ws, 'Node1_Loopback', templateVars['Node1_Loopback'], ['no', 'yes'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
            if not templateVars['Node2_ID'] == None:
                validating.number_check(row_num, ws, 'Node2_ID', templateVars['Node2_ID'], 101, 4001)
                validating.ip_address(row_num, ws, 'Node2_Router_ID', templateVars['Node2_Router_ID'])
                validating.values(row_num, ws, 'Node2_Loopback', templateVars['Node2_Loopback'], ['no', 'yes'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "logical_node_profile.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'L3Out_%s_Node_Profile_%s.tf' % (templateVars['L3Out'], templateVars['Node_Profile'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "logical_node_to_fabric_node.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Modify Variables for Template
        templateVars['Node_ID'] = templateVars['Node1_ID']
        templateVars['rtr_id'] = templateVars['Node1_Router_ID']
        templateVars['rtr_id_loopback'] = templateVars['Node1_Loopback']

        # Process the template through the Sites
        dest_file = 'L3Out_%s_Node_Profile_%s.tf' % (templateVars['L3Out'], templateVars['Node_Profile'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

        if not templateVars['Node2_ID'] == None:
            # Modify Variables for Template
            templateVars['Node_ID'] = templateVars['Node2_ID']
            templateVars['rtr_id'] = templateVars['Node2_Router_ID']
            templateVars['rtr_id_loop_back'] = templateVars['Node2_Loopback']

            # Process the template through the Sites
            dest_file = 'L3Out_%s_Node_Profile_%s.tf' % (templateVars['L3Out'], templateVars['Node_Profile'])
            dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
            process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def subject_add(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Tenant': '',
                         'Subject': '',
                         'Contract_Type': '',
                         'Contract': '',
                         'Reverse_Filter_Ports': '',
                         'QoS_Class': '',
                         'Target_DSCP': '',
                         'Filters_to_Assign': ''}
        optional_args = {'Description': '',
                         'Alias': '',
                         'Tags': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        templateVars['filters_count'] = 1
        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.dscp(row_num, ws, 'Target_DSCP', templateVars['Target_DSCP'])
            validating.name_rule(row_num, ws, 'Contract', templateVars['Contract'])
            validating.name_rule(row_num, ws, 'Subject', templateVars['Subject'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
            validating.qos_priority(row_num, ws, 'QoS_Class', templateVars['QoS_Class'])
            validating.values(row_num, ws, 'Contract_Type', templateVars['Contract_Type'], ['OOB', 'Standard', 'Taboo'])
            validating.values(row_num, ws, 'Reverse_Filter_Ports', templateVars['Reverse_Filter_Ports'], ['no', 'yes'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
            if not templateVars['Filters_to_Assign'] == None:
                if re.search(',', templateVars['Filters_to_Assign']):
                    templateVars['filters_count'] =+ 1
                    for x in templateVars['Filters_to_Assign'].split(','):
                        validating.name_rule(row_num, ws, 'Filters_to_Assign', x)
                else:
                    validating.name_rule(row_num, ws, 'Filters_to_Assign', templateVars['Filters_to_Assign'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "contract_subject.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Define the Template Source
        if templateVars['Contract_Type'] == 'OOB':
            template_file = "contract_subject.jinja2"
        elif templateVars['Contract_Type'] == 'Standard':
            template_file = "contract_subject.jinja2"
        elif templateVars['Contract_Type'] == 'Taboo':
            template_file = "contract_subject.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Contract_Type_%s_%s_Subj_%s.tf' % (templateVars['Contract_Type'], templateVars['Contract'], templateVars['Subject'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

# Terraform ACI Provider - Tenants Policies
# Class must be instantiated with Variables
class VMM_Policies(object):
    def __init__(self, ws):
        self.templateLoader = jinja2.FileSystemLoader(
            searchpath=(aci_template_path + 'Tenant_Policies/'))
        self.templateEnv = jinja2.Environment(loader=self.templateLoader)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def add_vmm(self, wb, ws, row_num, **kwargs):
        # Open the Network Policies Worksheet
        ws_net = wb['Network Policies']
        rows = ws_net.max_row

        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Tenant': '',
                         'VRF': '',
                         'Name': '',
                         'L3Out_Policy': '',
                         'L3_Domain': '',
                         'Ext_EPG': '',
                         'Ext_EPG_Policy': '',
                         'Subnet': '',
                         'Ext_Subnet_Policy': '',
                         'target_dscp': '',
                         'enforce_rtctrl': '',
                         'prio': '',
                         'epg_target_dscp': '',
                         'pref_gr_memb': '',
                         'match_t': '',
                         'flood': '',
                         'export-rtctrl': '',
                         'import-rtctrl': '',
                         'import-security': '',
                         'shared-security': '',
                         'shared-rtctrl': '',
                         'agg-export': '',
                         'agg-import': '',
                         'agg-shared': ''}
        optional_args = {'Description': '',
                         'EPG_Description': '',
                         'annotation': '',
                         'name_alias': '',
                         'leak_rtctrlProfile': '',
                         'damp_rtctrlProfile': '',
                         'fvBDPublicSubnetHolder': '',
                         'epg_annotation': '',
                         'epg_name_alias': '',
                         'cons_vzBrCP': '',
                         'vzCPIf': '',
                         'Master_fvEPg': '',
                         'prov_vzBrCP': '',
                         'vzTaboo': '',
                         'exception_tag': '',
                         'rtctrlProfile': '',
                         'sub_annotation': '',
                         'sub_name_alias': '',
                         'sub_rtctrlProfile': '',
                         'rtsumARtSummPol': ''}


        # Get the L3Out Policies from the Network Policies Tab
        func = 'L3Out_Policy'
        count = countKeys(ws_net, func)
        row_l3out = ''
        var_dict = findVars(ws_net, func, rows, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('L3Out_Policy'):
                row_l3out = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}

        func = 'Ext_EPG_Policy'
        count = countKeys(ws_net, func)
        row_epg = ''
        var_dict = findVars(ws_net, func, rows, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('Ext_EPG_Policy'):
                row_epg = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}

        func = 'Ext_Subnet_Policy'
        count = countKeys(ws_net, func)
        row_sub = ''
        var_dict = findVars(ws_net, func, rows, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('Ext_Subnet_Policy'):
                row_sub = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
            validating.name_rule(row_num, ws, 'VRF', templateVars['VRF'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            if not templateVars['Subnet'] == None:
                if re.search(',', templateVars['Subnet']):
                    sx = templateVars['Subnet'].split(',')
                    for x in sx:
                        validating.ip_address(row_num, ws, 'Subnet', x)
            validating.dscp(row_l3out, ws_net, 'target_dscp', templateVars['target_dscp'])
            validating.values(row_l3out, ws_net, 'enforce_rtctrl', templateVars['enforce_rtctrl'], ['export', 'export-import'])
            validating.dscp(row_epg, ws_net, 'epg_target_dscp', templateVars['epg_target_dscp'])
            validating.match_t(row_epg, ws_net, 'match_t', templateVars['match_t'])
            validating.qos_priority(row_epg, ws_net, 'prio', templateVars['prio'])
            validating.values(row_epg, ws_net, 'flood', templateVars['flood'], ['disabled', 'enabled'])
            validating.values(row_epg, ws_net, 'pref_gr_memb', templateVars['pref_gr_memb'], ['exclude', 'include'])
            validating.values(row_sub, ws_net, 'agg-export', templateVars['agg-export'], ['no', 'yes'])
            validating.values(row_sub, ws_net, 'agg-import', templateVars['agg-import'], ['no', 'yes'])
            validating.values(row_sub, ws_net, 'agg-shared', templateVars['agg-shared'], ['no', 'yes'])
            validating.values(row_sub, ws_net, 'export-rtctrl', templateVars['export-rtctrl'], ['no', 'yes'])
            validating.values(row_sub, ws_net, 'import-rtctrl', templateVars['import-rtctrl'], ['no', 'yes'])
            validating.values(row_sub, ws_net, 'import-security', templateVars['import-security'], ['no', 'yes'])
            validating.values(row_sub, ws_net, 'shared-security', templateVars['shared-security'], ['no', 'yes'])
            validating.values(row_sub, ws_net, 'shared-rtctrl', templateVars['shared-rtctrl'], ['no', 'yes'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Create aggregate templateVars
        aggregate = ''
        if templateVars['agg-export'] == 'yes':
            aggregate = aggregate + '"export-rtctrl"'
        if templateVars['agg-import'] == 'yes':
            aggregate = aggregate + ', ' + '"import-rtctrl"'
        if templateVars['agg-shared'] == 'yes':
            aggregate = aggregate + ', ' + '"shared-rtctrl"'

        else:
            templateVars['aggregate'] = '[%s]' % (aggregate)

        # Create scope templateVars
        scope = ''
        if templateVars['export-rtctrl'] == 'yes':
            scope = scope + '"export-rtctrl"'
        if templateVars['import-rtctrl'] == 'yes':
            scope = scope + ', ' + '"import-rtctrl"'
        if templateVars['import-security'] == 'yes':
            scope = scope + ', ' + '"import-security"'
        if templateVars['shared-security'] == 'yes':
            scope = scope + ', ' + '"shared-security"'
        if templateVars['shared-rtctrl'] == 'yes':
            scope = scope + ', ' + '"shared-rtctrl"'

        else:
            templateVars['scope'] = '[%s]' % (scope)

        # Define the Template Source
        template_file = "l3out.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'l3out_%s.tf' % (templateVars['Name'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "ext_epg.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'l3out_%s_epg_%s.tf' % (templateVars['Name'], templateVars['Ext_EPG'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        if re.search(',', templateVars['Subnet']):
            sx = templateVars['Subnet'].split(',')
            for x in sx:
                templateVars['Subnet'] = x
                templateVars['Subnet_'] = x.replace('.', '-')
                templateVars['Subnet_'] = x.replace('/', '_')

                # Define the Template Source
                template_file = "ext_subnet.jinja2"
                template = self.templateEnv.get_template(template_file)

                # Process the template through the Sites
                dest_file = 'l3out_%s_epg_%s_subnet_%s.tf' % (templateVars['Name'], templateVars['Ext_EPG'], templateVars['Subnet'])
                dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
                process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)
        else:
            templateVars['Subnet_'] = templateVars['Subnet'].replace('.', '-')
            templateVars['Subnet_'] = templateVars['Subnet'].replace('/', '_')

            # Define the Template Source
            template_file = "ext_subnet.jinja2"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = 'l3out_%s_epg_%s_subnet_%s.tf' % (templateVars['Name'], templateVars['Ext_EPG'], templateVars['Subnet'])
            dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
            process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

# Function to Count the Number of Keys
def countKeys(ws, func):
    count = 0
    for i in ws.rows:
        if any(i):
            if str(i[0].value) == func:
                count += 1
    return count

# Function to Create Interface Selectors
def create_selector(ws_sw, ws_sw_row_count, **templateVars):
    print(templateVars['port_count'])
    Port_Selector = ''
    for port in range(1, int(templateVars['port_count']) + 1):
        if port < 10:
            Port_Selector = 'Eth%s-0%s' % (templateVars['module'], port)
        elif port < 100:
            Port_Selector = 'Eth%s-%s' % (templateVars['module'], port)
        elif port > 99:
            Port_Selector = 'Eth%s_%s' % (templateVars['module'], port)
        modp = '%s/%s' % (templateVars['module'],port)
        # Copy the Port Selector to the Worksheet
        data = ['intf_selector',templateVars['Pod_ID'],templateVars['Node_ID'],templateVars['Name'],Port_Selector,modp,'','','','','','']
        ws_sw.append(data)
        rc = '%s:%s' % (ws_sw_row_count, ws_sw_row_count)
        for cell in ws_sw[rc]:
            if ws_sw_row_count % 2 == 0:
                cell.style = 'ws_odd'
            else:
                cell.style = 'ws_even'
        dv1_cell = 'A%s' % (ws_sw_row_count)
        dv2_cell = 'H%s' % (ws_sw_row_count)
        templateVars['dv1'].add(dv1_cell)
        templateVars['dv2'].add(dv2_cell)
        ws_sw_row_count += 1
    return ws_sw_row_count

# Function to Create Static Paths within EPGs
def create_static_paths(wb, wb_sw, row_num, wr_method, dest_dir, dest_file, template, **templateVars):
    wsheets = wb_sw.get_sheet_names()
    tf_file = ''
    for wsheet in wsheets:
        ws = wb_sw[wsheet]
        for row in ws.rows:
            if not (row[12].value == None or row[13].value == None):
                vlan_test = ''
                if re.search('^(individual|port-channel|vpc)$', row[7].value) and (re.search(r'\d+', str(row[12].value)) or re.search(r'\d+', str(row[13].value))):
                    if not row[12].value == None:
                        vlan = row[12].value
                        vlan_test = vlan_range(vlan, **templateVars)
                        if 'true' in vlan_test:
                            templateVars['mode'] = 'native'
                    if not 'true' in vlan_test:
                        templateVars['mode'] = 'regular'
                        if not row[13].value == None:
                            vlans = row[13].value
                            vlan_test = vlan_range(vlans, **templateVars)
                if vlan_test == 'true':
                    templateVars['Pod_ID'] = row[1].value
                    templateVars['Node_ID'] = row[2].value
                    templateVars['Interface_Profile'] = row[3].value
                    templateVars['Interface_Selector'] = row[4].value
                    templateVars['Port'] = row[5].value
                    templateVars['Policy_Group'] = row[6].value
                    templateVars['Port_Type'] = row[7].value
                    templateVars['Bundle_ID'] = row[9].value
                    Site_Group = templateVars['Site_Group']
                    pod = templateVars['Pod_ID']
                    node_id =  templateVars['Node_ID']
                    if templateVars['Port_Type'] == 'vpc':
                        ws_vpc = wb['Inventory']
                        for rx in ws_vpc.rows:
                            if rx[0].value == 'vpc_pair' and int(rx[1].value) == int(Site_Group) and str(rx[4].value) == str(node_id):
                                node1 = templateVars['Node_ID']
                                node2 = rx[5].value
                                templateVars['Policy_Group'] = '%s_vpc%s' % (row[3].value, templateVars['Bundle_ID'])
                                templateVars['tDn'] = 'topology/pod-%s/protpaths-%s-%s/pathep-[%s]' % (pod, node1, node2, templateVars['Policy_Group'])
                                templateVars['Static_Path'] = 'rspathAtt-[topology/pod-%s/protpaths-%s-%s/pathep-[%s]' % (pod, node1, node2, templateVars['Policy_Group'])
                                templateVars['GUI_Static'] = 'Pod-%s/Node-%s-%s/%s' % (pod, node1, node2, templateVars['Policy_Group'])
                                templateVars['Static_descr'] = 'Pod-%s_Nodes-%s-%s_%s' % (pod, node1, node2, templateVars['Policy_Group'])
                                tf_file = './ACI/%s/%s/%s' % (templateVars['Site_Name'], dest_dir, dest_file)
                                read_file = open(tf_file, 'r')
                                read_file.seek(0)
                                static_path_descr = 'resource "aci_epg_to_static_path" "%s_%s_%s"' % (templateVars['App_Profile'], templateVars['EPG'], templateVars['Static_descr'])
                                if not static_path_descr in read_file.read():
                                    create_tf_file(wr_method, dest_dir, dest_file, template, **templateVars)

                    elif templateVars['Port_Type'] == 'port-channel':
                        templateVars['Policy_Group'] = '%s_pc%s' % (row[3].value, templateVars['Bundle_ID'])
                        templateVars['tDn'] = 'topology/pod-%s/paths-%s/pathep-[%s]' % (pod, templateVars['Node_ID'], templateVars['Policy_Group'])
                        templateVars['Static_Path'] = 'rspathAtt-[topology/pod-%s/paths-%s/pathep-[%s]' % (pod, templateVars['Node_ID'], templateVars['Policy_Group'])
                        templateVars['GUI_Static'] = 'Pod-%s/Node-%s/%s' % (pod, templateVars['Node_ID'], templateVars['Policy_Group'])
                        templateVars['Static_descr'] = 'Pod-%s_Node-%s_%s' % (pod, templateVars['Node_ID'], templateVars['Policy_Group'])
                        read_file = open(tf_file, 'r')
                        read_file.seek(0)
                        static_path_descr = 'resource "aci_epg_to_static_path" "%s_%s_%s"' % (templateVars['App_Profile'], templateVars['EPG'], templateVars['Static_descr'])
                        if not static_path_descr in read_file.read():
                            create_tf_file(wr_method, dest_dir, dest_file, template, **templateVars)

                    elif templateVars['Port_Type'] == 'individual':
                        port = 'eth%s' % (templateVars['Port'])
                        templateVars['tDn'] = 'topology/pod-%s/paths-%s/pathep-[%s]' % (pod, templateVars['Node_ID'], port)
                        templateVars['Static_Path'] = 'rspathAtt-[topology/pod-%s/paths-%s/pathep-[%s]' % (pod, templateVars['Node_ID'], port)
                        templateVars['GUI_Static'] = 'Pod-%s/Node-%s/%s' % (pod, templateVars['Node_ID'], port)
                        templateVars['Static_descr'] = 'Pod-%s_Node-%s_%s' % (pod, templateVars['Node_ID'], templateVars['Interface_Selector'])
                        read_file = open(tf_file, 'r')
                        read_file.seek(0)
                        static_path_descr = 'resource "aci_epg_to_static_path" "%s_%s_%s"' % (templateVars['App_Profile'], templateVars['EPG'], templateVars['Static_descr'])
                        if not static_path_descr in read_file.read():
                            create_tf_file(wr_method, dest_dir, dest_file, template, **templateVars)

# Function to Create Destination Files
def create_file(wb, ws, row_num, wr_method, dest_dir, dest_file, **templateVars):
    if re.search('Grp_[A-F]', templateVars['Site_Group']):
        Group_ID = '%s' % (templateVars['Site_Group'])
        site_group = ast.literal_eval(os.environ[Group_ID])
        for x in range(1, 13):
            sitex = 'Site_%s' % (x)
            if not site_group[sitex] == None:
                Site_ID = 'Site_ID_%s' % (site_group[sitex])
                site_dict = ast.literal_eval(os.environ[Site_ID])

                # Create templateVars for Site_Name and APIC_URL
                templateVars['Site_Name'] = site_dict.get('Site_Name')
                templateVars['APIC_URL'] = site_dict.get('APIC_URL')

                # Create Terraform file from Template
                tf_file = './ACI/%s/%s/%s' % (templateVars['Site_Name'], dest_dir, dest_file)
                wr_file = open(tf_file, wr_method)
                wr_file.close()

    elif re.search(r'\d+', templateVars['Site_Group']):
        Site_ID = 'Site_ID_%s' % (templateVars['Site_Group'])
        site_dict = ast.literal_eval(os.environ[Site_ID])

        # Create templateVars for Site_Name and APIC_URL
        templateVars['Site_Name'] = site_dict.get('Site_Name')
        templateVars['APIC_URL'] = site_dict.get('APIC_URL')

        # Create Terraform file from Template
        tf_file = './ACI/%s/%s/%s' % (templateVars['Site_Name'], dest_dir, dest_file)
        wr_file = open(tf_file, wr_method)
        wr_file.close()
    else:
        print(f"\n-----------------------------------------------------------------------------\n")
        print(f"   Error on Worksheet {ws.title}, Row {row_num} Site_Group, value {templateVars['Site_Group']}.")
        print(f"   Unable to Determine if this is a Single or Group of Site(s).  Exiting....")
        print(f"\n-----------------------------------------------------------------------------\n")
        exit()

# Function to Create Terraform Files
def create_tf_file(wr_method, dest_dir, dest_file, template, **templateVars):
    # Make sure the Destination Directory Exists
    dest_dir = './ACI/%s/%s' % (templateVars['Site_Name'], dest_dir)
    if not os.path.isdir(dest_dir):
        mk_dir = 'mkdir -p %s' % (dest_dir)
        os.system(mk_dir)
    # Create File for the Template in the Destination Folder
    tf_file = '%s/%s' % (dest_dir, dest_file)
    wr_file = open(tf_file, wr_method)

    # Render Payload and Write to File
    payload = template.render(templateVars)
    wr_file.write(payload + '\n\n')
    wr_file.close()

# Function to find the Keys for each Section
def findKeys(ws, func_regex):
    func_list = OrderedSet()
    for i in ws.rows:
        if any(i):
            if re.search(func_regex, str(i[0].value)):
                func_list.add(str(i[0].value))
    return func_list

# Function to Assign the Variables to the Keys
def findVars(ws, func, rows, count):
    var_list = []
    var_dict = {}
    for i in range(1, rows + 1):
        if (ws.cell(row=i, column=1)).value == func:
            try:
                for x in range(2, 34):
                    if (ws.cell(row=i - 1, column=x)).value:
                        var_list.append(str(ws.cell(row=i - 1, column=x).value))
                    else:
                        x += 1
            except Exception as e:
                e = e
                pass
            break
    vcount = 1
    while vcount <= count:
        var_dict[vcount] = {}
        var_count = 0
        for z in var_list:
            var_dict[vcount][z] = ws.cell(row=i + vcount - 1, column=2 + var_count).value
            var_count += 1
        var_dict[vcount]['row'] = i + vcount - 1
        vcount += 1
    return var_dict

# Function to execute HTTP Post
def post(uri, oauth_token, payload, section=''):
    # Use this for Troubleshooting
    if print_payload:
        print(payload)

    tf_token = 'Bearer %s' % (oauth_token)
    MyHeaders = {'Authorization': tf_token,
               'Content-Type': 'application/vnd.api+json'
    }

    r = ''
    while r == '':
        try:
            r = requests.post('https://app.terraform.io/api/v2/{}', data=json.dumps(payload), headers=MyHeaders)%(uri)
            status = r.status_code
        except requests.exceptions.ConnectionError as e:
            print("Connection error, pausing before retrying. Error: {}"
                  .format(e))
            time.sleep(5)
        except Exception as e:
            print("Method {} failed. Exception: {}".format(section[:-5], e))
            status = 666
            return(status)

    # Use this for Troubleshooting
    if print_response_always:
        print(r.text)
    if status != 200 and print_response_on_fail:
        print(r.text)

    return status

# Function to validate input for each method
def process_kwargs(required_args, optional_args, **kwargs):
    # Validate all required kwargs passed
    # if all(item in kwargs for item in required_args.keys()) is not True:
    #    error_ = '\n***ERROR***\nREQUIRED Argument Not Found in Input:\n "%s"\nInsufficient required arguments.' % (item)
    #    raise InsufficientArgs(error_)
    error_count = 0
    error_list = []
    for item in required_args:
        if item not in kwargs.keys():
            error_count =+ 1
            error_list += [item]
    if error_count > 0:
        error_ = '\n\n***Begin ERROR***\n\n - The Following REQUIRED Key(s) Were Not Found in kwargs: "%s"\n\n****End ERROR****\n' % (error_list)
        raise InsufficientArgs(error_)

    error_count = 0
    error_list = []
    for item in optional_args:
        if item not in kwargs.keys():
            error_count =+ 1
            error_list += [item]
    if error_count > 0:
        error_ = '\n\n***Begin ERROR***\n\n - The Following Optional Key(s) Were Not Found in kwargs: "%s"\n\n****End ERROR****\n' % (error_list)
        raise InsufficientArgs(error_)

    # Load all required args values from kwargs
    error_count = 0
    error_list = []
    for item in kwargs:
        if item in required_args.keys():
            required_args[item] = kwargs[item]
            if required_args[item] == None:
                error_count =+ 1
                error_list += [item]

    if error_count > 0:
        error_ = '\n\n***Begin ERROR***\n\n - The Following REQUIRED Key(s) Argument(s) are Blank:\nPlease Validate "%s"\n\n****End ERROR****\n' % (error_list)
        raise InsufficientArgs(error_)

    for item in kwargs:
        if item in optional_args.keys():
            optional_args[item] = kwargs[item]
    # Combine option and required dicts for Jinja template render
    templateVars = {**required_args, **optional_args}
    return(templateVars)

# Function to Apply Method to Site(s)
def process_method(wb, ws, row_num, wr_method, dest_dir, dest_file, template, **templateVars):
    if re.search('Grp_[A-F]', templateVars['Site_Group']):
        Group_ID = '%s' % (templateVars['Site_Group'])
        site_group = ast.literal_eval(os.environ[Group_ID])
        for x in range(1, 13):
            sitex = 'Site_%s' % (x)
            if not site_group[sitex] == None:
                Site_ID = 'Site_ID_%s' % (site_group[sitex])
                site_dict = ast.literal_eval(os.environ[Site_ID])

                # Create templateVars for Site_Name and APIC_URL
                templateVars['Site_Name'] = site_dict.get('Site_Name')
                templateVars['APIC_URL'] = site_dict.get('APIC_URL')
                templateVars['APIC_Version'] = site_dict.get('APIC_Version')

                # Create Terraform file from Template
                create_tf_file(wr_method, dest_dir, dest_file, template, **templateVars)

    elif re.search(r'\d+', templateVars['Site_Group']):
        Site_ID = 'Site_ID_%s' % (templateVars['Site_Group'])
        site_dict = ast.literal_eval(os.environ[Site_ID])

        # Create templateVars for Site_Name and APIC_URL
        templateVars['Site_Name'] = site_dict.get('Site_Name')
        templateVars['APIC_URL'] = site_dict.get('APIC_URL')
        templateVars['APIC_Version'] = site_dict.get('APIC_Version')

        # Create Terraform file from Template
        create_tf_file(wr_method, dest_dir, dest_file, template, **templateVars)
    else:
        print(f"\n-----------------------------------------------------------------------------\n")
        print(f"   Error on Worksheet {ws.title}, Row {row_num} Site_Group, value {templateVars['Site_Group']}.")
        print(f"   Unable to Determine if this is a Single or Group of Site(s).  Exiting....")
        print(f"\n-----------------------------------------------------------------------------\n")
        exit()

# Function to Process Sensitive Variables
def process_sensitive_var(wb, ws, row_num, dest_dir, dest_file, template, **templateVars):
    # Initialize the Class
    lib_tf_ref = 'lib_terraform.Terraform_Cloud'
    class_init = '%s()' % (lib_tf_ref)

    # Check the Sites Tab for Variable Location
    if re.search('Grp_[A-F]', templateVars['Site_Group']):
        Group_ID = '%s' % (templateVars['Site_Group'])
        site_group = ast.literal_eval(os.environ[Group_ID])
        for x in range(1, 13):
            sitex = 'Site_%s' % (x)
            if not site_group[sitex] == None:
                Site_ID = 'Site_ID_%s' % (site_group[sitex])
                site_dict = ast.literal_eval(os.environ[Site_ID])

                kwargs = {}
                kwargs['Site_ID'] = site_dict.get('Site_ID')
                kwargs['Site_Name'] = site_dict.get('Site_Name')
                kwargs['APIC_URL'] = site_dict.get('APIC_URL')
                kwargs['APIC_Version'] = site_dict.get('APIC_Version')
                kwargs['APIC_Auth_Type'] = site_dict.get('APIC_Auth_Type')
                kwargs['Terraform_EQ'] = site_dict.get('Terraform_EQ')
                kwargs['Terraform_Version'] = site_dict.get('Terraform_Version')
                kwargs['Provider_EQ'] = site_dict.get('Provider_EQ')
                kwargs['Provider_Version'] = site_dict.get('Provider_Version')
                kwargs['Run_Location'] = site_dict.get('Run_Location')
                kwargs['State_Location'] = site_dict.get('State_Location')
                kwargs['Terraform_Cloud_Org'] = site_dict.get('Terraform_Cloud_Org')
                kwargs['Workspace_Prefix'] = site_dict.get('Workspace_Prefix')
                kwargs['VCS_Base_Repo'] = site_dict.get('VCS_Base_Repo')
                kwargs['Terraform_Agent_Pool_ID'] = site_dict.get('Terraform_Agent_Pool_ID')
                kwargs['Terraform_Agent_Pool_ID'] = site_dict.get('Terraform_Agent_Pool_ID')

                # Dicts for required and optional args
                required_args = {'Site_ID': '',
                                'Site_Name': '',
                                'APIC_URL': '',
                                'APIC_Version': '',
                                'APIC_Auth_Type': '',
                                'Terraform_EQ': '',
                                'Terraform_Version': '',
                                'Provider_EQ': '',
                                'Provider_Version': '',
                                'Run_Location': '',
                                'State_Location': ''}
                optional_args = {'Terraform_Cloud_Org': '',
                                'Workspace_Prefix': '',
                                'VCS_Base_Repo': '',
                                'Terraform_Agent_Pool_ID': ''}

                # Validate inputs, return dict of template vars
                templateVars2 = process_kwargs(required_args, optional_args, **kwargs)

                templateVars = {**templateVars, **templateVars2}

                # If the Run_Location is Terraform_Cloud Configure Variables in the Cloud
                if templateVars['Run_Location'] == 'Terraform_Cloud':

                    # Get terraform_cloud_token
                    kwargs['terraform_cloud_token'] = eval("%s.%s()" % (class_init, 'terraform_token'))

                    # Get var_ids
                    if not kwargs.get('Workspace_Prefix') == None:
                        kwargs['Workspace_Name'] = '%s_%s_ACI_%s' % (kwargs.get('Workspace_Prefix'),kwargs.get('Site_Name'), dest_dir)
                    else:
                        kwargs['Workspace_Name'] = '%s_ACI_%s' % (kwargs.get('Site_Name'), dest_dir)
                    kwargs['workspace_id'] = os.environ.get(kwargs['Workspace_Name'])
                    kwargs['Description'] = ''
                    var = templateVars['sensitive_var']
                    tf_var_dict = {}
                    tf_var_dict = tf_variables(class_init, dest_dir, var, tf_var_dict, **kwargs)

                else:
                    kwargs['Variable'] = templateVars['sensitive_var']
                    kwargs['Var_Value'] = eval("%s.%s(**kwargs)" % (class_init, 'var_value'))

    elif re.search(r'\d+', templateVars['Site_Group']):
        Site_ID = 'Site_ID_%s' % (templateVars['Site_Group'])
        site_dict = ast.literal_eval(os.environ[Site_ID])

        kwargs = {}
        kwargs['Site_ID'] = site_dict.get('Site_ID')
        kwargs['Site_Name'] = site_dict.get('Site_Name')
        kwargs['APIC_URL'] = site_dict.get('APIC_URL')
        kwargs['APIC_Version'] = site_dict.get('APIC_Version')
        kwargs['APIC_Auth_Type'] = site_dict.get('APIC_Auth_Type')
        kwargs['Terraform_EQ'] = site_dict.get('Terraform_EQ')
        kwargs['Terraform_Version'] = site_dict.get('Terraform_Version')
        kwargs['Provider_EQ'] = site_dict.get('Provider_EQ')
        kwargs['Provider_Version'] = site_dict.get('Provider_Version')
        kwargs['Run_Location'] = site_dict.get('Run_Location')
        kwargs['State_Location'] = site_dict.get('State_Location')
        kwargs['Terraform_Cloud_Org'] = site_dict.get('Terraform_Cloud_Org')
        kwargs['Workspace_Prefix'] = site_dict.get('Workspace_Prefix')
        kwargs['VCS_Base_Repo'] = site_dict.get('VCS_Base_Repo')
        kwargs['Terraform_Agent_Pool_ID'] = site_dict.get('Terraform_Agent_Pool_ID')

        # Dicts for required and optional args
        required_args = {'Site_ID': '',
                        'Site_Name': '',
                        'APIC_URL': '',
                        'APIC_Version': '',
                        'APIC_Auth_Type': '',
                        'Terraform_EQ': '',
                        'Terraform_Version': '',
                        'Provider_EQ': '',
                        'Provider_Version': '',
                        'Run_Location': '',
                        'State_Location': ''}
        optional_args = {'Terraform_Cloud_Org': '',
                        'Workspace_Prefix': '',
                        'VCS_Base_Repo': '',
                        'Terraform_Agent_Pool_ID': ''}

        # Validate inputs, return dict of template vars
        templateVars2 = process_kwargs(required_args, optional_args, **kwargs)

        templateVars = {**templateVars, **templateVars2}

        # If the Run_Location is Terraform_Cloud Configure Variables in the Cloud
        if templateVars['Run_Location'] == 'Terraform_Cloud':

            # Get terraform_cloud_token
            kwargs['terraform_cloud_token'] = eval("%s.%s()" % (class_init, 'terraform_token'))

            # Get var_ids
            tf_var_dict = {}
            if not kwargs.get('Workspace_Prefix') == None:
                kwargs['Workspace_Name'] = '%s_%s_ACI_%s' % (kwargs.get('Workspace_Prefix'),kwargs.get('Site_Name'), dest_dir)
            else:
                kwargs['Workspace_Name'] = '%s_ACI_%s' % (kwargs.get('Site_Name'), dest_dir)
            kwargs['workspace_id'] = os.environ.get(kwargs['Workspace_Name'])
            kwargs['Description'] = ''
            var = templateVars['sensitive_var']
            kwargs['Description'] = ''
            var = templateVars['sensitive_var']
            tf_var_dict = tf_variables(class_init, dest_dir, var, tf_var_dict, **kwargs)

        else:
            kwargs['Variable'] = templateVars['sensitive_var']
            kwargs['Var_Value'] = eval("%s.%s(**kwargs)" % (class_init, 'var_value'))

    else:
        print(f"\n-----------------------------------------------------------------------------\n")
        print(f"   Error on Worksheet {ws.title}, Row {row_num} Site_Group, value {templateVars['Site_Group']}.")
        print(f"   Unable to Determine if this is a Single or Group of Site(s).  Exiting....")
        print(f"\n-----------------------------------------------------------------------------\n")
        exit()

# Function to Add Static Port Bindings to Bridge Domains Terraform Files
def process_workbook(wb, ws, row_num, wr_method, dest_dir, dest_file, template, **templateVars):
    if re.search('Grp_[A-F]', templateVars['Site_Group']):
        Group_ID = '%s' % (templateVars['Site_Group'])
        site_group = ast.literal_eval(os.environ[Group_ID])
        for x in range(1, 13):
            sitex = 'Site_%s' % (x)
            if not site_group[sitex] == None:
                Site_ID = 'Site_ID_%s' % (site_group[sitex])
                site_dict = ast.literal_eval(os.environ[Site_ID])

                # Create templateVars for Site_Name and APIC_URL
                templateVars['Site_Name'] = site_dict.get('Site_Name')
                templateVars['Site_Group'] = site_dict.get('Site_ID')
                templateVars['APIC_URL'] = site_dict.get('APIC_URL')

                # Pull in the Site Workbook
                excel_workbook = '%s_intf_selectors.xlsx' % (templateVars['Site_Name'])
                try:
                    wb_sw = load_workbook(excel_workbook)
                except Exception as e:
                    print(f"Something went wrong while opening the workbook - {excel_workbook}... ABORT!")
                    sys.exit(e)

                # Process the Interface Selectors for Static Port Paths
                create_static_paths(wb, wb_sw, row_num, wr_method, dest_dir, dest_file, template, **templateVars)

    elif re.search(r'\d+', templateVars['Site_Group']):
        Site_ID = 'Site_ID_%s' % (templateVars['Site_Group'])
        site_dict = ast.literal_eval(os.environ[Site_ID])

        # Create templateVars for Site_Name and APIC_URL
        templateVars['Site_Name'] = site_dict.get('Site_Name')
        templateVars['Site_Group'] = site_dict.get('Site_ID')
        templateVars['APIC_URL'] = site_dict.get('APIC_URL')

        # Pull in the Site Workbook
        excel_workbook = '%s_intf_selectors.xlsx' % (templateVars['Site_Name'])
        try:
            wb_sw = load_workbook(excel_workbook)
        except Exception as e:
            print(f"Something went wrong while opening the workbook - {excel_workbook}... ABORT!")
            sys.exit(e)


        # Process the Interface Selectors for Static Port Paths
        create_static_paths(wb, wb_sw, row_num, wr_method, dest_dir, dest_file, template, **templateVars)

    else:
        print(f"\n-----------------------------------------------------------------------------\n")
        print(f"   Error on Worksheet {ws.title}, Row {row_num} Site_Group, value {templateVars['Site_Group']}.")
        print(f"   Unable to Determine if this is a Single or Group of Site(s).  Exiting....")
        print(f"\n-----------------------------------------------------------------------------\n")
        exit()

# Function to Determine Port Count on Modules
def query_module_type(row_num, module_type):
    if re.search('^M4', module_type):
        port_count = '4'
    elif re.search('^M6', module_type):
        port_count = '6'
    elif re.search('^M12', module_type):
        port_count = '12'
    elif re.search('X9716D-GX', module_type):
        port_count = '16'
    elif re.search('X9732C-EX', module_type):
        port_count = '32'
    elif re.search('X9736', module_type):
        port_count = '36'
    else:
        print(f'\n-----------------------------------------------------------------------------\n')
        print(f'   Error on Row {row_num}.  Unknown Switch Model {module_type}')
        print(f'   Please verify Input Information.  Exiting....')
        print(f'\n-----------------------------------------------------------------------------\n')
        exit()
    return port_count

# Function to Determine Port count from Switch Model
def query_switch_model(row_num, switch_type):
    modules = ''
    switch_type = str(switch_type)
    if re.search('^9396', switch_type):
        modules = '2'
        port_count = '48'
    elif re.search('^93', switch_type):
        modules = '1'

    if re.search('^9316', switch_type):
        port_count = '16'
    elif re.search('^(93120)', switch_type):
        port_count = '102'
    elif re.search('^(93108|93120|93216|93360)', switch_type):
        port_count = '108'
    elif re.search('^(93180|93240|9348|9396)', switch_type):
        port_count = '54'
    elif re.search('^(93240)', switch_type):
        port_count = '60'
    elif re.search('^9332', switch_type):
        port_count = '34'
    elif re.search('^(9336|93600)', switch_type):
        port_count = '36'
    elif re.search('^9364C-GX', switch_type):
        port_count = '64'
    elif re.search('^9364', switch_type):
        port_count = '66'
    elif re.search('^95', switch_type):
        port_count = '36'
        if switch_type == '9504':
            modules = '4'
        elif switch_type == '9508':
            modules = '8'
        elif switch_type == '9516':
            modules = '16'
        else:
            print(f'\n-----------------------------------------------------------------------------\n')
            print(f'   Error on Row {row_num}.  Unknown Switch Model {switch_type}')
            print(f'   Please verify Input Information.  Exiting....')
            print(f'\n-----------------------------------------------------------------------------\n')
            exit()
    else:
        print(f'\n-----------------------------------------------------------------------------\n')
        print(f'   Error on Row {row_num}.  Unknown Switch Model {switch_type}')
        print(f'   Please verify Input Information.  Exiting....')
        print(f'\n-----------------------------------------------------------------------------\n')
        exit()
    return modules,port_count

# Function to Read Excel Workbook Data
def read_in(excel_workbook):
    try:
        wb = load_workbook(excel_workbook)
        print("Workbook Loaded.")
    except Exception as e:
        print(f"Something went wrong while opening the workbook - {excel_workbook}... ABORT!")
        sys.exit(e)
    return wb

# Function to Define stdout_log output
def stdout_log(sheet, line):
    if log_level == 0:
        return
    elif ((log_level == (1) or log_level == (2)) and
            (sheet) and (line is None)):
        #print('*' * 80)
        print(f'\n-----------------------------------------------------------------------------\n')
        print(f'   Starting work on {sheet.title} Worksheet')
        print(f'\n-----------------------------------------------------------------------------\n')
        #print('*' * 80)
    elif log_level == (2) and (sheet) and (line is not None):
        print('Evaluating line %s from %s Worksheet...' % (line, sheet.title))
    else:
        return

# Function to Create/Update Variables in the Terrafrom Cloud
def tf_variables(class_init, folder, var, tf_var_dict, **kwargs):
    kwargs['Variable'] = var
    if print_response_always:
        print(f"Workspace ID is: {kwargs.get('workspace_id')} and variable is {kwargs.get('Variable')}")

    kwargs['HCL'] = 'false'
    if var == 'aciUrl':
        kwargs['Var_Value'] = 'https://%s' % (kwargs.get('APIC_URL'))
        kwargs['Sensitive'] = 'false'
    elif var == 'aciUser' or var == 'aciCertName':
        kwargs['Sensitive'] = 'false'
        if not os.environ.get(var) == None:
            kwargs['Var_Value'] = os.environ[var]
        else:
            kwargs['Var_Value'] = eval("%s.%s(**kwargs)" % (class_init, 'var_value'))
            os.environ[var] = kwargs['Var_Value']
    else:
        kwargs['Sensitive'] = 'true'
        if not os.environ.get(var) == None:
            kwargs['Var_Value'] = os.environ[var]
        else:
            kwargs['Var_Value'] = eval("%s.%s(**kwargs)" % (class_init, 'var_value'))
            os.environ[var] = kwargs['Var_Value']

    var_id = eval("%s.%s(**kwargs)" % (class_init, 'tf_variable'))
    folder_var = '%s_%s' % (folder, var)
    tf_var_dict[folder_var] = var_id

    return tf_var_dict

# Function to Aquire Workspace ID's from Terraform Cloud
def tf_workspace(class_init, folder, workspace_dict, **kwargs):
    kwargs['Working_Directory'] = 'ACI/%s/%s' % (kwargs.get('Site_Name'), folder)
    if not kwargs.get('Workspace_Prefix') == None:
        kwargs['Workspace_Name'] = '%s_%s_ACI_%s' % (kwargs.get('Workspace_Prefix'),kwargs.get('Site_Name'), folder)
    else:
        kwargs['Workspace_Name'] = '%s_ACI_%s' % (kwargs.get('Site_Name'), folder)
    print(f"Workspace Name is: {kwargs.get('Workspace_Name')}")
    # Get The Workspace ID
    workspace_id = eval("%s.%s(**kwargs)" % (class_init, 'tf_workspace'))

    tf_folder_id = 'Site_ID_%s_%s' % (kwargs.get('Site_ID'), folder)
    os.environ[tf_folder_id] = workspace_id
    workspace_dict[tf_folder_id] = workspace_id

    return workspace_dict

# Function to Expand a VLAN List to Individual VLANs
def vlan_list_full(vlan_list):
    full_vlan_list = []
    if re.search(r',', str(vlan_list)):
        vlist = vlan_list.split(',')
        for v in vlist:
            if re.fullmatch('^\\d{1,4}\\-\\d{1,4}$', v):
                a,b = v.split('-')
                a = int(a)
                b = int(b)
                vrange = range(a,b+1)
                for vl in vrange:
                    full_vlan_list.append(vl)
            elif re.fullmatch('^\\d{1,4}$', v):
                full_vlan_list.append(v)
    elif re.search('\\-', str(vlan_list)):
        a,b = vlan_list.split('-')
        a = int(a)
        b = int(b)
        vrange = range(a,b+1)
        for v in vrange:
            full_vlan_list.append(v)
    else:
        full_vlan_list.append(vlan_list)
    return full_vlan_list

# Add Prefix to VLAN Numbers for BD/EPG Names
def vlan_to_netcentric(vlan):
    vlan = int(vlan)
    if vlan < 10:
        vlan = str(vlan)
        netcentric = 'v000' + vlan
        return netcentric
    elif vlan < 100:
        vlan = str(vlan)
        netcentric = 'v00' + vlan
        return netcentric
    elif vlan < 1000:
        vlan = str(vlan)
        netcentric = 'v0' + vlan
        return netcentric
    else:
        vlan = str(vlan)
        netcentric = 'v' + vlan
        return netcentric

# Function to Expand a VLAN Range to a VLAN List
def vlan_range(vlan_list, **templateVars):
    results = 'unknown'
    while results == 'unknown':
        if re.search(',', str(vlan_list)):
            vx = vlan_list.split(',')
            for vrange in vx:
                if re.search('-', vrange):
                    vl = vrange.split('-')
                    min_ = int(vl[0])
                    max_ = int(vl[1])
                    if (int(templateVars['VLAN']) >= min_ and int(templateVars['VLAN']) <= max_):
                        results = 'true'
                        return results
                else:
                    if templateVars['VLAN'] == vrange:
                        results = 'true'
                        return results
            results = 'false'
            return results
        elif re.search('-', str(vlan_list)):
            vl = vlan_list.split('-')
            min_ = int(vl[0])
            max_ = int(vl[1])
            if (int(templateVars['VLAN']) >= min_ and int(templateVars['VLAN']) <= max_):
                results = 'true'
                return results
        else:
            if int(templateVars['VLAN']) == int(vlan_list):
                results = 'true'
                return results
        results = 'false'
        return results
