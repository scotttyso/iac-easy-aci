#!/usr/bin/env python3
"""Switch Migration Script - 
This Script is built to Deploy Converged Infrastructure from a YAML Configuration File.
The Script uses argparse to take in the following CLI arguments:
    d or dir:      Source Directory for "show run all" Configuration Files to Read.
    f or file:     A File with a List of Hosts to Login to and pull the configuration from.
    u or username: Username to Login to the switches with.
    v or vdc:      Flag to enable support for multi-vdcs, like 7Ks.
"""
#=================================================================
# Print Color Functions
#=================================================================
def prCyan(skk): print("\033[96m {}\033[00m" .format(skk))
def prGreen(skk): print("\033[92m {}\033[00m" .format(skk))
def prLightPurple(skk): print("\033[94m {}\033[00m" .format(skk))
def prLightGray(skk): print("\033[97m {}\033[00m" .format(skk))
def prPurple(skk): print("\033[95m {}\033[00m" .format(skk))
def prRed(skk): print("\033[91m {}\033[00m" .format(skk))
def prYellow(skk): print("\033[93m {}\033[00m" .format(skk))


#======================================================
# Source Modules
#======================================================
try:
    from dotmap import DotMap
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side 
    import argparse
    import easy_functions
    import json
    import pexpect
    import platform
    import os
    import re
    import sys
except ImportError as e:
    prRed(f'!!! ERROR !!!\n{e.__class__.__name__}')
    prRed(f" Module {e.name} is required to run this script")
    prRed(f" Install the module using the following: `pip install {e.name}`")

# Define Regular Expressions to be used
re_bpdu   = re.compile('^  spanning-tree bpduguard enable$\n')
re_cdpe   = re.compile('^  cdp enable$\n')
re_dhcp   = re.compile(r'^  ip dhcp relay address (\d{1,3}.\d{1,3}.\d{1,3}.\d{1,3}) $\n')
re_desc   = re.compile('^  description (.+)$\n')
re_host   = re.compile('^hostname (.+)$\n')
re_swname = re.compile('^switchname (.+)$\n')
re_hsv4   = re.compile(r'^    ip (\d{1,3}.\d{1,3}.\d{1,3}.\d{1,3})$\n')
re_hsv4s  = re.compile(r'^    ip (\d{1,3}.\d{1,3}.\d{1,3}.\d{1,3}) secondary$\n')
re_intf   = re.compile(r'^interface ((port\-channel\d+|Ethernet\d+[\d\/]+))$\n')
re_ipv4   = re.compile(r'^  ip address (\d{1,3}.\d{1,3}.\d{1,3}.\d{1,3}(?:/\d{1,2}|))$\n')
re_ipv4s  = re.compile(r'^  ip address (\d{1,3}.\d{1,3}.\d{1,3}.\d{1,3}(?:/\d{1,2}|)) secondary$\n')
re_ivln   = re.compile(r'^interface Vlan(\d+)$\n')
re_ldpr   = re.compile('^  lldp transmit$\n')
re_ldpt   = re.compile('^  lldp receive$\n')
re_mtu_   = re.compile(r'^  mtu (\d+)$\n')
re_nego   = re.compile('^  ((no negotiate auto|negotiate auto))$\n')
re_poch1  = re.compile(r'^  channel-group (\d+) mode ((active|on|passive))$\n')
re_poch2  = re.compile(r'^  channel-group (\d+)$\n')
re_sped   = re.compile('^  speed ((auto|[0-9]+))$\n')
re_swav   = re.compile(r'^  switchport access vlan (\d+)$\n')
re_swma   = re.compile('^  switchport mode access$\n')
re_swmt   = re.compile('^  switchport mode trunk$\n')
re_swpt   = re.compile('^  switchport$\n')
re_tknv   = re.compile(r'^  switchport trunk native vlan (\d{1,4})$\n')
re_tkv1   = re.compile(r'^  switchport trunk allowed vlan (\d{1,4}[\-,]+.+\d{1,4})$\n')
re_tkv2   = re.compile(r'^  switchport trunk allowed vlan (\d{1,4})$\n')
re_tkv3   = re.compile(r'^  switchport trunk allowed vlan add (\d{1,4}[\-,]+.+\d{1,4})$\n')
re_vlan   = re.compile(r'^vlan (\d{1,4})$\n')
re_vlnm   = re.compile('^  name (.+)$\n')
re_vlst   = re.compile(r'^vlan (\d{1,4}[\-,]+.+\d{1,4})$\n')
re_vpc_   = re.compile(r'^  vpc ((\d+|peer\-link))$\n')
re_vrf_   = re.compile('^  vrf member (.+)$\n')
re_vrfc   = re.compile('^vrf context (.+)$\n')
reipv6m   = re.compile('^  ipv6 multicast multipath s-g-hash\n')

# Create Workbook Format
bd1              = Side(style="thick", color="8EA9DB")
bd2              = Side(style="medium", color="8EA9DB")
wsh1             = NamedStyle(name="wsh1")
wsh1.alignment   = Alignment(horizontal="center", vertical="center", wrap_text="True")
wsh1.border      = Border(left=bd1, top=bd1, right=bd1, bottom=bd1)
wsh1.font        = Font(bold=True, size=15, color="FFFFFF")
wsh2             = NamedStyle(name="wsh2")
wsh2.alignment   = Alignment(horizontal="center", vertical="center", wrap_text="True")
wsh2.border      = Border(left=bd2, top=bd2, right=bd2, bottom=bd2)
wsh2.fill        = PatternFill("solid", fgColor="305496")
wsh2.font        = Font(bold=True, size=15, color="FFFFFF")
ws_odd           = NamedStyle(name="ws_odd")
ws_odd.alignment = Alignment(horizontal="center", vertical="center")
ws_odd.border    = Border(left=bd2, top=bd2, right=bd2, bottom=bd2)
ws_odd.fill      = PatternFill("solid", fgColor="D9E1F2")
ws_odd.font      = Font(bold=False, size=12, color="44546A")
ws_even          = NamedStyle(name="ws_even")
ws_even.alignment= Alignment(horizontal="center", vertical="center")
ws_even.border   = Border(left=bd2, top=bd2, right=bd2, bottom=bd2)
ws_even.font     = Font(bold=False, size=12, color="44546A")

#=================================================================
# Function to Create Export Workbooks
#=================================================================
def create_workbooks(jsonMap):
    wb = Workbook()
    wb.add_named_style(wsh1)
    wb.add_named_style(wsh2)
    wb.add_named_style(ws_odd)
    wb.add_named_style(ws_even)

    dest_file = 'tenant_export.xlsx'
    ws1 = wb.active
    ws1.title = "Tenants"
    ws2 = wb.create_sheet(title = "VRFs")
    ws3 = wb.create_sheet(title = "Bridge Domains")
    ws4 = wb.create_sheet(title = "DHCP Relay")
    ws5 = wb.create_sheet(title = "VLANs")
    ws6 = wb.create_sheet(title = "Duplicate BDs")
    ws2 = wb["VRFs"]
    ws3 = wb["Bridge Domains"]
    ws4 = wb["DHCP Relay"]
    ws5 = wb["VLANs"]
    ws6 = wb["Duplicate BDs"]

    # Populate Tenant Worksheet
    for x in range(1,4):
        ltr = chr(ord('@')+(x+1))
        ws1.column_dimensions[ltr].width = 30
    data = ['Type','Tenant Name','Description']
    ws1.append(data)
    for cell in ws1["1:1"]: cell.style = 'wsh2'
    
    # Populate VRFs Worksheet
    for x in range(1,10):
        ltr = chr(ord('@')+(x+1))
        ws2.column_dimensions[ltr].width = 30

    data = ['Type','site_group','tenant','name','create','alias','description','annotations','global_alias']
    ws2.append(data)
    for cell in ws2['1:1']: cell.style = 'wsh2'
    ws_count = 2
    for k, v in jsonMap.vrfs.items():
        data = ['vrf_add','','',k,'create','','','','']
        ws2.append(data)
        rc = '%s:%s' % (ws_count, ws_count)
        for cell in ws2[rc]:
            if ws_count % 2 == 0: cell.style = 'ws_even'
            else: cell.style = 'ws_odd'
        ws_count += 1

    # Populate Bridge Domains Worksheet
    for x in range(1,13):
        ltr = chr(ord('@')+(x+1))
        ws3.column_dimensions[ltr].width = 30

    data = [
        'Type','site_group','tenant','name','description','bd_template','epg_template',
        'application_profile','vlans','gateway_ips','l3outs','subnet_templates'
    ]
    ws3.append(data)
    for cell in ws3['1:1']: cell.style = 'wsh2'
    ws_count = 2
    for k, v in jsonMap.bridge_domains.items():
        if v.get('description'): descr = v.description
        else: descr = ''
        if v.get('gateway_ips'): gwy = ','.join(v.gateway_ips)
        else: gwy = ''
        data = ['bd_add','','',v['name'],descr,'','','nets',k,gwy,'','']
        ws3.append(data)
        rc = '%s:%s' % (ws_count, ws_count)
        for cell in ws3[rc]:
            if ws_count % 2 == 0: cell.style = 'ws_even'
            else: cell.style = 'ws_odd'
        ws_count += 1

    # Populate DHCP Relay Worksheet
    for x in range(1,9):
        ltr = chr(ord('@')+(x+1))
        ws4.column_dimensions[ltr].width = 30

    data = ['Type','site_group','names','addresses','mode','owner','description','epg_type',]
    ws4.append(data)
    for cell in ws4["1:1"]: cell.style = 'wsh2'
    ws_count = 2
    for k, v in jsonMap.dhcp_relay.items():
        data = ['dhcp_relay','','',k,'','','','',]
        ws4.append(data)
        rc = '%s:%s' % (ws_count, ws_count)
        for cell in ws4[rc]:
            if ws_count % 2 == 0: cell.style = 'ws_even'
            else: cell.style = 'ws_odd'
        ws_count += 1

    # Populate VLANs Worksheet
    for x in range(1,2):
        ltr = chr(ord('@')+(x+1))
        ws5.column_dimensions[ltr].width = 150

    data = ['vlans']
    ws5.append(data)
    for cell in ws5["1:1"]: cell.style = 'wsh2'
    ws_count = 2
    data = [jsonMap.vlans]
    ws5.append(data)
    rc = '%s:%s' % (ws_count, ws_count)
    for cell in ws5[rc]:
        if ws_count % 2 == 0: cell.style = 'ws_even'
        else: cell.style = 'ws_odd'
    ws_count += 1

    # Populate Duplicate BDs Worksheet
    for x in range(1,2):
        ltr = chr(ord('@')+(x+1))
        ws6.column_dimensions[ltr].width = 150

    data = ['vlan','name']
    ws6.append(data)
    for cell in ws6["1:1"]: cell.style = 'wsh2'
    ws_count = 2
    for k, v in jsonMap.bd_duplicates.items():
        data = [v['vlan'],k]
        ws6.append(data)
        rc = '%s:%s' % (ws_count, ws_count)
        for cell in ws6[rc]:
            if ws_count % 2 == 0: cell.style = 'ws_even'
            else: cell.style = 'ws_odd'
        ws_count += 1

    # Save the Excel Workbook
    wb.save(dest_file)

    # Create Switch Export Workbook
    wb1 = Workbook()
    wb1.add_named_style(wsh1)
    wb1.add_named_style(wsh2)
    wb1.add_named_style(ws_odd)
    wb1.add_named_style(ws_even)

    dest_file = 'switch_export.xlsx'
    ws_count = 0
    # Populate Switch Worksheet
    for k, v in jsonMap.switches.items():
        int_count = 0
        if len(v.interfaces) >= 1: int_count += 1
        if len(v.port_channels) >= 1: int_count += 1
        if int_count > 0:
            if ws_count == 0:
                ws = wb1.active
                ws.title = k
            else:
                ws = wb1.create_sheet(title = k)
                ws = wb1[k]
            for x in range(1,21):
                ltr = chr(ord('@')+(x+1))
                ws.column_dimensions[ltr].width = 30
            data = [
                'Type','current_host','current_interfaces','interface_profile','interface_selector',
                'interface','policy_group_type','policy_group','description','pc_id','pc_mode','vpc_id',
                'mtu','speed','status','sw_mode','acccess/native','allowed_vlans',
                'cdp','lldp_rx','lldp_tx','bpdu'
            ]
            ws.append(data)
            for cell in ws["1:1"]: cell.style = 'wsh2'

            # Loop Thru Switch Interfaces
            ws_count = 2
            for a, b in v.port_channels.items():
                ptype = 'bundle'
                if not b.vpc == '': itype = 'vpc_add'
                else: itype = 'pc_add'
                intfs = ','.join(b.interfaces)
                data = [
                    itype,k,intfs,'',a,
                    a,ptype,'needed',b.description,a,b.pc_mode,b.vpc,
                    b.mtu,b.speed,b.status,b.mode,b.access,b.allowed_vlans,
                    b.cdp,b.lldp_rx,b.lldp_tx,b.bpdu
                ]
                ws.append(data)
                rc = '%s:%s' % (ws_count, ws_count)
                for cell in ws[rc]:
                    if ws_count % 2 == 0: cell.style = 'ws_even'
                    else: cell.style = 'ws_odd'
                ws_count += 1
            for a, b in v.interfaces.items():
                intf = a.split('net')[1]
                iselect = f'Eth{intf}'
                iselect = iselect.replace('/', '-')
                if b.pc_id == 'n/a': ptype = 'access'
                else: ptype = 'bundle'
                data = [
                    'intf_selector',k,a,'',iselect,
                    intf,ptype,'needed',b.description,b.pc_id,b.pc_mode,'',
                    b.mtu,b.speed,b.status,b.mode,b.access,b.allowed_vlans,
                    b.cdp,b.lldp_rx,b.lldp_tx,b.bpdu
                ]
                ws.append(data)
                rc = '%s:%s' % (ws_count, ws_count)
                for cell in ws[rc]:
                    if ws_count % 2 == 0: cell.style = 'ws_even'
                    else: cell.style = 'ws_odd'
                ws_count += 1

    # Save the Excel Workbook
    wb1.save(dest_file)

#=================================================================
# Function to Parse the Configurations
#=================================================================
def parse_config_file(jsonMap, dir, file, status):
    # Start by Creating Default Variables
    str_bpdg = False
    str_cdp_ = False
    str_dhcp = ''
    str_desc = ''
    str_host = ''
    str_hsv4 = ''
    str_hsv4s = ''
    str_intf = ''
    str_ipv4 = ''
    str_ipv4s = ''
    str_ivln = ''
    str_lldr = False
    str_lldt = False
    str_mtu_ = '1500'
    str_nego = 'negotiate auto'
    str_poch = 'n/a'
    str_pomd = 'n/a'
    str_sped = 'auto'
    str_swav = '1'
    str_swmd = 'access'
    str_swpt = False
    str_tknv = '1'
    str_tkvl = '1'
    str_vlan = ''
    str_vlst = ''
    str_vlnm = ''
    str_vpc_ = 'n/a'
    str_vrf_ = 'default'
    str_vrfc = ''

    int_status = DotMap()
    # Read the Conifguration File and Gather Configuration Information
    if status == True:
        status_file = file.replace('.cfg', '.status')
        st_file = open(os.path.join(dir, status_file), 'r')
        for line in st_file.readlines():
            if re.search('(Eth[0-9]+/[0-9]+(/[0-9]+)?|Po[0-9]+)[ \\t]+.*[ \\t]+([a-zA-Z]+)[ \\t]+(trunk|[0-9]+|routed)[ \\t]+(full|auto|half)', line):
                intf = re.search('(Eth[0-9]+/[0-9]+(/[0-9]+)?|Po[0-9]+)[ \\t]+.*[ \\t]+([a-zA-Z]+)[ \\t]+(trunk|[0-9]+|routed)[ \\t]+(full|auto|half)', line).group(1)
                int_stat = re.search('(Eth[0-9]+/[0-9]+(/[0-9]+)?|Po[0-9]+)[ \\t]+.*[ \\t]+([a-zA-Z]+)[ \\t]+(trunk|[0-9]+|routed)[ \\t]+(full|auto|half)', line).group(3)
                int_status[intf] = int_stat

    bd_count = 0
    prGreen(f'Reading File: {file}')
    cfg_file = open(os.path.join(dir, file), 'r')
    for line in cfg_file.readlines():
        if re.fullmatch(re_host, line):
            # Set Hostname String
            str_host = re.fullmatch(re_host, line).group(1)
            # Append Hostname to Switches Dictionary
            if not str_host in jsonMap.switches.keys():
                jsonMap.switches[str_host] = DotMap(interfaces = DotMap(), port_channels = DotMap())
        elif re.fullmatch(re_swname, line):
            # Set Hostname String
            str_host = re.fullmatch(re_swname, line).group(1)
            # Append Hostname to Switches Dictionary
            if not str_host in jsonMap.switches.keys():
                jsonMap.switches[str_host] = DotMap(interfaces = DotMap(), port_channels = DotMap())
        elif re.fullmatch(re_vlst, line):
            # Matched the VLAN List... Now Parse for Data Export
            str_vlst = re.fullmatch(re_vlst, line).group(1)
            # Expand VLAN Ranges into Full VLAN List
            vlan_full = easy_functions.vlan_list_full(str_vlst)
            # Append Expanded VLAN List to vlans List
            jsonMap.vlans.extend(vlan_full)
        elif re.fullmatch(re_vlan, line):
            # Matched a VLAN... Now Parse for Data Export
            str_vlan = int(re.fullmatch(re_vlan, line).group(1))
            if not str_vlan in jsonMap.bridge_domains.keys():
                jsonMap.bridge_domains[str_vlan].name = 'unknown'
        elif re.fullmatch(re_vlnm, line):
            # Matched VLAN Name... Now Parse for Data Export
            str_vlnm = re.fullmatch(re_vlnm, line).group(1)
            if not str_vlan == '':
                if jsonMap.bridge_domains[str_vlan].name == 'unknown':
                    jsonMap.bridge_domains[str_vlan].name = str_vlnm
                elif not jsonMap.bridge_domains[str_vlan].name == str_vlnm:
                    jsonMap.bd_duplicates[str_vlnm] = DotMap(vlan = str_vlan)
        elif re.fullmatch(re_vrfc, line):
            str_vrfc = re.fullmatch(re_vrfc, line).group(1)
            if not str_vrfc in jsonMap.vrfs.keys():
                jsonMap.vrfs.update({str_vrfc:{}})
        elif re.fullmatch(re_ivln, line):
            # Matched an Interface VLAN... Now Parse for Data Export
            str_ivln = int(re.fullmatch(re_ivln, line).group(1))
        elif re.fullmatch(re_mtu_, line):
            # Matched the Interface MTU... Now Parse for Data Export
            str_mtu_ = re.fullmatch(re_mtu_, line).group(1)
        elif re.fullmatch(re_sped, line):
            # Matched the Interface Speed... Now Parse for Data Export
            str_sped = re.fullmatch(re_sped, line).group(1)
        elif re.fullmatch(re_nego, line):
            # Matched the Interface Negotiate Mode... Now Parse for Data Export
            str_nego = re.fullmatch(re_nego, line).group(1)
        elif re.fullmatch(re_vrf_, line):
            # Matched a VRF Context... Now Parse for Data Export
            str_vrf_ = re.fullmatch(re_vrf_, line).group(1)
        elif re.fullmatch(re_ipv4, line):
            # Matched an IPv4 Address/prefix... Now Parse for Data Export
            str_ipv4 = re.fullmatch(re_ipv4, line).group(1)
        elif re.fullmatch(re_ipv4s, line):
            # Matched an IPv4 Secondary Address/prefix... Now Parse for Data Export
            str_ipv4s = re.fullmatch(re_ipv4s, line).group(1)
        elif re.fullmatch(re_hsv4, line):
            # Matched an HSRP IPv4 Address... Now Parse for Data Export
            str_hsv4 = re.fullmatch(re_hsv4, line).group(1)
        elif re.fullmatch(re_hsv4s, line):
            # Matched an HSRP IPv4 Secondary Address/prefix... Now Parse for Data Export
            str_hsv4s = re.fullmatch(re_hsv4s, line).group(1)
        elif re.fullmatch(re_dhcp, line):
            # Matched an IPv4 DHCP Relay definition... Now Parse for Data Export
            str_dhcp = re.fullmatch(re_dhcp, line).group(1)
            if not str_dhcp in jsonMap.dhcp_relay.keys():
                jsonMap.dhcp_relay[str_dhcp] = DotMap(name = '')
        elif re.fullmatch(re_intf, line): str_intf = re.fullmatch(re_intf, line).group(1)
        elif re.fullmatch(re_bpdu, line): str_bpdg = 'BPDU_fg'
        elif re.fullmatch(re_cdpe, line): str_cdp_ = True
        elif re.fullmatch(re_ldpr, line): str_lldr = True
        elif re.fullmatch(re_ldpt, line): str_lldt = True
        elif re.fullmatch(re_swav, line): str_swav = re.fullmatch(re_swav, line).group(1)
        elif re.fullmatch(re_swma, line): str_swmd = 'access'
        elif re.fullmatch(re_swmt, line): str_swmd = 'trunk'
        elif re.fullmatch(re_tknv, line): str_tknv = re.fullmatch(re_tknv, line).group(1)
        elif re.fullmatch(re_tkv1, line): str_tkvl = re.fullmatch(re_tkv1, line).group(1)
        elif re.fullmatch(re_tkv2, line): str_tkvl = re.fullmatch(re_tkv2, line).group(1)
        elif re.fullmatch(re_tkv3, line): str_tkvl = str_tkvl + ',' + re.fullmatch(re_tkv3, line).group(1)
        elif re.fullmatch(re_swpt, line): str_swpt = True
        elif re.fullmatch(re_poch1, line):
            str_poch = re.fullmatch(re_poch1, line).group(1)
            str_pomd = re.fullmatch(re_poch1, line).group(2)
        elif re.fullmatch(re_poch2, line):
            str_poch = re.fullmatch(re_poch2, line).group(1)
            str_pomd = 'on'
        elif re.fullmatch(re_vpc_, line): str_vpc_ = re.fullmatch(re_vpc_, line).group(1)
        elif re.fullmatch(re_desc, line): str_desc = re.fullmatch(re_desc, line).group(1)
        elif line == "\n":
            # Found blank line, which means the end of the interface, time to create the output
            if not str_ipv4 == '' and not str_ivln == '':
                if str_hsv4:
                    a,b = str_ipv4.split('/')
                    gtwy = str(str_hsv4) + '/' + str(b)
                else: gtwy = str(str_ipv4)
                if not str_ivln in jsonMap.bridge_domains.keys():
                    jsonMap.bridge_domains[str_ivln] = DotMap(description = str_desc,name = 'unknown')
                jsonMap.bridge_domains[str_ivln].gateway_ips = [gtwy]
                if str_ipv4s:
                    if str_hsv4s:
                        a,b = str_ipv4s.split('/')
                        gtwy = str(str_hsv4s) + '/' + str(b)
                    else: gtwy = str(str_ipv4)
                    jsonMap.bridge_domains[str_ivln].gateway_ips.append(gtwy)
            elif 'channel' in str_intf:
                if str_swpt == True:
                    mtu1 = 9000
                    mtu2 = int(str_mtu_)
                    if mtu2 >= mtu1: str_mtu_ = '9000'
                    if str_swmd == 'trunk': str_swav = str_tknv
                    pc_intf = str_intf.split('l')[1]
                    str_status = 'unknown'
                    if len(int_status) > 0:
                        if int_status.get(f'Po{pc_intf}'): str_status = int_status[f'Po{pc_intf}']
                    jsonMap.switches[str_host].port_channels[pc_intf] = DotMap(
                        access       = str_swav,
                        allowed_vlans= str_tkvl,
                        bpdu         = str_bpdg,
                        cdp          = str_cdp_,
                        description  = str_desc,
                        interfaces   = [],
                        lldp_rx      = str_lldr,
                        lldp_tx      = str_lldt,
                        mode         = str_swmd,
                        mtu          = str_mtu_,
                        pc_mode      = str_pomd,
                        speed        = str_sped,
                        status       = str_status,
                        vpc          = str_vpc_
                    )
            elif 'Ethernet' in str_intf:
                if str_swpt == True:
                    mtu1 = 9000
                    mtu2 = int(str_mtu_)
                    if mtu2 >= mtu1: str_mtu_ = '9000'
                    if str_nego == 'no negotiate auto': str_nego = 'noNeg'
                    else: str_nego = 'Auto'
                    if   str_sped == '100':    str_sped = '100M_%s' % (str_nego)
                    elif str_sped == '1000':   str_sped = '1G_%s' % (str_nego)
                    elif str_sped == '2500':   str_sped = '2.5G_%s' % (str_nego)
                    elif str_sped == '5000':   str_sped = '5G_%s' % (str_nego)
                    elif str_sped == '10000':  str_sped = '10G_%s' % (str_nego)
                    elif str_sped == '25000':  str_sped = '25G_%s' % (str_nego)
                    elif str_sped == '40000':  str_sped = '40G_%s' % (str_nego)
                    elif str_sped == '50000':  str_sped = '50G_%s' % (str_nego)
                    elif str_sped == '100000': str_sped = '100G_%s' % (str_nego)
                    elif str_sped == '200000': str_sped = '200G_%s' % (str_nego)
                    elif str_sped == '400000': str_sped = '400G_%s' % (str_nego)
                    else: str_sped = 'inherit_%s' % (str_nego)
                    if re.search(r'\d+', str_poch):
                        jsonMap.switches[str_host].port_channels[str_poch].interfaces.append(str_intf)
                    if str_swmd == 'access': swav = str_swav
                    else: swav = str_tknv
                    str_status = 'unknown'
                    if len(int_status) > 0:
                        short_int = str_intf.replace('ernet', '')
                        if int_status.get(short_int): str_status = int_status[short_int]
                    jsonMap.switches[str_host].interfaces[str_intf] = DotMap(
                        access       = swav,
                        allowed_vlans= str_tkvl,
                        bpdu         = str_bpdg,
                        cdp          = str_cdp_,
                        description  = str_desc,
                        lldp_rx      = str_lldr,
                        lldp_tx      = str_lldt,
                        mode         = str_swmd,
                        mtu          = str_mtu_,
                        pc_id        = str_poch,
                        pc_mode      = str_pomd,
                        speed        = str_sped,
                        status       = str_status
                    )
            
            # Reset the Variables back to Blank except str_host
            str_bpdg = False
            str_cdp_ = False
            str_dhcp = ''
            str_desc = ''
            str_hsv4 = ''
            str_hsv4s = ''
            str_intf = ''
            str_ipv4 = ''
            str_ipv4s = ''
            str_ivln = ''
            str_lldr = False
            str_lldt = False
            str_mtu_ = '1500'
            str_nego = 'negotiate auto'
            str_poch = 'n/a'
            str_pomd = 'n/a'
            str_sped = 'auto'
            str_swav = 'n/a'
            str_swmd = 'access'
            str_swpt = False
            str_tknv = 'n/a'
            str_tkvl = 'n/a'
            str_vlan = ''
            str_vlst = ''
            str_vlnm = ''
            str_vpc_ = 'n/a'
            str_vrf_ = 'default'
            str_vrfc = ''
    
    # Return the Dictionary
    return jsonMap

#=================================================================
# The Main Module
#=================================================================
def main():
    # User Input Arguments
    Parser = argparse.ArgumentParser(description='Configuration Migration')
    Parser.add_argument('-d', '--dir',
        default = 'CONFIG',
        help = 'Source Directory for "show run all" Configuration Files to Read.'
    )
    Parser.add_argument('-f', '--file',
        help = 'A File with a List of Hosts to Login to and pull the configuration from.'
    )
    Parser.add_argument('-u', '--username',
        help = 'Username to Login to the switches with.'
    )
    Parser.add_argument('-v', '--vdc',
        action='store_true',
        help = 'Flag to enable support for multi-vdcs, like 7Ks.'
    )
    args = Parser.parse_args()

    kwargs = DotMap()
    kwargs.args = args
    kwargs.op_system = platform.system()
    jsonMap = DotMap(
        dhcp_relay    = DotMap(),
        bd_duplicates = DotMap(),
        bridge_domains= DotMap(),
        switches      = DotMap(),
        vlans         = [],
        vrf           = DotMap()
    )
    #=====================================================
    # Login to Devices if File with List is defined
    #=====================================================
    if args.file:
        kwargs.password = 'switch_password'
        kwargs.username = args.username
        try:
            if os.path.isfile(args.file):
                prLightGray(f'\n-----------------------------------------------------------------------------\n')
                prLightGray(f'   {args.dir} exists.  Beginning Script Execution...')
                prLightGray(f'\n-----------------------------------------------------------------------------\n')

                #=====================================================
                # Get Show run all
                #=====================================================
                args.dir = os.getcwd()
                kwargs.enable_prompt= "[0-9a-zA-Z\\-\\_\\.]{3,30}>"
                kwargs.host_prompt  = "[0-9a-zA-Z\\-\\_\\.]{3,30}#"
                for e in open(args.file, 'r'):
                    kwargs.hostname     = e.strip()
                    file_ext = ['.txt','.cfg','.config']
                    child, kwargs = easy_functions.child_login(kwargs)
                    child.sendline("term length 0")
                    child.expect(kwargs.host_prompt)
                    child.sendline("show interface status")
                    child.expect("show interface status")
                    child.expect(kwargs.host_prompt)
                    with open(f'{kwargs.hostname}.status', 'w') as f:
                        f.write(child.before)
                        f.close()

                    file = f'{kwargs.hostname}.status'
                    child.sendline("show run all")
                    child.expect("show run all")
                    child.expect(kwargs.host_prompt)
                    with open(f'{kwargs.hostname}.cfg', 'w') as f:
                        f.write(child.before)
                        f.close()

                    file = f'{kwargs.hostname}.cfg'
                    jsonMap = parse_config_file(jsonMap, args.dir, file, True)
                    os.remove(f'{kwargs.hostname}.cfg')
                    os.remove(f'{kwargs.hostname}.status')

                    #=====================================================
                    # Get Show run all
                    #=====================================================
                    if args.vdc == True:
                        vdcs = DotMap()
                        child.sendline('show vdc')
                        child.expect('show vdc')
                        vdc_output = False
                        while vdc_output == False:
                            vdc_regex = '([0-9])[ \t]+([0-9a-zA-Z\\-\\_\\.]+)[ \t]+active'
                            i = child.expect([vdc_regex, kwargs.host_prompt, pexpect.TIMEOUT])
                            if i == 0:
                                print(child.match)
                                vdcs[(child.match).group(1)] = (child.match).group(2)
                            elif i == 1: vdc_output = True
                        if len(vdcs) > 1:
                            for k, v in vdcs:
                                child.sendline(f'switchto vdc {v}')
                                child.expect('switchto vdc')
                                child.expect(f'{v}#')
                                child.sendline("show interface status")
                                child.expect("show interface status")
                                child.expect(f'{v}#')
                                with open(f'{v}.status', 'w') as f:
                                    f.write(child.before)
                                    f.close()

                                file = f'{v}.status'
                                child.sendline('show run all')
                                child.expect('show run all')
                                child.expect(f'{v}#')
                                with open(f'{v}.cfg', 'w') as f:
                                    f.write(child.before)
                                    f.close()
                                file = f'{v}.cfg'
                                jsonMap = parse_config_file(jsonMap, args.dir, file, True)
                                os.remove(f'{v}.cfg')
                                os.remove(f'{v}.status')
                    
                    child.sendline('exit')
                    child.expect('closed')
                    child.close()
            else:
                prLightGray(f'\n-----------------------------------------------------------------------------\n')
                prLightGray(f'   {args.file} does not exist.  Exiting....')
                prLightGray(f'\n-----------------------------------------------------------------------------\n')
                sys.exit(1)
        except IOError:
            prLightGray(f'\n-----------------------------------------------------------------------------\n')
            prLightGray(f'   {args.file} does not exist.  Exiting....')
            prLightGray(f'\n-----------------------------------------------------------------------------\n')
            sys.exit(1)
    else:
        # Check Configuration File(s) Directory Exists
        try:
            if os.path.isdir(args.dir):
                prLightGray(f'\n-----------------------------------------------------------------------------\n')
                prLightGray(f'   {args.dir} exists.  Beginning Script Execution...')
                prLightGray(f'\n-----------------------------------------------------------------------------\n')
            else:
                prLightGray(f'\n-----------------------------------------------------------------------------\n')
                prLightGray(f'   {args.dir} does not exist.  Exiting....')
                prLightGray(f'\n-----------------------------------------------------------------------------\n')
                sys.exit(1)
        except IOError:
            prLightGray(f'\n-----------------------------------------------------------------------------\n')
            prLightGray(f'   {args.dir} does not exist.  Exiting....')
            prLightGray(f'\n-----------------------------------------------------------------------------\n')
            sys.exit(1)

        #Get Configuration Files
        for file in os.listdir(args.dir):
            file_ext = ['.txt','.cfg','.config']
            for ext in file_ext:
                if file.endswith(ext):
                    jsonMap = parse_config_file(jsonMap, args.dir, file, False)
    
    # Sort vlans and compact list
    jsonMap.vlans.sort()
    vlans = set(jsonMap.vlans)
    vlan_list = []
    for vlan in vlans: vlan_list.append(vlan)
    jsonMap.vlans = easy_functions.vlan_list_format(vlan_list)
    
    # Sort Dictionary keys
    dlist = ['bd_duplicates','bridge_domains','switches','vrfs']
    for item in dlist:
        dic2 = {}    
        for i in sorted(jsonMap[item]):
            dic2[i] = jsonMap[item][i]
        jsonMap[item] = dic2
        dic2 = {}    

    # Create the Workbooks
    create_workbooks(jsonMap)

    #End Script
    prLightGray(f'\n-----------------------------------------------------------------------------\n')
    prLightGray(f'   Completed Running Script.  Exiting....')
    prLightGray(f'\n-----------------------------------------------------------------------------\n')
    sys.exit(0)

if __name__ == '__main__':
    main()
