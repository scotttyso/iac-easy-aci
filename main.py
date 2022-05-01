#!/usr/bin/env python3
"""ACI/NDO IAC - 
This Script is to Create Terraform HCL configuration from an Excel Spreadsheet.
It uses argparse to take in the following CLI arguments:
    d or dir:           Base Directory to use for creation of the HCL Configuration Files
    w or workbook:   Name of Excel Workbook file for the Data Source
"""

#======================================================
# Source Modules
#======================================================
from classes import access, admin, fabric, site_policies, system_settings
from class_tenants import tenants
from easy_functions import countKeys, findKeys, findVars
from easy_functions import read_easy_jsonData, read_in, stdout_log
# from openpyxl import load_workbook
# from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from pathlib import Path
from git import Repo
import argparse
import json
import os
import re
import stdiomask
import subprocess
import time

#======================================================
# Global Variables
#======================================================
excel_workbook = None
home = Path.home()
Parser = argparse.ArgumentParser(description='IaC Easy ACI Deployment Module')
workspace_dict = {}

#======================================================
# Regular Expressions to Control wich rows in the
# Spreadsheet should be processed.
#======================================================
access_regex = re.compile('^(aep_profile|bpdu|cdp|(fibre|port)_(channel|security)|l2_interface|l3_domain|(leaf|spine)_pg|link_level|lldp|mcp|pg_(access|breakout|bundle|spine)|phys_dom|stp|vlan_pool)$')
admin_regex = re.compile('^(auth|export_policy|remote_host|security)$')
# admin_regex = re.compile('^(auth|export_policy|radius|remote_host|security|tacacs)$')
system_settings_regex = re.compile('^(apic_preference|bgp_(asn|rr)|global_aes)$')
bridge_domains_regex = re.compile('^add_bd$')
contracts_regex = re.compile('(^(contract|filter|subject)_(add|entry|to_epg)$)')
dhcp_regex = re.compile('^dhcp_add$')
epgs_regex = re.compile('^((app|epg)_add)$')
fabric_regex = re.compile('^(date_time|dns_profile|ntp(_key)?|smart_(callhome|destinations|smtp_server)|snmp_(clgrp|community|destinations|policy|user)|syslog(_destinations)?)$')
inventory_regex = re.compile('^(apic_inb|switch|vpc_pair)$')
l3out_regex = re.compile('^(add_l3out|ext_epg|node_(prof|intf|path)|bgp_peer)$')
mgmt_tenant_regex = re.compile('^(add_bd|mgmt_epg|oob_ext_epg)$')
sites_regex = re.compile('^(site_id|group_id)$')
tenants_regex = re.compile('^(tenant_add)$')
# tenants_regex = re.compile('^((tenant|vrf)_add|vrf_community)$')
vmm_regex = re.compile('^add_vmm$')

#======================================================
# Function to run 'terraform plan' and
# 'terraform apply' in the each folder of the
# Destination Directory.
#======================================================
def apply_aci_terraform(folders):

    print(f'\n-----------------------------------------------------------------------------\n')
    print(f'  Found the Followng Folders with uncommitted changes:\n')
    for folder in folders:
        print(f'  - {folder}')
    print(f'\n  Beginning Terraform Plan and Apply in each folder.')
    print(f'\n-----------------------------------------------------------------------------\n')

    time.sleep(7)

    response_p = ''
    response_a = ''
    for folder in folders:
        path = './%s' % (folder)
        lock_count = 0
        p = subprocess.Popen(['terraform', 'init', '-plugin-dir=../../../terraform-plugins/providers/'],
                             cwd=path,
                             stdout=subprocess.PIPE,
                             stderr=subprocess.STDOUT)
        for line in iter(p.stdout.readline, b''):
            print(line)
            if re.search('does not match configured version', line.decode("utf-8")):
                lock_count =+ 1

        if lock_count > 0:
            p = subprocess.Popen(['terraform', 'init', '-upgrade', '-plugin-dir=../../../terraform-plugins/providers/'], cwd=path)
            p.wait()
        p = subprocess.Popen(['terraform', 'plan', '-out=main.plan'], cwd=path)
        p.wait()
        while True:
            print(f'\n-----------------------------------------------------------------------------\n')
            print(f'  Terraform Plan Complete.  Please Review the Plan and confirm if you want')
            print(f'  to move forward.  "A" to Apply the Plan. "S" to Skip.  "Q" to Quit.')
            print(f'  Current Working Directory: {folder}')
            print(f'\n-----------------------------------------------------------------------------\n')
            response_p = input('  Please Enter ["A", "S" or "Q"]: ')
            if re.search('^(A|S)$', response_p):
                break
            elif response_p == 'Q':
                exit()
            else:
                print(f'\n-----------------------------------------------------------------------------\n')
                print(f'  A Valid Response is either "A", "S" or "Q"...')
                print(f'\n-----------------------------------------------------------------------------\n')

        if response_p == 'A':
            p = subprocess.Popen(['terraform', 'apply', '-parallelism=1', 'main.plan'], cwd=path)
            p.wait()

        while True:
            if response_p == 'A':
                response_p = ''
                print(f'\n-----------------------------------------------------------------------------\n')
                print(f'  Terraform Apply Complete.  Please Review for any errors and confirm if you')
                print(f'  want to move forward.  "M" to Move to the Next Section. "Q" to Quit..')
                print(f'\n-----------------------------------------------------------------------------\n')
                response_a = input('  Please Enter ["M" or "Q"]: ')
            elif response_p == 'S':
                break
            if response_a == 'M':
                break
            elif response_a == 'Q':
                exit()
            else:
                print(f'\n-----------------------------------------------------------------------------\n')
                print(f'  A Valid Response is either "M" or "Q"...')
                print(f'\n-----------------------------------------------------------------------------\n')

#======================================================
# Function to Check the Git Status of the Destination Folder
#======================================================
def check_git_status():
    random_folders = []
    git_path = './'
    result = subprocess.Popen(['python3', '-m', 'git_status_checker'], stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
    while(True):
        # returns None while subprocess is running
        retcode = result.poll()
        line = result.stdout.readline()
        line = line.decode('utf-8')
        if re.search(r'M (.*/).*.tf\n', line):
            folder = re.search(r'M (.*/).*.tf\n', line).group(1)
            if not re.search(r'ACI.templates', folder):
                if not folder in random_folders:
                    random_folders.append(folder)
        elif re.search(r'\?\? (.*/).*.tf\n', line):
            folder = re.search(r'\?\? (.*/).*.tf\n', line).group(1)
            if not re.search(r'ACI.templates', folder):
                if not folder in random_folders:
                    random_folders.append(folder)
        elif re.search(r'\?\? (ACI/.*/)\n', line):
            folder = re.search(r'\?\? (ACI/.*/)\n', line).group(1)
            if not (re.search(r'ACI.templates', folder) or re.search(r'\.terraform', folder)):
                if os.path.isdir(folder):
                    folder = [folder]
                    random_folders = random_folders + folder
                else:
                    group_x = [os.path.join(folder, o) for o in os.listdir(folder) if os.path.isdir(os.path.join(folder,o))]
                    random_folders = random_folders + group_x
        if retcode is not None:
            break

    if not random_folders:
        print(f'\n-----------------------------------------------------------------------------\n')
        print(f'   There were no uncommitted changes in the environment.')
        print(f'   Proceedures Complete!!! Closing Environment and Exiting Script.')
        print(f'\n-----------------------------------------------------------------------------\n')
        exit()

    strict_folders = []
    folder_order = ['Access', 'System', 'Tenant_common', 'Tenant_infra', 'Tenant_mgmt', 'Fabric', 'Admin', 'VLANs', 'Tenant_infra',]
    for folder in folder_order:
        for fx in random_folders:
            if folder in fx:
                if 'ACI' in folder:
                    strict_folders.append(fx)
    for folder in strict_folders:
        if folder in random_folders:
            random_folders.remove(folder)
    for folder in random_folders:
        if 'ACI' in folder:
            strict_folders.append(folder)

    # print(strict_folders)
    return strict_folders

def get_user_pass():
    print(f'\n-----------------------------------------------------------------------------\n')
    print(f'   Beginning Proceedures to Apply Terraform Resources to the environment')
    print(f'\n-----------------------------------------------------------------------------\n')

    user = input('Enter APIC username: ')
    while True:
        try:
            password = stdiomask.getpass(prompt='Enter APIC password: ')
            break
        except Exception as e:
            print('Something went wrong. Error received: {}'.format(e))

    os.environ['TF_VAR_aciUser'] = '%s' % (user)
    os.environ['TF_VAR_aciPass'] = '%s' % (password)

#======================================================
# Function to Read the Access Worksheet
#======================================================
def process_access(easyDict, easy_jsonData, wb):
    # Evaluate Access Worksheet
    class_init = 'access'
    class_folder = 'access'
    func_regex = access_regex
    ws = wb['Access']
    easyDict = read_worksheet(class_init, class_folder, easyDict, easy_jsonData, func_regex, wb, ws)
    return easyDict

#======================================================
# Function to Read the Admin Worksheet
#======================================================
def process_admin(easyDict, easy_jsonData, wb):
    # Evaluate Admin Worksheet
    class_init = 'admin'
    class_folder = 'admin'
    func_regex = admin_regex
    ws = wb['Admin']
    easyDict = read_worksheet(class_init, class_folder, easyDict, easy_jsonData, func_regex, wb, ws)
    return easyDict

#======================================================
# Function to Read the Bridge Domains Worksheet
#======================================================
def process_bridge_domains(easyDict, easy_jsonData, wb):
    # Evaluate Bridge_Domains Worksheet
    class_init = 'tenants'
    class_folder = 'tenants'
    func_regex = bridge_domains_regex
    ws = wb['Bridge_Domains']
    easyDict = read_worksheet(class_init, class_folder, easyDict, easy_jsonData, func_regex, wb, ws)
    return easyDict

#======================================================
# Function to Read the Contracts Worksheet
#======================================================
def process_contracts(easyDict, easy_jsonData, wb):
    # Evaluate Contracts Worksheet
    class_init = 'tenants'
    class_folder = 'tenants'
    func_regex = contracts_regex
    ws = wb['Contracts']
    easyDict = read_worksheet(class_init, class_folder, easyDict, easy_jsonData, func_regex, wb, ws)
    return easyDict

#======================================================
# Function to Read the EPGs Worksheet
#======================================================
def process_epgs(easyDict, easy_jsonData, wb):
    # Evaluate EPGs Worksheet
    class_init = 'tenants'
    class_folder = 'tenants'
    func_regex = epgs_regex
    ws = wb['EPGs']
    easyDict = read_worksheet(class_init, class_folder, easyDict, easy_jsonData, func_regex, wb, ws)
    return easyDict

#======================================================
# Function to Read the Fabric Worksheet
#======================================================
def process_fabric(easyDict, easy_jsonData, wb):
    # Evaluate Fabric Worksheet
    class_init = 'fabric'
    class_folder = 'fabric'
    func_regex = fabric_regex
    ws = wb['Fabric']
    easyDict = read_worksheet(class_init, class_folder, easyDict, easy_jsonData, func_regex, wb, ws)
    return easyDict

#======================================================
# Function to Read the Fabric Worksheet
#======================================================
def process_inventory(easyDict, easy_jsonData, wb):
    # Evaluate Inventory Worksheet
    class_init = 'access'
    class_folder = 'access'
    func_regex = inventory_regex
    ws = wb['Inventory']
    easyDict = read_worksheet(class_init, class_folder, easyDict, easy_jsonData, func_regex, wb, ws)
    return easyDict

#======================================================
# Function to Read the L3Out Worksheet
#======================================================
def process_l3out(easyDict, easy_jsonData, wb):
    # Evaluate L3Out Worksheet
    class_init = 'tenants'
    class_folder = 'tenants'
    func_regex = l3out_regex
    ws = wb['L3Out']
    easyDict = read_worksheet(class_init, class_folder, easyDict, easy_jsonData, func_regex, wb, ws)
    return easyDict

#======================================================
# Function to Read the Sites Worksheet
#======================================================
def process_sites(easyDict, easy_jsonData, wb):
    # Evaluate Sites Worksheet
    class_init = 'site_policies'
    class_folder = 'sites'
    func_regex = sites_regex
    ws = wb['Sites']
    easyDict = read_worksheet(class_init, class_folder, easyDict, easy_jsonData, func_regex, wb, ws)
    return easyDict

#======================================================
# Function to Read the System Settings Worksheet
#======================================================
def process_system_settings(easyDict, easy_jsonData, wb):
    # Evaluate System_Settings Worksheet
    class_init = 'system_settings'
    class_folder = 'system_settings'
    func_regex = system_settings_regex
    ws = wb['System Settings']
    easyDict = read_worksheet(class_init, class_folder, easyDict, easy_jsonData, func_regex, wb, ws)
    return easyDict

#======================================================
# Function to Read the Tenants Worksheet
#======================================================
def process_tenants(easyDict, easy_jsonData, wb):
    # Evaluate Tenants Worksheet
    class_init = 'tenants'
    class_folder = 'tenants'
    func_regex = tenants_regex
    ws = wb['Tenants']
    easyDict = read_worksheet(class_init, class_folder, easyDict, easy_jsonData, func_regex, wb, ws)
    return easyDict

#======================================================
# Function to Read the Worksheet and Create Templates
#======================================================
def read_worksheet(class_init, class_folder, easyDict, easy_jsonData, func_regex, wb, ws):
    rows = ws.max_row
    func_list = findKeys(ws, func_regex)
    stdout_log(ws, None)
    for func in func_list:
        count = countKeys(ws, func)
        var_dict = findVars(ws, func, rows, count)
        for pos in var_dict:
            row_num = var_dict[pos]['row']
            del var_dict[pos]['row']
            for x in list(var_dict[pos].keys()):
                if var_dict[pos][x] == '':
                    del var_dict[pos][x]
            stdout_log(ws, row_num)
            var_dict[pos].update(
                {
                    'class_folder':class_folder,
                    'easyDict':easyDict,
                    'easy_jsonData':easy_jsonData,
                    'row_num':row_num,
                    'wb':wb,
                    'ws':ws
                }
            )
            easyDict = eval(f"{class_init}(class_folder).{func}(**var_dict[pos])")
    
    # Return the easyDict
    return easyDict

#======================================================
# Function to Update the Workbook.  This is not
# currently utilized.  Likely to Depricate.
#======================================================
def wb_update(wr_ws, status, i):
    # build green and red style sheets for excel
    bd1 = Side(style="thick", color="8EA9DB")
    bd2 = Side(style="medium", color="8EA9DB")
    wsh1 = NamedStyle(name="wsh1")
    wsh1.alignment = Alignment(horizontal="center", vertical="center", wrap_text="True")
    wsh1.border = Border(left=bd1, top=bd1, right=bd1, bottom=bd1)
    wsh1.font = Font(bold=True, size=15, color="FFFFFF")
    wsh2 = NamedStyle(name="wsh2")
    wsh2.alignment = Alignment(horizontal="center", vertical="center", wrap_text="True")
    wsh2.border = Border(left=bd2, top=bd2, right=bd2, bottom=bd2)
    wsh2.fill = PatternFill("solid", fgColor="305496")
    wsh2.font = Font(bold=True, size=15, color="FFFFFF")
    green_st = NamedStyle(name="ws_odd")
    green_st.alignment = Alignment(horizontal="center", vertical="center")
    green_st.border = Border(left=bd2, top=bd2, right=bd2, bottom=bd2)
    green_st.fill = PatternFill("solid", fgColor="D9E1F2")
    green_st.font = Font(bold=False, size=12, color="44546A")
    red_st = NamedStyle(name="ws_even")
    red_st.alignment = Alignment(horizontal="center", vertical="center")
    red_st.border = Border(left=bd2, top=bd2, right=bd2, bottom=bd2)
    red_st.font = Font(bold=False, size=12, color="44546A")
    yellow_st = NamedStyle(name="ws_even")
    yellow_st.alignment = Alignment(horizontal="center", vertical="center")
    yellow_st.border = Border(left=bd2, top=bd2, right=bd2, bottom=bd2)
    yellow_st.font = Font(bold=False, size=12, color="44546A")
    # green_st = xlwt.easyxf('pattern: pattern solid;')
    # green_st.pattern.pattern_fore_colour = 3
    # red_st = xlwt.easyxf('pattern: pattern solid;')
    # red_st.pattern.pattern_fore_colour = 2
    # yellow_st = xlwt.easyxf('pattern: pattern solid;')
    # yellow_st.pattern.pattern_fore_colour = 5
    # if stanzas to catch the status code from the request
    # and then input the appropriate information in the workbook
    # this then writes the changes to the doc
    if status == 200:
        wr_ws.write(i, 1, 'Success (200)', green_st)
    if status == 400:
        print("Error 400 - Bad Request - ABORT!")
        print("Probably have a bad URL or payload")
        wr_ws.write(i, 1, 'Bad Request (400)', red_st)
        pass
    if status == 401:
        print("Error 401 - Unauthorized - ABORT!")
        print("Probably have incorrect credentials")
        wr_ws.write(i, 1, 'Unauthorized (401)', red_st)
        pass
    if status == 403:
        print("Error 403 - Forbidden - ABORT!")
        print("Server refuses to handle your request")
        wr_ws.write(i, 1, 'Forbidden (403)', red_st)
        pass
    if status == 404:
        print("Error 404 - Not Found - ABORT!")
        print("Seems like you're trying to POST to a page that doesn't"
              " exist.")
        wr_ws.write(i, 1, 'Not Found (400)', red_st)
        pass
    if status == 666:
        print("Error - Something failed!")
        print("The POST failed, see stdout for the exception.")
        wr_ws.write(i, 1, 'Unkown Failure', yellow_st)
        pass
    if status == 667:
        print("Error - Invalid Input!")
        print("Invalid integer or other input.")
        wr_ws.write(i, 1, 'Unkown Failure', yellow_st)
        pass

#======================================================
# The Main Module
#======================================================
def main():
    description = None
    if description is not None:
        Parser.description = description
    Parser.add_argument('-d', '--dir', default='ACI',
                        help='The Directory to Publish the Terraform Files to.'
    )
    Parser.add_argument('-wb', '--workbook', default='ACI_Base_Workbookv2.xlsx',
                        help='The Workbook to read for Input.'
    )
    Parser.add_argument('-ws', '--worksheet', default=None,
                        help='The Workbook to read for Input.'
    )
    args = Parser.parse_args()

    jsonFile = 'templates/variables/easy_variables.json'
    jsonOpen = open(jsonFile, 'r')
    easy_jsonData = json.load(jsonOpen)
    jsonOpen.close()

    # Ask user for required Information: ACI_DEPLOY_FILE
    if os.path.isfile(args.workbook):
        excel_workbook = args.workbook
    else:
        print('\nWorkbook not Found.  Please enter a valid /path/filename for the source workbook you will be using.')
        while True:
            print('Please enter a valid /path/filename for the source you will be using.')
            excel_workbook = input('/Path/Filename: ')
            if os.path.isfile(excel_workbook):
                print(f'\n-----------------------------------------------------------------------------\n')
                print(f'   {excel_workbook} exists.  Will Now Check for API Variables...')
                print(f'\n-----------------------------------------------------------------------------\n')
                break
            else:
                print('\nWorkbook not Found.  Please enter a valid /path/filename for the source you will be using.')

    # Load Workbook
    wb = read_in(excel_workbook)

    # Create Dictionary for Worksheets in the Workbook

    easyDict = {
        'access':{
            'domains':[],
            'firmware':[],
            'global_policies':[],
            'interface_policies':[],
            'inventory':[],
            'leaf_interface_policy_groups':[],
            'leaf_policy_groups':[],
            'leaf_profiles':[],
            'spine_interface_policy_groups':[],
            'spine_policy_groups':[],
            'spine_profiles':[],
            'vmm':[],
            'vpc_domains':[],
        },
        'admin':{
            'authentication':[],
            'configuration_backups':[],
            'global_security':[],
            'radius':[],
            'tacacs':[],
        },
        'fabric':{
            'date_and_time':[],
            'dns_profiles':[],
            'smartcallhome':[],
            'snmp_policies':[],
            'syslog':[],
        },
        'inventory':{},
        'sites':{},
        'system_settings':{
            'apic_connectivity_preference':[],
            'bgp_asn':[],
            'bgp_rr':[],
            'global_aes_encryption_settings':[]
        },
        'tenants':{
            'application_epgs':[],
            'application_profiles':[],
            'bfd_interface_policies':[],
            'bgp_policies':[],
            'bridge_domains':[],
            'dhcp_option_policies':[],
            'dhcp_relay_policies':[],
            'endpoint_retention_policies':[],
            'filters':[],
            'hsrp_policies':[],
            'l3out_floating_svi':[],
            'l3out_hsrp':[],
            'l3out_static_route':[],
            'l3outs':[],
            'ospf_policies':[],
            'route_map_match_rules':[],
            'route_map_set_rules':[],
            'route_maps_for_route_control':[],
            'schemas':[],
            'tenants':[],
            'vrfs':[],
        },
        'wb':wb
    }

    # Run Proceedures for Worksheets in the Workbook
    easyDict = process_sites(easyDict, easy_jsonData, wb)

    # Either Run All Remaining Proceedures or Just Specific based on sys.argv[2:]
    if not args.worksheet == None:
        if re.search('site', str(args.worksheet)):
            process_sites(easyDict, easy_jsonData, wb)
        elif re.search('access', str(args.worksheet)):
            process_access(easyDict, easy_jsonData, wb)
        elif re.search('admin', str(args.worksheet)):
            process_admin(easyDict, easy_jsonData, wb)
        elif re.search('inventory', str(args.worksheet)):
            process_inventory(easyDict, easy_jsonData, wb)
        elif re.search('system_settings', str(args.worksheet)):
            process_system_settings(easyDict, easy_jsonData, wb)
        elif re.search('fabric', str(args.worksheet)):
            process_fabric(easyDict, easy_jsonData, wb)
        elif re.search('tenant', str(args.worksheet)):
            process_tenants(easyDict, easy_jsonData, wb)
        elif re.search('contract', str(args.worksheet)):
            process_contracts(easyDict, easy_jsonData, wb)
        elif re.search('l3out', str(args.worksheet)):
            process_l3out(easyDict, easy_jsonData, wb)
        elif re.search('bd', str(args.worksheet)):
            process_bridge_domains(easyDict, easy_jsonData, wb)
        elif re.search('epg', str(args.worksheet)):
            process_epgs(easyDict, easy_jsonData, wb)
        else:
            print(f'\n-----------------------------------------------------------------------------\n')
            print(f'   {args.worksheet} is not a valid worksheet.  If you are trying to run')
            print(f'   a single worksheet please re-enter the -ws argument.  Exiting...')
            print(f'\n-----------------------------------------------------------------------------\n')
            exit()
    else:
        easyDict = process_system_settings(easyDict, easy_jsonData, wb)
        easyDict = process_fabric(easyDict, easy_jsonData, wb)
        easyDict = process_admin(easyDict, easy_jsonData, wb)
        # easyDict = process_tenants(easyDict, easy_jsonData, wb)
        # easyDict = process_epgs(easyDict, easy_jsonData, wb)
        easyDict.pop('wb')
        print(json.dumps(easyDict, indent = 4))
        exit()
        read_easy_jsonData(easy_jsonData, **easyDict)
        easyDict = process_bridge_domains(easyDict, easy_jsonData, wb)
        easyDict = process_access(easyDict, easy_jsonData, wb)
        easyDict = process_inventory(easyDict, easy_jsonData, wb)
        easyDict = process_l3out(easyDict, easy_jsonData, wb)
        easyDict = process_contracts(easyDict, easy_jsonData, wb)

    folders = check_git_status()
    get_user_pass()
    apply_aci_terraform(folders)
    # else:
    #     print('hello')
    #     path = './'
    #     repo = Repo.init(path)

    #     index = Repo.init(path.index)

    #     index.commit('Testing Commit')

    print(f'\n-----------------------------------------------------------------------------\n')
    print(f'  Proceedures Complete!!! Closing Environment and Exiting Script.')
    print(f'\n-----------------------------------------------------------------------------\n')
    exit()

if __name__ == '__main__':
    main()
