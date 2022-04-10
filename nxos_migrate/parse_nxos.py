#!/usr/bin/env python3

import ipaddress
import openpyxl
import pandas as pd
import numpy as np
import os, re, sys, traceback, validators
from openpyxl import Workbook
from openpyxl.styles import Alignment, colors, Border, Font, NamedStyle, PatternFill, Protection, Side 
from openpyxl.utils.dataframe import dataframe_to_rows

# Define Regular Expressions to be used in function definations and searches
re_bpdu = re.compile('^  spanning-tree bpduguard enable$\n')
re_cdpe = re.compile('^  cdp enable$\n')
re_dhcp = re.compile(r'^  ip dhcp relay address (\d{1,3}.\d{1,3}.\d{1,3}.\d{1,3}) $\n')
re_desc = re.compile('^  description (.+)$\n')
re_host = re.compile('^hostname (.+)$\n')
re_hsv4 = re.compile(r'^    ip (\d{1,3}.\d{1,3}.\d{1,3}.\d{1,3})$\n')
re_hsv4s = re.compile(r'^    ip (\d{1,3}.\d{1,3}.\d{1,3}.\d{1,3}) secondary$\n')
re_intf = re.compile(r'^interface ((port\-channel\d+|Ethernet\d+[\d\/]+))$\n')
re_ipv4 = re.compile(r'^  ip address (\d{1,3}.\d{1,3}.\d{1,3}.\d{1,3}(?:/\d{1,2}|))$\n')
re_ipv4s = re.compile(r'^  ip address (\d{1,3}.\d{1,3}.\d{1,3}.\d{1,3}(?:/\d{1,2}|)) secondary$\n')
re_ivln = re.compile(r'^interface Vlan(\d+)$\n')
re_ldpr = re.compile('^  lldp transmit$\n')
re_ldpt = re.compile('^  lldp receive$\n')
re_mtu_ = re.compile(r'^  mtu (\d+)$\n')
re_nego = re.compile('^  ((no negotiate auto|negotiate auto))$\n')
re_poch = re.compile(r'^  channel-group (\d+) mode ((active|on|passive))$\n')
re_sped = re.compile('^  speed ((auto|[0-9]+))$\n')
re_swav = re.compile(r'^  switchport access vlan (\d+)$\n')
re_swma = re.compile('^  switchport mode access$\n')
re_swmt = re.compile('^  switchport mode trunk$\n')
re_swpt = re.compile('^  switchport$\n')
re_tknv = re.compile(r'^  switchport trunk native vlan (\d{1,4})$\n')
re_tkv1 = re.compile(r'^  switchport trunk allowed vlan (\d{1,4}[\-,]+.+\d{1,4})$\n')
re_tkv2 = re.compile(r'^  switchport trunk allowed vlan (\d{1,4})$\n')
re_vlan = re.compile(r'^vlan (\d{1,4})$\n')
re_vlnm = re.compile('^  name (.+)$\n')
re_vlst = re.compile(r'^vlan (\d{1,4}[\-,]+.+\d{1,4})$\n')
re_vpc_ = re.compile(r'^  vpc ((\d+|peer\-link))$\n')
re_vrf_ = re.compile('^  vrf member (.+)$\n')
re_vrfc = re.compile('^vrf context (.+)$\n')
reipv6m = re.compile('^  ipv6 multicast multipath s-g-hash\n')
def function_expand_vlst(vlst):
    vlist = str_vlst.split(',')
    for v in vlist:
        if re.fullmatch('^\\d{1,4}\\-\\d{1,4}$', v):
            a,b = v.split('-')
            a = int(a)
            b = int(b)
            vrange = range(a,b+1)
            for vl in vrange:
                function_wr_vlan(vl)
        elif re.fullmatch('^\\d{1,4}$', v):
            v = int(v)
            function_wr_vlan(v)
    
def function_vlan_to_bd(vlan):
    if vlan < 10:
        vlan = str(vlan)
        bd = 'v000' + vlan + '_bd'
        return bd
    elif vlan < 100:
        vlan = str(vlan)
        bd = 'v00' + vlan + '_bd'
        return bd
    elif vlan < 1000:
        vlan = str(vlan)
        bd = 'v0' + vlan + '_bd'
        return bd
    else:
        vlan = str(vlan)
        bd = 'v' + vlan + '_bd'
        return bd

def func_wr_poch(str_host, str_intf, str_vpc_, str_mtu_, str_sped, str_swmd, str_swav, str_tknv, str_tkvl, str_desc):
    wr_poch.write('%s\t%s\t%s\t%s\t%s\t%s\t%s\t%s\t%s\t%s\n' % (str_host, str_intf, str_vpc_, str_mtu_, str_sped,
                   str_swmd, str_swav, str_tknv, str_tkvl, str_desc))

def function_wr_name(vlan,name):
    if vlan < 10:
        wr_name.write('v000%s_bd\t%s\n' % (vlan, name))
    elif vlan < 100:
        vlan = str(vlan)
        wr_name.write('v00%s_bd\t%s\n' % (vlan, name))
    elif vlan < 1000:
        vlan = str(vlan)
        wr_name.write('v0%s_bd\t%s\n' % (vlan, name))
    else:
        vlan = str(vlan)
        wr_name.write('v%s_bd\t%s\n' % (vlan, name))

def function_wr_vlan(vlan):
    if vlan < 10:
        vlan = str(vlan)
        wr_vlan.write('v000%s_bd\n' % (vlan))
    elif vlan < 100:
        vlan = str(vlan)
        wr_vlan.write('v00%s_bd\n' % (vlan))
    elif vlan < 1000:
        vlan = str(vlan)
        wr_vlan.write('v0%s_bd\n' % (vlan))
    else:
        vlan = str(vlan)
        wr_vlan.write('v%s_bd\n' % (vlan))

# Start by Creating Default Variables
str_bpdg = 'no'
str_cdp_ = 'no'
str_dhcp = ''
str_desc = ''
str_host = ''
str_hsv4 = ''
str_hsv4s = ''
str_intf = ''
str_ipv4 = ''
str_ipv4s = ''
str_ivln = ''
str_lldr = 'no'
str_lldt = 'no'
str_mtu_ = ''
str_nego = 'negotiate auto'
str_poch = 'n/a'
str_pomd = 'n/a'
str_sped = 'auto'
str_swav = 'n/a'
str_swmd = 'access'
str_swpt = 'no'
str_tknv = 'n/a'
str_tkvl = 'n/a'
str_vlan = ''
str_vlst = ''
str_vlnm = ''
str_vpc_ = 'n/a'
str_vrf_ = 'default'
str_vrfc = ''

# Import the Configuration File
config_file = sys.argv[1]
try:
    if os.path.isfile(config_file):
        print(f'\n-----------------------------------------------------------------------------\n')
        print(f'   {config_file} exists.  Beginning Script Execution...')
        print(f'\n-----------------------------------------------------------------------------\n')
    else:
        print(f'\n-----------------------------------------------------------------------------\n')
        print(f'   {config_file} does not exist.  Exiting....')
        print(f'\n-----------------------------------------------------------------------------\n')
        exit()
except IOError:
    print(f'\n-----------------------------------------------------------------------------\n')
    print(f'   {config_file} does not exist.  Exiting....')
    print(f'\n-----------------------------------------------------------------------------\n')
    exit()

file = open(config_file, 'r')
wr_dhcp = open('dhcp.csv', 'w')
wr_poch = open('int_poch.csv', 'w')
wr_name = open('vlan_name.csv', 'w')
wr_vlan = open('vlan_list.csv', 'w')
wr_vrf = open('vrf_list.csv', 'w')


bd1 = Side(style="thick", color="8EA9DB")
bd2 = Side(style="medium", color="8EA9DB")
wsh1 = NamedStyle(name="wsh1")
wsh1.alignment = Alignment(horizontal="center", vertical="center", wrap_text="True")
wsh1.border = Border(left=bd1, top=bd1, right=bd1, bottom=bd1)
wsh1.font = Font(bold=True, size=15, color="FFFFFF")
wsh2 = NamedStyle(name="wsh2")
wsh2.alignment = Alignment(horizontal="center", vertical="center", wrap_text="True")
wsh2.border = Border(left=bd2, top=bd2, right=bd2, bottom=bd2)
wsh2.fill = PatternFill("solid", fgColor="305496")
wsh2.font = Font(bold=True, size=15, color="FFFFFF")
ws_odd = NamedStyle(name="ws_odd")
ws_odd.alignment = Alignment(horizontal="center", vertical="center")
ws_odd.border = Border(left=bd2, top=bd2, right=bd2, bottom=bd2)
ws_odd.fill = PatternFill("solid", fgColor="D9E1F2")
ws_odd.font = Font(bold=False, size=12, color="44546A")
ws_even = NamedStyle(name="ws_even")
ws_even.alignment = Alignment(horizontal="center", vertical="center")
ws_even.border = Border(left=bd2, top=bd2, right=bd2, bottom=bd2)
ws_even.font = Font(bold=False, size=12, color="44546A")


wb = Workbook()
wb.add_named_style(wsh1)
wb.add_named_style(wsh2)
wb.add_named_style(ws_odd)
wb.add_named_style(ws_even)

dest_file = 'export.xlsx'
ws1 = wb.active
ws1.title = "Tenant"
ws2 = wb.create_sheet(title = "VRF")
ws3 = wb.create_sheet(title = "Bridge Domain")
ws4 = wb.create_sheet(title = "Subnet")
ws5 = wb.create_sheet(title = "DHCP Relay")
ws6 = wb.create_sheet(title = "Access Interfaces")
ws7 = wb.create_sheet(title = "Static Port Mappings")

ws2 = wb["VRF"]
ws3 = wb["Bridge Domain"]
ws4 = wb["Subnet"]
ws5 = wb["DHCP Relay"]
ws6 = wb["Access Interfaces"]
ws6 = wb["Static Port Mappings"]
ws1.column_dimensions['A'].width = 15
ws1.column_dimensions['B'].width = 20
ws1.column_dimensions['C'].width = 40
ws2.column_dimensions['A'].width = 15
ws2.column_dimensions['B'].width = 20
ws2.column_dimensions['C'].width = 20
ws2.column_dimensions['D'].width = 40
ws3.column_dimensions['A'].width = 15
ws3.column_dimensions['B'].width = 20
ws3.column_dimensions['C'].width = 20
ws3.column_dimensions['D'].width = 20
ws3.column_dimensions['E'].width = 15
ws3.column_dimensions['F'].width = 22
ws3.column_dimensions['G'].width = 40
ws4.column_dimensions['A'].width = 15
ws4.column_dimensions['B'].width = 20
ws4.column_dimensions['C'].width = 20
ws4.column_dimensions['D'].width = 20
ws4.column_dimensions['E'].width = 20
ws4.column_dimensions['F'].width = 20
ws4.column_dimensions['G'].width = 25
ws4.column_dimensions['H'].width = 20
ws4.column_dimensions['I'].width = 40
ws5.column_dimensions['A'].width = 15
ws5.column_dimensions['B'].width = 20
ws5.column_dimensions['C'].width = 20
ws5.column_dimensions['D'].width = 20
ws5.column_dimensions['E'].width = 40
ws6.column_dimensions['A'].width = 10
ws6.column_dimensions['B'].width = 20
ws6.column_dimensions['C'].width = 20
ws6.column_dimensions['D'].width = 20
ws6.column_dimensions['E'].width = 15
ws6.column_dimensions['F'].width = 12
ws6.column_dimensions['G'].width = 10
ws6.column_dimensions['H'].width = 10
ws6.column_dimensions['I'].width = 10
ws6.column_dimensions['J'].width = 15
ws6.column_dimensions['K'].width = 14
ws6.column_dimensions['L'].width = 13
ws6.column_dimensions['M'].width = 40
ws6.column_dimensions['N'].width = 12

data = ['Type','Tenant Name','Description']
ws1.append(data)
for cell in ws1["1:1"]:
    cell.style = 'wsh2'
data = ['Type','Tenant','VRF Name','Description']
ws2.append(data)
for cell in ws2["1:1"]:
    cell.style = 'wsh2'
data = ['Type','Tenant','VRF Name','Bridge Domain','Extend Outside ACI','Preferred or vzAny','Description']
ws3.append(data)
for cell in ws3["1:1"]:
    cell.style = 'wsh2'
data = ['Type','Tenant','VRF Name','Bridge Domain','Gateway IPv4','Gateway_Type','Advertise External', 'Layer3 Out','Description']
ws4.append(data)
for cell in ws4["1:1"]:
    cell.style = 'wsh2'
data = ['Type','Tenant','VRF Name','IPv4 Address','Description']
ws5.append(data)
for cell in ws5["1:1"]:
    cell.style = 'wsh2'
data = ['Type','Current Host','Current Interface','Access Entity Profile','Policy Group Name','LACP Policy','MTU','Speed',\
        'CDP Enabled','LLDP Receive','LLDP Transmit','BPDU Guard','Port-Channel Description','Port Description']
ws6.append(data)
for cell in ws6["1:1"]:
    cell.style = 'wsh2'
ws1_row_count = 2
ws2_row_count = 2
ws3_row_count = 2
ws4_row_count = 2
ws5_row_count = 2
ws6_row_count = 2

# Read the Conifguration File and Gather Vlan Information
lines = file.readlines()

line_count = 0
ethn_count = 0
for line in lines:
    if re.fullmatch(re_host, line):
        str_host = re.fullmatch(re_host, line).group(1)
        line_count += 1
    elif re.fullmatch(re_vlst, line):
        # Matched the VLAN List... Now Parse for Data Export
        str_vlst = re.fullmatch(re_vlst, line).group(1)
        # Expand VLAN Ranges into Full VLAN List
        function_expand_vlst(str_vlst)
        line_count += 1
    elif re.fullmatch(re_vlan, line):
        # Matched a VLAN... Now Parse for Data Export
        str_vlan = int(re.fullmatch(re_vlan, line).group(1))
        line_count += 1
    elif re.fullmatch(re_vlnm, line):
        # Matched VLAN Name... Now Parse for Data Export
        str_vlnm = re.fullmatch(re_vlnm, line).group(1)
        function_wr_name(str_vlan,str_vlnm)
        line_count += 1
    elif re.fullmatch(re_vrfc, line):
        str_vrfc = re.fullmatch(re_vrfc, line).group(1)
        line_count += 1
    elif re.fullmatch(reipv6m, line):
        data = ['vrf_add','',str_vrfc,str_desc]
        ws2.append(data)
        rc = '%s:%s' % (ws2_row_count, ws2_row_count)
        for cell in ws2[rc]:
            if ws2_row_count % 2 == 0:
                cell.style = 'ws_even'
            else:
                cell.style = 'ws_odd'
        ws2_row_count += 1
        line_count += 1
        str_desc = ''
        str_vrfc = ''
        line_count += 1

    elif re.fullmatch(re_ivln, line):
        # Matched an Interface VLAN... Now Parse for Data Export
        str_ivln = int(re.fullmatch(re_ivln, line).group(1))
        line_count += 1
    elif re.fullmatch(re_mtu_, line):
        # Matched the Interface MTU... Now Parse for Data Export
        str_mtu_ = re.fullmatch(re_mtu_, line).group(1)
        line_count += 1
    elif re.fullmatch(re_sped, line):
        # Matched the Interface Speed... Now Parse for Data Export
        str_sped = re.fullmatch(re_sped, line).group(1)
        line_count += 1
    elif re.fullmatch(re_nego, line):
        # Matched the Interface Negotiate Mode... Now Parse for Data Export
        str_nego = re.fullmatch(re_nego, line).group(1)
        line_count += 1
    elif re.fullmatch(re_vrf_, line):
        # Matched a VRF Context... Now Parse for Data Export
        str_vrf_ = re.fullmatch(re_vrf_, line).group(1)
        line_count += 1
    elif re.fullmatch(re_ipv4, line):
        # Matched an IPv4 Address/prefix... Now Parse for Data Export
        str_ipv4 = re.fullmatch(re_ipv4, line).group(1)
        line_count += 1
    elif re.fullmatch(re_ipv4s, line):
        # Matched an IPv4 Secondary Address/prefix... Now Parse for Data Export
        str_ipv4s = re.fullmatch(re_ipv4s, line).group(1)
        line_count += 1
    elif re.fullmatch(re_hsv4, line):
        # Matched an HSRP IPv4 Address... Now Parse for Data Export
        str_hsv4 = re.fullmatch(re_hsv4, line).group(1)
        line_count += 1
    elif re.fullmatch(re_hsv4s, line):
        # Matched an HSRP IPv4 Secondary Address/prefix... Now Parse for Data Export
        str_hsv4s = re.fullmatch(re_hsv4s, line).group(1)
        line_count += 1
    elif re.fullmatch(re_dhcp, line):
        # Matched an IPv4 DHCP Relay definition... Now Parse for Data Export
        str_dhcp = re.fullmatch(re_dhcp, line).group(1)
        wr_dhcp.write('%s,%s\n' % (str_vrf_, str_dhcp))
        line_count += 1
    elif re.fullmatch(re_intf, line):
        str_intf = re.fullmatch(re_intf, line).group(1)
        line_count += 1
    elif re.fullmatch(re_bpdu, line):
        str_bpdg = 'BPDU_fg'
        line_count += 1
    elif re.fullmatch(re_cdpe, line):
        str_cdp_ = 'yes'
        line_count += 1
    elif re.fullmatch(re_ldpr, line):
        str_lldr = 'yes'
        line_count += 1
    elif re.fullmatch(re_ldpt, line):
        str_lldt = 'yes'
        line_count += 1
    elif re.fullmatch(re_swav, line):
        str_swav = re.fullmatch(re_swav, line).group(1)
        line_count += 1
    elif re.fullmatch(re_swma, line):
        str_swmd = 'access'
        line_count += 1
    elif re.fullmatch(re_swmt, line):
        str_swmd = 'trunk'
        line_count += 1
    elif re.fullmatch(re_tknv, line):
        str_tknv = re.fullmatch(re_tknv, line).group(1)
        line_count += 1
    elif re.fullmatch(re_tkv1, line):
        str_tkvl = re.fullmatch(re_tkv1, line).group(1)
        line_count += 1
    elif re.fullmatch(re_tkv2, line):
        str_tkvl = re.fullmatch(re_tkv2, line).group(1)
        line_count += 1
    elif re.fullmatch(re_swpt, line):
        str_swpt = 'yes'
        line_count += 1
    elif re.fullmatch(re_poch, line):
        str_poch = re.fullmatch(re_poch, line).group(1)
        str_pomd = re.fullmatch(re_poch, line).group(2)
        line_count += 1
    elif re.fullmatch(re_vpc_, line):
        str_vpc_ = re.fullmatch(re_vpc_, line).group(1)
        line_count += 1
    elif re.fullmatch(re_desc, line):
        # Found a Description on the Interface
        str_desc = re.fullmatch(re_desc, line).group(1)
        line_count += 1
    elif line == "\n":
        # Found blank line, which means the end of the interface, time to create the output
        if str_ipv4 and str_ivln:
            bd = function_vlan_to_bd(str_ivln)
            bd = bd.strip()
            if str_hsv4:
                a,b = str_ipv4.split('/')
                gtwy = str(str_hsv4) + '/' + str(b)
            else:
                gtwy = str(str_ipv4)
            data = ['subnet_add','',str_vrf_,bd,gtwy,'primary','','',str_desc]
            ws4.append(data)
            rc = '%s:%s' % (ws4_row_count, ws4_row_count)
            for cell in ws4[rc]:
                if ws4_row_count % 2 == 0:
                    cell.style = 'ws_even'
                else:
                    cell.style = 'ws_odd'
            ws4_row_count += 1
            wr_vrf.write('%s,%s\n' % (bd, str_vrf_))
            if str_ipv4s:
                if str_hsv4s:
                    a,b = str_ipv4s.split('/')
                    gtwy = str(str_hsv4s) + '/' + str(b)
                else:
                    gtwy = str(str_ipv4)
                data = ['subnet_add','',str_vrf_,bd,gtwy,'secondary','','',str_desc]
                ws4.append(data)
                rc = '%s:%s' % (ws4_row_count, ws4_row_count)
                for cell in ws4[rc]:
                    if ws4_row_count % 2 == 0:
                        cell.style = 'ws_even'
                    else:
                        cell.style = 'ws_odd'
                ws4_row_count += 1
            line_count += 1
        elif 'channel' in str_intf:
            if str_swpt == 'yes':
                mtu1 = 9000
                mtu2 = int(str_mtu_)
                if mtu2 >= mtu1:
                    str_mtu_ = '9000'
                func_wr_poch(str_host, str_intf, str_vpc_, str_mtu_, str_sped, str_swmd, str_swav, str_tknv, str_tkvl, str_desc)
        elif 'Ethernet' in str_intf:
            if ethn_count == 0:
                wr_poch.close()
                read_poch = open('int_poch.csv', 'r')
                po_lines = read_poch.readlines()
                ethn_count += 1
            if str_swpt == 'yes':
                mtu1 = 9000
                mtu2 = int(str_mtu_)
                if mtu2 >= mtu1:
                    str_mtu_ = '9000'
                if str_nego == 'no negotiate auto':
                    str_nego = 'noNeg'
                else:
                    str_nego = 'Auto'
                if str_sped == '100':
                    str_sped = '100M_%s' % (str_nego)
                elif str_sped == '1000':
                    str_sped = '1G_%s' % (str_nego)
                elif str_sped == '2500':
                    str_sped = '2.5G_%s' % (str_nego)
                elif str_sped == '5000':
                    str_sped = '5G_%s' % (str_nego)
                elif str_sped == '10000':
                    str_sped = '10G_%s' % (str_nego)
                elif str_sped == '25000':
                    str_sped = '25G_%s' % (str_nego)
                elif str_sped == '40000':
                    str_sped = '40G_%s' % (str_nego)
                elif str_sped == '50000':
                    str_sped = '50G_%s' % (str_nego)
                elif str_sped == '100000':
                    str_sped = '100G_%s' % (str_nego)
                elif str_sped == '200000':
                    str_sped = '200G_%s' % (str_nego)
                elif str_sped == '400000':
                    str_sped = '400G_%s' % (str_nego)
                else:
                    str_sped = 'inherit_%s' % (str_nego)
                if re.search(r'(\d+|peer)', str_poch):
                    for line in po_lines:
                        x = line.split('\t')
                        desc = x[8].strip()
                        y = x[1].split('l')
                        if str_poch == y[1]:
                            if str_swmd == 'access':
                                swav = x[5]
                            else:
                                swav = x[6]
                            if x[2] == 'n/a':
                                type = 'pcg_add'
                            else:
                                type = 'vpc_add'
                            data = [type,'','',str_host,str_intf,str_poch,x[2],x[3],str_sped,x[5],swav,x[8],str_cdp_,str_lldr,
                                    str_lldt,str_bpdg,desc,str_desc]
                            ws6.append(data)
                            rc = '%s:%s' % (ws6_row_count, ws6_row_count)
                            for cell in ws6[rc]:
                                if ws6_row_count % 2 == 0:
                                    cell.style = 'ws_even'
                                else:
                                    cell.style = 'ws_odd'
                            ws6_row_count += 1
                else:
                    type = 'apg_add'
                    if str_swmd == 'access':
                        swav = str_swav
                    else:
                        swav = str_tknv
                    data = [type,'','',str_host,str_intf,str_poch,str_vpc_,str_mtu_,str_sped,str_swmd,swav,str_tkvl,str_cdp_,str_lldr,
                            str_lldt,str_bpdg,'n/a',str_desc]
                    ws6.append(data)
                    rc = '%s:%s' % (ws6_row_count, ws6_row_count)
                    for cell in ws6[rc]:
                        if ws6_row_count % 2 == 0:
                            cell.style = 'ws_even'
                        else:
                            cell.style = 'ws_odd'
                    ws6_row_count += 1
        
        # Reset the Variables back to Blank except str_host
        str_bpdg = 'no'
        str_cdp_ = 'no'
        str_dhcp = ''
        str_desc = ''
        str_hsv4 = ''
        str_hsv4s = ''
        str_intf = ''
        str_ipv4 = ''
        str_ipv4s = ''
        str_ivln = ''
        str_lldr = 'no'
        str_lldt = 'no'
        str_mtu_ = ''
        str_nego = 'negotiate auto'
        str_poch = 'n/a'
        str_pomd = 'n/a'
        str_sped = 'auto'
        str_swav = 'n/a'
        str_swmd = 'access'
        str_swpt = 'no'
        str_tknv = 'n/a'
        str_tkvl = 'n/a'
        str_vlan = ''
        str_vlst = ''
        str_vlnm = ''
        str_vpc_ = 'n/a'
        str_vrf_ = 'default'
        str_vrfc = ''
        line_count += 1
    else:
        line_count += 1


file.close()
wr_vlan.close()
wr_name.close()
wr_dhcp.close()
wr_vrf.close()

#Get VLAN's that don't have a name and those that do and combine into one file
bg_list1 = open('vlan_list.csv', 'r') 
bg_list2 = open('vlan_name.csv', 'r')
bg_list3 = open('vlan_comb.csv', 'w')
vlan_lines = bg_list1.readlines()
name_lines = bg_list2.readlines()
for lineg1 in vlan_lines:
    matched = 0
    lineg1 = lineg1.strip()
    for lineg2 in name_lines:
        lineg2 = lineg2.strip()
        if lineg1 in lineg2:
            matched +=1
    if matched == 0:
        bg_list3.write('%s\n' % (lineg1))
for line in name_lines:
    line.strip()
    bg_list3.write('%s' % (line))

bg_list1.close()
bg_list2.close()
bg_list3.close()

#Sort the combined VLANs in final output file
bg_list3 = open('vlan_comb.csv', 'r')
vrf_list = open('vrf_list.csv', 'r')
bddm = bg_list3.readlines()
vrfl = vrf_list.readlines()
bddm.sort()
for line in range(len(bddm)):
    bddm[line]
    if re.search('\t', bddm[line]):
        bd,descr = bddm[line].split('\t')
    else:
        bd = bddm[line]
        descr = ''
    bd = bd.strip()
    descr = descr.strip()
    vrf_bd = ''
    for x in vrfl:
        x = x.strip()
        y = x.split(',')
        if y[0] == bd:
            vrf_bd = y[1]
    if vrf_bd == '':
        vrf_bd = 'default'
    data = ['nca_bd','',vrf_bd,bd,'yes',descr]
    ws3.append(data)
    rc = '%s:%s' % (ws3_row_count, ws3_row_count)
    for cell in ws3[rc]:
        if ws3_row_count % 2 == 0:
            cell.style = 'ws_even'
        else:
            cell.style = 'ws_odd'
    ws3_row_count += 1

bg_list3.close()
vrf_list.close()

dhcp_relay_uniq = 'cat dhcp.csv | sort | uniq > dhcp_sort.csv'
os.system(dhcp_relay_uniq)

file_relays = open('dhcp_sort.csv', 'r')
read_relays = file_relays.readlines()
for line in read_relays:
    vrf,relay_ip = line.split(',')
    data = ['dhcp_relay','',vrf,relay_ip]
    ws5.append(data)
    rc = '%s:%s' % (ws5_row_count, ws5_row_count)
    for cell in ws5[rc]:
        if ws5_row_count % 2 == 0:
            cell.style = 'ws_even'
        else:
            cell.style = 'ws_odd'
    ws5_row_count += 1
file_relays.close()
remove_extra_files = 'rm dhcp.csv dhcp_sort.csv int_poch.csv vlan_comb.csv vlan_list.csv vlan_name.csv vrf_list.csv'
os.system(remove_extra_files)

#for row in range(2,ws4.max_row+1):
#    for column in 'CD':
#        cell_name = "%s%s" % (column,row)
#        #print(cell_name)
#        print(ws4[cell_name].value)

# Save the Excel Workbook
wb.save(dest_file)

if not str_host == '':
    rename_excel = 'mv export.xlsx %s_export.xlsx' % (str_host)
    os.system(rename_excel)

#End Script
print(f'\n-----------------------------------------------------------------------------\n')
print(f'   Completed Running Script.  Exiting....')
print(f'\n-----------------------------------------------------------------------------\n')
exit()