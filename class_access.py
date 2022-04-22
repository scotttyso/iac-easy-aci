#!/usr/bin/env python3

import ast
import jinja2
import os
import pkg_resources
import re
import validating
from class_terraform import terraform_cloud
from easy_functions import countKeys, create_selector, create_tf_file
from easy_functions import findKeys, findVars
from easy_functions import process_kwargs, query_module_type, query_switch_model
from easy_functions import sensitive_var_site_group, stdout_log
from easy_functions import vlan_list_full
from easy_functions import write_to_site
from easy_functions import write_to_template
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

aci_template_path = pkg_resources.resource_filename('class_access', 'templates/')

# Exception Classes
class InsufficientArgs(Exception):
    pass

class ErrException(Exception):
    pass

class InvalidArg(Exception):
    pass

class LoginFailed(Exception):
    pass

# Terraform ACI Provider - Access Policies
# Class must be instantiated with Variables
class access(object):
    def __init__(self, type):
        self.templateLoader = jinja2.FileSystemLoader(
            searchpath=(aci_template_path + '%s/') % (type))
        self.templateEnv = jinja2.Environment(loader=self.templateLoader)
        self.type = type

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def aep_profile(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Infra_VLAN': ''}
        optional_args = {'Description': '',
                         'Physical_Domains': '',
                         'L3_Domains': '',
                         'VMM_Domains': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.values(row_num, ws, 'Infra_VLAN', templateVars['Infra_VLAN'], ['no', 'yes'])
            if not templateVars['Physical_Domains'] == None:
                templateVars['domains'] = 'yes'
                if re.search(r',', templateVars['Physical_Domains']):
                    x = templateVars['Physical_Domains'].split(',')
                    for domain in x:
                        validating.name_rule(row_num, ws, 'Physical_Domains', domain)
                else:
                    validating.name_rule(row_num, ws, 'Physical_Domains', templateVars['Physical_Domains'])
            if not templateVars['L3_Domains'] == None:
                templateVars['domains'] = 'yes'
                if re.search(r',', templateVars['L3_Domains']):
                    x = templateVars['L3_Domains'].split(',')
                    for domain in x:
                        validating.name_rule(row_num, ws, 'L3_Domains', domain)
                else:
                    validating.name_rule(row_num, ws, 'L3_Domains', templateVars['L3_Domains'])
            if not templateVars['VMM_Domains'] == None:
                templateVars['domains'] = 'yes'
                if re.search(r',', templateVars['VMM_Domains']):
                    x = templateVars['VMM_Domains'].split(',')
                    for domain in x:
                        validating.name_rule(row_num, ws, 'VMM_Domains', domain)
                else:
                    validating.name_rule(row_num, ws, 'VMM_Domains', templateVars['VMM_Domains'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        templateVars['phys_count'] = 0
        templateVars['l3_count'] = 0
        templateVars['vmm_count'] = 0
        if not templateVars['Physical_Domains'] == None:
            if re.search(r',', templateVars['Physical_Domains']):
                x = templateVars['Physical_Domains'].split(',')
                templateVars['Physical_Domains'] = []
                for domain in x:
                    templateVars['Physical_Domains'].append(domain)
                    templateVars['phys_count'] =+ 1
            else:
                templateVars['Physical_Domains'] = [templateVars['Physical_Domains']]
                templateVars['phys_count'] =+ 1
        if not templateVars['L3_Domains'] == None:
            if re.search(r',', templateVars['L3_Domains']):
                x = templateVars['L3_Domains'].split(',')
                templateVars['L3_Domains'] = []
                for domain in x:
                    templateVars['L3_Domains'].append(domain)
                    templateVars['l3_count'] =+ 1
            else:
                templateVars['L3_Domains'] = [templateVars['L3_Domains']]
                templateVars['l3_count'] =+ 1
        if not templateVars['VMM_Domains'] == None:
            if re.search(r',', templateVars['VMM_Domains']):
                x = templateVars['VMM_Domains'].split(',')
                templateVars['VMM_Domains'] = []
                for domain in x:
                    templateVars['VMM_Domains'].append(domain)
                    templateVars['vmm_count'] =+ 1
            else:
                templateVars['VMM_Domains'] = [templateVars['VMM_Domains']]
                templateVars['vmm_count'] =+ 1

        # Define the Template Source
        template_file = "global_aep.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Policies_Global_AEP_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def apic_inb(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Node_ID': '',
                         'Pod_ID': '',
                         'Inband_EPG': ''}
        optional_args = {'Inband_IPv4': '',
                         'Inband_GWv4': '',
                         'Inband_IPv6': '',
                         'Inband_GWv6': '',}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        # Configure the Generic Template Variables
        templateVars['Device_Type'] = 'apic'
        templateVars['Type'] = 'in_band'
        templateVars['EPG'] = templateVars['Inband_EPG']
        templateVars['IPv4'] = templateVars['Inband_IPv4']
        templateVars['GWv4'] = templateVars['Inband_GWv4']
        templateVars['IPv6'] = templateVars['Inband_IPv6']
        templateVars['GWv6'] = templateVars['Inband_GWv6']

        # Initialize the Class
        lib_aci_ref = 'Access_Policies'
        class_init = '%s(ws)' % (lib_aci_ref)

        # Assign the APIC Inband Management IP's
        eval("%s.%s(wb, ws, row_num, **templateVars)" % (class_init, 'mgmt_static'))

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def cdp(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Admin_State': ''}
        optional_args = {'Description': '',
                         'Alias': ''}
        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.values(row_num, ws, 'Admin_State', templateVars['Admin_State'], ['disabled', 'enabled'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)


        # Define the Template Source
        template_file = "policy_intf_cdp.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Policies_Interface_CDP_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def fibre_channel(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Port_Mode': '',
                         'Trunk_Mode': '',
                         'Speed': '',
                         'Auto_Max_Speed': '',
                         'Fill_Pattern': '',
                         'Buffer_Credit': ''}
        optional_args = {'Description': '',
                         'Alias': ''}
        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.number_check(row_num, ws, 'Buffer_Credit', templateVars['Buffer_Credit'], 16, 64)
            validating.values(row_num, ws, 'Port_Mode', templateVars['Port_Mode'], ['f', 'np'])
            validating.values(row_num, ws, 'Trunk_Mode', templateVars['Trunk_Mode'], ['auto', 'trunk-off', 'trunk-on'])
            validating.values(row_num, ws, 'Speed', templateVars['Speed'], ['auto', '4G', '8G', '16G', '32G'])
            validating.values(row_num, ws, 'Auto_Max_Speed', templateVars['Auto_Max_Speed'], ['4G', '8G', '16G', '32G'])
            validating.values(row_num, ws, 'Fill_Pattern', templateVars['Fill_Pattern'], ['ARBFF', 'IDLE'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)


        # Define the Template Source
        template_file = "policy_intf_fc.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Policies_Interface_FC_Interface_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def intf_profile(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Switch_Role': '',
                         'Name': '',
                         'Dest_Folder': ''}
        optional_args = {'Description': ''}
        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.name_rule(row_num, ws, 'Dest_Folder', templateVars['Dest_Folder'])
            validating.values(row_num, ws, 'Switch_Role', templateVars['Switch_Role'], ['leaf', 'spine'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        if templateVars['Switch_Role'] == 'leaf':
            # Define the Template Source
            template_file = "leaf_interface_profile.jinja2"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = '%s_Interface_Profile.tf' % (templateVars['Name'])
        elif templateVars['Switch_Role'] == 'spine':
            # Define the Template Source
            template_file = "spine_interface_profile.jinja2"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = '%s_Interface_Profile.tf' % (templateVars['Name'])

        if not templateVars['Dest_Folder'] == None:
            dest_dir = '%s' % (templateVars['Dest_Folder'])
        else:
            dest_dir = 'Access'

        write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def intf_selector(self, wb, ws, row_num, wr_file, **kwargs):
        if not kwargs.get('Policy_Group') == None:
            # Dicts for required and optional args
            required_args = {'Site_Group': '',
                             'Site_Name': '',
                             'Switch_Role': '',
                             'Pod_ID': '',
                             'Node_ID': '',
                             'Interface_Profile': '',
                             'Interface_Selector': '',
                             'Port': '',
                             'Policy_Group': '',
                             'Port_Type': ''}
            optional_args = {'LACP_Policy': '',
                             'Bundle_ID': '',
                             'Description': '',
                             'Switchport_Mode': '',
                             'Access_or_Native': '',
                             'Trunk_Allowed_VLANs': ''}

            kwargs['Switch_Site'] = kwargs.get('Site_Group')
            if not kwargs.get('Port_Type') == None:
                if re.search('(port-channel|vpc)', kwargs.get('Port_Type')):

                    temp_descr = kwargs['Description']
                    # Open the Access Worksheet and Find the Policy Group
                    ws_pg = wb['Access']
                    rows = ws_pg.max_row
                    row_bundle = ''
                    func = 'pg_access'
                    count = countKeys(ws_pg, func)
                    var_dict = findVars(ws_pg, func, rows, count)
                    for pos in var_dict:
                        if var_dict[pos].get('Name') == kwargs.get('Policy_Group'):
                            row_bundle = var_dict[pos]['row']
                            del var_dict[pos]['row']
                            kwargs = {**kwargs, **var_dict[pos]}
                            break

                    # Open the Network Policies Worksheet to get the Interface_Policy
                    ws_net = wb['Network Policies']
                    rows = ws_net.max_row

                    # Get the Interface Policies from the Network Policies Tab
                    func = 'intf_polgrp'
                    count = countKeys(ws_net, func)
                    row_pg = ''
                    var_dict = findVars(ws_net, func, rows, count)
                    for pos in var_dict:
                        if var_dict[pos].get('Policy_Name') == kwargs.get('Interface_Policy'):
                            row_pg = var_dict[pos]['row']
                            del var_dict[pos]['row']
                            kwargs = {**kwargs, **var_dict[pos]}
                            break

                    # Validate inputs, return dict of template vars

                    if kwargs.get('Port_Type') == 'vpc':
                        kwargs['Lag_Type'] = 'node'
                        ws_vpc = wb['Inventory']
                        for row in ws_vpc.rows:
                            if row[0].value == 'vpc_pair' and int(row[1].value) == int(kwargs.get('Switch_Site')) and str(row[5].value) == str(kwargs.get('Node_ID')):
                                kwargs['VPC_Name'] = row[2].value
                                kwargs['Name'] = '%s_vpc%s' % (row[2].value, kwargs.get('Bundle_ID'))

                            elif row[0].value == 'vpc_pair' and str(row[1].value) == str(kwargs.get('Switch_Site')) and str(row[6].value) == str(kwargs.get('Node_ID')):
                                kwargs['VPC_Name'] = row[2].value
                                kwargs['Name'] = '%s_vpc%s' % (row[2].value, kwargs.get('Bundle_ID'))
                    elif kwargs.get('Port_Type') == 'port-channel':
                        kwargs['Lag_Type'] = 'link'
                        kwargs['Name'] = '%s_pc%s' % (kwargs.get('Interface_Profile'), kwargs.get('Bundle_ID'))

                    kwargs['Description'] = temp_descr
                    # Create the Bundle Policy Group
                    kwargs['Site_Group'] = kwargs.get('Switch_Site')
                    lib_aci_ref = 'Access_Policies'
                    class_init = '%s(ws)' % (lib_aci_ref)
                    func = 'pg_bundle'
                    eval("%s.%s(wb, ws, row_num, **kwargs)" % (class_init, func))

            # Validate inputs, return dict of template vars
            templateVars = process_kwargs(required_args, optional_args, **kwargs)
            # leafx = Name
            xa = templateVars['Port'].split('/')
            xcount = len(xa)
            templateVars['Module_From'] = xa[0]
            templateVars['Module_To'] = xa[0]
            templateVars['Port_From'] = xa[1]
            templateVars['Port_To'] = xa[1]
            templateVars['Selector_Type'] = 'range'

            if templateVars['Switch_Role'] == 'leaf':
                templateVars['Policy_Group'] = kwargs.get('Name')
                # Define the Template Source
                template_file = "leaf_portselect.jinja2"
                template = self.templateEnv.get_template(template_file)

                # Process the template and write to file
                payload = template.render(templateVars)
                wr_file.write(payload + '\n\n')

                # Define the Template Source
                if xcount == 3:
                    templateVars['SubPort_From'] = xa[2]
                    templateVars['SubPort_To'] = xa[2]
                    template_file = "leaf_portblock_sub.jinja2"
                else:
                    template_file = "leaf_portblock.jinja2"
                template = self.templateEnv.get_template(template_file)

                # Process the template and write to file
                payload = template.render(templateVars)
                wr_file.write(payload + '\n\n')

            elif templateVars['Switch_Role'] == 'spine':
                # Define the Template Source
                template_file = "spine_portselect.jinja2"
                template = self.templateEnv.get_template(template_file)

                # Process the template and write to file
                payload = template.render(templateVars)
                wr_file.write(payload + '\n\n')

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def l2_interface(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'QinQ': '',
                         'Reflective_Relay': '',
                         'VLAN_Scope': ''}
        optional_args = {'Description': '',
                         'Alias': ''}
        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.values(row_num, ws, 'QinQ', templateVars['QinQ'], ['disabled', 'enabled'])
            validating.values(row_num, ws, 'Reflective_Relay', templateVars['Reflective_Relay'], ['disabled', 'enabled'])
            validating.values(row_num, ws, 'VLAN_Scope', templateVars['VLAN_Scope'], ['global', 'portlocal'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)


        # Define the Template Source
        template_file = "policy_intf_l2_interface.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Policies_Interface_L2_Interface_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def l3_domain(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'VLAN_Pool': ''}
        optional_args = { }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.name_rule(row_num, ws, 'VLAN_Pool', templateVars['VLAN_Pool'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)


        # Define the Template Source
        template_file = "domain_l3.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Policies_Domain_L3_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def leaf_pg(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Auth_8021X': '',
                         'BFD_IPv4': '',
                         'BFD_IPv6': '',
                         'BFD_MH_IPv4': '',
                         'BFD_MH_IPv6': '',
                         'CDP_Policy': '',
                         'CoPP_Leaf_Policy': '',
                         'CoPP_Pre_Filter': '',
                         'Flash_Config': '',
                         'Fast_Link_Failover': '',
                         'FC_SAN_Policy': '',
                         'FC_Node_Policy': '',
                         'Forward_Scale': '',
                         'LLDP_Policy': '',
                         'Monitoring_Policy': '',
                         'Netflow_Node': '',
                         'PoE_Policy': '',
                         'STP_Policy': ''}
        optional_args = {'Description': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.name_rule(row_num, ws, 'Auth_8021X', templateVars['Auth_8021X'])
            validating.name_rule(row_num, ws, 'BFD_IPv4', templateVars['BFD_IPv4'])
            validating.name_rule(row_num, ws, 'BFD_IPv6', templateVars['BFD_IPv6'])
            validating.name_rule(row_num, ws, 'BFD_MH_IPv4', templateVars['BFD_MH_IPv4'])
            validating.name_rule(row_num, ws, 'BFD_MH_IPv6', templateVars['BFD_MH_IPv6'])
            validating.name_rule(row_num, ws, 'CDP_Policy', templateVars['CDP_Policy'])
            validating.name_rule(row_num, ws, 'CoPP_Leaf_Policy', templateVars['CoPP_Leaf_Policy'])
            validating.name_rule(row_num, ws, 'CoPP_Pre_Filter', templateVars['CoPP_Pre_Filter'])
            validating.name_rule(row_num, ws, 'Flash_Config', templateVars['Flash_Config'])
            validating.name_rule(row_num, ws, 'Fast_Link_Failover', templateVars['Fast_Link_Failover'])
            validating.name_rule(row_num, ws, 'FC_SAN_Policy', templateVars['FC_SAN_Policy'])
            validating.name_rule(row_num, ws, 'FC_Node_Policy', templateVars['FC_Node_Policy'])
            validating.name_rule(row_num, ws, 'Forward_Scale', templateVars['Forward_Scale'])
            validating.name_rule(row_num, ws, 'LLDP_Policy', templateVars['LLDP_Policy'])
            validating.name_rule(row_num, ws, 'Monitoring_Policy', templateVars['Monitoring_Policy'])
            validating.name_rule(row_num, ws, 'Netflow_Node', templateVars['Netflow_Node'])
            validating.name_rule(row_num, ws, 'PoE_Policy', templateVars['PoE_Policy'])
            validating.name_rule(row_num, ws, 'STP_Policy', templateVars['STP_Policy'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "leaf_policy_group.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Leaf_Policy_Group_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def link_level(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Auto_Neg': '',
                         'Speed': '',
                         'Port_Delay': '',
                         'Debounce_Interval': '',
                         'FEC_Mode': ''}
        optional_args = {'Description': '',
                         'Alias': ''}
        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.number_check(row_num, ws, 'Port_Delay', templateVars['Port_Delay'], 0, 10000)
            validating.number_check(row_num, ws, 'Debounce_Interval', templateVars['Debounce_Interval'], 0, 5000)
            validating.values(row_num, ws, 'Auto_Neg', templateVars['Auto_Neg'], ['off', 'on'])
            validating.values(row_num, ws, 'Speed', templateVars['Speed'], ['inherit', '100M', '1G', '10G', '25G', '40G', '50G', '100G', '200G', '400G'])
            validating.values(row_num, ws, 'FEC_Mode', templateVars['FEC_Mode'], ['inherit', 'auto-fec', 'cl74-fc-fec', 'cl91-rs-fec', 'cons16-rs-fec', 'disable-fec', 'ieee-rs-fec', 'kp-fec'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)


        # Define the Template Source
        template_file = "policy_intf_link_level.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Policies_Interface_Link_Level_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def lldp(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Receive_State': '',
                         'Transmit_State': ''}
        optional_args = {'Description': '',
                         'Alias': ''}
        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.values(row_num, ws, 'Receive_State', templateVars['Receive_State'], ['disabled', 'enabled'])
            validating.values(row_num, ws, 'Transmit_State', templateVars['Transmit_State'], ['disabled', 'enabled'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)


        # Define the Template Source
        template_file = "policy_intf_lldp.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Policies_Interface_LLDP_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def mcp(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Admin_State': ''}
        optional_args = {'Description': '',
                         'Alias': ''}
        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.values(row_num, ws, 'Admin_State', templateVars['Admin_State'], ['disabled', 'enabled'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)


        # Define the Template Source
        template_file = "policy_intf_mcp.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Policies_Interface_MCP_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def mgmt_static(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Node_ID': '',
                         'Pod_ID': '',
                         'Device_Type': '',
                         'Type': '',
                         'EPG': ''}
        optional_args = {'IPv4': '',
                         'GWv4': '',
                         'IPv6': '',
                         'GWv6': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.hostname(row_num, ws, 'Name', templateVars['Name'])
            validating.name_rule(row_num, ws, 'EPG', templateVars['EPG'])
            validating.number_check(row_num, ws, 'Pod_ID', templateVars['Pod_ID'], 1, 15)
            if templateVars['Device_Type'] == 'apic':
                validating.number_check(row_num, ws, 'Node_ID', templateVars['Node_ID'], 1, 7)
            else:
                validating.number_check(row_num, ws, 'Node_ID', templateVars['Node_ID'], 101, 4001)
            if not templateVars['IPv4'] == None:
                validating.mgmt_network(row_num, ws, 'IPv4', templateVars['IPv4'], 'GWv4', templateVars['GWv4'])
            if not templateVars['IPv6'] == None:
                validating.mgmt_network(row_num, ws, 'IPv6', templateVars['IPv6'], 'GWv6', templateVars['GWv6'])
            else:
                templateVars['IPv6'] = '::'
                templateVars['GWv6'] = '::'
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "static_node_mgmt_address.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = '%s_%s_EPG_%s_Static_Address.tf' % (templateVars['Name'], templateVars['Type'], templateVars['EPG'])
        dest_dir = 'Tenant_mgmt'
        write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def pg_access(self, wb, ws, row_num, **kwargs):
        # Open the Network Policies Worksheet
        ws_net = wb['Network Policies']
        rows = ws_net.max_row

        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'AEP_Policy': '',
                         'CDP_Policy': '',
                         'Link_Level': '',
                         'LLDP_Policy': '',
                         'MCP_Policy': '',
                         'STP_Policy': '',
                         'Interface_Policy': '',
                         'Policy_Name': ''}
        optional_args = {'Description': '',
                         'Alias': '',
                         'Pol_802_1x': '',
                         'poeIfPol': '',
                         'monFabricPol': '',
                         'dwdmIfPol': '',
                         'coppIfPol': '',
                         'qosDppPol_egress': '',
                         'qosDppPol_ingress': '',
                         'Fibre_Channel': '',
                         'L2_Interface': '',
                         'fabricLinkFlapPol': '',
                         'qosLlfcIfPol': '',
                         'macsecIfPol': '',
                         'netflowMonitorPol': '',
                         'Port_Security': '',
                         'qosPfcIfPol': '',
                         'qosSdIfPol': '',
                         'stormctrlIfPol': ''}

        # Get the Application Profile Policies from the Network Policies Tab
        func = 'intf_polgrp'
        count = countKeys(ws_net, func)
        row_pg = ''
        var_dict = findVars(ws_net, func, rows, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('Interface_Policy'):
                row_pg = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}
                break

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.name_rule(row_num, ws, 'AEP_Policy', templateVars['AEP_Policy'])
            validating.name_rule(row_num, ws, 'CDP_Policy', templateVars['CDP_Policy'])
            validating.name_rule(row_num, ws, 'Link_Level', templateVars['Link_Level'])
            validating.name_rule(row_num, ws, 'LLDP_Policy', templateVars['LLDP_Policy'])
            validating.name_rule(row_num, ws, 'MCP_Policy', templateVars['MCP_Policy'])
            validating.name_rule(row_num, ws, 'STP_Policy', templateVars['STP_Policy'])
            if not templateVars['Fibre_Channel'] == None:
                validating.name_rule(row_pg, ws_net, 'Fibre_Channel', templateVars['Fibre_Channel'])
            validating.name_rule(row_pg, ws_net, 'L2_Interface', templateVars['L2_Interface'])
            validating.name_rule(row_pg, ws_net, 'Port_Security', templateVars['Port_Security'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
            if not templateVars['Pol_802_1x'] == None:
                validating.name_rule(row_num, ws, 'Pol_802_1x', templateVars['Pol_802_1x'])
                templateVars['Pol_802_1x'] = 'uni/infra/portauthpol-%s' % (templateVars['Pol_802_1x'])
            if not templateVars['poeIfPol'] == None:
                validating.name_rule(row_num, ws, 'poeIfPol', templateVars['poeIfPol'])
                templateVars['poeIfPol'] = 'uni/infra/poeIfP-%s' % (templateVars['poeIfPol'])
            if not templateVars['monFabricPol'] == None:
                validating.name_rule(row_num, ws, 'monFabricPol', templateVars['monFabricPol'])
                templateVars['monFabricPol'] = 'uni/fabric/monfab-%s' % (templateVars['monFabricPol'])
            if not templateVars['dwdmIfPol'] == None:
                validating.name_rule(row_num, ws, 'dwdmIfPol', templateVars['dwdmIfPol'])
                templateVars['dwdmIfPol'] = 'uni/infra/dwdmifpol-%s' % (templateVars['dwdmIfPol'])
            if not templateVars['coppIfPol'] == None:
                validating.name_rule(row_pg, ws_net, 'coppIfPol', templateVars['coppIfPol'])
                templateVars['coppIfPol'] = 'uni/infra/coppifpol-%s' % (templateVars['coppIfPol'])
            if not templateVars['qosDppPol_egress'] == None:
                validating.name_rule(row_pg, ws_net, 'qosDppPol_egress', templateVars['qosDppPol_egress'])
                templateVars['qosDppPol_egress'] = 'uni/infra/qosdpppol-%s' % (templateVars['qosDppPol_egress'])
            if not templateVars['qosDppPol_ingress'] == None:
                validating.name_rule(row_pg, ws_net, 'qosDppPol_ingress', templateVars['qosDppPol_ingress'])
                templateVars['qosDppPol_ingress'] = 'uni/infra/qosdpppol-%s' % (templateVars['qosDppPol_ingress'])
            if not templateVars['fabricLinkFlapPol'] == None:
                validating.name_rule(row_pg, ws_net, 'fabricLinkFlapPol', templateVars['fabricLinkFlapPol'])
                templateVars['fabricLinkFlapPol'] = 'uni/infra/linkflappol-%s' % (templateVars['fabricLinkFlapPol'])
            if not templateVars['qosLlfcIfPol'] == None:
                validating.name_rule(row_pg, ws_net, 'qosLlfcIfPol', templateVars['qosLlfcIfPol'])
                templateVars['qosLlfcIfPol'] = 'uni/infra/llfc-%s' % (templateVars['qosLlfcIfPol'])
            if not templateVars['macsecIfPol'] == None:
                validating.name_rule(row_pg, ws_net, 'macsecIfPol', templateVars['macsecIfPol'])
                templateVars['macsecIfPol'] = 'uni/infra/macsecifp-%s' % (templateVars['macsecIfPol'])
            if not templateVars['netflowMonitorPol'] == None:
                validating.name_rule(row_pg, ws_net, 'netflowMonitorPol', templateVars['netflowMonitorPol'])
                templateVars['netflowMonitorPol'] = 'uni/infra/poeIfP-%s' % (templateVars['netflowMonitorPol'])
            if not templateVars['qosPfcIfPol'] == None:
                validating.name_rule(row_pg, ws_net, 'qosPfcIfPol', templateVars['qosPfcIfPol'])
                templateVars['qosPfcIfPol'] = 'uni/infra/pfc-%s' % (templateVars['qosPfcIfPol'])
            if not templateVars['qosSdIfPol'] == None:
                validating.name_rule(row_pg, ws_net, 'qosSdIfPol', templateVars['qosSdIfPol'])
                templateVars['qosSdIfPol'] = 'uni/infra/qossdpol-%s' % (templateVars['qosSdIfPol'])
            if not templateVars['stormctrlIfPol'] == None:
                validating.name_rule(row_pg, ws_net, 'stormctrlIfPol', templateVars['stormctrlIfPol'])
                templateVars['stormctrlIfPol'] = 'uni/infra/stormctrlifp-%s' % (templateVars['stormctrlIfPol'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "leaf_intf_pg_access.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Leaf_Interface_PG_Access_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def pg_bundle(self, wb, ws, row_num, **kwargs):
        # Assign the kwargs to a initial var for each process
        initial_kwargs = kwargs

        # Open the Network Policies Worksheet
        ws_net = wb['Network Policies']
        rows = ws_net.max_row

        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Lag_Type': '',
                         'AEP_Policy': '',
                         'CDP_Policy': '',
                         'Link_Level': '',
                         'LACP_Policy': '',
                         'LLDP_Policy': '',
                         'MCP_Policy': '',
                         'STP_Policy': '',
                         'Interface_Policy': '',
                         'Policy_Name': ''}
        optional_args = {'Description': '',
                         'Alias': '',
                         'monFabricPol': '',
                         'coppIfPol': '',
                         'qosDppPol_egress': '',
                         'qosDppPol_ingress': '',
                         'Fibre_Channel': '',
                         'L2_Interface': '',
                         'fabricLinkFlapPol': '',
                         'qosLlfcIfPol': '',
                         'macsecIfPol': '',
                         'netflowMonitorPol': '',
                         'Port_Security': '',
                         'qosPfcIfPol': '',
                         'qosSdIfPol': '',
                         'stormctrlIfPol': ''}

        # Get the Application Profile Policies from the Network Policies Tab
        func = 'intf_polgrp'
        count = countKeys(ws_net, func)
        row_pg = ''
        var_dict = findVars(ws_net, func, rows, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('Interface_Policy'):
                row_pg = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}
                break

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.name_rule(row_num, ws, 'AEP_Policy', templateVars['AEP_Policy'])
            validating.name_rule(row_num, ws, 'CDP_Policy', templateVars['CDP_Policy'])
            validating.name_rule(row_num, ws, 'Link_Level', templateVars['Link_Level'])
            validating.name_rule(row_num, ws, 'LACP_Policy', templateVars['LACP_Policy'])
            validating.name_rule(row_num, ws, 'LLDP_Policy', templateVars['LLDP_Policy'])
            validating.name_rule(row_num, ws, 'MCP_Policy', templateVars['MCP_Policy'])
            validating.name_rule(row_num, ws, 'STP_Policy', templateVars['STP_Policy'])
            if not templateVars['Fibre_Channel'] == None:
                validating.name_rule(row_pg, ws_net, 'Fibre_Channel', templateVars['Fibre_Channel'])
            validating.name_rule(row_pg, ws_net, 'L2_Interface', templateVars['L2_Interface'])
            validating.name_rule(row_pg, ws_net, 'Port_Security', templateVars['Port_Security'])
            validating.values(row_num, ws, 'Lag_Type', templateVars['Lag_Type'], ['link', 'node'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
            if not templateVars['monFabricPol'] == None:
                validating.name_rule(row_num, ws, 'monFabricPol', templateVars['monFabricPol'])
                templateVars['monFabricPol'] = 'uni/fabric/monfab-%s' % (templateVars['monFabricPol'])
            if not templateVars['coppIfPol'] == None:
                validating.name_rule(row_pg, ws_net, 'coppIfPol', templateVars['coppIfPol'])
                templateVars['coppIfPol'] = 'uni/infra/coppifpol-%s' % (templateVars['coppIfPol'])
            if not templateVars['qosDppPol_egress'] == None:
                validating.name_rule(row_pg, ws_net, 'qosDppPol_egress', templateVars['qosDppPol_egress'])
                templateVars['qosDppPol_egress'] = 'uni/infra/qosdpppol-%s' % (templateVars['qosDppPol_egress'])
            if not templateVars['qosDppPol_ingress'] == None:
                validating.name_rule(row_pg, ws_net, 'qosDppPol_ingress', templateVars['qosDppPol_ingress'])
                templateVars['qosDppPol_ingress'] = 'uni/infra/qosdpppol-%s' % (templateVars['qosDppPol_ingress'])
            if not templateVars['fabricLinkFlapPol'] == None:
                validating.name_rule(row_pg, ws_net, 'fabricLinkFlapPol', templateVars['fabricLinkFlapPol'])
                templateVars['fabricLinkFlapPol'] = 'uni/infra/linkflappol-%s' % (templateVars['fabricLinkFlapPol'])
            if not templateVars['qosLlfcIfPol'] == None:
                validating.name_rule(row_pg, ws_net, 'qosLlfcIfPol', templateVars['qosLlfcIfPol'])
                templateVars['qosLlfcIfPol'] = 'uni/infra/llfc-%s' % (templateVars['qosLlfcIfPol'])
            if not templateVars['macsecIfPol'] == None:
                validating.name_rule(row_pg, ws_net, 'macsecIfPol', templateVars['macsecIfPol'])
                templateVars['macsecIfPol'] = 'uni/infra/macsecifp-%s' % (templateVars['macsecIfPol'])
            if not templateVars['netflowMonitorPol'] == None:
                validating.name_rule(row_pg, ws_net, 'netflowMonitorPol', templateVars['netflowMonitorPol'])
                templateVars['netflowMonitorPol'] = 'uni/infra/poeIfP-%s' % (templateVars['netflowMonitorPol'])
            if not templateVars['qosPfcIfPol'] == None:
                validating.name_rule(row_pg, ws_net, 'qosPfcIfPol', templateVars['qosPfcIfPol'])
                templateVars['qosPfcIfPol'] = 'uni/infra/pfc-%s' % (templateVars['qosPfcIfPol'])
            if not templateVars['qosSdIfPol'] == None:
                validating.name_rule(row_pg, ws_net, 'qosSdIfPol', templateVars['qosSdIfPol'])
                templateVars['qosSdIfPol'] = 'uni/infra/qossdpol-%s' % (templateVars['qosSdIfPol'])
            if not templateVars['stormctrlIfPol'] == None:
                validating.name_rule(row_pg, ws_net, 'stormctrlIfPol', templateVars['stormctrlIfPol'])
                templateVars['stormctrlIfPol'] = 'uni/infra/stormctrlifp-%s' % (templateVars['stormctrlIfPol'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "leaf_intf_pg_bundle.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Leaf_Interface_PG_Bundle_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def pg_breakout(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Breakout_Map': ''}
        optional_args = {'Description': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.values(row_num, ws, 'Breakout_Map', templateVars['Breakout_Map'], ['100g-2x', '100g-4x', '10g-4x', '25g-4x', '50g-8x'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "leaf_intf_pg_breakout.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Leaf_Interface_PG_Breakout_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def pg_spine(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'AEP_Policy': '',
                         'CDP_Policy': '',
                         'Link_Level': ''}
        optional_args = {'Description': '',
                         'Alias': '',
                         'fabricLinkFlapPol': '',
                         'macsecIfPol': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.name_rule(row_num, ws, 'AEP_Policy', templateVars['AEP_Policy'])
            validating.name_rule(row_num, ws, 'CDP_Policy', templateVars['CDP_Policy'])
            validating.name_rule(row_num, ws, 'Link_Level', templateVars['Link_Level'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
            if not templateVars['fabricLinkFlapPol'] == None:
                validating.name_rule(row_num, ws, 'fabricLinkFlapPol', templateVars['fabricLinkFlapPol'])
                templateVars['fabricLinkFlapPol'] = 'uni/infra/linkflappol-%s' % (templateVars['fabricLinkFlapPol'])
            if not templateVars['macsecIfPol'] == None:
                validating.name_rule(row_num, ws, 'macsecIfPol', templateVars['macsecIfPol'])
                templateVars['macsecIfPol'] = 'uni/infra/macsecifpol-%s' % (templateVars['macsecIfPol'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "spine_intf_pg_access.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Spine_Interface_PG_Access_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def phys_dom(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'VLAN_Pool': ''}
        optional_args = { }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.name_rule(row_num, ws, 'VLAN_Pool', templateVars['VLAN_Pool'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)


        # Define the Template Source
        template_file = "domain_phys.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Policies_Domain_Phys_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def port_channel(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Mode': '',
                         'Min_Links': '',
                         'Max_Links': '',
                         'Fast_Select': '',
                         'Graceful': '',
                         'Load_Defer': '',
                         'Suspend_Individual': '',
                         'Symmetric_Hash': ''}
        optional_args = {'Description': '',
                         'Alias': ''}
        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.number_check(row_num, ws, 'Min_Links', templateVars['Min_Links'], 1, 16)
            validating.number_check(row_num, ws, 'Max_Links', templateVars['Max_Links'], 1, 16)
            validating.values(row_num, ws, 'Mode', templateVars['Mode'], ['active', 'explicit-failover', 'mac-pin', 'mac-pin-nicload', 'off', 'passive'])
            validating.values(row_num, ws, 'Fast_Select', templateVars['Fast_Select'], ['no', 'yes'])
            validating.values(row_num, ws, 'Graceful', templateVars['Graceful'], ['no', 'yes'])
            validating.values(row_num, ws, 'Load_Defer', templateVars['Load_Defer'], ['no', 'yes'])
            validating.values(row_num, ws, 'Suspend_Individual', templateVars['Suspend_Individual'], ['no', 'yes'])
            validating.values(row_num, ws, 'Symmetric_Hash', templateVars['Symmetric_Hash'], ['no', 'yes'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Create ctrl templateVars
        ctrl_count = 0
        Ctrl = ''
        if templateVars['Fast_Select'] == 'yes':
            Ctrl = '"fast-sel-hot-stdby"'
            ctrl_count =+ 1
        if templateVars['Graceful'] == 'yes' and ctrl_count > 0:
            Ctrl = Ctrl + ',' + '"graceful-conv"'
            ctrl_count =+ 1
        elif templateVars['Graceful'] == 'yes':
            ctrl = '"graceful-conv"'
            ctrl_count =+ 1
        if templateVars['Load_Defer'] == 'yes' and ctrl_count > 0:
            Ctrl = Ctrl + ',' + '"load-defer"'
            ctrl_count =+ 1
        elif templateVars['Load_Defer'] == 'yes':
            Ctrl = '"load-defer"'
            ctrl_count =+ 1
        if templateVars['Suspend_Individual'] == 'yes' and ctrl_count > 0:
            Ctrl = Ctrl + ',' + '"susp-individual"'
            ctrl_count =+ 1
        elif templateVars['Suspend_Individual'] == 'yes':
            Ctrl = '"susp-individual"'
            ctrl_count =+ 1
        if templateVars['Symmetric_Hash'] == 'yes' and ctrl_count > 0:
            Ctrl = Ctrl + ',' + '"symmetric-hash"'
            ctrl_count =+ 1
        elif templateVars['Symmetric_Hash'] == 'yes':
            Ctrl = '"symmetric-hash"'
            ctrl_count =+ 1
        if ctrl_count > 0:
            templateVars['Ctrl'] = '[%s]' % (Ctrl)
        else:
            templateVars['Ctrl'] = '["unspecified"]'

        # Define the Template Source
        template_file = "policy_intf_lacp.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Policies_Interface_Port_Channel_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def port_cnvt(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Node_ID': '',
                         'Port': ''}
        optional_args = { }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.hostname(row_num, ws, 'Name', templateVars['Name'])
            validating.number_check(row_num, ws, 'Node_ID', templateVars['Node_ID'], 101, 4001)
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Create Port Name Var
        zz = templateVars['Port'].split('/')
        templateVars['Port_Name'] = '%s_%s' % (zz[0], zz[1])

        # Define the Template Source
        template_file = "downlink.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Downlink_Convert_%s.tf' % (templateVars['Port_Name'])
        dest_dir = 'Access/%s' % (templateVars['Name'])
        write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def port_security(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Timeout': '',
                         'Maximum_Endpoints': ''}
        optional_args = {'Description': '',
                         'Alias': ''}
        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.number_check(row_num, ws, 'Timeout', templateVars['Timeout'], 60, 3600)
            validating.number_check(row_num, ws, 'Maximum_Endpoints', templateVars['Maximum_Endpoints'], 0, 12000)
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)


        # Define the Template Source
        template_file = "policy_intf_port_sec.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Policies_Interface_Port_Security_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def spine_pg(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'BFD_IPv4': '',
                         'BFD_IPv6': '',
                         'CDP_Policy': '',
                         'CoPP_Pre_Filter': '',
                         'CoPP_Spine_Policy': '',
                         'LLDP_Policy': ''}
        optional_args = {'Description': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.name_rule(row_num, ws, 'BFD_IPv4', templateVars['BFD_IPv4'])
            validating.name_rule(row_num, ws, 'BFD_IPv6', templateVars['BFD_IPv6'])
            validating.name_rule(row_num, ws, 'CDP_Policy', templateVars['CDP_Policy'])
            validating.name_rule(row_num, ws, 'CoPP_Pre_Filter', templateVars['CoPP_Pre_Filter'])
            validating.name_rule(row_num, ws, 'CoPP_Spine_Policy', templateVars['CoPP_Spine_Policy'])
            validating.name_rule(row_num, ws, 'LLDP_Policy', templateVars['LLDP_Policy'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "spine_policy_group.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Spine_Policy_Group_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def stp(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Filter': '',
                         'Guard': ''}
        optional_args = {'Description': '',
                         'Alias': ''}
        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.values(row_num, ws, 'Filter', templateVars['Filter'], ['disabled', 'enabled'])
            validating.values(row_num, ws, 'Guard', templateVars['Guard'], ['disabled', 'enabled'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Create ctrl templateVars
        ctrl_count = 0
        Ctrl = ''
        if templateVars['Filter'] == 'enabled':
            Ctrl = '"bpdu-filter"'
            ctrl_count =+ 1
        if templateVars['Guard'] == 'enabled' and ctrl_count > 0:
            Ctrl = Ctrl + ',' + '"bpdu-guard"'
            ctrl_count =+ 1
        elif templateVars['Guard'] == 'enabled':
            Ctrl = '"bpdu-guard"'
            ctrl_count =+ 1
        if ctrl_count > 0:
            templateVars['Ctrl'] = '[%s]' % (Ctrl)
        else:
            templateVars['Ctrl'] = '["unspecified"]'

        # Define the Template Source
        template_file = "policy_intf_stp.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Policies_Interface_STP_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def switch(self, wb, ws, row_num, **kwargs):
        # Initialize the Class
        lib_aci_ref = 'Access_Policies'
        class_init = '%s(ws)' % (lib_aci_ref)

        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Serial': '',
                         'Node_ID': '',
                         'Name': '',
                         'Profiles': '',
                         'Node_Type': '',
                         'Pod_ID': '',
                         'Switch_Role': '',
                         'Switch_Type': '',
                         'Is_Virtual': '',
                         'Tier-2': '',
                         'Inband_EPG': '',
                         'OOB_EPG': ''}
        optional_args = {'Policy_Group': '',
                         'Remote_ID': '',
                         'Fabric_ID': '',
                         'MG_Name': '',
                         'Inband_IPv4': '',
                         'Inband_GWv4': '',
                         'Inband_IPv6': '',
                         'Inband_GWv6': '',
                         'OOB_IPv4': '',
                         'OOB_GWv4': '',
                         'OOB_IPv6': '',
                         'OOB_GWv6': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        # Use Switch_Type to Determine the Number of ports on the switch
        modules,port_count = query_switch_model(row_num, templateVars['Switch_Type'])

        try:
            # Validate Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.hostname(row_num, ws, 'Name', templateVars['Name'])
            validating.modules(row_num, templateVars['Name'], templateVars['Switch_Role'], modules)
            if not templateVars['MG_Name'] == None:
                validating.name_rule(row_num, ws, 'MG_Name', templateVars['MG_Name'])
            validating.name_rule(row_num, ws, 'Inband_EPG', templateVars['Inband_EPG'])
            validating.name_rule(row_num, ws, 'OOB_EPG', templateVars['OOB_EPG'])
            validating.number_check(row_num, ws, 'Node_ID', templateVars['Node_ID'], 101, 4001)
            validating.number_check(row_num, ws, 'Pod_ID', templateVars['Pod_ID'], 1, 12)
            validating.number_check(row_num, ws, 'Fabric_ID', templateVars['Fabric_ID'], 1, 12)
            validating.port_count(row_num, templateVars['Name'], templateVars['Switch_Role'], port_count)
            validating.values(row_num, ws, 'Profiles', templateVars['Profiles'], ['no', 'yes'])
            validating.values(row_num, ws, 'Node_Type', templateVars['Node_Type'], ['remote-leaf-wan', 'unspecified'])
            validating.values(row_num, ws, 'Switch_Role', templateVars['Switch_Role'], ['leaf', 'spine'])
            validating.values(row_num, ws, 'Is_Virtual', templateVars['Is_Virtual'], ['no', 'yes'])
            validating.values(row_num, ws, 'Tier-2', templateVars['Tier-2'], ['no', 'yes'])
            if templateVars['Profiles'] == 'yes':
                validating.name_rule(row_num, ws, 'Policy_Group', templateVars['Policy_Group'])
            if not templateVars['Remote_ID'] == None:
                validating.number_check(row_num, ws, 'Remote_ID', templateVars['Remote_ID'], 1, 255)
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Configure the Generic Template Variables for Node Management Inband IP's
        templateVars['Device_Type'] = 'switch'
        templateVars['Type'] = 'in_band'
        templateVars['EPG'] = templateVars['Inband_EPG']
        templateVars['IPv4'] = templateVars['Inband_IPv4']
        templateVars['GWv4'] = templateVars['Inband_GWv4']
        templateVars['IPv6'] = templateVars['Inband_IPv6']
        templateVars['GWv6'] = templateVars['Inband_GWv6']

        if re.search('\d', templateVars['IPv4']) or re.search('\d', templateVars['IPv6']):
            # Assign the Switch Inband Management IP's
            eval("%s.%s(wb, ws, row_num, **templateVars)" % (class_init, 'mgmt_static'))

        # Configure the Generic Template Variables for Node Management Out-of-Band IP's
        templateVars['Device_Type'] = 'switch'
        templateVars['Type'] = 'out_of_band'
        templateVars['EPG'] = templateVars['OOB_EPG']
        templateVars['IPv4'] = templateVars['OOB_IPv4']
        templateVars['GWv4'] = templateVars['OOB_GWv4']
        templateVars['IPv6'] = templateVars['OOB_IPv6']
        templateVars['GWv6'] = templateVars['OOB_GWv6']

        if re.search('\d', templateVars['IPv4']) or re.search('\d', templateVars['IPv6']):
            # Assign the Switch Out-of-Band Management IP's
            eval("%s.%s(wb, ws, row_num, **templateVars)" % (class_init, 'mgmt_static'))

        if not templateVars['MG_Name'] == None:
            # Define the Template Source
            template_file = "maint_group_nodeblk.jinja2"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = 'Maintenance_Group_%s.tf' % (templateVars['MG_Name'])
            dest_dir = 'Admin'
            write_to_site(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

        Site_ID = 'Site_ID_%s' % (templateVars['Site_Group'])
        site_dict = ast.literal_eval(os.environ[Site_ID])

        # Create kwargs for Site Variables
        templateVars['Site_ID'] = site_dict.get('Site_ID')
        templateVars['Site_Name'] = site_dict.get('Site_Name')
        templateVars['APIC_URL'] = site_dict.get('APIC_URL')
        templateVars['APIC_Version'] = site_dict.get('APIC_Version')
        templateVars['APIC_Auth_Type'] = site_dict.get('APIC_Auth_Type')
        templateVars['Terraform_EQ'] = site_dict.get('Terraform_EQ')
        templateVars['Terraform_Version'] = site_dict.get('Terraform_Version')
        templateVars['Provider_EQ'] = site_dict.get('Provider_EQ')
        templateVars['Provider_Version'] = site_dict.get('Provider_Version')
        templateVars['Run_Location'] = site_dict.get('Run_Location')
        templateVars['State_Location'] = site_dict.get('State_Location')
        templateVars['Terraform_Cloud_Org'] = site_dict.get('Terraform_Cloud_Org')
        templateVars['Workspace_Prefix'] = site_dict.get('Workspace_Prefix')
        templateVars['VCS_Base_Repo'] = site_dict.get('VCS_Base_Repo')
        templateVars['Terraform_Agent_Pool_ID'] = site_dict.get('Terraform_Agent_Pool_ID')

        self.templateLoader = jinja2.FileSystemLoader(searchpath=(aci_template_path + 'Access_Policies/'))
        self.templateEnv = jinja2.Environment(loader=self.templateLoader)
        excel_wkbook = '%s_intf_selectors.xlsx' % (templateVars['Site_Name'])

        wb_sw = load_workbook(excel_wkbook)

        # Check if there is a Worksheet for the Switch Already
        if not templateVars['Name'] in wb_sw.sheetnames:
            ws_sw = wb_sw.create_sheet(title = templateVars['Name'])
            ws_sw = wb_sw[templateVars['Name']]
            ws_sw.column_dimensions['A'].width = 15
            ws_sw.column_dimensions['B'].width = 10
            ws_sw.column_dimensions['c'].width = 10
            ws_sw.column_dimensions['d'].width = 20
            ws_sw.column_dimensions['E'].width = 20
            ws_sw.column_dimensions['F'].width = 10
            ws_sw.column_dimensions['G'].width = 20
            ws_sw.column_dimensions['H'].width = 20
            ws_sw.column_dimensions['I'].width = 20
            ws_sw.column_dimensions['J'].width = 15
            ws_sw.column_dimensions['K'].width = 30
            ws_sw.column_dimensions['L'].width = 20
            ws_sw.column_dimensions['M'].width = 20
            ws_sw.column_dimensions['N'].width = 30
            dv1 = DataValidation(type="list", formula1='"intf_selector"', allow_blank=True)
            dv2 = DataValidation(type="list", formula1='"access,breakout,port-channel,vpc"', allow_blank=True)
            ws_sw.add_data_validation(dv1)
            ws_sw.add_data_validation(dv2)
            ws_header = '%s Interface Selectors' % (templateVars['Name'])
            data = [ws_header]
            ws_sw.append(data)
            ws_sw.merge_cells('A1:N1')
            for cell in ws_sw['1:1']:
                cell.style = 'Heading 1'
            data = ['','Notes: Breakout Policy Group Names are 2x100g_pg, 4x10g_pg, 4x25g_pg, 4x100g_pg, 8x50g_pg.']
            ws_sw.append(data)
            ws_sw.merge_cells('B2:N2')
            for cell in ws_sw['2:2']:
                cell.style = 'Heading 2'
            data = ['Type','Pod_ID','Node_ID','Interface_Profile','Interface_Selector','Port','Policy_Group','Port_Type','LACP_Policy','Bundle_ID','Description','Switchport_Mode','Access_or_Native','Trunk_Allowed_VLANs']
            ws_sw.append(data)
            for cell in ws_sw['3:3']:
                cell.style = 'Heading 3'

            ws_sw_row_count = 4
            templateVars['dv1'] = dv1
            templateVars['dv2'] = dv2
            templateVars['port_count'] = port_count
            sw_type = str(templateVars['Switch_Type'])
            sw_name = str(templateVars['Name'])
            if re.search('^(93[0-9][0-9])', sw_type):
                for module in range(1, 2):
                    templateVars['module'] = module
                    ws_sw_row_count = create_selector(ws_sw, ws_sw_row_count, **templateVars)
            if re.search('^(9396|95[0-1][4-8])', sw_type):
                row_count = 1
                for row in ws.rows:
                    if re.search('9396', sw_type):
                        start, end = 2, 2
                    else:
                        start, end = 1, int(modules)
                    if str(row[0].value) == sw_type and str(row[2].value) == sw_name:
                        for module in range(start, end + 2):
                            templateVars['module'] = module
                            module_type = row[module + 2].value
                            if module_type == None:
                                module_type = 'none'
                            elif re.search('(X97|M(4|6|12)P)', module_type):
                                templateVars['port_count'] = query_module_type(row_count, module_type)
                                ws_sw_row_count = create_selector(ws_sw, ws_sw_row_count, **templateVars)
                        row_count += 1
                        break
            wb_sw.save(excel_wkbook)
        else:
            ws_sw = wb_sw[templateVars['Name']]

        # Define the Template Source
        template_file = "inventory.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = '%s.tf' % (templateVars['Name'])
        dest_dir = '%s' % (templateVars['Name'])
        write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        if templateVars['Profiles'] == 'yes':
            templateVars['Description'] = None
            templateVars['Dest_Folder'] = templateVars['Name']
            eval("%s.%s(wb, ws, row_num, **templateVars)" % (class_init, 'intf_profile'))

            templateVars['Selector_Name'] = templateVars['Name']
            templateVars['Association_Type'] = 'range'
            templateVars['Nodeblk_Name'] = 'blk%s-%s' % (templateVars['Node_ID'], templateVars['Node_ID'])
            templateVars['Node_ID_From'] = templateVars['Node_ID']
            templateVars['Node_ID_To'] = templateVars['Node_ID']
            templateVars['Interface_Profile'] = templateVars['Name']
            eval("%s.%s(wb, ws, row_num, **templateVars)" % (class_init, 'sw_profile'))

            sw_intf_profile = './ACI/%s/%s/%s_Interface_Profile.tf' % (templateVars['Site_Name'], templateVars['Name'], templateVars['Name'])
            wr_file = open(sw_intf_profile, 'a+')
            lib_aci_ref = 'Access_Policies'
            rows_sw = ws_sw.max_row
            func_regex = re.compile('^intf_selector$')
            func_list = findKeys(ws_sw, func_regex)
            class_init = '%s(ws_sw)' % (lib_aci_ref)
            stdout_log(ws_sw, None)
            for func in func_list:
                count = countKeys(ws_sw, func)
                var_dict = findVars(ws_sw, func, rows_sw, count)
                for pos in var_dict:
                    row_num = var_dict[pos]['row']
                    del var_dict[pos]['row']
                    for x in list(var_dict[pos].keys()):
                        if var_dict[pos][x] == '':
                            del var_dict[pos][x]
                    stdout_log(ws_sw, row_num)
                    var_dict[pos]['Site_Group'] = templateVars['Site_ID']
                    var_dict[pos]['Switch_Role'] = templateVars['Switch_Role']
                    var_dict[pos]['Site_Name'] = templateVars['Site_Name']
                    eval("%s.%s(wb, ws_sw, row_num, wr_file, **var_dict[pos])" % (class_init, func))
            wr_file.close()
            ws_wr = wb_sw.get_sheet_names()
            for sheetName in ws_wr:
                if sheetName in ['Sites']:
                    sheetToDelete = wb_sw.get_sheet_by_name(sheetName)
                    wb_sw.remove_sheet(sheetToDelete)
                    wb_sw.save(excel_wkbook)
            wb_sw.close()

        if re.search('Grp_[A-F]', templateVars['Site_Group']):
            print(f"\n-----------------------------------------------------------------------------\n")
            print(f"   Error on Worksheet {ws.title}, Row {row_num} Site_Group, value {templateVars['Site_Group']}.")
            print(f"   A Leaf can only be assigned to one Site.  Exiting....")
            print(f"\n-----------------------------------------------------------------------------\n")
            exit()
        elif re.search(r'\d+', templateVars['Site_Group']):
            Site_ID = 'Site_ID_%s' % (templateVars['Site_Group'])
            site_dict = ast.literal_eval(os.environ[Site_ID])

            # Set Destination Directory
            dest_dir = '%s' % (templateVars['Name'])

            # Create kwargs for Site Variables
            kwargs['Site_ID'] = site_dict.get('Site_ID')
            kwargs['Site_Name'] = site_dict.get('Site_Name')
            kwargs['APIC_URL'] = site_dict.get('APIC_URL')
            kwargs['APIC_Version'] = site_dict.get('APIC_Version')
            kwargs['APIC_Auth_Type'] = site_dict.get('APIC_Auth_Type')
            kwargs['Terraform_EQ'] = site_dict.get('Terraform_EQ')
            kwargs['Terraform_Version'] = site_dict.get('Terraform_Version')
            kwargs['Provider_EQ'] = site_dict.get('Provider_EQ')
            kwargs['Provider_Version'] = site_dict.get('Provider_Version')
            kwargs['Run_Location'] = site_dict.get('Run_Location')
            kwargs['State_Location'] = site_dict.get('State_Location')
            kwargs['Terraform_Cloud_Org'] = site_dict.get('Terraform_Cloud_Org')
            kwargs['Workspace_Prefix'] = site_dict.get('Workspace_Prefix')
            kwargs['VCS_Base_Repo'] = site_dict.get('VCS_Base_Repo')
            kwargs['Terraform_Agent_Pool_ID'] = site_dict.get('Terraform_Agent_Pool_ID')

            # Dicts for required and optional args
            required_args = {'Site_ID': '',
                                'Site_Name': '',
                                'APIC_URL': '',
                                'APIC_Version': '',
                                'APIC_Auth_Type': '',
                                'Terraform_EQ': '',
                                'Terraform_Version': '',
                                'Provider_EQ': '',
                                'Provider_Version': '',
                                'Run_Location': '',
                                'State_Location': ''}
            optional_args = {'Terraform_Cloud_Org': '',
                                'Workspace_Prefix': '',
                                'VCS_Base_Repo': '',
                                'Terraform_Agent_Pool_ID': ''}

            # Validate inputs, return dict of template vars
            templateVars = process_kwargs(required_args, optional_args, **kwargs)

            # If the State_Location is Terraform_Cloud Configure Workspaces in the Cloud
            if templateVars['State_Location'] == 'Terraform_Cloud':
                # Initialize the Class
                lib_tf_ref = 'lib_terraform.Terraform_Cloud'
                class_init = '%s()' % (lib_tf_ref)

                # Get terraform_cloud_token
                kwargs['terraform_cloud_token'] = eval("%s.%s()" % (class_init, 'terraform_token'))

                # Get terraform_cloud_token
                kwargs['terraform_oath_token'] = eval("%s.%s(**kwargs)" % (class_init, 'oath_token'))

                # Get workspace_ids
                workspace_dict = {}
                workspace_dict = terraform_cloud.tfcWorkspace(class_init, dest_dir, workspace_dict, **kwargs)

            # If the Run_Location is Terraform_Cloud Configure Variables in the Cloud
            if templateVars['Run_Location'] == 'Terraform_Cloud':
                # Set Variable List
                if templateVars['APIC_Auth_Type'] == 'user_pass':
                    var_list = ['aciUrl', 'aciUser', 'aciPass']
                else:
                    var_list = ['aciUrl', 'aciCertName', 'aciPrivateKey']

                # Get var_ids
                tf_var_dict = {}
                folder_id = 'Site_ID_%s_%s' % (templateVars['Site_ID'], dest_dir)
                kwargs['workspace_id'] = workspace_dict[folder_id]
                kwargs['Description'] = ''
                for var in var_list:
                    tf_var_dict = terraform_cloud.tfcVariables(class_init, dest_dir, var, tf_var_dict, **kwargs)

        else:
            print(f"\n-----------------------------------------------------------------------------\n")
            print(f"   Error on Worksheet {ws.title}, Row {row_num} Site_Group, value {templateVars['Site_Group']}.")
            print(f"   Unable to Determine if this is a Single or Group of Site(s).  Exiting....")
            print(f"\n-----------------------------------------------------------------------------\n")
            exit()

        self.templateLoader = jinja2.FileSystemLoader(
            searchpath=(aci_template_path))
        self.templateEnv = jinja2.Environment(loader=self.templateLoader)

        # Add the Default Files to the Tenant Directory
        file_list = ['.gitignore_.gitignore', 'main.jinja2_main.tf', 'variables.jinja2_variables.tf']
        for file in file_list:
            x = file.split('_')
            template_file = x[0]
            dest_file = x[1]
            template = self.templateEnv.get_template(template_file)
            create_tf_file('w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def sw_profile(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Switch_Role': '',
                         'Name': '',
                         'Selector_Name': '',
                         'Association_Type': '',
                         'Nodeblk_Name': '',
                         'Node_ID_From': '',
                         'Node_ID_To': '',
                         'Policy_Group': '',
                         'Interface_Profile': '',
                         'Dest_Folder': ''}
        optional_args = {'Description': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.name_rule(row_num, ws, 'Selector_Name', templateVars['Selector_Name'])
            validating.name_rule(row_num, ws, 'Nodeblk_Name', templateVars['Nodeblk_Name'])
            validating.name_rule(row_num, ws, 'Policy_Group', templateVars['Policy_Group'])
            validating.name_rule(row_num, ws, 'Interface_Profile', templateVars['Interface_Profile'])
            validating.name_rule(row_num, ws, 'Dest_Folder', templateVars['Dest_Folder'])
            validating.number_check(row_num, ws, 'Node_ID_From', templateVars['Node_ID_From'], 101, 4001)
            validating.number_check(row_num, ws, 'Node_ID_To', templateVars['Node_ID_To'], 101, 4001)
            validating.values(row_num, ws, 'Switch_Role', templateVars['Switch_Role'], ['leaf', 'spine'])
            validating.values(row_num, ws, 'Association_Type', templateVars['Association_Type'], ['ALL', 'range', 'ALL_IN_POD'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        if templateVars['Switch_Role'] == 'leaf':
            template_file = "leaf_profile.jinja2"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = '%s_leaf_profile.tf' % (templateVars['Name'])
        else:
            template_file = "spine_profile.jinja2"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = '%s_spine_profile.tf' % (templateVars['Name'])

        if not templateVars['Dest_Folder'] == None:
            dest_dir = '%s' % (templateVars['Dest_Folder'])
        else:
            dest_dir = 'Access'
        write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def vlan_pool(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Allocation_Mode': '',
                         'VLAN_Grp1': '',
                         'VGRP1_Allocation': ''}
        optional_args = {'Description': '',
                         'Alias': '',
                         'VLAN_Grp1': '',
                         'VGRP1_Allocation': '',
                         'VLAN_Grp2': '',
                         'VGRP2_Allocation': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.values(row_num, ws, 'Allocation_Mode', templateVars['Allocation_Mode'], ['dynamic', 'static'])
            validating.values(row_num, ws, 'VGRP1_Allocation', templateVars['VGRP1_Allocation'], ['dynamic', 'static'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
            if not templateVars['VGRP2_Allocation'] == None:
                validating.values(row_num, ws, 'VGRP2_Allocation', templateVars['VGRP2_Allocation'], ['dynamic', 'static'])
            validating.vlans(row_num, ws, 'VLAN_Grp1', templateVars['VLAN_Grp1'])
            if not templateVars['VLAN_Grp2'] == None:
                validating.vlans(row_num, ws, 'VLAN_Grp2', templateVars['VLAN_Grp2'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        if templateVars['Name'] == None:
            Error_Return = 'Error on Worksheet %s Row %s.  Could not Determine the Name of the VLAN Pool.' % (ws.title, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "vlan_pool.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'vlan_pool_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "data_vlan_pool.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'data_vlan_pool_%s.tf' % (templateVars['Name'])
        dest_dir = 'VLANs'
        write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # Add VLAN(s) to VLAN Pool FIle
        if re.search('Grp_[A-F]', templateVars['Site_Group']):
            Group_ID = '%s' % (templateVars['Site_Group'])
            site_group = ast.literal_eval(os.environ[Group_ID])
            for x in range(1, 13):
                sitex = 'Site_%s' % (x)
                if not site_group[sitex] == None:
                    Site_ID = 'Site_ID_%s' % (site_group[sitex])
                    site_dict = ast.literal_eval(os.environ[Site_ID])

                    # Create templateVars for Site_Name and APIC_URL
                    templateVars['Site_Name'] = site_dict.get('Site_Name')
                    templateVars['APIC_URL'] = site_dict.get('APIC_URL')

                    # Create Blank VLAN Pool VLAN(s) File
                    dest_file = './ACI/%s/VLANs/vlp_%s.tf' % (templateVars['Site_Name'], templateVars['Name'])
                    wr_file = open(dest_file, 'w')
                    wr_file.close()
                    dest_file = 'vlan_pool_%s.tf' % (templateVars['Name'])
                    dest_dir = 'VLANs'
                    template_file = "add_vlan_to_pool.jinja2"
                    template = self.templateEnv.get_template(template_file)

                    for z in range(1, 3):
                        vgroup = 'VLAN_Grp%s' % (z)
                        vgrp = 'VGRP%s_Allocation' % (z)
                        templateVars['Allocation_Mode'] = templateVars[vgrp]
                        if re.search(r'\d+', str(templateVars[vgroup])):
                            vlan_list = vlan_list_full(templateVars[vgroup])
                            for v in vlan_list:
                                vlan = str(v)
                                if re.fullmatch(r'\d+', vlan):
                                    templateVars['VLAN_ID'] = int(vlan)

                                    # Add VLAN to VLAN Pool File
                                    create_tf_file('a+', dest_dir, dest_file, template, **templateVars)

        elif re.search(r'\d+', templateVars['Site_Group']):
            Site_ID = 'Site_ID_%s' % (templateVars['Site_Group'])
            site_dict = ast.literal_eval(os.environ[Site_ID])

            # Create templateVars for Site_Name and APIC_URL
            templateVars['Site_Name'] = site_dict.get('Site_Name')
            templateVars['APIC_URL'] = site_dict.get('APIC_URL')

            # Create Blank VLAN Pool VLAN(s) File
            dest_file = './ACI/%s/VLANs/vlan_pool_%s.tf' % (templateVars['Site_Name'], templateVars['Name'])
            wr_file = open(dest_file, 'w')
            wr_file.close()
            dest_file = 'vlan_pool_%s.tf' % (templateVars['Name'])
            dest_dir = 'VLANs'
            template_file = "add_vlan_to_pool.jinja2"
            template = self.templateEnv.get_template(template_file)

            for z in range(1, 3):
                vgroup = 'VLAN_Grp%s' % (z)
                vgrp = 'VGRP%s_Allocation' % (z)
                templateVars['Allocation_Mode'] = templateVars[vgrp]
                if re.search(r'\d+', str(templateVars[vgroup])):
                    vlan_list = vlan_list_full(templateVars[vgroup])
                    for v in vlan_list:
                        vlan = str(v)
                        if re.fullmatch(r'\d+', vlan):
                            templateVars['VLAN_ID'] = int(vlan)

                            # Add VLAN to VLAN Pool File
                            create_tf_file('a+', dest_dir, dest_file, template, **templateVars)
        else:
            print(f"\n-----------------------------------------------------------------------------\n")
            print(f"   Error on Worksheet {ws.title}, Row {row_num} Site_Group, value {templateVars['Site_Group']}.")
            print(f"   Unable to Determine if this is a Single or Group of Site(s).  Exiting....")
            print(f"\n-----------------------------------------------------------------------------\n")
            exit()

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def vpc_pair(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'VPC_ID': '',
                         'Name': '',
                         'Node1_ID': '',
                         'Node2_ID': ''}
        optional_args = { }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.number_check(row_num, ws, 'VPC_ID', templateVars['VPC_ID'], 1, 1000)
            validating.number_check(row_num, ws, 'Node1_ID', templateVars['Node1_ID'], 101, 4001)
            validating.number_check(row_num, ws, 'Node2_ID', templateVars['Node2_ID'], 101, 4001)
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "vpc_domain.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'vpc_domain_%s.tf' % (templateVars['VPC_ID'])
        dest_dir = 'Access'
        write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

