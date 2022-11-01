#!/usr/bin/env python3

#=============================================================================
# Source Modules
#=============================================================================
from copy import deepcopy
from collections import OrderedDict
from openpyxl import load_workbook
from requests.api import delete
import easy_functions
import jinja2
import json
import os
import pkg_resources
import re
import requests
import stdiomask
import sys
import validating
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Global options for debugging
print_payload = False
print_response_always = False
print_response_on_fail = True

# Log levels 0 = None, 1 = Class only, 2 = Line
log_level = 2

# Global path to main Template directory
template_path = pkg_resources.resource_filename('classes', 'templates/')

class LoginFailed(Exception):
    pass

#=====================================================================================
# Please Refer to the "Notes" in the relevant column headers in the input Spreadhseet
# for detailed information on the Arguments used by this Class.
#=====================================================================================
class access(object):
    def __init__(self, type):
        self.type = type

    #=============================================================================
    # Function - Domain - Layer 3
    #=============================================================================
    def l3_domains(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['access.domains.Layer3']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)

        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'access,physical_and_external_domains,l3_domains'
        kwargs['easyDict'] = easy_functions.ez_append(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Domains - Physical
    #=============================================================================
    def phys_domains(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['access.domains.Physical']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)

        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'access,physical_and_external_domains,physical_domains'
        kwargs['easyDict'] = easy_functions.ez_append(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Global Policies - AAEP Profiles
    #=============================================================================
    def global_aaep(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['access.global.attachableAccessEntityProfile']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)

        # Split Domain Values to Lists
        domain_list = ['physical_domains', 'l3_domains', 'vmm_domains']
        for i in domain_list:
            if not polVars[f'{i}'] == None:
                polVars[f'{i}'] = polVars[f'{i}'].split(',')

        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'access,policies,global,attachable_access_entity_profiles'
        kwargs['easyDict'] = easy_functions.ez_append(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Leaf Policy Group
    #=============================================================================
    def leaf_pg(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['access.switches.leafPolicyGroup']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)

        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'access,switches,leaf,policy_groups'
        kwargs['easyDict'] = easy_functions.ez_append(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Policy Group - Access
    #=============================================================================
    def pg_access(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['access.policyGroups.leafAccessPort']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)

        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'access,interfaces,leaf,policy_groups,access'
        kwargs['easyDict'] = easy_functions.ez_append(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Policy Group - Breakout
    #=============================================================================
    def pg_breakout(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['access.policyGroups.leafBreakOut']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)

        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'access,interfaces,leaf,policy_groups,breakout'
        kwargs['easyDict'] = easy_functions.ez_append(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Policy Group - VPC/Port-Channel
    #=============================================================================
    def pg_bundle(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['access.policyGroups.leafBundles']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        polVars.pop('template_name')
        polVars['names'] = polVars['names'].split(',')

        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'access,interfaces,leaf,policy_groups,bundle,names'
        kwargs['policy'] = 'template_name'
        kwargs['policy_name'] = kwargs['template_name']
        kwargs['easyDict'] = easy_functions.ez_append_arg(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Policy Group - VPC/Port-Channel Template
    #=============================================================================
    def pg_template(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['access.policyGroups.leafBundleTemplate']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        if not polVars['netflow_monitor_policies'] == None:
            polVars['netflow_monitor_policies'] = polVars['netflow_monitor_policies'].split(',')

        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'access,interfaces,leaf,policy_groups,bundle'
        kwargs['easyDict'] = easy_functions.ez_append(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Policy Group - Spine
    #=============================================================================
    def pg_spine(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['access.policyGroups.spineAccessPort']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)

        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'access,interfaces,spine,policy_groups'
        kwargs['easyDict'] = easy_functions.ez_append(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Interface Policies - CDP
    #=============================================================================
    def pol_cdp(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['access.policies.cdpInterface']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)

        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'access,policies,interface,cdp_interface'
        kwargs['easyDict'] = easy_functions.ez_append(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Interface Policies - Fibre Channel
    #=============================================================================
    def pol_fc(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['access.policies.fibreChannelInterface']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)

        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'access,policies,interface,fibre_channel_interface'
        kwargs['easyDict'] = easy_functions.ez_append(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Interface Policies - L2 Interfaces
    #=============================================================================
    def pol_l2(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['access.policies.L2Interface']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)

        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'access,policies,interface,l2_interface'
        kwargs['easyDict'] = easy_functions.ez_append(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Interface Policies - Link Level (Speed)
    #=============================================================================
    def pol_link_level(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['access.policies.linkLevel']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)

        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'access,policies,interface,link_level'
        kwargs['easyDict'] = easy_functions.ez_append(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Interface Policies - LLDP
    #=============================================================================
    def pol_lldp(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['access.policies.lldpInterface']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)

        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'access,policies,interface,lldp_interface'
        kwargs['easyDict'] = easy_functions.ez_append(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Interface Policies - Mis-Cabling Protocol
    #=============================================================================
    def pol_mcp(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['access.policies.mcpInterface']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)

        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'access,policies,interface,mcp_interface'
        kwargs['easyDict'] = easy_functions.ez_append(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Interface Policies - Port Channel
    #=============================================================================
    def pol_port_ch(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['access.policies.PortChannel']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)

        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'access,policies,interface,port_channel'
        kwargs['easyDict'] = easy_functions.ez_append(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Interface Policies - Port Security
    #=============================================================================
    def pol_port_sec(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['access.policies.portSecurity']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)

        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'access,policies,interface,port_security'
        kwargs['easyDict'] = easy_functions.ez_append(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Interface Policies - Spanning Tree
    #=============================================================================
    def pol_stp(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['access.policies.spanningTreeInterface']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)

        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'access,policies,interface,spanning_tree_interface'
        kwargs['easyDict'] = easy_functions.ez_append(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Recommended Settings
    #=============================================================================
    def pre_built(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['access.preBuiltPolicies']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        remove_temp = kwargs['easyDict']['remove_default_args']
        kwargs['easyDict']['remove_default_args'] = False
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        kwargs['easyDict']['remove_default_args'] == remove_temp
        
        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'access,policies,interface,create_pre_built_interface_policies'
        kwargs['easyDict'] = easy_functions.ez_update(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Spine Policy Group
    #=============================================================================
    def spine_pg(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['access.switches.spinePolicyGroup']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)

        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'access,switches,spine,policy_groups'
        kwargs['easyDict'] = easy_functions.ez_append(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - VLAN Pools
    #=============================================================================
    def vlan_pools(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['access.pools.Vlan']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)

        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'access,pools,vlan'
        kwargs['easyDict'] = easy_functions.ez_append(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Virtual Networking - Controllers
    #=============================================================================
    def vmm_controllers(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['access.vmm.Controllers']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        polVars.pop('domain_name')

        # Add Dictionary to Policy
        kwargs['class_path'] = 'virtual_networking,vmm,controllers'
        kwargs['policy'] = 'name'
        kwargs['policy_name'] = kwargs['domain_name']
        kwargs['easyDict'] = easy_functions.ez_append_subtype(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Virtual Networking - Credentials
    #=============================================================================
    def vmm_creds(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['access.vmm.Credentials']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)

        # Check Environment for VMM Credentials Password
        sensitive_list = ['vmm_password']
        polVars, kwargs = easy_functions.env_sensitive(sensitive_list, jsonData, polVars, **kwargs)

        # Add Dictionary to Policy
        kwargs['class_path'] = 'virtual_networking,vmm,credentials'
        kwargs['policy'] = 'name'
        kwargs['policy_name'] = kwargs['domain_name']
        kwargs['easyDict'] = easy_functions.ez_append_subtype(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Virtual Networking - Domains
    #=============================================================================
    def vmm_domain(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['access.vmm.Domains']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)

        # Convert to Lists
        if not polVars["uplink_names"] == None:
            polVars["uplink_names"] = polVars["uplink_names"].split(',')
        polVars = easy_functions.ez_remove_empty(polVars)

        newDict = {
            'controllers':[],
            'credentials':[],
            'enhanced_lag_policy':[],
            'domain':[polVars],
            'name':polVars['name'],
            'site_group':polVars['site_group'],
            'vswitch_policy':[]
        }
        polVars = newDict
        
        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'virtual_networking,vmm'
        kwargs['easyDict'] = easy_functions.ez_append(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Virtual Networking - Controllers
    #=============================================================================
    def vmm_elagp(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['access.vmm.enhancedLag']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        polVars.pop('domain_name')

        # Add Dictionary to Policy
        kwargs['class_path'] = 'virtual_networking,vmm,enhanced_lag_policy'
        kwargs['policy'] = 'name'
        kwargs['policy_name'] = kwargs['domain_name']
        kwargs['easyDict'] = easy_functions.ez_append_subtype(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Virtual Networking - Controllers
    #=============================================================================
    def vmm_vswitch(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['access.vmm.vswitchPolicy']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        polVars.pop('domain_name')

        # Add Dictionary to Policy
        kwargs['class_path'] = 'virtual_networking,vmm,vswitch_policy'
        kwargs['policy'] = 'name'
        kwargs['policy_name'] = kwargs['domain_name']
        kwargs['easyDict'] = easy_functions.ez_append_subtype(polVars, **kwargs)
        return kwargs['easyDict']

#=====================================================================================
# Please Refer to the "Notes" in the relevant column headers in the input Spreadhseet
# for detailed information on the Arguments used by this Class.
#=====================================================================================
class admin(object):
    def __init__(self, type):
        self.type = type

    #=============================================================================
    # Function - Authentication
    #=============================================================================
    def auth(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['admin.Authentication']['allOf'][1]['properties']
        
        # Check for Variable values that could change required arguments
        args_list = []
        if kwargs.get('console_realm'):
            if not kwargs['console_realm'] == 'local': args_list.append('console_login_domain')
        else: kwargs['console_realm'] == 'local'
        if kwargs.get('default_realm'):
            if not kwargs['default_realm'] == 'local': args_list.append('default_login_domain')
        else: kwargs['default_realm'] == 'local'
        jsonData = easy_functions.args_add(args_list, jsonData)
        
        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        if polVars['console_login_domain'] == None: polVars['console_login_domain'] = ''
        if polVars['default_login_domain'] == None: polVars['default_login_domain'] = ''

        # Reset jsonData
        jsonData = easy_functions.args_remove(args_list, jsonData)
        newDict = {
            'name':'default',
            'console':{'login_domain':polVars['console_login_domain'],'realm':kwargs['console_realm']},
            'default':{'login_domain':polVars['default_login_domain'],'realm':kwargs['default_realm']},
            'icmp_reachability':{
                'retries':1,
                'timeout':5,
                'use_icmp_reachable_providers_only':polVars.get('use_icmp_reachable_providers_only')
            },
            'remote_user_login_policy':kwargs['remote_user_login_policy'],
            'site_group':polVars['site_group']
        }
        polVars = newDict
        polVars['icmp_reachability'] = easy_functions.ez_remove_empty(polVars['icmp_reachability'])
        
        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'admin,aaa,authentication,aaa'
        kwargs['easyDict'] = easy_functions.ez_append(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Configuration Backup - Export Policies
    #=============================================================================
    def export_policy(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['admin.exportPolicy']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)

        polVars['schedule'] = {'days':kwargs['days'], 'hour':kwargs['hour'], 'minute':kwargs['minute']}
        polVars.update({'configuration_export': []})
        remove_list = ['days', 'hour', 'minute']
        for i in remove_list:
            if not polVars.get(i) == None: polVars.pop(i)
    
        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'admin,import-export,configuration_backups'
        kwargs['easyDict'] = easy_functions.ez_append(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Global Security Settings
    #=============================================================================
    def mg_policy(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['admin.firmware.Policy']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)

        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'admin,firmware,maintenance_group_policies'
        kwargs['easyDict'] = easy_functions.ez_append(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Configuration Backup  - Remote Host
    #=============================================================================
    def maint_group(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['admin.firmware.MaintenanceGroups']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)

        # Convert to Lists
        polVars['node_list'] = [eval(i) for i in polVars['node_list'].split(',')]

        # Add Dictionary to Policy
        kwargs['class_path'] = 'admin,firmware,maintenance_groups'
        kwargs['easyDict'] = easy_functions.ez_append(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - RADIUS Authentication
    #=============================================================================
    def radius(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['admin.Radius']['allOf'][1]['properties']
        
        # Check for Variable values that could change required arguments
        if kwargs.get('server_monitoring'):
            if kwargs['server_monitoring'] == 'enabled':
                args_list = ['username']
                jsonData = easy_functions.args_add(args_list, jsonData)
        else: kwargs['server_monitoring'] == 'disabled'
        
        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        polVars['hosts'] = polVars['hosts'].split(',')
        pop_list = ['server_monitoring', 'username']

        # Check if the Sensitive Variables are in the Environment.  If not Add them.
        if kwargs['server_monitoring'] == 'enabled':
            jsonData = easy_functions.args_remove(args_list, jsonData)
            polVars['server_monitoring'] = {
                'admin_state': kwargs['server_monitoring'],
                'username': kwargs['username']
            }
            sensitive_list = ['radius_key', 'radius_monitoring_password']
        else: sensitive_list = ['radius_key']
        polVars, kwargs = easy_functions.env_sensitive(sensitive_list, jsonData, polVars, **kwargs)
        for i in pop_list:
            if not polVars.get(i) == None: polVars.pop(i)

        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'admin,aaa,authentication,radius'
        kwargs['easyDict'] = easy_functions.ez_append(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Configuration Backup  - Remote Host
    #=============================================================================
    def remote_host(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['admin.remoteHost']['allOf'][1]['properties']

        # Check for Variable values that could change required arguments
        args_list = []
        if kwargs.get('authentication_type'):
            if kwargs['authentication_type'] == 'usePassword': args_list.append('username')
        else:
            kwargs['authentication_type'] == 'usePassword'
            args_list.append('username')
        jsonData = easy_functions.args_add(args_list, jsonData)

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        polVars["remote_hosts"] = polVars["remote_hosts"].split(',')
        polVars.pop('scheduler_name')
        
        # Check if the Sensitive Variables are in the Environment.  If not Add them.
        if kwargs['authentication_type'] == 'usePassword':
            sensitive_list = ['remote_password']
        else: sensitive_list = ['ssh_key_contents', 'ssh_key_passphrase']
        polVars, kwargs = easy_functions.env_sensitive(sensitive_list, jsonData, polVars, **kwargs)

        # Reset jsonData
        jsonData = easy_functions.args_remove(args_list, jsonData)
        
        # Add Dictionary to Policy
        kwargs['class_path'] = 'admin,import-export,configuration_backups'
        kwargs['policy'] = 'name'
        kwargs['policy_name'] = kwargs['scheduler_name']
        kwargs['easyDict'] = easy_functions.ez_merge(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Global Security Settings
    #=============================================================================
    def security(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['admin.globalSecurity']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        polVars['lockout_user'] = {
          'enable_lockout': kwargs['enable_lockout'],
          'lockout_duration': kwargs['lockout_duration'],
          'max_failed_attempts': kwargs['max_failed_attempts'],
          'max_failed_attempts_window': kwargs['max_failed_attempts_window']
        }
        pop_list = ['enable_lockout', 'lockout_duration', 'max_failed_attempts', 'max_failed_attempts_window']
        for i in pop_list:
            if not polVars.get(i) == None: polVars.pop(i)
        
        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'admin,aaa,security'
        kwargs['easyDict'] = easy_functions.ez_append(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Smart CallHome Policy
    #=============================================================================
    def smart_callhome(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['admin.smartCallHome']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        polVars['name'] = 'default'
        polVars['smtp_server'] = {}
        polVars['smart_destinations'] = []

        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'admin,external-data-collectors,smart_callhome'
        kwargs['easyDict'] = easy_functions.ez_append(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Smart CallHome Policy - Smart Destinations
    #=============================================================================
    def smart_destinations(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['admin.smartDestinations']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)

        # Add Dictionary to Policy
        kwargs['class_path'] = 'admin,external-data-collectors,smart_callhome,smart_destinations'
        kwargs['policy'] = 'name'
        kwargs['policy_name'] = 'default'
        kwargs['easyDict'] = easy_functions.ez_append_subtype(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Smart CallHome Policy - SMTP Server
    #=============================================================================
    def smart_smtp_server(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['admin.smartSmtpServer']['allOf'][1]['properties']

        # Check for Variable values that could change required arguments
        if kwargs.get('secure_smtp'):
             if kwargs['secure_smtp'] == 'true':
                jsonData = easy_functions.args_add(['username'], jsonData)
        else: kwargs['secure_smtp'] == 'false'

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)

        # Check if the Sensitive Variables are in the Environment.  If not Add them.
        if kwargs['secure_smtp'] == True:
            sensitive_list = ['smtp_password']
            polVars, kwargs = easy_functions.env_sensitive(sensitive_list, jsonData, polVars, **kwargs)
            jsonData = easy_functions.args_remove(['username'], jsonData)

        # Add Dictionary to Policy
        kwargs['class_path'] = 'admin,external-data-collectors,smart_callhome,smtp_server'
        kwargs['policy'] = 'name'
        kwargs['policy_name'] = 'default'
        kwargs['easyDict'] = easy_functions.ez_update_subtype(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Syslog Policy
    #=============================================================================
    def syslog(self, **kwargs):
       # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['admin.Syslog']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        Additions = {
            'console_destination': {
                'admin_state':kwargs['console_admin_state'],
                'severity':kwargs['console_severity'],
            },
            'include_types': {
                'audit_logs':kwargs['audit_logs'],
                'events':kwargs['events'],
                'faults':kwargs['faults'],
                'session_logs':kwargs['session_logs']
            },
            'local_file_destination': {
                'admin_state':kwargs['local_admin_state'],
                'severity':kwargs['local_severity'],
            },
            'name':'default',
            'remote_destinations': []
        }
        Additions = {'name':'default', 'remote_destinations': []}
        polVars.update(Additions)
        pop_list = [
            'local_admin_state', 'local_severity', 'console_admin_state', 'console_severity',
            'audit_logs', 'events', 'faults', 'session_logs'
        ]
        for i in pop_list:
            if not polVars.get(i) == None: polVars.pop(i)
        
        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'admin,external-data-collectors,syslog'
        kwargs['easyDict'] = easy_functions.ez_append(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Syslog Policy - Syslog Destinations
    #=============================================================================
    def syslog_destinations(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['admin.syslogRemoteDestinations']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        polVars['hosts'] = polVars['hosts'].split(',')

        # Add Dictionary to Policy
        kwargs['class_path'] = 'admin,external-data-collectors,syslog,remote_destinations'
        kwargs['policy'] = 'name'
        kwargs['policy_name'] = 'default'
        kwargs['easyDict'] = easy_functions.ez_append_subtype(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - TACACS+ Authentication
    #=============================================================================
    def tacacs(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['admin.Tacacs']['allOf'][1]['properties']
        
        # Check for Variable values that could change required arguments
        if kwargs.get('server_monitoring'):
            if kwargs['server_monitoring'] == 'enabled':
                jsonData = easy_functions.args_add(['username'], jsonData)
        else: kwargs['server_monitoring'] == 'disabled'
        
        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        polVars['accounting_include'] = {
            'audit_logs':polVars['audit_logs'],
            'events':polVars['events'],
            'faults':polVars['faults'],
            'session_logs':polVars['session_logs']
        }
        polVars['hosts'] = polVars['hosts'].split(',')
        pop_list = ['audit_logs', 'events' ,'faults' ,'session_logs', 'server_monitoring', 'username']

        # Check if the Sensitive Variables are in the Environment.  If not Add them.
        if kwargs['server_monitoring'] == 'enabled':
            jsonData = easy_functions.args_remove(['username'], jsonData)
            polVars['server_monitoring'] = {
                'admin_state': kwargs['server_monitoring'],
                'username': kwargs['username']
            }
            sensitive_list = ['tacacs_key', 'radius_monitoring_password']
        else: sensitive_list = ['tacacs_key']
        polVars, kwargs = easy_functions.env_sensitive(sensitive_list, jsonData, polVars, **kwargs)
        for i in pop_list:
            if not polVars.get(i) == None: polVars.pop(i)

        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'admin,aaa,authentication,tacacs'
        kwargs['easyDict'] = easy_functions.ez_append(polVars, **kwargs)
        return kwargs['easyDict']

#=====================================================================================
# Please Refer to the "Notes" in the relevant column headers in the input Spreadhseet
# for detailed information on the Arguments used by this Class.
#=====================================================================================
class apicLogin(object):
    def __init__(self, apic, user, pword):
        self.apic = apic
        self.user = user
        self.pword = pword

    def login(self):
        # Load login json payload
        payload = '''
        {{
            "aaaUser": {{
                "attributes": {{
                    "name": "{user}",
                    "pwd": "{pword}"
                }}
            }}
        }}
        '''.format(user=self.user, pword=self.pword)
        payload = json.loads(payload,
                             object_pairs_hook=OrderedDict)
        s = requests.Session()
        # Try the request, if exception, exit program w/ error
        try:
            # Verify is disabled as there are issues if it is enabled
            r = s.post('https://{}/api/aaaLogin.json'.format(self.apic),
                       data=json.dumps(payload), verify=False)
            # Capture HTTP status code from the request
            status = r.status_code
            # Capture the APIC cookie for all other future calls
            cookies = r.cookies
            # Log login status/time(?) somewhere
            if status == 400:
                print("Error 400 - Bad Request - ABORT!")
                print("Probably have a bad URL")
                sys.exit()
            if status == 401:
                print("Error 401 - Unauthorized - ABORT!")
                print("Probably have incorrect credentials")
                sys.exit()
            if status == 403:
                print("Error 403 - Forbidden - ABORT!")
                print("Server refuses to handle your request")
                sys.exit()
            if status == 404:
                print("Error 404 - Not Found - ABORT!")
                print("Seems like you're trying to POST to a page that doesn't"
                      " exist.")
                sys.exit()
        except Exception as e:
            print("Something went wrong logging into the APIC - ABORT!")
            # Log exit reason somewhere
            raise LoginFailed(e)
        self.cookies = cookies
        return cookies

#=====================================================================================
# Please Refer to the "Notes" in the relevant column headers in the input Spreadhseet
# for detailed information on the Arguments used by this Class.
#=====================================================================================
class fabric(object):
    def __init__(self, type):
        self.type = type

    #=============================================================================
    # Function - Date and Time Policy
    #=============================================================================
    def date_time(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['fabric.DateandTime']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        if polVars['server_state'] == 'disabled':
            polVars['master_mode'] = 'disabled'
        polVars.update({'name':'default'})
        
        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'fabric,policies,pod,date_and_time'
        kwargs['easyDict'] = easy_functions.ez_append(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - DNS Profiles
    #=============================================================================
    def dns_profile(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['fabric.dnsProfiles']['allOf'][1]['properties']

        if not kwargs['dns_domains'] == None:
            kwargs['dns_domains'] = kwargs['dns_domains'].split(',')
            if not kwargs['default_domain'] == None:
                if not kwargs['default_domain'] in kwargs['dns_domains']:
                    kwargs['dns_domains'].append(kwargs['default_domain'])

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        polVars['name'] = 'default'
        
        # Convert to Lists
        dns_providers = polVars["dns_providers"].split(',')
        polVars["dns_providers"] = []
        for i in dns_providers:
            if polVars.get('preferred'):
                if i == polVars['preferred']:
                    polVars["dns_providers"].append({'dns_provider':i,'preferred':True})
                else: polVars["dns_providers"].append({'dns_provider':i,'preferred':False})
            else: polVars["dns_providers"].append({'dns_provider':i,'preferred':False})
        dns_domains = polVars['dns_domains']
        polVars['dns_domains'] = []
        for i in dns_domains:
            if polVars.get('default_domain'):
                if i == polVars['default_domain']:
                    polVars["dns_domains"].append({'domain':i,'default_domain':True})
                else: polVars["dns_domains"].append({'domain':i,'default_domain':False})
            else: polVars["dns_domains"].append({'domain':i,'default_domain':False})
        pop_list = ['preferred','default_domain']
        for i in pop_list:
            if not polVars.get(i) == None: polVars.pop(i)

        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'fabric,policies,global,dns_profiles'
        kwargs['easyDict'] = easy_functions.ez_append(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Date and Time Policy - NTP Servers
    #=============================================================================
    def ntp(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['fabric.Ntp']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)

        # Add Dictionary to Policy
        kwargs['class_path'] = 'fabric,policies,pod,date_and_time,ntp_servers'
        kwargs['policy'] = 'name'
        kwargs['policy_name'] = 'default'
        kwargs['easyDict'] = easy_functions.ez_append_subtype(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Date and Time Policy - NTP Keys
    #=============================================================================
    def ntp_key(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['fabric.NtpKeys']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)

        # Check if the Sensitive Variables are in the Environment.  If not Add them.
        sensitive_list = [f'ntp_key_{kwargs["key_id"]}']
        polVars, kwargs = easy_functions.env_sensitive(sensitive_list, jsonData, polVars, **kwargs)

        # Add Dictionary to Policy
        kwargs['class_path'] = 'fabric,policies,pod,date_and_time,authentication_keys'
        kwargs['policy'] = 'name'
        kwargs['policy_name'] = 'default'
        kwargs['easyDict'] = easy_functions.ez_append_subtype(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - SNMP Policy - Client Groups
    #=============================================================================
    def snmp_clgrp(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['fabric.snmpClientGroups']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        if not polVars["clients"] == None:
            clients = []
            for i in polVars["clients"].split(','):
                clients.append({'address':i})
            polVars["clients"] = clients

        # Add Dictionary to Policy
        kwargs['class_path'] = 'fabric,policies,pod,snmp,snmp_client_groups'
        kwargs['policy'] = 'name'
        kwargs['policy_name'] = 'default'
        kwargs['easyDict'] = easy_functions.ez_append_subtype(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - SNMP Policy - Communities
    #=============================================================================
    def snmp_community(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['fabric.snmpCommunities']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        polVars['community_variable'] = int(polVars['community_variable'])

        # Check if the Sensitive Variables are in the Environment.  If not Add them.
        sensitive_list = [f'snmp_community_{kwargs["community_variable"]}']
        polVars, kwargs = easy_functions.env_sensitive(sensitive_list, jsonData, polVars, **kwargs)

        # Add Dictionary to Policy
        kwargs['class_path'] = 'fabric,policies,pod,snmp,snmp_communities'
        kwargs['policy'] = 'name'
        kwargs['policy_name'] = 'default'
        kwargs['easyDict'] = easy_functions.ez_append_subtype(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - SNMP Policy - SNMP Trap Destinations
    #=============================================================================
    def snmp_destinations(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['fabric.snmpDestinations']['allOf'][1]['properties']

        # Check for Variable values that could change required arguments
        args_list = []
        if kwargs.get('version'):
            if re.fullmatch('(v1|v2c)', kwargs['version']): args_list.append('community_variable')
            elif 'v3' in kwargs['version']: args_list.extend(['username', 'v3_security_level'])
        else: kwargs['version'] = 'v2c'
        jsonData = easy_functions.args_add(args_list, jsonData)

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)

        # Reset Arguments
        jsonData = easy_functions.args_remove(args_list, jsonData)
        
        # Add Dictionary to Policy
        kwargs['class_path'] = 'fabric,policies,pod,snmp,snmp_destinations'
        kwargs['policy'] = 'name'
        kwargs['policy_name'] = 'default'
        kwargs['easyDict'] = easy_functions.ez_append_subtype(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - SNMP Policy
    #=============================================================================
    def snmp_policy(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['fabric.snmpPolicy']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        polVars['name'] = 'default'
        polVars['snmp_client_groups'] = []
        polVars['snmp_communities'] = []
        polVars['snmp_destinations'] = []
        polVars['users'] = []
        polVars['include_types'] = {
            'audit_logs':polVars.get('audit_logs'),
            'events':polVars.get('events'),
            'faults':polVars.get('faults'),
            'session_logs':polVars.get('session_logs')
        }
        pop_list = ['audit_logs', 'events' ,'faults' ,'session_logs' ]
        for i in pop_list:
            if not polVars.get(i) == None: polVars.pop(i)
        polVars['include_types'] = easy_functions.ez_remove_empty(polVars['include_types'])
        if len(polVars['include_types']) == 0:
            polVars.pop('include_types')

        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'fabric,policies,pod,snmp'
        kwargs['easyDict'] = easy_functions.ez_append(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - SNMP Policy - SNMP Users
    #=============================================================================
    def snmp_user(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['fabric.snmpUsers']['allOf'][1]['properties']

        # Check for Variable values that could change required arguments
        if kwargs.get('privacy_key'):
            if not kwargs['privacy_key'] == 'none':
                jsonData = easy_functions.args_add(['privacy_key'], jsonData)
        else:
            kwargs['privacy_key'] = 'none'

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)

        # Check if the Sensitive Variables are in the Environment.  If not Add them.
        if not kwargs['privacy_type'] == 'none':
            jsonData = easy_functions.args_remove(['privacy_key'], jsonData)
            sensitive_list = [
                f'snmp_authorization_key_{kwargs["authorization_key"]}',
                f'snmp_privacy_key_{kwargs["privacy_key"]}'
            ]
        else: sensitive_list = [f'snmp_authorization_key_{kwargs["authorization_key"]}']
        polVars, kwargs = easy_functions.env_sensitive(sensitive_list, jsonData, polVars, **kwargs)

        # Add Dictionary to Policy
        kwargs['class_path'] = 'fabric,policies,pod,snmp,users'
        kwargs['policy'] = 'name'
        kwargs['policy_name'] = 'default'
        kwargs['easyDict'] = easy_functions.ez_append_subtype(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Recommended Settings
    #=============================================================================
    def recommended_settings(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['fabric.recommendedSettings']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        remove_temp = kwargs['easyDict']['remove_default_args']
        kwargs['easyDict']['remove_default_args'] = False
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        kwargs['easyDict']['remove_default_args'] == remove_temp
        
        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'fabric,recommended_settings'
        kwargs['easyDict'] = easy_functions.ez_update(polVars, **kwargs)
        return kwargs['easyDict']

#=====================================================================================
# Please Refer to the "Notes" in the relevant column headers in the input Spreadhseet
# for detailed information on the Arguments used by this Class.
#=====================================================================================
class ndoLogin(object):
    def __init__(self, ndo, domain, pword, user):
        self.domain = domain
        self.ndo = ndo
        self.pword = pword
        self.user = user

    def login(self):
        # Load login json payload
        payload = '''
        {{
            "username": "{user}",
            "password": "{pword}",
            "domainId": "{domain}"
        }}
        '''.format(user=self.user, pword=self.pword, domain=self.domain)
        payload = json.loads(payload)
        s = requests.Session()
        # Try the request, if exception, exit program w/ error
        try:
            # Verify is disabled as there are issues if it is enabled
            newHeaders = {'Content-type': 'application/json'}
            r = requests.post('https://{}/login'.format(self.ndo),
                       data=json.dumps(payload), verify=False, headers=newHeaders)

            # Capture HTTP status code from the request
            status = r.status_code

            # Capture the APIC cookie for all other future calls
            cookies = r.cookies
            # Log login status/time(?) somewhere
            if status == 400:
                print("Error 400 - Bad Request - ABORT!")
                print("Probably have a bad URL")
                sys.exit()
            if status == 401:
                print("Error 401 - Unauthorized - ABORT!")
                print("Probably have incorrect credentials")
                sys.exit()
            if status == 403:
                print("Error 403 - Forbidden - ABORT!")
                print("Server refuses to handle your request")
                sys.exit()
            if status == 404:
                print("Error 404 - Not Found - ABORT!")
                print("Seems like you're trying to POST to a page that doesn't"
                      " exist.")
                sys.exit()
        except Exception as e:
            print("Something went wrong logging into Nexus Dashboard Orchestor - ABORT!")
            # Log exit reason somewhere
            raise LoginFailed(e)
        self.cookies = cookies
        return cookies

#=====================================================================================
# Please Refer to the "Notes" in the relevant column headers in the input Spreadhseet
# for detailed information on the Arguments used by this Class.
#=====================================================================================
class switches(object):
    def __init__(self, type):
        self.type = type
        self.templateLoader = jinja2.FileSystemLoader(
            searchpath=(template_path + 'switches/'))
        self.templateEnv = jinja2.Environment(loader=self.templateLoader)
    #=============================================================================
    # Function - Interface Selectors
    #=============================================================================
    def intf_selector(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['access.profiles.interfaceSelectors']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        polVars['interface_description'] = polVars['description']
        if len(polVars['interface'].split(',')) > 2: polVars['sub_port'] = True
        else: polVars['sub_port'] = None
        pop_list = [
            'access_or_native_vlan', 'description', 'interface_profile', 'interface_selector',
            'node_id', 'pod_id',  'switchport_mode', 'trunk_port_allowed_vlans'
        ]
        for i in pop_list:
            if not polVars.get(i) == None: polVars.pop(i)
        pgt = polVars['policy_group_type']
        if pgt == 'spine_pg': polVars.pop('policy_group_type')

        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'switch,switch_profiles,interfaces'
        kwargs['policy'] = 'name'
        kwargs['policy_name'] = kwargs['interface_profile']
        kwargs['easyDict'] = easy_functions.ez_append_subtype(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Port Conversion
    #=============================================================================
    def port_cnvt(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['access.switches.portConvert']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)

        def process_site(site_dict, polVars, **kwargs):
            if site_dict['auth_type'] == 'username':
                if not site_dict['login_domain'] == None:
                    apic_user = f"apic#{site_dict['login_domain']}\\{site_dict['username']}"
                else: apic_user = site_dict['username']
                
                # Check if the Sensitive Variables are in the Environment.  If not Add them.
                sensitive_list = ['apicPass']
                polVars, kwargs = easy_functions.env_sensitive(sensitive_list, jsonData, polVars, **kwargs)
                apic_pass = os.environ.get('TF_VAR_apicPass')
                node_list = easy_functions.vlan_list_full(polVars['node_list'])
                port_list = easy_functions.vlan_list_full(polVars['port_list'])

                controller = site_dict['controller']
                fablogin = apicLogin(controller, apic_user, apic_pass)
                cookies = fablogin.login()

                for node in node_list:
                    polVars['node_id'] = node
                    for port in port_list:
                        # Locate template for method
                        template_file = "check_ports.json"
                        template = self.templateEnv.get_template(template_file)
                        # Render template w/ values from dicts
                        payload = template.render(polVars)
                        uri = 'ncapi/config'
                        # port_modes = get(controller, payload, cookies, uri, template_file)

                        # Locate template for method
                        polVars['port'] = f"1/{port}"
                        template_file = "port_convert.json"
                        template = self.templateEnv.get_template(template_file)
                        # Render template w/ values from dicts
                        payload = template.render(polVars)
                        uri = 'ncapi/config'
                        easy_functions.apic_post(controller, 'conf', payload, cookies, uri, template_file)

        # Loop Through the Site Groups
        if re.search('Grp_', polVars['site_group']):
            site_group = kwargs['easyDict']['site_groups'][kwargs['site_group']]
            for site in site_group['sites']:
                # Process the Site Port Conversions
                siteDict = kwargs['easyDict']['sites'][site]['site_settings']
                process_site(siteDict, polVars, **kwargs)
        else:
            # Process the Site Port Conversions
            siteDict = kwargs['easyDict']['sites'][kwargs['site_group']]['site_settings']
            process_site(siteDict, polVars, **kwargs)

        return kwargs['easyDict']

    #=============================================================================
    # Function - Switch Inventory
    #=============================================================================
    def switch(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['access.profiles.switchProfiles']['allOf'][1]['properties']

        if re.search('Grp_[A-F]', kwargs['site_group']):
            print(f"\n-----------------------------------------------------------------------------\n")
            print(f"   Error on Worksheet {kwargs['ws'].title}, Row {kwargs['row_num']} site_group, value {kwargs['site_group']}.")
            print(f"   A Leaf can only be assigned to one Site.  Exiting....")
            print(f"\n-----------------------------------------------------------------------------\n")
            exit()

        # If IP Gateways Present, add as Required Variable
        args_list = []
        atype_list = ['ipv4', 'ipv6']
        mgmt_list = ['inband', 'ooband']
        for mgmt in mgmt_list:
            for atype in atype_list:
                if kwargs.get(f'{mgmt}_{atype}'):
                    if not kwargs[f'{mgmt}_{atype}'] == None:
                        args_list.extend([f'{mgmt}_{atype}', f'{mgmt}_{atype}_gateway'])
        jsonData = easy_functions.args_add(args_list, jsonData)

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)

        # Reset Required Arguments
        jsonData = easy_functions.args_remove(args_list, jsonData)

        # If device is a Leaf, determine if it is in a VPC Pair
        if not polVars['node_type'] == 'spine':
            if not polVars['vpc_name'] == None:
                if kwargs['easyDict']['sites'][kwargs['site_group']]['switch'].get('vpc_domains'):
                    if len(kwargs['easyDict']['sites'][kwargs['site_group']]['switch']['vpc_domains']) > 0:
                        vpc_count = 0
                        for i in kwargs['easyDict']['sites'][kwargs['site_group']]['switch']['vpc_domains']:
                            if i['name'] == polVars['vpc_name']:
                                i['switches'].append(polVars['node_id'])
                                vpc_count =+ 1
                        if vpc_count == 0:
                            # Add Policy Variables to easyDict
                            vpcArgs = {
                                'name':polVars['vpc_name'],
                                'domain_id':polVars['vpc_domain_id'],
                                'site_group':kwargs['site_group'],
                                'switches':[polVars['node_id']],
                                'vpc_domain_policy':'default',
                            }
                            kwargs['class_path'] = 'switch,vpc_domains'
                            kwargs['easyDict'] = easy_functions.ez_append(vpcArgs, **kwargs)
                else:
                    # Add Policy Variables to easyDict
                    vpcArgs = {
                        'name':polVars['vpc_name'],
                        'domain_id':polVars['vpc_domain_id'],
                        'site_group':kwargs['site_group'],
                        'switches':[polVars['node_id']],
                        'vpc_domain_policy':'default',
                    }
                    kwargs['class_path'] = 'switch,vpc_domains'
                    kwargs['easyDict'] = easy_functions.ez_append(vpcArgs, **kwargs)

        # Modify the Format of the IP Addressing
        polVars.update({
            'interfaces':[],
            'name':polVars['switch_name'],
        })
        if not polVars['inband_ipv4_gateway'] == None or not polVars['inband_ipv6_gateway'] == None:
            polVars['inband_addressing'] = {'management_epg':polVars['inband_mgmt_epg']}
        if not polVars['ooband_ipv4_gateway'] == None or not polVars['ooband_ipv6_gateway'] == None:
            polVars['ooband_addressing'] = {'management_epg':polVars['ooband_mgmt_epg']}
        if not polVars['inband_ipv4_gateway'] == None:
            polVars['inband_addressing'].update({
                'ipv4_address':polVars['inband_ipv4'], 'ipv4_gateway':polVars['inband_ipv4_gateway']
            })
        if not polVars['inband_ipv6_gateway'] == None:
            polVars['inband_addressing'].update({
                'ipv6_address':polVars['inband_ipv6'], 'ipv6_gateway':polVars['inband_ipv6_gateway']
            })
        if not polVars['ooband_ipv4_gateway'] == None:
            polVars['ooband_addressing'].update({
                'ipv4_address':polVars['ooband_ipv4'], 'ipv4_gateway':polVars['ooband_ipv4_gateway']
            })
        if not polVars['inband_ipv6_gateway'] == None:
            polVars['inband_addressing'].update({
                'ipv6_address':polVars['ooband_ipv6'], 'ipv6_gateway':polVars['ooband_ipv6_gateway']
            })
        ptypes = ['ipv4', 'ipv6']
        mtypes = ['inband', 'ooband']
        for mtype in mtypes:
            polVars.pop(f'{mtype}_mgmt_epg')
            for ptype in ptypes:
                polVars.pop(f'{mtype}_{ptype}')
                polVars.pop(f'{mtype}_{ptype}_gateway')

        pop_list = ['vpc_name', 'vpc_domain_id']
        for i in pop_list:
            if not polVars.get(i) == None: polVars.pop(i)

        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'switch,switch_profiles'
        kwargs['easyDict'] = easy_functions.ez_append(polVars, **kwargs)

        # Create or Modify the Interface Selector Workbook
        siteDict = kwargs['easyDict']['sites'][kwargs['site_group']]['site_settings']
        kwargs['excel_workbook'] = '%s_interface_selectors.xlsx' % (siteDict['site_name'])
        kwargs['wb_sw'] = load_workbook(kwargs['excel_workbook'])
        easy_functions.interface_selector_workbook(polVars, **kwargs)

        # Remove Site Worksheet if it Exists
        ws_site = kwargs['wb_sw'].get_sheet_names()
        for sheetName in ws_site:
            if sheetName in ['Sites']:
                sheetToDelete = kwargs['wb_sw'].get_sheet_by_name(sheetName)
                kwargs['wb_sw'].remove_sheet(sheetToDelete)
                kwargs['wb_sw'].save(kwargs['excel_workbook'])

        # Set the wb and ws before it is over-written
        wb = kwargs['wb']
        ws = kwargs['ws']

        # Evaluate The Interface Selectors Worksheet in the Site Workbook
        wb = kwargs['wb_sw']
        class_init = 'switches'
        class_folder = 'switches'
        func_regex = '^intf_selector$'
        ws = wb[f"{polVars['switch_name']}"]
        rows = ws.max_row
        func_list = easy_functions.findKeys(ws, func_regex)
        easy_functions.stdout_log(ws, None, 'begin')
        for func in func_list:
            count = easy_functions.countKeys(ws, func)
            var_dict = easy_functions.findVars(ws, func, rows, count)
            for pos in var_dict:
                row_num = var_dict[pos]['row']
                del var_dict[pos]['row']
                easy_functions.stdout_log(ws, row_num, 'begin')
                var_dict[pos].update(
                    {
                        'class_folder':class_folder,
                        'easyDict':kwargs['easyDict'],
                        'easy_jsonData':kwargs['easy_jsonData'],
                        'row_num':row_num,
                        'wb':wb,
                        'ws':ws
                    }
                )
                easyDict = eval(f"{class_init}(class_folder).{func}(**var_dict[pos])")

        # Set the wb and ws back
        kwargs['wb'] = wb
        kwargs['ws'] = ws
        kwargs['wb_sw'].close()
        return easyDict

    #=============================================================================
    # Function - Switch Modules
    #=============================================================================
    def sw_modules(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['access.profiles.switchModules']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)

        # Split the Node List into Nodes
        polVars['node_list'] = polVars['node_list'].split(',')
 
        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'switch,spine_modules'
        kwargs['easyDict'] = easy_functions.ez_append(polVars, **kwargs)
        return kwargs['easyDict']

#=====================================================================================
# Please Refer to the "Notes" in the relevant column headers in the input Spreadhseet
# for detailed information on the Arguments used by this Class.
#=====================================================================================
class site_policies(object):
    def __init__(self, type):
        self.type = type

    #=============================================================================
    # Function - Site Settings
    #=============================================================================
    def site_id(self, **kwargs):
        args = kwargs['args']
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['site.Identifiers']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)

        kwargs["multi_select"] = False
        # Prompt User for the Version of the Controller
        if polVars['controller_type'] == 'apic':
            # Obtain the APIC version from the API
            polVars['easyDict'] = kwargs['easyDict']
            polVars['jsonData'] = jsonData
            polVars["Variable"] = 'apicPass'
            apic_pass = easy_functions.sensitive_var_value(**polVars)
            pop_list = ['easyDict', 'jsonData', 'Variable']
            for i in pop_list:
                if not polVars.get(i) == None: polVars.pop(i)

            if args.skip_version_check == True:
                polVars['version'] = '4.2(7m)'
            else:
                if not kwargs['login_domain'] == None:
                    apic_user = f"apic#{kwargs['login_domain']}\\{kwargs['username']}"
                else:
                    apic_user = kwargs['username']
                fablogin = apicLogin(kwargs['controller'], apic_user, apic_pass)
                cookies = fablogin.login()

                # Locate template for method
                template_file = "aaaRefresh.json"
                uri = 'api/aaaRefresh'
                uriResponse = easy_functions.apic_api(
                    kwargs['controller'], 'get', {}, cookies, uri, template_file
                )
                verJson = uriResponse.json()
                polVars['version'] = verJson['imdata'][0]['aaaLogin']['attributes']['version']
        else:
            # Obtain the NDO version from the API
            polVars['easyDict'] = kwargs['easyDict']
            polVars['jsonData'] = jsonData
            polVars["Variable"] = 'ndoPass'
            ndo_domain = kwargs['login_domain']
            ndo_user = kwargs['username']
            ndo_pass = easy_functions.sensitive_var_value(**polVars)
            pop_list = ['easyDict', 'jsonData', 'Variable']
            for i in pop_list:
                if not polVars.get(i) == None: polVars.pop(i)

            if kwargs['args'].skip_version_check == 'True':
                polVars['version'] = '3.2(7l)'
            else:
                fablogin = ndoLogin(kwargs['controller'], ndo_domain, ndo_pass, ndo_user)
                cookies = fablogin.login()

                # Locate template for method and obtain running Version
                template_file = "aaaRefresh.json"
                uri = 'mso/api/v1/platform/version'
                uriResponse = easy_functions.ndo_api(
                    kwargs['controller'], 'get', cookies, uri, template_file
                )
                verJson = uriResponse.json()
                polVars['version'] = verJson['version']

        if polVars['controller_type'] == 'apic': 
            site_wb = '%s_interface_selectors.xlsx' % (kwargs['site_name'])
            if not os.path.isfile(site_wb):
                kwargs['wb'].save(filename=site_wb)
                wb_wr = load_workbook(site_wb)
                ws_wr = wb_wr.get_sheet_names()
                for sheetName in ws_wr:
                    if sheetName not in ['Sites']:
                        sheetToDelete = wb_wr.get_sheet_by_name(sheetName)
                        wb_wr.remove_sheet(sheetToDelete)
                wb_wr.save(filename=site_wb)

        # Attach the Site Dictionary
        siteDict = kwargs['easy_jsonData']['easy_aci']['allOf'][1]['properties']['siteDict']
        if not kwargs['easyDict'].get('sites'):
            kwargs['easyDict'].update(deepcopy({'sites':{}}))
        if not kwargs['easyDict'].get('tmp'):
            kwargs['easyDict'].update(deepcopy({'tmp':{}}))
        kwargs['easyDict']['sites'].update(deepcopy({polVars['site_id']:siteDict}))
        
        # Attach the Site Settings
        kwargs['easyDict']['sites'][polVars['site_id']]['site_settings'].update(deepcopy(polVars))
        return kwargs['easyDict']

    #=============================================================================
    # Function - Site Settings
    #=============================================================================
    def site_settings(self, **kwargs):
        tfc_config = 0
        easyDict = kwargs['easyDict']
        jsonData = kwargs['easy_jsonData']['easy_aci']['allOf'][1]['properties']
        for k, v in easyDict['sites'].items():
            polVars = {}
            polVars['annotation'] = 'orchestrator:terraform:easy-aci-v%s' % (jsonData['version'])
            polVars['annotations'] = [{
                'key':'orchestrator',
                'value':f'terraform:easy-aci:v{jsonData["version"]}'
            }]
            polVars['controller_type'] = v['site_settings']['controller_type']
            if v['site_settings']['controller_type'] == 'apic':
                polVars['apicHostname'] = v['site_settings']['controller']
                polVars['apic_version'] = v['site_settings']['version']
                if v['site_settings']['auth_type'] == 'username':
                    if not v['site_settings']['login_domain'] == None:
                        login_domain = v['site_settings']['login_domain']
                        username = v['site_settings']['username']
                        polVars['apicUser'] = f"apic#{login_domain}\\{username}"
                    else: polVars['apicUser'] = v['site_settings']['username']
            else:
                polVars['ndoHostname'] = v['site_settings']['controller']
                polVars['ndoUser'] = v['site_settings']['username']
                polVars['ndo_version'] = v['site_settings']['version']
                if not v['site_settings']['login_domain'] == None:
                    polVars['ndoDomain'] = v['site_settings']['login_domain']
            
            # Assign Management EPGs
            polVars['management_epgs'] = kwargs['easyDict']['tmp']['management_epgs']
            
            # siteDirs = next(os.walk(os.path.join(args.dir, site_name)))[1]
            kwargs['auth_type'] = v['site_settings']['auth_type']
            kwargs['class_type'] = 'sites'
            kwargs['controller_type'] = v['site_settings']['controller_type']
            kwargs["initial_write"] = True
            kwargs['site'] = k
            kwargs['site_name'] = v['site_settings']['site_name']
            kwargs['template_file'] = 'variables.j2'
            kwargs['tf_file'] = 'variables.auto.tfvars'

            kwargs["dest_dir"] = ''
            polVars = OrderedDict(sorted(polVars.items()))
            easy_functions.write_to_site(polVars, **kwargs)

            polVars = {
                "aci_provider_version": easyDict['latest_versions']['aci_provider_version'],
                "ndo_provider_version": easyDict['latest_versions']['ndo_provider_version'],
                "terraform_version": easyDict['latest_versions']['terraform_version'],
                "utils_provider_version": easyDict['latest_versions']['utils_provider_version']
            }
            kwargs["template_file"] = 'provider.j2'
            kwargs["tf_file"] = 'provider.tf'
            polVars = OrderedDict(sorted(polVars.items()))
            easy_functions.write_to_site(polVars, **kwargs)

            if v['site_settings']['run_location'] == 'tfc' and v['site_settings']['configure_terraform_cloud'] == True:
                tfc_config += 1
                
        if tfc_config > 0: terraform_cloud().create_terraform_workspaces(**kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Site Groups
    #=============================================================================
    def group_id(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['site.Groups']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)

        polVars['row_num'] = kwargs['row_num']
        polVars['ws'] = kwargs['ws']
        validating.site_groups(**polVars)
        polVars.pop('row_num')
        polVars.pop('ws')

        sites = []
        for x in range(1, 11):
            if not kwargs[f'site_{x}'] == None:
                sites.append(kwargs[f'site_{x}'])

        # Save the Site Information into Environment Variables
        os.environ[kwargs['site_group']] = '%s' % (polVars)


        # Add Site Group to easyDict
        polVars = {
            'site_group':kwargs['site_group'],
            'sites':sites,
        }
        if not kwargs['easyDict'].get('site_groups'):
            kwargs['easyDict'].update(deepcopy({'site_groups':{}}))
        kwargs['easyDict']['site_groups'].update(deepcopy({kwargs['site_group']:{'sites':sites}}))
        return kwargs['easyDict']

#=====================================================================================
# Please Refer to the "Notes" in the relevant column headers in the input Spreadhseet
# for detailed information on the Arguments used by this Class.
#=====================================================================================
class system_settings(object):
    def __init__(self, type):
        self.type = type

    #=============================================================================
    # Function - APIC Connectivity Preference
    #=============================================================================
    def apic_preference(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['system.apicConnectivityPreference']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)

        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'system_settings,apic_connectivity_preference'
        kwargs['easyDict'] = easy_functions.ez_update(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - BGP Autonomous System Number
    #=============================================================================
    def bgp_asn(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['system.bgpASN']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)

        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'system_settings,bgp'
        kwargs['easyDict'] = easy_functions.ez_append(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - BGP Route Reflectors
    #=============================================================================
    def bgp_rr(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['system.bgpRouteReflector']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        polVars['pod_id'] = int(polVars['pod_id'])
        polVars['route_reflector_nodes'] = [eval(i) for i in polVars['route_reflector_nodes'].split(',')]

        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'system_settings,bgp_route_reflectors,pods'
        kwargs['easyDict'] = easy_functions.ez_append(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Recommended Settings
    #=============================================================================
    def recommended_settings(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['system.recommendedSettings']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        remove_temp = kwargs['easyDict']['remove_default_args']
        kwargs['easyDict']['remove_default_args'] = False
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        kwargs['easyDict']['remove_default_args'] == remove_temp
        
        # Check if the Sensitive Variables are in the Environment.  If not Add them.
        if kwargs['global_aes_encryption_settings'] == True:
            sensitive_list = ['aes_passphrase']
            polVars, kwargs = easy_functions.env_sensitive(sensitive_list, jsonData, polVars, **kwargs)

        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'system_settings,recommended_settings'
        kwargs['easyDict'] = easy_functions.ez_update(polVars, **kwargs)
        return kwargs['easyDict']

#=====================================================================================
# Please Refer to the "Notes" in the relevant column headers in the input Spreadhseet
# for detailed information on the Arguments used by this Class.
#=====================================================================================
class tenants(object):
    def __init__(self, type):
        self.type = type

    #=============================================================================
    # Function - APIC Inband Configuration
    #=============================================================================
    def apic_inb(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['tenants.apic.InbandMgmt']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        kwargs['tenant'] = 'mgmt'
        polVars['management_epg'] = polVars['inband_epg']
        polVars.pop('inband_epg')

        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'node_management_addresses,static_node_management_addresses,apics_inband'
        kwargs['easyDict'] = easy_functions.ez_tenants_append(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Application Profiles
    #=============================================================================
    def app_add(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['tenants.applicationProfiles']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        polVars['monitoring_policy'] = 'default'

        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'application_profiles'
        kwargs['easyDict'] = easy_functions.ez_tenants_append(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Bridge Domains
    #=============================================================================
    def bd_add(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['tenants.bridgeDomains']['allOf'][1]['properties']

        args_add = []
        if not kwargs['epg_template'] == None:
            args_add.extend(['application_profile', 'epg_template'])
        jsonData = easy_functions.args_add(args_add, jsonData)
        
        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)

        # Remove Items in the Pop List
        jsonData = easy_functions.args_remove(args_add, jsonData)

        if not polVars['annotations'] == None:
            polVars['annotations'] = easy_functions.annotations_split(polVars['annotations'])
        polVars['gateway_ips'] = polVars['gateway_ips'].split(',')
        polVars['l3outs'] = polVars['l3outs'].split(',')
        polVars['subnet_templates'] = polVars['subnet_templates'].split(',')
        subs = []
        if len(polVars['subnet_templates']) == 1:
            for i in polVars['gateway_ips']:
                subs.append({'gateway_ip':i,'template':polVars['subnet_templates'][0]})
        elif len(polVars['subnet_templates']) >= 1:
            if len(polVars['subnet_templates']) == len(polVars['gateway_ips']):
                for i in range(0, len(polVars['subnet_templates']) + 1):
                    subs.append({'gateway_ip':polVars['gateway_ips'][i],'template':polVars['subnet_templates'][i]})
            else: validating.error_policy_match('gateway_ips', 'subnet_templates', **kwargs)
        if len(polVars['subnet_templates']) > 0:
            for i in polVars['subnet_templates']:
                easy_functions.confirm_templates_exist('subnets', i, **kwargs)
        else: subs = None
        # Re-Classify the Application EPG Template
        if not polVars['epg_template'] == None:
            easy_functions.confirm_templates_exist('application_epgs', polVars['epg_template'], **kwargs)
            polVars['application_epg'] = {'application_profile':polVars['application_profile'],'template':polVars['epg_template']}
            if not polVars['vlans'] == None:
                polVars['epg_to_aaep_vlans'] = [eval(i) for i in polVars['vlans'].split(',')]
                polVars.pop('vlans')
        if not polVars.get('vlans') == None:
            polVars['vlans'] = [eval(i) for i in polVars['vlans'].split(',')]
        
        # Re-classify the Bridge Domain Template
        easy_functions.confirm_templates_exist('bridge_domains', polVars['bd_template'], **kwargs)
        polVars['subnets'] = subs
        polVars['template'] = polVars['bd_template']

        kwargs['class_path'] = 'templates,bridge_domains,l3_configurations,associated_l3outs'
        kwargs['policy'] = 'template_name'
        kwargs['policy_name'] = polVars['bd_template']
        kwargs['easyDict'] = easy_functions.ez_append_l3out({'l3outs':polVars['l3outs']}, **kwargs)

        pop_list = ['bd_template', 'epg_template', 'gateway_ips', 'l3outs', 'subnet_templates', 'tenant']
        for i in pop_list:
            if not polVars.get(i) == None: polVars.pop(i)

        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'networking,bridge_domains'
        kwargs['easyDict'] = easy_functions.ez_tenants_append(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Bridge Domains - Templates
    #=============================================================================
    def bd_template(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['tenants.bd.Templates']['allOf'][1]['properties']

        # If DHCP Names defined add scope as required Argument
        args_list = []
        if not kwargs['names'] == None:
            args_list.extend(['names', 'scope'])
        jsonData = easy_functions.args_add(args_list, jsonData)
        
        # Build Dictionary of Policy Variables
        remove_temp = kwargs['easyDict']['remove_default_args']
        kwargs['easyDict']['remove_default_args'] = False
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        kwargs['easyDict']['remove_default_args'] == remove_temp

        # Remove Items in the Pop List
        jsonData = easy_functions.args_remove(args_list, jsonData)

        # Begin Modifying the Variables for Output File Format
        # DHCP Relay Policy
        if not polVars['names'] == None:
            dhcp = {
                'dhcp_option_policy': kwargs['dhcp_option_policy'],
                'names': kwargs['names'].split(','),
                'scope': kwargs['scope']
            }
            dhcp = easy_functions.ez_remove_empty(dhcp)
            dhcp = [dhcp]
        else: dhcp = None

        # L3Out and VRF Configuration
        l3outs = [{'tenant': polVars['vrf_tenant']}]
        vrf = {'name': polVars['vrf'], 'tenant': polVars['vrf_tenant']}
        #l3outs = easy_functions.ez_remove_empty(l3outs)
        vrf = easy_functions.ez_remove_empty(vrf)
        if not polVars['ndo_settings'] == None:
            if kwargs['easyDict']['tmp']['ndo_settings'].get(polVars['ndo_settings']):
                ndo_settings = kwargs['easyDict']['tmp']['ndo_settings'][polVars['ndo_settings']]
            else: validating.error_schema('ndo_settings', **kwargs)
            vrf['schema'] = ndo_settings['vrf_schema']
            vrf['template'] = ndo_settings['vrf_template']
            ndo = {'template': ndo_settings['template'], 'sites': ndo_settings['sites']}
        else: ndo = None

        # Remove Arguments from Dictionary
        pop_list = ['dhcp_option_policy', 'names', 'scope', 'vrf', 'vrf_tenant']
        for i in pop_list:
            if not polVars.get(i) == None: polVars.pop(i)

        # Configure Bridge Domain Policy Tabs
        l3_tab = {
            'associated_l3outs':l3outs,
            'ep_move_detection_mode':polVars.get('ep_move_detection_mode'),
            'nd_policy':polVars.get('nd_policy'),
            'route_profile':polVars.get('route_profile'),
            'unicast_routing':polVars.get('unicast_routing')
        }
        pop_list = ['ep_move_detection_mode', 'nd_policy', 'route_profile', 'unicast_routing']
        for i in pop_list:
            if not polVars.get(i) == None: polVars.pop(i)
        l3_tab = easy_functions.ez_remove_empty(l3_tab)

        if not polVars['netflow_monitor_policies'] == None:
            polVars['netflow_monitor_policies'] = polVars['netflow_monitor_policies'].split(',')
        if not polVars['rogue_coop_exception_list'] == None:
            polVars['rogue_coop_exception_list'] = polVars['rogue_coop_exception_list'].split(',')
        general_tab = deepcopy(polVars)
        pop_list = [
            'disable_ip_data_plane_learning_for_pbr', 'first_hop_security_policy',
            'intersite_l2_stretch', 'intersite_bum_traffic_allow', 'optimize_wan_bandwidth',
            'netflow_monitor_policies', 'rogue_coop_exception_list'
        ]
        for i in pop_list: 
            if not general_tab.get(i) == None: general_tab.pop(i)
        general_tab = easy_functions.ez_remove_empty(general_tab)
        pop_list = list(general_tab.keys())
        general_tab['vrf'] = vrf

        atr_tab = deepcopy(polVars)
        for i in pop_list:
            if not atr_tab.get(i) == None: atr_tab.pop(i)
        atr_tab = easy_functions.ez_remove_empty(atr_tab)

        polVars = {
            'advanced/troubleshooting': atr_tab,
            'dhcp_relay_labels': dhcp,
            'general': general_tab,
            'l3_configurations': l3_tab,
            'ndo': ndo,
            'site_group': kwargs['site_group'],
            'template_name': kwargs['template_name']
        }
        polVars = easy_functions.ez_remove_empty(polVars)

        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'templates,bridge_domains'
        kwargs['easyDict'] = easy_functions.ez_append(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Bridge Domain - Subnets Templates
    #=============================================================================
    def subnet_template(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['tenants.bd.SubnetsTemplates']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        remove_temp = kwargs['easyDict']['remove_default_args']
        kwargs['easyDict']['remove_default_args'] = False
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        kwargs['easyDict']['remove_default_args'] == remove_temp

        # Modify the polVars scope and subnet_control
        polVars['scope'] = {
            'advertise_externally':polVars['advertise_externally'],
            'shared_between_vrfs':polVars['shared_between_vrfs']
        }
        polVars['subnet_control'] = {
            'neighbor_discovery':polVars['neighbor_discovery'],
            'no_default_svi_gateway':polVars['no_default_svi_gateway'],
            'querier_ip':polVars['querier_ip']
        }
        pop_list = [
            'advertise_externally', 'neighbor_discovery', 'no_default_svi_gateway',
            'querier_ip', 'shared_between_vrfs',
        ]
        for i in pop_list:
            if not polVars.get(i) == None: polVars.pop(i)
        
        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'templates,subnets'
        kwargs['easyDict'] = easy_functions.ez_append(polVars, **kwargs)
        return kwargs['easyDict']
        
    #=============================================================================
    # Function - Application Profiles
    #=============================================================================
    def bgp_pfx(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['tenants.policies.bgpPrefix']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        polVars.pop('tenant')
        
        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'policies,protocol,bgp,bgp_peer_prefix'
        kwargs['easyDict'] = easy_functions.ez_tenants_append(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - L3out - BGP Peer Connectivity Profile - Templates
    #=============================================================================
    def bgp_template(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['tenants.l3out.bgpPeerConnectivityProfile.Templates']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)

        # Modify the polVars Address Type Controls, BGP Controls, Peer Controls, and Private AS Controls
        polVars['address_type_controls'] = {
            'af_mcast':polVars.get('af_mcast'),
            'af_ucast':polVars.get('af_ucast')
        }
        polVars['address_type_controls'] = easy_functions.ez_remove_empty(polVars['address_type_controls'])
        if len(polVars['address_type_controls']) == 0:
            polVars.pop('address_type_controls')
        polVars['bgp_controls'] = {
            'allow_self_as':polVars.get('allow_self_as'),
            'as_override':polVars.get('as_override'),
            'disable_peer_as_check':polVars.get('disable_peer_as_check'),
            'next_hop_self':polVars.get('next_hop_self'),
            'send_community':polVars.get('send_community'),
            'send_domain_path':polVars.get('send_domain_path'),
            'send_extended_community':polVars.get('send_extended_community')
        }
        polVars['bgp_controls'] = easy_functions.ez_remove_empty(polVars['bgp_controls'])
        if len(polVars['bgp_controls']) == 0:
            polVars.pop('bgp_controls')
        polVars['peer_controls'] = {
            'bidirectional_forwarding_detection':polVars.get('bidirectional_forwarding_detection'),
            'disable_connected_check':polVars.get('disable_connected_check')
        }
        polVars['peer_controls'] = easy_functions.ez_remove_empty(polVars['peer_controls'])
        if len(polVars['peer_controls']) == 0:
            polVars.pop('peer_controls')
        polVars['private_as_control'] = {
            'remove_all_private_as':polVars.get('remove_all_private_as'),
            'remove_private_as':polVars.get('remove_private_as'),
            'replace_private_as_with_local_as':polVars.get('replace_private_as_with_local_as')
        }
        polVars['private_as_control'] = easy_functions.ez_remove_empty(polVars['private_as_control'])
        if len(polVars['private_as_control']) == 0:
            polVars.pop('private_as_control')
        pop_list = [
            'af_mcast', 'af_ucast', 'allow_self_as', 'as_override',
            'bidirectional_forwarding_detection', 'disable_connected_check',
            'disable_peer_as_check', 'next_hop_self', 'template_name',
            'remove_all_private_as', 'remove_private_as', 'replace_private_as_with_local_as',
            'send_community', 'send_domain_path', 'send_extended_community'
        ]
        for i in pop_list:
            if not polVars.get(i) == None: polVars.pop(i)

        # Add Policy Variables to easyDict
        if not kwargs['easyDict']['tmp'].get('bgp_templates'):
            kwargs['easyDict']['tmp'].update(deepcopy({'bgp_templates':{}}))
        kwargs['easyDict']['tmp']['bgp_templates'].update(deepcopy({f"{kwargs['template_name']}":polVars}))
        return kwargs['easyDict']

    #=============================================================================
    # Function - Contracts
    #=============================================================================
    def contract_add(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['tenants.contract.Contracts']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        if not polVars['annotations'] == None:
            polVars['annotations'] = easy_functions.annotations_split(polVars['annotations'])
        polVars['subjects'] = []

        # Add NDO Settings if Defined
        if not polVars['ndo_settings'] == None:
            if kwargs['easyDict']['tmp']['ndo_settings'].get(polVars['ndo_settings']):
                ndo_settings = kwargs['easyDict']['tmp']['ndo_settings'][polVars['ndo_settings']]
            else: validating.error_schema('ndo_settings', **kwargs)
            polVars['ndo'] = {
                'schema': ndo_settings['schema'],
                'sites': ndo_settings['sites'],
                'template': ndo_settings['template']
            }
        else: polVars['ndo'] = None
        
        # Remove Unused Variables
        pop_list = ['tenant']
        for i in pop_list:
            if not polVars.get(i) == None: polVars.pop(i)
        
        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'contracts,contracts'
        kwargs['easyDict'] = easy_functions.ez_tenants_append(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Application Profiles
    #=============================================================================
    def contract_assign(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['tenants.contract.ContractAssignments']['allOf'][1]['properties']

        # Add Required Arguments based on the type of Contract
        args_list = []
        if kwargs['target_type'] == 'external_epg':
            args_list = ['l3out', 'external_epgs']
        elif kwargs['target_type'] == 'epg':
            args_list = ['application_epgs', 'application_profile']
        elif re.search('^(inb|oob)$', kwargs['target_type']):
            args_list.append('application_epgs')
        elif kwargs['target_type'] == 'vrf':
            args_list.append('vrfs')
        jsonData = easy_functions.args_add(args_list, jsonData)
        
        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        if polVars['target_tenant'] == None:
            polVars['target_tenant'] = polVars['tenant']
        polVars['name'] = polVars['contract']

        # Remove Items in the Pop List
        jsonData = easy_functions.args_remove(args_list, jsonData)

        # Add NDO Settings if Defined
        if not polVars['ndo_settings'] == None:
            if kwargs['easyDict']['tmp']['ndo_settings'].get(polVars['ndo_settings']):
                ndo_settings = kwargs['easyDict']['tmp']['ndo_settings'][polVars['ndo_settings']]
            else: validating.error_schema('ndo_settings', **kwargs)
            polVars['ndo'] = {
                'schema': ndo_settings['schema'],
                'sites': ndo_settings['sites'],
                'template': ndo_settings['template']
            }
        else: polVars['ndo'] = None

        # Remove Unused Variables
        pop_list = ['contract', 'target_tenant', 'target_type']
        if kwargs['target_tenant'] == kwargs['tenant']: pop_list.append('tenant')
        if re.search('^(epg|inb|oob)$', kwargs['target_type']): pop_list.extend(['application_profile', 'application_epgs'])
        elif kwargs['target_type'] == 'external_epg': pop_list.extend(['external_epgs', 'l3out'])
        elif kwargs['target_type'] == 'vrf': pop_list.extend(['vrfs'])
        for i in pop_list:
            if not polVars.get(i) == None: polVars.pop(i)

        # Attach the Contract to the EPG and VRF Resource(s)
        if re.search('^(epg|inb|oob)$', kwargs['target_type']):
            for i in kwargs['application_epgs'].split(','):
                kwargs['class_path'] = 'application_profiles,application_epgs,contracts'
                kwargs['policy1'] = 'name'
                kwargs['policy_name1'] = kwargs['application_profile']
                kwargs['policy2'] = 'name'
                kwargs['policy_name2'] = i
                kwargs['easyDict'] = easy_functions.ez_tenants_append_sub_subtype(polVars, **kwargs)
        elif kwargs['target_type'] == 'external_epg':
            for i in kwargs['external_epgs'].split(','):
                kwargs['class_path'] = 'networking,l3outs,external_epgs,contracts'
                kwargs['policy1'] = 'name'
                kwargs['policy_name1'] = kwargs['l3out']
                kwargs['policy2'] = 'name'
                kwargs['policy_name2'] = i
                kwargs['easyDict'] = easy_functions.ez_tenants_append_sub_subtype(polVars, **kwargs)
        elif kwargs['target_type'] == 'vrf':
            for i in kwargs['vrfs'].split(','):
                kwargs['class_path'] = 'networking,vrfs,contracts'
                kwargs['policy1'] = 'name'
                kwargs['policy_name1'] = i
                kwargs['policy2'] = 'epg_esg_collection_for_vrfs'
                kwargs['policy_name2'] = 'epg_esg_collection_for_vrfs'
                kwargs['easyDict'] = easy_functions.ez_tenants_append_sub_subtype(polVars, **kwargs)
        
        # Return EasyDict
        return kwargs['easyDict']

    #=============================================================================
    # Function - Contracts - Add Subject
    #=============================================================================
    def contract_filters(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['tenants.contract.ContractFilters']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        polVars['directives'] = {
            'enable_policy_compression':polVars.get('enable_policy_compression'),
            'log':polVars.get('log_packets')
        }
        polVars['directives'] = easy_functions.ez_remove_empty(polVars['directives'])
        if len(polVars['directives']) == 0:
            polVars.pop('directives')
        polVars['filters'] = polVars['filters_to_assign'].split(',')
        pop_list = ['contract_name', 'enable_policy_compression', 'filters_to_assign', 'log_packets', 'tenant', ]
        for i in pop_list:
            if not polVars.get(i) == None: polVars.pop(i)

        # Add Dictionary to Policy
        kwargs['class_path'] = 'contracts,contracts,subjects'
        kwargs['policy'] = 'name'
        kwargs['policy_name'] = kwargs['contract_name']
        kwargs['easyDict'] = easy_functions.ez_tenants_append_subtype(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - EIGRP Interface Policy
    #=============================================================================
    def dhcp_relay(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['tenants.policies.dhcpRelay']['allOf'][1]['properties']

        # Add/Remove Required Arguments as Necessary
        args_list = []
        if kwargs['epg_type'] == 'external_epg':
            args_list.append('l3out')
        else: args_list.append('application_epg')
        jsonData = easy_functions.args_add(args_list, jsonData)

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        jsonData = easy_functions.args_remove(args_list, jsonData)
        names = polVars['names'].split(',')
        addresses = polVars['addresses'].split(',')
        if not len(names) == len(addresses):
            validating.error_policy_match('names', 'addresses', **kwargs)
        name_list = []
        for i in range(len(names)): name_list.append([names[i], addresses[i]])
        polVars = {
            'name_addr_list':name_list,
            'application_profile':polVars['application_profile'],
            'epg':polVars['epg'],
            'epg_type':polVars['epg_type'],
            'l3out':polVars['l3out'],
            'site_group':polVars['site_group'],
            'tenant':polVars['tenant']
        }
        # Add Policy Variables to easyDict
        if kwargs['owner'] == 'tenant':
            kwargs['class_path'] = 'tenants,policies,protocol,dhcp,relay_policies'
            kwargs['easyDict'] = easy_functions.ez_tenants_append(polVars, **kwargs)
        else:
            kwargs['class_path'] = 'access,policies,global,dhcp_relay'
            kwargs['easyDict'] = easy_functions.ez_append(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - EIGRP Interface Policy
    #=============================================================================
    def eigrp_interface(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['tenants.policies.eigrpInterface']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        polVars.pop('tenant')

        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'policies,protocol,eigrp,eigrp_interface'
        kwargs['easyDict'] = easy_functions.ez_tenants_append(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - EIGRP Interface Profile
    #=============================================================================
    def eigrp_profile(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['tenants.l3out.eigrpInterfaceProfile']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)

        polVars.pop('profile_name')

        # Add Policy Variables to easyDict
        if not kwargs['easyDict']['tmp'].get('eigrp_interface_profiles'):
            kwargs['easyDict']['tmp'].update(deepcopy({'eigrp_interface_profiles':{}}))
        kwargs['easyDict']['tmp']['eigrp_interface_profiles'].update(deepcopy({f"{kwargs['template_name']}":polVars}))
        return kwargs['easyDict']

    #=============================================================================
    # Function - Application EPG
    #=============================================================================
    def epg_add(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['tenants.applicationEpgs']['allOf'][1]['properties']

        # Confirm if the EPG Template Exists
        if not kwargs.get('epg_template'):
            print(f"Error on {kwargs['ws'].title}, row {kwargs['row']}, epg_template is required.")
            exit()
        easy_functions.confirm_templates_exist('application_epgs', kwargs['epg_template'], **kwargs)
        def get_epg_template(site, **kwargs):
            tempcount = 0
            for i in kwargs['easyDict']['sites'][site]['templates']['application_epgs']:
                if i['template_name'] == kwargs['epg_template']:
                    epg_template = i
                    tempcount += 1
            if tempcount == 0:
                validating.error_template_not_found('epg_template', **kwargs)
            return epg_template
        if 'Grp_' in kwargs['site_group']:
            if kwargs['easyDict']['site_groups'].get(kwargs['site_group']):
                sites = kwargs['easyDict']['site_groups'][kwargs['site_group']]['sites']
                epg_template = get_epg_template(sites[0], **kwargs)
            else:
                validating.error_site_group('site_group', **kwargs)
        else: epg_template = get_epg_template(kwargs['site_group'], **kwargs)
        kwargs.update(epg_template)
        if not kwargs['easyDict']['tmp'].get('management_epgs'):
            kwargs['easyDict']['tmp']['management_epgs'] = []
        args_add = []
        args_remove = []
        if kwargs['epg_type'] == 'inb':
            args_add.append('vlans')
            kwargs['easyDict']['tmp']['management_epgs'].append({'name':kwargs['name'],'type':kwargs['epg_type']})
        if kwargs['epg_type'] == 'oob':
            args_remove.append('bridge_domain')
            kwargs['easyDict']['tmp']['management_epgs'].append({'name':kwargs['name'],'type':kwargs['epg_type']})
        jsonData = easy_functions.args_add(args_add, jsonData)
        jsonData = easy_functions.args_remove(args_remove, jsonData)
        
        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)

        # Remove Items in the Pop List
        jsonData = easy_functions.args_add(args_remove, jsonData)
        jsonData = easy_functions.args_remove(args_add, jsonData)
        if not polVars['annotations'] == None:
            polVars['annotations'] = easy_functions.annotations_split(polVars['annotations'])
        polVars['monitoring_policy'] = 'default'
        if not polVars['vlans'] == None:
            polVars['vlans'] =[eval(i) for i in polVars['vlans'].split(',')]

        # Add Application EPG Template
        polVars['template'] = polVars['epg_template']
        pop_list = ['application_profile', 'epg_template', 'tenant']
        for i in pop_list:
            if not polVars.get(i) == None: polVars.pop(i)

        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'application_profiles,application_epgs'
        kwargs['policy'] = 'name'
        kwargs['policy_name'] = kwargs['application_profile']
        kwargs['easyDict'] = easy_functions.ez_tenants_append_subtype(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Application EPG - Templates
    #=============================================================================
    def epg_template(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['tenants.applicationEpg.Templates']['allOf'][1]['properties']

        args_list = []
        if not kwargs['epg_to_aaeps'] == None: args_list.extend(['epg_to_aaeps', 'epg_to_aaep_mode'])
        if not kwargs['vmm_domains'] == None: args_list.extend(['vmm_domains', 'vmm_template'])
        jsonData = easy_functions.args_add(args_list, jsonData)
        
        # Build Dictionary of Policy Variables
        remove_temp = kwargs['easyDict']['remove_default_args']
        kwargs['easyDict']['remove_default_args'] = False
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        kwargs['easyDict']['remove_default_args'] == remove_temp
        polVars['epg_type'] = kwargs['epg_type']

        # Remove Items in the Pop List
        jsonData = easy_functions.args_remove(args_list, jsonData)
        
        # Assign EPG to AAEP Mapping if it isn't Empty
        if not polVars['epg_to_aaeps'] == None:
            aaep_list = []
            for i in polVars['epg_to_aaeps'].split(','):
                aaep_list.append(deepcopy({
                    'aaep': i,
                    'mode': polVars['epg_to_aaep_mode'],
                }))
            polVars['epg_to_aaeps'] = aaep_list
        
        # Add Domain Mappings
        if not polVars['physical_domains'] == None or not polVars['vmm_domains'] == None:
            polVars['domains'] = []
            if not polVars['physical_domains'] == None:
                for i in polVars['physical_domains'].split(','):
                    polVars['domains'].append(deepcopy({'name':i}))
            if not polVars['vmm_domains'] == None:
                for i in polVars['vmm_domains'].split(','):
                    if kwargs['easyDict']['tmp'].get('vmm_templates'):
                        if kwargs['easyDict']['tmp']['vmm_templates'].get(polVars['vmm_template']):
                            vmm_template = deepcopy(kwargs['easyDict']['tmp']['vmm_templates'][polVars['vmm_template']])
                        else: validating.error_template_not_found('vmm_template', **kwargs)
                    else: validating.error_template_not_found('vmm_template', **kwargs)
                    vmm_template.update(deepcopy({'name':i,'domain_type':'vmm'}))
                    polVars['domains'].append(deepcopy(vmm_template))
        
        # Add NDO Settings if Defined
        if not polVars['ndo_settings'] == None:
            if kwargs['easyDict']['tmp']['ndo_settings'].get(polVars['ndo_settings']):
                ndo_settings = kwargs['easyDict']['tmp']['ndo_settings'][polVars['ndo_settings']]
            else: validating.error_schema('ndo_settings', **kwargs)
            polVars['ndo'] = {'template': ndo_settings['template'], 'sites': ndo_settings['sites']}
        else: polVars['ndo'] = None
        
        # Remove Unused Items
        pop_list = ['epg_to_aaep_mode', 'physical_domains', 'vmm_domains', 'vmm_template']
        for i in pop_list:
            if not polVars.get(i) == None: polVars.pop(i)

        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'templates,application_epgs'
        kwargs['easyDict'] = easy_functions.ez_append(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Application EPG - VMM Templates
    #=============================================================================
    def epg_vmm_temp(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['tenants.applicationEpg.VMMTemplates']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        polVars['security'] = {
            'allow_promiscuous': kwargs['allow_promiscuous'],
            'forged_transmits': kwargs['forged_transmits'],
            'mac_changes': kwargs['mac_changes'],
        }
        pop_list = ['allow_promiscuous', 'forged_transmits', 'mac_changes', 'template_name']
        for i in pop_list:
            if not polVars.get(i) == None: polVars.pop(i)

        # Add Policy Variables to easyDict
        if not kwargs['easyDict']['tmp'].get('vmm_templates'):
            kwargs['easyDict']['tmp'].update(deepcopy({'vmm_templates':{}}))
        kwargs['easyDict']['tmp']['vmm_templates'].update(deepcopy({f"{kwargs['template_name']}":polVars}))
        return kwargs['easyDict']

    #=============================================================================
    # Function - L3Out - Exteranl EPG
    #=============================================================================
    def ext_epg(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['tenants.l3out.externalEpg']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)

        # Attach the External EPG Policy Additional Attributes
        if kwargs['easyDict']['tmp'].get('external_epg_templates'):
            if kwargs['easyDict']['tmp']['external_epg_templates'].get(polVars['external_epg_template']):
                polVars.update(kwargs['easyDict']['tmp']['external_epg_templates'][polVars['external_epg_template']])
            else: validating.error_template_not_found('external_epg_template', **kwargs)
        else: validating.error_template_not_found('external_epg_template', **kwargs)

        pop_list = ['external_epg_template', 'l3out', 'template_name', 'tenant']
        for i in pop_list:
            if not polVars.get(i) == None: polVars.pop(i)

        # Add Dictionary to Policy
        kwargs['class_path'] = 'networking,l3outs,external_epgs'
        kwargs['policy'] = 'name'
        kwargs['policy_name'] = kwargs['l3out']
        kwargs['easyDict'] = easy_functions.ez_tenants_append_subtype(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - L3Out - External EPG - Templates
    #=============================================================================
    def ext_epg_temp(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['tenants.l3out.externalEpg.Templates']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        polVars = easy_functions.ez_remove_empty(polVars)

        # Add Policy Variables to easyDict
        if not kwargs['easyDict']['tmp'].get('external_epg_templates'):
            kwargs['easyDict']['tmp'].update(deepcopy({'external_epg_templates':{}}))
        kwargs['easyDict']['tmp']['external_epg_templates'].update(deepcopy({f"{kwargs['template_name']}":polVars}))
        return kwargs['easyDict']

    #=============================================================================
    # Function - Bridge Domain - Subnets
    #=============================================================================
    def ext_epg_sub(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['tenants.l3out.externalEpg.Subnet']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        polVars['subnets'] = polVars['subnets'].split(',')

        # Modify the polVars aggregate, external_epg_classification, and route_control
        polVars['aggregate'] = {
            'aggregate_export':polVars.get('aggregate_export'),
            'aggregate_import':polVars.get('aggregate_import'),
            'aggregate_shared_routes':polVars.get('aggregate_shared_routes')
        }
        polVars['aggregate'] = easy_functions.ez_remove_empty(polVars['aggregate'])
        polVars['external_epg_classification'] = {
            'external_subnets_for_external_epg':polVars.get('external_subnets_for_external_epg'),
            'shared_security_import_subnet':polVars.get('shared_security_import_subnet')
        }
        polVars['external_epg_classification'] = easy_functions.ez_remove_empty(polVars['external_epg_classification'])
        polVars['route_control'] = {
            'export_route_control_subnet':polVars.get('export_route_control_subnet'),
            'import_route_control_subnet':polVars.get('import_route_control_subnet'),
            'shared_route_control_subnet':polVars.get('shared_route_control_subnet')
        }
        polVars['route_control'] = easy_functions.ez_remove_empty(polVars['route_control'])
        pop_list = list(polVars['aggregate'].keys())
        pop_list.extend(list(polVars['external_epg_classification'].keys()))
        pop_list.extend(list(polVars['route_control'].keys()))
        pop_list.extend(['external_epg', 'l3out', 'tenant'])
        for i in pop_list:
            if not polVars.get(i) == None: polVars.pop(i)
        if len(polVars['aggregate']) == 0:
            polVars.pop('aggregate')
        if len(polVars['external_epg_classification']) == 0:
            polVars.pop('external_epg_classification')
        if len(polVars['route_control']) == 0:
            polVars.pop('route_control')

        # Add Dictionary to Policy
        kwargs['class_path'] = 'networking,l3outs,external_epgs,subnets'
        kwargs['policy1'] = 'name'
        kwargs['policy_name1'] = kwargs['l3out']
        kwargs['policy2'] = 'name'
        kwargs['policy_name2'] = kwargs['external_epg']
        kwargs['easyDict'] = easy_functions.ez_tenants_append_sub_subtype(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Contract Filter
    #=============================================================================
    def filter_add(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['tenants.contract.Filters']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        if not polVars['annotations'] == None:
            polVars['annotations'] = easy_functions.annotations_split(polVars['annotations'])
        polVars['filter_entries'] = []
        pop_list = ['tenant']
        for i in pop_list:
            if not polVars.get(i) == None: polVars.pop(i)

        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'contracts,filters'
        kwargs['easyDict'] = easy_functions.ez_tenants_append(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Contract Filter - Filter Entry
    #=============================================================================
    def filter_entry(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['tenants.contract.filterEntry']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        if polVars.get('ip_protocol'):
            if polVars['ip_protocol'] == 'tcp':
                polVars['tcp_session_rules'] = {
                    'acknowledgement': polVars.get('acknowledgement'),
                    'established': polVars.get('established'),
                    'finish': polVars.get('finish'),
                    'reset': polVars.get('reset'),
                    'synchronize': polVars.get('synchronize'),
                }
                polVars['tcp_session_rules'] = easy_functions.ez_remove_empty(polVars['tcp_session_rules'])
                if len(polVars['tcp_session_rules']) == 0:
                    polVars.pop('tcp_session_rules')
        pop_list = ['acknowledgement', 'established', 'filter_name', 'finish', 'reset', 'synchronize', 'tenant']
        for i in pop_list:
            if not polVars.get(i) == None: polVars.pop(i)

        # Add Dictionary to Policy
        kwargs['class_path'] = 'contracts,filters,filter_entries'
        kwargs['policy'] = 'name'
        kwargs['policy_name'] = kwargs['filter_name']
        kwargs['easyDict'] = easy_functions.ez_tenants_append_subtype(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - L3Out
    #=============================================================================
    def l3out_add(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['tenants.L3Out']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        polVars['external_epgs'] = []

        # Attach the L3Out Template, for the Additional Attributes
        if kwargs['easyDict']['tmp'].get('l3out_templates'):
            if kwargs['easyDict']['tmp']['l3out_templates'].get(polVars['l3out_template']):
                polVars.update(kwargs['easyDict']['tmp']['l3out_templates'][polVars['l3out_template']])
            else: validating.error_template_not_found('l3out_template', **kwargs)
        else: validating.error_template_not_found('l3out_template', **kwargs)

        # Attach the OSPF Routing Profile if defined
        if not polVars['ospf_external_profile'] == None:
            if kwargs['easyDict']['tmp'].get('ospf_routing'):
                ospfp = kwargs['easyDict']['tmp']['ospf_routing']
                if kwargs['easyDict']['tmp']['ospf_routing'].get(polVars['ospf_external_profile']):
                    polVars['ospf_external_profile'] = ospfp[polVars['ospf_external_profile']]
                else: validating.error_template_not_found('ospf_external_profile', **kwargs)
            else: validating.error_template_not_found('ospf_external_profile', **kwargs)
        
        # Attach the Nexus Dashboard Orchestrator Settings if defined
        if not polVars['ndo_settings'] == None:
            if kwargs['easyDict']['tmp']['ndo_settings'].get(polVars['ndo_settings']):
                ndo_settings = kwargs['easyDict']['tmp']['ndo_settings'][polVars['ndo_settings']]
            else: validating.error_schema('ndo_settings', **kwargs)
            polVars['ndo'] = {'template': ndo_settings['template'], 'sites': ndo_settings['sites']}
        else: polVars['ndo'] = None

        pop_list = [ 'l3out_template', 'ndo_settings', 'ospf_external_profile', 'tenant']
        for i in pop_list:
            if not polVars.get(i) == None: polVars.pop(i)

        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'networking,l3outs'
        kwargs['easyDict'] = easy_functions.ez_tenants_append(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - L3Out - Templates
    #=============================================================================
    def l3out_template(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['tenants.L3Out.Templates']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)

        # Modify the polVars Route Control Enforcement
        polVars['route_control_enforcement'] = {
            'export':polVars.get('export'),
            'import':polVars.get('import')
        }
        polVars['route_control_enforcement'] = easy_functions.ez_remove_empty(polVars['route_control_enforcement'])
        pop_list = ['export', 'import', 'template_name']
        for i in pop_list:
            if not polVars.get(i) == None: polVars.pop(i)
        polVars = easy_functions.ez_remove_empty(polVars)

        # Add Policy Variables to easyDict
        if not kwargs['easyDict']['tmp'].get('l3out_templates'):
            kwargs['easyDict']['tmp'].update(deepcopy({'l3out_templates':{}}))
        kwargs['easyDict']['tmp']['l3out_templates'].update(deepcopy({f"{kwargs['template_name']}":polVars}))
        return kwargs['easyDict']

    #=============================================================================
    # Function - Tenants
    #=============================================================================
    def ndo_schema(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['tenants.ndoSchemas']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        polVars['sites'] = polVars['sites'].split(',')

        # Add Policy Variables to easyDict
        if not kwargs['easyDict']['tmp'].get('ndo_settings'):
            kwargs['easyDict']['tmp'].update(deepcopy({'ndo_settings':{}}))
        kwargs['easyDict']['tmp']['ndo_settings'].update(deepcopy({polVars['ndo_settings']:polVars}))
        if not kwargs['easyDict']['tmp'].get('ndo_schemas'):
            kwargs['easyDict']['tmp'].update({'ndo_schemas':[]})
        scount = 0
        for item in kwargs['easyDict']['tmp']['ndo_schemas']:
            if item['name'] == polVars['schema']:
                item['templates'].append(deepcopy({'name':polVars['template'],'sites':polVars['sites']}))
                scount += 1
        if scount == 0:
            polVars['name'] = polVars['schema']
            polVars['templates'] = []
            polVars['templates'].append(deepcopy(
                {'name':polVars['template'],'sites':polVars['sites']}
            ))
            pop_list = ['ndo_settings', 'schema', 'template', 'sites', 'vrf_schema', 'vrf_template']
            for i in pop_list:
                if not polVars.get(i) == None: polVars.pop(i)
            kwargs['easyDict']['tmp']['ndo_schemas'].append(polVars)
        return kwargs['easyDict']

    #=============================================================================
    # Function - L3Out - Exteranl EPG
    #=============================================================================
    def node_interface(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['tenants.l3out.logicalNodeInterfaceProfile']['allOf'][1]['properties']

        args_list = []
        if not kwargs['bgp_peers'] == None: args_list.append('bgp_template')
        jsonData = easy_functions.args_add(args_list, jsonData)
        
        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)

        # Remove Items in the Pop List
        jsonData = easy_functions.args_remove(args_list, jsonData)
        
        # Attach the Node Interface Template, for the Additional Attributes
        if kwargs['easyDict']['tmp'].get('node_intf_templates'):
            if kwargs['easyDict']['tmp']['node_intf_templates'].get(polVars['interface_template']):
                polVars.update(kwargs['easyDict']['tmp']['node_intf_templates'][polVars['interface_template']])
            else: validating.error_template_not_found('interface_template', **kwargs)
        else: validating.error_template_not_found('interface_template', **kwargs)

        # Attach the Node Interface Configuration Template, for the Additional Attributes
        if kwargs['easyDict']['tmp'].get('node_intf_config'):
            if kwargs['easyDict']['tmp']['node_intf_config'].get(polVars['interface_config_template']):
                polVars.update(kwargs['easyDict']['tmp']['node_intf_config'][polVars['interface_config_template']])
            else: validating.error_template_not_found('interface_config_template', **kwargs)
        else: validating.error_template_not_found('interface_config_template', **kwargs)

        # Attach the BGP Peers if defined
        addresses = deepcopy(polVars['bgp_peers'])
        if not addresses == None:
            if kwargs['easyDict']['tmp'].get('bgp_templates'):
                if kwargs['easyDict']['tmp']['bgp_templates'].get(polVars['bgp_template']):
                    polVars['bgp_peers'] = kwargs['easyDict']['tmp']['bgp_templates'][polVars['bgp_template']]
                    polVars['bgp_peers']['addresses'] = addresses.split(',')
                    if not polVars['bgp_password'] == None:
                        # Check Environment for VMM Credentials Password
                        sensitive_list = [f"bgp_password_{polVars['bgp_password']}"]
                        polVars, kwargs = easy_functions.env_sensitive(sensitive_list, jsonData, polVars, **kwargs)
                        polVars['bgp_peers']['password'] = polVars['bgp_password']
                else: validating.error_template_not_found('bgp_template', **kwargs)
            else: validating.error_template_not_found('bgp_template', **kwargs)

        # Attach the EIGRP Interface Profile if defined
        if not polVars['eigrp_interface_profile'] == None:
            if kwargs['easyDict']['tmp'].get('eigrp_interface_profiles'):
                ospfp = kwargs['easyDict']['tmp']['eigrp_interface_profiles']
                if kwargs['easyDict']['tmp']['eigrp_interface_profiles'].get(polVars['eigrp_interface_profile']):
                    polVars['eigrp_interface_profile'] = ospfp[polVars['eigrp_interface_profile']]
                else: validating.error_template_not_found('eigrp_interface_profile', **kwargs)
            else: validating.error_template_not_found('eigrp_interface_profile', **kwargs)

        # Attach the OSPF Interface Profile if defined
        if not polVars['ospf_interface_profile'] == None:
            if kwargs['easyDict']['tmp'].get('ospf_interface_profiles'):
                ospfp = kwargs['easyDict']['tmp']['ospf_interface_profiles']
                if kwargs['easyDict']['tmp']['ospf_interface_profiles'].get(polVars['ospf_interface_profile']):
                    polVars['ospf_interface_profile'] = ospfp[polVars['ospf_interface_profile']]
                else: validating.error_template_not_found('ospf_interface_profile', **kwargs)
            else: validating.error_template_not_found('ospf_interface_profile', **kwargs)

        pop_list = [
            'bgp_template', 'eigrp_interface_profile', 'interface_config_template', 'interface_template',
            'l3out', 'node_profile', 'ospf_interface_profile', 'template_name', 'tenant'
        ]
        for i in pop_list:
            if not polVars.get(i) == None: polVars.pop(i)

        # Add Dictionary to Policy
        kwargs['class_path'] = 'networking,l3outs,logical_node_profiles,logical_interface_profiles'
        kwargs['policy1'] = 'name'
        kwargs['policy_name1'] = kwargs['l3out']
        kwargs['policy2'] = 'name'
        kwargs['policy_name2'] = kwargs['node_profile']
        kwargs['easyDict'] = easy_functions.ez_tenants_append_sub_subtype(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - L3Out - Logical Node Interface Profile - Interface Configuration
    #=============================================================================
    def node_intf_cfg(self, **kwargs):
        # Get Variables from Library
        dict1 = kwargs['easy_jsonData']['tenants.l3out.logicalNodeInterfaceProfile.InterfaceConfiguration']
        jsonData = dict1['allOf'][1]['properties']

        args_list = []
        if re.search('^(l3-port|sub-interface)$', kwargs['interface_type']):
            args_list.append('auto_state')
            if kwargs['interface_type'] == 'l3-port':
                args_list.extend(['encap_scope', 'mode', 'vlan'])
            jsonData = easy_functions.args_remove(args_list, jsonData)
        
        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        if polVars['interface_type'] == 'ext-svi':
            polVars['svi_addresses'] = {}
            svi = polVars['svi_addresses']
            
            # Check for Link Local Addresses
            if not polVars['link_local_addresses'] == None:
                if not len(polVars['link_local_addresses'].split(',')) == 2:
                    validating.error_interface_address('link_local_addresses', **kwargs)
                svi['link_local_addresses'] = polVars['link_local_addresses'].split(',')
            
            # Check for Primary Addresses
            if not polVars['primary_preferred_addresses'] == None:
                if not len(polVars['primary_preferred_addresses'].split(',')) == 2:
                    validating.error_interface_address('primary_preferred_addresses', **kwargs)
                svi['primary_preferred_addresses'] = polVars['primary_preferred_addresses'].split(',')
            
            # Check for Secondary Addresses
            if not polVars['secondary_addresses'] == None:
                if not len(polVars['secondary_addresses'].split(',')) % 2  == 0:
                    validating.error_interface_address('secondary_addresses', **kwargs)
                svi['secondary_addresses'] = polVars['secondary_addresses'].split(',')
        else:
            if not polVars['link_local_addresses'] == None:
                if not len(polVars['link_local_addresses'].split(',')) == 1:
                    validating.error_interface_address('link_local_addresses', **kwargs)
                polVars['link_local_address'] = polVars['link_local_addresses']
            
            # Check for Primary Addresses
            if not polVars['primary_preferred_addresses'] == None:
                if not len(polVars['primary_preferred_addresses'].split(',')) == 1:
                    validating.error_interface_address('primary_preferred_addresses', **kwargs)
                polVars['primary_preferred_address'] = polVars['primary_preferred_addresses']
            
            # Check for Secondary Addresses
            if not polVars['secondary_addresses'] == None:
                polVars['secondary_addresses'] = polVars['secondary_addresses'].split(',')

        pop_list = ['link_local_addresses', 'primary_preferred_addresses', 'template_name']
        for i in pop_list:
            if not polVars.get(i) == None: polVars.pop(i)
        polVars = easy_functions.ez_remove_empty(polVars)

        # If Arguments were Removed, Add them back
        jsonData = easy_functions.args_add(args_list, jsonData)

       # Add Policy Variables to easyDict
        if not kwargs['easyDict']['tmp'].get('node_intf_config'):
            kwargs['easyDict']['tmp'].update(deepcopy({'node_intf_config':{}}))
        kwargs['easyDict']['tmp']['node_intf_config'].update(
            deepcopy({f"{kwargs['template_name']}":polVars})
        )
        return kwargs['easyDict']

    #=============================================================================
    # Function - L3Out - Logical Node Interface Profile - Templates
    #=============================================================================
    def node_intf_temp(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['tenants.l3out.logicalNodeInterfaceProfile.Templates']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        polVars = easy_functions.ez_remove_empty(polVars)

       # Add Policy Variables to easyDict
        if not kwargs['easyDict']['tmp'].get('node_intf_templates'):
            kwargs['easyDict']['tmp'].update(deepcopy({'node_intf_templates':{}}))
        kwargs['easyDict']['tmp']['node_intf_templates'].update(deepcopy({f"{kwargs['template_name']}":polVars}))
        return kwargs['easyDict']

    #=============================================================================
    # Function - L3Out - Logical Node Profile
    #=============================================================================
    def node_profile(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['tenants.l3out.logicalNodeProfile']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        polVars['node_router_ids'] = polVars['node_router_ids'].split(',')
        polVars['node_list'] = [int(s) for s in str(polVars['node_list']).split(',')]
        polVars['nodes'] = []
        for x in range(0, len(polVars['node_list'])):
            if polVars.get('use_router_id_as_loopback'):
                polVars['nodes'].append({
                    'node_id':polVars['node_list'][x],
                    'router_id':polVars['node_router_ids'][x],
                    'use_router_id_as_loopback':polVars.get('use_router_id_as_loopback')
                })
            else:
                polVars['nodes'].append({'node_id':polVars['node_list'][x],'router_id':polVars['node_router_ids'][x]})

        # Remove Arguments
        pop_list = ['l3out', 'node_list', 'node_router_ids', 'tenant', 'use_router_id_as_loopback']
        for i in pop_list:
            if not polVars.get(i) == None: polVars.pop(i)

        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'networking,l3outs,logical_node_profiles'
        kwargs['policy'] = 'name'
        kwargs['policy_name'] = kwargs['l3out']
        kwargs['easyDict'] = easy_functions.ez_tenants_append_subtype(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Application Profiles
    #=============================================================================
    def ospf_interface(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['tenants.policies.ospfInterface']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        polVars.pop('tenant')

        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'policies,protocol,ospf,ospf_interface'
        kwargs['easyDict'] = easy_functions.ez_tenants_append(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - OSPF Interface Profile Templates
    #=============================================================================
    def ospf_profile(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['tenants.l3out.ospfInterfaceProfile']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        polVars.pop('template_name')
        polVars = easy_functions.ez_remove_empty(polVars)

        # Add Policy Variables to easyDict
        if not kwargs['easyDict']['tmp'].get('ospf_interface_profiles'):
            kwargs['easyDict']['tmp'].update(deepcopy({'ospf_interface_profiles':{}}))
        kwargs['easyDict']['tmp']['ospf_interface_profiles'].update(deepcopy({f"{kwargs['template_name']}":polVars}))
        return kwargs['easyDict']

    #=============================================================================
    # Function - OSPF Routing Profile Templates
    #=============================================================================
    def ospf_routing(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['tenants.l3out.ospfRoutingProfile']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)

        # Modify the polVars OSPF Area Control
        polVars['ospf_area_control'] = {
            'originate_summary_lsa':polVars.get('originate_summary_lsa'),
            'send_redistribution_lsas_into_nssa_area':polVars.get('send_redistribution_lsas_into_nssa_area'),
            'suppress_forwarding_address':polVars.get('suppress_forwarding_address')
        }
        pop_list = [
            'originate_summary_lsa',
            'send_redistribution_lsas_into_nssa_area',
            'suppress_forwarding_address', 'template_name'
        ]
        for i in pop_list:
            if not polVars.get(i) == None: polVars.pop(i)
        polVars = easy_functions.ez_remove_empty(polVars)

        # Add Policy Variables to easyDict
        if not kwargs['easyDict']['tmp'].get('ospf_routing'):
            kwargs['easyDict']['tmp'].update(deepcopy({'ospf_routing':{}}))
        kwargs['easyDict']['tmp']['ospf_routing'].update(deepcopy({f"{kwargs['template_name']}":polVars}))
        return kwargs['easyDict']

    #=============================================================================
    # Function - Tenants
    #=============================================================================
    def tenant_add(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['tenants.Tenants']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        if not polVars['annotations'] == None:
            polVars['annotations'] = easy_functions.annotations_split(polVars['annotations'])
        polVars['monitoring_policy'] = 'default'
        if not polVars['ndo_settings'] == None:
            if kwargs['easyDict']['tmp'].get('ndo_settings'):
                if kwargs['easyDict']['tmp']['ndo_settings'].get(polVars['ndo_settings']):
                    for i in kwargs['easyDict']['tmp']['ndo_schemas']:
                        if i['name'] == polVars['ndo_settings']:
                            polVars['ndo'] = deepcopy(i)
                else: validating.error_schema('ndo_settings', **kwargs)
            else: validating.error_schema('ndo_settings', **kwargs)
            polVars['ndo']['users'] = polVars['users'].split(',')
            polVars.pop('ndo_settings')
            polVars.pop('users')
            ndo_sites = []
            for item in polVars['ndo']['templates']:
                for i in item['sites']:
                    if not i in ndo_sites:
                        ndo_sites.append(i)
            if len(ndo_sites) > 0:
                polVars['ndo']['sites'] = []
                for item in ndo_sites:
                    for i in kwargs['easyDict']['tmp']['ndo_sites']:
                        if i['name'] == item:
                            polVars['ndo']['sites'].append(deepcopy(i))
        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'tenants'
        kwargs['easyDict'] = easy_functions.ez_append(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - Tenants
    #=============================================================================
    def tenant_site(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['tenants.Sites']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        polVars = easy_functions.ez_remove_empty(polVars)

        # Add Policy Variables to easyDict
        if not kwargs['easyDict']['tmp'].get('ndo_sites'):
            kwargs['easyDict']['tmp'].update({'ndo_sites':[]})
        kwargs['easyDict']['tmp']['ndo_sites'].append(polVars)
        return kwargs['easyDict']

    #=============================================================================
    # Function - VRFs
    #=============================================================================
    def vrf_add(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['tenants.Vrfs']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        if not polVars['annotations'] == None:
            polVars['annotations'] = easy_functions.annotations_split(polVars['annotations'])
        polVars['epg_esg_collection_for_vrfs'] = {
            'contracts': [],
            'label_match_criteria': 'AtleastOne'
        }
        if not polVars['ndo_template'] == None:
            polVars['ndo'] = {'template': polVars['ndo_template']}
        pop_list = ['ndo_template', 'tenant']
        for i in pop_list:
            if not polVars.get(i) == None: polVars.pop(i)

        # Attach the VRF Template Additional Attributes
        if kwargs['easyDict']['tmp'].get('vrf_templates'):
            if kwargs['easyDict']['tmp']['vrf_templates'].get(polVars['vrf_template']):
                polVars.update(kwargs['easyDict']['tmp']['vrf_templates'][polVars['vrf_template']])
            else: validating.error_template_not_found('vrf_template', **kwargs)
        else: validating.error_template_not_found('vrf_template', **kwargs)
        polVars.pop('vrf_template')

        # Add Policy Variables to easyDict
        kwargs['class_path'] = 'networking,vrfs'
        kwargs['easyDict'] = easy_functions.ez_tenants_append(polVars, **kwargs)
        return kwargs['easyDict']
        
    #=============================================================================
    # Function - VRF - Communities
    #=============================================================================
    def vrf_community(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['tenants.vrf.Community']['allOf'][1]['properties']

        # Build Dictionary of Policy Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)

        # Check if the Sensitive Variables are in the Environment.  If not Add them.
        sensitive_list = [f'vrf_snmp_community_{kwargs["community_variable"]}']
        polVars, kwargs = easy_functions.env_sensitive(sensitive_list, jsonData, polVars, **kwargs)
        polVars.pop('tenant')
        polVars.pop('vrf')

        # Add Dictionary to Policy
        kwargs['class_path'] = 'networking,vrfs,communities'
        kwargs['policy'] = 'name'
        kwargs['policy_name'] = kwargs['vrf']
        kwargs['easyDict'] = easy_functions.ez_tenants_append_subtype(polVars, **kwargs)
        return kwargs['easyDict']

    #=============================================================================
    # Function - VRF - Templates
    #=============================================================================
    def vrf_template(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['tenants.vrf.Templates']['allOf'][1]['properties']

        # Build Dictionary of Template Variables
        polVars = easy_functions.process_kwargs(jsonData, **kwargs)
        polVars.pop('template_name')
        per_list = ['bgp_timers_per_address_family', 'eigrp_timers_per_address_family', 'ospf_timers_per_address_family']
        for i in per_list:
            if not polVars[i] == None:
                dict_list = []
                for v in polVars[i].split(','):
                    if '_' in v: dict_list.append({ 'address_family': v.split('_')[0], 'policy': v.split('_')[1] })
                polVars[i] = dict_list

        # Add Policy Variables to easyDict
        if not kwargs['easyDict']['tmp'].get('vrf_templates'):
            kwargs['easyDict']['tmp'].update(deepcopy({'vrf_templates':{}}))
        kwargs['easyDict']['tmp']['vrf_templates'].update(deepcopy({f"{kwargs['template_name']}":polVars}))
        return kwargs['easyDict']

#=====================================================================================
# Please Refer to the "Notes" in the relevant column headers in the input Spreadhseet
# for detailed information on the Arguments used by this Class.
#=====================================================================================
class terraform_cloud(object):
    def __init__(self):
        self.templateLoader = jinja2.FileSystemLoader(
            searchpath=(template_path + 'terraform/'))
        self.templateEnv = jinja2.Environment(loader=self.templateLoader)

    #=============================================================================
    # Function - Create Terraform Cloud Workspaces
    #=============================================================================
    def create_terraform_workspaces(self, **kwargs):
        easyDict = kwargs['easyDict']
        jsonData = kwargs['easy_jsonData']['site.Identifiers']['allOf'][1]['properties']
        tfcb_config = []
        valid = False
        while valid == False:
            polVars = {}
            polVars["Description"] = f'Terraform Cloud Workspaces.'
            polVars["varInput"] = f'Do you want to Proceed with creating Workspaces in Terraform Cloud?'
            polVars["varDefault"] = 'Y'
            polVars["varName"] = 'Terraform Cloud Workspaces'
            runTFCB = easy_functions.varBoolLoop(**polVars)
            valid = True
        if runTFCB == True:
            polVars["terraform_cloud_token"] = terraform_cloud().terraform_token()
            
            # Obtain Terraform Cloud Organization
            if os.environ.get('tfc_organization') is None:
                polVars["tfc_organization"] = terraform_cloud().tfc_organization(**polVars)
                os.environ['tfc_organization'] = polVars["tfc_organization"]
            else:  polVars["tfc_organization"] = os.environ.get('tfc_organization')
            tfcb_config.append({'tfc_organization':polVars["tfc_organization"]})
            
            # Obtain Terraform Cloud Agent_Pool
            if os.environ.get('agentPoolId') is None:
                polVars["agentPoolId"] = terraform_cloud().tfc_agent_pool(**polVars)
                os.environ['agentPoolId'] = polVars["tfc_organization"]
            else:  polVars["agentPoolId"] = os.environ.get('agentPoolId')
            tfcb_config.append({'agentPoolId':polVars["agentPoolId"]})
            
            # Obtain Version Control Provider
            if os.environ.get('tfc_vcs_provider') is None:
                tfc_vcs_provider,polVars["tfc_oath_token"] = terraform_cloud(
                ).tfc_vcs_providers(**polVars)
                polVars["tfc_vcs_provider"] = tfc_vcs_provider
                os.environ['tfc_vcs_provider'] = tfc_vcs_provider
                os.environ['tfc_oath_token'] = polVars["tfc_oath_token"]
            else:
                polVars["tfc_vcs_provider"] = os.environ.get('tfc_vcs_provider')
                polVars["tfc_oath_token"] = os.environ['tfc_oath_token']

            # Set Some of the default Variables that user is not Prompted for
            polVars["allowDestroyPlan"] = False
            polVars["executionMode"] = 'agent'
            polVars["queueAllRuns"] = False
            polVars["speculativeEnabled"] = True
            polVars["triggerPrefixes"] = []

            # Set the Terraform Version for the Workspaces
            polVars["terraformVersion"] = kwargs['easyDict']['latest_versions']['terraform_version']

            for k, v in easyDict['sites'].items():
                run_loc = v['site_settings']['run_location']
                ctc = v['site_settings']['configure_terraform_cloud']
                if run_loc == 'tfc' and ctc == True:
                    site = v['site_settings']['site_id']
                    site_name = v['site_settings']['site_name']
                    polVars['site_name'] = site_name
                    polVars['site_group'] = site

                    # Obtain Version Control Base Repo for Workspace
                    polVars["vcsBaseRepo"] = terraform_cloud().tfc_vcs_repository(**polVars)
                    os.environ['vcsBaseRepo'] = polVars["vcsBaseRepo"]
                    
                    # Loop through the Site Folders
                    polVars["autoApply"] = True
                    polVars["Description"] = f'Site: {site_name}'
                    polVars["globalRemoteState"] = False
                    polVars["site_name"] = site_name

                    # Create Terraform Cloud Workspace
                    tfcb_config,polVars = terraform_cloud().tfcWorkspace(tfcb_config, **polVars)

                    #==============================================
                    # Add Sensitive Variables to Workspace
                    #==============================================
                    var_list = []
                    if easyDict['sites'].get(site):
                        if easyDict['sites'][site].get('sensitive_vars'):
                            var_list.extend(easyDict['sites'][site]['sensitive_vars'])

                    if v['site_settings']['controller_type'] == 'apic' and v['site_settings']['auth_type'] == 'username':
                        var_list.append('apicPass')
                    elif v['site_settings']['controller_type'] == 'apic':
                        var_list.extend(['certName', 'privateKey'])
                    else: var_list.append('ndoPass')
                    var_list.sort()
                    for var in var_list:
                        if 'cert' in var or 'private' in var:
                            polVars["Multi_Line_Input"] = True
                        else: polVars["Multi_Line_Input"] = False
                        print(f'* Adding {var} to {polVars["workspaceName"]}')
                        kwargs['class_path'] = 'tfcVariables'
                        polVars["Description"] = ''
                        polVars["easyDict"] = easyDict
                        polVars['jsonData'] = jsonData
                        polVars["Variable"] = var
                        polVars["varId"] = var
                        polVars["varKey"] = var
                        easy_functions.sensitive_var_site_group(**polVars)
                        osvar = f'TF_VAR_{var}'
                        polVars["varValue"] = os.environ.get(osvar)
                        polVars["Sensitive"] = True
                        terraform_cloud().tfcVariables(**polVars)
        else:
            print(f'\n{"-"*91}\n')
            print(f'  Skipping Step to Create Terraform Cloud Workspaces.')
            print(f'\n{"-"*91}\n')
     
    #=============================================================================
    # Function - Terraform Cloud - API Token
    #=============================================================================
    def terraform_token(self):
        # -------------------------------------------------------------------------------------------------------------------------
        # Check to see if the TF_VAR_terraform_cloud_token is already set in the Environment, and if not prompt the user for Input
        #--------------------------------------------------------------------------------------------------------------------------
        if os.environ.get('TF_VAR_terraform_cloud_token') is None:
            print(f'\n----------------------------------------------------------------------------------------\n')
            print(f'  The Run or State Location was set to Terraform Cloud.  To Store the Data in Terraform')
            print(f'  Cloud we will need a User or Org Token to authenticate to Terraform Cloud.  If you ')
            print(f'  have not already obtained a token see instructions in how to obtain a token Here:\n')
            print(f'   - https://www.terraform.io/docs/cloud/users-teams-organizations/api-tokens.html')
            print(f'\n----------------------------------------------------------------------------------------\n')

            while True:
                user_response = input('press enter to continue: ')
                if user_response == '':
                    break

            # Request the TF_VAR_terraform_cloud_token Value from the User
            while True:
                try:
                    secure_value = stdiomask.getpass(prompt=f'Enter the value for the Terraform Cloud Token: ')
                    break
                except Exception as e:
                    print('Something went wrong. Error received: {}'.format(e))

            # Add the TF_VAR_terraform_cloud_token to the Environment
            os.environ['TF_VAR_terraform_cloud_token'] = '%s' % (secure_value)
            terraform_cloud_token = secure_value
        else:
            terraform_cloud_token = os.environ.get('TF_VAR_terraform_cloud_token')

        return terraform_cloud_token

    #=============================================================================
    # Function - Terraform Cloud - VCS Repository
    #=============================================================================
    def tfc_agent_pool(self, **polVars):
        #-------------------------------
        # Configure the Variables URL
        #-------------------------------
        tfc_org = polVars['tfc_organization']
        url = f"https://app.terraform.io/api/v2/organizations/{tfc_org}/agent-pools"
        tf_token = f"Bearer {polVars['terraform_cloud_token']}"
        tf_header = {'Authorization': tf_token, 'Content-Type': 'application/vnd.api+json'}

        #----------------------------------------------------------------------------------
        # Get the Contents of the Workspace to Search for the Variable
        #----------------------------------------------------------------------------------
        status,json_data = easy_functions.tfc_api(url, 'get', {}, tf_header, 'Get Agent Pools')

        #--------------------------------------------------------------
        # Parse the JSON Data to see if the Variable Exists or Not.
        #--------------------------------------------------------------

        if status == 200:
            json_data = json_data['data']
            pool_list = []
            pool_dict = {}
            for item in json_data:
                pool_list.append(item['attributes']['name'])
                pool_dict.update({item['attributes']['name']:item['id']})

            # print(vcsProvider)
            polVars["multi_select"] = False
            polVars["var_description"] = "Terraform Cloud Agent Pools:"
            polVars["jsonVars"] = sorted(pool_list)
            polVars["varType"] = 'Agent Pools'
            polVars["defaultVar"] = ''
            agentPool = easy_functions.variablesFromAPI(**polVars)

            agentPool = pool_dict[agentPool]
            return agentPool
        else:
            print(status)

    #=============================================================================
    # Function - Terraform Cloud - Organization
    #=============================================================================
    def tfc_organization(self, **polVars):
        #-------------------------------
        # Configure the Variables URL
        #-------------------------------
        url = 'https://app.terraform.io/api/v2/organizations'
        tf_token = f"Bearer {polVars['terraform_cloud_token']}"
        tf_header = {'Authorization': tf_token, 'Content-Type': 'application/vnd.api+json'}

        #----------------------------------------------------------------------------------
        # Get the Contents of the Workspace to Search for the Variable
        #----------------------------------------------------------------------------------
        status,json_data = easy_functions.tfc_api(url, 'get', {}, tf_header, 'Get Terraform Cloud Organizations')

        #--------------------------------------------------------------
        # Parse the JSON Data to see if the Variable Exists or Not.
        #--------------------------------------------------------------

        if status == 200:
            json_data = json_data['data']
            tfcOrgs = []
            for item in json_data:
                for k, v in item.items():
                    if k == 'id':
                        tfcOrgs.append(v)

            # print(tfcOrgs)
            polVars["multi_select"] = False
            polVars["var_description"] = 'Terraform Cloud Organizations:'
            polVars["jsonVars"] = tfcOrgs
            polVars["varType"] = 'Terraform Cloud Organization'
            polVars["defaultVar"] = ''
            tfc_organization = easy_functions.variablesFromAPI(**polVars)
            return tfc_organization
        else:
            print(status)

    #=============================================================================
    # Function - Terraform Cloud - VCS Repository
    #=============================================================================
    def tfc_vcs_repository(self, **polVars):
        #-------------------------------
        # Configure the Variables URL
        #-------------------------------
        oauth_token = polVars["tfc_oath_token"]
        url = f'https://app.terraform.io/api/v2/oauth-tokens/{oauth_token}/authorized-repos?oauth_token_id={oauth_token}'
        tf_token = 'Bearer %s' % (polVars['terraform_cloud_token'])
        tf_header = {'Authorization': tf_token,
                'Content-Type': 'application/vnd.api+json'
        }

        #----------------------------------------------------------------------------------
        # Get the Contents of the Workspace to Search for the Variable
        #----------------------------------------------------------------------------------
        status,json_data = easy_functions.tfc_api(url, 'get', {}, tf_header, 'Get VCS Repos')

        #--------------------------------------------------------------
        # Parse the JSON Data to see if the Variable Exists or Not.
        #--------------------------------------------------------------

        if status == 200:
            json_data = json_data['data']
            repo_list = []
            for item in json_data:
                for k, v in item.items():
                    if k == 'id':
                        repo_list.append(v)

            # Obtain the VCS Base Repository
            polVars["multi_select"] = False
            polVars["var_description"] = f'Site: "{polVars["site_name"]}" Terraform Cloud VCS Base Repository:'
            polVars["jsonVars"] = sorted(repo_list)
            polVars["varType"] = 'VCS Base Repository'
            polVars["defaultVar"] = ''
            vcsBaseRepo = easy_functions.variablesFromAPI(**polVars)

            return vcsBaseRepo
        else:
            print(status)

    #=============================================================================
    # Function - Terraform Cloud - VCS Providers
    #=============================================================================
    def tfc_vcs_providers(self, **polVars):
        #-------------------------------
        # Configure the Variables URL
        #-------------------------------
        tfc_org = polVars["tfc_organization"]
        url = f'https://app.terraform.io/api/v2/organizations/{tfc_org}/oauth-clients'
        tf_token = f"Bearer {polVars['terraform_cloud_token']}"
        tf_header = {'Authorization': tf_token, 'Content-Type': 'application/vnd.api+json'}

        #----------------------------------------------------------------------------------
        # Get the Contents of the Workspace to Search for the Variable
        #----------------------------------------------------------------------------------
        status,json_data = easy_functions.tfc_api(url, 'get', {}, tf_header, 'Get VCS Repos')

        #--------------------------------------------------------------
        # Parse the JSON Data to see if the Variable Exists or Not.
        #--------------------------------------------------------------

        if status == 200:
            json_data = json_data['data']
            vcsProvider = []
            vcsAttributes = []
            for item in json_data:
                for k, v in item.items():
                    if k == 'id':
                        vcs_id = v
                    elif k == 'attributes':
                        vcs_name = v['name']
                    elif k == 'relationships':
                        oauth_token = v["oauth-tokens"]["data"][0]["id"]
                vcsProvider.append(vcs_name)
                vcs_repo = {
                    'id':vcs_id,
                    'name':vcs_name,
                    'oauth_token':oauth_token
                }
                vcsAttributes.append(vcs_repo)

            # print(vcsProvider)
            polVars["multi_select"] = False
            polVars["var_description"] = "Terraform Cloud VCS Provider:"
            polVars["jsonVars"] = vcsProvider
            polVars["varType"] = 'VCS Provider'
            polVars["defaultVar"] = ''
            vcsRepoName = easy_functions.variablesFromAPI(**polVars)

            for i in vcsAttributes:
                if i["name"] == vcsRepoName:
                    tfc_oauth_token = i["oauth_token"]
                    vcsBaseRepo = i["id"]
            # print(f'vcsBaseRepo {vcsBaseRepo} and tfc_oauth_token {tfc_oauth_token}')
            return vcsBaseRepo,tfc_oauth_token
        else:
            print(status)

    #=============================================================================
    # Function - Terraform Cloud - GET Workspaces
    #=============================================================================
    def tfcWorkspace(self, tfcb_config, **polVars):
        get_workspace = False
        site = polVars['site_group']
        site_name = polVars['site_name']
        tfc_org = polVars['tfc_organization']
        while get_workspace == False:
            polVars["Description"] = f'   Name of the Workspace to Create in Terraform Cloud for:\n'\
                f'  - Site: "{site_name}"'
            polVars["varDefault"] = site_name
            polVars["varInput"] = f'Terraform Cloud Workspace Name. [{site_name}]: '
            polVars["varName"] = f'Workspace Name'
            polVars["maximum"] = 90
            polVars["minimum"] = 1
            polVars["pattern"] = '^[a-zA-Z0-9\\-\\_]+$'
            polVars["workspaceName"] = easy_functions.varStringLoop(**polVars)
            tfc_workspace = polVars["workspaceName"]
            valid = False
            while valid == False:
                polVars["Description"] = f'Terraform Cloud Workspace Working Directory.'
                polVars["varInput"] = f'Do you need to Configure a Terraform Working Directory for Workspace "{tfc_workspace}"?'
                polVars["varDefault"] = 'Y'
                polVars["varName"] = 'Terraform Cloud Workspace Working Directory'
                tfcWD = easy_functions.varBoolLoop(**polVars)
                valid = True
            if tfcWD == True:
                polVars["Description"] = f'   Workspace Working Directory for:\n'\
                    f'  - Site: "{site_name}"'
                polVars["varDefault"] = site_name
                polVars["varInput"] = f'Terraform Cloud Workspace "{tfc_workspace}" Working Directory. [{site_name}]: '
                polVars["varName"] = f'Workspace Working Directory'
                polVars["maximum"] = 90
                polVars["minimum"] = 1
                polVars["pattern"] = '^[/a-zA-Z0-9\\-\\_\\.]+$'
                polVars["workingDirectory"] = easy_functions.varStringLoop(**polVars)
            else: polVars["workingDirectory"] = ''

            #-------------------------------
            # Configure the Workspace URL
            #-------------------------------
            url = f'https://app.terraform.io/api/v2/organizations/{tfc_org}/workspaces/{tfc_workspace}'
            tf_token = 'Bearer %s' % (polVars['terraform_cloud_token'])
            tf_header = {'Authorization': tf_token, 'Content-Type': 'application/vnd.api+json'}

            #----------------------------------------------------------------------------------
            # Get the Contents of the Organization to Search for the Workspace
            #----------------------------------------------------------------------------------
            status,json_data = easy_functions.tfc_api(url, 'get', {}, tf_header, 'workspace_check')

            #--------------------------------------------------------------
            # Parse the JSON Data to see if the Workspace Exists or Not.
            #--------------------------------------------------------------
            key_count = 0
            if status == 200:
                creation_date = json_data['data']['attributes']['created-at']
                vcs_url = ''
                if json_data['data']['attributes'].get('vcs-repo'):
                    if json_data['data']['attributes']['vcs-repo'].get('repository-http-url'):
                        vcs_url = json_data['data']['attributes']['vcs-repo']['repository-http-url']
                        vcs_ = True
                    else: vcs_ = False
                else: vcs_ = False
                valid = False
                while valid == False:
                    polVars["Description"] = f'Terraform Cloud Workspace for Site "{site_name}".'
                    if vcs_ == True:
                        polVars["varInput"] = f'   The Workspace "{tfc_workspace}" already exists.\n'\
                          f'   Creation Date:  {creation_date}\n'\
                          f'   It is Associated to the Following Repository:\n'\
                          f'   {vcs_url}\n\n'\
                          f'Do you want to proceed using this workspace?'
                    else:
                        polVars["varInput"] = f'The Workspace "{tfc_workspace}" already exists.\n'\
                          f'   Creation Date:  {creation_date}\n\n'\
                          f'Do you want to proceed using this workspace?'
                    polVars["varDefault"] = 'Y'
                    polVars["varName"] = 'Existing Workspace'
                    Continue = easy_functions.varBoolLoop(**polVars)
                    valid = True
                if Continue == True:
                    workspace_id = json_data['data']['id']
                    key_count =+ 1
                    get_workspace = True
            elif status == 404: get_workspace = True

        #--------------------------------------------
        # If the Workspace was not found Create it.
        #--------------------------------------------

        if key_count == 0:
            #-------------------------------
            # Create Workspace
            #-------------------------------
            url = f"https://app.terraform.io/api/v2/organizations/{tfc_org}/workspaces/"
            tf_token = 'Bearer %s' % (polVars['terraform_cloud_token'])
            tf_header = {'Authorization': tf_token, 'Content-Type': 'application/vnd.api+json'}

            # Define the Template Source
            template_file = 'workspace.jinja2'
            template = self.templateEnv.get_template(template_file)

            # Create the Payload
            payload = template.render(polVars)
            if print_payload: print(payload)

            # Post the Contents to Terraform Cloud
            json_data = easy_functions.tfc_api(url, 'post', payload, tf_header, template_file)

            # Get the Workspace ID from the JSON Dump
            key_count =+ 1
        else:
            #-----------------------------------
            # Configure the PATCH Variables URL
            #-----------------------------------
            url = f'https://app.terraform.io/api/v2/workspaces/{workspace_id}/'
            tf_token = f"Bearer {polVars['terraform_cloud_token']}"
            tf_header = {'Authorization': tf_token, 'Content-Type': 'application/vnd.api+json'}

            # Define the Template Source
            template_file = 'workspace.jinja2'
            template = self.templateEnv.get_template(template_file)

            # Create the Payload
            payload = template.render(polVars)
            if print_payload: print(payload)

            # PATCH the Contents to Terraform Cloud
            json_data = easy_functions.tfc_api(url, 'patch', payload, tf_header, template_file)
            # Get the Workspace ID from the JSON Dump
            key_count =+ 1

        if not key_count > 0:
            print(f'\n-----------------------------------------------------------------------------\n')
            print(f'\n   Unable to Determine the Workspace ID for "{polVars["workspaceName"]}".')
            print(f'\n   Exiting...')
            print(f'\n-----------------------------------------------------------------------------\n')
            exit()
        
        # Get the Workspace ID from the JSON Dump
        tfcb_config.append({site:polVars["workspaceName"]})
        polVars['workspace_id'] = json_data['data']['id']
        return tfcb_config,polVars

    #=============================================================================
    # Function - Terraform Cloud - Workspace Remove
    #=============================================================================
    def tfcWorkspace_remove(self, **polVars):
        #-------------------------------
        # Configure the Workspace URL
        #-------------------------------
        tfc_org = polVars['tfc_organization']
        tfc_workspace = polVars['workspaceName']
        url = f'https://app.terraform.io/api/v2/organizations/{tfc_org}/workspaces/{tfc_workspace}'
        tf_token = f"Bearer {polVars['terraform_cloud_token']}"
        tf_header = {'Authorization': tf_token, 'Content-Type': 'application/vnd.api+json'}

        #----------------------------------------------------------------------------------
        # Delete the Workspace of the Organization to Search for the Workspace
        #----------------------------------------------------------------------------------
        response = delete(url, headers=tf_header)

        #--------------------------------------------------------------
        # Parse the JSON Data to see if the Workspace Exists or Not.
        #--------------------------------------------------------------
        del_count = 0
        if response.status_code == 200:
            print(f'\n-----------------------------------------------------------------------------\n')
            print(f'    Successfully Deleted Workspace "{polVars["workspaceName"]}".')
            print(f'\n-----------------------------------------------------------------------------\n')
            del_count =+ 1
        elif response.status_code == 204:
            print(f'\n-----------------------------------------------------------------------------\n')
            print(f'    Successfully Deleted Workspace "{polVars["workspaceName"]}".')
            print(f'\n-----------------------------------------------------------------------------\n')
            del_count =+ 1

        if not del_count > 0:
            print(f'\n-----------------------------------------------------------------------------\n')
            print(f'    Unable to Determine the Workspace ID for "{polVars["workspaceName"]}".')
            print(f'\n-----------------------------------------------------------------------------\n')

    #=============================================================================
    # Function - Terraform Cloud - Workspace Variables
    #=============================================================================
    def tfcVariables(self, **polVars):
        #-------------------------------
        # Configure the Variables URL
        #-------------------------------
        url = f"https://app.terraform.io/api/v2/workspaces/{polVars['workspace_id']}/vars"
        tf_token = f"Bearer {polVars['terraform_cloud_token']}"
        tf_header = {'Authorization': tf_token, 'Content-Type': 'application/vnd.api+json'}

        #----------------------------------------------------------------------------------
        # Get the Contents of the Workspace to Search for the Variable
        #----------------------------------------------------------------------------------
        status,json_data = easy_functions.tfc_api(url, 'get', {}, tf_header, 'variable_check')

        #--------------------------------------------------------------
        # Parse the JSON Data to see if the Variable Exists or Not.
        #--------------------------------------------------------------
        json_text = json.dumps(json_data)
        key_count = 0
        var_id = ''
        if 'id' in json_text:
            for keys in json_data['data']:
                if keys['attributes']['key'] == polVars['Variable']:
                    var_id = keys['id']
                    key_count =+ 1

        #--------------------------------------------
        # If the Variable was not found Create it.
        # If it is Found Update the Value
        #--------------------------------------------
        if not key_count > 0:
            # Define the Template Source
            template_file = 'variables.jinja2'
            template = self.templateEnv.get_template(template_file)

            # Create the Payload
            payload = template.render(polVars)
            if print_payload: print(payload)

            # Post the Contents to Terraform Cloud
            json_data = easy_functions.tfc_api(url, 'post', payload, tf_header, template_file)

            # Get the Workspace ID from the JSON Dump
            var_id = json_data['data']['id']
            key_count =+ 1

        else:
            #-----------------------------------
            # Configure the PATCH Variables URL
            #-----------------------------------
            url = 'https://app.terraform.io/api/v2/workspaces/%s/vars/%s' %  (polVars['workspace_id'], var_id)
            tf_token = 'Bearer %s' % (polVars['terraform_cloud_token'])
            tf_header = {'Authorization': tf_token, 'Content-Type': 'application/vnd.api+json'}

            # Define the Template Source
            template_file = 'variables.jinja2'
            template = self.templateEnv.get_template(template_file)

            # Create the Payload
            polVars.pop('varId')
            payload = template.render(polVars)
            if print_payload: print(payload)

            # PATCH the Contents to Terraform Cloud
            json_data = easy_functions.tfc_api(url, 'patch', payload, tf_header, template_file)
            # Get the Workspace ID from the JSON Dump
            var_id = json_data['data']['id']
            key_count =+ 1

        if not key_count > 0:
            print(f'\n-----------------------------------------------------------------------------\n')
            print(f"\n   Unable to Determine the Variable ID for {polVars['Variable']}.")
            print(f"\n   Exiting...")
            print(f'\n-----------------------------------------------------------------------------\n')
            exit()
        return var_id
