#!/usr/bin/env python3

import ast
import ipaddress
import jinja2
import json
import os
import pkg_resources
import re
import validating
from class_terraform import terraform_cloud
from easy_functions import countKeys, create_tf_file, findVars
from easy_functions import process_kwargs, process_workbook
from easy_functions import sensitive_var_site_group
from easy_functions import write_to_site, write_to_template
from easy_functions import update_easyDict
from openpyxl import load_workbook

aci_template_path = pkg_resources.resource_filename('class_tenants', 'templates/')

# Exception Classes
class InsufficientArgs(Exception):
    pass

class ErrException(Exception):
    pass

class InvalidArg(Exception):
    pass

class LoginFailed(Exception):
    pass

class tenants(object):
    def __init__(self, type):
        self.templateLoader = jinja2.FileSystemLoader(
            searchpath=(aci_template_path + '%s/') % (type))
        self.templateEnv = jinja2.Environment(loader=self.templateLoader)
        self.type = type

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def app_add(self, **kwargs):
        # Set Locally Used Variables
        wb = kwargs['wb']
        ws = kwargs['ws']
        row_num = kwargs['row_num']

        # Dicts for Application Profile; required and optional args
        required_args = {
            'site_group': '',
            'tenant': '',
            'name': '',
        }
        optional_args = {
            'alias': '',
            'annotations': '',
            'description': '',
            'qos_class': '',
        }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'site_group', templateVars['site_group'])
            validating.name_rule(row_num, ws, 'tenant', templateVars['tenant'])
            validating.name_rule(row_num, ws, 'name', templateVars['name'])
            if not templateVars['alias'] == None:
                validating.name_rule(row_num, ws, 'alias', templateVars['alias'])
            if not templateVars['annotations'] == None:
                for i in templateVars['annotations']:
                    for k, v in i.items():
                        validating.name_rule(row_num, ws, 'annotations', k)
                        validating.name_rule(row_num, ws, 'annotations', v)
            if not templateVars['description'] == None:
                validating.description(row_num, ws, 'description', templateVars['description'])
            if not templateVars['qos_class'] == None:
                validating.qos_priority(row_num, ws, 'qos_class', templateVars['qos_class'])
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(errorReturn)

        dataDict = {
            'alias':kwargs['alias'],
            'annotations': kwargs['annotations'],
            'description':kwargs['description'],
            'monitoring_policy':'default',
            'tenant':kwargs['tenant'],
            'qos_class':kwargs['qos_class']
        }

        # Add Dictionary to easyDict
        class_type = 'tenants'
        data_type = 'application_profiles'
        if not any(kwargs['site_group'] in d for d in kwargs['easyDict'][class_type][data_type]):
            kwargs['easyDict']['tenants'][data_type].append({kwargs['site_group']:[]})
            
        count = 0
        for i in kwargs['easyDict'][class_type][data_type]:
            for k, v in i.items():
                if kwargs['site_group'] == k:
                    kwargs['easyDict'][class_type][data_type][count][kwargs['site_group']].append(dataDict)
            count += 1

        # Return Dictionary
        return kwargs['easyDict']

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def bd_add(self, wb, ws, row_num, **kwargs):
        # Assign the kwargs to a initial var for each process
        initial_kwargs = kwargs

        # Open the Network Policies Worksheet
        ws_net = wb['Network Policies']
        rows = ws_net.max_row

        # Dicts for Bridge Domain required and optional args
        required_args = {'site_group': '',
                         'Tenant': '',
                         'Bridge_Domain': '',
                         'BD_Policy': '',
                         'Policy_Name': '',
                         # BD Policy Required Args
                         'bd_type': '',
                         'host_routing': '',
                         'ep_clear': '',
                         'unk_mac': '',
                         'unk_mcast': '',
                         'v6unk_mcast': '',
                         'multi_dst': '',
                         'mcast_allow': '',
                         'ipv6_mcast': '',
                         'arp_flood': '',
                         'limit_learn': '',
                         'fvEpRetPol': '',
                         'unicast_route': '',
                         'intersight_l2': '',
                         'intersight_bum': '',
                         'optimize_wan': '',
                         'monEPGPol': '',
                         'ip_dp_learning': ''}
        optional_args = {'alias': '',
                         'description': '',
                         'annotation': '',
                         'Custom_MAC': '',
                         'Link_Local_IPv6': '',
                         'VRF_Tenant': '',
                         'VRF': '',
                         'Subnet': '',
                         'Subnet_description': '',
                         'Subnet_Policy': '',
                         'L3Out_Tenant': '',
                         'L3Out': '',
                         # BD Policy Optional Args
                         'dhcpRelayP': '',
                         'igmpIfPol': '',
                         'igmpSnoopPol': '',
                         'mldSnoopPol': '',
                         'ep_move': '',
                         'rtctrlProfile': '',
                         'ndIfPol': '',
                         'fhsBDPol': '',
                         'netflowMonitorPol': ''}

        # Get the BD Policies from the Network Policies Tab
        func = 'bd'
        count = countKeys(ws_net, func)
        row_bd = ''
        var_dict = findVars(ws_net, func, rows, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('BD_Policy'):
                row_bd = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}
                break

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'site_group', templateVars['site_group'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
            validating.name_rule(row_num, ws, 'Bridge_Domain', templateVars['Bridge_Domain'])
            if not templateVars['alias'] == None:
                validating.name_rule(row_num, ws, 'alias', templateVars['alias'])
            if not templateVars['description'] == None:
                validating.description(row_num, ws, 'description', templateVars['description'])
            if not templateVars['annotation'] == None:
                if re.match(',', templateVars['annotation']):
                    for tag in templateVars['annotation'].split(','):
                        validating.name_rule(row_num, ws, 'annotation', tag)
                else:
                    validating.name_rule(row_num, ws, 'annotation', templateVars['annotation'])
            if not templateVars['Custom_MAC'] == None:
                validating.mac_address(row_num, ws, 'Custom_MAC', templateVars['Custom_MAC'])
            if not templateVars['Link_Local_IPv6'] == None:
                validating.ip_address(row_num, ws, 'Link_Local_IPv6', templateVars['Link_Local_IPv6'])
            if not templateVars['BD_Policy'] == templateVars['Policy_Name']:
                validating.error_policy_names(row_num, ws, templateVars['BD_Policy'], templateVars['Policy_Name'])
            if not templateVars['VRF'] == None:
                validating.name_rule(row_num, ws, 'VRF_Tenant', templateVars['VRF_Tenant'])
                validating.name_rule(row_num, ws, 'VRF', templateVars['VRF'])
            if not templateVars['Subnet'] == None:
                validating.ip_address(row_num, ws, 'Subnet', templateVars['Subnet'])
            if not templateVars['description'] == None:
                validating.description(row_num, ws, 'description', templateVars['description'])
            if not templateVars['L3Out'] == None:
                validating.name_rule(row_num, ws, 'L3Out_Tenant', templateVars['L3Out_Tenant'])
                validating.name_rule(row_num, ws, 'L3Out', templateVars['L3Out'])
            validating.values(row_bd, ws_net, 'bd_type', templateVars['bd_type'], ['fc', 'regular'])
            validating.values(row_bd, ws_net, 'ep_clear', templateVars['ep_clear'], ['no', 'yes'])
            validating.values(row_bd, ws_net, 'host_routing', templateVars['host_routing'], ['no', 'yes'])
            validating.values(row_bd, ws_net, 'mcast_allow', templateVars['mcast_allow'], ['no', 'yes'])
            validating.values(row_bd, ws_net, 'ipv6_mcast', templateVars['ipv6_mcast'], ['no', 'yes'])
            validating.values(row_bd, ws_net, 'arp_flood', templateVars['arp_flood'], ['no', 'yes'])
            validating.values(row_bd, ws_net, 'limit_learn', templateVars['limit_learn'], ['no', 'yes'])
            validating.values(row_bd, ws_net, 'unicast_route', templateVars['unicast_route'], ['no', 'yes'])
            validating.values(row_bd, ws_net, 'limit_learn', templateVars['limit_learn'], ['no', 'yes'])
            validating.values(row_bd, ws_net, 'intersight_l2', templateVars['intersight_l2'], ['no', 'yes'])
            validating.values(row_bd, ws_net, 'intersight_bum', templateVars['intersight_bum'], ['no', 'yes'])
            validating.values(row_bd, ws_net, 'optimize_wan', templateVars['optimize_wan'], ['no', 'yes'])
            validating.values(row_bd, ws_net, 'ip_dp_learning', templateVars['ip_dp_learning'], ['no', 'yes'])
            validating.values(row_bd, ws_net, 'unk_mac', templateVars['unk_mac'], ['flood', 'proxy'])
            validating.values(row_bd, ws_net, 'unk_mcast', templateVars['unk_mcast'], ['flood', 'opt-flood'])
            validating.values(row_bd, ws_net, 'v6unk_mcast', templateVars['v6unk_mcast'], ['flood', 'opt-flood'])
            validating.values(row_bd, ws_net, 'multi_dst', templateVars['multi_dst'], ['bd-flood', 'drop', 'encap-flood'])
            if not templateVars['ep_move'] == None:
                validating.values(row_bd, ws_net, 'ep_move', templateVars['ep_move'], ['garp'])
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(errorReturn)

        if templateVars['dhcpRelayP'] == 'default':
            templateVars['dhcpRelayP'] = 'uni/tn-common/relayp-default'
        if templateVars['fhsBDPol'] == 'default':
            templateVars['fhsBDPol'] = 'uni/tn-common/bdpol-default'
        if templateVars['fvEpRetPol'] == 'default':
            templateVars['fvEpRetPol'] = 'uni/tn-common/epRPol-default'
        if templateVars['igmpIfPol'] == 'default':
            templateVars['igmpIfPol'] = 'uni/tn-common/igmpIfPol-default'
        if templateVars['igmpSnoopPol'] == 'default':
            templateVars['igmpSnoopPol'] = 'uni/tn-common/snPol-default'
        if templateVars['mldSnoopPol'] == 'default':
            templateVars['mldSnoopPol'] = 'uni/tn-common/mldsnoopPol-default'
        if templateVars['monEPGPol'] == 'default':
            templateVars['monEPGPol'] = 'uni/tn-common/monepg-default'
        if templateVars['ndIfPol'] == 'default':
            templateVars['ndIfPol'] = 'uni/tn-common/ndifpol-default'
        if templateVars['netflowMonitorPol'] == 'default':
            templateVars['netflowMonitorPol'] = 'uni/tn-common/monitorpol-default'

        # Define the Template Source
        template_file = "bd.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Bridge_Domain_%s.tf' % (templateVars['Bridge_Domain'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

       # Reset kwargs back to initial kwargs
        kwargs = initial_kwargs

        # Initialize the Class
        lib_aci_ref = 'Tenant_Policies'
        class_init = '%s(ws)' % (lib_aci_ref)

        # Create the Subnet if it Exists
        if not kwargs.get('Subnet') == None:
            eval("%s.%s(wb, ws, row_num, **kwargs)" % (class_init, 'add_subnet'))

        if not templateVars['Tenant'] == templateVars['VRF_Tenant']:
            templateVars['bd_Tenant'] = templateVars['Tenant']

            # Process the template through the Sites
            templateVars['Tenant'] = templateVars['VRF_Tenant']
            template_file = "data_tenant.jinja2"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = 'data_Tenant_%s.tf' % (templateVars['VRF_Tenant'])
            dest_dir = 'Tenant_%s' % (templateVars['bd_Tenant'])
            write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

            # Define the Template Source
            template_file = "data_vrf.jinja2"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = 'data_Tenant_%s_VRF_%s.tf' % (templateVars['VRF_Tenant'], templateVars['VRF'])
            dest_dir = 'Tenant_%s' % (templateVars['bd_Tenant'])
            write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

            templateVars['Tenant'] = templateVars['bd_Tenant']

        if not templateVars['L3Out'] == None:
            if not templateVars['Tenant'] == templateVars['L3Out_Tenant']:
                templateVars['bd_Tenant'] = templateVars['Tenant']

                # Process the template through the Sites
                templateVars['Tenant'] = templateVars['L3Out_Tenant']
                template_file = "data_tenant.jinja2"
                template = self.templateEnv.get_template(template_file)

                # Process the template through the Sites
                dest_file = 'data_Tenant_%s.tf' % (templateVars['L3Out_Tenant'])
                dest_dir = 'Tenant_%s' % (templateVars['bd_Tenant'])
                write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

                # Process the template through the Sites
                template_file = "data_l3out.jinja2"
                template = self.templateEnv.get_template(template_file)

                # Process the template through the Sites
                dest_file = 'data_Tenant_%s_L3Out_%s.tf' % (templateVars['L3Out_Tenant'], templateVars['L3Out'])
                dest_dir = 'Tenant_%s' % (templateVars['bd_Tenant'])
                write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def bgp_peer(self, wb, ws, row_num, **kwargs):
        # Open the Network Policies Worksheet
        ws_net = wb['Network Policies']
        rows = ws_net.max_row

        # Dicts for required and optional args
        required_args = {'site_group': '',
                         'Path_Policy_Name': '',
                         'Peer_Interface': '',
                         'Peer_Address': '',
                         'Remote_ASN': '',
                         'eBGP_Multihop_TTL': '',
                         'Weight': '',
                         'Local_ASN_Config': '',
                         'Admin_State': '',
                         'BGP_Peer_Policy': '',
                         'Tenant': '',
                         'L3Out': '',
                         'Profile_Name': '',
                         'Node_Profile': '',
                         'Interface_Profile': '',
                         'Interface_Type': '',
                         'Pod_ID': '',
                         'Node1_ID': '',
                         'Interface_or_PG': '',
                         'allow_self_as': '',
                         'as_override': '',
                         'disable_peer_as_check': '',
                         'next_hop_self': '',
                         'send_community': '',
                         'send_ext_community': '',
                         'allowed_self_as_count': '',
                         'bfd': '',
                         'disable_connected_check': '',
                         'AF_Mcast': '',
                         'AF_Ucast': '',
                         'remove_all_private_as': '',
                         'remove_private_as': '',
                         'private_to_local': ''}
        optional_args = {'description': '',
                         'BGP_Password': '',
                         'BGP_Peer_Prefix_Policy': '',
                         'Local_ASN': '',
                         'Node2_ID': '',
                         'Policy_Name': '',
                         'Prefix_Tenant': '',
                         'PFX_description': '',
                         'Action': '',
                         'Maximum_Prefixes': '',
                         'Restart_Time': '',
                         'Threshold': ''}

        # Get the Node Policies from the Network Policies Tab
        rows = ws.max_row
        func = 'l3out_path'
        count = countKeys(ws, func)
        row_path = ''
        var_dict = findVars(ws, func, rows, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('Path_Policy_Name'):
                row_path = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}

        # Get the Node Policies from the Network Policies Tab
        func = 'bgp_profile'
        bgp_count = countKeys(ws_net, func)
        row_bgp = ''
        var_dict = findVars(ws_net, func, rows, bgp_count)
        for pos in var_dict:
            if var_dict[pos].get('Profile_Name') == kwargs.get('BGP_Peer_Policy'):
                row_bgp = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}

        if not kwargs.get('BGP_Peer_Prefix_Policy') == None:
            # Get the Node Policies from the Network Policies Tab
            func = 'pfx_policy'
            pfx_count = countKeys(ws_net, func)
            row_pfx = ''
            var_dict = findVars(ws_net, func, rows, pfx_count)
            for pos in var_dict:
                if var_dict[pos].get('Policy_Name') == kwargs.get('BGP_Peer_Prefix_Policy'):
                    row_pfx = var_dict[pos]['row']
                    del var_dict[pos]['row']
                    kwargs = {**kwargs, **var_dict[pos]}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'site_group', templateVars['site_group'])
            validating.values(row_num, ws, 'Peer_Interface', templateVars['Peer_Interface'], ['Interface', 'Loopback'])
            validating.ip_address(row_num, ws, 'Peer_Address', templateVars['Peer_Address'])
            validating.number_check(row_num, ws, 'Remote_ASN', templateVars['Remote_ASN'], 1, 4294967295)
            if not templateVars['description'] == None:
                validating.description(row_num, ws, 'description', templateVars['description'])
            validating.number_check(row_num, ws, 'eBGP_Multihop_TTL', templateVars['eBGP_Multihop_TTL'], 1, 255)
            validating.number_check(row_num, ws, 'Weight', templateVars['Weight'], 0, 65535)
            if not templateVars['BGP_Peer_Prefix_Policy'] == None:
                validating.name_rule(row_num, ws, 'BGP_Peer_Prefix_Policy', templateVars['BGP_Peer_Prefix_Policy'])
            validating.values(row_num, ws, 'Local_ASN_Config', templateVars['Local_ASN_Config'], ['dual-as', 'no-prepend', 'none', 'replace-as'])
            if not templateVars['Local_ASN'] == None:
                validating.number_check(row_num, ws, 'Local_ASN', templateVars['Local_ASN'], 1, 4294967295)
            validating.values(row_num, ws, 'Admin_State', templateVars['Admin_State'], ['disabled', 'enabled'])
            validating.name_rule(row_num, ws, 'BGP_Peer_Policy', templateVars['BGP_Peer_Policy'])

            validating.name_rule(row_path, ws, 'Tenant', templateVars['Tenant'])
            validating.name_rule(row_path, ws, 'L3Out', templateVars['L3Out'])
            validating.name_rule(row_path, ws, 'Node_Profile', templateVars['Node_Profile'])
            validating.name_rule(row_path, ws, 'Interface_Profile', templateVars['Interface_Profile'])
            validating.values(row_path, ws, 'Interface_Type', templateVars['Interface_Type'], ['ext-svi', 'l3-port', 'sub-interface'])
            validating.number_check(row_path, ws, 'Pod_ID', templateVars['Pod_ID'], 1, 15)
            validating.number_check(row_path, ws, 'Node1_ID', templateVars['Node1_ID'], 101, 4001)
            if not templateVars['Node2_ID'] == None:
                validating.number_check(row_path, ws, 'Node2_ID', templateVars['Node2_ID'], 101, 4001)

            validating.number_check(row_bgp, ws_net, 'allowed_self_as_count', templateVars['allowed_self_as_count'], 1, 10)
            validating.values(row_bgp, ws_net, 'allow_self_as', templateVars['allow_self_as'], ['no', 'yes'])
            validating.values(row_bgp, ws_net, 'as_override', templateVars['as_override'], ['no', 'yes'])
            validating.values(row_bgp, ws_net, 'disable_peer_as_check', templateVars['disable_peer_as_check'], ['no', 'yes'])
            validating.values(row_bgp, ws_net, 'next_hop_self', templateVars['next_hop_self'], ['no', 'yes'])
            validating.values(row_bgp, ws_net, 'send_community', templateVars['send_community'], ['no', 'yes'])
            validating.values(row_bgp, ws_net, 'send_ext_community', templateVars['send_ext_community'], ['no', 'yes'])
            validating.values(row_bgp, ws_net, 'bfd', templateVars['bfd'], ['no', 'yes'])
            validating.values(row_bgp, ws_net, 'AF_Mcast', templateVars['AF_Mcast'], ['no', 'yes'])
            validating.values(row_bgp, ws_net, 'AF_Ucast', templateVars['AF_Ucast'], ['no', 'yes'])
            validating.values(row_bgp, ws_net, 'remove_all_private_as', templateVars['remove_all_private_as'], ['no', 'yes'])
            validating.values(row_bgp, ws_net, 'remove_private_as', templateVars['remove_private_as'], ['no', 'yes'])
            validating.values(row_bgp, ws_net, 'private_to_local', templateVars['private_to_local'], ['no', 'yes'])

            if not templateVars['BGP_Peer_Prefix_Policy'] == None:
                validating.name_rule(row_pfx, ws_net, 'Prefix_Tenant', templateVars['Prefix_Tenant'])
                validating.number_check(row_pfx, ws_net, 'Maximum_Prefixes', templateVars['Maximum_Prefixes'], 1, 300000)
                validating.number_check(row_pfx, ws_net, 'Restart_Time', templateVars['Restart_Time'], 1, 65535)
                validating.number_check(row_pfx, ws_net, 'Threshold', templateVars['Threshold'], 1, 100)
                validating.values(row_pfx, ws_net, 'Action', templateVars['Action'], ['log', 'reject', 'restart', 'shut'])
                if not templateVars['PFX_description'] == None:
                    validating.description(row_pfx, ws_net, 'PFX_description', templateVars['PFX_description'])
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(errorReturn)

        if re.search(r'\.', templateVars['Peer_Address']):
            templateVars['Peer_Address_'] = templateVars['Peer_Address'].replace('.', '-')
        else:
            templateVars['Peer_Address_'] = templateVars['Peer_Address'].replace(':', '-')

        ctrl_count = 0
        Ctrl = ''
        if templateVars['allow_self_as'] == 'yes':
            Ctrl = 'allow-self-as'
            ctrl_count =+ 1
        if templateVars['as_override'] == 'yes' and ctrl_count > 0:
            Ctrl = Ctrl + ',' + 'as-override'
            ctrl_count =+ 1
        elif templateVars['as_override'] == 'yes':
            Ctrl = 'as-override'
            ctrl_count =+ 1
        if templateVars['disable_peer_as_check'] == 'yes' and ctrl_count > 0:
            Ctrl = Ctrl + ',' + 'dis-peer-as-check'
            ctrl_count =+ 1
        elif templateVars['disable_peer_as_check'] == 'yes':
            Ctrl = 'dis-peer-as-check'
            ctrl_count =+ 1
        if templateVars['next_hop_self'] == 'yes' and ctrl_count > 0:
            Ctrl = Ctrl + ',' + 'nh-self'
            ctrl_count =+ 1
        elif templateVars['next_hop_self'] == 'yes':
            Ctrl = 'nh-self'
            ctrl_count =+ 1
        if templateVars['send_community'] == 'yes' and ctrl_count > 0:
            Ctrl = Ctrl + ',' + 'send-com'
            ctrl_count =+ 1
        elif templateVars['send_community'] == 'yes':
            Ctrl = 'send-com'
            ctrl_count =+ 1
        if templateVars['send_ext_community'] == 'yes' and ctrl_count > 0:
            Ctrl = Ctrl + ',' + 'send-ext-com'
            ctrl_count =+ 1
        elif templateVars['send_ext_community'] == 'yes':
            Ctrl = 'send-ext-com'
            ctrl_count =+ 1
        if ctrl_count > 0:
            templateVars['Ctrl'] = '%s' % (Ctrl)
        else:
            templateVars['Ctrl'] = ''

        ctrl_count = 0
        Ctrl = ''
        if templateVars['bfd'] == 'yes':
            Ctrl = 'bfd'
            ctrl_count =+ 1
        if templateVars['disable_connected_check'] == 'yes' and ctrl_count > 0:
            Ctrl = Ctrl + ',' + 'dis-conn-check'
            ctrl_count =+ 1
        elif templateVars['disable_connected_check'] == 'yes':
            Ctrl = 'dis-conn-check'
            ctrl_count =+ 1
        if ctrl_count > 0:
            templateVars['Peer_Ctrl'] = '%s' % (Ctrl)
        else:
            templateVars['Peer_Ctrl'] = ''

        ctrl_count = 0
        Ctrl = ''
        if templateVars['AF_Mcast'] == 'yes':
            Ctrl = 'af-mcast'
            ctrl_count =+ 1
        if templateVars['AF_Ucast'] == 'yes' and ctrl_count > 0:
            Ctrl = Ctrl + ',' + 'af-ucast'
            ctrl_count =+ 1
        elif templateVars['AF_Ucast'] == 'yes':
            Ctrl = 'af-ucast'
            ctrl_count =+ 1
        if ctrl_count > 0:
            templateVars['Address_Fam_Ctrl'] = '%s' % (Ctrl)
        else:
            templateVars['Address_Fam_Ctrl'] = ''

        ctrl_count = 0
        Ctrl = ''
        if templateVars['remove_all_private_as'] == 'yes':
            Ctrl = 'remove-all'
            ctrl_count =+ 1
        if templateVars['remove_private_as'] == 'yes' and ctrl_count > 0:
            Ctrl = Ctrl + ',' + 'remove-exclusive'
            ctrl_count =+ 1
        elif templateVars['remove_private_as'] == 'yes':
            Ctrl = 'remove-exclusive'
            ctrl_count =+ 1
        if templateVars['private_to_local'] == 'yes' and ctrl_count > 0:
            Ctrl = Ctrl + ',' + 'replace-as'
            ctrl_count =+ 1
        elif templateVars['private_to_local'] == 'yes':
            Ctrl = 'replace-as'
            ctrl_count =+ 1
        if ctrl_count > 0:
            templateVars['Private_AS_Ctrl'] = '%s' % (Ctrl)
        else:
            templateVars['Private_AS_Ctrl'] = ''

        if not templateVars['BGP_Password'] == None:
            x = templateVars['BGP_Password'].split('r')
            key_number = x[1]
            templateVars['sensitive_var'] = 'BGP_Password%s' % (key_number)

            # Define the Template Source
            template_file = "variables.jinja2"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = 'variable_%s.tf' % (templateVars['sensitive_var'])
            dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
            write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

            sensitive_var_site_group(wb, ws, row_num, dest_dir, dest_file, template, **templateVars)

        # Create Global Variables for First Template
        if not templateVars['Node2_ID'] == None:
            templateVars['PATH'] = 'protpaths-%s-%s' % (templateVars['Node1_ID'], templateVars['Node2_ID'])
        else:
            templateVars['PATH'] = 'paths-%s' % (templateVars['Node1_ID'])

        # Define the Template Source
        if templateVars['PATH'] == 'Loopback':
            template_file = "bgp_peer_connectivity_profile.jinja2"
        else:
            if templateVars['Local_ASN_Config'] == 'none':
                templateVars['Local_ASN_Config'] == None
            template_file = "bgp_peer_interface.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'L3Out_%s_Node_Profile_%s.tf' % (templateVars['L3Out'], templateVars['Node_Profile'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        write_to_site(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

        if not templateVars['BGP_Peer_Prefix_Policy'] == None:

            # Define the Template Source
            if templateVars['Restart_Time'] == 65535:
                templateVars['Restart_Time'] = 'infinite'
            template_file = "bgp_peer_prefix.jinja2"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = 'BGP_Peer_Prefix_%s.tf' % (templateVars['Policy_Name'])
            dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
            print(f'dest file {dest_file} and dest_dir {dest_dir} and Tenant = {templateVars["Tenant"]}')
            write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

            if not templateVars['Tenant'] == templateVars['Prefix_Tenant']:
                # Define the Template Source
                template_file = "data_tenant.jinja2"
                template = self.templateEnv.get_template(template_file)

                temp_Tenant = templateVars['Tenant']
                templateVars['Tenant'] = templateVars['Prefix_Tenant']
                # Process the template through the Sites
                dest_file = 'data_Tenant_%s.tf' % (templateVars['Tenant'])
                dest_dir = 'Tenant_%s' % (temp_Tenant)
                write_to_site(wb, ws_net, row_pfx, 'w', dest_dir, dest_file, template, **templateVars)

                templateVars['Tenant'] = temp_Tenant

                # Define the Template Source
                template_file = "data_bgp_peer_prefix.jinja2"
                template = self.templateEnv.get_template(template_file)

                # Process the template through the Sites
                dest_file = 'data_BGP_Peer_Prefix_%s.tf' % (templateVars['Policy_Name'])
                dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
                write_to_site(wb, ws_net, row_pfx, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def ctx_comm(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'site_group': '',
                         'Tenant': '',
                         'VRF': '',
                         'Ctx_Community': ''}
        optional_args = {'description': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'site_group', templateVars['site_group'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
            validating.name_rule(row_num, ws, 'VRF', templateVars['VRF'])
            validating.snmp_string(row_num, ws, 'Ctx_Community', templateVars['Ctx_Community'])
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(errorReturn)

        # Define the Template Source
        template_file = "snmp_ctx_community.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'VRF_%s.tf' % (templateVars['VRF'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        write_to_site(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def contract_add(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'site_group': '',
                         'Tenant': '',
                         'Contract_Type': '',
                         'Contract': '',
                         'Scope': '',
                         'QoS_Class': '',
                         'Target_DSCP': ''}
        optional_args = {'description': '',
                         'alias': '',
                         'annotation': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'site_group', templateVars['site_group'])
            validating.dscp(row_num, ws, 'Target_DSCP', templateVars['Target_DSCP'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
            validating.name_rule(row_num, ws, 'Contract', templateVars['Contract'])
            validating.qos_priority(row_num, ws, 'QoS_Class', templateVars['QoS_Class'])
            validating.values(row_num, ws, 'Contract_Type', templateVars['Contract_Type'], ['OOB', 'Standard', 'Taboo'])
            validating.values(row_num, ws, 'Scope', templateVars['Scope'], ['application-profile', 'context', 'global', 'tenant'])
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(errorReturn)

        # Define the Template Source
        if templateVars['Contract_Type'] == 'OOB':
            template_file = "contract_oob.jinja2"
        elif templateVars['Contract_Type'] == 'Standard':
            template_file = "contract.jinja2"
        elif templateVars['Contract_Type'] == 'Taboo':
            template_file = "contract_taboo.jinja2"
        dest_file = 'Contract_Type_%s_%s.tf' % (templateVars['Contract_Type'], templateVars['Contract'])
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def contract_to_epg(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'site_group': '',
                         'Contract_Tenant': '',
                         'Contract_Type': '',
                         'Contract': '',
                         'Tenant': '',
                         'App_Profile': '',
                         'EPG': ''}
        optional_args = { }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'site_group', templateVars['site_group'])
            validating.name_rule(row_num, ws, 'Contract_Tenant', templateVars['Contract_Tenant'])
            validating.name_rule(row_num, ws, 'Contract', templateVars['Contract'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
            validating.name_rule(row_num, ws, 'App_Profile', templateVars['App_Profile'])
            validating.name_rule(row_num, ws, 'EPG', templateVars['EPG'])
            validating.values(row_num, ws, 'Contract_Type', templateVars['Contract_Type'], ['consumer', 'provider'])
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(errorReturn)

        # Define the Template Source
        template_file = "epg_to_contract.jinja2"
        dest_file = 'App_Profile_%s_EPG_%s.tf' % (templateVars['App_Profile'], templateVars['EPG'])
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        write_to_site(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def epg_add(self, **kwargs):
        # Set Locally Used Variables
        wb = kwargs['wb']
        ws = kwargs['ws']
        row_num = kwargs['row_num']

        # Open the Network Policies Worksheet
        ws_net = wb['Network Policies']
        rows = ws_net.max_row

        # Get the EPG Policies from the Network Policies Tab
        func = 'epg'
        count = countKeys(ws_net, func)
        row_epg = ''
        var_dict = findVars(ws_net, func, rows, count)
        for pos in var_dict:
            if var_dict[pos].get('policy_name') == kwargs.get('epg_policy'):
                row_epg = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}

        # Dicts for EPG required and optional args
        required_args = {
            'site_group': '',
            'tenant': '',
            'application_profile': '',
            'name': '',
            'bridge_domain': '',
            'contract_exception_tag': '',
            'useg_epg': '',
            'qos_class': '',
            'intra_epg_isolation': '',
            'preferred_group_member': '',
            'flood_in_encapsulation': '',
            'label_match_criteria': '',
            'epg_admin_state': '',
            'has_multicast_source': ''
        }
        optional_args = {
            'alias': '',
            'description': '',
            'annotations': '',
            'global_alias': '',
            'fibre_channel_domain_association': '',
            'physical_domains': '',
            'vmm_domains': '',
            'VLAN': '',
            'PVLAN': '',
            'epg_to_aep': '',
            'epg_contract_master': '',
            'contract_exception_tag': '',
            'custom_qos': '',
            'data_plane_policer': '',
            'fhs_trust_control_policy': '',
            'vzGraphCont': '',
        }


        if kwargs['custom_qos'] == 'default':
            kwargs['custom_qos'] = 'uni/tn-common/qoscustom-default'
        if kwargs['data_plane_policer'] == 'default':
            kwargs['data_plane_policer'] = 'uni/tn-common/qosdpppol-default'
        if kwargs['fhs_trust_control_policy'] == 'default':
            kwargs['fhs_trust_control_policy'] = 'uni/tn-common/trustctrlpol-default'

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'site_group', templateVars['site_group'])
            validating.name_rule(row_num, ws, 'application_profile', templateVars['application_profile'])
            validating.name_rule(row_num, ws, 'bridge_domain', templateVars['bridge_domain'])
            validating.name_rule(row_num, ws, 'name', templateVars['name'])
            validating.name_rule(row_num, ws, 'tenant', templateVars['tenant'])
            validating.qos_priority(row_epg, ws_net, 'qos_class', templateVars['qos_class'])
            validating.values(row_epg, ws_net, 'epg_admin_state', templateVars['epg_admin_state'], ['admin_up', 'admin_shut'])
            validating.values(row_epg, ws_net, 'flood_in_encapsulation', templateVars['flood_in_encapsulation'], ['disabled', 'enabled'])
            validating.values(row_epg, ws_net, 'intra_epg_isolation', templateVars['intra_epg_isolation'], ['enforced', 'unenforced'])
            validating.values(row_epg, ws_net, 'label_match_criteria', templateVars['label_match_criteria'], ['All', 'AtleastOne', 'AtmostOne', 'None'])
            validating.values(row_epg, ws_net, 'preferred_group_member', templateVars['preferred_group_member'], ['exclude', 'include'])
            validating.values(row_epg, ws_net, 'useg_epg', templateVars['useg_epg'], ['true', 'false'])
            if not templateVars['alias'] == None:
                validating.name_rule(row_num, ws, 'alias', templateVars['alias'])
            if not templateVars['description'] == None:
                validating.description(row_num, ws, 'description', templateVars['description'])
            if not templateVars['annotations'] == None:
                for i in templateVars['annotations']:
                    for k, v in i.items():
                        validating.name_rule(row_num, ws, 'annotations', k)
                        validating.name_rule(row_num, ws, 'annotations', v)
            if not templateVars['physical_domains'] == None:
                if re.match(',', templateVars['physical_domains']):
                    for phys in templateVars['physical_domains'].split(','):
                        validating.name_rule(row_num, ws, 'physical_domains', phys)
                else:
                    validating.name_rule(row_num, ws, 'physical_domains', templateVars['physical_domains'])
            if not templateVars['vmm_domains'] == None:
                if re.match(',', templateVars['vmm_domains']):
                    for phys in templateVars['vmm_domains'].split(','):
                        validating.name_rule(row_num, ws, 'vmm_domains', phys)
                else:
                    validating.name_rule(row_num, ws, 'vmm_domains', templateVars['vmm_domains'])
            if not templateVars['VLAN'] == None:
                validating.vlans(row_num, ws, 'VLAN', templateVars['VLAN'])
            if not templateVars['PVLAN'] == None:
                validating.vlans(row_num, ws, 'PVLAN', templateVars['PVLAN'])
            if not templateVars['epg_to_aep'] == None:
                validating.name_rule(row_num, ws, 'epg_to_aep', templateVars['epg_to_aep'])
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(errorReturn)

        if not templateVars['VLAN'] == None:
            # Define the Template Source
            template_file = "static_path.jinja2"
            template = self.templateEnv.get_template(template_file)

            dest_file = 'App_Profile_%s_EPG_%s.tf' % (templateVars['App_Profile'], templateVars['EPG'])
            dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
            process_workbook(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

        if not templateVars['epg_to_aep'] == None:
            if re.search(',', templateVars['epg_to_aep']):
                # Define the Template Source
                aep_list = templateVars['epg_to_aep'].split(',')
                for aep in aep_list:
                    templateVars['AAEP'] = aep

                    # Define the Template Source
                    template_file = "policies_global_aep_generic.jinja2"
                    template = self.templateEnv.get_template(template_file)

                    # Process the template through the Sites
                    dest_file = 'Policies_Global_AEP_%s_generic.tf' % (templateVars['AAEP'])
                    dest_dir = 'Access'
                    write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

                    # Define the Template Source
                    template_file = "data_access_generic.jinja2"
                    template = self.templateEnv.get_template(template_file)

                    # Process the template through the Sites
                    dest_file = 'data_AEP_%s.tf' % (templateVars['AAEP'])
                    dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
                    write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

                    # Define the Template Source
                    template_file = "epgs_using_function.jinja2"
                    template = self.templateEnv.get_template(template_file)

                    # Process the template through the Sites
                    dest_file = 'App_Profile_%s_EPG_%s.tf' % (templateVars['App_Profile'], templateVars['EPG'])
                    dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
                    write_to_site(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)
            else:
                templateVars['AAEP'] = templateVars['EPG_to_AAEP']
                # Define the Template Source
                template_file = "policies_global_aep_generic.jinja2"
                template = self.templateEnv.get_template(template_file)

                # Process the template through the Sites
                dest_file = 'Policies_Global_AEP_%s_generic.tf' % (templateVars['AAEP'])
                dest_dir = 'Access'
                write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

                # Define the Template Source
                template_file = "data_access_generic.jinja2"
                template = self.templateEnv.get_template(template_file)

                # Process the template through the Sites
                dest_file = 'data_AEP_%s.tf' % (templateVars['AAEP'])
                dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
                write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

                # Define the Template Source
                template_file = "epgs_using_function.jinja2"
                template = self.templateEnv.get_template(template_file)

                # Process the template through the Sites
                dest_file = 'App_Profile_%s_EPG_%s.tf' % (templateVars['App_Profile'], templateVars['EPG'])
                dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
                write_to_site(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

        # dest_file = 'epg_%s_%s_static_bindings.tf' % (templateVars['App_Profile'], templateVars['EPG'])
        # dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        # create_file(wb, ws, row_num, 'w', dest_dir, dest_file, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def ext_epg(self, wb, ws, row_num, **kwargs):
        # Open the Network Policies Worksheet
        ws_net = wb['Network Policies']
        rows = ws_net.max_row

        # Dicts for required and optional args
        required_args = {'site_group': '',
                         'Tenant': '',
                         'L3Out': '',
                         'Ext_EPG': '',
                         'Ext_EPG_Policy': '',
                         'Subnets': '',
                         'Ext_Subnet_Policy': '',
                         'prio': '',
                         'target_dscp': '',
                         'pref_gr_memb': '',
                         'match_t': '',
                         'flood': '',
                         'export-rtctrl': '',
                         'import-rtctrl': '',
                         'import-security': '',
                         'shared-security': '',
                         'shared-rtctrl': '',
                         'agg-export': '',
                         'agg-import': '',
                         'agg-shared': ''}
        optional_args = {'alias': '',
                         'description': '',
                         'annotation': '',
                         'cons_vzBrCP': '',
                         'vzCPIf': '',
                         'Master_fvEPg': '',
                         'prov_vzBrCP': '',
                         'vzTaboo': '',
                         'exception_tag': '',
                         'rtctrlProfile': '',
                         'sub_rtctrlProfile': '',
                         'rtsumARtSummPol': ''}


        # Get the L3Out Policies from the Network Policies Tab
        func = 'ext_epg'
        count = countKeys(ws_net, func)
        row_epg = ''
        var_dict = findVars(ws_net, func, rows, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('Ext_EPG_Policy'):
                row_epg = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}

        func = 'ext_subnet'
        count = countKeys(ws_net, func)
        row_sub = ''
        var_dict = findVars(ws_net, func, rows, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('Ext_Subnet_Policy'):
                row_sub = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'site_group', templateVars['site_group'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
            if not templateVars['Subnets'] == None:
                if re.search(',', templateVars['Subnets']):
                    sx = templateVars['Subnets'].split(',')
                    for x in sx:
                        validating.ip_address(row_num, ws, 'Subnets', x)
                else:
                    validating.ip_address(row_num, ws, 'Subnets', templateVars['Subnets'])
            validating.dscp(row_epg, ws_net, 'target_dscp', templateVars['target_dscp'])
            validating.match_t(row_epg, ws_net, 'match_t', templateVars['match_t'])
            validating.qos_priority(row_epg, ws_net, 'prio', templateVars['prio'])
            validating.values(row_epg, ws_net, 'flood', templateVars['flood'], ['disabled', 'enabled'])
            validating.values(row_epg, ws_net, 'pref_gr_memb', templateVars['pref_gr_memb'], ['exclude', 'include'])
            validating.values(row_sub, ws_net, 'agg-export', templateVars['agg-export'], ['no', 'yes'])
            validating.values(row_sub, ws_net, 'agg-import', templateVars['agg-import'], ['no', 'yes'])
            validating.values(row_sub, ws_net, 'agg-shared', templateVars['agg-shared'], ['no', 'yes'])
            validating.values(row_sub, ws_net, 'export-rtctrl', templateVars['export-rtctrl'], ['no', 'yes'])
            validating.values(row_sub, ws_net, 'import-rtctrl', templateVars['import-rtctrl'], ['no', 'yes'])
            validating.values(row_sub, ws_net, 'import-security', templateVars['import-security'], ['no', 'yes'])
            validating.values(row_sub, ws_net, 'shared-security', templateVars['shared-security'], ['no', 'yes'])
            validating.values(row_sub, ws_net, 'shared-rtctrl', templateVars['shared-rtctrl'], ['no', 'yes'])
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(errorReturn)

        # Create aggregate templateVars
        aggregate = ''
        aggregate_count = 0
        if templateVars['agg-export'] == 'yes':
            aggregate = '"agg-export"'
            aggregate_count =+ 1
        if templateVars['agg-import'] == 'yes':
            if aggregate_count == 0:
                aggregate = '"agg-import"'
                aggregate_count =+ 1
            else:
                aggregate = aggregate + ', ' + '"agg-import"'
                aggregate_count =+ 1
        if templateVars['agg-shared'] == 'yes':
            if aggregate_count == 0:
                aggregate = '"agg-import"'
                aggregate_count =+ 1
            else:
                aggregate = aggregate + ', ' + '"agg-shared"'
                aggregate_count =+ 1

        if aggregate_count == 0:
            templateVars['aggregate'] = None
        else:
            templateVars['aggregate'] = '[%s]' % (aggregate)

        # Create scope templateVars
        scope = ''
        scope_count = 0
        if templateVars['export-rtctrl'] == 'yes':
            scope = '"export-rtctrl"'
            scope_count =+ 1
        if templateVars['import-rtctrl'] == 'yes':
            if scope_count == 0:
                scope = '"import-rtctrl"'
                scope_count =+ 1
            else:
                scope = scope + ', ' + '"import-rtctrl"'
                scope_count =+ 1
        if templateVars['import-security'] == 'yes':
            if scope_count == 0:
                scope = '"import-security"'
                scope_count =+ 1
            else:
                scope = scope + ', ' + '"import-security"'
                scope_count =+ 1
        if templateVars['shared-security'] == 'yes':
            if scope_count == 0:
                scope = '"shared-security"'
                scope_count =+ 1
            else:
                scope = scope + ', ' + '"shared-security"'
                scope_count =+ 1
        if templateVars['shared-rtctrl'] == 'yes':
            if scope_count == 0:
                scope = '"shared-rtctrl"'
                scope_count =+ 1
            else:
                scope = scope + ', ' + '"shared-rtctrl"'
                scope_count =+ 1

        if scope_count == 0:
            templateVars['scope'] = None
        else:
            templateVars['scope'] = '[%s]' % (scope)

        # Define the Template Source
        template_file = "epg_ext.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'L3Out_%s_EXTERNAL_EPG_%s.tf' % (templateVars['L3Out'], templateVars['Ext_EPG'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        if re.search(',', templateVars['Subnets']):
            sx = templateVars['Subnets'].split(',')
            for x in sx:
                templateVars['Subnet'] = x
                if re.search(':', x):
                    templateVars['Subnet_'] = x.replace(':', '-')
                    templateVars['Subnet_'] = templateVars['Subnet_'].replace('/', '_')
                else:
                    templateVars['Subnet_'] = x.replace('.', '-')
                    templateVars['Subnet_'] = templateVars['Subnet_'].replace('/', '_')

                # Define the Template Source
                template_file = "ext_subnet.jinja2"
                template = self.templateEnv.get_template(template_file)

                # Process the template through the Sites
                dest_file = 'L3Out_%s_EXTERNAL_EPG_%s.tf' % (templateVars['L3Out'], templateVars['Ext_EPG'])
                dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
                write_to_site(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)
        else:
            templateVars['Subnet'] = templateVars['Subnets']
            if re.search(':', templateVars['Subnet']):
                templateVars['Subnet_'] = templateVars['Subnet'].replace(':', '-')
                templateVars['Subnet_'] = templateVars['Subnet_'].replace('/', '_')
            else:
                templateVars['Subnet_'] = templateVars['Subnet'].replace('.', '-')
                templateVars['Subnet_'] = templateVars['Subnet_'].replace('/', '_')

            # Define the Template Source
            template_file = "ext_subnet.jinja2"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = 'L3Out_%s_EXTERNAL_EPG_%s.tf' % (templateVars['L3Out'], templateVars['Ext_EPG'])
            dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
            write_to_site(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def extepg_oob(self, wb, ws, row_num, **kwargs):
        # Open the Network Policies Worksheet
        ws_net = wb['Network Policies']
        rows = ws_net.max_row

        # Dicts for required and optional args
        required_args = {'site_group': '',
                         'Ext_EPG': '',
                         'QoS_Class': '',
                         'Subnets': ''}
        optional_args = {'annotation': '',
                         'consumed_Contracts': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'site_group', templateVars['site_group'])
            validating.name_rule(row_num, ws, 'Ext_EPG', templateVars['Ext_EPG'])
            validating.qos_priority(row_num, ws, 'QoS_Class', templateVars['QoS_Class'])
            if not templateVars['annotation'] == None:
                if re.search(',', templateVars['annotation']):
                    for x in templateVars['annotation'].split(','):
                        validating.description(row_num, ws, 'annotation', x)
                else:
                    validating.description(row_num, ws, 'annotation', templateVars['annotation'])
            if not templateVars['consumed_Contracts'] == None:
                if re.search(',', templateVars['consumed_Contracts']):
                    templateVars['provide_count'] =+ 1
                    for x in templateVars['consumed_Contracts'].split(','):
                        validating.name_rule(row_num, ws, 'consumed_Contracts', x)
                else:
                    validating.name_rule(row_num, ws, 'consumed_Contracts', templateVars['consumed_Contracts'])
            if not templateVars['Subnet'] == None:
                if re.search(',', templateVars['Subnets']):
                    sx = templateVars['Subnets'].split(',')
                    for x in sx:
                        validating.ip_address(row_num, ws, 'Subnets', x)
                else:
                    validating.ip_address(row_num, ws, 'Subnets', templateVars['Subnets'])
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(errorReturn)

        # Define the Template Source
        template_file = "epg_ext_oob.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'EPG_Mgmt_OOB_External_EPG_%s.tf' % (templateVars['Ext_EPG'])
        dest_dir = 'Tenant_mgmt'
        write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        if re.search(',', templateVars['Subnets']):
            sx = templateVars['Subnets'].split(',')
            for x in sx:
                templateVars['Subnet'] = x
                if re.search(':', x):
                    templateVars['Subnet_'] = x.replace(':', '-')
                    templateVars['Subnet_'] = templateVars['Subnet_'].replace('/', '_')
                else:
                    templateVars['Subnet_'] = x.replace('.', '-')
                    templateVars['Subnet_'] = templateVars['Subnet_'].replace('/', '_')

                # Define the Template Source
                template_file = "ext_subnet_oob.jinja2"
                template = self.templateEnv.get_template(template_file)

                # Process the template through the Sites
                dest_file = 'EPG_Mgmt_OOB_External_EPG_%s.tf' % (templateVars['Ext_EPG'])
                dest_dir = 'Tenant_mgmt'
                write_to_site(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)
        else:
            templateVars['Subnet'] = templateVars['Subnets']
            if re.search(':', templateVars['Subnet']):
                templateVars['Subnet_'] = templateVars['Subnet'].replace(':', '-')
                templateVars['Subnet_'] = templateVars['Subnet_'].replace('/', '_')
            else:
                templateVars['Subnet_'] = templateVars['Subnet'].replace('.', '-')
                templateVars['Subnet_'] = templateVars['Subnet_'].replace('/', '_')

            # Define the Template Source
            template_file = "ext_subnet_oob.jinja2"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = 'EPG_Mgmt_OOB_External_EPG_%s.tf' % (templateVars['Ext_EPG'])
            dest_dir = 'Tenant_mgmt'
            write_to_site(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def filter_add(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'site_group': '',
                         'Tenant': '',
                         'Filter': ''}
        optional_args = {'description': '',
                         'alias': '',
                         'annotation': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'site_group', templateVars['site_group'])
            validating.name_rule(row_num, ws, 'Filter', templateVars['Filter'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
            if not templateVars['alias'] == None:
                validating.name_rule(row_num, ws, 'alias', templateVars['alias'])
            if not templateVars['description'] == None:
                validating.description(row_num, ws, 'description', templateVars['description'])
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(errorReturn)

        # Define the Template Source
        template_file = "contract_filter.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Contract_Filter_%s.tf' % (templateVars['Filter'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def filter_entry(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'site_group': '',
                         'Tenant': '',
                         'Filter': '',
                         'Filter_Entry': '',
                         'EtherType': '',
                         'IP_Protocol': '',
                         'ARP_Flag': '',
                         'ICMPv4_Type': '',
                         'ICMPv6_Type': '',
                         'Match_DSCP': '',
                         'Match_Only_Frags': '',
                         'Source_From': '',
                         'Source_To': '',
                         'Dest_From': '',
                         'Dest_To': '',
                         'Stateful': '',
                         'TCP_Session_Rules': ''}
        optional_args = {'description': '',
                         'alias': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'site_group', templateVars['site_group'])
            validating.dscp(row_num, ws, 'Match_DSCP', templateVars['Match_DSCP'])
            validating.filter_ports(row_num, ws, 'Source_From', templateVars['Source_From'])
            validating.filter_ports(row_num, ws, 'Source_To', templateVars['Source_To'])
            validating.filter_ports(row_num, ws, 'Dest_From', templateVars['Dest_From'])
            validating.filter_ports(row_num, ws, 'Dest_To', templateVars['Dest_To'])
            validating.name_rule(row_num, ws, 'Filter', templateVars['Filter'])
            validating.name_rule(row_num, ws, 'Filter_Entry', templateVars['Filter_Entry'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
            validating.values(row_num, ws, 'Match_Only_Frags', templateVars['Match_Only_Frags'], ['no', 'yes'])
            validating.values(row_num, ws, 'Stateful', templateVars['Stateful'], ['no', 'yes'])
            validating.values(row_num, ws, 'EtherType', templateVars['EtherType'], ['arp', 'fcoe', 'ip', 'ipv4', 'ipv6', 'trill', 'mac_security', 'mpls_ucast', 'unspecified'])
            validating.values(row_num, ws, 'IP_Protocol', templateVars['IP_Protocol'], ['egp', 'eigrp', 'igp', 'icmp', 'icmpv6', 'igmp', 'l2tp', 'ospfigp', 'pim', 'tcp', 'udp', 'unspecified'])
            validating.values(row_num, ws, 'ARP_Flag', templateVars['ARP_Flag'], ['req', 'reply', 'unspecified'])
            validating.values(row_num, ws, 'ICMPv4_Type', templateVars['ICMPv4_Type'], ['dst-unreach', 'echo', 'echo-rep', 'src-quench', 'time-exceeded', 'unspecified'])
            validating.values(row_num, ws, 'ICMPv6_Type', templateVars['ICMPv6_Type'], ['dst-unreach', 'echo-req', 'echo-rep', 'nbr-solicit', 'nbr-advert', 'redirect', 'time-exceeded', 'unspecified'])
            validating.values(row_num, ws, 'TCP_Session_Rules', templateVars['TCP_Session_Rules'], ['ack', 'est', 'fin', 'rst', 'syn', 'unspecified'])
            if not templateVars['alias'] == None:
                validating.name_rule(row_num, ws, 'alias', templateVars['alias'])
            if not templateVars['description'] == None:
                validating.description(row_num, ws, 'description', templateVars['description'])
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(errorReturn)

        if templateVars['TCP_Session_Rules'] == 'unspecified':
            templateVars['TCP_Session_Rules'] = None

        # Define the Template Source
        template_file = "contract_filter_entry.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Contract_Filter_%s.tf' % (templateVars['Filter'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        write_to_site(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def l3out_add(self, wb, ws, row_num, **kwargs):
        # Open the Network Policies Worksheet
        ws_net = wb['Network Policies']
        rows = ws_net.max_row

        # Dicts for required and optional args
        required_args = {'site_group': '',
                         'Tenant': '',
                         'L3Out': '',
                         'VRF_Tenant': '',
                         'VRF': '',
                         'L3_Domain': '',
                         'target_dscp': '',
                         'Run_BGP': '',
                         'export': '',
                         'import': '',
                         'target_dscp': ''}
        optional_args = {'description': '',
                         'alias': '',
                         'annotation': '',
                         'EIGRP_Routing_Policy': '',
                         'OSPF_Routing_Policy': '',
                         'leak_rtctrlProfile': '',
                         'damp_rtctrlProfile': '',
                         'fvBDPublicSubnetHolder': ''}


        # Get the L3Out Policies from the Network Policies Tab
        func = 'l3Out'
        count = countKeys(ws_net, func)
        row_l3out = ''
        var_dict = findVars(ws_net, func, rows, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('L3Out_Policy'):
                row_l3out = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'site_group', templateVars['site_group'])
            validating.dscp(row_l3out, ws_net, 'target_dscp', templateVars['target_dscp'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
            validating.name_rule(row_num, ws, 'L3Out', templateVars['L3Out'])
            validating.name_rule(row_num, ws, 'VRF', templateVars['VRF'])
            validating.name_rule(row_num, ws, 'VRF_Tenant', templateVars['VRF_Tenant'])
            validating.values(row_num, ws, 'export', templateVars['export'], ['no', 'yes'])
            validating.values(row_num, ws, 'import', templateVars['import'], ['no', 'yes'])
            validating.values(row_num, ws, 'Run_BGP', templateVars['Run_BGP'], ['no', 'yes'])
            if not templateVars['description'] == None:
                validating.description(row_num, ws, 'description', templateVars['description'])
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(errorReturn)

        # Process the template through the Sites
        template_file = "data_domain_l3_profile.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'data_domain_l3_profile_%s.tf' % (templateVars['L3_Domain'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        ctrl_count = 0
        Ctrl = ''
        if templateVars['export'] == 'yes':
            Ctrl = '"export"'
            ctrl_count =+ 1
        if templateVars['import'] == 'yes' and ctrl_count > 0:
            Ctrl = Ctrl + ',' + '"import"'
            ctrl_count =+ 1
        elif templateVars['import'] == 'yes':
            Ctrl = '"import"'
            ctrl_count =+ 1
        if ctrl_count > 0:
            templateVars['enforce_rtctrl'] = '[%s]' % (Ctrl)
        else:
            templateVars['enforce_rtctrl'] = '["unspecified"]'

        # Define the Template Source
        template_file = "l3out.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'L3Out_%s.tf' % (templateVars['L3Out'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        if templateVars['Run_BGP'] == 'yes':
            # Define the Template Source
            template_file = "bgp_external_policy.jinja2"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = 'L3Out_%s.tf' % (templateVars['L3Out'])
            dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
            write_to_site(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

        #--------------------------
        # EIGRP Routing Policy
        #--------------------------
        if not templateVars['EIGRP_Routing_Policy'] == None:

            # Dicts for required and optional args
            required_args = {'site_group': '',
                            'Tenant': '',
                            'L3Out': '',
                            'Policy_Name': '',
                            'AS_Number': ''}
            optional_args = { }
            # Get the L3Out Policies from the Network Policies Tab
            func = 'eigrp_routing'
            count = countKeys(ws_net, func)
            row_eigrp = ''
            var_dict = findVars(ws_net, func, rows, count)
            for pos in var_dict:
                if var_dict[pos].get('Policy_Name') == kwargs.get('EIGRP_Routing_Policy'):
                    row_eigrp = var_dict[pos]['row']
                    del var_dict[pos]['row']
                    kwargs = {**kwargs, **var_dict[pos]}

            # Validate inputs, return dict of template vars
            templateVars = process_kwargs(required_args, optional_args, **kwargs)

            try:
                # Validate Required Arguments
                validating.number_check(row_eigrp, ws_net, 'AS_Number', templateVars['AS_Number'], 1, 65534)
            except Exception as err:
                errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws_net, row_eigrp)
                raise ErrException(errorReturn)

            # Define the Template Source
            template_file = "l3out_eigrp_external_policy.jinja2"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = 'L3Out_%s.tf' % (templateVars['L3Out'])
            dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
            write_to_site(wb, ws_net, row_eigrp, 'a+', dest_dir, dest_file, template, **templateVars)

        #--------------------------
        # OSPF Routing Policy
        #--------------------------
        if not templateVars['OSPF_Routing_Policy'] == None:

            # Dicts for required and optional args
            required_args = {'site_group': '',
                            'Tenant': '',
                            'L3Out': '',
                            'Policy_Name': '',
                            'Area_ID': '',
                            'Area_Type': '',
                            'Cost': '',
                            'Redistribute_NSSA': '',
                            'Originate_Summary': '',
                            'Suppress_FA': ''}
            optional_args = { }
            # Get the L3Out Policies from the Network Policies Tab
            func = 'ospf_routing'
            count = countKeys(ws_net, func)
            row_ospf = ''
            var_dict = findVars(ws_net, func, rows, count)
            for pos in var_dict:
                if var_dict[pos].get('Policy_Name') == kwargs.get('OSPF_Routing_Policy'):
                    row_ospf = var_dict[pos]['row']
                    del var_dict[pos]['row']
                    kwargs = {**kwargs, **var_dict[pos]}

            # Validate inputs, return dict of template vars
            templateVars = process_kwargs(required_args, optional_args, **kwargs)

            try:
                # Validate Required Arguments
                validating.number_check(row_ospf, ws_net, 'Cost', templateVars['Cost'], 0, 16777215)
                validating.values(row_ospf, ws_net, 'Area_Type', templateVars['Area_Type'], ['nssa', 'regular', 'stub'])
                validating.values(row_ospf, ws_net, 'Redistribute_NSSA', templateVars['Redistribute_NSSA'], ['no', 'yes'])
                validating.values(row_ospf, ws_net, 'Originate_Summary', templateVars['Originate_Summary'], ['no', 'yes'])
                validating.values(row_ospf, ws_net, 'Suppress_FA', templateVars['Suppress_FA'], ['no', 'yes'])
            except Exception as err:
                errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws_net, row_ospf)
                raise ErrException(errorReturn)

            ctrl_count = 0
            Ctrl = ''
            if templateVars['Redistribute_NSSA'] == 'yes':
                Ctrl = 'redistribute'
                ctrl_count =+ 1
            if templateVars['Originate_Summary'] == 'yes' and ctrl_count > 0:
                Ctrl = Ctrl + ',' + 'summary'
                ctrl_count =+ 1
            elif templateVars['Originate_Summary'] == 'yes':
                Ctrl = 'summary'
                ctrl_count =+ 1
            if templateVars['Suppress_FA'] == 'yes' and ctrl_count > 0:
                Ctrl = Ctrl + ',' + 'supress-fa'
                ctrl_count =+ 1
            elif templateVars['Suppress_FA'] == 'yes':
                Ctrl = 'supress-fa'
                ctrl_count =+ 1
            if ctrl_count > 0:
                templateVars['Ctrl'] = '%s' % (Ctrl)
            else:
                templateVars['Ctrl'] = 'unspecified'

            # Define the Template Source
            template_file = "l3out_ospf_external_policy.jinja2"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = 'L3Out_%s.tf' % (templateVars['L3Out'])
            dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
            write_to_site(wb, ws_net, row_ospf, 'a+', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def mgmt_epg(self, wb, ws, row_num, **kwargs):
        # Dicts for Bridge Domain required and optional args
        required_args = {'site_group': '',
                         'Type': '',
                         'EPG': '',
                         'QoS_Class': ''}
        optional_args = {'annotation': '',
                         'VLAN': '',
                         'Bridge_Domain': '',
                         'Tenant': '',
                         'consumed_Contracts': '',
                         'provided_Contracts': '',
                         'match_t': '',
                         'Contract_Interfaces': '',
                         'Taboo_Contracts': '',
                         'Subnets': '',
                         'Static_Routes': '',}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        templateVars['consume_count'] = 1
        templateVars['provide_count'] = 1
        templateVars['interface_count'] = 1
        templateVars['taboo_count'] = 1
        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'site_group', templateVars['site_group'])
            validating.name_rule(row_num, ws, 'EPG', templateVars['EPG'])
            validating.qos_priority(row_num, ws, 'QoS_Class', templateVars['QoS_Class'])
            validating.values(row_num, ws, 'Type', templateVars['Type'], ['in_band', 'out_of_band'])
            if templateVars['Type'] == 'in_band':
                validating.vlans(row_num, ws, 'VLAN', templateVars['VLAN'])
                validating.name_rule(row_num, ws, 'Bridge_Domain', templateVars['Bridge_Domain'])
            if not templateVars['Tenant'] == None:
                validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
                if re.search(',', templateVars['provided_Contracts']):
                    templateVars['provide_count'] =+ 1
                    for x in templateVars['provided_Contracts'].split(','):
                        validating.name_rule(row_num, ws, 'provided_Contracts', x)
                else:
                    validating.name_rule(row_num, ws, 'provided_Contracts', templateVars['provided_Contracts'])
                if templateVars['Type'] == 'in_band':
                    if not templateVars['consumed_Contracts'] == None:
                        if re.search(',', templateVars['consumed_Contracts']):
                            templateVars['provide_count'] =+ 1
                            for x in templateVars['consumed_Contracts'].split(','):
                                validating.name_rule(row_num, ws, 'consumed_Contracts', x)
                        else:
                            validating.name_rule(row_num, ws, 'consumed_Contracts', templateVars['consumed_Contracts'])
                    if not templateVars['Contract_Interfaces'] == None:
                        if re.search(',', templateVars['Contract_Interfaces']):
                            for x in templateVars['Contract_Interfaces'].split(','):
                                templateVars['interface_count'] =+ 1
                                validating.not_empty(row_num, ws, 'Contract_Interfaces', x)
                        else:
                            validating.not_empty(row_num, ws, 'Contract_Interfaces', templateVars['Contract_Interfaces'])
                    if not templateVars['Taboo_Contracts'] == None:
                        if re.search(',', templateVars['Taboo_Contracts']):
                            templateVars['taboo_count'] =+ 1
                            for x in templateVars['Taboo_Contracts'].split(','):
                                validating.not_empty(row_num, ws, 'Taboo_Contracts', x)
                        else:
                            validating.not_empty(row_num, ws, 'Taboo_Contracts', templateVars['Taboo_Contracts'])
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(errorReturn)

        if not templateVars['Tenant'] == 'mgmt':
            dest_dir = 'Tenant_mgmt'

            template_file = 'data_tenant.jinja2'
            template = self.templateEnv.get_template(template_file)
            dest_file = 'data_Tenant_%s.tf' % (templateVars['Tenant'])
            write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

            templateVars['Contract_Type'] = 'Standard'
            if not templateVars['consumed_Contracts'] == None:
                if re.search(',', templateVars['consumed_Contracts']):
                    for x in templateVars['consumed_Contracts'].split(','):
                        templateVars['Contract'] = x
                        template_file = 'data_contract.jinja2'
                        template = self.templateEnv.get_template(template_file)
                        dest_file = 'data_Contract_Type_%s_%s.tf' % (templateVars['Contract_Type'], x)
                        write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)
                else:
                    templateVars['Contract'] = templateVars['consumed_Contracts']
                    template_file = 'data_contract.jinja2'
                    template = self.templateEnv.get_template(template_file)
                    dest_file = 'data_Contract_Type_%s_%s.tf' % (templateVars['Contract_Type'], templateVars['consumed_Contracts'])
                    write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

            if not templateVars['Contract_Interfaces'] == None:
                if re.search(',', templateVars['Contract_Interfaces']):
                    for x in templateVars['Contract_Interfaces'].split(','):
                        templateVars['Contract'] = x
                        template_file = 'data_contract.jinja2'
                        template = self.templateEnv.get_template(template_file)
                        dest_file = 'data_Contract_Type_%s_%s.tf' % (templateVars['Contract_Type'], x)
                        write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)
                else:
                    templateVars['Contract'] = templateVars['Contract_Interfaces']
                    template_file = 'data_contract.jinja2'
                    template = self.templateEnv.get_template(template_file)
                    dest_file = 'data_Contract_Type_%s_%s.tf' % (templateVars['Contract_Type'], templateVars['Contract_Interfaces'])
                    write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

            if not templateVars['provided_Contracts'] == None:
                if templateVars['Type'] == 'in_band':
                    if re.search(',', templateVars['provided_Contracts']):
                        for x in templateVars['provided_Contracts'].split(','):
                            templateVars['Contract'] = x
                            template_file = 'data_contract.jinja2'
                            template = self.templateEnv.get_template(template_file)
                            dest_file = 'data_Contract_Type_%s_%s.tf' % (templateVars['Contract_Type'], x)
                            write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)
                    else:
                        templateVars['Contract'] = templateVars['provided_Contracts']
                        template_file = 'data_contract.jinja2'
                        template = self.templateEnv.get_template(template_file)
                        dest_file = 'data_Contract_Type_%s_%s.tf' % (templateVars['Contract_Type'], templateVars['provided_Contracts'])
                        write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

            templateVars['Contract_Type'] = 'Taboo'
            if not templateVars['Taboo_Contracts'] == None:
                if re.search(',', templateVars['Taboo_Contracts']):
                    for x in templateVars['Taboo_Contracts'].split(','):
                        templateVars['Contract'] == x
                        template_file = 'data_contract_taboo.jinja2'
                        template = self.templateEnv.get_template(template_file)
                        dest_file = 'data_Contract_Type_%s_%s.tf' % (templateVars['Contract_Type'], x)
                        write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)
                else:
                    templateVars['Contract'] = templateVars['Taboo_Contracts']
                    template_file = 'data_contract_taboo.jinja2'
                    template = self.templateEnv.get_template(template_file)
                    dest_file = 'data_Contract_Type_%s_%s.tf' % (templateVars['Contract_Type'], templateVars['Contract_Type'])
                    write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source and Destination File
        template_file = "epg_mgmt.jinja2"
        template = self.templateEnv.get_template(template_file)
        dest_file = 'EPG_Mgmt_Type_%s_EPG_%s.tf' % (templateVars['Type'], templateVars['EPG'])

        # Process the template through the Sites
        dest_dir = 'Tenant_mgmt'
        write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source and Destination File
        template_file = "var_mgmt.jinja2"
        template = self.templateEnv.get_template(template_file)

        if templateVars['Type'] == 'in_band':
            templateVars['var_name'] = 'in_band'
            dest_file = 'var_Mgmt_EPG_%s.tf' % ('inb')
        else:
            templateVars['var_name'] = 'out_of_band'
            dest_file = 'var_Mgmt_EPG_%s.tf' % ('oob')

        # Process the template through the Sites
        dest_dir = 'Access'
        write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)
        dest_dir = 'Admin'
        write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)
        dest_dir = 'Fabric'
        write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def node_intf(self, wb, ws, row_num, **kwargs):
        # Open the Network Policies Worksheet
        ws_net = wb['Network Policies']
        rows = ws_net.max_row

        # Dicts for required and optional args
        required_args = {'site_group': '',
                         'Tenant': '',
                         'L3Out': '',
                         'Node_Profile': '',
                         'Interface_Profile': '',
                         'QoS_Class': '',
                         'Node_Intf_Policy': '',
                         'Policy_Name': '',
                         'tag': ''}
        optional_args = {'description': '',
                         'alias': '',
                         'EIGRP_Intf_Profile': '',
                         'OSPF_Intf_Profile': '',
                         'ndIfPol': '',
                         'egress_qosDppPol': '',
                         'ingress_qosDppPol': '',
                         'qosCustomPol': '',
                         'igmp_policy': '',
                         'netflowMonitorPol': ''}

        # Get the Node Policies from the Network Policies Tab
        func = 'node_intf'
        count = countKeys(ws_net, func)
        row_node = ''
        var_dict = findVars(ws_net, func, rows, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('Node_Intf_Policy'):
                row_node = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'site_group', templateVars['site_group'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
            validating.name_rule(row_num, ws, 'L3Out', templateVars['L3Out'])
            validating.name_rule(row_num, ws, 'Node_Profile', templateVars['Node_Profile'])
            validating.name_rule(row_num, ws, 'Interface_Profile', templateVars['Interface_Profile'])
            validating.qos_priority(row_num, ws, 'QoS_Class', templateVars['QoS_Class'])
            if not templateVars['alias'] == None:
                validating.name_rule(row_num, ws, 'alias', templateVars['alias'])
            if not templateVars['description'] == None:
                validating.description(row_num, ws, 'description', templateVars['description'])
            if not templateVars['EIGRP_Intf_Profile'] == None:
                validating.name_rule(row_num, ws, 'EIGRP_Intf_Profile', templateVars['EIGRP_Intf_Profile'])
            if not templateVars['OSPF_Intf_Profile'] == None:
                validating.name_rule(row_num, ws, 'OSPF_Intf_Profile', templateVars['OSPF_Intf_Profile'])
            if not templateVars['tag'] == None:
                validating.tag_check(row_node, ws_net, 'tag', templateVars['tag'])
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(errorReturn)

        # Define the Template Source
        template_file = "logical_interface_profile.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'L3Out_%s_Node_Profile_%s.tf' % (templateVars['L3Out'], templateVars['Node_Profile'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        write_to_site(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

        #--------------------------
        # OSPF Interface Profile
        #--------------------------
        if not templateVars['OSPF_Intf_Profile'] == None:

            # Dicts for required and optional args
                             # OSPF Interface Profile
            required_args = {'site_group': '',
                             'Tenant': '',
                             'L3Out': '',
                             'Node_Profile': '',
                             'Interface_Profile': '',
                             'Auth_Type': '',
                             'Interface_Policy_Name': '',
                             # OSPF Interface Policy
                             'OSPF_Policy_Name': '',
                             'Policy_Tenant': '',
                             'Network_Type': '',
                             'Priority': '',
                             'Cost': '',
                             'Advertise_Subnet': '',
                             'BFD': '',
                             'MTU_Ignore': '',
                             'Passive_Interface': '',
                             'Hello_Interval': '',
                             'Dead_Interval': '',
                             'Retransmit_Interval': '',
                             'Transmit_Delay': ''}
                             # OSPF Interface Profile
            optional_args = {'Auth_Key_ID': '',
                             'OSPF_Auth_Key': '',
                             'description': '',
                             # OSPF Interface Policy
                             'OSPF_description': ''}

            # Get the OSPF Profile Attributes from the Network Policies Tab
            func = 'ospf_profile'
            count = countKeys(ws_net, func)
            row_ospf = ''
            var_dict = findVars(ws_net, func, rows, count)
            for pos in var_dict:
                if var_dict[pos].get('Policy_Name') == kwargs.get('OSPF_Intf_Profile'):
                    row_ospf = var_dict[pos]['row']
                    del var_dict[pos]['row']
                    kwargs = {**kwargs, **var_dict[pos]}

           # Get the OSPF Policy Attributes from the Network Policies Tab
            func = 'ospf_policy'
            count = countKeys(ws_net, func)
            row_intf = ''
            var_dict = findVars(ws_net, func, rows, count)
            for pos in var_dict:
                if var_dict[pos].get('OSPF_Policy_Name') == kwargs.get('Interface_Policy_Name'):
                    row_intf = var_dict[pos]['row']
                    del var_dict[pos]['row']
                    kwargs = {**kwargs, **var_dict[pos]}
            # print(kwargs)
            # Validate inputs, return dict of template vars
            templateVars = process_kwargs(required_args, optional_args, **kwargs)

            try:
                # Validate OSPF Profile Required Arguments
                validating.values(row_ospf, ws_net, 'Auth_Type', templateVars['Auth_Type'], ['md5', 'none', 'simple'])
                if not templateVars['Auth_Key_ID'] == None:
                    validating.number_check(row_ospf, ws_net, 'Auth_Key_ID', templateVars['Auth_Key_ID'], 1, 255)
                if not templateVars['description'] == None:
                    validating.description(row_ospf, ws_net, 'description', templateVars['description'])
                # Validate OSPF Policy Required Arguments
                validating.name_rule(row_intf, ws_net, 'OSPF_Policy_Name', templateVars['OSPF_Policy_Name'])
                validating.number_check(row_intf, ws_net, 'Priority', templateVars['Priority'], 0, 255)
                validating.number_check(row_intf, ws_net, 'Cost', templateVars['Cost'], 0, 65535)
                validating.number_check(row_intf, ws_net, 'Hello_Interval', templateVars['Hello_Interval'], 1, 65535)
                validating.number_check(row_intf, ws_net, 'Dead_Interval', templateVars['Dead_Interval'], 1, 65535)
                validating.number_check(row_intf, ws_net, 'Retransmit_Interval', templateVars['Retransmit_Interval'], 1, 65535)
                validating.number_check(row_intf, ws_net, 'Transmit_Delay', templateVars['Transmit_Delay'], 1, 65535)
                validating.values(row_intf, ws_net, 'Network_Type', templateVars['Network_Type'], ['broadcast', 'p2p', 'unspecified'])
                validating.values(row_intf, ws_net, 'Advertise_Subnet', templateVars['Advertise_Subnet'], ['no', 'yes'])
                validating.values(row_intf, ws_net, 'BFD', templateVars['BFD'], ['no', 'yes'])
                validating.values(row_intf, ws_net, 'MTU_Ignore', templateVars['MTU_Ignore'], ['no', 'yes'])
                validating.values(row_intf, ws_net, 'Passive_Interface', templateVars['Passive_Interface'], ['no', 'yes'])
                if not templateVars['OSPF_description'] == None:
                    validating.description(row_intf, ws_net, 'OSPF_description', templateVars['OSPF_description'])
            except Exception as err:
                errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws_net, row_ospf)
                raise ErrException(errorReturn)

            if templateVars['Auth_Type'] == 'none':
                templateVars['Auth_Key_ID'] = 1

            # Define the Template Source
            template_file = "l3out_ospf_interface_profile.jinja2"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = 'L3Out_%s_Node_Profile_%s.tf' % (templateVars['L3Out'], templateVars['Node_Profile'])
            dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
            write_to_site(wb, ws_net, row_ospf, 'a+', dest_dir, dest_file, template, **templateVars)

            if not templateVars['OSPF_Auth_Key'] == None:
                x = templateVars['OSPF_Auth_Key'].split('r')
                key_number = x[1]
                templateVars['sensitive_var'] = 'OSPF_Auth_Key%s' % (key_number)

                # Define the Template Source
                template_file = "variables.jinja2"
                template = self.templateEnv.get_template(template_file)

                # Process the template through the Sites
                dest_file = 'variable_%s.tf' % (templateVars['sensitive_var'])
                dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
                write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

                sensitive_var_site_group(wb, ws, row_num, dest_dir, dest_file, template, **templateVars)

            #--------------------------
            # OSPF Interface Policy
            #--------------------------

            ctrl_count = 0
            Ctrl = ''
            if templateVars['Advertise_Subnet'] == 'yes':
                Ctrl = '"advert-subnet"'
                ctrl_count =+ 1
            if templateVars['BFD'] == 'yes' and ctrl_count > 0:
                Ctrl = Ctrl + ', ' + '"bfd"'
                ctrl_count =+ 1
            elif templateVars['BFD'] == 'yes':
                Ctrl = '"bfd"'
                ctrl_count =+ 1
            if templateVars['MTU_Ignore'] == 'yes' and ctrl_count > 0:
                Ctrl = Ctrl + ', ' + '"mtu-ignore"'
                ctrl_count =+ 1
            elif templateVars['MTU_Ignore'] == 'yes':
                Ctrl = '"mtu-ignore"'
                ctrl_count =+ 1
            if templateVars['Passive_Interface'] == 'yes' and ctrl_count > 0:
                Ctrl = Ctrl + ', ' + '"passive"'
                ctrl_count =+ 1
            elif templateVars['Passive_Interface'] == 'yes':
                Ctrl = '"passive"'
                ctrl_count =+ 1
            if ctrl_count > 0:
                templateVars['Ctrl'] = '[%s]' % (Ctrl)
            else:
                templateVars['Ctrl'] = '[unspecified]'

            if templateVars['Cost'] == 0:
                templateVars['Cost'] = 'unspecified'

            # Define the Template Source
            template_file = "ospf_interface_policy.jinja2"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = 'Policies_OSPF_Interface_%s.tf' % (templateVars['OSPF_Policy_Name'])
            dest_dir = 'Tenant_%s' % (templateVars['Policy_Tenant'])
            write_to_site(wb, ws_net, row_ospf, 'w', dest_dir, dest_file, template, **templateVars)

            if not templateVars['Tenant'] == templateVars['Policy_Tenant']:
                # Define the Template Source
                template_file = "data_ospf_interface_policy.jinja2"
                template = self.templateEnv.get_template(template_file)

                # Process the template through the Sites
                dest_file = 'data_Policies_OSPF_Interface_%s.tf' % (templateVars['OSPF_Policy_Name'])
                dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
                write_to_site(wb, ws_net, row_ospf, 'w', dest_dir, dest_file, template, **templateVars)

                templateVars['L3Out_Tenant'] = templateVars['Tenant']
                templateVars['Tenant'] = templateVars['Policy_Tenant']
                # Define the Template Source
                template_file = "data_tenant.jinja2"
                template = self.templateEnv.get_template(template_file)

                # Process the template through the Sites
                dest_file = 'data_Tenant_%s.tf' % (templateVars['Tenant'])
                dest_dir = 'Tenant_%s' % (templateVars['L3Out_Tenant'])
                write_to_site(wb, ws_net, row_ospf, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def node_path(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
                         # Logical Interface Profile
        required_args = {'site_group': '',
                         'Path_Policy_Name': '',
                         'MTU': '',
                         'Target_DSCP': '',
                         'SideA_Address': '',
                         'SideA_IPv6_DAD': '',
                         # L3Out Path Profile
                         'Tenant': '',
                         'L3Out': '',
                         'Policy_Name': '',
                         'Node_Profile': '',
                         'Interface_Profile': '',
                         'Interface_Type': '',
                         'Pod_ID': '',
                         'Node1_ID': '',
                         'Interface_or_PG': ''}
                         # Logical Interface Profile
        optional_args = {'Encap_Scope': '',
                         'Mode': '',
                         'VLAN': '',
                         'description': '',
                         'Auto_State': '',
                         'MAC_Address': '',
                         'SideA_Secondary': '',
                         'SideA_Link_Local': '',
                         'SideB_Address': '',
                         'SideB_IPv6_DAD': '',
                         'SideB_Secondary': '',
                         'SideB_Link_Local': '',
                         'Node2_ID': ''}

        # Get the Node Policies from the Network Policies Tab
        rows = ws.max_row
        func = 'l3out_path'
        count = countKeys(ws, func)
        row_path = ''
        var_dict = findVars(ws, func, rows, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('Path_Policy_Name'):
                row_path = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments Logincal Interface Profile
            validating.site_group(row_num, ws, 'site_group', templateVars['site_group'])
            validating.name_rule(row_path, ws, 'Path_Policy_Name', templateVars['Path_Policy_Name'])
            if not templateVars['Encap_Scope'] == None:
                validating.values(row_num, ws, 'Encap_Scope', templateVars['Encap_Scope'], ['ctx', 'local'])
            if not templateVars['Mode'] == None:
                validating.values(row_num, ws, 'Mode', templateVars['Mode'], ['native', 'regular', 'untagged'])
            if not templateVars['VLAN'] == None:
                validating.vlans(row_num, ws, 'VLAN', templateVars['VLAN'])
            if not templateVars['description'] == None:
                validating.description(row_num, ws, 'description', templateVars['description'])
            if not templateVars['Auto_State'] == None:
                validating.values(row_num, ws, 'Auto_State', templateVars['Auto_State'], ['disabled', 'enabled'])
            if not templateVars['MTU'] == 'inherit':
                validating.number_check(row_path, ws, 'MTU', templateVars['MTU'], 1300, 9216)
            if not templateVars['MAC_Address'] == None:
                validating.mac_address(row_num, ws, 'MAC_Address', templateVars['MAC_Address'])
            validating.dscp(row_num, ws, 'Target_DSCP', templateVars['Target_DSCP'])
            validating.ip_address(row_num, ws, 'SideA_Address', templateVars['SideA_Address'])
            validating.values(row_num, ws, 'SideA_IPv6_DAD', templateVars['SideA_IPv6_DAD'], ['disabled', 'enabled'])
            if not templateVars['SideA_Secondary'] == None:
                validating.ip_address(row_num, ws, 'SideA_Secondary', templateVars['SideA_Secondary'])
            if not templateVars['SideA_Link_Local'] == None:
                validating.ip_address(row_num, ws, 'SideA_Link_Local', templateVars['SideA_Link_Local'])
            if not templateVars['SideB_Address'] == None:
                validating.ip_address(row_num, ws, 'SideB_Address', templateVars['SideB_Address'])
            if not templateVars['SideB_Address'] == None:
                validating.values(row_num, ws, 'SideB_IPv6_DAD', templateVars['SideB_IPv6_DAD'], ['disabled', 'enabled'])
            if not templateVars['SideB_Secondary'] == None:
                validating.ip_address(row_num, ws, 'SideB_Secondary', templateVars['SideB_Secondary'])
            if not templateVars['SideB_Link_Local'] == None:
                validating.ip_address(row_num, ws, 'SideB_Link_Local', templateVars['SideB_Link_Local'])
            validating.name_rule(row_path, ws, 'Tenant', templateVars['Tenant'])
            validating.name_rule(row_path, ws, 'L3Out', templateVars['L3Out'])
            validating.name_rule(row_path, ws, 'Node_Profile', templateVars['Node_Profile'])
            validating.name_rule(row_path, ws, 'Interface_Profile', templateVars['Interface_Profile'])
            validating.values(row_path, ws, 'Interface_Type', templateVars['Interface_Type'], ['ext-svi', 'l3-port', 'sub-interface'])
            validating.number_check(row_path, ws, 'Pod_ID', templateVars['Pod_ID'], 1, 15)
            validating.number_check(row_path, ws, 'Node1_ID', templateVars['Node1_ID'], 101, 4001)
            if not templateVars['Node2_ID'] == None:
                validating.number_check(row_path, ws, 'Node2_ID', templateVars['Node2_ID'], 101, 4001)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(errorReturn)

        # Create Global Variables for First Template
        if not templateVars['Node2_ID'] == None:
            templateVars['Address'] = None
            templateVars['IPv6_DAD'] = templateVars['SideA_IPv6_DAD']
            templateVars['Link_Local'] = None
            templateVars['PATH'] = 'protpaths-%s-%s' % (templateVars['Node1_ID'], templateVars['Node2_ID'])
        else:
            templateVars['Address'] = templateVars['SideA_Address']
            templateVars['IPv6_DAD'] = templateVars['SideA_IPv6_DAD']
            templateVars['Link_Local'] = templateVars['SideA_Link_Local']
            templateVars['PATH'] = 'paths-%s' % (templateVars['Node1_ID'])
        # Define the Template Source
        template_file = "l3out_path_attachment.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'L3Out_%s_Node_Profile_%s.tf' % (templateVars['L3Out'], templateVars['Node_Profile'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        write_to_site(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

        # Create Global Variables for First Template
        if not templateVars['Node2_ID'] == None:
            templateVars['PATH'] = 'protpaths-%s-%s' % (templateVars['Node1_ID'], templateVars['Node2_ID'])

            templateVars['Address'] = templateVars['SideA_Address']
            templateVars['IPv6_DAD'] = templateVars['SideA_IPv6_DAD']
            templateVars['Link_Local'] = templateVars['SideA_Link_Local']
            templateVars['Side'] = 'A'

            # Define the Template Source
            template_file = "l3out_path_attach_vpc.jinja2"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = 'L3Out_%s_Node_Profile_%s.tf' % (templateVars['L3Out'], templateVars['Node_Profile'])
            dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
            write_to_site(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

            templateVars['Address'] = templateVars['SideB_Address']
            templateVars['IPv6_DAD'] = templateVars['SideB_IPv6_DAD']
            templateVars['Link_Local'] = templateVars['SideB_Link_Local']
            templateVars['Side'] = 'B'

            # Define the Template Source
            template_file = "l3out_path_attach_vpc.jinja2"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = 'L3Out_%s_Node_Profile_%s.tf' % (templateVars['L3Out'], templateVars['Node_Profile'])
            dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
            write_to_site(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        if not templateVars['SideA_Secondary'] == None:
            if templateVars['Node2_ID'] == None:
                templateVars['Secondary'] = templateVars['SideA_Secondary']
                templateVars['IPv6_DAD'] = templateVars['SideA_IPv6_DAD']
                templateVars['PATH'] = 'paths-%s' % (templateVars['Node1_ID'])
                template_file = "l3out_path_attach_secondary.jinja2"
                template = self.templateEnv.get_template(template_file)

                # Process the template through the Sites
                dest_file = 'L3Out_%s_Node_Profile_%s.tf' % (templateVars['L3Out'], templateVars['Node_Profile'])
                dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
                write_to_site(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def node_prof(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'site_group': '',
                         'Tenant': '',
                         'L3Out': '',
                         'Node_Profile': '',
                         'Target_DSCP': '',
                         'Color_Tag': '',
                         'Pod_ID': '',
                         'Node1_ID': '',
                         'Node1_Router_ID': '',
                         'Node1_Loopback': ''}
        optional_args = {'alias': '',
                         'description': '',
                         'Node2_ID': '',
                         'Node2_Router_ID': '',
                         'Node2_Loopback': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'site_group', templateVars['site_group'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
            validating.name_rule(row_num, ws, 'L3Out', templateVars['L3Out'])
            validating.name_rule(row_num, ws, 'Node_Profile', templateVars['Node_Profile'])
            validating.dscp(row_num, ws, 'Target_DSCP', templateVars['Target_DSCP'])
            validating.tag_check(row_num, ws, 'Color_Tag', templateVars['Color_Tag'])
            validating.number_check(row_num, ws, 'Node1_ID', templateVars['Node1_ID'], 101, 4001)
            validating.ip_address(row_num, ws, 'Node1_Router_ID', templateVars['Node1_Router_ID'])
            validating.values(row_num, ws, 'Node1_Loopback', templateVars['Node1_Loopback'], ['no', 'yes'])
            if not templateVars['alias'] == None:
                validating.name_rule(row_num, ws, 'alias', templateVars['alias'])
            if not templateVars['description'] == None:
                validating.description(row_num, ws, 'description', templateVars['description'])
            if not templateVars['Node2_ID'] == None:
                validating.number_check(row_num, ws, 'Node2_ID', templateVars['Node2_ID'], 101, 4001)
                validating.ip_address(row_num, ws, 'Node2_Router_ID', templateVars['Node2_Router_ID'])
                validating.values(row_num, ws, 'Node2_Loopback', templateVars['Node2_Loopback'], ['no', 'yes'])
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(errorReturn)

        # Define the Template Source
        template_file = "logical_node_profile.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'L3Out_%s_Node_Profile_%s.tf' % (templateVars['L3Out'], templateVars['Node_Profile'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "logical_node_to_fabric_node.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Modify Variables for Template
        templateVars['Node_ID'] = templateVars['Node1_ID']
        templateVars['rtr_id'] = templateVars['Node1_Router_ID']
        templateVars['rtr_id_loopback'] = templateVars['Node1_Loopback']

        # Process the template through the Sites
        dest_file = 'L3Out_%s_Node_Profile_%s.tf' % (templateVars['L3Out'], templateVars['Node_Profile'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        write_to_site(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

        if not templateVars['Node2_ID'] == None:
            # Modify Variables for Template
            templateVars['Node_ID'] = templateVars['Node2_ID']
            templateVars['rtr_id'] = templateVars['Node2_Router_ID']
            templateVars['rtr_id_loop_back'] = templateVars['Node2_Loopback']

            # Process the template through the Sites
            dest_file = 'L3Out_%s_Node_Profile_%s.tf' % (templateVars['L3Out'], templateVars['Node_Profile'])
            dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
            write_to_site(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def subject_add(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'site_group': '',
                         'Tenant': '',
                         'Subject': '',
                         'Contract_Type': '',
                         'Contract': '',
                         'Reverse_Filter_Ports': '',
                         'QoS_Class': '',
                         'Target_DSCP': '',
                         'Filters_to_Assign': ''}
        optional_args = {'description': '',
                         'alias': '',
                         'annotation': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        templateVars['filters_count'] = 1
        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'site_group', templateVars['site_group'])
            validating.dscp(row_num, ws, 'Target_DSCP', templateVars['Target_DSCP'])
            validating.name_rule(row_num, ws, 'Contract', templateVars['Contract'])
            validating.name_rule(row_num, ws, 'Subject', templateVars['Subject'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
            validating.qos_priority(row_num, ws, 'QoS_Class', templateVars['QoS_Class'])
            validating.values(row_num, ws, 'Contract_Type', templateVars['Contract_Type'], ['OOB', 'Standard', 'Taboo'])
            validating.values(row_num, ws, 'Reverse_Filter_Ports', templateVars['Reverse_Filter_Ports'], ['no', 'yes'])
            if not templateVars['alias'] == None:
                validating.name_rule(row_num, ws, 'alias', templateVars['alias'])
            if not templateVars['description'] == None:
                validating.description(row_num, ws, 'description', templateVars['description'])
            if not templateVars['Filters_to_Assign'] == None:
                if re.search(',', templateVars['Filters_to_Assign']):
                    templateVars['filters_count'] =+ 1
                    for x in templateVars['Filters_to_Assign'].split(','):
                        validating.name_rule(row_num, ws, 'Filters_to_Assign', x)
                else:
                    validating.name_rule(row_num, ws, 'Filters_to_Assign', templateVars['Filters_to_Assign'])
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(errorReturn)

        # Define the Template Source
        template_file = "contract_subject.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Define the Template Source
        if templateVars['Contract_Type'] == 'OOB':
            template_file = "contract_subject.jinja2"
        elif templateVars['Contract_Type'] == 'Standard':
            template_file = "contract_subject.jinja2"
        elif templateVars['Contract_Type'] == 'Taboo':
            template_file = "contract_subject.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Contract_Type_%s_%s_Subj_%s.tf' % (templateVars['Contract_Type'], templateVars['Contract'], templateVars['Subject'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def subnet_add(self, wb, ws, row_num, **kwargs):
        # Open the Network Policies Worksheet
        ws_net = wb['Network Policies']
        rows = ws_net.max_row

        # Dicts for Subnet required and optional args
        required_args = {'site_group': '',
                         'Tenant': '',
                         'Bridge_Domain': '',
                         'Subnet': '',
                         'Subnet_Policy': '',
                         'Policy_Name': '',
                         'virtual': '',
                         'preferred': '',
                         'scope': '',
                         'nd': '',
                         'no-default-gateway': '',
                         'querier': ''}
        optional_args = {'Subnet_description': '',
                         'l3extOut': '',
                         'rtctrlProfile': '',
                         'ndPfxPol': ''}

        # Get the Subnet Policies from the Network Policies Tab
        func = 'subnet'
        count = countKeys(ws_net, func)
        row_subnet = ''
        var_dict = findVars(ws_net, func, rows, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('Subnet_Policy'):
                row_subnet = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}
                break

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'site_group', templateVars['site_group'])
            validating.ip_address(row_num, ws, 'Subnet', templateVars['Subnet'])
            if not templateVars['Subnet_description'] == None:
                validating.description(row_num, ws, 'Subnet_description', templateVars['Subnet_description'])
            validating.values(row_subnet, ws_net, 'virtual', templateVars['virtual'], ['no', 'yes'])
            validating.values(row_subnet, ws_net, 'preferred', templateVars['preferred'], ['no', 'yes'])
            validating.values(row_subnet, ws_net, 'scope', templateVars['scope'], ['private', 'public', 'shared', 'private-shared', 'public-shared'])
            validating.values(row_subnet, ws_net, 'nd', templateVars['nd'], ['no', 'yes'])
            validating.values(row_subnet, ws_net, 'no-default-gateway', templateVars['no-default-gateway'], ['no', 'yes'])
            validating.values(row_subnet, ws_net, 'querier', templateVars['querier'], ['no', 'yes'])
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(errorReturn)

        if templateVars['l3extOut'] == 'default':
            templateVars['l3extOut'] = 'uni/tn-common/out-default'
        if templateVars['ndPfxPol'] == 'default':
            templateVars['ndPfxPol'] = 'uni/tn-common/ndpfxpol-default'

        # Create ctrl templateVars
        ctrl_count = 0
        Ctrl = ''
        if templateVars['nd'] == 'yes':
            Ctrl = '"nd"'
            ctrl_count =+ 1
        if templateVars['no-default-gateway'] == 'yes' and ctrl_count > 0:
            Ctrl = Ctrl + ', ' + '"no-default-gateway"'
            ctrl_count =+ 1
        elif templateVars['no-default-gateway'] == 'yes':
            Ctrl = '"no-default-gateway"'
            ctrl_count =+ 1
        if templateVars['querier'] == 'yes' and ctrl_count > 0:
            Ctrl = Ctrl + ', ' + '"querier"'
            ctrl_count =+ 1
        elif templateVars['querier'] == 'yes':
            Ctrl = '"querier"'
            ctrl_count =+ 1

        if ctrl_count > 0:
            templateVars['Ctrl'] = '[%s]' % (Ctrl)
        else:
            templateVars['Ctrl'] = '["unspecified"]'

        # Modify scope templateVars
        if re.search('^(private|public|shared)$', templateVars['scope']):
            templateVars['scope'] = '"%s"' % (templateVars['scope'])
        elif re.search('^(private|public)\\-shared$', templateVars['scope']):
            x = templateVars['scope'].split('-')
            templateVars['scope'] = '"%s", "%s"' % (x[0], x[1])

        # As period and colon are not allowed in description need to modify Subnet to work for description and filename
        if ':' in templateVars['Subnet']:
            network = "%s" % (ipaddress.IPv6Network(templateVars['Subnet'], strict=False))
            templateVars['Subnet_'] = network
            templateVars['Subnet_'] = templateVars['Subnet_'].replace(':', '-')
            templateVars['Subnet_'] = templateVars['Subnet_'].replace('/', '_')
        else:
            network = "%s" % (ipaddress.IPv4Network(templateVars['Subnet'], strict=False))
            templateVars['Subnet_'] = network
            templateVars['Subnet_'] = templateVars['Subnet_'].replace('.', '-')
            templateVars['Subnet_'] = templateVars['Subnet_'].replace('/', '_')

        # Define the Template Source
        template_file = "subnet.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Bridge_Domain_%s.tf' % (templateVars['Bridge_Domain'],)
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        write_to_site(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def tenant_add(self, **kwargs):
        # Set Locally Used Variables
        wb = kwargs['wb']
        ws = kwargs['ws']
        row_num = kwargs['row_num']

        # Dicts for required and optional args
        required_args = {
            'site_group': '',
            'tenant': ''
        }
        optional_args = {
            'alias': '',
            'annotations': '',
            'description': '',
            'users': '',
        }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'site_group', templateVars['site_group'])
            validating.name_rule(row_num, ws, 'tenant', templateVars['tenant'])
            if not templateVars['alias'] == None:
                validating.name_rule(row_num, ws, 'alias', templateVars['alias'])
            if not templateVars['annotations'] == None:
                for i in templateVars['annotations']:
                    for k, v in i.items():
                        validating.name_rule(row_num, ws, 'annotations', k)
                        validating.name_rule(row_num, ws, 'annotations', v)
            if not templateVars['description'] == None:
                validating.description(row_num, ws, 'description', templateVars['description'])
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(errorReturn)

        upDates = {
            'monitoring_policy':'default',
            'sites':[],
        }
        kwargs.update(upDates)

        # Add Dictionary to easyDict
        templateVars['class_type'] = 'tenants'
        templateVars['data_type'] = 'tenants'
        kwargs['easyDict'] = update_easyDict(templateVars, **kwargs)
        return kwargs['easyDict']

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def vrf_add(self, wb, ws, row_num, **kwargs):
        # Open the Network Policies Worksheet
        ws_net = wb['Network Policies']
        rows = ws_net.max_row

        # Dicts for required and optional args
        required_args = {'site_group': '',
                         'Tenant': '',
                         'VRF': '',
                         'VRF_Policy': '',
                         'Policy_Name': '',
                         'pc_enf_pref': '',
                         'pc_enf_dir': '',
                         'bd_enforce': '',
                         'enf_type': '',
                         'fvEpRetPol': '',
                         'monEPGPol': '',
                         'ip_dp_learning': '',
                         'knw_mcast_act': ''}
        optional_args = {'alias': '',
                         'description': '',
                         'annotation': '',
                         'cons_vzBrCP': '',
                         'vzCPIf': '',
                         'prov_vzBrCP': '',
                         'bgpCtxPol': '',
                         'bgpCtxAfPol': '',
                         'ospfCtxPol': '',
                         'ospfCtxAfPol': '',
                         'eigrpCtxAfPol': '',
                         'l3extRouteTagPol': '',
                         'l3extVrfValidationPol': ''}


        # Get the VRF Policies from the Network Policies Tab
        func = 'VRF'
        count = countKeys(ws_net, func)
        row_vrf = ''
        var_dict = findVars(ws_net, func, rows, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('VRF_Policy'):
                row_vrf = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'site_group', templateVars['site_group'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
            validating.name_rule(row_num, ws, 'VRF', templateVars['VRF'])
            if not templateVars['alias'] == None:
                validating.name_rule(row_num, ws, 'alias', templateVars['alias'])
            if not templateVars['description'] == None:
                validating.description(row_num, ws, 'description', templateVars['description'])
            if not templateVars['annotation'] == None:
                if re.match(',', templateVars['annotation']):
                    for tag in templateVars['annotation'].split(','):
                        validating.name_rule(row_num, ws, 'annotation', tag)
                else:
                    validating.name_rule(row_num, ws, 'annotation', templateVars['annotation'])
            validating.values(row_vrf, ws_net, 'bd_enforce', templateVars['bd_enforce'], ['no', 'yes'])
            validating.values(row_vrf, ws_net, 'ip_dp_learning', templateVars['ip_dp_learning'], ['disabled', 'enabled'])
            validating.values(row_vrf, ws_net, 'knw_mcast_act', templateVars['knw_mcast_act'], ['deny', 'permit'])
            validating.values(row_vrf, ws_net, 'pc_enf_dir', templateVars['pc_enf_dir'], ['egress', 'ingress'])
            validating.values(row_vrf, ws_net, 'pc_enf_pref', templateVars['pc_enf_pref'], ['enforced', 'unenforced'])
            validating.values(row_vrf, ws_net, 'enf_type', templateVars['enf_type'], ['contract', 'pref_grp', 'vzAny'])
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(errorReturn)

        if templateVars['cons_vzBrCP'] == 'default':
            templateVars['cons_vzBrCP'] = 'uni/tn-common/brc-default'
        if templateVars['prov_vzBrCP'] == 'default':
            templateVars['prov_vzBrCP'] = 'uni/tn-common/brc-default'
        if templateVars['vzCPIf'] == 'default':
            templateVars['vzCPIf'] = 'uni/tn-common/cif-default'
        if templateVars['bgpCtxPol'] == 'default':
            templateVars['bgpCtxPol'] = 'uni/tn-common/bgpCtxP-default'
        if templateVars['bgpCtxAfPol'] == 'default':
            templateVars['bgpCtxAfPol'] = 'uni/tn-common/bgpCtxAfP-default'
        if templateVars['eigrpCtxAfPol'] == 'default':
            templateVars['eigrpCtxAfPol'] = 'uni/tn-common/eigrpCtxAfP-default'
        if templateVars['ospfCtxPol'] == 'default':
            templateVars['ospfCtxPol'] = 'uni/tn-common/ospfCtxP-default'
        if templateVars['ospfCtxAfPol'] == 'default':
            templateVars['ospfCtxAfPol'] = 'uni/tn-common/ospfCtxP-default'
        if templateVars['fvEpRetPol'] == 'default':
            templateVars['fvEpRetPol'] = 'uni/tn-common/epRPol-default'
        if templateVars['monEPGPol'] == 'default':
            templateVars['monEPGPol'] = 'uni/tn-common/monepg-default'
        if templateVars['l3extVrfValidationPol'] == 'default':
            templateVars['l3extVrfValidationPol'] = 'uni/tn-common/vrfvalidationpol-default'

        # Define the Template Source
        template_file = "vrf.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'VRF_%s.tf' % (templateVars['VRF'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        write_to_site(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        if templateVars['enf_type'] == 'pref_grp':
            # Define the Template Source
            template_file = "pref_grp.jinja2"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = 'VRF_%s.tf' % (templateVars['VRF'])
            dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
            write_to_site(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

        elif templateVars['enf_type'] == 'vzAny':
            # Define the Template Source
            template_file = "vzAny.jinja2"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = 'VRF_%s.tf' % (templateVars['VRF'])
            dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
            write_to_site(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "snmp_ctx.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'VRF_%s.tf' % (templateVars['VRF'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        write_to_site(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

    def wr_auto_tfvars(self, **easyDict):
        functionList = ['tenants']
        for func in functionList:
            func_type = 'tenants'
            # jsonDump = json.dumps(easyDict[func_type][func], indent=4)
            # print(jsonDump)
            # exit()
            for item in easyDict[func_type][func]:
                for k, v in item.items():
                    for i in v:
                        templateVars = i
                        templateVars['row_num'] = '%s_section' % (func)
                        templateVars['site_group'] = k
                        templateVars['ws'] = easyDict['wb']['Tenants']
                        
                        # Add Variables for Template Functions
                        templateVars["initial_write"] = True
                        templateVars['policy_type'] = func.replace('_', ' ').capitalize()
                        templateVars["template_file"] = 'template_open.jinja2'
                        templateVars['template_type'] = func
                        templateVars['tfvars_file'] = func
                        
                        # Write to the Template file and Return Dictionary
                        write_to_site(self, **templateVars)

                        templateVars["initial_write"] = False
                        templateVars["template_file"] = f'{func}.jinja2'

                        # Write to the Template file and Return Dictionary
                        write_to_site(self, **templateVars)

                        templateVars["initial_write"] = False
                        templateVars["template_file"] = 'template_close.jinja2'

                        # Write to the Template file and Return Dictionary
                        write_to_site(self, **templateVars)
